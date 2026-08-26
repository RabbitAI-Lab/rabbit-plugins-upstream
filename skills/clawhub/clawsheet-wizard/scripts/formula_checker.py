#!/usr/bin/env python3
"""ClawSheet Wizard — formel-tjekker (unik feature)."""
import sys
import re


def check(path: str, fix: bool = False) -> str:
    try:
        import openpyxl
    except ImportError:
        sys.exit("FEJL: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)
    out = [f"🧮 Formel-tjek: {path}", ""]
    errors = []
    for ws, ws_vals in zip(wb.worksheets, wb_values.worksheets):
        for row, row_vals in zip(ws.iter_rows(), ws_vals.iter_rows()):
            for cell, cell_vals in zip(row, row_vals):
                v = cell.value
                cached = cell_vals.value
                # 1) Fejl-streng i selve formlen
                if isinstance(v, str) and v.startswith("="):
                    if re.search(r"#REF!|#DIV/0!|#VALUE!|#NAME\?|#NULL!", v):
                        errors.append(f"  [{ws.title}!{cell.coordinate}] {v[:80]}")
                    # cirkulær reference (formel refererer egen celle)
                    refs = re.findall(r"([A-Z]{1,3}\d+)", v)
                    if cell.coordinate in refs:
                        errors.append(f"  [{ws.title}!{cell.coordinate}] CIRKULÆR: {v[:80]}")
                # 2) Cached fejl-værdi fra sidste beregning (fanger #DIV/0! osv.)
                if isinstance(cached, str) and cached.startswith("#") and cached in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#N/A"):
                    errors.append(f"  [{ws.title}!{cell.coordinate}] CACHED-FEJL {cached}: {str(v)[:60]}")
    if errors:
        out.append(f"⚠️ {len(errors)} fejl fundet:")
        out.extend(errors)
        if fix:
            wb2 = openpyxl.load_workbook(path, data_only=False)
            # data_only=True version for at se cached værdier — vi markerer kun her
            out.append("\n💡 Kør med --report for detaljeret rapport; auto-fix kræver manuel gennemgang af formler.")
    else:
        out.append("✅ Ingen formel-fejl fundet.")
    out.append(f"\n{len(wb.sheetnames)} ark tjekket: {', '.join(wb.sheetnames)}")
    return "\n".join(out)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("BRUG: python3 formula_checker.py bog.xlsx [--report rapport.md]")
    report = check(sys.argv[1], fix="--fix" in sys.argv)
    if "--report" in sys.argv:
        p = sys.argv[sys.argv.index("--report") + 1]
        with open(p, "w", encoding="utf-8") as f:
            f.write(report)
        print(f"✅ Rapport: {p}")
    else:
        print(report)
