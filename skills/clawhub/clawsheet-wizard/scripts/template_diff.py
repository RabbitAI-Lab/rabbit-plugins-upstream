#!/usr/bin/env python3
"""ClawSheet Wizard — skabelon-diff (unik feature)."""
import sys


def diff_workbooks(orig: str, new: str) -> str:
    try:
        import openpyxl
    except ImportError:
        sys.exit("FEJL: pip install openpyxl")

    wb1 = openpyxl.load_workbook(orig, data_only=False)
    wb2 = openpyxl.load_workbook(new, data_only=False)
    out = [f"🔍 Skabelon-diff: {orig} ↔ {new}", ""]

    # Ark-navne
    s1, s2 = set(wb1.sheetnames), set(wb2.sheetnames)
    if s1 != s2:
        out.append(f"  Ark ændret: +{sorted(s2 - s1)} -{sorted(s1 - s2)}")

    changes = 0
    for sheet in wb1.sheetnames:
        if sheet not in wb2.sheetnames:
            continue
        ws1, ws2 = wb1[sheet], wb2[sheet]
        for row in ws1.iter_rows():
            for cell in row:
                c2 = ws2[cell.coordinate]
                if cell.value != c2.value:
                    changes += 1
                    if changes <= 25:
                        out.append(f"  [{sheet}!{cell.coordinate}]")
                        out.append(f"    FØR: {str(cell.value)[:80]}")
                        out.append(f"    EFTER: {str(c2.value)[:80]}")

    out.append(f"\n{changes} ændringer i alt" + ("" if changes <= 25 else " (viser første 25)"))
    return "\n".join(out)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit("BRUG: python3 template_diff.py original.xlsx redigeret.xlsx")
    print(diff_workbooks(sys.argv[1], sys.argv[2]))
