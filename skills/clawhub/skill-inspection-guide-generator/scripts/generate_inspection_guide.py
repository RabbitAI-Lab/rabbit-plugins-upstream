#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检验指导书生成脚本

功能：
1. 根据JSON格式的检验数据生成Excel检验指导书
2. 支持使用用户提供的模板或默认模板
3. 自动命名并保存文件

使用方式：
python generate_inspection_guide.py --product_name "产品名称" --inspection_data '<JSON数据>' [--template_path "模板路径"] [--output_dir "输出目录"]
"""

import argparse
import json
import os
import sys
from datetime import datetime
from typing import Dict, List, Optional, Any

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：缺少 openpyxl 库，请执行：pip install openpyxl==3.1.2")
    sys.exit(1)

try:
    from docx import Document
except ImportError:
    print("警告：缺少 python-docx 库，无法处理Word模板。如需处理Word模板，请执行：pip install python-docx==1.1.0")
    Document = None


# 默认模板字段定义
DEFAULT_HEADERS = [
    ("检验项目", "item", True),
    ("检验方法", "method", True),
    ("检验标准", "standard", True),
    ("检验频次", "frequency", True),
    ("备注", "remark", False),
    ("检验工具", "tool", False),
    ("抽样方案", "sampling", False),
    ("责任部门", "department", False),
]

# 样式定义
HEADER_FONT = Font(name='微软雅黑', size=11, bold=True)
CELL_FONT = Font(name='微软雅黑', size=10)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def parse_inspection_data(data_str: str) -> Dict[str, Any]:
    """
    解析检验数据JSON字符串
    
    Args:
        data_str: JSON格式的检验数据字符串
        
    Returns:
        解析后的字典数据
        
    Raises:
        ValueError: JSON格式错误或缺少必需字段
    """
    try:
        data = json.loads(data_str)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON格式错误: {str(e)}")
    
    # 验证必需字段
    if "product_name" not in data:
        raise ValueError("缺少必需字段: product_name")
    
    if "inspection_items" not in data:
        raise ValueError("缺少必需字段: inspection_items")
    
    if not isinstance(data["inspection_items"], list):
        raise ValueError("inspection_items 必须是数组类型")
    
    # 验证每个检验项的必填字段
    required_fields = ["item", "method", "standard", "frequency"]
    for idx, item in enumerate(data["inspection_items"]):
        if not isinstance(item, dict):
            raise ValueError(f"inspection_items[{idx}] 必须是对象类型")
        
        missing_fields = [f for f in required_fields if f not in item or not item[f]]
        if missing_fields:
            raise ValueError(f"inspection_items[{idx}] 缺少必填字段: {', '.join(missing_fields)}")
    
    return data


def extract_word_table_headers(doc_path: str) -> Optional[List[str]]:
    """
    从Word文档中提取表格表头
    
    Args:
        doc_path: Word文档路径
        
    Returns:
        表头列表，如果无法提取则返回None
    """
    if Document is None:
        print("警告：无法处理Word模板，python-docx库未安装")
        return None
    
    try:
        doc = Document(doc_path)
        tables = doc.tables
        
        if not tables:
            return None
        
        # 使用第一个表格
        table = tables[0]
        headers = []
        
        # 提取第一行作为表头
        for cell in table.rows[0].cells:
            header = cell.text.strip()
            if header:
                headers.append(header)
        
        return headers if headers else None
        
    except Exception as e:
        print(f"读取Word模板失败: {str(e)}")
        return None


def map_headers_to_fields(headers: List[str]) -> Dict[str, str]:
    """
    将模板表头映射到字段名
    
    Args:
        headers: 表头列表
        
    Returns:
        表头到字段名的映射字典
    """
    header_mapping = {}
    
    # 标准映射关系
    standard_mapping = {
        "检验项目": "item",
        "检验方法": "method",
        "检验标准": "standard",
        "检验频次": "frequency",
        "备注": "remark",
        "检验工具": "tool",
        "抽样方案": "sampling",
        "责任部门": "department",
        "检验项": "item",
        "项目": "item",
        "方法": "method",
        "标准": "standard",
        "频次": "frequency",
    }
    
    for header in headers:
        header_clean = header.strip()
        # 尝试精确匹配
        if header_clean in standard_mapping:
            header_mapping[header] = standard_mapping[header_clean]
        else:
            # 尝试模糊匹配
            for key, field in standard_mapping.items():
                if key in header_clean or header_clean in key:
                    header_mapping[header] = field
                    break
    
    return header_mapping


def create_workbook_with_template(template_path: str) -> tuple:
    """
    使用模板创建工作簿
    
    Args:
        template_path: 模板文件路径
        
    Returns:
        (工作簿对象, 表头映射字典, 表头行索引)
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")
    
    # 判断文件类型
    if template_path.lower().endswith('.docx'):
        # Word模板
        headers = extract_word_table_headers(template_path)
        if headers:
            # 创建新工作簿，使用Word表格的表头
            wb = Workbook()
            ws = wb.active
            ws.title = "检验指导书"
            
            header_mapping = map_headers_to_fields(headers)
            
            # 写入表头
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
                cell.border = THIN_BORDER
            
            return wb, header_mapping, 1
        else:
            print("警告：无法从Word模板提取表头，将使用默认模板")
            return create_default_workbook()
    
    elif template_path.lower().endswith(('.xlsx', '.xls')):
        # Excel模板
        try:
            wb = load_workbook(template_path)
            ws = wb.active
            
            # 查找表头行
            header_row = None
            for row_idx in range(1, min(10, ws.max_row + 1)):
                row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
                # 检查是否包含必需字段
                row_text = ' '.join(str(v) for v in row_values if v)
                if any(keyword in row_text for keyword in ["检验项目", "项目", "检验方法"]):
                    header_row = row_idx
                    break
            
            if header_row is None:
                header_row = 1
            
            # 提取表头
            headers = []
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=header_row, column=col_idx).value
                if header:
                    headers.append(str(header))
            
            header_mapping = map_headers_to_fields(headers)
            
            return wb, header_mapping, header_row
            
        except Exception as e:
            print(f"读取Excel模板失败: {str(e)}，将使用默认模板")
            return create_default_workbook()
    
    else:
        print(f"不支持的模板格式: {template_path}，将使用默认模板")
        return create_default_workbook()


def create_default_workbook() -> tuple:
    """
    创建默认格式的工作簿
    
    Returns:
        (工作簿对象, 表头映射字典, 表头行索引)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "检验指导书"
    
    # 写入表头
    for col_idx, (header, field, _) in enumerate(DEFAULT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    
    # 设置列宽
    column_widths = [20, 25, 30, 15, 20, 15, 15, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 创建表头映射
    header_mapping = {header: field for header, field, _ in DEFAULT_HEADERS}
    
    return wb, header_mapping, 1


def fill_inspection_data(
    wb: Workbook,
    header_mapping: Dict[str, str],
    header_row: int,
    inspection_items: List[Dict[str, str]]
) -> None:
    """
    填充检验数据到工作簿
    
    Args:
        wb: 工作簿对象
        header_mapping: 表头映射字典
        header_row: 表头所在行
        inspection_items: 检验项列表
    """
    ws = wb.active
    
    # 获取列映射：字段名 -> 列索引
    field_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col_idx).value
        if header and header in header_mapping:
            field_to_col[header_mapping[header]] = col_idx
    
    # 确保必填字段存在
    required_fields = ["item", "method", "standard", "frequency"]
    missing_fields = [f for f in required_fields if f not in field_to_col]
    if missing_fields:
        raise ValueError(f"模板缺少必需字段: {', '.join(missing_fields)}")
    
    # 填充数据
    start_row = header_row + 1
    
    for item_idx, item_data in enumerate(inspection_items):
        row_idx = start_row + item_idx
        
        for field, col_idx in field_to_col.items():
            value = item_data.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER


def generate_inspection_guide(
    product_name: str,
    inspection_data: Dict[str, Any],
    template_path: Optional[str] = None,
    output_dir: str = "."
) -> str:
    """
    生成检验指导书
    
    Args:
        product_name: 产品名称
        inspection_data: 检验数据字典
        template_path: 模板路径（可选）
        output_dir: 输出目录
        
    Returns:
        生成的文件路径
    """
    # 创建工作簿
    if template_path and os.path.exists(template_path):
        wb, header_mapping, header_row = create_workbook_with_template(template_path)
    else:
        wb, header_mapping, header_row = create_default_workbook()
    
    # 填充数据
    inspection_items = inspection_data.get("inspection_items", [])
    fill_inspection_data(wb, header_mapping, header_row, inspection_items)
    
    # 生成文件名
    date_str = datetime.now().strftime("%Y%m%d")
    filename = f"{product_name}_检验指导书_{date_str}.xlsx"
    
    # 确保输出目录存在
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 保存文件
    output_path = os.path.join(output_dir, filename)
    wb.save(output_path)
    
    return output_path


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="检验指导书生成脚本",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument(
        "--product_name",
        required=True,
        help="产品名称"
    )
    
    parser.add_argument(
        "--inspection_data",
        required=True,
        help="JSON格式的检验数据"
    )
    
    parser.add_argument(
        "--template_path",
        default=None,
        help="模板文件路径（可选）"
    )
    
    parser.add_argument(
        "--output_dir",
        default=".",
        help="输出目录（默认为当前目录）"
    )
    
    args = parser.parse_args()
    
    try:
        # 解析检验数据
        inspection_data = parse_inspection_data(args.inspection_data)
        
        # 使用命令行参数中的产品名称（优先级更高）
        product_name = args.product_name
        
        # 生成检验指导书
        output_path = generate_inspection_guide(
            product_name=product_name,
            inspection_data=inspection_data,
            template_path=args.template_path,
            output_dir=args.output_dir
        )
        
        print(f"检验指导书已生成: {output_path}")
        
    except ValueError as e:
        print(f"错误: {str(e)}")
        sys.exit(1)
    except Exception as e:
        print(f"生成失败: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
