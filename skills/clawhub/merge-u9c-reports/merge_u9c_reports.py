import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
import re
import glob

# 文件类型识别规则（通过文件名前缀匹配）
FILE_TYPE_RULES = {
    'PR': {
        'pattern': r'^PR\d+.*\.xlsx$',
        'sheet_name': '请购单',
        'date_col': 0,
        'status_col': 1,
        'type_col': 2,
        'no_col': 3,
        'initiator_col': 4,
        'arrive_col': 5,
        'hand_col': 6,
        'handler_col': 7,
        'remark_col': None,
        'section_title': '请购单（昨日停滞）',
    },
    'Receivement': {
        'pattern': r'^Receivement\d+.*\.xlsx$',
        'sheet_name': '收货单',
        'date_col': 0,
        'status_col': 1,
        'type_col': 2,
        'no_col': 3,
        'initiator_col': 10,
        'arrive_col': 5,
        'hand_col': 6,
        'handler_col': 7,
        'remark_col': 11,
        'section_title': '收货单（昨日停滞）',
    },
    'SO': {
        'pattern': r'^SO\d+.*\.xlsx$',
        'sheet_name': '销售订单',
        'date_col': 0,
        'status_col': 1,
        'type_col': 2,
        'no_col': 3,
        'initiator_col': 4,
        'arrive_col': 5,
        'hand_col': 6,
        'handler_col': 7,
        'remark_col': 12,
        'section_title': '销售订单（昨日停滞）',
    },
    'Ship': {
        'pattern': r'^Ship\d+.*\.xlsx$',
        'sheet_name': '出货单',
        'date_col': 0,
        'status_col': 1,
        'type_col': 2,
        'no_col': 3,
        'initiator_col': 4,
        'arrive_col': 5,
        'hand_col': 6,
        'handler_col': 7,
        'remark_col': None,
        'section_title': '出货单(超过7天)',
    },
    'PayReqFundHead': {
        'pattern': r'^PayReqFundHead\d+.*\.xlsx$',
        'sheet_name': '付款请款单表头',
        'date_col': 0,
        'status_col': 1,
        'type_col': 2,
        'no_col': 3,
        'initiator_col': 4,
        'arrive_col': 5,
        'hand_col': 6,
        'handler_col': 7,
        'remark_col': 12,
        'section_title': '付款请款单（昨日停滞）',
    },
}

OUTPUT_COLS = ['业务日期', '状态', '单据类型', '单据编号', '发起人', '送达时间', '在手时间', '当前处理人', '备注']

# 固定的输出顺序
OUTPUT_ORDER = ['PR', 'Receivement', 'SO', 'Ship', 'PayReqFundHead']


def identify_file_type(filename):
    """通过文件名识别报表类型"""
    basename = os.path.basename(filename)
    for type_key, rule in FILE_TYPE_RULES.items():
        if re.match(rule['pattern'], basename):
            return type_key
    return None


def scan_report_files(folder_path):
    """扫描文件夹中的所有报表文件，返回分类字典"""
    files_dict = {}

    # 扫描所有 xlsx 文件
    xlsx_files = glob.glob(os.path.join(folder_path, '*.xlsx'))

    for file_path in xlsx_files:
        file_type = identify_file_type(file_path)
        if file_type:
            if file_type not in files_dict:
                files_dict[file_type] = []
            files_dict[file_type].append(file_path)
            print(f'识别: {os.path.basename(file_path)} -> {file_type}')

    return files_dict


def extract_data_from_file(file_path, file_type):
    """从单个文件提取数据，返回 DataFrame"""
    rule = FILE_TYPE_RULES[file_type]
    df = pd.read_excel(file_path, header=0, sheet_name=0)

    # 跳过最后一行（合计行）
    if len(df) > 0 and df.iloc[-1].isna().sum() > len(df.columns) / 2:
        df = df.iloc[:-1]

    rows = []
    for _, row in df.iterrows():
        # 跳过标题行（防止重复）
        status_val = str(row.iloc[rule['status_col']]).strip() if pd.notna(row.iloc[rule['status_col']]) else ''
        if status_val in ['状态', '行状态', '单据状态']:
            continue

        # 提取各字段
        date_val = row.iloc[rule['date_col']] if rule['date_col'] is not None else ''
        type_val = row.iloc[rule['type_col']] if rule['type_col'] is not None else ''
        no_val = row.iloc[rule['no_col']] if rule['no_col'] is not None else ''
        initiator_val = row.iloc[rule['initiator_col']] if rule['initiator_col'] is not None else ''
        arrive_val = row.iloc[rule['arrive_col']] if rule['arrive_col'] is not None else ''
        hand_val = row.iloc[rule['hand_col']] if rule['hand_col'] is not None else ''
        handler_val = row.iloc[rule['handler_col']] if rule['handler_col'] is not None else ''
        remark_val = row.iloc[rule['remark_col']] if rule['remark_col'] is not None else ''

        # 转换日期
        if pd.notna(date_val):
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val)
        else:
            date_str = ''

        rows.append([
            date_str,
            status_val,
            str(type_val) if pd.notna(type_val) else '',
            str(no_val) if pd.notna(no_val) else '',
            str(initiator_val) if pd.notna(initiator_val) else '',
            str(arrive_val) if pd.notna(arrive_val) else '',
            str(hand_val) if pd.notna(hand_val) else '',
            str(handler_val) if pd.notna(handler_val) else '',
            str(remark_val) if pd.notna(remark_val) else '',
        ])

    return pd.DataFrame(rows, columns=OUTPUT_COLS), rule['section_title']


def merge_u9c_reports_from_folder(folder_path, output_file, date_str=None):
    """
    自动扫描文件夹中的所有报表文件并合并。

    参数:
        folder_path: 报表文件所在文件夹
        output_file: 输出文件路径
        date_str: 日期字符串（用于 sheet 名称），格式如 '20260814'
    """
    print(f'扫描文件夹: {folder_path}')
    files_dict = scan_report_files(folder_path)

    if not files_dict:
        print('错误: 未找到任何报表文件')
        return None

    print(f'\n找到报表类型: {list(files_dict.keys())}')

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = date_str if date_str else datetime.now().strftime('%Y%m%d')

    # 样式定义
    title_font = Font(name='微软雅黑', size=14, bold=True)
    section_font = Font(name='微软雅黑', size=11, bold=True)
    header_font = Font(name='微软雅黑', size=10, bold=True)
    data_font = Font(name='微软雅黑', size=10)
    header_fill = PatternFill('solid', start_color='D9E1F2')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    current_row = 1

    # 写入总标题
    ws.cell(row=current_row, column=1, value='U9C超时停滞单据')
    ws.cell(row=current_row, column=1).font = title_font
    current_row += 1

    # 按固定顺序处理各类型
    total_rows = 0
    for file_type in OUTPUT_ORDER:
        if file_type not in files_dict:
            continue

        file_list = files_dict[file_type]
        section_title = FILE_TYPE_RULES[file_type]['section_title']

        # 收集该类型所有文件的数据
        all_dfs = []
        for file_path in file_list:
            df, _ = extract_data_from_file(file_path, file_type)
            if len(df) > 0:
                all_dfs.append(df)
                print(f'  处理: {os.path.basename(file_path)} ({len(df)} 行)')

        if not all_dfs:
            continue

        # 合并该类型的所有数据
        combined_df = pd.concat(all_dfs, ignore_index=True)

        # 写入分节标题
        ws.cell(row=current_row, column=1, value=section_title)
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 1

        # 写入表头
        for col_idx, col_name in enumerate(OUTPUT_COLS, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        current_row += 1

        # 写入数据行
        for _, row_data in combined_df.iterrows():
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value if value else None)
                cell.font = data_font
                cell.border = border
                cell.alignment = left_align if col_idx == 9 else center_align
            current_row += 1
            total_rows += 1

        # 分节后空一行
        current_row += 1

    # 写入底部说明
    ws.cell(row=current_row, column=1, value='根据《关于T+系统的管理规定》4.3.5条款，给予2分/次的行政处罚。')
    ws.cell(row=current_row, column=1).font = data_font

    # 调整列宽
    col_widths = [12, 10, 16, 22, 10, 20, 16, 16, 30]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    # 冻结首行
    ws.freeze_panes = 'A3'

    wb.save(output_file)
    print(f'\n输出文件已保存: {output_file}')
    print(f'共处理 {len(files_dict)} 种类型，{total_rows} 行数据')
    return output_file


def merge_u9c_reports_from_files(file_paths, output_file, date_str=None):
    """
    从指定的文件列表合并报表。

    参数:
        file_paths: 文件路径列表
        output_file: 输出文件路径
        date_str: 日期字符串（用于 sheet 名称）
    """
    files_dict = {}

    for file_path in file_paths:
        file_type = identify_file_type(file_path)
        if file_type:
            if file_type not in files_dict:
                files_dict[file_type] = []
            files_dict[file_type].append(file_path)
            print(f'识别: {os.path.basename(file_path)} -> {file_type}')

    if not files_dict:
        print('错误: 未找到任何报表文件')
        return None

    # 后续处理与 merge_u9c_reports_from_folder 相同
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = date_str if date_str else datetime.now().strftime('%Y%m%d')

    title_font = Font(name='微软雅黑', size=14, bold=True)
    section_font = Font(name='微软雅黑', size=11, bold=True)
    header_font = Font(name='微软雅黑', size=10, bold=True)
    data_font = Font(name='微软雅黑', size=10)
    header_fill = PatternFill('solid', start_color='D9E1F2')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    current_row = 1

    ws.cell(row=current_row, column=1, value='U9C超时停滞单据')
    ws.cell(row=current_row, column=1).font = title_font
    current_row += 1

    total_rows = 0
    for file_type in OUTPUT_ORDER:
        if file_type not in files_dict:
            continue

        file_list = files_dict[file_type]
        section_title = FILE_TYPE_RULES[file_type]['section_title']

        all_dfs = []
        for file_path in file_list:
            df, _ = extract_data_from_file(file_path, file_type)
            if len(df) > 0:
                all_dfs.append(df)
                print(f'  处理: {os.path.basename(file_path)} ({len(df)} 行)')

        if not all_dfs:
            continue

        combined_df = pd.concat(all_dfs, ignore_index=True)

        ws.cell(row=current_row, column=1, value=section_title)
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 1

        for col_idx, col_name in enumerate(OUTPUT_COLS, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        current_row += 1

        for _, row_data in combined_df.iterrows():
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value if value else None)
                cell.font = data_font
                cell.border = border
                cell.alignment = left_align if col_idx == 9 else center_align
            current_row += 1
            total_rows += 1

        current_row += 1

    ws.cell(row=current_row, column=1, value='根据《关于T+系统的管理规定》4.3.5条款，给予2分/次的行政处罚。')
    ws.cell(row=current_row, column=1).font = data_font

    col_widths = [12, 10, 16, 22, 10, 20, 16, 16, 30]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    ws.freeze_panes = 'A3'

    wb.save(output_file)
    print(f'\n输出文件已保存: {output_file}')
    print(f'共处理 {len(files_dict)} 种类型，{total_rows} 行数据')
    return output_file


if __name__ == '__main__':
    import sys

    if len(sys.argv) > 1:
        # 参数模式：指定文件夹
        folder = sys.argv[1]
        output = sys.argv[2] if len(sys.argv) > 2 else r'C:\Users\jinshan\Desktop\合并输出.xlsx'
        date_str = sys.argv[3] if len(sys.argv) > 3 else None
        merge_u9c_reports_from_folder(folder, output, date_str)
    else:
        # 测试模式：使用 Downloads 文件夹
        folder = r'C:\Users\jinshan\Downloads'
        output = r'C:\Users\jinshan\Desktop\合并输出.xlsx'
        merge_u9c_reports_from_folder(folder, output, '20260814')
