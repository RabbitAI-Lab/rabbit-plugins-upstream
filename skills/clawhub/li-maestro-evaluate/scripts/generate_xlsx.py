#!/usr/bin/env python3
"""Generate standardized .xlsx threat model output from MAESTRO evaluation.

Usage:
    python generate_xlsx.py <project-dir> [--output OUTPUT]

Creates 4 sheets:
  Sheet 1: 风险总表 (Full Risk Register) — all threats with auto-filter
  Sheet 2: 分类统计 (Category Summary) — threats grouped by 三级九子类
  Sheet 3: 缓解措施 (Mitigation Summary) — all mitigations
  Sheet 4: 说明 (Legend) — definitions, scales, disclaimer
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ── constants ────────────────────────────────────────────────────────

AI_RISK_CATEGORIES = {
    "技术内生安全风险": {
        "数据安全风险": "E1", "算法安全风险": "E2", "模型安全风险": "E3",
    },
    "技术应用安全风险": {
        "网络系统安全风险": "A1", "供应链安全风险": "A2", "隐私保护风险": "A3",
    },
    "应用衍生安全风险": {
        "滥用风险": "D1", "信任风险": "D2", "合规风险": "D3",
    },
}

# reverse: code → (大类, 子类)
CODE_TO_CATEGORY = {}
for parent, subs in AI_RISK_CATEGORIES.items():
    for sub, code in subs.items():
        CODE_TO_CATEGORY[code] = (parent, sub)

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
CRITICAL_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
LOW_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ── helpers ──────────────────────────────────────────────────────────

def _read_json(path):
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return None


def _read_md(path):
    try:
        with open(path, encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return None


def _cjk_width(v, max_width=50, min_width=8):
    """Calculate column width accounting for CJK characters (approx 2x width)."""
    s = str(v or "")
    ascii_len = sum(1 for c in s if ord(c) < 128)
    cjk_len = len(s) - ascii_len
    w = ascii_len * 1.1 + cjk_len * 2.0
    return max(min(w + 2, max_width), min_width)


def _auto_width(ws, col_count, max_width=50, min_width=8):
    for c in range(1, col_count + 1):
        widths = []
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
            widths.append(_cjk_width(row[0], max_width, min_width))
        ws.column_dimensions[get_column_letter(c)].width = max(widths)


def _style_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_body(ws, start_row, end_row, col_count):
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _apply_risk_fill(ws, col, start_row, end_row):
    for r in range(start_row, end_row + 1):
        val = str(ws.cell(row=r, column=col).value or "").lower()
        if "critical" in val:
            ws.cell(row=r, column=col).fill = CRITICAL_FILL
        elif "high" in val:
            ws.cell(row=r, column=col).fill = HIGH_FILL
        elif "medium" in val:
            ws.cell(row=r, column=col).fill = MEDIUM_FILL
        elif "low" in val:
            ws.cell(row=r, column=col).fill = LOW_FILL


# ── parsers ──────────────────────────────────────────────────────────

def _parse_table_rows(md_text, header_keywords, min_cols=5):
    rows = []
    if not md_text:
        return rows
    in_table = False
    for line in md_text.splitlines():
        stripped = line.strip()
        if not stripped.startswith("|"):
            if in_table:
                break
            continue
        if any(kw in stripped for kw in header_keywords):
            in_table = True
            continue
        if in_table and stripped.startswith("|---"):
            continue
        if in_table:
            cells = [c.strip() for c in stripped.strip("| ").split("|")]
            if len(cells) >= min_cols:
                rows.append(cells)
    return rows


def _extract_ai_risk_codes(md_text):
    """Scan threat register for AI Risk Codes (E1-E3, A1-A3, D1-D3)."""
    codes = []
    if not md_text:
        return codes
    for row in _parse_table_rows(md_text, ["Local ID", "Threat Name", "Name"], 4):
        for cell in row:
            m = re.search(r"\b(E[123]|A[123]|D[123])\b", cell)
            if m:
                codes.append(m.group(1))
                break
    # also scan vertical card format
    for m in re.finditer(r"\*\*AI Risk Code\*\*\s*\|\s*(E[123]|A[123]|D[123])", md_text):
        codes.append(m.group(1))
    return codes


# ── builders ─────────────────────────────────────────────────────────

def build_xlsx(project_dir, output_path=None):
    project_dir = Path(project_dir)
    state = _read_json(project_dir / "state.json")
    analysis_mode = (state or {}).get("analysis_mode", "full_assessment")

    if analysis_mode == "mvtm_checklist":
        # MVTM mode: single checklist file contains all data
        mvtm_data = _read_md(project_dir / "01-mvtm-checklist.md") or ""
        phase6 = mvtm_data
        phase7 = mvtm_data
    else:
        # Full Assessment mode: individual phase files
        phase6 = _read_md(project_dir / "06-threat-register.md")
        phase7 = _read_md(project_dir / "07-mitigations.md")

    risk_cls = _read_md(project_dir / "11-ai-risk-classification.md")

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════
    # Sheet 1: 风险总表 (Full Risk Register)
    # ══════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "风险总表"
    headers1 = ["Local ID", "Threat Name", "Description", "Layer", "ASI ID",
                 "Severity", "Likelihood", "Risk Level", "AI Risk Category",
                 "AI Risk Code", "Mitigation ID", "Implementation Status", "Risk Owner"]
    ws1.append(headers1)
    _style_header(ws1, 1, len(headers1))

    threats = _parse_table_rows(phase6, ["Local ID", "Threat Name"], 5)
    for t in threats[:500]:
        row = t[:13] if len(t) >= 13 else t + [""] * (13 - len(t))
        ws1.append(row)

    n_risk = ws1.max_row
    if n_risk > 1:
        _style_body(ws1, 2, n_risk, len(headers1))
        _apply_risk_fill(ws1, 8, 2, n_risk)
        tab = Table(ws1.min_row, ws1.min_column, n_risk, len(headers1))
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws1.add_table(tab)
    _auto_width(ws1, len(headers1))

    # ══════════════════════════════════════════════════════════════════
    # Sheet 2: 分类统计 (Category Summary) — populated from real data
    # ══════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("分类统计")
    headers2 = ["大类", "子类", "Risk Code", "Critical", "High", "Medium", "Low", "Total"]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))

    # scan actual threats for their AI Risk Codes
    actual_codes = _extract_ai_risk_codes(phase6)

    row_idx = 2
    for cat, subs in AI_RISK_CATEGORIES.items():
        for sub, code in subs.items():
            critical = sum(1 for c in actual_codes if c == code)
            ws2.append([cat, sub, code, critical, 0, 0, 0, critical])
            row_idx += 1

    _style_body(ws2, 2, row_idx - 1, len(headers2))
    _auto_width(ws2, len(headers2))

    # ══════════════════════════════════════════════════════════════════
    # Sheet 3: 缓解措施 (Mitigation Summary)
    # ══════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("缓解措施")
    headers3 = ["Mitigation ID", "Catalog ID", "Type", "Cost",
                "Effectiveness", "Status", "Targeted Threats", "AI Risk Code"]
    ws3.append(headers3)
    _style_header(ws3, 1, len(headers3))

    mitigations = _parse_table_rows(phase7, ["Mitigation ID", "Catalog ID"], 4)
    for m in mitigations[:300]:
        row = m[:8] if len(m) >= 8 else m + [""] * (8 - len(m))
        ws3.append(row)

    n_mit = ws3.max_row
    if n_mit > 1:
        _style_body(ws3, 2, n_mit, len(headers3))
    _auto_width(ws3, len(headers3))

    # ══════════════════════════════════════════════════════════════════
    # Sheet 4: 说明 (Legend)
    # ══════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("说明")

    ws4.cell(row=1, column=1).value = "三级九子类定义"
    ws4.cell(row=1, column=1).font = Font(bold=True, size=12, color="2F5496")
    ws4.cell(row=2, column=1).value = "大类"
    ws4.cell(row=2, column=2).value = "子类"
    ws4.cell(row=2, column=3).value = "Code"
    _style_header(ws4, 2, 3)
    legend_data = [
        ["技术内生安全风险", "数据安全风险", "E1"],
        ["技术内生安全风险", "算法安全风险", "E2"],
        ["技术内生安全风险", "模型安全风险", "E3"],
        ["技术应用安全风险", "网络系统安全风险", "A1"],
        ["技术应用安全风险", "供应链安全风险", "A2"],
        ["技术应用安全风险", "隐私保护风险", "A3"],
        ["应用衍生安全风险", "滥用风险", "D1"],
        ["应用衍生安全风险", "信任风险", "D2"],
        ["应用衍生安全风险", "合规风险", "D3"],
    ]
    for i, row in enumerate(legend_data, start=3):
        for j, val in enumerate(row, start=1):
            ws4.cell(row=i, column=j).value = val
    _style_body(ws4, 3, 11, 3)

    offset = 13
    ws4.cell(row=offset, column=1).value = "严重性等级"
    ws4.cell(row=offset, column=1).font = Font(bold=True, size=12, color="2F5496")
    sev_rows = [
        ["严重性", "说明"],
        ["Critical", "完全失效，后果严重"],
        ["High", "对运营有重大影响"],
        ["Medium", "中等影响，可管理"],
        ["Low", "轻微影响，易处理"],
    ]
    for i, row in enumerate(sev_rows, start=offset + 1):
        for j, val in enumerate(row, start=1):
            ws4.cell(row=i, column=j).value = val
    _style_header(ws4, offset + 1, 2)
    _style_body(ws4, offset + 2, offset + 5, 2)

    offset2 = offset + 7
    ws4.cell(row=offset2, column=1).value = "可能性等级"
    ws4.cell(row=offset2, column=1).font = Font(bold=True, size=12, color="2F5496")
    lik_rows = [
        ["可能性", "说明"],
        ["Very Likely", "攻击路径已知且易于利用 (>90%)"],
        ["Likely", "存在直接攻击路径 (50-90%)"],
        ["Possible", "需要中等技能或单一控制失效 (10-50%)"],
        ["Unlikely", "需要高级攻击者且多重控制失效 (<10%)"],
    ]
    for i, row in enumerate(lik_rows, start=offset2 + 1):
        for j, val in enumerate(row, start=1):
            ws4.cell(row=i, column=j).value = val
    _style_header(ws4, offset2 + 1, 2)
    _style_body(ws4, offset2 + 2, offset2 + 5, 2)

    offset3 = offset2 + 7
    ws4.cell(row=offset3, column=1).value = "免责声明"
    ws4.cell(row=offset3, column=1).font = Font(bold=True, size=12, color="2F5496")
    ws4.merge_cells(start_row=offset3 + 1, start_column=1, end_row=offset3 + 3, end_column=4)
    disclaimer_cell = ws4.cell(row=offset3 + 1, column=1)
    disclaimer_cell.value = (
        "本威胁模型系借助AI辅助，基于OWASP MAESTRO Playbook框架生成，"
        "AI风险分类映射至《人工智能安全治理框架》2.0版。"
        "在用于生产环境风险决策之前，必须经合格安全专业人员审核。"
    )
    disclaimer_cell.font = Font(italic=True, size=10, color="808080")
    disclaimer_cell.alignment = Alignment(wrap_text=True, vertical="top")

    _auto_width(ws4, 4, max_width=70)

    # ── save ──
    if output_path is None:
        output_path = project_dir / "11-ai-risk-classification.xlsx"
    wb.save(str(output_path))
    print(f"OK: {output_path}")
    return output_path


# ── cli ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate .xlsx threat model from MAESTRO evaluation output"
    )
    parser.add_argument("project_dir", help="Path to <project> directory containing state.json and phase files")
    parser.add_argument("--output", "-o", help="Output .xlsx path (default: <project_dir>/11-ai-risk-classification.xlsx)")
    args = parser.parse_args()

    if not Path(args.project_dir).is_dir():
        print(f"ERROR: directory not found: {args.project_dir}")
        sys.exit(1)

    state = _read_json(Path(args.project_dir) / "state.json")
    analysis_mode = (state or {}).get("analysis_mode", "full_assessment")
    if analysis_mode == "mvtm_checklist":
        prereqs = ["state.json", "01-mvtm-checklist.md"]
    else:
        prereqs = ["state.json", "06-threat-register.md", "07-mitigations.md"]
    missing = [f for f in prereqs if not (Path(args.project_dir) / f).is_file()]
    if missing:
        print(f"WARNING: missing prerequisite files: {', '.join(missing)} (output may be incomplete)")

    out = build_xlsx(args.project_dir, args.output)
    print(f"Done: {out}")


if __name__ == "__main__":
    main()
