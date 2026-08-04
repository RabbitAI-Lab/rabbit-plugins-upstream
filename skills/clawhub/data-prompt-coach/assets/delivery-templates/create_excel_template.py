"""生成 Excel 骨架模板 — data-prompt-coach 交付物模板"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT = Path(__file__).parent / "excel-template.xlsx"

wb = Workbook()

# 样式
header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2F5496")
data_font = Font(name="Microsoft YaHei", size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Sheet 1: Data 数据骨架
ws_data = wb.active
ws_data.title = "Data"

headers = ["序号", "来源/文档名", "字段1", "字段2", "字段3", "异常标记", "抓取时间"]
widths = [8, 30, 14, 14, 14, 20, 18]

for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws_data.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
    ws_data.column_dimensions[get_column_letter(col_idx)].width = w

# 示例行（占位）
sample = [1, "{来源示例}", "{字段1示例}", "{字段2示例}", "{字段3示例}", "", "=NOW()"]
for col_idx, val in enumerate(sample, 1):
    cell = ws_data.cell(row=2, column=col_idx, value=val)
    cell.font = data_font
    cell.border = border
    cell.alignment = center

ws_data.freeze_panes = "A2"

# Sheet 2: Summary 汇总骨架
ws_sum = wb.create_sheet("Summary")
sum_headers = ["统计指标", "公式", "说明"]
sum_widths = [20, 30, 40]
for col_idx, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
    cell = ws_sum.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
    ws_sum.column_dimensions[get_column_letter(col_idx)].width = w

sum_rows = [
    ["总数", "=COUNTA(Data!A:A)-1", "数据总行数（不含表头）"],
    ["异常数", '=COUNTIF(Data!F:F,"待人工核查")', "异常标记列中待人工处理的数量"],
]
for r_idx, row in enumerate(sum_rows, 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws_sum.cell(row=r_idx, column=c_idx, value=val)
        cell.font = data_font
        cell.border = border
        cell.alignment = center

ws_sum.freeze_panes = "A2"

# Sheet 3: 字段定义说明（元数据）
ws_meta = wb.create_sheet("字段定义")
meta_headers = ["字段名", "类型", "必填", "默认值", "校验规则", "说明"]
meta_widths = [14, 12, 8, 14, 24, 40]
for col_idx, (h, w) in enumerate(zip(meta_headers, meta_widths), 1):
    cell = ws_meta.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
    ws_meta.column_dimensions[get_column_letter(col_idx)].width = w

meta_rows = [
    ["序号", "number", "是", "", "", "行号，可追溯"],
    ["来源/文档名", "text", "是", "", "", "数据来源 URL 或文档名"],
    ["字段1", "text", "是", "", "", "替换为实际字段名"],
    ["字段2", "text", "否", "未填写", "", "替换为实际字段名"],
    ["字段3", "text", "否", "未填写", "", "替换为实际字段名"],
    ["异常标记", "text", "否", "", "", "字段缺失/格式异常时填'待人工核查'"],
    ["抓取时间", "datetime", "是", "=NOW()", "", "数据抓取/处理时间"],
]
for r_idx, row in enumerate(meta_rows, 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws_meta.cell(row=r_idx, column=c_idx, value=val)
        cell.font = data_font
        cell.border = border
        cell.alignment = center

ws_meta.freeze_panes = "A2"

wb.save(OUTPUT)
print(f"Excel 模板已生成: {OUTPUT}")
print(f"包含 3 个 sheet: Data（数据骨架）/ Summary（汇总骨架）/ 字段定义（元数据）")
