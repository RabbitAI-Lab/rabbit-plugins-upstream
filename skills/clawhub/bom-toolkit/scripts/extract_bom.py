#!/usr/bin/env python3
"""
BOM提取与合并脚本
从多个供应商BOM中提取物料并合并，或从协同表拆分机械/电气BOM。

Usage:
    python extract_bom.py --merge supplier1.xlsx supplier2.xlsx --output=merged.xlsx
    python extract_bom.py --split 协同表.xlsx --output-dir=./split_output
"""
import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl required. Install: pip install openpyxl")
    sys.exit(1)


# 电气品牌关键词
ELECTRICAL_BRANDS = {
    '施耐德', '西门子', '汇川', '欧姆龙', '菲尼克斯', '三菱', 'ABB',
    '台达', '安川', '松下', '基恩士', 'SMC', 'FESTO', '实点科技',
    'REER', '伯恩施坦', 'MOONS', '天逸', '胜蓝', '正泰', '德力西',
    '禾川', '信捷', '维控', '繁易', '昆仑通态', '威纶通',
}

# 机械品牌关键词
MECHANICAL_BRANDS = {
    '米思米', '怡合达', 'THK', 'HIWIN', 'NSK', 'SKF',
    '上银', '银泰', 'TBI', 'THK',
}

# 电气物料关键词
ELECTRICAL_KEYWORDS = {'PLC', '伺服', '传感器', '断路器', '线缆', '接触器', '继电器', '变频器', 'HMI'}

# 机械物料关键词
MECHANICAL_KEYWORDS = {'钣金', '机加工', '导向件', '传动件', '轴承', '齿轮', '同步带', '丝杠'}


def read_bom(filepath, sheet_name=None):
    """读取BOM文件"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    # 找表头
    header_row = None
    for row_idx in range(1, min(20, ws.max_row + 1)):
        values = [str(ws.cell(row_idx, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        if any(k in values for k in ['序号', '品牌', '名称', '型号']):
            header_row = row_idx
            break
    
    if not header_row:
        return [], []
    
    headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    
    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = {}
        for i, h in enumerate(headers):
            val = ws.cell(row_idx, i + 1).value
            row[h] = val
        if any(row.values()):
            rows.append(row)
    
    wb.close()
    return rows, headers


def classify_material(row):
    """判断物料属于机械还是电气"""
    brand = str(row.get('品牌', '') or '').strip()
    name = str(row.get('名称', '') or '').strip()
    category = str(row.get('类别', '') or '').strip()
    sub_category = str(row.get('物料小类', '') or '').strip()
    
    # 按类别列判断
    if '电气' in category:
        return 'electrical'
    if '机械' in category:
        return 'mechanical'
    
    # 按品牌判断
    if brand in ELECTRICAL_BRANDS or any(b in brand for b in ELECTRICAL_BRANDS):
        return 'electrical'
    if brand in MECHANICAL_BRANDS or any(b in brand for b in MECHANICAL_BRANDS):
        return 'mechanical'
    
    # 按关键词判断
    for kw in ELECTRICAL_KEYWORDS:
        if kw in name or kw in sub_category:
            return 'electrical'
    for kw in MECHANICAL_KEYWORDS:
        if kw in name or kw in sub_category:
            return 'mechanical'
    
    return 'unknown'


def merge_boms(file_list, output_file):
    """合并多个供应商BOM"""
    merged = {}  # (品牌, 名称, 型号) -> {fields}
    
    for filepath in file_list:
        rows, _ = read_bom(filepath)
        for row in rows:
            brand = str(row.get('品牌', '') or '').strip()
            name = str(row.get('名称', '') or '').strip()
            model = str(row.get('型号', '') or '').strip()
            
            if not name:
                continue
            
            key = (brand, name, model)
            qty = row.get('数量', 0) or 0
            
            if key in merged:
                merged[key]['数量'] = (merged[key].get('数量', 0) or 0) + qty
            else:
                merged[key] = dict(row)
                merged[key]['数量'] = qty
    
    # 输出
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '合并BOM'
    
    headers = ['序号', '品牌', '名称', '型号', '规格', '数量', '单位']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = openpyxl.styles.Font(bold=True)
    
    for idx, (key, item) in enumerate(merged.items(), 1):
        ws.cell(idx + 1, 1, idx)
        ws.cell(idx + 1, 2, item.get('品牌', ''))
        ws.cell(idx + 1, 3, item.get('名称', ''))
        ws.cell(idx + 1, 4, item.get('型号', ''))
        ws.cell(idx + 1, 5, item.get('规格', ''))
        ws.cell(idx + 1, 6, item.get('数量', 0))
        ws.cell(idx + 1, 7, item.get('单位', ''))
    
    wb.save(output_file)
    print(f"Merged {len(file_list)} files, {len(merged)} unique items -> {output_file}")


def split_bom(filepath, output_dir):
    """拆分机电协同表为机械BOM和电气BOM"""
    rows, headers = read_bom(filepath)
    
    mechanical = []
    electrical = []
    unknown = []
    
    for row in rows:
        category = classify_material(row)
        if category == 'mechanical':
            mechanical.append(row)
        elif category == 'electrical':
            electrical.append(row)
        else:
            unknown.append(row)
    
    os.makedirs(output_dir, exist_ok=True)
    
    base = os.path.splitext(os.path.basename(filepath))[0]
    
    for label, data in [('机械BOM', mechanical), ('电气BOM', electrical)]:
        if not data:
            continue
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = label
        
        use_headers = headers if headers else ['序号', '品牌', '名称', '型号', '规格', '数量', '单位']
        for i, h in enumerate(use_headers, 1):
            ws.cell(1, i, h).font = openpyxl.styles.Font(bold=True)
        
        for idx, row in enumerate(data, 1):
            for i, h in enumerate(use_headers, 1):
                ws.cell(idx + 1, i, row.get(h, ''))
        
        out_file = os.path.join(output_dir, f'{base}_{label}.xlsx')
        wb.save(out_file)
        print(f"{label}: {len(data)} items -> {out_file}")
    
    if unknown:
        print(f"Unclassified: {len(unknown)} items (review manually)")


def main():
    parser = argparse.ArgumentParser(description='BOM提取与合并')
    parser.add_argument('--merge', nargs='+', help='合并多个BOM文件')
    parser.add_argument('--split', help='拆分机电协同表')
    parser.add_argument('--output', default='output.xlsx', help='输出文件')
    parser.add_argument('--output-dir', default='./split_output', help='拆分输出目录')
    
    args = parser.parse_args()
    
    if args.merge:
        merge_boms(args.merge, args.output)
    elif args.split:
        split_bom(args.split, args.output_dir)
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
