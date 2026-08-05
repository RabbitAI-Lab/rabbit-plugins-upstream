#!/usr/bin/env python3
"""
BOM版本对比脚本
比较两个版本的BOM Excel文件差异，输出变更报告。

Usage:
    python compare_bom.py --old=旧BOM.xlsx --new=新BOM.xlsx --output=变更报告.xlsx
    python compare_bom.py --old=旧BOM.xlsx --new=新BOM.xlsx --sheet=电气BOM
"""
import argparse
import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("Error: openpyxl required. Install: pip install openpyxl")
    sys.exit(1)


# 颜色定义
FILL_CHANGE = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色=变更
FILL_NEW = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')     # 绿色=新增
FILL_DELETE = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 红色=删除


def read_bom(filepath, sheet_name=None):
    """读取BOM文件，返回物料字典 {(品牌, 名称, 型号): {字段: 值}}"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    
    # 找到表头行
    header_row = None
    for row_idx in range(1, min(20, ws.max_row + 1)):
        values = [str(ws.cell(row_idx, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        if any(k in values for k in ['序号', '品牌', '名称', '型号']):
            header_row = row_idx
            break
    
    if not header_row:
        return {}, []
    
    headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    
    # 找到关键列
    key_cols = {}
    for keyword in ['序号', '品牌', '名称', '型号', '规格', '数量', '单位', '物料小类']:
        for i, h in enumerate(headers):
            if keyword in h:
                key_cols[keyword] = i + 1
                break
    
    materials = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        seq = ws.cell(row_idx, key_cols.get('序号', 1)).value
        if not seq:
            continue
        
        brand = str(ws.cell(row_idx, key_cols.get('品牌', 2)).value or '').strip()
        name = str(ws.cell(row_idx, key_cols.get('名称', 3)).value or '').strip()
        model = str(ws.cell(row_idx, key_cols.get('型号', 4)).value or '').strip()
        
        key = (brand, name, model)
        materials[key] = {
            '序号': str(seq),
            '品牌': brand,
            '名称': name,
            '型号': model,
            '规格': str(ws.cell(row_idx, key_cols.get('规格', 5)).value or '').strip(),
            '数量': ws.cell(row_idx, key_cols.get('数量', 6)).value or 0,
            '单位': str(ws.cell(row_idx, key_cols.get('单位', 7)).value or '').strip(),
            '物料小类': str(ws.cell(row_idx, key_cols.get('物料小类', 8)).value or '').strip(),
        }
    
    wb.close()
    return materials, headers


def compare_bom(old_file, new_file, sheet_name=None):
    """对比两个BOM"""
    old_bom, _ = read_bom(old_file, sheet_name)
    new_bom, _ = read_bom(new_file, sheet_name)
    
    changes = []
    
    # 检查新增和变更
    for key, new_item in new_bom.items():
        if key not in old_bom:
            changes.append({
                'type': '新增',
                'item': new_item,
                'field': '',
                'old_value': '',
                'new_value': '',
            })
        else:
            old_item = old_bom[key]
            for field in ['品牌', '名称', '型号', '规格', '数量', '单位']:
                if str(old_item.get(field, '')) != str(new_item.get(field, '')):
                    changes.append({
                        'type': '变更',
                        'item': new_item,
                        'field': field,
                        'old_value': old_item.get(field, ''),
                        'new_value': new_item.get(field, ''),
                    })
    
    # 检查删除
    for key, old_item in old_bom.items():
        if key not in new_bom:
            changes.append({
                'type': '删除',
                'item': old_item,
                'field': '',
                'old_value': '',
                'new_value': '',
            })
    
    return changes


def write_report(changes, output_file):
    """输出变更报告Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'BOM变更报告'
    
    headers = ['变更类型', '序号', '物料小类', '名称', '品牌', '变更字段', '旧值', '新值', '备注']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = openpyxl.styles.Font(bold=True)
    
    for row_idx, change in enumerate(changes, 2):
        item = change['item']
        ws.cell(row_idx, 1, change['type'])
        ws.cell(row_idx, 2, item.get('序号', ''))
        ws.cell(row_idx, 3, item.get('物料小类', ''))
        ws.cell(row_idx, 4, item.get('名称', ''))
        ws.cell(row_idx, 5, item.get('品牌', ''))
        ws.cell(row_idx, 6, change['field'])
        ws.cell(row_idx, 7, change['old_value'])
        ws.cell(row_idx, 8, change['new_value'])
        
        # 颜色标记
        fill = None
        if change['type'] == '变更':
            fill = FILL_CHANGE
        elif change['type'] == '新增':
            fill = FILL_NEW
        elif change['type'] == '删除':
            fill = FILL_DELETE
        
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row_idx, col).fill = fill
    
    wb.save(output_file)
    print(f"Report saved: {output_file}")
    print(f"Total changes: {len(changes)}")


def main():
    parser = argparse.ArgumentParser(description='BOM版本对比')
    parser.add_argument('--old', required=True, help='旧版BOM文件')
    parser.add_argument('--new', required=True, help='新版BOM文件')
    parser.add_argument('--output', default='bom_changes.xlsx', help='输出报告文件')
    parser.add_argument('--sheet', help='指定Sheet名称')
    
    args = parser.parse_args()
    changes = compare_bom(args.old, args.new, args.sheet)
    write_report(changes, args.output)


if __name__ == '__main__':
    main()
