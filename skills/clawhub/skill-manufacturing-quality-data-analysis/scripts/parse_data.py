#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
制造业质量现场数据分析 - 数据解析脚本
支持 Excel、CSV、Word、PDF 格式的数据解析
自动识别关键字段：产品、工序、班次、不良类型、不良数量、日期
"""

import argparse
import json
import os
import sys
from datetime import datetime
from typing import Dict, List, Any, Tuple

try:
    import pandas as pd
    import openpyxl
    from openpyxl.utils import get_column_letter
    from docx import Document
    import PyPDF2
except ImportError as e:
    print(json.dumps({"status": "error", "message": f"缺少必要依赖库: {e}"}))
    sys.exit(1)


# 关键字段映射表（包含常见同义词）
KEY_FIELDS = {
    'product': ['产品', '产品名称', '产品型号', '型号', 'product', '产品编号'],
    'process': ['工序', '制程', '工位', 'process', '工序名称', '工序编号'],
    'shift': ['班次', '班别', 'shift', '班组'],
    'defect_type': ['不良类型', '不良现象', '缺陷类型', 'defect_type', '不良原因', '不良项目'],
    'defect_count': ['不良数量', '不良数', '缺陷数量', 'defect_count', '不良个数', '次品数'],
    'date': ['日期', '检验日期', 'date', '生产日期', '检验时间'],
    'total_count': ['总数', '检验总数', 'total_count', '样本数', '检验数量']
}


def find_field_mapping(columns: List[str]) -> Dict[str, str]:
    """
    根据列名找到关键字段映射
    """
    mapping = {}
    for key, synonyms in KEY_FIELDS.items():
        for col in columns:
            for synonym in synonyms:
                if synonym in col:
                    mapping[key] = col
                    break
            if key in mapping:
                break
    return mapping


def parse_excel(file_path: str) -> Dict[str, Any]:
    """
    解析 Excel 文件（支持多 Sheet）
    """
    result = {
        'file_type': 'excel',
        'file_name': os.path.basename(file_path),
        'sheets': [],
        'field_mapping': {},
        'issues': [],
        'requires_confirmation': False,
        'data': {}
    }

    try:
        # 使用 openpyxl 获取宏观结构信息
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames

        for sheet_name in sheet_names:
            sheet_info = {
                'name': sheet_name,
                'has_data': False,
                'header_row': None,
                'columns': [],
                'row_count': 0
            }

            # 读取前 20 行以确定表头位置
            sheet = wb[sheet_name]
            data_rows = []
            for idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
                data_rows.append(row)

            # 自动检测表头位置（查找包含最多非空值的行）
            header_row = 0
            max_non_empty = 0
            for idx, row in enumerate(data_rows):
                non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())
                if non_empty > max_non_empty:
                    max_non_empty = non_empty
                    header_row = idx

            if max_non_empty >= 2:  # 至少 2 列有数据才认为是有效表头
                sheet_info['header_row'] = header_row
                sheet_info['columns'] = [str(cell) if cell else '' for cell in data_rows[header_row]]

                # 计算总行数
                sheet_info['row_count'] = sum(1 for _ in sheet.iter_rows()) - header_row - 1
                sheet_info['has_data'] = sheet_info['row_count'] > 0

            result['sheets'].append(sheet_info)

        wb.close()

        # 使用 pandas 读取所有有效 Sheet 的数据
        valid_sheets = [s for s in result['sheets'] if s['has_data']]

        if not valid_sheets:
            result['issues'].append('未检测到有效数据 Sheet')
            result['requires_confirmation'] = True
            return result

        # 读取数据
        for sheet_info in valid_sheets:
            sheet_name = sheet_info['name']
            header_row = sheet_info['header_row']

            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header_row,
                engine='openpyxl'
            )

            # 清理列名（空值处理）
            df.columns = [str(col).strip() if pd.notna(col) else f'col_{i}' for i, col in enumerate(df.columns)]

            # 移除空行
            df = df.dropna(how='all')

            # 获取字段映射
            field_mapping = find_field_mapping(df.columns.tolist())

            if not field_mapping:
                result['issues'].append(f'Sheet [{sheet_name}] 未识别到关键字段')
                result['requires_confirmation'] = True
            else:
                # 更新全局字段映射
                result['field_mapping'].update(field_mapping)

            result['data'][sheet_name] = {
                'columns': df.columns.tolist(),
                'field_mapping': field_mapping,
                'row_count': len(df),
                'sample_data': df.head(5).to_dict(orient='records')
            }

        # 跨 Sheet 字段关联分析
        result['cross_sheet_analysis'] = analyze_cross_sheets(result)

        # 检查是否需要用户确认
        if result['cross_sheet_analysis']['needs_confirmation']:
            result['requires_confirmation'] = True
            result['issues'].append('检测到跨 Sheet 数据关联，请确认字段含义')

    except Exception as e:
        result['status'] = 'error'
        result['message'] = f'Excel 解析失败: {str(e)}'
        return result

    result['status'] = 'success'
    return result


def parse_csv(file_path: str) -> Dict[str, Any]:
    """
    解析 CSV 文件
    """
    result = {
        'file_type': 'csv',
        'file_name': os.path.basename(file_path),
        'field_mapping': {},
        'issues': [],
        'requires_confirmation': False,
        'data': {}
    }

    try:
        # 尝试不同编码
        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
        df = None
        used_encoding = None

        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, encoding=encoding)
                used_encoding = encoding
                break
            except:
                continue

        if df is None:
            result['status'] = 'error'
            result['message'] = 'CSV 文件编码无法识别，请确保为 utf-8、gbk 或 gb2312'
            return result

        # 清理列名
        df.columns = [str(col).strip() if pd.notna(col) else f'col_{i}' for i, col in enumerate(df.columns)]

        # 移除空行
        df = df.dropna(how='all')

        # 获取字段映射
        field_mapping = find_field_mapping(df.columns.tolist())

        if not field_mapping:
            result['issues'].append('未识别到关键字段')
            result['requires_confirmation'] = True
        else:
            result['field_mapping'] = field_mapping

        result['data']['sheet1'] = {
            'columns': df.columns.tolist(),
            'field_mapping': field_mapping,
            'row_count': len(df),
            'sample_data': df.head(5).to_dict(orient='records')
        }

        # 检查数据量是否足够
        if len(df) < 3:
            result['issues'].append(f'数据量过少（仅 {len(df)} 条），建议收集更多数据')
            result['requires_confirmation'] = True

    except Exception as e:
        result['status'] = 'error'
        result['message'] = f'CSV 解析失败: {str(e)}'
        return result

    result['status'] = 'success'
    return result


def parse_word(file_path: str) -> Dict[str, Any]:
    """
    解析 Word 文档中的表格
    """
    result = {
        'file_type': 'word',
        'file_name': os.path.basename(file_path),
        'field_mapping': {},
        'issues': [],
        'requires_confirmation': False,
        'data': {}
    }

    try:
        doc = Document(file_path)
        tables = doc.tables

        if not tables:
            result['status'] = 'error'
            result['message'] = 'Word 文档中未检测到表格'
            return result

        for idx, table in enumerate(tables):
            # 提取表格数据
            rows = []
            for row in table.rows:
                row_data = [cell.text.strip() for cell in row.cells]
                rows.append(row_data)

            if len(rows) < 2:
                continue

            # 假设第一行为表头
            headers = rows[0]
            data_rows = rows[1:]

            # 清理表头
            headers = [h if h else f'col_{i}' for i, h in enumerate(headers)]

            # 转换为 DataFrame
            df = pd.DataFrame(data_rows, columns=headers)
            df = df.dropna(how='all')

            # 获取字段映射
            field_mapping = find_field_mapping(headers)

            if not field_mapping:
                result['issues'].append(f'表格 {idx+1} 未识别到关键字段')
                result['requires_confirmation'] = True
            else:
                result['field_mapping'].update(field_mapping)

            result['data'][f'table_{idx+1}'] = {
                'columns': headers,
                'field_mapping': field_mapping,
                'row_count': len(df),
                'sample_data': df.head(5).to_dict(orient='records')
            }

        if not result['data']:
            result['status'] = 'error'
            result['message'] = '未提取到有效表格数据'
            return result

    except Exception as e:
        result['status'] = 'error'
        result['message'] = f'Word 解析失败: {str(e)}'
        return result

    result['status'] = 'success'
    return result


def parse_pdf(file_path: str) -> Dict[str, Any]:
    """
    解析 PDF 文档（基础文本提取）
    注意：PDF 表格解析较为复杂，建议使用 Excel 或 Word 格式
    """
    result = {
        'file_type': 'pdf',
        'file_name': os.path.basename(file_path),
        'field_mapping': {},
        'issues': [],
        'requires_confirmation': False,
        'data': {}
    }

    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ''

            for page in pdf_reader.pages:
                text += page.extract_text()

            # 简单的表格提取（基于制表符或空格对齐）
            lines = text.split('\n')
            table_data = []
            for line in lines:
                if '|' in line or '\t' in line:
                    # 尝试按分隔符分割
                    if '|' in line:
                        row = [cell.strip() for cell in line.split('|')]
                    else:
                        row = [cell.strip() for cell in line.split('\t')]
                    if len(row) > 1:
                        table_data.append(row)

            if len(table_data) < 2:
                result['status'] = 'warning'
                result['message'] = 'PDF 表格解析较为复杂，建议使用 Excel 或 Word 格式'
                result['issues'].append('PDF 格式建议转换为 Excel 或 Word 以获得更好的解析效果')
                result['requires_confirmation'] = True
                return result

            # 假设第一行为表头
            headers = table_data[0]
            data_rows = table_data[1:]

            headers = [h if h else f'col_{i}' for i, h in enumerate(headers)]
            df = pd.DataFrame(data_rows, columns=headers)
            df = df.dropna(how='all')

            field_mapping = find_field_mapping(headers)

            if not field_mapping:
                result['issues'].append('未识别到关键字段')
                result['requires_confirmation'] = True
            else:
                result['field_mapping'] = field_mapping

            result['data']['table1'] = {
                'columns': headers,
                'field_mapping': field_mapping,
                'row_count': len(df),
                'sample_data': df.head(5).to_dict(orient='records')
            }

    except Exception as e:
        result['status'] = 'error'
        result['message'] = f'PDF 解析失败: {str(e)}'
        return result

    result['status'] = 'success'
    return result


def analyze_cross_sheets(parse_result: Dict[str, Any]) -> Dict[str, Any]:
    """
    分析跨 Sheet 数据关联
    """
    analysis = {
        'needs_confirmation': False,
        'related_fields': {},
        'suggestions': []
    }

    data_sheets = {k: v for k, v in parse_result['data'].items() if v['row_count'] > 0}

    if len(data_sheets) <= 1:
        return analysis

    # 检查关键字段在不同 Sheet 中的分布
    all_fields = {}
    for sheet_name, sheet_data in data_sheets.items():
        for key, col_name in sheet_data['field_mapping'].items():
            if key not in all_fields:
                all_fields[key] = []
            all_fields[key].append((sheet_name, col_name))

    # 检查是否有冲突或需要确认的字段
    for key, locations in all_fields.items():
        if len(locations) > 1:
            analysis['related_fields'][key] = locations
            analysis['suggestions'].append(
                f"字段 [{key}] 在多个 Sheet 中出现: {', '.join([f'{s}[{c}]' for s, c in locations])}"
            )
            analysis['needs_confirmation'] = True

    return analysis


def main():
    parser = argparse.ArgumentParser(description='制造业质量现场数据分析 - 数据解析')
    parser.add_argument('--file_path', required=True, help='数据文件路径')
    args = parser.parse_args()

    if not os.path.exists(args.file_path):
        print(json.dumps({
            "status": "error",
            "message": f"文件不存在: {args.file_path}"
        }, ensure_ascii=False))
        sys.exit(1)

    file_ext = os.path.splitext(args.file_path)[1].lower()

    if file_ext in ['.xlsx', '.xls']:
        result = parse_excel(args.file_path)
    elif file_ext == '.csv':
        result = parse_csv(args.file_path)
    elif file_ext in ['.docx', '.doc']:
        result = parse_word(args.file_path)
    elif file_ext == '.pdf':
        result = parse_pdf(args.file_path)
    else:
        print(json.dumps({
            "status": "error",
            "message": f"不支持的文件格式: {file_ext}，支持的格式: Excel, CSV, Word, PDF"
        }, ensure_ascii=False))
        sys.exit(1)

    # 输出结果
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
