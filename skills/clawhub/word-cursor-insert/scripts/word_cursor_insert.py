#!/usr/bin/env python3
# word_cursor_insert.py
# Insert tabular data (CSV or Excel — same thing: a referenced table file) or
# plain text into the cursor of the currently open Word document, context-aware
# and fast.
#
# WHY PYTHON (not PowerShell): in this environment PowerShell COM stdout is
# garbled (only the console title + BEL echo shows), so results can't be read
# back directly. Python stdout is clean. Also, cursor context is read with a
# BOUNDED window (never doc.Range(0, sel.Start)) so cost is O(window),
# independent of document size.
#
# Usage:
#   Insert a table at the cursor (Excel or CSV — auto-detected by extension):
#       word_cursor_insert.py 表.xlsx [--no-title]
#       word_cursor_insert.py 表.csv  [--no-title]
#   Insert plain text inline at cursor (e.g. fill a blank in a sentence):
#       word_cursor_insert.py --text "7" [--newline]   # short values are
#                                                    # idempotent: re-running
#                                                    # will NOT duplicate them
#                                                    # (use --no-idempotent to force)
#
# Requirements: pywin32 installed in the managed venv.
# openpyxl is needed only when the table source is an .xlsx/.xls file.

import sys
import csv
import win32com.client

WIN = 300  # chars of context to read before/after the cursor


def read_csv(path):
    for enc in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            with open(path, encoding=enc, newline="") as f:
                return list(csv.reader(f))
        except Exception:
            continue
    raise IOError("cannot decode csv: " + path)


def read_xlsx_sheet(path, sheet=None):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def read_table(path):
    """Read a tabular file (.csv / .xls / .xlsx) into a list of rows.
    Excel and CSV are the same operation — only the parser differs."""
    if path.lower().endswith((".xlsx", ".xls", ".xlsm")):
        return read_xlsx_sheet(path)
    return read_csv(path)


def connect_word():
    return win32com.client.GetActiveObject("Word.Application")


def context_report(doc, sel):
    start = sel.Start
    end = sel.End
    before = doc.Range(max(start - WIN, 0), start).Text if start > 0 else ""
    doc_end = doc.Content.End
    after = doc.Range(end, min(end + WIN, doc_end)).Text if end < doc_end else ""
    in_table = bool(sel.Information(12))  # wdWithInTable = 12
    print("CONTEXT before:", repr(before[-200:]))
    print("CONTEXT after :", repr(after[:200]))
    print("IN_TABLE:", in_table)
    return in_table


def insert_text(doc, sel, text, newline, idempotent=True):
    # Bounded near-cursor idempotency (one COM read, O(window)): skip only if the
    # EXACT text already sits within ~150 chars of the cursor.
    #   - Long text: exact match is safe (a unique paragraph won't falsely match).
    #   - Short text (<=8): require a word boundary so "7" won't match inside "177".
    # Use --no-idempotent to force insertion regardless.
    if idempotent and text:
        W = 150
        s, e = sel.Start, sel.End
        dend = doc.Content.End
        before = doc.Range(max(s - W, 0), s).Text if s > 0 else ""
        after = doc.Range(e, min(e + W, dend)).Text if e < dend else ""
        near = before + after
        dup = False
        if text in near:
            if len(text) <= 8:
                for m in range(len(near) - len(text) + 1):
                    if near[m:m + len(text)] == text:
                        prev = near[m - 1] if m > 0 else ""
                        nxt = near[m + len(text)] if m + len(text) < len(near) else ""
                        if not (prev.isalnum() and nxt.isalnum()):
                            dup = True
                            break
            else:
                dup = True
        if dup:
            print("SKIP text already present near cursor (%d字, 幂等跳过)" % len(text))
            return
    if newline:
        sel.TypeParagraph()
    sel.TypeText(text)
    print("OK text inserted at cursor (%d字)" % len(text))


def insert_table(doc, sel, data, use_title):
    first = data[0]
    nonempty = [c for c in first if str(c).strip()]
    if use_title and len(nonempty) == 1 and len(data) > 2:
        title = nonempty[0].strip()
        header = data[1]
        body = data[2:]
    else:
        title = None
        header = data[0]
        body = data[1:]

    if title:
        sel.TypeParagraph()
        sel.TypeText(title)
        doc.Paragraphs(doc.Paragraphs.Count).Range.Font.Bold = True
    sel.TypeParagraph()
    tbl = doc.Tables.Add(sel.Range, len(body) + 1, len(header))
    for c, h in enumerate(header):
        tbl.Cell(1, c + 1).Range.Text = str(h)
    for r, row in enumerate(body):
        for c in range(len(header)):
            val = row[c] if c < len(row) else ""
            tbl.Cell(r + 2, c + 1).Range.Text = str(val)
    tbl.Rows(1).Range.Font.Bold = True
    tbl.Borders.InsideLineStyle = 1   # wdLineStyleSingle
    tbl.Borders.OutsideLineStyle = 1
    tbl.AutoFitBehavior(2)            # wdAutoFitWindow
    print("OK table title=%r rows=%d cols=%d" % (title, tbl.Rows.Count, tbl.Columns.Count))
    if body:
        print("FIRST:", " | ".join(str(x) for x in body[0]))
        print("LAST :", " | ".join(str(x) for x in body[-1]))


def main():
    args = sys.argv[1:]
    if not args:
        print('USAGE: word_cursor_insert.py <table.xlsx|table.csv> [--no-title] | '
              '--text "..." [--newline]')
        sys.exit(2)

    text_mode = None
    newline = False
    no_title = False
    no_idem = False
    table_path = None
    i = 0
    while i < len(args):
        a = args[i]
        if a == "--text":
            text_mode = args[i + 1]
            i += 2
        elif a == "--file":
            with open(args[i + 1], encoding="utf-8") as _f:
                text_mode = _f.read().rstrip("\r\n")
            i += 2
        elif a == "--newline":
            newline = True
            i += 1
        elif a == "--no-title":
            no_title = True
            i += 1
        elif a == "--no-idempotent":
            no_idem = True
            i += 1
        else:
            table_path = a
            i += 1

    word = connect_word()
    doc = word.ActiveDocument
    sel = word.Selection

    if text_mode is not None:
        context_report(doc, sel)
        insert_text(doc, sel, text_mode, newline, idempotent=not no_idem)
    elif table_path:
        data = read_table(table_path)
        insert_table(doc, sel, data, not no_title)
    else:
        print("ERROR: provide a table file path or --text")
        sys.exit(2)


if __name__ == "__main__":
    main()
