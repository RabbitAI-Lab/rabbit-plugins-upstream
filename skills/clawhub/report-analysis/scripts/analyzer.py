# -*- coding: utf-8 -*-
"""《评优激励通报》核心分析模块：统一数据口径，供 3 个功能复用。
功能A：具体内容查询   功能B：具体网格/渠道查询   功能C：总结分析(Word)
"""
import openpyxl
from collections import defaultdict, Counter

import os
SRC = os.environ.get("YJ_SRC", "")  # 必须通过环境变量 YJ_SRC 指定 Excel 路径（本地无默认值）
MAIN_SHEET = os.environ.get("YJ_SHEET", "")  # 留空则自动匹配「渠道完成情况通报」工作表
COUNTY_SHEET = os.environ.get("YJ_COUNTY_SHEET", "")  # 留空则自动匹配含「县」且含「汇总」的工作表


def resolve_sheet(wb, prefer, fuzzy):
    if prefer and prefer in wb.sheetnames:
        return prefer
    for name in wb.sheetnames:
        if all(k in name for k in fuzzy):
            return name
    raise KeyError(f"未找到工作表（fuzzy={fuzzy}，现有：{wb.sheetnames}）")


def _wb():
    if not SRC:
        raise SystemExit("未指定数据文件：请通过环境变量 YJ_SRC 传入《评优激励通报》Excel 路径，"
                         "如 YJ_SRC=D:/xx/2026年3季度合作伙伴评优激励通报20260811.xlsx")
    return openpyxl.load_workbook(SRC, data_only=True, read_only=True)

# 主表列索引（1-based，与技能文档一致）
C = dict(
    county=1, grid=2, code3=3, code4=4, name=5, photo=6, raw_bean=7,
    raw_amt=8, gate_amt=9, term_amt=10, focus_amt=11, complaint_amt=12,
    weakgrid_amt=13, final_amt=14,
    tiger=15, tiger_ratio=16, ai5=17, rights_up=18, member88=19, gate_level=20,
    term_qty=21, term_contract=22, term_ratio=24, term_coef=26,
    stock_pts=27, prod_pts=28, focus_coef=31,
    complaint_photo=32, complaint_val=33, complaint_pnl=34,
    weakgrid_sign=35, weakgrid_done=36, weakgrid_pnl=37,
    newnet_qty=38, newnet_fuse=39, newnet_fuse_ratio=40,
    newterm_qty=41, newterm_fuse=42, newterm_fuse_ratio=43,
    bb_qty=44, bb_fuse=45, bb_fuse_ratio=46, app_coef=47,
    star=90,
)


def load_rows():
    wb = _wb()
    sheet = resolve_sheet(wb, MAIN_SHEET, ["渠道完成情况通报"])
    ws = wb[sheet]
    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r[0] is None or str(r[0]).strip() == "":
            continue
        rows.append(r)
    return rows


def num(v):
    """安全转数字"""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "/", "-", "--", "未出数，先按不扣罚计", "未出数"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def load_analyzed():
    """返回渠道级明细列表 + 汇总统计"""
    rows = load_rows()
    chans = []
    for r in rows:
        raw = num(r[C["raw_amt"] - 1])
        final = num(r[C["final_amt"] - 1])
        loss = raw - final
        chan = dict(
            county=str(r[C["county"] - 1]).strip(),
            grid=str(r[C["grid"] - 1]).strip(),
            code=str(r[C["code3"] - 1]).strip(),
            name=str(r[C["name"] - 1]).strip(),
            raw=raw, final=final, loss=loss,
            loss_rate=(loss / raw * 100) if raw > 0 else 0.0,
            gate_amt=num(r[C["gate_amt"] - 1]),
            term_amt=num(r[C["term_amt"] - 1]),
            focus_amt=num(r[C["focus_amt"] - 1]),
            weakgrid_amt=num(r[C["weakgrid_amt"] - 1]),
            gate_level=str(r[C["gate_level"] - 1]).strip(),
            tiger=num(r[C["tiger"] - 1]),
            ai5=num(r[C["ai5"] - 1]),
            rights_up=num(r[C["rights_up"] - 1]),
            member88=num(r[C["member88"] - 1]),
            term_ratio=num(r[C["term_ratio"] - 1]),
            focus_coef=num(r[C["focus_coef"] - 1]),
            weakgrid_sign=str(r[C["weakgrid_sign"] - 1]).strip(),
            weakgrid_done=num(r[C["weakgrid_done"] - 1]),
            newnet=num(r[C["newnet_qty"] - 1]),
            newnet_fuse=num(r[C["newnet_fuse_ratio"] - 1]) * 100,
            newterm=num(r[C["newterm_qty"] - 1]),
            newterm_fuse=num(r[C["newterm_fuse_ratio"] - 1]) * 100,
            bb=num(r[C["bb_qty"] - 1]),
            bb_fuse=num(r[C["bb_fuse_ratio"] - 1]) * 100,
            app_coef=num(r[C["app_coef"] - 1]),
            star=num(r[C["star"] - 1]),
            term_qty=num(r[C["term_qty"] - 1]),
            term_contract=num(r[C["term_contract"] - 1]),
            stock_pts=num(r[C["stock_pts"] - 1]),
            prod_pts=num(r[C["prod_pts"] - 1]),
            complaint_val=r[C["complaint_val"] - 1],
        )
        chans.append(chan)
    return chans


def summarize(chans):
    """总体统计：9 类欠缺指标"""
    n = len(chans)
    s = dict(n=n)
    s["raw_total"] = sum(c["raw"] for c in chans)
    s["final_total"] = sum(c["final"] for c in chans)
    s["loss_total"] = s["raw_total"] - s["final_total"]
    s["loss_rate"] = s["loss_total"] / s["raw_total"] * 100 if s["raw_total"] else 0

    zero_gate = [c for c in chans if c["gate_level"] == "门槛未完成"]
    s["gate_notdone"] = len(zero_gate)
    s["gate_notdone_rate"] = len(zero_gate) / n * 100

    s["zero_ai5"] = sum(1 for c in chans if c["ai5"] == 0)
    s["zero_tiger"] = sum(1 for c in chans if c["tiger"] == 0)
    s["zero_member88"] = sum(1 for c in chans if c["member88"] == 0)
    s["zero_term"] = sum(1 for c in chans if c["term_ratio"] == 0)
    s["avg_term"] = sum(c["term_ratio"] for c in chans) / n * 100 if n else 0

    s["avg_newnet_fuse"] = sum(c["newnet_fuse"] for c in chans) / n
    s["avg_newterm_fuse"] = sum(c["newterm_fuse"] for c in chans) / n
    s["avg_bb_fuse"] = sum(c["bb_fuse"] for c in chans) / n

    s["coef081"] = sum(1 for c in chans if abs(c["focus_coef"] - 0.81) < 0.001)
    s["coef09"] = sum(1 for c in chans if abs(c["focus_coef"] - 0.9) < 0.001)
    s["coef1"] = sum(1 for c in chans if abs(c["focus_coef"] - 1.0) < 0.001)

    s["signed"] = sum(1 for c in chans if c["weakgrid_sign"] == "是")
    s["signed_zero_done"] = sum(1 for c in chans if c["weakgrid_sign"] == "是" and c["weakgrid_done"] == 0)

    s["zero_final"] = sum(1 for c in chans if c["final"] == 0)
    s["zero_final_rate"] = s["zero_final"] / n * 100

    s["complaint_undone"] = sum(1 for c in chans if c["complaint_val"] == 0)  # 未出数占位
    s["counties"] = Counter(c["county"] for c in chans)
    return s


def by_county(chans):
    d = defaultdict(list)
    for c in chans:
        d[c["county"]].append(c)
    out = []
    for county, lst in d.items():
        raw = sum(x["raw"] for x in lst)
        final = sum(x["final"] for x in lst)
        out.append(dict(
            county=county, n=len(lst), raw=raw, final=final, loss=raw - final,
            loss_rate=(raw - final) / raw * 100 if raw else 0,
            gate_notdone=sum(1 for x in lst if x["gate_level"] == "门槛未完成"),
            zero_final=sum(1 for x in lst if x["final"] == 0),
            avg_term=sum(x["term_ratio"] for x in lst) / len(lst) * 100,
        ))
    out.sort(key=lambda x: -x["loss"])
    return out


def by_grid(chans):
    d = defaultdict(list)
    for c in chans:
        d[(c["county"], c["grid"])].append(c)
    out = []
    for (county, grid), lst in d.items():
        raw = sum(x["raw"] for x in lst)
        final = sum(x["final"] for x in lst)
        out.append(dict(
            county=county, grid=grid, n=len(lst), raw=raw, final=final, loss=raw - final,
            loss_rate=(raw - final) / raw * 100 if raw else 0,
            gate_notdone=sum(1 for x in lst if x["gate_level"] == "门槛未完成"),
            zero_final=sum(1 for x in lst if x["final"] == 0),
            chans=lst,
        ))
    out.sort(key=lambda x: -x["loss"])
    return out


def load_county_summary():
    """从县（市）汇总表读取权威的逐道损收（万元），返回 {县名: dict, '全州': dict}
    keys: raw, gate, gate_loss, term, term_loss, focus, focus_loss,
          complaint, complaint_loss, weakgrid, weakgrid_loss, app, app_loss
    """
    wb = _wb()
    sheet = resolve_sheet(wb, COUNTY_SHEET, ["县", "汇总"])
    ws = wb[sheet]
    out = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r[0] is None:
            continue
        out[str(r[0]).strip()] = dict(
            raw=num(r[1]), gate=num(r[2]), gate_loss=num(r[3]),
            term=num(r[4]), term_loss=num(r[5]),
            focus=num(r[6]), focus_loss=num(r[7]),
            complaint=num(r[8]), complaint_loss=num(r[9]),
            weakgrid=num(r[10]), weakgrid_loss=num(r[11]),
            app=num(r[12]), app_loss=num(r[13]),
        )
    return out


def fmt_wan(v):
    if round(v) == 0:
        return "0元"
    return f"{v / 10000:.2f}万" if abs(v) >= 10000 else f"{v:,.0f}元"


if __name__ == "__main__":
    chans = load_analyzed()
    s = summarize(chans)
    print(f"渠道总数: {s['n']}")
    print(f"原始金额合计: {fmt_wan(s['raw_total'])}  最终金额: {fmt_wan(s['final_total'])}")
    print(f"总损失: {fmt_wan(s['loss_total'])}  损失率: {s['loss_rate']:.1f}%")
    print(f"门槛未完成: {s['gate_notdone']} 家 ({s['gate_notdone_rate']:.1f}%)")
    print(f"AI五件套为0: {s['zero_ai5']}  金虎为0: {s['zero_tiger']}  88会员为0: {s['zero_member88']}")
    print(f"终端合约率为0: {s['zero_term']}  均值: {s['avg_term']:.1f}%")
    print(f"APP融合率均值 新入网:{s['avg_newnet_fuse']:.1f}% 新终端:{s['avg_newterm_fuse']:.1f}% 宽带:{s['avg_bb_fuse']:.1f}%")
    print(f"牵引系数 0.81:{s['coef081']}  0.9:{s['coef09']}  1.0:{s['coef1']}")
    print(f"弱势网格签约: {s['signed']}  签约后4项为0: {s['signed_zero_done']}")
    print(f"最终为0元渠道: {s['zero_final']} ({s['zero_final_rate']:.1f}%)")
    print("\n=== 按县 ===")
    for x in by_county(chans):
        print(f"  {x['county']}: {x['n']}家 原始{fmt_wan(x['raw'])} 最终{fmt_wan(x['final'])} "
              f"损失{fmt_wan(x['loss'])}({x['loss_rate']:.0f}%) 门槛未完成{x['gate_notdone']} 0元{x['zero_final']}")
    print("\n=== 损失TOP网格 ===")
    for x in by_grid(chans)[:8]:
        print(f"  {x['county']}-{x['grid']}: {x['n']}家 损失{fmt_wan(x['loss'])} 门槛未完成{x['gate_notdone']}")
