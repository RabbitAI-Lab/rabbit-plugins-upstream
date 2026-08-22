# -*- coding: utf-8 -*-
'''《评优激励通报》预警分级 + 名单导出（功能D）
分级规则（数据驱动）：
  红牌：门槛未完成 且 四项门槛业务全为0（完全无业务）；或 最终金额为负（扣罚超可发额）
  黄牌：门槛未完成 但有业务量（有量未达档）
  绿牌：门槛达档（达标渠道）
导出 Excel：全渠道分级 / 红牌重点整改名单 / 黄牌名单 / 县市分级汇总
用法：python gen_warn.py [输出路径]
'''
import os
import sys
from collections import defaultdict

from analyzer import load_analyzed, fmt_wan

OUT = os.environ.get('YJ_WARN_OUT', '') or (
    r'D:/AI/WorkBuddy/2026-08-20-16-02-35/analysis/2026年3季度评优激励-预警分级名单.xlsx')


def level_of(c):
    '''返回 (等级, 说明)'''
    if c['gate_level'] != '门槛未完成':
        return '绿牌', '达标渠道'
    if c['final'] < 0:
        return '红牌', '最终金额为负，扣罚超可发额'
    biz = c['tiger'] + c['ai5'] + c['rights_up'] + c['member88']
    if biz <= 0:
        return '红牌', '四项门槛业务全为0，完全无业务'
    return '黄牌', '有业务量但未达门槛档位线'


def gen_rows(chans):
    rows = []
    for c in chans:
        lv, note = level_of(c)
        rows.append(dict(
            county=c['county'], grid=c['grid'], code=c['code'], name=c['name'],
            level=lv, note=note,
            raw=c['raw'], final=c['final'], loss=c['loss'],
            loss_rate=c['loss_rate'],
            gate=c['gate_level'],
            tiger=c['tiger'], ai5=c['ai5'], rights=c['rights_up'], m88=c['member88'],
            term_ratio=c['term_ratio'] * 100, app_coef=c['app_coef'],
        ))
    order = {'红牌': 0, '黄牌': 1, '绿牌': 2}
    rows.sort(key=lambda r: (order[r['level']], -r['loss']))
    return rows


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LV_FILL = {
    '红牌': PatternFill('solid', fgColor='FFC7CE'),
    '黄牌': PatternFill('solid', fgColor='FFEB9C'),
    '绿牌': PatternFill('solid', fgColor='C6EFCE'),
}
HEAD_FILL = PatternFill('solid', fgColor='305496')
HEAD_FONT = Font(color='FFFFFF', bold=True, size=10)
BODY_FONT = Font(size=10)
THIN = Border(*[Side(style='thin', color='D9D9D9')] * 4)


def style_sheet(ws, headers, widths, level_col=None):
    for j, h in enumerate(headers, 1):
        cell = ws.cell(1, j, h)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
    ws.freeze_panes = 'A2'


COL_KEY = {  # 中文列名 -> 数据键
    '县': 'county', '网格': 'grid', '渠道编码': 'code', '渠道名称': 'name',
    '预警': 'level', '分级说明': 'note', '原始金额': 'raw', '最终金额': 'final',
    '损收': 'loss', '损收率': 'loss_rate', '门槛': 'gate',
    '金虎': 'tiger', 'AI5': 'ai5', '权益': 'rights', '88会员': 'm88',
    '合约率': 'term_ratio', 'APP系数': 'app_coef',
}


def write_sheet(ws, headers, data, level_col=None):
    style_sheet(ws, headers, [12, 16, 12, 20, 8, 26, 12, 12, 12, 9, 9, 9, 9, 9, 9, 9, 9, 9])
    for i, r in enumerate(data, 2):
        for j, h in enumerate(headers, 1):
            k = COL_KEY[h]
            v = r[k]
            if k in ('raw', 'final', 'loss'):
                v = round(v)
            elif k in ('loss_rate', 'term_ratio'):
                v = round(v, 1)
            cell = ws.cell(i, j, v)
            cell.font = BODY_FONT
            cell.border = THIN
            if k == 'loss_rate':
                cell.number_format = '0.0"%"'
            if k == 'term_ratio':
                cell.number_format = '0.0"%"'
            if level_col and j == level_col:
                cell.fill = LV_FILL.get(r['level'], PatternFill())
                cell.alignment = Alignment(horizontal='center')
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(data) + 1}'


def build():
    chans = load_analyzed()
    rows = gen_rows(chans)

    wb = Workbook()
    # Sheet1 全渠道分级
    ws1 = wb.active
    ws1.title = '全渠道分级'
    headers1 = ['县', '网格', '渠道编码', '渠道名称', '预警', '分级说明', '原始金额', '最终金额',
                '损收', '损收率', '门槛', '金虎', 'AI5', '权益', '88会员', '合约率', 'APP系数']
    write_sheet(ws1, headers1, rows, level_col=5)

    # Sheet2/3 红黄牌
    for lv, title in (('红牌', '红牌-重点整改'), ('黄牌', '黄牌-关注帮扶')):
        ws = wb.create_sheet(title)
        sub = [r for r in rows if r['level'] == lv]
        write_sheet(ws, headers1, sub, level_col=5)

    # Sheet4 县市分级汇总
    ws4 = wb.create_sheet('县市分级汇总')
    agg = defaultdict(lambda: {'红牌': 0, '黄牌': 0, '绿牌': 0})
    for r in rows:
        agg[r['county']][r['level']] += 1
    headers4 = ['县', '红牌', '黄牌', '绿牌', '合计', '红牌占比']
    style_sheet(ws4, headers4, [14, 10, 10, 10, 10, 12])
    order = {'红牌': 0, '黄牌': 1, '绿牌': 2}
    i = 2
    for county, cnt in sorted(agg.items()):
        total = sum(cnt.values())
        ws4.cell(i, 1, county)
        for j, lv in enumerate(('红牌', '黄牌', '绿牌'), 2):
            cell = ws4.cell(i, j, cnt[lv])
            cell.fill = LV_FILL[lv]
            cell.font = BODY_FONT
            cell.border = THIN
        ws4.cell(i, 5, total).font = BODY_FONT
        ws4.cell(i, 5).border = THIN
        cell = ws4.cell(i, 6, round(cnt['红牌'] / total * 100, 1))
        cell.number_format = '0.0"%"'
        cell.font = BODY_FONT
        cell.border = THIN
        i += 1
    # 合计行
    ws4.cell(i, 1, '合计')
    tot = {lv: sum(agg[c][lv] for c in agg) for lv in ('红牌', '黄牌', '绿牌')}
    for j, lv in enumerate(('红牌', '黄牌', '绿牌'), 2):
        ws4.cell(i, j, tot[lv])
    ws4.cell(i, 5, sum(tot.values()))
    ws4.cell(i, 6, round(tot['红牌'] / sum(tot.values()) * 100, 1))
    for j in range(1, 7):
        ws4.cell(i, j).font = Font(bold=True, size=10)

    wb.save(OUT)
    return rows


def main():
    rows = build()
    cnt = defaultdict(int)
    for r in rows:
        cnt[r['level']] += 1
    print(f'已导出: {OUT}')
    print(f'  红牌 {cnt["红牌"]} 家 | 黄牌 {cnt["黄牌"]} 家 | 绿牌 {cnt["绿牌"]} 家')
    red = [r for r in rows if r['level'] == '红牌']
    print('  红牌TOP5（按损收降序）:')
    for r in red[:5]:
        print(f'    {r["county"]}-{r["grid"]} {r["name"]} 损失{fmt_wan(r["loss"])} [{r["note"]}]')
    return rows


if __name__ == '__main__':
    main()
