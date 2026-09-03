"""Write the reconciliation Excel workbook (one row per reference)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from common import clean_text

HEADERS = [
    "序号", "来源类型", "原文链接/文件名", "原始提供的APA信息",
    "标题", "作者", "年份", "期刊/来源", "卷/期/页码", "DOI",
    "文献类型", "APA引用（第7版）", "核对结果", "修正说明/备注",
]

WIDTHS = [6, 12, 40, 40, 40, 28, 8, 24, 14, 26, 14, 55, 12, 40]


def _authors_str(rec):
    names = []
    for a in rec.get("authors", []):
        names.append(a.get("name", ""))
    return "; ".join(n for n in names if n)


def records_to_excel(records, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "文献对账表"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    check_fills = {
        "已修正": PatternFill("solid", fgColor="FFF2CC"),
        "原样正确": PatternFill("solid", fgColor="E2EFDA"),
        "待人工确认": PatternFill("solid", fgColor="FCE4D6"),
    }

    for r, rec in enumerate(records, start=2):
        vol_issue_page = ""
        if rec.get("volume"):
            vol_issue_page += clean_text(rec["volume"])
            if rec.get("issue"):
                vol_issue_page += f"({rec['issue']})"
        if rec.get("pages"):
            vol_issue_page += f", {clean_text(rec['pages'])}"

        orig_ref = ""
        if rec.get("original_url"):
            orig_ref = rec["original_url"]
        elif rec.get("original_filename"):
            orig_ref = rec["original_filename"]

        row = [
            rec.get("id", r - 1),
            rec.get("source_type", ""),
            orig_ref,
            rec.get("original_apa", "") or "",
            rec.get("title", "") or "",
            _authors_str(rec),
            rec.get("year", "") or "",
            rec.get("journal", "") or "",
            vol_issue_page,
            rec.get("doi", "") or "",
            rec.get("ref_type_name", "") or "",
            rec.get("apa", "") or "",
            rec.get("check_result", ""),
            "；".join(rec.get("notes", [])),
        ]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (3, 4, 5, 8, 12, 14)))
        cr = rec.get("check_result")
        if cr in check_fills:
            ws.cell(row=r, column=13).fill = check_fills[cr]

    ws.freeze_panes = "A2"
    wb.save(path)
    return path
