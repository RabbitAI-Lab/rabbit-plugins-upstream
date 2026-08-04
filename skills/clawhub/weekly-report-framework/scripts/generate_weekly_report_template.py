#!/usr/bin/env python3
"""
装备所周报模板生成器 V2.0
支持生成3个研究室的周报模板：机电系统(ME)、工业视觉(IV)、物流与自动化(LA)

用法:
  python generate_weekly_report_template.py                  # 生成全部3个研究室
  python generate_weekly_report_template.py --lab=IV         # 仅生成工业视觉
  python generate_weekly_report_template.py --lab=ME,LA      # 生成机电+物流
  python generate_weekly_report_template.py --list           # 列出可用研究室
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
import argparse

# 研究室配置
LAB_CONFIG = {
    "ME": {
        "name": "机电系统研究室",
        "prefix": "ME",
        "output_dir": "knowledge/项目周报知识库_机电系统",
    },
    "IV": {
        "name": "工业视觉研究室",
        "prefix": "IV",
        "output_dir": "knowledge/项目周报知识库_工业视觉",
    },
    "LA": {
        "name": "物流与自动化研究室",
        "prefix": "LA",
        "output_dir": "knowledge/项目周报知识库_物流与自动化",
    },
}


def calculate_week_range():
    """计算本周日期范围（周一至周日）"""
    today = datetime.now()
    weekday = today.weekday()  # 0=周一
    monday = today - timedelta(days=weekday)
    sunday = monday + timedelta(days=6)
    return monday.strftime("%Y-%m-%d"), sunday.strftime("%Y-%m-%d"), sunday.strftime("%Y%m%d")


def create_report_sheet(ws, sheet_name, room_name, prefix):
    """创建周报Sheet"""
    monday, sunday, _ = calculate_week_range()

    # 表头12列
    headers = [
        "序号", "项目", "状态", "项目经理", "本周成员输出",
        "本周进度", "本周洞察", "下周计划", "上周进度",
        "风险/阻塞", "备注", "信息来源"
    ]

    # 列宽设置
    col_widths = [10, 28, 14, 12, 32, 45, 32, 32, 45, 22, 15, 22]

    # 第1行：大标题
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = f"{room_name}周报"
    title_cell.font = Font(name='微软雅黑', size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 第2行：汇报周期
    ws.merge_cells('A2:L2')
    period_cell = ws['A2']
    period_cell.value = f"汇报周期：{monday} 至 {sunday}"
    period_cell.font = Font(name='微软雅黑', size=11)
    period_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25

    # 第3行：列头
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name='微软雅黑', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill
        cell.border = header_border

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths[col_idx - 1]

    ws.row_dimensions[3].height = 35

    # 示例数据行（第4行）
    example_data = [
        f"{prefix}2601",  # 序号（项目编码）
        "示例项目",  # 项目
        "设备装配",  # 状态
        "张三",     # 项目经理
        "智研院：\n张三：完成算法优化\n\n事业部：\n李四：提供测试样品",  # 本周成员输出
        f"实际情况：设备装配阶段。完成模组安装。\n本周进展：\n1、7.8 模组安装完成（已完成）\n2、7.9 调试进行中（进行中）",  # 本周进度
        "1、进度符合预期\n2、调试需加快",  # 本周洞察
        "1、7.15 完成调试\n2、7.20 开始联调",  # 下周计划
        "实际情况：设备装配阶段。\n本周进展：\n1、7.1 机械结构安装完成（已完成）",  # 上周进度
        "调试进度延迟2天",  # 风险/阻塞
        "",  # 备注
        "项目群聊"  # 信息来源
    ]

    data_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, value in enumerate(example_data, 1):
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.font = Font(name='微软雅黑', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = data_border

    # 序号列和项目列居中
    ws.cell(row=4, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=4, column=2).alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[4].height = 120

    # 说明行（第5行）
    note_cell = ws.cell(row=5, column=1, value="※ 此为示例数据，请删除后填写实际项目信息")
    note_cell.font = Font(name='微软雅黑', size=9, color='FF0000')
    ws.merge_cells('A5:L5')

    return ws


def create_instructions_sheet(ws, room_name, prefix):
    """创建填表说明Sheet"""
    ws.column_dimensions['A'].width = 80

    instructions = [
        (f"{room_name}周报填表说明", True, 14),
        ("", False, 10),
        ("一、项目编号规则", True, 12),
        (f"已立项项目：{prefix}+年份(2位)+2位数字，如 {prefix}2601、{prefix}2602", False, 10),
        ("未立项项目：年份(2位)+3位数字，如 26001、26002", False, 10),
        ("", False, 10),
        ("二、项目状态（11阶段+3特殊）", True, 12),
        ("正常阶段：前期调研→风险评估→方案设计→项目立项→详细设计→采购加工→设备装配→设备调试→现场交付→项目结项→项目售后", False, 10),
        ("特殊状态：暂停、终止、取消", False, 10),
        ("", False, 10),
        ("三、12列字段说明", True, 12),
        ("A列-序号：项目编码（如IV2601），不是纯数字", False, 10),
        ("B列-项目：项目名称，垂直居中+水平居中", False, 10),
        ("C列-状态：项目当前阶段", False, 10),
        ("D列-项目经理：项目负责人姓名", False, 10),
        ("E列-本周成员输出：按归属分组（智研院→空行→事业部&工厂）", False, 10),
        ("F列-本周进度：实际情况+本周进展+问题点（序号1、，同一天用；合并）", False, 10),
        ("G列-本周洞察：上周计划核对→建议", False, 10),
        ("H列-下周计划：AI预判+具体日期节点", False, 10),
        ("I列-上周进度：从上周周报复制", False, 10),
        ("J列-风险/阻塞：自动提取", False, 10),
        ("K列-备注：留空", False, 10),
        ("L列-信息来源：群聊名称", False, 10),
        ("", False, 10),
        ("四、F列格式示例", True, 12),
        ("实际情况：设备调试阶段。项目整体状态描述。", False, 10),
        ("本周进展：", False, 10),
        ("1、7.8 焊接参数优化完成，良率提升至85%（已完成）", False, 10),
        ("2、7.9 移载机构精度验收通过，偏差±0.3mm（已完成）", False, 10),
        ("问题点：", False, 10),
        ("1、双排成像仍有轻微干扰，需继续优化", False, 10),
    ]

    for row_idx, (text, is_bold, size) in enumerate(instructions, 1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(name='微软雅黑', size=size, bold=is_bold)
        cell.alignment = Alignment(vertical='top', wrap_text=True)


def generate_template(lab_code: str, workspace_root: str = None) -> str:
    """生成单个研究室的周报模板"""
    if lab_code not in LAB_CONFIG:
        raise ValueError(f"未知的研究室代码: {lab_code}，可用: {list(LAB_CONFIG.keys())}")

    config = LAB_CONFIG[lab_code]
    room_name = config["name"]
    prefix = config["prefix"]
    output_dir = config["output_dir"]

    # 确定输出路径
    if workspace_root:
        output_dir = os.path.join(workspace_root, output_dir)

    # 确保目录存在
    os.makedirs(output_dir, exist_ok=True)

    # 创建Excel工作簿
    wb = openpyxl.Workbook()

    # Sheet1: 已立项项目
    ws1 = wb.active
    ws1.title = "已立项项目"
    create_report_sheet(ws1, "已立项项目", room_name, prefix)

    # Sheet2: 未立项项目
    ws2 = wb.create_sheet("未立项项目")
    create_report_sheet(ws2, "未立项项目", room_name, prefix)

    # Sheet3: 填表说明
    ws3 = wb.create_sheet("填表说明")
    create_instructions_sheet(ws3, room_name, prefix)

    # 保存文件
    output_file = os.path.join(output_dir, f"周报模板_{room_name}.xlsx")
    wb.save(output_file)

    return output_file


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description="装备所周报模板生成器 V2.0")
    parser.add_argument("--lab", type=str, default=None,
                        help="研究室代码（ME/IV/LA），多个用逗号分隔，不指定则生成全部")
    parser.add_argument("--list", action="store_true",
                        help="列出可用研究室")
    parser.add_argument("--workspace", type=str, default=None,
                        help="workspace根目录（默认为当前目录）")

    args = parser.parse_args()

    if args.list:
        print("可用研究室：")
        for code, config in LAB_CONFIG.items():
            print(f"  {code} - {config['name']}")
        return

    # 确定要生成的研究室
    if args.lab:
        lab_codes = [c.strip().upper() for c in args.lab.split(",")]
    else:
        lab_codes = list(LAB_CONFIG.keys())

    # 计算本周日期范围
    monday, sunday, filename_date = calculate_week_range()
    print(f"汇报周期: {monday} 至 {sunday}")
    print(f"文件名日期: {filename_date}")
    print()

    # 生成模板
    generated_files = []
    for code in lab_codes:
        try:
            output_file = generate_template(code, args.workspace)
            generated_files.append((code, output_file))
            print(f"✅ [{code}] 模板已生成: {output_file}")
        except Exception as e:
            print(f"❌ [{code}] 生成失败: {e}")

    print()
    print(f"完成: {len(generated_files)}/{len(lab_codes)} 个模板")
    return generated_files


if __name__ == "__main__":
    main()
