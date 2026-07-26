#!/usr/bin/env python3
"""
process_imdf.py — IMDF + VA Data Processor for smart-venue-map skill

Usage:
  python process_imdf.py --imdf-zip path/to/IMDF.zip [--va-xlsx path/to/data.xlsx] [--va-json path/to/data.json] --out-dir ./outputs

Outputs:
  <out-dir>/imdf_b64.txt     — gzip+base64 encoded IMDF (for embedding in HTML)
  <out-dir>/dash_data.json   — processed VA/sensor data
  <out-dir>/imdf_compact.json — intermediate compact IMDF (for inspection)
"""

import argparse, zipfile, json, gzip, base64, os, sys
from pathlib import Path

EXCLUDE_UNIT_CATEGORIES = {'parking', 'parkingspot'}

def round_coords(coords, dp=5):
    if isinstance(coords[0], (int, float)):
        return [round(c, dp) for c in coords]
    return [round_coords(c, dp) for c in coords]

def load_geojson(zf, name):
    try:
        with zf.open(name) as f:
            fc = json.load(f)
        return fc.get('features', [])
    except KeyError:
        print(f"  [warn] {name} not found in zip")
        return []

def extract_prop(feat, *keys):
    props = feat.get('properties') or {}
    for k in keys:
        if k in props and props[k] is not None:
            return props[k]
    return None

def process_imdf(zip_path):
    print(f"Processing IMDF: {zip_path}")
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        def find(suffix):
            matches = [n for n in names if n.endswith(suffix)]
            return matches[0] if matches else suffix

        # Levels
        level_features = load_geojson(zf, find('level.geojson'))
        levels = []
        for f in level_features:
            fid = f.get('id') or extract_prop(f, 'id')
            props = f.get('properties') or {}
            levels.append({
                'id': fid,
                'name': props.get('short_name') or props.get('name') or str(props.get('ordinal', '?')),
                'ordinal': props.get('ordinal', 0),
            })
        print(f"  Levels: {len(levels)}")

        # Units
        unit_features = load_geojson(zf, find('unit.geojson'))
        units_raw = len(unit_features)
        units = []
        for f in unit_features:
            fid = f.get('id') or extract_prop(f, 'id')
            props = f.get('properties') or {}
            cat = (props.get('category') or '').lower()
            if cat in EXCLUDE_UNIT_CATEGORIES:
                continue
            geom = f.get('geometry')
            if geom and 'coordinates' in geom:
                geom = dict(geom)
                geom['coordinates'] = round_coords(geom['coordinates'])
            level_id = extract_prop(f, 'level_id', 'level', 'levelId')
            name = props.get('name') or props.get('short_name')
            units.append({'id': fid, 'level_id': level_id, 'category': cat, 'name': name, 'geometry': geom})
        print(f"  Units: {units_raw} -> {len(units)} (filtered {units_raw - len(units)} parking)")

        # Openings
        opening_features = load_geojson(zf, find('opening.geojson'))
        openings = []
        for f in opening_features:
            geom = f.get('geometry')
            if geom and 'coordinates' in geom:
                geom = dict(geom)
                geom['coordinates'] = round_coords(geom['coordinates'])
            level_id = extract_prop(f, 'level_id', 'level', 'levelId')
            openings.append({'level_id': level_id, 'geometry': geom})
        print(f"  Openings: {len(openings)}")

        # Center from footprint
        footprint_features = load_geojson(zf, find('footprint.geojson'))
        center = [0, 0]
        if footprint_features:
            geom = footprint_features[0].get('geometry', {})
            coords = geom.get('coordinates', [[]])
            all_pts = []
            def flatten(c):
                if isinstance(c[0], (int, float)): all_pts.append(c)
                else: [flatten(x) for x in c]
            flatten(coords)
            if all_pts:
                center = [
                    sum(p[1] for p in all_pts) / len(all_pts),
                    sum(p[0] for p in all_pts) / len(all_pts)
                ]

    return {'levels': levels, 'units': units, 'openings': openings, 'center': center}

def encode_imdf(imdf):
    raw = json.dumps(imdf, separators=(',', ':')).encode('utf-8')
    compressed = gzip.compress(raw, compresslevel=9)
    b64 = base64.b64encode(compressed).decode('ascii')
    print(f"  Encoded: {len(raw)//1024}KB -> {len(compressed)//1024}KB gz -> {len(b64)//1024}KB b64 ({len(raw)/len(compressed):.1f}x compression)")
    return b64

def process_va_xlsx(xlsx_path):
    try:
        import openpyxl
    except ImportError:
        print("  [error] openpyxl not installed. Run: pip install openpyxl --break-system-packages")
        sys.exit(1)

    print(f"Processing VA data: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"  Columns: {headers}")

    def col(*names):
        for n in names:
            try: return headers.index(n)
            except ValueError: pass
        return None

    cam_col    = col('camera_id', 'camera', 'cam_id', 'CameraId', 'camera_name')
    date_col   = col('date', 'Date', 'capture_date')
    hour_col   = col('hour', 'Hour', 'capture_hour')
    gender_col = col('gender', 'Gender', 'sex')
    age_col    = col('age_group', 'age', 'Age', 'age_range')
    glass_col  = col('eyewear', 'glasses', 'Eyewear')

    if cam_col is None:
        print("  [error] Cannot find camera_id column in:", headers)
        sys.exit(1)

    cam_data = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        cam = str(row[cam_col]) if row[cam_col] is not None else None
        if not cam: continue
        if cam not in cam_data:
            cam_data[cam] = {'total': 0, 'gender': {}, 'age': {}, 'eyewear': {},
                             'footfall_hour': {}, 'footfall_date': {}, 'footfall_dh': {}}
        cd = cam_data[cam]
        cd['total'] += 1

        if gender_col is not None and row[gender_col]:
            g = str(row[gender_col]).upper()
            cd['gender'][g] = cd['gender'].get(g, 0) + 1
        if age_col is not None and row[age_col]:
            a = str(row[age_col])
            cd['age'][a] = cd['age'].get(a, 0) + 1
        if glass_col is not None and row[glass_col]:
            e = str(row[glass_col])
            cd['eyewear'][e] = cd['eyewear'].get(e, 0) + 1
        if hour_col is not None and row[hour_col] is not None:
            h = str(int(row[hour_col]))
            cd['footfall_hour'][h] = cd['footfall_hour'].get(h, 0) + 1
        if date_col is not None and row[date_col]:
            d = str(row[date_col])[:10]
            cd['footfall_date'][d] = cd['footfall_date'].get(d, 0) + 1
            if hour_col is not None and row[hour_col] is not None:
                h = str(int(row[hour_col]))
                if d not in cd['footfall_dh']: cd['footfall_dh'][d] = {}
                cd['footfall_dh'][d][h] = cd['footfall_dh'][d].get(h, 0) + 1

    print(f"  Cameras: {len(cam_data)} | Rows: {i+1:,}")
    return cam_data

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--imdf-zip', required=True)
    ap.add_argument('--va-xlsx')
    ap.add_argument('--va-json')
    ap.add_argument('--out-dir', default='./outputs')
    args = ap.parse_args()

    out = Path(args.out_dir)
    out.mkdir(parents=True, exist_ok=True)

    imdf = process_imdf(args.imdf_zip)
    compact_path = out / 'imdf_compact.json'
    with open(compact_path, 'w') as f:
        json.dump(imdf, f, separators=(',', ':'))
    print(f"  -> {compact_path} ({os.path.getsize(compact_path)//1024}KB)")

    b64 = encode_imdf(imdf)
    b64_path = out / 'imdf_b64.txt'
    b64_path.write_text(b64)
    print(f"  -> {b64_path} ({len(b64)//1024}KB)")

    cam_data = {}
    if args.va_xlsx:
        cam_data = process_va_xlsx(args.va_xlsx)
    elif args.va_json:
        with open(args.va_json) as f:
            loaded = json.load(f)
        cam_data = loaded.get('cam_data', loaded)
    else:
        print("  [info] No VA data provided — map will work without analytics")

    va_path = out / 'dash_data.json'
    with open(va_path, 'w') as f:
        json.dump({'cameras': sorted(cam_data.keys()), 'cam_data': cam_data}, f, separators=(',', ':'))
    print(f"  -> {va_path} ({os.path.getsize(va_path)//1024}KB)")
    print("\nDone! Pass imdf_b64.txt and dash_data.json to the HTML generator.")

if __name__ == '__main__':
    main()
