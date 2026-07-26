#!/usr/bin/env python3
"""将 Excel 数据库文件转换为精简 CSV（仅保留搜索所需列）"""
import openpyxl
import csv
import os

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
os.makedirs(DATA_DIR, exist_ok=True)

# === 1. CPCD 产品数据 ===
print("=== 转换 CPCD 产品数据 ===")
wb = openpyxl.load_workbook(
    os.path.join(os.path.dirname(__file__), "..",
                 "CPCD数据库V260413-国家温室气体排放因子数据库第二版",
                 "CPCD产品数据-20260413.xlsx"),
    read_only=True
)
ws = wb.active
cpcd_headers = [
    "product_name", "product_name_en", "product_id", "category",
    "model", "cf_value", "cf_unit", "quality_score",
    "data_year", "functional_unit"
]
# 对应原始列索引: 2,3,1,0,4,5,6,7,11,12
col_map_cpcd = [2, 3, 1, 0, 4, 5, 6, 7, 11, 12]

with open(os.path.join(DATA_DIR, "cpcd.csv"), "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(cpcd_headers)
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 15:
            continue
        pn = row[2]
        if not pn:
            continue
        out = []
        for i, ci in enumerate(col_map_cpcd):
            v = row[ci] if ci < len(row) else ""
            out.append(str(v) if v is not None else "")
        writer.writerow(out)
        count += 1
wb.close()
print(f"CPCD: {count} 行")

# === 2. GHG 排放因子数据库 ===
print("\n=== 转换 GHG 排放因子数据库 ===")
ghg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..",
                        "CPCD数据库V260413-国家温室气体排放因子数据库第二版",
                        "国家温室气体排放因子数据库-第二版.xlsx")
wb = openpyxl.load_workbook(ghg_path, read_only=True, data_only=True)
ghg_headers = [
    "category_path", "emission_type", "fuel_type", "factor_value",
    "unit_cn", "source"
]

with open(os.path.join(DATA_DIR, "ghg_factor.csv"), "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(ghg_headers)
    count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 11:
                continue
            cat = row[0]
            if not cat:
                continue
            out = [str(row[i]) if row[i] is not None else "" for i in [0,1,2,3,4,10]]
            writer.writerow(out)
            count += 1
wb.close()
print(f"GHG: {count} 行")

# === 3. Database Overview (ecoinvent) - 仅 Cut-Off AO sheet ===
print("\n=== 转换 Database Overview (ecoinvent Cut-Off AO) ===")
eco_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Database Overview v3.12.xlsx")
wb = openpyxl.load_workbook(eco_path, read_only=True, data_only=True)
ws = wb["Cut-Off AO"]
# Cut-Off AO 列: 0=Activity UUID, 1=Activity Name, 2=Geography, 3=Time Period,
#   4=Sector, 5=ISIC Classification, 6=ISIC Section, 7=Product UUID,
#   8=Product Group, 9=Product Name, 10=CPC Classification, 11=HS2017 Classification,
#   12=Unit, 13=Product Information, 14=CAS Number, 15=Classification(?), 16=?, 17=?
# app.py 映射: activity_name=row[3]... 但那是旧的 Excel 格式
# Database Overview 列名不同，按列名映射

# 读取 header
first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
col_names = [str(c).strip() if c else "" for c in first_row]
print(f"Cut-Off AO columns ({len(col_names)}): {col_names}")

# 建立列名到索引映射
col_idx = {name: i for i, name in enumerate(col_names)}

eco_headers = ["activity_name", "product_name", "geography", "time_period",
               "sector", "unit", "cas_number"]
# 映射到 Cut-Off AO 列名
eco_col_map = {
    "activity_name": "Activity Name",
    "product_name": "Reference Product Name",
    "geography": "Geography",
    "time_period": "Time Period",
    "sector": "Sector",
    "unit": "Unit",
    "cas_number": "CAS Number",
}

with open(os.path.join(DATA_DIR, "ecoinvent.csv"), "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(eco_headers)
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 跳过没有 activity_name 的行
        act_idx = col_idx.get("Activity Name", 1)
        if act_idx >= len(row) or not row[act_idx]:
            continue
        out = []
        for h in eco_headers:
            src_col = eco_col_map[h]
            idx = col_idx.get(src_col, -1)
            if idx >= 0 and idx < len(row) and row[idx] is not None:
                out.append(str(row[idx]))
            else:
                out.append("")
        writer.writerow(out)
        count += 1
wb.close()
print(f"Ecoinvent (Cut-Off AO): {count} 行")

print("\n=== 转换完成 ===")
for fn in ["cpcd.csv", "ghg_factor.csv", "ecoinvent.csv"]:
    fp = os.path.join(DATA_DIR, fn)
    size = os.path.getsize(fp) / 1024
    print(f"  {fn}: {size:.1f} KB")
