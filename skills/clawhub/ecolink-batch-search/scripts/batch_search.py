#!/usr/bin/env python3
"""EcoLink 批量碳足迹搜索脚本（本地数据库模式）

完整搜索管道（与网页端 app.py 对齐）：
  1. LLM 分析材料 → 翻译 + 别名 + 复合材料判断
  2. 本地 CSV 数据库匹配 → 中文名 + 英文名 + 别名多路搜索
  3. 无结果 → LLM 化学品分解 → 搜索各组分
  4. 仍无结果 → LLM 替代品推荐 → 搜索替代品
  5. 输出包含结果类型、匹配原因
  6. 可选：生成 HTML 预览页面，用户勾选后导出 CSV

用法:
    python batch_search.py --input "钢材,水泥,PVC" --output results.csv --no-llm
    python batch_search.py --input "钢材,水泥" --output results.csv --no-llm --preview
    python batch_search.py --file products.xlsx --column "产品名称" --output results.csv --no-llm
"""
import argparse
import csv
import json
import sys
import time
import os
import re
from collections import defaultdict
from difflib import SequenceMatcher

# ============================================================
# 数据加载（本地 CSV）
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "..", "data")

# 全局数据
all_rows = []           # ecoinvent 数据行
cpcd_rows = []          # CPCD 数据行
ghg_factor_rows = []    # GHG 因子库数据行
product_name_to_indices = defaultdict(list)
keyword_index = defaultdict(set)


def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


def tokenize(text):
    stop = {'market', 'for', 'production', 'the', 'of', 'and', 'or', 'at', 'in',
            'to', 'a', 'an', 'from', 'by', 'with'}
    tokens = re.findall(r'[a-zA-Z]+', text.lower())
    return [t for t in tokens if t not in stop and len(t) > 1]


def load_data():
    """从 CSV 文件加载三个数据库到内存"""
    global all_rows, cpcd_rows, ghg_factor_rows
    global product_name_to_indices, keyword_index

    # --- ecoinvent ---
    eco_path = os.path.join(DATA_DIR, "ecoinvent.csv")
    if os.path.exists(eco_path):
        with open(eco_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for idx, row in enumerate(reader):
                row['_idx'] = idx
                all_rows.append(row)
                pn = row.get('product_name', '').lower().strip()
                if pn:
                    product_name_to_indices[pn].append(idx)
                for token in tokenize(pn):
                    keyword_index[token].add(idx)
                # 也对 activity_name 建索引
                an = row.get('activity_name', '').lower().strip()
                for token in tokenize(an):
                    keyword_index[token].add(idx)
        print(f"[数据] ecoinvent: {len(all_rows)} 行")
    else:
        print(f"[数据] ecoinvent.csv 未找到: {eco_path}")

    # --- CPCD ---
    cpcd_path = os.path.join(DATA_DIR, "cpcd.csv")
    if os.path.exists(cpcd_path):
        with open(cpcd_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get('product_name'):
                    cpcd_rows.append(row)
        print(f"[数据] CPCD: {len(cpcd_rows)} 行")
    else:
        print(f"[数据] cpcd.csv 未找到: {cpcd_path}")

    # --- GHG 因子库 ---
    ghg_path = os.path.join(DATA_DIR, "ghg_factor.csv")
    if os.path.exists(ghg_path):
        with open(ghg_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get('category_path'):
                    ghg_factor_rows.append(row)
        print(f"[数据] GHG因子库: {len(ghg_factor_rows)} 行")
    else:
        print(f"[数据] ghg_factor.csv 未找到: {ghg_path}")


# ============================================================
# 本地搜索函数（与 app.py 对齐）
# ============================================================

def local_search(query, max_results=80):
    """ecoinvent 本地模糊搜索"""
    query_lower = query.lower().strip()

    if query_lower in product_name_to_indices:
        return [all_rows[i] for i in product_name_to_indices[query_lower]][:max_results]

    query_tokens = set(tokenize(query_lower))
    if not query_tokens:
        return []

    candidates = {}
    for token in query_tokens:
        if token in keyword_index:
            for idx in keyword_index[token]:
                row = all_rows[idx]
                pn_lower = row.get('product_name', '').lower()
                row_tokens = set(tokenize(pn_lower))
                overlap = len(query_tokens & row_tokens)
                union = len(query_tokens | row_tokens)
                jaccard = overlap / union if union > 0 else 0
                seq_score = SequenceMatcher(None, query_lower, pn_lower).ratio()
                base_score = jaccard * 0.5 + seq_score * 0.5

                length_penalty = len(query_lower) / max(len(pn_lower), 1)
                length_penalty = min(length_penalty, 1.0)
                score = base_score * (0.6 + 0.4 * length_penalty)

                activity_lower = row.get('activity_name', '').lower()
                if any(w in pn_lower or w in activity_lower
                       for w in ['waste', 'treatment of', 'landfill', 'incineration',
                                 'open burning', 'open dump']):
                    score *= 0.5

                if score > 0.2:
                    candidates[idx] = max(candidates.get(idx, 0), score)

    sorted_cands = sorted(candidates.items(), key=lambda x: x[1], reverse=True)
    return [all_rows[idx] for idx, _ in sorted_cands[:max_results]]


def search_cpcd(query, translation=None, max_results=10):
    """CPCD 数据库搜索"""
    if not cpcd_rows:
        return []
    query_lower = query.lower().strip()
    trans_lower = (translation or '').lower().strip()
    results = []
    for row in cpcd_rows:
        pn = row['product_name'].lower()
        pn_en = row.get('product_name_en', '').lower()
        model = row.get('model', '').lower()
        score = 0
        if query_lower == pn or query_lower == pn_en:
            score = 1.0
        elif trans_lower and (trans_lower == pn or trans_lower == pn_en):
            score = 0.95
        elif model and (query_lower == model or (trans_lower and trans_lower == model)):
            score = 0.9
        elif model and query_lower in model:
            len_ratio = len(model) / max(len(query_lower), 1)
            score = 0.7 if len_ratio <= 1.5 else 0.15
        elif model and trans_lower and trans_lower in model:
            len_ratio = len(model) / max(len(trans_lower), 1)
            score = 0.65 if len_ratio <= 1.5 else 0.13
        elif model and model in query_lower:
            score = 0.5
        elif model and trans_lower and model in trans_lower:
            score = 0.45
        elif query_lower in pn or query_lower in pn_en:
            matched = pn if query_lower in pn else pn_en
            len_ratio = len(matched) / max(len(query_lower), 1)
            score = 0.7 if len_ratio <= 1.5 else 0.15
        elif trans_lower and (trans_lower in pn or trans_lower in pn_en):
            matched = pn if trans_lower in pn else pn_en
            len_ratio = len(matched) / max(len(trans_lower), 1)
            score = 0.65 if len_ratio <= 1.5 else 0.13
        elif pn in query_lower:
            score = 0.5
        elif trans_lower and pn in trans_lower:
            score = 0.45
        else:
            seq = max(SequenceMatcher(None, query_lower, pn).ratio(),
                      SequenceMatcher(None, query_lower, pn_en).ratio())
            if trans_lower:
                seq = max(seq, SequenceMatcher(None, trans_lower, pn).ratio(),
                          SequenceMatcher(None, trans_lower, pn_en).ratio())
            if seq > 0.7:
                cn_seq = SequenceMatcher(None, query_lower, pn).ratio()
                if cn_seq > 0.6 or (trans_lower and SequenceMatcher(None, trans_lower, pn_en).ratio() > 0.7):
                    score = seq * 0.5
        if score > 0.2:
            results.append((score, row))
    results.sort(key=lambda x: x[0], reverse=True)
    return [r for _, r in results[:max_results]]


def search_ghg_factor(query, translation=None, max_results=10):
    """GHG 因子库搜索"""
    if not ghg_factor_rows:
        return []
    query_lower = query.lower().strip()
    trans_lower = (translation or '').lower().strip()
    results = []
    for row in ghg_factor_rows:
        cat = row.get('category_path', '').lower()
        fuel = row.get('fuel_type', '').lower()
        emission = row.get('emission_type', '').lower()
        score = 0
        if query_lower == fuel or query_lower == emission:
            score = 1.0
        elif trans_lower and (trans_lower == fuel or trans_lower == emission):
            score = 0.95
        elif query_lower in fuel or query_lower in emission:
            score = 0.7
        elif trans_lower and (trans_lower in fuel or trans_lower in emission):
            score = 0.6
        elif query_lower in cat:
            score = 0.25
        elif trans_lower and trans_lower in cat:
            score = 0.22
        else:
            seq_fuel = SequenceMatcher(None, query_lower, fuel).ratio()
            seq_em = SequenceMatcher(None, query_lower, emission).ratio()
            seq_cat = SequenceMatcher(None, query_lower, cat).ratio()
            direct_best = max(seq_fuel, seq_em)
            best_seq = max(direct_best, seq_cat * 0.6)
            if trans_lower:
                seq_fuel2 = SequenceMatcher(None, trans_lower, fuel).ratio()
                seq_em2 = SequenceMatcher(None, trans_lower, emission).ratio()
                best_seq = max(best_seq, seq_fuel2, seq_em2)
            if best_seq > 0.7 and direct_best > 0.6:
                score = best_seq * 0.5
        if score > 0.2:
            results.append((score, row))
    results.sort(key=lambda x: x[0], reverse=True)
    return [r for _, r in results[:max_results]]


def search_local_all(material, translation="", alt_names=None, top_n=5):
    """本地搜索三个数据库，返回与 search_raw 相同格式的结果"""
    search_terms = [material]
    if translation:
        search_terms.append(translation)
    if alt_names:
        search_terms.extend(alt_names)

    # ecoinvent: 多路搜索合并去重
    ecoinvent_results = []
    seen_eco = set()
    for term in search_terms:
        for r in local_search(term, max_results=30):
            key = (r.get('product_name', ''), r.get('geography', ''), r.get('activity_name', ''))
            if key not in seen_eco:
                seen_eco.add(key)
                ecoinvent_results.append({
                    'product_name': r.get('product_name', ''),
                    'activity_name': r.get('activity_name', ''),
                    'geography': r.get('geography', ''),
                    'unit': r.get('unit', ''),
                    'sector': r.get('sector', ''),
                    'time_period': r.get('time_period', ''),
                })

    # CPCD: 多路搜索
    cpcd_results = []
    seen_cpcd = set()
    for term in search_terms:
        for r in search_cpcd(term, translation=translation, max_results=top_n):
            pid = r.get('product_id', '')
            if pid and pid not in seen_cpcd:
                seen_cpcd.add(pid)
                cpcd_results.append({
                    'product_name': r.get('product_name', ''),
                    'model': r.get('model', ''),
                    'cf_value': r.get('cf_value', ''),
                    'cf_unit': r.get('cf_unit', ''),
                    'quality_score': r.get('quality_score', ''),
                    'functional_unit': r.get('functional_unit', ''),
                    'data_year': r.get('data_year', ''),
                })

    # GHG 因子库
    ghg_results = []
    seen_ghg = set()
    for term in search_terms:
        for r in search_ghg_factor(term, translation=translation, max_results=top_n):
            key = (r.get('category_path', ''), r.get('fuel_type', ''))
            if key not in seen_ghg:
                seen_ghg.add(key)
                ghg_results.append({
                    'fuel_type': r.get('fuel_type', ''),
                    'factor_value': r.get('factor_value', ''),
                    'unit_cn': r.get('unit_cn', ''),
                    'category_path': r.get('category_path', ''),
                    'source': r.get('source', ''),
                })

    return {
        'ecoinvent': ecoinvent_results[:top_n * 3],
        'cpcd': cpcd_results[:top_n],
        'ghg_factor': ghg_results[:top_n],
    }


def has_any_result(raw_data):
    """检查是否有结果"""
    if not raw_data or "error" in raw_data:
        return False
    return bool(raw_data.get("ecoinvent") or raw_data.get("cpcd") or raw_data.get("ghg_factor"))


# ============================================================
# LLM 调用
# ============================================================

def call_llm(api_url, api_key, model, messages, temperature=0.1, timeout=30):
    """调用 LLM API（OpenAI 兼容格式）"""
    try:
        import requests
        resp = requests.post(
            api_url,
            json={"model": model, "messages": messages, "temperature": temperature, "max_tokens": 2048},
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            timeout=timeout,
        )
        resp.raise_for_status()
        data = resp.json()
        return data["choices"][0]["message"]["content"].strip()
    except Exception:
        return None


def parse_llm_json(content):
    """解析 LLM 返回的 JSON"""
    if not content:
        return None
    text = content.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


# ============================================================
# LLM 功能函数
# ============================================================

def llm_analyze_material(material, api_url, api_key, model):
    system_prompt = (
        "你是一个碳足迹LCA数据库匹配专家。你的任务是：\n"
        "1. 将用户输入的原材料/产品名称翻译成准确的英文名称（多个常用英文名全部列出）\n"
        "2. 判断该产品是否为复合材料/组合产品\n"
        "3. 如果是复合产品，拆解出各组分材料的英文名称和百分比\n\n"
        "请严格按以下JSON格式返回：\n"
        '{"translation": "主要英文名", "alt_names": ["别名1"], "is_composite": false, '
        '"components": [], "reason": "分析说明（中文）"}\n\n'
        '复合材料示例：\n'
        '{"translation": "concrete", "alt_names": [], "is_composite": true, '
        '"components": [{"name": "cement", "percentage": 30}, {"name": "sand", "percentage": 40}, '
        '{"name": "gravel", "percentage": 30}], "reason": "混凝土是由水泥、砂和石子按比例混合的复合材料"}'
    )
    content = call_llm(api_url, api_key, model, [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": material},
    ])
    return parse_llm_json(content)


def llm_decompose_chemical(material, translation, api_url, api_key, model):
    name_display = f"{material} ({translation})" if translation else material
    system_prompt = (
        "你是一个化学专家。用户查询的化学品在碳足迹LCA数据库中没有直接数据。\n"
        "你的任务是将该化学品分解为可以通过化学反应方程式组合得到的基础组分，并计算质量比例。\n\n"
        "分解规则：\n"
        "1. 优先分解为数据库中最可能存在的简单化学品/元素\n"
        "2. 无机化合物分解为氧化物或单质（如 CaCO3 → CaO + CO2）\n"
        "3. 有机化合物分解为常见有机原料\n"
        "4. 合金/混合物分解为各组分金属\n"
        "5. 组分必须是具体的化学品，不能是宽泛类别\n\n"
        "请严格按以下JSON格式返回：\n"
        '{"molecular_formula": "分子式", "molar_mass": 分子量, "is_chemical": true,\n'
        ' "components": [{"name": "组分英文名", "formula": "分子式", "molar_mass": 分子量,\n'
        '   "mass_fraction": 0.5, "search_terms": ["搜索词1"]}],\n'
        ' "reaction": "化学反应方程式", "reason": "分解说明（中文）"}\n\n'
        "如果不是化学品，设 is_chemical 为 false 并返回空 components"
    )
    content = call_llm(api_url, api_key, model, [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": f"化学品名称: {name_display}"},
    ])
    return parse_llm_json(content)


def llm_suggest_alternatives(material, api_url, api_key, model):
    system_prompt = (
        "你是一个碳足迹LCA数据库匹配专家。用户查询的原材料/产品在数据库中找不到精确匹配。\n"
        "请推荐在LCA数据库中可能存在的**具体**替代品或相近产品。\n\n"
        "请严格按以下JSON格式返回：\n"
        '{"alternatives": [{"name": "替代品英文名", "reason": "推荐理由（中文）"}]}\n\n'
        "推荐规则：\n"
        "1. 替代品必须是**具体的产品/化学品**，不能是宽泛的类别\n"
        "   禁止：chemical, inorganic / metal / plastic / polymer / material 等\n"
        "2. 化学品推荐同族/同类具体化学品\n"
        "3. 工业产品推荐功能/用途最接近的具体产品\n"
        "4. 金属/合金推荐成分最接近的具体合金\n"
        "5. 最多推荐3个替代品\n"
        "6. 如果确实找不到任何具体的替代品，返回空数组"
    )
    content = call_llm(api_url, api_key, model, [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": f"用户在碳足迹数据库中找不到: {material}\n请推荐具体的替代品。"},
    ])
    return parse_llm_json(content)


# ============================================================
# --strict 模式
# ============================================================

def llm_validate_results(material, translation, ecoinvent_rows, api_url, api_key, model):
    if not ecoinvent_rows:
        return None
    top_products = []
    for i, r in enumerate(ecoinvent_rows[:10]):
        top_products.append(f"{i+1}. [{r.get('product_name','')}] | Activity: {r.get('activity_name','')} | Geo: {r.get('geography','')}")
    product_list = "\n".join(top_products)
    system_prompt = (
        "你是一个碳足迹LCA数据库匹配质量审核专家。判断搜索结果是否与用户查询真正相关。\n\n"
        "请严格按以下JSON格式返回：\n"
        '{"relevant": true, "better_terms": []}\n\n'
        "如果不相关：\n"
        '{"relevant": false, "reason": "不相关的原因简述（中文）", "better_terms": ["更好的搜索词1", "搜索词2"]}\n\n'
        "判断规则：\n"
        "1. 产品名和activity名必须与查询材料在功能/用途/化学性质上直接相关\n"
        "2. 仅因为包含某个通用单词（如food、steel）不算相关\n"
        "3. better_terms应为LCA数据库中可能存在的标准产品名称"
    )
    content = call_llm(api_url, api_key, model, [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": f"用户查询: {material}（翻译: {translation or 'N/A'}）\n\n搜索结果（前10条）:\n{product_list}"},
    ])
    return parse_llm_json(content)


def llm_filter_results(material, translation, ecoinvent_rows, api_url, api_key, model):
    if not ecoinvent_rows or len(ecoinvent_rows) <= 1:
        return ecoinvent_rows
    product_list_parts = []
    for i, r in enumerate(ecoinvent_rows[:15]):
        product_list_parts.append(f"{i}. [{r.get('product_name','')}] | {r.get('activity_name','')} | {r.get('geography','')}")
    product_list = "\n".join(product_list_parts)
    system_prompt = (
        "你是碳足迹数据库搜索结果过滤器。从搜索结果中逐条判断哪些真正相关。\n\n"
        "请严格按以下JSON格式返回：\n"
        '{"keep": [0, 2, 5], "removed_reason": "简要说明移除了哪些及原因（中文）"}\n\n'
        "严格过滤规则：\n"
        "1. 产品必须在功能、用途、材料本质上与查询相关\n"
        "2. 仅共享某个通用单词（如cold、steel、transport）不算相关\n"
        "3. 移除含 waste/treatment/landfill/incineration 的废弃处理类条目\n"
        "4. 加工/制造服务类（cutting, shaping, drilling）不是原材料，应移除\n"
        "5. keep数组中的数字是结果编号（从0开始），只包含确实相关的\n"
        "6. 全不相关时keep为空数组"
    )
    content = call_llm(api_url, api_key, model, [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": f"用户查询: {material}\n英文翻译: {translation or 'N/A'}\n\n搜索结果（编号从0开始）:\n{product_list}\n\n请逐条审查，返回应该保留的结果编号。"},
    ], temperature=0.05)
    if not content:
        return ecoinvent_rows
    parsed = parse_llm_json(content)
    if not parsed:
        return ecoinvent_rows
    keep_indices = parsed.get("keep", [])
    if not isinstance(keep_indices, list):
        return ecoinvent_rows
    max_idx = min(len(ecoinvent_rows), 15)
    valid = set(i for i in keep_indices if isinstance(i, int) and 0 <= i < max_idx)
    if not valid:
        return []
    filtered = [ecoinvent_rows[i] for i in range(max_idx) if i in valid]
    if len(ecoinvent_rows) > 15:
        filtered.extend(ecoinvent_rows[15:])
    return filtered if filtered else ecoinvent_rows


# ============================================================
# 结果格式化
# ============================================================

# 输出列定义
COLUMNS = [
    "搜索词", "结果类型", "AI翻译", "数据来源", "产品名称", "Activity名称",
    "型号", "地理区域", "碳足迹/因子值", "单位", "质量评分",
    "功能单元", "数据年份", "匹配原因", "说明", "专家点评",
]

# 替代品黑名单
BLACKLIST = {
    'chemical, inorganic', 'chemical, organic',
    'chemical, basic', 'chemical, fine',
    'chemical, unspecified', 'metal', 'plastic',
    'polymer', 'material',
}


def fmt_row(material, result_type, translation, source, data, reason=""):
    base = {col: "" for col in COLUMNS}
    base["搜索词"] = material
    base["结果类型"] = result_type
    base["AI翻译"] = translation
    base["数据来源"] = source
    base["匹配原因"] = reason
    if source == "ecoinvent":
        base["产品名称"] = data.get("product_name", "")
        base["Activity名称"] = data.get("activity_name", "")
        base["地理区域"] = data.get("geography", "")
        base["单位"] = data.get("unit", "")
        base["质量评分"] = data.get("sector", "")
        base["数据年份"] = data.get("time_period", "")
    elif source == "CPCD":
        base["产品名称"] = data.get("product_name", "")
        base["型号"] = data.get("model", "")
        base["碳足迹/因子值"] = data.get("cf_value", "")
        base["单位"] = data.get("cf_unit", "")
        base["质量评分"] = data.get("quality_score", "")
        base["功能单元"] = data.get("functional_unit", "")
        base["数据年份"] = data.get("data_year", "")
    elif source == "GHG因子库":
        base["产品名称"] = data.get("fuel_type", "")
        base["碳足迹/因子值"] = data.get("factor_value", "")
        base["单位"] = data.get("unit_cn", "")
        base["说明"] = data.get("category_path", "")
        base["专家点评"] = data.get("source", "")
    return base


def format_direct_results(material, translation, raw_data, top_n, reason=""):
    rows = []
    if "error" in raw_data:
        row = {col: "" for col in COLUMNS}
        row["搜索词"] = material
        row["数据来源"] = "error"
        row["产品名称"] = f"请求失败: {raw_data['error']}"
        return [row]
    for r in raw_data.get("ecoinvent", [])[:top_n]:
        rows.append(fmt_row(material, "direct", translation, "ecoinvent", r, reason))
    for r in raw_data.get("cpcd", [])[:top_n]:
        rows.append(fmt_row(material, "direct", translation, "CPCD", r, reason))
    for r in raw_data.get("ghg_factor", [])[:top_n]:
        rows.append(fmt_row(material, "direct", translation, "GHG因子库", r, reason))
    return rows


def format_composite_results(material, translation, components_info, top_n, reason=""):
    rows = []
    for comp in components_info:
        comp_name = comp.get("name", "")
        comp_pct = comp.get("percentage", 0)
        prefix = f"[组分: {comp_name} {comp_pct}%] "
        raw_data = comp.get("raw_data", {})
        for r in raw_data.get("ecoinvent", [])[:top_n]:
            row = fmt_row(material, "composite", translation, "ecoinvent", r, reason)
            row["产品名称"] = prefix + row["产品名称"]
            rows.append(row)
        for r in raw_data.get("cpcd", [])[:top_n]:
            row = fmt_row(material, "composite", translation, "CPCD", r, reason)
            row["产品名称"] = prefix + row["产品名称"]
            rows.append(row)
    return rows


def format_decomposition_results(material, translation, decomp, top_n, reason=""):
    rows = []
    for comp in decomp.get("components", []):
        comp_name = comp.get("name", "")
        comp_formula = comp.get("formula", "")
        mass_frac = comp.get("mass_fraction", 0)
        prefix = f"[分解: {comp_name}({comp_formula}) {mass_frac*100:.1f}%] "
        raw_data = comp.get("raw_data", {})
        for r in raw_data.get("ecoinvent", [])[:top_n]:
            row = fmt_row(material, "decomposition", translation, "ecoinvent", r, reason)
            row["产品名称"] = prefix + row["产品名称"]
            rows.append(row)
        for r in raw_data.get("cpcd", [])[:top_n]:
            row = fmt_row(material, "decomposition", translation, "CPCD", r, reason)
            row["产品名称"] = prefix + row["产品名称"]
            rows.append(row)
    return rows


def format_alternative_results(material, translation, alt_name, alt_reason, raw_data, top_n, reason=""):
    rows = []
    full_reason = f"替代品: {alt_name}（{alt_reason}）"
    if reason:
        full_reason = reason + " | " + full_reason
    for r in raw_data.get("ecoinvent", [])[:top_n]:
        rows.append(fmt_row(material, "alternative", translation, "ecoinvent", r, full_reason))
    for r in raw_data.get("cpcd", [])[:top_n]:
        rows.append(fmt_row(material, "alternative", translation, "CPCD", r, full_reason))
    for r in raw_data.get("ghg_factor", [])[:top_n]:
        rows.append(fmt_row(material, "alternative", translation, "GHG因子库", r, full_reason))
    return rows


def format_no_result(material, translation, reason=""):
    row = {col: "" for col in COLUMNS}
    row["搜索词"] = material
    row["AI翻译"] = translation
    row["数据来源"] = "-"
    row["产品名称"] = "无匹配结果"
    row["匹配原因"] = reason
    return [row]


# ============================================================
# 主搜索管道
# ============================================================

def search_one_material(material, no_llm, api_url, api_key, model, top_n, delay, strict=False):
    translation = ""
    alt_names = []
    reason = ""

    # Step 1: LLM 分析
    analysis = None
    if not no_llm:
        analysis = llm_analyze_material(material, api_url, api_key, model)
        if analysis:
            translation = analysis.get("translation", "")
            alt_names = analysis.get("alt_names", [])
            reason = analysis.get("reason", "")
            time.sleep(delay * 0.3)

    # Step 2: 本地数据库匹配
    raw_data = search_local_all(material, translation=translation, alt_names=alt_names, top_n=top_n)

    if has_any_result(raw_data):
        # 复合材料
        if analysis and analysis.get("is_composite") and analysis.get("components"):
            print(f"  [复合材料]", flush=True)
            comp_results = []
            for comp in analysis["components"]:
                comp_name = comp.get("name", "")
                comp_data = search_local_all(comp_name, top_n=top_n)
                comp["raw_data"] = comp_data
                comp_results.append(comp)
                time.sleep(delay)
            rows = format_composite_results(material, translation, comp_results, top_n, reason)
            if rows:
                return rows, True
            return format_direct_results(material, translation, raw_data, top_n, reason), True

        # --strict 模式
        if strict and not no_llm and raw_data.get("ecoinvent"):
            print(f"  [验证]", flush=True)
            validation = llm_validate_results(material, translation, raw_data["ecoinvent"], api_url, api_key, model)
            time.sleep(delay * 0.3)
            if validation and not validation.get("relevant"):
                better_terms = validation.get("better_terms", [])
                if better_terms:
                    print(f"  [重试:{better_terms[0][:20]}]", flush=True)
                    retry_data = search_local_all(material, translation=better_terms[0], top_n=top_n)
                    time.sleep(delay)
                    if has_any_result(retry_data) and retry_data.get("ecoinvent"):
                        raw_data = retry_data
                        reason = f"验证后重试: {validation.get('reason', '')}"
            if raw_data.get("ecoinvent"):
                print(f"  [过滤]", flush=True)
                filtered_eco = llm_filter_results(material, translation, raw_data["ecoinvent"], api_url, api_key, model)
                time.sleep(delay * 0.3)
                raw_data["ecoinvent"] = filtered_eco
                if not has_any_result(raw_data):
                    return format_no_result(material, translation, reason + "（过滤后无结果）"), False

        return format_direct_results(material, translation, raw_data, top_n, reason), True

    # Step 3: 化学品分解
    if not no_llm:
        print(f"  [尝试分解]", flush=True)
        decomp = llm_decompose_chemical(material, translation, api_url, api_key, model)
        time.sleep(delay * 0.3)
        if decomp and decomp.get("is_chemical") and decomp.get("components"):
            decomp_reason = decomp.get("reason", reason)
            found_any = False
            for comp in decomp["components"]:
                search_terms = comp.get("search_terms", [comp.get("name", "")])
                comp_raw = {}
                for term in search_terms:
                    comp_raw = search_local_all(term, top_n=top_n)
                    if has_any_result(comp_raw):
                        found_any = True
                        break
                    time.sleep(delay)
                comp["raw_data"] = comp_raw
            if found_any:
                rows = format_decomposition_results(material, translation, decomp, top_n, decomp_reason)
                if rows:
                    return rows, True

    # Step 4: 替代品推荐
    if not no_llm:
        print(f"  [替代品]", flush=True)
        alt_result = llm_suggest_alternatives(material, api_url, api_key, model)
        time.sleep(delay * 0.3)
        if alt_result and alt_result.get("alternatives"):
            for alt in alt_result["alternatives"]:
                alt_name = alt.get("name", "")
                alt_reason = alt.get("reason", "")
                if not alt_name or alt_name.lower() in BLACKLIST:
                    continue
                alt_raw = search_local_all(alt_name, top_n=top_n)
                time.sleep(delay)
                if has_any_result(alt_raw):
                    rows = format_alternative_results(
                        material, translation, alt_name, alt_reason, alt_raw, top_n, reason
                    )
                    if rows:
                        return rows, True

    return format_no_result(material, translation, reason), False


# ============================================================
# HTML 预览生成
# ============================================================

def generate_html_preview(all_rows_data, output_path):
    """生成交互式 HTML 预览页面，用户勾选后导出 CSV"""
    html_path = output_path.rsplit(".", 1)[0] + "_preview.html"

    # 按搜索词分组
    grouped = defaultdict(list)
    for row in all_rows_data:
        grouped[row["搜索词"]].append(row)

    # 将数据嵌入 HTML
    data_json = json.dumps(all_rows_data, ensure_ascii=False)
    columns_json = json.dumps(COLUMNS, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>EcoLink 搜索结果预览</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f5f7fa; color: #333; padding: 20px; }}
h1 {{ text-align: center; margin-bottom: 20px; color: #2c3e50; }}
.toolbar {{ background: #fff; padding: 15px 20px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); margin-bottom: 20px; display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }}
.toolbar button {{ padding: 8px 16px; border: none; border-radius: 5px; cursor: pointer; font-size: 14px; }}
.btn-primary {{ background: #4caf50; color: #fff; }}
.btn-primary:hover {{ background: #45a049; }}
.btn-secondary {{ background: #2196f3; color: #fff; }}
.btn-secondary:hover {{ background: #1e88e5; }}
.btn-danger {{ background: #f44336; color: #fff; }}
.btn-danger:hover {{ background: #e53935; }}
.stats {{ font-size: 14px; color: #666; margin-left: auto; }}
.group {{ background: #fff; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); margin-bottom: 16px; overflow: hidden; }}
.group-header {{ background: #34495e; color: #fff; padding: 10px 16px; font-size: 15px; font-weight: 600; display: flex; justify-content: space-between; align-items: center; }}
.group-header label {{ cursor: pointer; }}
table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
th {{ background: #ecf0f1; padding: 8px 10px; text-align: left; position: sticky; top: 0; }}
td {{ padding: 7px 10px; border-bottom: 1px solid #eee; max-width: 300px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
tr:hover {{ background: #f8f9fa; }}
tr.unchecked {{ opacity: 0.4; }}
.source-ecoinvent {{ color: #2196f3; font-weight: 500; }}
.source-cpcd {{ color: #4caf50; font-weight: 500; }}
.source-ghg {{ color: #ff9800; font-weight: 500; }}
.badge {{ display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 11px; color: #fff; }}
.badge-direct {{ background: #4caf50; }}
.badge-composite {{ background: #9c27b0; }}
.badge-decomposition {{ background: #ff9800; }}
.badge-alternative {{ background: #2196f3; }}
</style>
</head>
<body>
<h1>EcoLink 搜索结果预览</h1>
<div class="toolbar">
    <button class="btn-primary" onclick="exportCSV()">&#128229; 导出选中为 CSV</button>
    <button class="btn-secondary" onclick="selectAll(true)">&#9745; 全选</button>
    <button class="btn-danger" onclick="selectAll(false)">&#9744; 全不选</button>
    <span class="stats" id="stats"></span>
</div>
<div id="content"></div>

<script>
const DATA = {data_json};
const COLUMNS = {columns_json};
const checked = new Set(DATA.map((_, i) => i));

function render() {{
    const grouped = {{}};
    DATA.forEach((row, i) => {{
        const key = row["搜索词"];
        if (!grouped[key]) grouped[key] = [];
        grouped[key].push({{row, idx: i}});
    }});

    let html = '';
    for (const [term, items] of Object.entries(grouped)) {{
        const checkedCount = items.filter(x => checked.has(x.idx)).length;
        html += `<div class="group">`;
        html += `<div class="group-header">
            <span>${{term}} (${{items.length}} 条结果)</span>
            <label><input type="checkbox" ${{checkedCount === items.length ? 'checked' : ''}} onchange="toggleGroup('${{term}}', this.checked)"> 全选</label>
        </div>`;
        html += `<table><tr><th style="width:30px"></th>`;
        const showCols = ["结果类型","数据来源","产品名称","Activity名称","型号","地理区域","碳足迹/因子值","单位","质量评分","匹配原因"];
        showCols.forEach(c => html += `<th>${{c}}</th>`);
        html += `</tr>`;
        items.forEach(({{row, idx}}) => {{
            const cls = checked.has(idx) ? '' : ' class="unchecked"';
            const srcCls = row["数据来源"] === "ecoinvent" ? "source-ecoinvent" :
                          row["数据来源"] === "CPCD" ? "source-cpcd" :
                          row["数据来源"] === "GHG因子库" ? "source-ghg" : "";
            const type = row["结果类型"];
            const badgeCls = type ? `badge-${{type}}` : "";
            html += `<tr${{cls}}>`;
            html += `<td><input type="checkbox" ${{checked.has(idx)?'checked':''}} onchange="toggleRow(${{idx}}, this.checked)"></td>`;
            html += `<td><span class="badge ${{badgeCls}}">${{type}}</span></td>`;
            html += `<td class="${{srcCls}}">${{row["数据来源"]}}</td>`;
            html += `<td title="${{esc(row["产品名称"])}}">${{esc(row["产品名称"])}}</td>`;
            html += `<td title="${{esc(row["Activity名称"])}}">${{esc(row["Activity名称"])}}</td>`;
            html += `<td>${{esc(row["型号"])}}</td>`;
            html += `<td>${{esc(row["地理区域"])}}</td>`;
            html += `<td>${{esc(row["碳足迹/因子值"])}}</td>`;
            html += `<td>${{esc(row["单位"])}}</td>`;
            html += `<td>${{esc(row["质量评分"])}}</td>`;
            html += `<td title="${{esc(row["匹配原因"])}}">${{esc(row["匹配原因"])}}</td>`;
            html += `</tr>`;
        }});
        html += `</table></div>`;
    }}
    document.getElementById('content').innerHTML = html;
    updateStats();
}}

function esc(s) {{ return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/"/g,'&quot;').substring(0,100); }}

function toggleRow(idx, val) {{
    if (val) checked.add(idx); else checked.delete(idx);
    render();
}}

function toggleGroup(term, val) {{
    DATA.forEach((row, i) => {{
        if (row["搜索词"] === term) {{
            if (val) checked.add(i); else checked.delete(i);
        }}
    }});
    render();
}}

function selectAll(val) {{
    if (val) DATA.forEach((_, i) => checked.add(i));
    else checked.clear();
    render();
}}

function updateStats() {{
    document.getElementById('stats').textContent = `已选 ${{checked.size}} / ${{DATA.length}} 条`;
}}

function exportCSV() {{
    const selected = DATA.filter((_, i) => checked.has(i));
    if (!selected.length) {{ alert('请至少选择一条结果'); return; }}
    let csv = '\\uFEFF' + COLUMNS.join(',') + '\\n';
    selected.forEach(row => {{
        csv += COLUMNS.map(col => {{
            let v = (row[col] || '').toString().replace(/"/g, '""');
            return '"' + v + '"';
        }}).join(',') + '\\n';
    }});
    const blob = new Blob([csv], {{type: 'text/csv;charset=utf-8'}});
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = 'ecolink_results.csv';
    a.click();
    URL.revokeObjectURL(url);
}}

render();
</script>
</body>
</html>"""

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    return html_path


# ============================================================
# 文件 I/O
# ============================================================

def read_input_file(filepath, column):
    ext = os.path.splitext(filepath)[1].lower()
    materials = []
    if ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if column not in header:
            print(f"错误: 列 '{column}' 不存在。可用列: {', '.join(h for h in header if h)}")
            sys.exit(1)
        col_idx = header.index(column)
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[col_idx]
            if val and str(val).strip():
                materials.append(str(val).strip())
        wb.close()
    elif ext == ".csv":
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if column not in reader.fieldnames:
                print(f"错误: 列 '{column}' 不存在。可用列: {', '.join(reader.fieldnames)}")
                sys.exit(1)
            for row in reader:
                val = row.get(column, "").strip()
                if val:
                    materials.append(val)
    else:
        print(f"错误: 不支持的文件格式 '{ext}'")
        sys.exit(1)
    return materials


def write_output(rows, output_path, fmt="csv"):
    if fmt == "excel":
        try:
            import openpyxl
        except ImportError:
            print("错误: Excel 输出需要 openpyxl，已切换为 CSV")
            fmt = "csv"
            output_path = output_path.rsplit(".", 1)[0] + ".csv"

    if fmt == "excel":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EcoLink批量搜索结果"
        ws.append(COLUMNS)
        for row in rows:
            ws.append([row.get(col, "") for col in COLUMNS])
        wb.save(output_path)
    else:
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
    return output_path


# ============================================================
# 主函数
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="EcoLink 批量碳足迹搜索（本地数据库模式）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--input", type=str, help="逗号分隔的产品列表")
    parser.add_argument("--file", type=str, help="输入文件路径 (Excel/CSV)")
    parser.add_argument("--column", type=str, default="产品名称", help="列名 (默认: 产品名称)")
    parser.add_argument("--output", type=str, required=True, help="输出文件路径")
    parser.add_argument("--format", type=str, default="csv", choices=["csv", "excel"])
    parser.add_argument("--top", type=int, default=3, help="每库取前N条 (默认: 3)")
    parser.add_argument("--delay", type=float, default=0.5, help="请求间隔秒 (默认: 0.5)")
    parser.add_argument("--no-llm", action="store_true", help="不调用LLM（零费用，仅匹配）")
    parser.add_argument("--strict", action="store_true", help="严格模式：LLM验证+逐条过滤")
    parser.add_argument("--api-key", type=str, default="", help="LLM API Key")
    parser.add_argument("--api-url", type=str, default="https://openrouter.ai/api/v1/chat/completions", help="LLM API URL")
    parser.add_argument("--model", type=str, default="deepseek/deepseek-chat-v3-0324", help="LLM 模型名称")
    parser.add_argument("--preview", action="store_true", help="生成 HTML 预览页面")
    args = parser.parse_args()

    # 加载数据库
    print("正在加载本地数据库...")
    load_data()
    if not all_rows and not cpcd_rows and not ghg_factor_rows:
        print("错误: 没有加载到任何数据库数据，请检查 skill/data/ 目录")
        sys.exit(1)

    # 获取产品列表
    if args.input:
        materials = [m.strip() for m in args.input.split(",") if m.strip()]
    elif args.file:
        materials = read_input_file(args.file, args.column)
    else:
        print("错误: 请指定 --input 或 --file")
        sys.exit(1)

    if not materials:
        print("错误: 没有找到任何产品名称")
        sys.exit(1)

    no_llm = args.no_llm
    api_key = args.api_key

    if not no_llm and not api_key:
        print("错误: 请提供 --api-key 或加上 --no-llm")
        print("  --api-key sk-xxx  用您自己的 LLM API")
        print("  --no-llm          不调LLM，纯匹配（零费用）")
        sys.exit(1)

    total = len(materials)
    llm_label = "关闭" if no_llm else f"{args.model}"

    print(f"\n{'='*55}")
    print(f"  EcoLink 批量搜索（本地数据库模式）")
    print(f"  产品: {total} | LLM: {llm_label} | 输出: {args.format}")
    print(f"  数据库: ecoinvent {len(all_rows)} | CPCD {len(cpcd_rows)} | GHG {len(ghg_factor_rows)}")
    strict_label = "开启" if args.strict else "关闭"
    print(f"  严格模式: {strict_label} | 预览: {'开启' if args.preview else '关闭'}")
    print(f"{'='*55}\n")

    all_result_rows = []
    found_count = 0
    not_found = []
    type_stats = {"direct": 0, "composite": 0, "decomposition": 0, "alternative": 0, "none": 0}

    for i, material in enumerate(materials, 1):
        print(f"[{i}/{total}] {material} ...", end=" ", flush=True)

        rows, found = search_one_material(
            material, no_llm, args.api_url, api_key, args.model, args.top, args.delay,
            strict=args.strict
        )
        all_result_rows.extend(rows)

        if found:
            found_count += 1
            rtype = rows[0]["结果类型"] if rows else "direct"
            type_stats[rtype] = type_stats.get(rtype, 0) + 1
            sources = set(r["数据来源"] for r in rows if r["数据来源"] not in ("-", "error"))
            print(f"OK [{rtype}] {len(rows)}条 [{', '.join(sources)}]")
        else:
            not_found.append(material)
            type_stats["none"] += 1
            print("NO MATCH")

        if i < total:
            time.sleep(args.delay)

    # 写入结果
    actual_path = write_output(all_result_rows, args.output, args.format)

    # HTML 预览
    preview_path = None
    if args.preview and all_result_rows:
        preview_path = generate_html_preview(all_result_rows, args.output)
        print(f"\n预览页面: {preview_path}")

    print(f"\n{'='*55}")
    print(f"  搜索完成！ 总计: {total} | 有结果: {found_count} | 无结果: {len(not_found)}")
    print(f"  直接匹配: {type_stats.get('direct', 0)} | 复合材料: {type_stats.get('composite', 0)}"
          f" | 化学分解: {type_stats.get('decomposition', 0)} | 替代品: {type_stats.get('alternative', 0)}")
    print(f"  结果文件: {actual_path}")
    if preview_path:
        print(f"  预览文件: {preview_path}")
    print(f"{'='*55}")

    if not_found:
        print(f"\n未找到: {', '.join(not_found[:20])}")
        if len(not_found) > 20:
            print(f"  ... 等共 {len(not_found)} 个")


if __name__ == "__main__":
    main()
