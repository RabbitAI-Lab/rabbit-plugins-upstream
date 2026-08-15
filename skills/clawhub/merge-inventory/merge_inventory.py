import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import sys

# 列索引映射（基于表1表头所在行，0基）
COL_SEQ = 1
COL_CODE = 5
COL_NAME = 12
COL_SPEC = 16
COL_UNIT = 19
COL_QTY = 20
COL_PRICE = 25
COL_AMT = 29
COL_BATCH = 33
COL_CUST_DATE = 36
COL_CUST_NO = 39
COL_SUPPLIER_DATE = 43
COL_SUPPLIER_NO = 45

STANDARD_COLS = [
    '序号', '存货编码', '存货名称', '规格', '单位',
    '数量', '单价', '含税金额', '批次号',
    '客户送货日期', '客户送货单号', '供应商送货日期', '供应商送货单号'
]


def extract_blocks_from_sheet(df):
    """从单个 sheet 的 DataFrame 中提取所有数据块（表头→数据→合计）。"""
    total_rows = len(df)
    blocks = []
    i = 0
    while i < total_rows:
        row = df.iloc[i]
        if pd.notna(row[COL_SEQ]) and str(row[COL_SEQ]).strip() == '序号':
            data_rows = []
            j = i + 1
            while j < total_rows:
                next_row = df.iloc[j]
                next_seq = next_row[COL_SEQ]
                if pd.isna(next_seq):
                    break
                seq_val = str(next_seq).strip()
                if seq_val == '序号':
                    break  # 下一个块
                if seq_val == '合计':
                    total_row = [
                        '合计',
                        '',
                        '',
                        next_row[COL_SPEC] if pd.notna(next_row[COL_SPEC]) else '大写',
                        '', '', '',
                        next_row[COL_AMT] if pd.notna(next_row[COL_AMT]) else '',
                        '', '', '', '', '',
                    ]
                    data_rows.append(total_row)
                    j += 1
                    break
                else:
                    data_rows.append([
                        seq_val,
                        str(next_row[COL_CODE]).strip() if pd.notna(next_row[COL_CODE]) else '',
                        str(next_row[COL_NAME]).strip() if pd.notna(next_row[COL_NAME]) else '',
                        str(next_row[COL_SPEC]).strip() if pd.notna(next_row[COL_SPEC]) else '',
                        str(next_row[COL_UNIT]).strip() if pd.notna(next_row[COL_UNIT]) else '',
                        next_row[COL_QTY] if pd.notna(next_row[COL_QTY]) else '',
                        next_row[COL_PRICE] if pd.notna(next_row[COL_PRICE]) else '',
                        next_row[COL_AMT] if pd.notna(next_row[COL_AMT]) else '',
                        str(next_row[COL_BATCH]).strip() if pd.notna(next_row[COL_BATCH]) else '',
                        str(next_row[COL_CUST_DATE]).strip() if pd.notna(next_row[COL_CUST_DATE]) else '',
                        str(next_row[COL_CUST_NO]).strip() if pd.notna(next_row[COL_CUST_NO]) else '',
                        str(next_row[COL_SUPPLIER_DATE]).strip() if pd.notna(next_row[COL_SUPPLIER_DATE]) else '',
                        str(next_row[COL_SUPPLIER_NO]).strip() if pd.notna(next_row[COL_SUPPLIER_NO]) else '',
                    ])
                    j += 1
            if data_rows:
                block_df = pd.DataFrame(data_rows, columns=STANDARD_COLS)
                blocks.append(block_df)
            i = j
        else:
            i += 1
    return blocks


def merge_inventory_tables(input_file, output_file, sheet=None, filter_placeholder=False):
    """
    将入库单 Excel 转换为标准化表格（表2格式）。

    参数:
      input_file: 输入 Excel 路径
      output_file: 输出 Excel 路径
      sheet: 指定处理的 sheet 名；None 时自动处理所有 sheet(多sheet文件)或第一个(单sheet)
      filter_placeholder: 是否过滤'食堂采购专用'占位行(存货编码71000064且金额0)
    """
    xl = pd.ExcelFile(input_file)
    sheet_names = xl.sheet_names

    if sheet is not None:
        target_sheets = [sheet]
    elif len(sheet_names) == 1:
        target_sheets = sheet_names
    else:
        target_sheets = sheet_names  # 多 sheet 全部处理

    all_blocks = []
    for sname in target_sheets:
        df = pd.read_excel(input_file, header=None, sheet_name=sname)
        blocks = extract_blocks_from_sheet(df)
        all_blocks.extend(blocks)

    # 可选过滤占位行
    if filter_placeholder:
        filtered = []
        for b in all_blocks:
            b2 = b[~((b['存货编码'] == '71000064') & (b['含税金额'] == 0))]
            filtered.append(b2)
        all_blocks = filtered

    # 合并所有块，块之间插入空行
    merged_rows = []
    for idx, block in enumerate(all_blocks):
        if idx > 0:
            merged_rows.append([''] * len(STANDARD_COLS))
        merged_rows.extend(block.values.tolist())

    out_df = pd.DataFrame(merged_rows, columns=STANDARD_COLS)

    # 写入 Excel（openpyxl 格式，支持样式）
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '入库明细汇总'

    header_font = Font(bold=True, name='Arial', size=11)
    header_fill = PatternFill('solid', start_color='D9E1F2')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(STANDARD_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for row_idx, row_data in enumerate(out_df.values.tolist(), start=2):
        is_total = str(row_data[0]) == '合计'
        row_font = Font(bold=is_total, name='Arial', size=10)
        total_fill = PatternFill('solid', start_color='FFF2CC') if is_total else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value != '' else None)
            cell.font = row_font
            if total_fill:
                cell.fill = total_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, col_name in enumerate(STANDARD_COLS, start=1):
        max_len = len(str(col_name))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 30)

    ws.freeze_panes = 'A2'

    wb.save(output_file)
    print(f'输出文件已保存: {output_file}')
    print(f'共合并 {len(all_blocks)} 个入库单数据块，{len(out_df)} 行数据（含空行分隔）')
    return output_file


if __name__ == '__main__':
    input_path = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\jinshan\Desktop\1.xlsx'
    output_path = sys.argv[2] if len(sys.argv) > 2 else r'C:\Users\jinshan\Desktop\merged_output.xlsx'
    merge_inventory_tables(input_path, output_path)
