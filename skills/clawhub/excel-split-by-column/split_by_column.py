# -*- coding: utf-8 -*-
"""
按列拆分 Excel 表格
用法:
    python split_by_column.py <输入.xlsx> [--column 列名] [--mode sheets|files] [--outdir 输出目录]

功能:
1. 读取主表, 按指定列的值把数据分组
2. mode=sheets: 每组写成一个 sheet, 输出单个 xlsx
3. mode=files:  每组写成一个独立 xlsx 文件, 输出到指定目录
4. 自动处理 sheet 名非法字符 / 重复分组名 / 超 31 字符截断
5. 每个子表带表头、冻结首行, 输出拆分统计
"""
import sys
import re
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要 openpyxl, 请先安装: pip install openpyxl")
    sys.exit(1)

ILLEGAL_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def sanitize_sheet_name(name, used, max_len=28):
    """生成合法且唯一的 sheet 名"""
    name = ILLEGAL_SHEET_CHARS.sub("_", str(name).strip() or "未填写")
    name = name[:max_len].strip() or "未填写"
    base = name
    i = 2
    while name.lower() in used:
        suffix = f"_{i}"
        name = base[: max_len - len(suffix)] + suffix
        i += 1
    used.add(name.lower())
    return name


def style_sheet(ws, headers, rows):
    """写入表头和数据并美化"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="16A34A")
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(list(row))
    for c_idx, h in enumerate(headers, 1):
        max_len = len(str(h or ""))
        for row in rows[:200]:
            if c_idx - 1 < len(row) and row[c_idx - 1] is not None:
                l = len(str(row[c_idx - 1]))
                if l > max_len:
                    max_len = l
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 4, 40)
    ws.freeze_panes = "A2"


def main():
    parser = argparse.ArgumentParser(description="按列拆分 Excel")
    parser.add_argument("input", help="输入 xlsx 文件")
    parser.add_argument("--column", "-c", default="", help="拆分依据的列名 (默认自动检测: 部门/城市/分类/状态类列)")
    parser.add_argument("--mode", "-m", choices=["sheets", "files"], default="sheets", help="sheets=拆到多个sheet; files=拆成多个文件")
    parser.add_argument("--outdir", "-o", default="", help="files 模式的输出目录")
    parser.add_argument("--sheet", default="", help="源 sheet 名 (默认第一个)")
    parser.add_argument("--keep-empty", action="store_true", help="保留拆分列值为空的行到'未填写'组")
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        print(f"错误: 找不到文件 {src}")
        sys.exit(1)

    wb = load_workbook(src, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # 定位表头
    header_idx = None
    for i, row in enumerate(rows[:5]):
        if any(c is not None and str(c).strip() for c in row):
            header_idx = i
            break
    if header_idx is None:
        print("错误: 空表")
        sys.exit(1)
    headers = [str(h).strip() if h is not None else "" for h in rows[header_idx]]
    data_rows = [r for r in rows[header_idx + 1:] if any(c is not None and str(c).strip() for c in r)]

    # 确定拆分列
    col_idx = None
    if args.column:
        target = args.column.strip()
        for i, h in enumerate(headers):
            if h == target or target in h:
                col_idx = i
                break
        if col_idx is None:
            print(f"错误: 找不到列 '{args.column}', 可用列: {', '.join(h for h in headers if h)}")
            sys.exit(1)
    else:
        keywords = ["部门", "城市", "大区", "区域", "分类", "类别", "状态", "类型", "项目", "校区", "门店", "组别"]
        best = -1
        best_unique = 0
        for i, h in enumerate(headers):
            if any(k in h for k in keywords):
                uniq = len({str(r[i]).strip() if i < len(r) and r[i] is not None else "" for r in data_rows})
                if 1 < uniq <= 30 and uniq > best_unique:
                    best = i
                    best_unique = uniq
        if best >= 0:
            col_idx = best
        else:
            # 兜底: 唯一值在 2~30 之间的文本列
            for i, h in enumerate(headers):
                uniq = {str(r[i]).strip() if i < len(r) and r[i] is not None else "" for r in data_rows}
                if 1 < len(uniq) <= 30:
                    col_idx = i
                    break
        if col_idx is None:
            print("错误: 无法自动识别拆分列, 请用 --column 指定")
            sys.exit(1)

    col_name = headers[col_idx]

    # 分组
    groups = {}
    for row in data_rows:
        raw = row[col_idx] if col_idx < len(row) else None
        key = str(raw).strip() if raw is not None and str(raw).strip() else "未填写"
        groups.setdefault(key, []).append(row)

    if not args.keep_empty and "未填写" in groups and len(groups) > 1:
        print(f"提示: 有 {len(groups['未填写'])} 行 '{col_name}' 为空, 已归入'未填写'组")

    # 排序分组名
    group_names = sorted(groups.keys())

    # 输出
    stats = []
    if args.mode == "sheets":
        out_file = src.with_name(src.stem + "_split.xlsx")
        out_wb = Workbook()
        out_wb.remove(out_wb.active)
        used = set()
        for gname in group_names:
            sheet_name = sanitize_sheet_name(gname, used)
            gws = out_wb.create_sheet(sheet_name)
            style_sheet(gws, headers, groups[gname])
            stats.append((gname, sheet_name, len(groups[gname])))
        # 汇总 sheet 放最前
        sum_ws = out_wb.create_sheet("拆分汇总", 0)
        sum_ws.append([f"按「{col_name}」拆分", ""])
        sum_ws.append(["分组值", "sheet 名", "行数"])
        for gname, sname, n in stats:
            sum_ws.append([gname, sname, n])
        sum_ws.append(["合计", "", len(data_rows)])
        sum_ws["A1"].font = Font(bold=True, size=14)
        for c in range(1, 4):
            cell = sum_ws.cell(row=2, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="16A34A")
        sum_ws.column_dimensions["A"].width = 28
        sum_ws.column_dimensions["B"].width = 30
        sum_ws.column_dimensions["C"].width = 10
        out_wb.save(out_file)
        result_path = out_file
    else:
        outdir = Path(args.outdir) if args.outdir else src.parent / (src.stem + "_split")
        outdir.mkdir(parents=True, exist_ok=True)
        used = set()
        for gname in group_names:
            safe = sanitize_sheet_name(gname, used)
            out_file = outdir / f"{safe}.xlsx"
            owb = Workbook()
            ows = owb.active
            ows.title = safe[:31]
            style_sheet(ows, headers, groups[gname])
            owb.save(out_file)
            stats.append((gname, safe, len(groups[gname])))
        result_path = outdir

    print("=" * 50)
    print(f"拆分完成: 按「{col_name}」列 → {len(stats)} 组, 共 {len(data_rows)} 行")
    for gname, sname, n in stats:
        print(f"  {gname:<20} → {sname:<28} {n:>5} 行")
    print(f"  {'合计':<20} {'':<30} {len(data_rows):>5} 行")
    print(f"输出: {result_path}")
    print("=" * 50)


if __name__ == "__main__":
    main()
