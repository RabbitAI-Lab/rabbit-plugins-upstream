# -*- coding: utf-8 -*-
"""
WMS 仓库收费对账脚本
将Excel收费清单与WMS OpenAPI中的实际业务记录逐笔核对

用法:
  python reconcile.py --excel "C:/path/to/收费清单.xlsx" \
    --app-key "xxx" --app-secret "xxx" --warehouse-id "<仓库ID>" \
    --base-url "https://jou.topwms.com/api/open/erp" \
    --date-from "2026-06-01 00:00:00" --date-to "2026-08-17 23:59:59" \
    [--skip-rent] [--mark-excel] [--output-dir "./output"]

参数说明:
  --excel       Excel收费清单路径（必填）
  --app-key     WMS系统 AppKey（必填，向仓库方获取）
  --app-secret  WMS系统 AppSecret（必填，向仓库方获取）
  --warehouse-id 仓库ID（必填，向仓库方获取）
  --base-url    API地址（默认 https://jou.topwms.com/api/open/erp）
  --date-from   账期开始时间（默认从Excel读取）
  --date-to     账期结束时间（默认从Excel读取）
  --skip-rent   跳过仓租费核对（默认跳过）
  --mark-excel  标红Excel异常行并增加异常说明列
  --output-dir  输出目录（默认为脚本所在目录）

免责声明: 本脚本为第三方独立工具，与TopWMS无关联/背书关系。
AppKey/AppSecret仅在本机使用，不存储、不上传。
"""
import argparse
import hashlib
import json
import os
import sys
import time
from datetime import datetime
from collections import Counter

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ============ API函数 ============

def generate_signature(json_body_str, app_secret):
    body_md5 = hashlib.md5(json_body_str.encode('utf-8')).hexdigest()
    to_sign = body_md5 + app_secret
    return hashlib.md5(to_sign.encode('utf-8')).hexdigest()


def call_api(base_url, app_key, app_secret, uri, params=None):
    if params is None:
        params = {}
    params['requestTimestamp'] = int(time.time())
    json_body = json.dumps(params, ensure_ascii=False, separators=(',', ':'))
    signature = generate_signature(json_body, app_secret)
    headers = {
        'Content-Type': 'application/json',
        'AppKey': app_key,
        'Signature': signature,
    }
    resp = requests.post(base_url + uri, data=json_body.encode('utf-8'), headers=headers, timeout=30)
    try:
        return resp.json()
    except Exception:
        return None


def fetch_all_pages(base_url, app_key, app_secret, uri, base_params, list_key='list'):
    all_items = []
    cursor = None
    page = 1
    while True:
        params = dict(base_params)
        params['pageSize'] = 50
        if cursor:
            params['cursor'] = cursor
        result = call_api(base_url, app_key, app_secret, uri, params)
        if not result or result.get('result') != 'success':
            print(f"  API失败: {result}")
            break
        data = result.get('data', {})
        items = data.get(list_key, [])
        all_items.extend(items)
        total = data.get('total', 0)
        cursor = data.get('cursor')
        print(f"  第{page}页: {len(items)}条, 累计{len(all_items)}/{total}")
        if not cursor or len(items) == 0:
            break
        page += 1
        time.sleep(0.3)
    return all_items


# ============ Excel读取 ============

def read_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = None
        for row in ws.iter_rows(min_row=1, values_only=True):
            if all(c is None for c in row):
                continue
            if headers is None:
                headers = [str(c).strip() if c else '' for c in row]
                continue
            row_dict = {}
            for h, v in zip(headers, row):
                if h:
                    row_dict[h] = v
            if row_dict:
                rows.append(row_dict)
        sheets[sheet_name] = rows
        print(f"  Sheet '{sheet_name}': {len(rows)}行, 列: {headers}")
    wb.close()
    return sheets


# ============ 出库费核对 ============

def reconcile_outbound(excel_rows, api_orders):
    """以运单号为准，核对运单状态是否最终完成"""
    print("\n" + "=" * 60)
    print("核对出库费明细 — 以运单号为准，验证业务完成状态")
    print("=" * 60)

    tracking_map = {}
    for order in api_orders:
        ln = order.get('logisticsNo', '')
        if ln:
            tracking_map[ln.strip()] = order

    results = []
    for i, row in enumerate(excel_rows):
        excel_tracking = str(row.get('运单号', '')).strip() if row.get('运单号') else ''
        order_sn = str(row.get('业务单号', '')).strip() if row.get('业务单号') else ''
        excel_fee = float(row['消费金额']) if row.get('消费金额') is not None else None

        if not excel_tracking:
            results.append({
                'index': i + 1, 'order_sn': order_sn, 'tracking_no': '',
                'excel_fee': excel_fee, 'status': '异常',
                'issue': 'Excel中无运单号', 'checks': {},
            })
            continue

        api_order = tracking_map.get(excel_tracking)
        if not api_order:
            results.append({
                'index': i + 1, 'order_sn': order_sn, 'tracking_no': excel_tracking,
                'excel_fee': excel_fee, 'status': '未找到',
                'issue': f'WMS中未找到运单号 {excel_tracking}', 'checks': {},
            })
            continue

        stage = api_order.get('stage', '')
        gmt_out = api_order.get('gmtOutStorage', '')
        waybill_url = api_order.get('waybillOssUrl', '')
        hold_up = api_order.get('holdUpStatus', '')
        closed_reason = api_order.get('closedReason', '')

        checks = {
            'stage_has_out': stage == 'has_out_storage',
            'has_out_time': bool(gmt_out),
            'has_waybill': bool(waybill_url),
            'no_holdup': str(hold_up) == '0' or hold_up == 0,
            'not_closed': not closed_reason,
        }

        issues = []
        if not checks['stage_has_out']:
            issues.append(f'订单状态={stage}（非has_out_storage）')
        if not checks['has_out_time']:
            issues.append('无出库时间')
        if not checks['has_waybill']:
            issues.append('无面单(waybillOssUrl为空)')
        if not checks['no_holdup']:
            issues.append(f'有拦截(holdUpStatus={hold_up})')
        if not checks['not_closed']:
            issues.append(f'订单已关闭(closedReason={closed_reason})')

        status = '已完成' if not issues else '未完成'
        results.append({
            'index': i + 1, 'order_sn': order_sn,
            'order_id': api_order.get('orderId', ''),
            'tracking_no': excel_tracking,
            'excel_fee': excel_fee,
            'api_fee': api_order.get('outStorageFee', ''),
            'stage': stage, 'gmt_out_storage': gmt_out,
            'has_waybill': bool(waybill_url), 'hold_up_status': hold_up,
            'closed_reason': closed_reason,
            'status': status, 'issue': '; '.join(issues) if issues else '',
            'checks': checks,
        })

    matched = sum(1 for r in results if r['status'] == '已完成')
    print(f"  已完成: {matched}/{len(results)}")
    return results


# ============ 退件处理费核对 ============

def reconcile_return(excel_rows, api_returns):
    """以业务单号为准，核对退货物料是否已实际入库"""
    print("\n" + "=" * 60)
    print("核对退件处理费明细 — 以业务单号为准，验证入库")
    print("=" * 60)

    return_map = {}
    for ret in api_returns:
        sn = ret.get('returnSn', '')
        if sn:
            return_map[sn.strip()] = ret

    results = []
    for i, row in enumerate(excel_rows):
        rma_no = str(row.get('业务单号', '')).strip() if row.get('业务单号') else ''
        excel_fee = float(row['消费金额']) if row.get('消费金额') is not None else None
        tracking_no = str(row.get('运单号', '')).strip() if row.get('运单号') else ''

        if not rma_no:
            results.append({
                'index': i + 1, 'rma_no': '', 'tracking_no': tracking_no,
                'excel_fee': excel_fee, 'status': '异常',
                'issue': 'Excel中无业务单号', 'checks': {},
            })
            continue

        api_ret = return_map.get(rma_no)
        if not api_ret:
            results.append({
                'index': i + 1, 'rma_no': rma_no, 'tracking_no': tracking_no,
                'excel_fee': excel_fee, 'status': '未找到',
                'issue': f'WMS中未找到退货单(returnSn={rma_no})', 'checks': {},
            })
            continue

        tab = api_ret.get('tab', '')
        gmt_sign = api_ret.get('gmtSign', '')
        goods_sku_list = api_ret.get('goodsSkuList', [])

        total_in_storage = 0
        total_sign = 0
        sku_details = []
        for sku_item in goods_sku_list:
            in_qty = sku_item.get('inStorageQuantity', 0) or 0
            sign_qty = sku_item.get('signQuantity', 0) or 0
            discard_qty = sku_item.get('discardQuantity', 0) or 0
            total_in_storage += in_qty
            total_sign += sign_qty
            sku_details.append({
                'sku': sku_item.get('goodsSkuOuterId', ''),
                'inStorageQuantity': in_qty,
                'signQuantity': sign_qty,
                'discardQuantity': discard_qty,
            })

        checks = {
            'tab_finished': tab == 'finished',
            'has_sign_time': bool(gmt_sign),
            'has_inbound': total_in_storage > 0,
        }

        issues = []
        if not checks['tab_finished']:
            issues.append(f'退货单状态={tab}（非finished）')
        if not checks['has_sign_time']:
            issues.append('无签收时间')
        if not checks['has_inbound']:
            issues.append(f'退货商品未入库(inStorageQuantity=0)')

        status = '已入库' if not issues else '未入库'
        results.append({
            'index': i + 1, 'rma_no': rma_no,
            'return_id': api_ret.get('warehouseReturnOrderId', ''),
            'tracking_no': tracking_no,
            'platform_order': api_ret.get('platformOrderSn', ''),
            'excel_fee': excel_fee, 'tab': tab,
            'gmt_sign': gmt_sign,
            'total_in_storage': total_in_storage,
            'total_sign': total_sign,
            'sku_details': sku_details,
            'status': status, 'issue': '; '.join(issues) if issues else '',
            'checks': checks,
        })

    matched = sum(1 for r in results if r['status'] == '已入库')
    print(f"  已入库: {matched}/{len(results)}")
    return results


# ============ 包装费核对（以WMS为准） ============

def reconcile_packaging(excel_rows, api_orders):
    """以WMS packagingCost为准，对比Excel收费"""
    print("\n" + "=" * 60)
    print("核对包装费明细 — 以WMS系统packagingCost为准")
    print("=" * 60)

    tracking_map = {}
    order_sn_map = {}
    for order in api_orders:
        ln = order.get('logisticsNo', '')
        if ln:
            tracking_map[ln.strip()] = order
        sn = order.get('platformOrderSn', '')
        if sn:
            order_sn_map[sn.strip()] = order

    results = []
    excel_total = 0
    wms_total = 0

    for i, row in enumerate(excel_rows):
        excel_tracking = str(row.get('运单号', '')).strip() if row.get('运单号') else ''
        order_sn = str(row.get('业务单号', '')).strip() if row.get('业务单号') else ''
        excel_fee_raw = row.get('消费金额')
        excel_fee = abs(float(excel_fee_raw)) if excel_fee_raw is not None else 0
        excel_total += excel_fee

        api_order = None
        match_by = ''
        if excel_tracking:
            api_order = tracking_map.get(excel_tracking)
            match_by = '运单号'
        if not api_order and order_sn:
            api_order = order_sn_map.get(order_sn)
            match_by = '业务单号'

        if not api_order:
            results.append({
                'index': i + 1, 'order_sn': order_sn, 'tracking_no': excel_tracking,
                'excel_fee': excel_fee, 'wms_fee': None, 'diff': None,
                'status': '未找到', 'match_by': '',
                'issue': 'WMS中未找到该订单',
            })
            continue

        wms_fee_raw = api_order.get('packagingCost', None)
        wms_fee = float(wms_fee_raw) if wms_fee_raw is not None else 0
        wms_total += wms_fee
        diff = round(excel_fee - wms_fee, 3)

        if abs(diff) < 0.001:
            status = '一致'
            issue = ''
        else:
            status = '不一致'
            issue = f'Excel收费{excel_fee}，WMS记录{wms_fee}，差额{diff}（以WMS为准）'

        results.append({
            'index': i + 1, 'order_sn': order_sn,
            'tracking_no': excel_tracking or api_order.get('logisticsNo', ''),
            'excel_fee': excel_fee, 'wms_fee': wms_fee, 'diff': diff,
            'status': status, 'match_by': match_by, 'issue': issue,
        })

    matched = sum(1 for r in results if r['status'] == '一致')
    print(f"  一致: {matched}/{len(results)}")
    print(f"  Excel总计: {round(excel_total, 3)}, WMS总计: {round(wms_total, 3)}, 差额: {round(excel_total - wms_total, 3)}")
    return results, round(excel_total, 3), round(wms_total, 3)


# ============ 销毁费核对 ============

def reconcile_destroy(excel_rows, api_orders):
    """核对销毁费关联单号是否在WMS中存在"""
    print("\n" + "=" * 60)
    print("核对销毁费明细 — 验证关联单号")
    print("=" * 60)

    # 建立所有可能的单号索引
    all_sn_map = {}
    for order in api_orders:
        for key in ['platformOrderSn', 'orderId', 'logisticsNo']:
            val = order.get(key, '')
            if val:
                all_sn_map[str(val).strip()] = order

    results = []
    for i, row in enumerate(excel_rows):
        excel_fee = float(row.get('消费金额', 0) or 0) if row.get('消费金额') else 0
        # 尝试多种列名
        ref_no = ''
        for col in ['补收销毁费单号', '业务单号', '关联单号', '备注']:
            val = row.get(col, '')
            if val:
                ref_no = str(val).strip()
                break

        api_order = all_sn_map.get(ref_no)
        if api_order:
            status = '已找到'
            issue = f'关联单号{ref_no}在WMS出库单中找到(orderId={api_order.get("orderId","")})'
        else:
            status = '未找到'
            issue = f'关联单号{ref_no}在WMS中未找到'

        results.append({
            'index': i + 1, 'ref_no': ref_no, 'excel_fee': excel_fee,
            'status': status, 'issue': issue,
        })

    matched = sum(1 for r in results if r['status'] == '已找到')
    print(f"  已找到: {matched}/{len(results)}")
    return results


# ============ HTML报告生成 ============

def generate_html_report(outbound_results, return_results, pkg_results, pkg_excel_total, pkg_wms_total,
                         destroy_results, bill_summary, output_path):
    total_items = len(outbound_results) + len(return_results) + len(pkg_results)
    total_ok = (sum(1 for r in outbound_results if r['status'] == '已完成') +
                sum(1 for r in return_results if r['status'] == '已入库') +
                sum(1 for r in pkg_results if r['status'] == '一致'))
    total_issues = total_items - total_ok

    outbound_ok = sum(1 for r in outbound_results if r['status'] == '已完成')
    outbound_issues = [r for r in outbound_results if r['status'] != '已完成']
    return_ok = sum(1 for r in return_results if r['status'] == '已入库')
    pkg_ok = sum(1 for r in pkg_results if r['status'] == '一致')
    pkg_mismatches = [r for r in pkg_results if r['status'] == '不一致']

    # 包装费差异模式
    diff_patterns = Counter()
    for r in pkg_mismatches:
        pattern = f"Excel {r['excel_fee']} → WMS {r['wms_fee']}"
        diff_patterns[pattern] += 1

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>WMS对账报告 - {bill_summary.get('bill_no', 'N/A')}</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: 'Microsoft YaHei', 'Segoe UI', sans-serif; background: #f5f5f5; color: #333; line-height: 1.6; }}
.container {{ max-width: 1200px; margin: 0 auto; padding: 20px; }}
.header {{ background: linear-gradient(135deg, #1a237e, #3949ab); color: white; padding: 30px; border-radius: 10px; margin-bottom: 20px; }}
.header h1 {{ font-size: 24px; margin-bottom: 10px; }}
.header .info {{ display: flex; flex-wrap: wrap; gap: 20px; font-size: 14px; opacity: 0.9; }}
.header .info span {{ background: rgba(255,255,255,0.15); padding: 4px 12px; border-radius: 4px; }}
.summary {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 15px; margin-bottom: 20px; }}
.card {{ background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
.card .label {{ font-size: 13px; color: #666; margin-bottom: 5px; }}
.card .value {{ font-size: 28px; font-weight: bold; }}
.card .value.green {{ color: #2e7d32; }}
.card .value.red {{ color: #c62828; }}
.card .value.blue {{ color: #1565c0; }}
.card .value.orange {{ color: #e65100; }}
.section {{ background: white; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); margin-bottom: 20px; overflow: hidden; }}
.section-header {{ padding: 15px 20px; background: #f0f0f0; border-bottom: 1px solid #e0e0e0; display: flex; justify-content: space-between; align-items: center; }}
.section-header h2 {{ font-size: 16px; color: #333; }}
.section-header .badge {{ padding: 3px 10px; border-radius: 12px; font-size: 12px; font-weight: bold; }}
.badge.green {{ background: #e8f5e9; color: #2e7d32; }}
.badge.red {{ background: #ffebee; color: #c62828; }}
.section-body {{ padding: 15px 20px; }}
table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
table th {{ background: #f5f5f5; padding: 8px 10px; text-align: left; border-bottom: 2px solid #e0e0e0; white-space: nowrap; }}
table td {{ padding: 8px 10px; border-bottom: 1px solid #f0f0f0; }}
table tr:hover {{ background: #fafafa; }}
.note {{ background: #fff3e0; border-left: 4px solid #ff9800; padding: 12px 15px; margin: 10px 0; border-radius: 4px; font-size: 14px; }}
.note.danger {{ background: #ffebee; border-left-color: #f44336; }}
.note.success {{ background: #e8f5e9; border-left-color: #4caf50; }}
.footer {{ text-align: center; padding: 20px; color: #999; font-size: 12px; }}
</style>
</head>
<body>
<div class="container">
    <div class="header">
        <h1>WMS 仓库收费对账报告</h1>
        <div class="info">
            <span>📋 账单编号: {bill_summary.get('bill_no', 'N/A')}</span>
            <span>📅 账期: {bill_summary.get('date_range', 'N/A')}</span>
            <span>💰 账单总额: {bill_summary.get('bill_total', 'N/A')} CNY</span>
            <span>🕐 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</span>
        </div>
    </div>

    <div class="note">
        <strong>对账方法：</strong>
        <ul style="margin-left:20px;margin-top:5px;">
            <li>出库费：以<strong>运单号</strong>为准，核对WMS中运单状态是否最终完成（已出库+有出库时间+有面单+无拦截+未关闭）</li>
            <li>退件处理费：以<strong>业务单号</strong>为准，核对退货物料是否已实际入库（inStorageQuantity &gt; 0）</li>
            <li>包装费：以<strong>WMS系统packagingCost</strong>为准，对比Excel收费金额，不一致的标红</li>
            <li>仓租费：暂不核对</li>
        </ul>
    </div>

    <div class="summary">
        <div class="card"><div class="label">总核对笔数</div><div class="value blue">{total_items}</div></div>
        <div class="card"><div class="label">通过/一致</div><div class="value green">{total_ok}</div></div>
        <div class="card"><div class="label">异常/不一致</div><div class="value red">{total_issues}</div></div>
        <div class="card"><div class="label">包装费差额</div><div class="value red">{round(pkg_excel_total - pkg_wms_total, 3)}</div></div>
    </div>
"""

    # 出库费
    html += f"""
    <div class="section">
        <div class="section-header">
            <h2>一、出库费核对（以运单号为准，{len(outbound_results)}笔）</h2>
            <span class="badge {'green' if outbound_ok == len(outbound_results) else 'red'}">{outbound_ok}/{len(outbound_results)} 已完成</span>
        </div>
        <div class="section-body">
            <div class="note {'success' if outbound_ok == len(outbound_results) else 'danger'}">
                检查项：① stage=has_out_storage ② 有出库时间 ③ 有面单 ④ 无拦截 ⑤ 未关闭
            </div>"""
    if outbound_issues:
        html += """<table><thead><tr><th>序号</th><th>业务单号</th><th>运单号</th><th>Excel金额</th><th>状态</th><th>拦截</th><th>问题</th></tr></thead><tbody>"""
        for r in outbound_issues:
            html += f"<tr><td>{r['index']}</td><td>{r.get('order_sn','')}</td><td>{r.get('tracking_no','')}</td><td>{r.get('excel_fee','')}</td><td style='color:#c62828;font-weight:bold;'>{r['status']}</td><td>{r.get('hold_up_status','')}</td><td>{r.get('issue','')}</td></tr>"
        html += "</tbody></table>"
    html += "</div></div>"

    # 退件处理费
    html += f"""
    <div class="section">
        <div class="section-header">
            <h2>二、退件处理费核对（以业务单号为准，验证入库，{len(return_results)}笔）</h2>
            <span class="badge {'green' if return_ok == len(return_results) else 'red'}">{return_ok}/{len(return_results)} 已入库</span>
        </div>
        <div class="section-body">
            <table><thead><tr><th>序号</th><th>业务单号(RMA)</th><th>Excel金额</th><th>WMS状态</th><th>签收时间</th><th>入库数量</th><th>SKU明细</th><th>核对结果</th></tr></thead><tbody>"""
    for r in return_results:
        status_color = '#2e7d32' if r['status'] == '已入库' else '#c62828'
        sku_str = '<br>'.join([f"{s['sku']}: 入库{s['inStorageQuantity']}, 签收{s['signQuantity']}" for s in r.get('sku_details', [])])
        html += f"<tr><td>{r['index']}</td><td>{r.get('rma_no','')}</td><td>{r.get('excel_fee','')}</td><td>{r.get('tab','')}</td><td>{r.get('gmt_sign','')}</td><td>{r.get('total_in_storage','')}</td><td style='font-size:11px;'>{sku_str}</td><td style='color:{status_color};font-weight:bold;'>{r['status']}</td></tr>"
    html += "</tbody></table></div></div>"

    # 包装费
    html += f"""
    <div class="section">
        <div class="section-header">
            <h2>三、包装费核对（以WMS为准，{len(pkg_results)}笔）</h2>
            <span class="badge {'green' if pkg_ok == len(pkg_results) else 'red'}">{pkg_ok}/{len(pkg_results)} 一致</span>
        </div>
        <div class="section-body">
            <div class="note danger">
                <strong>金额汇总：</strong>Excel总计 {pkg_excel_total} CNY，WMS总计 {pkg_wms_total} CNY，差额 {round(pkg_excel_total - pkg_wms_total, 3)} CNY
            </div>"""
    if diff_patterns:
        html += "<table><thead><tr><th>差异模式</th><th>笔数</th></tr></thead><tbody>"
        for pattern, count in diff_patterns.most_common():
            html += f"<tr><td>{pattern}</td><td>{count}</td></tr>"
        html += "</tbody></table>"

    html += "<table><thead><tr><th>序号</th><th>业务单号</th><th>运单号</th><th>Excel收费</th><th>WMS收费</th><th>差额</th><th>状态</th></tr></thead><tbody>"
    for r in pkg_results:
        if r['status'] == '未找到':
            html += f"<tr><td>{r['index']}</td><td>{r.get('order_sn','')}</td><td>{r.get('tracking_no','')}</td><td>{r.get('excel_fee','')}</td><td>N/A</td><td>N/A</td><td style='color:#c62828;font-weight:bold;'>未找到</td></tr>"
        else:
            color = '#2e7d32' if r['status'] == '一致' else '#c62828'
            html += f"<tr><td>{r['index']}</td><td>{r.get('order_sn','')}</td><td>{r.get('tracking_no','')}</td><td>{r.get('excel_fee','')}</td><td>{r.get('wms_fee','')}</td><td>{r.get('diff',''):>+}</td><td style='color:{color};font-weight:bold;'>{r['status']}</td></tr>"
    html += "</tbody></table></div></div>"

    # 销毁费
    if destroy_results:
        html += f"""
    <div class="section">
        <div class="section-header">
            <h2>四、销毁费核对（{len(destroy_results)}笔）</h2>
        </div>
        <div class="section-body">
            <table><thead><tr><th>序号</th><th>关联单号</th><th>Excel金额</th><th>状态</th><th>说明</th></tr></thead><tbody>"""
        for r in destroy_results:
            color = '#2e7d32' if r['status'] == '已找到' else '#c62828'
            html += f"<tr><td>{r['index']}</td><td>{r.get('ref_no','')}</td><td>{r.get('excel_fee','')}</td><td style='color:{color};font-weight:bold;'>{r['status']}</td><td>{r.get('issue','')}</td></tr>"
        html += "</tbody></table></div></div>"

    html += f"""
    <div class="footer">本报告由WMS对账技能自动生成 | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
</div>
</body>
</html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"\nHTML报告已保存至: {output_path}")


# ============ Excel标红 ============

def mark_excel_red(excel_path, pkg_results, output_path=None):
    """在Excel包装费明细Sheet中标红异常行并增加异常说明列"""
    print("\n标红Excel异常行...")
    wb = openpyxl.load_workbook(excel_path)

    if '包装费明细' not in wb.sheetnames:
        print("  未找到'包装费明细'Sheet，跳过标红")
        wb.close()
        return

    ws = wb['包装费明细']
    header_row = [cell.value for cell in ws[1]]
    new_col_idx = len(header_row) + 1

    # 写入表头
    header_cell = ws.cell(row=1, column=new_col_idx, value="异常说明")
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    red_font = Font(color="C62828", bold=True)
    red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    for i, result in enumerate(pkg_results):
        excel_row = i + 2
        if result['status'] == '不一致':
            desc = f"异常：Excel收费{result['excel_fee']}，WMS系统记录{result['wms_fee']}，差额{result['diff']}（以WMS为准）"
        elif result['status'] == '未找到':
            desc = "异常：WMS系统中未找到该订单"
        else:
            continue  # 一致的不标红

        cell = ws.cell(row=excel_row, column=new_col_idx, value=desc)
        cell.font = red_font
        cell.fill = red_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for col in range(1, new_col_idx + 1):
            c = ws.cell(row=excel_row, column=col)
            c.font = red_font
            if col != new_col_idx:
                c.fill = red_fill

    ws.column_dimensions[openpyxl.utils.get_column_letter(new_col_idx)].width = 50

    # 保存
    save_path = output_path or excel_path
    try:
        wb.save(save_path)
        print(f"  已保存至: {save_path}")
    except PermissionError:
        base = os.path.dirname(excel_path)
        name = os.path.basename(excel_path)
        name_no_ext = os.path.splitext(name)[0]
        save_path = os.path.join(base, f"{name_no_ext}_标红异常.xlsx")
        wb.save(save_path)
        print(f"  原文件被占用，已保存副本: {save_path}")
    finally:
        wb.close()

    return save_path


# ============ 主流程 ============

def main():
    parser = argparse.ArgumentParser(description='WMS仓库收费对账')
    parser.add_argument('--excel', required=True, help='Excel收费清单路径')
    parser.add_argument('--app-key', required=True, help='WMS系统 AppKey（向仓库方获取）')
    parser.add_argument('--app-secret', required=True, help='WMS系统 AppSecret（向仓库方获取）')
    parser.add_argument('--warehouse-id', required=True, help='仓库ID')
    parser.add_argument('--base-url', default='https://jou.topwms.com/api/open/erp', help='API地址')
    parser.add_argument('--date-from', default=None, help='账期开始时间 (YYYY-MM-DD HH:MM:SS)')
    parser.add_argument('--date-to', default=None, help='账期结束时间 (YYYY-MM-DD HH:MM:SS)')
    parser.add_argument('--skip-rent', action='store_true', default=True, help='跳过仓租费核对')
    parser.add_argument('--mark-excel', action='store_true', help='标红Excel异常行')
    parser.add_argument('--output-dir', default='.', help='输出目录')
    args = parser.parse_args()

    print("=" * 60)
    print("WMS 仓库收费对账脚本启动")
    print("=" * 60)

    os.makedirs(args.output_dir, exist_ok=True)

    # 1. 读取Excel
    print("\n[1/5] 读取Excel收费清单...")
    excel_data = read_excel(args.excel)

    # 从账单明细读取账期
    bill_rows = excel_data.get('账单明细', [])
    bill_summary = {}
    if bill_rows:
        bill = bill_rows[0]
        bill_summary['bill_no'] = bill.get('账单编号', '')
        bill_summary['bill_total'] = bill.get('账单金额', '')
        # 尝试读取账期
        date_from = args.date_from or bill.get('账期开始', '') or bill.get('开始时间', '')
        date_to = args.date_to or bill.get('账期结束', '') or bill.get('结束时间', '')
        bill_summary['date_range'] = f"{date_from} ~ {date_to}"
    else:
        date_from = args.date_from
        date_to = args.date_to
        bill_summary['date_range'] = f"{date_from} ~ {date_to}"

    if not date_from or not date_to:
        print("错误：无法确定账期，请通过 --date-from 和 --date-to 参数指定")
        sys.exit(1)

    # 2. 获取API数据
    print("\n[2/5] 从WMS API获取业务数据...")

    print("\n  获取出库单...")
    api_orders = fetch_all_pages(args.base_url, args.app_key, args.app_secret,
                                 "/order/search_order_page", {
                                     "warehouseId": args.warehouse_id,
                                     "gmtCreateFrom": date_from,
                                     "gmtCreateTo": date_to,
                                 })
    print(f"  出库单总数: {len(api_orders)}")

    print("\n  获取退货单...")
    api_returns = fetch_all_pages(args.base_url, args.app_key, args.app_secret,
                                  "/warehouse_return_order/search_page", {
                                      "warehouseId": args.warehouse_id,
                                      "gmtCreateFrom": date_from,
                                      "gmtCreateTo": date_to,
                                  })
    print(f"  退货单总数: {len(api_returns)}")

    # 3. 逐项核对
    print("\n[3/5] 开始逐笔核对...")

    outbound_results = reconcile_outbound(excel_data.get('出库费明细', []), api_orders)
    return_results = reconcile_return(excel_data.get('退件处理费明细', []), api_returns)
    pkg_results, pkg_excel_total, pkg_wms_total = reconcile_packaging(
        excel_data.get('包装费明细', []), api_orders)
    destroy_results = reconcile_destroy(excel_data.get('销毁费明细', []), api_orders)

    # 4. 生成报告
    print("\n[4/5] 生成对账报告...")
    html_path = os.path.join(args.output_dir, 'reconciliation_report.html')
    generate_html_report(outbound_results, return_results, pkg_results,
                         pkg_excel_total, pkg_wms_total, destroy_results,
                         bill_summary, html_path)

    # 保存JSON
    detail = {
        'bill_summary': bill_summary,
        'api_stats': {'出库单': len(api_orders), '退货单': len(api_returns)},
        'outbound_fee': outbound_results,
        'return_fee': return_results,
        'packaging_fee': pkg_results,
        'packaging_summary': {
            'excel_total': pkg_excel_total,
            'wms_total': pkg_wms_total,
            'diff': round(pkg_excel_total - pkg_wms_total, 3),
        },
        'destroy_fee': destroy_results,
    }
    json_path = os.path.join(args.output_dir, 'reconciliation_detail.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(detail, f, ensure_ascii=False, indent=2, default=str)
    print(f"详细JSON已保存至: {json_path}")

    # 5. 标红Excel
    if args.mark_excel:
        print("\n[5/5] 标红Excel异常行...")
        marked_path = mark_excel_red(args.excel, pkg_results,
                                     os.path.join(args.output_dir, 'reconciliation_marked.xlsx'))
    else:
        print("\n[5/5] 跳过Excel标红（未指定 --mark-excel）")

    # 摘要
    print("\n" + "=" * 60)
    print("对账完成摘要")
    print("=" * 60)
    outbound_ok = sum(1 for r in outbound_results if r['status'] == '已完成')
    return_ok = sum(1 for r in return_results if r['status'] == '已入库')
    pkg_ok = sum(1 for r in pkg_results if r['status'] == '一致')
    print(f"  出库费: {outbound_ok}/{len(outbound_results)} 已完成")
    print(f"  退件处理费: {return_ok}/{len(return_results)} 已入库")
    print(f"  包装费: {pkg_ok}/{len(pkg_results)} 一致 (Excel {pkg_excel_total} vs WMS {pkg_wms_total}, 差额 {round(pkg_excel_total - pkg_wms_total, 3)})")
    print(f"  销毁费: {sum(1 for r in destroy_results if r['status'] == '已找到')}/{len(destroy_results)} 已找到")
    if args.mark_excel:
        print(f"  标红Excel: 已保存")


if __name__ == '__main__':
    main()
