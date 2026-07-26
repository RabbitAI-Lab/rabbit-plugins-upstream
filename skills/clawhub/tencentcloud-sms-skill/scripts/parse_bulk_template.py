#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
腾讯云短信 — 解析群发 Excel 模板 (parse_bulk_template)

解析群发 Excel 模板文件，提取手机号和模板变量列表，供 send_sms.py 批量发送使用。
通过 --template-id 参数自动查询模板实际变量数量，校验 Excel 中填充的变量列数是否满足。

校验规则：
  - 用户填充的变量数 >= 模板实际变量数 → 自动兼容，忽略多余列
  - 用户填充的变量数 < 模板实际变量数 → 报错提示缺少几个变量

模板文件格式：
  A 列：客户手机号
  B 列起：短信内容变量1、变量2、变量3 ...（按模板变量 {1}、{2}... 顺序）

需要依赖: openpyxl（脚本会自动检测并安装）

用法:
    python parse_bulk_template.py --file "/path/to/群发模板.xlsx" --template-id 1110
    python parse_bulk_template.py --file "/path/to/群发模板.xlsx" --template-id 1110 --international 0
    python parse_bulk_template.py --file "/path/to/群发模板.xlsx" --template-id 1111 --international 1
"""

import argparse
import json
import os
import re
import subprocess
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _common import output_json, output_error  # noqa: E402


def ensure_openpyxl():
    """检测并自动安装 openpyxl 依赖。"""
    try:
        import openpyxl  # noqa: F401  # pylint: disable=unused-import
    except ImportError:
        print("[INFO] openpyxl not found. Installing...", file=sys.stderr)
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
            stdout=sys.stderr,
            stderr=sys.stderr,
        )
        print("[INFO] openpyxl installed successfully.", file=sys.stderr)


def query_template_variable_count(template_id, international=0):
    """查询模板 ID 实际拥有的变量数量。

    通过调用 DescribeSmsTemplateList API 获取模板内容，
    然后从 TemplateContent 中解析 {1}, {2}, {3}... 占位符的数量。

    Args:
        template_id: 模板 ID
        international: 0=国内短信, 1=国际/港澳台短信

    Returns:
        (variable_count, template_content) 模板变量数量和模板内容
    """
    from _common import ensure_dependencies, get_credentials, build_client

    ensure_dependencies()

    from tencentcloud.common.exception.tencent_cloud_sdk_exception import (
        TencentCloudSDKException,
    )
    from tencentcloud.sms.v20210111 import models

    cred = get_credentials()
    client = build_client(cred)

    try:
        params = {
            "International": international,
            "TemplateIdSet": [int(template_id)],
        }
        req = models.DescribeSmsTemplateListRequest()
        req.from_json_string(json.dumps(params))
        resp = client.DescribeSmsTemplateList(req)
        result = json.loads(resp.to_json_string())

        template_set = result.get("DescribeTemplateStatusSet", [])
        if not template_set:
            output_error(
                "TEMPLATE_NOT_FOUND",
                f"未找到模板 ID {template_id}（international={international}），请确认模板 ID 是否正确。",
            )

        tpl = template_set[0]
        content = tpl.get("TemplateContent", "")

        # 提取 {1}, {2}, {3} ... 形式的变量占位符
        placeholders = re.findall(r"\{(\d+)\}", content)
        if placeholders:
            var_count = max(int(p) for p in placeholders)
        else:
            var_count = 0

        return var_count, content

    except TencentCloudSDKException as e:
        from _common import handle_api_error
        handle_api_error(e)
    except Exception as e:  # pylint: disable=broad-except
        output_error("TEMPLATE_QUERY_ERROR", f"查询模板变量数失败: {e}")


def normalize_phone(phone_str, international=False):
    """标准化手机号为 E.164 格式。

    Args:
        phone_str: 原始手机号字符串
        international: 是否为国际/港澳台短信

    Returns:
        E.164 格式的手机号字符串
    """
    phone = str(phone_str).strip().replace(" ", "").replace("-", "")

    # 已经是 E.164 格式
    if phone.startswith("+"):
        return phone

    if not international:
        # 国内短信：补 +86 前缀
        if phone.startswith("86"):
            return f"+{phone}"
        return f"+86{phone}"
    else:
        # 国际短信：模板中号码格式为 国家码+手机号（如 8521414xxxx）
        # 直接加 + 前缀
        return f"+{phone}"


def _load_excel_rows(file_path):
    """读取 Excel 全部行（含表头）。"""
    import openpyxl

    if not os.path.isfile(file_path):
        output_error("FILE_NOT_FOUND", f"文件不存在: {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:  # pylint: disable=broad-except
        output_error("FILE_READ_ERROR", f"无法读取 Excel 文件: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    wb.close()
    if not rows:
        output_error("EMPTY_FILE", "Excel 文件为空")
    return rows


def _count_variable_columns(header):
    """校验表头并统计「短信内容变量」列数。"""
    if not header[0] or "手机号" not in str(header[0]):
        output_error(
            "INVALID_FORMAT",
            f"Excel 第一行 A 列应为'客户手机号'，当前值: {header[0]}",
        )
    count = 0
    for i in range(1, len(header)):
        col_name = str(header[i]).strip() if header[i] else ""
        if col_name.startswith("短信内容变量"):
            count += 1
        else:
            break
    return count


def _resolve_actual_var_count(file_var_count, expected_var_count, template_content):
    """根据 Excel 列数 vs 模板要求，决定实际使用的变量列数。

    不足直接 output_error 退出；充足则返回应使用的列数（取模板要求）。
    """
    if expected_var_count is None or expected_var_count <= 0:
        return file_var_count
    if file_var_count < expected_var_count:
        shortage = expected_var_count - file_var_count
        error_msg = (
            f"Excel 中填充了 {file_var_count} 个变量列，"
            f"但模板实际需要 {expected_var_count} 个变量，"
            f"还缺少 {shortage} 个变量。"
        )
        if template_content:
            error_msg += f"\n模板内容: {template_content}"
        error_msg += f"\n请补充缺少的变量列（变量{file_var_count + 1}"
        if shortage > 1:
            error_msg += f" ~ 变量{expected_var_count}"
        error_msg += "）后重新上传。"
        output_error("VARIABLE_COUNT_MISMATCH", error_msg)
    return expected_var_count


def _extract_records(data_rows, actual_var_count, international):
    """从数据行中提取手机号 + 模板变量记录。"""
    records = []
    for row_idx, row in enumerate(data_rows, start=2):
        phone_raw = row[0]
        if phone_raw is None or str(phone_raw).strip() == "":
            continue
        phone = normalize_phone(phone_raw, international)
        template_params = []
        for i in range(1, actual_var_count + 1):
            if i < len(row) and row[i] is not None:
                template_params.append(str(row[i]).strip())
            else:
                template_params.append("")
        records.append({
            "phone": phone,
            "template_params": template_params,
            "row": row_idx,
        })
    if not records:
        output_error("NO_DATA", "Excel 文件中没有有效的数据行")
    return records


def _collect_warnings(records, file_var_count, expected_var_count):
    """收集数据级警告（变量数不一致 / 重复号码 / 自动裁剪提示）。"""
    warnings = []
    param_counts = set(len(r["template_params"]) for r in records)
    if len(param_counts) > 1:
        warnings.append(f"模板变量个数不一致: {param_counts}")

    phones = [r["phone"] for r in records]
    duplicates = set(p for p in phones if phones.count(p) > 1)
    if duplicates:
        warnings.append(f"存在重复手机号: {list(duplicates)[:5]}")

    if expected_var_count is not None and file_var_count > expected_var_count:
        warnings.append(
            f"Excel 中有 {file_var_count} 个变量列，模板只需 {expected_var_count} 个，"
            f"已自动忽略多余的 {file_var_count - expected_var_count} 列。"
        )
    return warnings


def _build_parse_result(file_path, international, records, actual_var_count,
                       file_var_count, expected_var_count, template_content,
                       warnings):
    """组装最终输出字典。"""
    result = {
        "file": os.path.basename(file_path),
        "type": "国际/港澳台短信" if international else "国内短信",
        "total_records": len(records),
        "variable_count": actual_var_count,
        "records": records,
    }
    if expected_var_count is not None and file_var_count > expected_var_count:
        result["variable_auto_trimmed"] = {
            "excel_columns": file_var_count,
            "template_required": expected_var_count,
            "ignored_columns": file_var_count - expected_var_count,
        }
    if expected_var_count is not None:
        result["template_expected_variables"] = expected_var_count
        if template_content:
            result["template_content"] = template_content
    if warnings:
        result["warnings"] = warnings
    result["summary"] = {
        "total_recipients": len(records),
        "sample_phones": [r["phone"] for r in records[:5]],
        "sample_params": records[0]["template_params"] if records else [],
        "variable_columns": actual_var_count,
    }
    return result


def parse_excel(file_path, international=False, expected_var_count=None, template_content=None):
    """解析群发 Excel 模板文件。

    Args:
        file_path: Excel 文件路径
        international: 是否为国际/港澳台短信
        expected_var_count: 模板实际要求的变量数量（通过 API 查询获得），
                          传入后会进行校验：
                          - Excel 变量列数 >= expected_var_count → 自动兼容，只取前 N 列
                          - Excel 变量列数 < expected_var_count → 报错提示不足
        template_content: 模板内容（用于报错时展示）

    Returns:
        解析结果字典
    """
    rows = _load_excel_rows(file_path)
    file_var_count = _count_variable_columns(rows[0])
    actual_var_count = _resolve_actual_var_count(
        file_var_count, expected_var_count, template_content,
    )
    records = _extract_records(rows[1:], actual_var_count, international)
    warnings = _collect_warnings(records, file_var_count, expected_var_count)
    return _build_parse_result(
        file_path, international, records, actual_var_count,
        file_var_count, expected_var_count, template_content, warnings,
    )


def build_parser():
    parser = argparse.ArgumentParser(
        description="腾讯云短信 — 解析群发 Excel 模板"
    )
    parser.add_argument(
        "--file", required=True,
        help="群发 Excel 模板文件路径",
    )
    parser.add_argument(
        "--international", type=int, default=0, choices=[0, 1],
        help="短信类型：0=国内短信（默认），1=国际/港澳台短信",
    )
    parser.add_argument(
        "--template-id", required=True,
        help="模板 ID，脚本会自动查询模板实际变量数量并校验 Excel 变量列数（必填）",
    )
    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    ensure_openpyxl()

    # 查询模板实际变量数（--template-id 为必填参数）
    expected_var_count, template_content = query_template_variable_count(
        args.template_id, international=args.international
    )

    result = parse_excel(
        args.file,
        international=bool(args.international),
        expected_var_count=expected_var_count,
        template_content=template_content,
    )
    output_json(result)


if __name__ == "__main__":
    main()
