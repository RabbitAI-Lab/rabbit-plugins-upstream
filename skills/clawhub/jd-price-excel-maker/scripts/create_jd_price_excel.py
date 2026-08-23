#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "电脑配置价格表"

# 设置表头
headers = ["类型", "详细配置", "数量", "单价（元）", "小计（元）", "京东链接", "备注"]
ws.append(headers)

# 表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 数据
data = [
    {
        "type": "主板",
        "config": "技嘉MS03-CE0 单路至强4/5代XX16",
        "qty": 1,
        "price": 4899,
        "link": "https://jingfen.jd.com/detail/irURM1lpmN0KpKz1NerVYWIWsPIMQ_3sdbW6yVNSeRn64200.html",
        "remark": "京东常卖价，店铺：北京契程工作站服务器店"
    },
    {
        "type": "CPU",
        "config": "intel Xeon 8481C 56核心112线程 基准频率2.0GHz",
        "qty": 1,
        "price": None,
        "link": "https://item.jd.com/10136850221470.html",
        "remark": "京东无单独CPU零售，此链接为搭载该CPU的楚霏准系统，常卖价¥24699起（不含显卡）"
    },
    {
        "type": "散热器",
        "config": "志强4/5代4677 M96风冷散热器",
        "qty": 1,
        "price": 499,
        "link": "https://item.jd.com/10068065166152.html",
        "remark": "COOLSERVER金钱豹 4677-M96/4U6热管/320W，参考中关村在线京东在售价格"
    },
    {
        "type": "内存",
        "config": "海力士/三星/镁光 32G DDR5 4800 RECC",
        "qty": 4,
        "price": 6499,
        "link": "https://jingfen.jd.com/detail/dKnpciiAGrSXiaDErkgrXiWO9iuFen_3CVJIOpio5q8H4sizk.html",
        "remark": "SK HYNIX海力士 DDR5 ECC RDIMM 32GB 4800，常卖价，单条价格"
    },
    {
        "type": "机箱",
        "config": "服务器专用塔式机箱",
        "qty": 1,
        "price": 1359,
        "link": "https://jingfen.jd.com/detail/BWQ9383svppZ8iteT2xcZ8iteT2xcQ_3rtcLEeeUh98aOjmXd.html",
        "remark": "银昕(SilverStone) RM44 4U服务器工作站机箱，机塔互换"
    },
    {
        "type": "电源",
        "config": "长城服务器专用1250W",
        "qty": 1,
        "price": 709,
        "link": "https://item.jd.com/10059966614196.html",
        "remark": "长城巨龙GW-EPS1250DA 1250W金牌全模组，参考什么值得买爆料价（店铺：长城GreatWall旗舰店）"
    }
]

# 边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 填充数据
for row_num, item in enumerate(data, 2):
    subtotal = item["price"] * item["qty"] if item["price"] else None
    row = [
        item["type"],
        item["config"],
        item["qty"],
        item["price"] if item["price"] else "—",
        subtotal if subtotal else "—",
        item["link"],
        item["remark"]
    ]
    ws.append(row)
    
    # 设置超链接
    link_cell = ws.cell(row=row_num, column=6)
    link_cell.hyperlink = item["link"]
    link_cell.value = item["link"]
    link_cell.font = Font(color="0563C1", underline="single")
    link_cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # 设置所有单元格样式
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# 添加总计行
total_row = len(data) + 2
ws.cell(row=total_row, column=1, value="合计（不含CPU）")
ws.cell(row=total_row, column=5, value="=SUM(E2:E7)")
ws.cell(row=total_row, column=1).font = Font(bold=True)
ws.cell(row=total_row, column=5).font = Font(bold=True)

# 为总计行添加边框
for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=total_row, column=col_num)
    cell.border = thin_border
    cell.alignment = Alignment(vertical="center")

# 添加说明
note_row = total_row + 2
notes = [
    "说明：",
    "1. 以上价格来源于网络搜索，获取时间为2026年8月20日，实际价格以京东页面显示为准。",
    "2. Intel Xeon 8481C为OEM定制CPU，京东平台无单独散片零售，仅能通过整机/准系统购买。",
    "3. 由于京东对自动化访问有限制，部分价格参考自中关村在线、什么值得买等平台的京东在售信息。",
    "4. 内存按海力士DDR5 ECC RDIMM 32GB 4800报价，用户原配置中“海力士/三星/镁光”为三选一品牌。",
    "5. 点击“京东链接”列中的链接可直接访问对应商品页面。"
]

for i, note in enumerate(notes):
    ws.cell(row=note_row + i, column=1, value=note)
    if i == 0:
        ws.cell(row=note_row + i, column=1).font = Font(bold=True)

# 合并合计单元格
ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right", vertical="center")

# 设置列宽
column_widths = {
    "A": 12,
    "B": 40,
    "C": 8,
    "D": 12,
    "E": 12,
    "F": 55,
    "G": 55
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 设置行高
ws.row_dimensions[1].height = 30
for row_num in range(2, total_row + 1):
    ws.row_dimensions[row_num].height = 60

# 保存文件
output_path = "/Users/wuhaorui/WorkBuddy/本地虾/电脑配置京东价格表.xlsx"
wb.save(output_path)
print(f"Excel文件已保存至: {output_path}")
