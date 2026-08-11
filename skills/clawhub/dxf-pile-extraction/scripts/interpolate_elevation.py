#!/usr/bin/env python3
"""
从含等高线的DXF中提取等高线数据，用IDW插值计算每根桩的岩面标高。
写入桩参数Excel的新列。

用法:
    python interpolate_elevation.py <dxf路径> <桩参数Excel路径> \\
        --layer g_sDgxLayer --sheet 中风化|微风化 \\
        [--k 12] [--power 2] [--column-offset 8]
"""
import sys
import re
import ezdxf
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from scipy.spatial import KDTree
from collections import Counter


def extract_contours(doc, layer_name):
    """从指定图层提取等高线（POLYLINE顶点）和标高标注。"""
    msp = doc.modelspace()

    # 提取等高线POLYLINE
    polylines = []
    for ent in msp:
        if ent.dxftype() not in ('POLYLINE', 'LWPOLYLINE'):
            continue
        try:
            if ent.dxf.layer != layer_name:
                continue
        except Exception:
            continue

        pts = []
        if ent.dxftype() == 'LWPOLYLINE':
            pts = [(p[0], p[1]) for p in ent.get_points()]
        else:  # POLYLINE
            pts = [(v.dxf.location.x, v.dxf.location.y)
                   for v in ent.vertices]

        if len(pts) >= 2:
            polylines.append({'points': pts, 'entity': ent})

    print(f"  等高线POLYLINE: {len(polylines)} 条")

    # 提取标高TEXT
    elevations = []
    for ent in msp:
        if ent.dxftype() != 'TEXT':
            continue
        try:
            if ent.dxf.layer != layer_name:
                continue
        except Exception:
            continue
        text = ent.dxf.text.strip()
        if re.match(r'^-?\d+\.?\d*$', text):
            elev = float(text)
            elevations.append({
                'elev': elev,
                'x': ent.dxf.insert.x,
                'y': ent.dxf.insert.y
            })

    print(f"  标高TEXT: {len(elevations)} 个")

    return polylines, elevations


def match_elev_to_polyline(polylines, elevations):
    """最近距离匹配标高到等高线。"""
    from collections import defaultdict

    matched = {}
    unmatched = 0

    for elev in elevations:
        min_dist = float('inf')
        best_pl = None
        for i, pl in enumerate(polylines):
            # 计算标高文本到等高线所有顶点的最小距离
            for px, py in pl['points']:
                d = ((elev['x'] - px) ** 2 + (elev['y'] - py) ** 2) ** 0.5
                if d < min_dist:
                    min_dist = d
                    best_pl = i

        if best_pl is not None and min_dist < 50:  # 5m阈值
            if best_pl not in matched:
                matched[best_pl] = []
            matched[best_pl].append(elev['elev'])
        else:
            unmatched += 1

    # 多条标高取平均
    for pl_idx in matched:
        matched[pl_idx] = sum(matched[pl_idx]) / len(matched[pl_idx])

    print(f"  匹配成功: {len(matched)} 条, 未匹配: {unmatched}")
    return matched


def create_interp_points(polylines, elevations_matched):
    """创建插值点列表。(x, y, elevation)。"""
    points = []
    for i, pl in enumerate(polylines):
        if i not in elevations_matched:
            continue
        elev = elevations_matched[i]
        for px, py in pl['points']:
            points.append(np.array([px, py, elev]))
    return np.array(points)


def idw_interpolate(interp_points, x, y, k=12, power=2):
    """IDW反距离加权插值。"""
    if len(interp_points) == 0:
        return None

    distances = np.sqrt((interp_points[:, 0] - x) ** 2 +
                        (interp_points[:, 1] - y) ** 2)

    # 取最近的k个点
    if len(distances) <= k:
        idx = np.argsort(distances)
    else:
        idx = np.argpartition(distances, k)[:k]
        idx = idx[np.argsort(distances[idx])]

    # 检查是否正好落在已知点上
    if distances[idx[0]] < 0.001:
        return interp_points[idx[0], 2]

    # IDW
    d = distances[idx] ** power
    d = np.maximum(d, 0.001)  # 避免除零
    weights = 1.0 / d
    return np.sum(interp_points[idx, 2] * weights) / np.sum(weights)


def main(dxf_path, excel_path, layer='g_sDgxLayer',
         sheet_type='中风化', k=12, power=2, column_offset=None):
    print(f"📖 读取等高线DXF: {dxf_path}")
    doc = ezdxf.readfile(dxf_path)

    # ====== 第1步：提取等高线数据 ======
    print(f"📐 提取等高线（图层: {layer}）...")
    polylines, elevations = extract_contours(doc, layer)
    if not polylines:
        print("❌ 未找到等高线，退出")
        return
    if not elevations:
        print("⚠️ 未找到标高标注，无法插值")
        return

    # ====== 第2步：匹配标高到等高线 ======
    print("🔗 匹配标高→等高线...")
    elevations_matched = match_elev_to_polyline(polylines, elevations)

    # ====== 第3步：创建插值点集 ======
    print("📊 创建插值点集...")
    interp_points = create_interp_points(polylines, elevations_matched)
    print(f"  插值点数: {len(interp_points)}")

    all_elevs = sorted(set(round(e, 2) for e in elevations_matched.values()))
    print(f"  标高等值: {min(all_elevs)} ~ {max(all_elevs)}m"
          f" ({len(all_elevs)}个值)")

    # ====== 第4步：读取Excel桩参数 ======
    print(f"📖 读取桩参数: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # 找桩顶绝对标高列（表头含"绝对标高"）
    abs_col = None
    x_col, y_col = None, None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=c).value or '')
        if '绝对' in h and '标高' in h:
            abs_col = c
        if '坐标X' in h or ('X' in h and '坐标' in str(ws.cell(row=1,
                                    column=c+1).value or '')):
            x_col = c
        if '坐标Y' in h or 'Y' in h:
            y_col = c

    # 备选：手动指定
    if abs_col is None:
        abs_col = 6   # 默认第6列=桩顶绝对标高
    if x_col is None:
        x_col = 7     # 默认第7列=X坐标
    if y_col is None:
        y_col = 8     # 默认第8列=Y坐标

    print(f"  桩顶绝对标高列: {abs_col}, X列: {x_col}, Y列: {y_col}")

    # ====== 第5步：逐桩插值 ======
    print("🔢 IDW插值计算...")
    header_row = 1  # 假设表头在第1行
    data_start = header_row + 1

    pile_count = 0
    results = []
    for r in range(data_start, ws.max_row + 1):
        x = ws.cell(row=r, column=x_col).value
        y = ws.cell(row=r, column=y_col).value
        elev_abs = ws.cell(row=r, column=abs_col).value

        if not isinstance(x, (int, float)) or not isinstance(y, (int, float)):
            continue

        rock_elev = idw_interpolate(interp_points, x, y, k=k, power=power)
        if rock_elev is None:
            continue

        rock_elev = round(float(rock_elev), 2)

        if isinstance(elev_abs, (int, float)):
            depth = round(elev_abs - rock_elev, 2)
        else:
            depth = None

        results.append((r, rock_elev, depth))
        pile_count += 1

    print(f"  ✓ 插值完成: {pile_count} 根桩")

    # ====== 第6步：写入Excel新列 ======
    print("📝 写入Excel新列...")

    # 确定写入位置
    if column_offset is not None:
        start_col = column_offset + 1
    else:
        start_col = ws.max_column + 1

    rock_col = start_col
    depth_col = start_col + 1

    # 写表头
    header_style = Font(name='宋体', size=11, bold=True)
    ws.cell(row=header_row, column=rock_col, value=f'{sheet_type}面\n标高(m)')
    ws.cell(row=header_row, column=rock_col).font = header_style
    ws.cell(row=header_row, column=rock_col).alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)

    ws.cell(row=header_row, column=depth_col, value=f'进入{sheet_type}\n深度(m)')
    ws.cell(row=header_row, column=depth_col).font = header_style
    ws.cell(row=header_row, column=depth_col).alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)

    # 写数据
    body_font = Font(name='宋体', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for r, rock_elev, depth in results:
        ws.cell(row=r, column=rock_col, value=rock_elev)
        ws.cell(row=r, column=depth_col, value=depth)
        for c in [rock_col, depth_col]:
            ws.cell(row=r, column=c).font = body_font
            ws.cell(row=r, column=c).alignment = center_align
            ws.cell(row=r, column=c).border = thin_border

    # 调整列宽
    ws.column_dimensions['A'].width = 10

    # ====== 第7步：统计分析 ======
    vals = [d for _, _, d in results if d is not None]
    print(f"\n=== {sheet_type}统计 ===")
    print(f"岩面标高: {min(r for _, r, _ in results if r is not None):.2f}"
          f" ~ {max(r for _, r, _ in results if r is not None):.2f}m")
    print(f"进入深度: {min(vals):.2f} ~ {max(vals):.2f}m"
          f" (均值 {np.mean(vals):.2f}m)")

    # ====== 第8步：更新汇总Sheet ======
    if '汇总' in wb.sheetnames:
        ws2 = wb['汇总']
        r = ws2.max_row + 3

        ws2.cell(row=r, column=1, value=f"{sheet_type}面标高统计").font = Font(
            name='宋体', size=11, bold=True)
        r += 1
        rock_elevs = [rv for _, rv, _ in results if rv is not None]
        ws2.cell(row=r, column=1,
                 value=f"范围: {min(rock_elevs):.2f} ~ {max(rock_elevs):.2f}m"
                       f"（均值 {np.mean(rock_elevs):.2f}m）")
        r += 1
        ws2.cell(row=r, column=1,
                 value=f"进入{sheet_type}深度: {min(vals):.2f} ~ {max(vals):.2f}m"
                       f"（均值 {np.mean(vals):.2f}m）")
        r += 1

        # 深度分档统计
        bins = [(0, 10), (10, 20), (20, 30), (30, 40), (40, 50), (50, 100)]
        for low, high in bins:
            count = sum(1 for v in vals if low <= v < high)
            if count > 0:
                ws2.cell(row=r, column=1, value=f"  {low}-{high}m:")
                ws2.cell(row=r, column=2, value=f"{count}根")
                r += 1

    wb.save(excel_path)
    print(f"\n✓ 输出已更新: {excel_path}")


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    dxf_path = sys.argv[1]
    excel_path = sys.argv[2]

    layer = 'g_sDgxLayer'
    sheet_type = '岩层'
    k, power = 12, 2
    column_offset = None

    i = 3
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg == '--layer' and i + 1 < len(sys.argv):
            layer = sys.argv[i + 1]; i += 1
        elif arg == '--sheet' and i + 1 < len(sys.argv):
            sheet_type = sys.argv[i + 1]; i += 1
        elif arg == '--k' and i + 1 < len(sys.argv):
            k = int(sys.argv[i + 1]); i += 1
        elif arg == '--power' and i + 1 < len(sys.argv):
            power = float(sys.argv[i + 1]); i += 1
        elif arg == '--column-offset' and i + 1 < len(sys.argv):
            column_offset = int(sys.argv[i + 1]); i += 1
        i += 1

    main(dxf_path, excel_path, layer, sheet_type, k, power, column_offset)
