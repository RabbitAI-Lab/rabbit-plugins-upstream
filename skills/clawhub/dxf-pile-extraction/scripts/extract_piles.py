#!/usr/bin/env python3
"""
从桩基础施工图DXF中提取桩参数：编号、坐标、桩径、桩顶标高。
输出Excel参数表（含桩参数Sheet + 汇总Sheet）。
"""
import sys
import re
import ezdxf
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict


def extract_block_circles(doc):
    """从含CIRCLE的匿名块中提取桩（模式1：块内圆）。"""
    pile_blocks = {}
    for block in doc.blocks:
        circles = [e for e in block if e.dxftype() == 'CIRCLE']
        if circles:
            pile_blocks[block.name] = [c for c in circles]
    if not pile_blocks:
        return None, "块内无CIRCLE"

    # 遍历INSERT，提取桩心世界坐标
    piles = []
    for insert in doc.modelspace().query('INSERT'):
        if insert.dxf.name not in pile_blocks:
            continue
        block = doc.blocks[insert.dxf.name]
        for ent in block:
            if ent.dxftype() != 'CIRCLE':
                continue
            lx = ent.dxf.center.x       # 块内圆心X
            ly = ent.dxf.center.y       # 块内圆心Y
            r = ent.dxf.radius           # 块内半径
            # 世界坐标 = 插入点 + 块内坐标 × 缩放
            wx = insert.dxf.insert.x + lx * insert.dxf.xscale
            wy = insert.dxf.insert.y + ly * insert.dxf.yscale
            wr = r * insert.dxf.xscale
            piles.append({
                'x': round(wx, 3), 'y': round(wy, 3),
                'r': round(wr, 1), 'diameter': round(wr * 2, 0),
                'layer': insert.dxf.layer
            })
    return piles, f"块内圆×{len(piles)}"


def extract_polyline_piles(doc):
    """从LWPOLYLINE提取桩（模式2：45°直径符号线，P1=桩心）。"""
    piles = []
    for ent in doc.modelspace():
        if ent.dxftype() != 'LWPOLYLINE':
            continue
        pts = list(ent.get_points())
        if len(pts) != 2:
            continue
        p1, p2 = pts[0], pts[1]
        x1, y1 = p1[0], p1[1]
        x2, y2 = p2[0], p2[1]
        length = ((x2 - x1) ** 2 + (y2 - y1) ** 2) ** 0.5
        dia = length / (2 ** 0.5)  # 桩径 = 线段长/√2
        piles.append({
            'x': round(x1, 3), 'y': round(y1, 3),
            'r': round(dia / 2, 1), 'diameter': round(dia, 0),
            'layer': ent.dxf.layer
        })
    return piles


def extract_number_labels(doc):
    """提取桩编号标签。"""
    labels = []
    for ent in doc.modelspace():
        etype = ent.dxftype()
        if etype not in ('TEXT', 'MTEXT'):
            continue
        try:
            layer = ent.dxf.layer
        except Exception:
            continue
        if '编号' not in layer and 'BH' not in layer and '桩' not in layer:
            continue
        text = (ent.dxf.text if etype == 'TEXT'
                else ent.text).strip()
        if not text:
            continue
        # 匹配 Z1~Z999 格式
        m = re.match(r'^(Z\d+|P\d+|[A-Z]?\d+)$', text)
        if not m:
            continue
        x = ent.dxf.insert.x
        y = ent.dxf.insert.y
        labels.append({'num': text, 'x': round(x, 3), 'y': round(y, 3),
                       'etype': etype, 'layer': layer})
    return labels


def extract_type_labels(doc):
    """提取桩型+标高标签（GZH01,H=-6.500格式）。"""
    labels = []
    for ent in doc.modelspace():
        if ent.dxftype() != 'TEXT':
            continue
        try:
            layer = ent.dxf.layer
        except Exception:
            continue
        if 'BH' not in layer and 'BASE' not in layer:
            continue
        text = ent.dxf.text.strip()
        m = re.match(r'(GZH\d+|YZ\d+|PHC\d+).*?[Hh]\s*=\s*(-?\d+\.?\d*)', text)
        if not m:
            continue
        labels.append({
            'type': m.group(1),
            'elev': float(m.group(2)),  # 相对标高
            'x': round(ent.dxf.insert.x, 3),
            'y': round(ent.dxf.insert.y, 3),
            'layer': layer
        })
    return labels


def extract_type_table(doc):
    """从TEXT_NOTE图层提取桩型参数表。"""
    type_map = {}
    notes = []
    for ent in doc.modelspace():
        if ent.dxftype() != 'TEXT':
            continue
        try:
            if 'NOTE' not in ent.dxf.layer and 'TEXT' not in ent.dxf.layer:
                continue
        except Exception:
            continue
        notes.append((ent.dxf.insert.y, ent.dxf.text.strip()))

    # 查找 GZHxx D=xxx 格式的行
    for _, text in sorted(notes, key=lambda x: -x[0]):
        m = re.match(r'(GZH\d+|YZ\d+|PHC\d+).*?[Dd]\s*=\s*(\d+)', text)
        if m:
            ttype = m.group(1)
            dia = int(m.group(2))
            # 提取 Ra（如果能找到）
            ra_m = re.search(r'[Rr]a\s*=\s*(\d+)\s*k?[Nn]', text)
            c_m = re.search(r'[Cc](\d+)\s*\(\s*P?\d*\s*\)', text)
            type_map[ttype] = {
                'diameter': dia,
                'ra': int(ra_m.group(1)) if ra_m else None,
                'concrete': c_m.group(0) if c_m else None,
            }
    return type_map


def greedy_match(source, targets, key):
    """贪心最优匹配：一一配对，避免冲突。"""
    # 计算所有距离对
    pairs = []
    for i, s in enumerate(source):
        for j, t in enumerate(targets):
            dist = ((s['x'] - t['x']) ** 2 + (s['y'] - t['y']) ** 2) ** 0.5
            pairs.append((dist, i, j))
    pairs.sort()

    used_s = set()
    used_t = set()
    result = {}
    for dist, si, tj in pairs:
        if si in used_s or tj in used_t:
            continue
        used_s.add(si)
        used_t.add(tj)
        result[si] = (tj, dist)
    return result


def format_excel(ws, row_count):
    """套用Excel样式。"""
    header_font = Font(name='宋体', size=11, bold=True)
    body_font = Font(name='宋体', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    for row in ws.iter_rows(min_row=1, max_row=row_count):
        for cell in row:
            cell.font = body_font
            cell.alignment = center_align
            cell.border = thin_border

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # 列宽
    col_widths = {1: 6, 2: 8, 3: 10, 4: 10, 5: 14, 6: 14, 7: 12, 8: 12}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def main(dxf_path, output_path, datum=23.40):
    print(f"📖 读取DXF: {dxf_path}")
    doc = ezdxf.readfile(dxf_path)

    # ====== 第1步：提取桩轮廓 ======
    print("📐 提取桩轮廓...")
    pile_outlines, msg = extract_block_circles(doc)
    if pile_outlines is None:
        print(f"  ⚠️ {msg}，尝试LWPOLYLINE模式...")
        pile_outlines = extract_polyline_piles(doc)
    print(f"  ✓ 提取到 {len(pile_outlines)} 个桩轮廓")

    if not pile_outlines:
        print("❌ 未检测到任何桩轮廓，退出")
        return

    # ====== 第2步：提取编号标签 ======
    print("#️⃣ 提取桩编号...")
    number_labels = extract_number_labels(doc)
    print(f"  ✓ 提取到 {len(number_labels)} 个编号")

    # ====== 第3步：提取桩型标签 ======
    print("🏷️ 提取桩型/标高标签...")
    type_labels = extract_type_labels(doc)
    print(f"  ✓ 提取到 {len(type_labels)} 个桩型标签")

    # ====== 第4步：匹配 ======
    print("🔗 匹配编号↔轮廓...")
    num_match = greedy_match(number_labels, pile_outlines, 'num')
    # 编号匹配：距离阈值约3m（MTEXT插入点偏移）
    valid_matches = {si: (tj, d) for si, (tj, d) in num_match.items() if d < 5.0}
    print(f"  ✓ {len(valid_matches)}/{len(number_labels)} 个编号匹配成功")

    print("🔗 匹配桩型↔轮廓...")
    type_match = greedy_match(type_labels, pile_outlines, 'type')
    valid_type_matches = {si: (tj, d) for si, (tj, d) in type_match.items() if d < 5.0}
    print(f"  ✓ {len(valid_type_matches)}/{len(type_labels)} 个桩型匹配成功")

    # ====== 第5步：提取桩型表 ======
    type_table = extract_type_table(doc)
    print(f"📋 桩型参数表: {len(type_table)} 种桩型")
    for t, info in sorted(type_table.items()):
        print(f"  {t}: D={info['diameter']}mm, Ra={info['ra']}, {info['concrete']}")

    # ====== 第6步：构建桩数据 ======
    print("📊 构建桩参数表...")
    # 以轮廓为基准，关联编号和桩型
    # 先建立 轮廓索引 → 编号 映射
    outline_to_num = {}
    for si, (tj, _) in valid_matches.items():
        outline_to_num[tj] = number_labels[si]['num']

    # 轮廓索引 → 桩型
    outline_to_type = {}
    for si, (tj, _) in valid_type_matches.items():
        outline_to_type[tj] = {
            'type': type_labels[si]['type'],
            'elev_rel': type_labels[si]['elev'],
        }

    pile_data = []
    for i, outline in enumerate(pile_outlines):
        num = outline_to_num.get(i, f"?{i+1}")
        tinfo = outline_to_type.get(i, {})
        ptype = tinfo.get('type', '?')
        elev_rel = tinfo.get('elev_rel', None)

        # 桩径：优先用参数表，否则用轮廓圆径
        if ptype in type_table:
            diameter = type_table[ptype]['diameter']
        else:
            diameter = int(outline['diameter'])

        elev_abs = round(elev_rel + datum, 3) if elev_rel is not None else None

        pile_data.append({
            'num': num,
            'type': ptype,
            'diameter': diameter,
            'elev_rel': elev_rel,
            'elev_abs': elev_abs,
            'x': outline['x'],
            'y': outline['y'],
        })

    # 按Y降序+X升序排列
    pile_data.sort(key=lambda p: (-p['y'], p['x']))

    # 重新编号
    for idx, p in enumerate(pile_data, 1):
        p['seq'] = idx

    print(f"  ✓ 总桩数: {len(pile_data)}")

    # ====== 第7步：输出Excel ======
    print(f"📝 输出Excel: {output_path}")
    wb = Workbook()
    ws = wb.active
    ws.title = "桩参数表"

    # 表头
    headers = ['序号', '桩编号', '桩型', '桩径(mm)',
                '桩顶相对\n标高(m)', '桩顶绝对\n标高(m)',
                '桩心坐标X', '桩心坐标Y']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    # 数据行
    for i, p in enumerate(pile_data):
        r = i + 2
        ws.cell(row=r, column=1, value=p['seq'])
        ws.cell(row=r, column=2, value=p['num'])
        ws.cell(row=r, column=3, value=p['type'])
        ws.cell(row=r, column=4, value=p['diameter'])
        ws.cell(row=r, column=5, value=p['elev_rel'] if p['elev_rel'] is not None else '')
        ws.cell(row=r, column=6, value=p['elev_abs'] if p['elev_abs'] is not None else '')
        ws.cell(row=r, column=7, value=p['x'])
        ws.cell(row=r, column=8, value=p['y'])

    format_excel(ws, len(pile_data) + 2)

    # ====== Sheet2: 汇总 ======
    ws2 = wb.create_sheet("汇总")

    ws2.cell(row=1, column=1, value="桩基础参数汇总").font = Font(name='宋体', size=14, bold=True)
    ws2.merge_cells('A1:B1')

    r = 3
    ws2.cell(row=r, column=1, value="总体").font = Font(name='宋体', size=11, bold=True)
    r += 1
    stats = [
        ("总桩数", len(pile_data)),
        ("桩型种类", len(set(p['type'] for p in pile_data))),
        ("标高基准", f"±0.00=绝对{datum}m"),
    ]
    for label, val in stats:
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=val)
        r += 1

    # 桩型分布
    r += 1
    ws2.cell(row=r, column=1, value="桩型分布").font = Font(name='宋体', size=11, bold=True)
    r += 1
    ws2.cell(row=r, column=1, value="桩型")
    ws2.cell(row=r, column=2, value="桩数")
    ws2.cell(row=r, column=3, value="桩径(mm)")
    ws2.cell(row=r, column=4, value="占比")
    r += 1
    from collections import Counter
    type_counts = Counter(p['type'] for p in pile_data)
    for t, c in sorted(type_counts.items()):
        dia = type_table.get(t, {}).get('diameter', '?')
        ratio = f"{c / len(pile_data) * 100:.1f}%"
        ws2.cell(row=r, column=1, value=t)
        ws2.cell(row=r, column=2, value=c)
        ws2.cell(row=r, column=3, value=dia)
        ws2.cell(row=r, column=4, value=ratio)
        r += 1

    # 标高分布
    r += 1
    ws2.cell(row=r, column=1, value="桩顶标高分布（相对）").font = Font(name='宋体', size=11, bold=True)
    r += 1
    elev_counts = Counter(p['elev_rel'] for p in pile_data if p['elev_rel'] is not None)
    for elev, count in sorted(elev_counts.items()):
        abs_elev = round(elev + datum, 3)
        ws2.cell(row=r, column=1, value=f"{elev}m (绝对{abs_elev}m)")
        ws2.cell(row=r, column=2, value=f"{count}根")
        r += 1

    # 为汇总表套样式
    for row in ws2.iter_rows(min_row=1, max_row=r):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for col_letter in ['A', 'B', 'C', 'D']:
        ws2.column_dimensions[col_letter].width = 25

    wb.save(output_path)

    # ====== 核查报告 ======
    print("\n=== 核查报告 ===")
    print(f"桩总数: {len(pile_data)}")
    # 检查重复坐标
    coords = [(p['x'], p['y']) for p in pile_data]
    dupes = []
    for i, c in enumerate(coords):
        for j in range(i+1, len(coords)):
            if abs(c[0]-coords[j][0]) < 0.01 and abs(c[1]-coords[j][1]) < 0.01:
                dupes.append((pile_data[i]['num'], pile_data[j]['num'], c))
    if dupes:
        print(f"⚠️ 重复坐标桩: {len(dupes)} 对")
        for a, b, c in dupes[:5]:
            print(f"  {a} ↔ {b} @({c[0]},{c[1]})")
    else:
        print("✓ 无重复坐标桩")

    print(f"\n输出文件: {output_path}")


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(__doc__)
        print("用法: python extract_piles.py <dxf路径> <输出Excel.xlsx> [--datum 23.40]")
        sys.exit(1)

    dxf_path = sys.argv[1]
    output_path = sys.argv[2]
    datum = 23.40

    for i in range(3, len(sys.argv)):
        if sys.argv[i] == '--datum' and i + 1 < len(sys.argv):
            datum = float(sys.argv[i + 1])

    main(dxf_path, output_path, datum)
