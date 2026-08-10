#!/usr/bin/env python3
"""
在DXF桩基平面图上标注桩编号（新图层）。
读取桩参数Excel中的桩编号+坐标，在DXF中创建新图层添加TEXT标注。

用法:
    python label_piles.py <输入.dxf> <桩编号Excel.xlsx> <输出带编号.dxf>
"""
import sys
import ezdxf
from openpyxl import load_workbook


def main(dxf_path, excel_path, output_path):
    print(f"📖 读取DXF: {dxf_path}")
    doc = ezdxf.readfile(dxf_path)
    msp = doc.modelspace()

    print(f"📖 读取桩参数: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # 找表头列
    num_col, x_col, y_col, dia_col = None, None, None, None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=c).value or '')
        if h == '桩编号' or '编号' in h:
            num_col = c
        if '坐标X' in h or ('X' in h and '坐标' in str(
                ws.cell(row=1, column=c + 1).value or '')):
            x_col = c
        if '坐标Y' in h:
            y_col = c
        if '桩径' in h or '桩径(mm)' in h:
            dia_col = c

    if None in (num_col, x_col, y_col):
        print("❌ 未找到必需列（桩编号、坐标X/Y），请检查表头")
        return

    # 读取桩数据
    piles = []
    for r in range(2, ws.max_row + 1):
        num = ws.cell(row=r, column=num_col).value
        x = ws.cell(row=r, column=x_col).value
        y = ws.cell(row=r, column=y_col).value
        dia = ws.cell(row=r, column=dia_col).value if dia_col else 800
        if isinstance(num, str) and isinstance(x, (int, float)):
            piles.append((num, float(x), float(y), int(dia) if dia else 800))

    print(f"  读取到 {len(piles)} 根桩")

    # 创建标注图层
    layer_name = 'PILE-NUMBER'
    if layer_name not in [l.dxf.name for l in doc.layers]:
        new_layer = doc.layers.new(name=layer_name)
        new_layer.color = 1  # 红色
        print(f"  创建新图层: {layer_name} (红色)")

    # 添加标注TEXT
    text_height = 210  # 文字高度（图纸单位）
    x_offset = 500     # 桩心右偏移
    y_offset = 300     # 桩心上偏移

    added = 0
    for num, px, py, dia in piles:
        # 偏移量与桩径成正比
        scaled_x_offset = x_offset * (dia / 800)
        scaled_y_offset = y_offset * (dia / 800)

        tx, ty = px + scaled_x_offset, py + scaled_y_offset

        text_entity = msp.add_text(
            num,
            dxfattribs={
                'layer': layer_name,
                'insert': (tx, ty, 0),
                'height': text_height,
                'color': 1,  # 红色
            }
        )
        added += 1

    doc.saveas(output_path)
    print(f"✓ 已标注 {added} 根桩 → {output_path}")


if __name__ == '__main__':
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2], sys.argv[3])
