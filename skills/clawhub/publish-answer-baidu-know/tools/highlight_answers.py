#!/usr/bin/env python3
"""Highlight correct answer options in Excel quiz file."""

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, InlineFont, TextBlock

# Option letter to index mapping
LETTER_TO_IDX = {"A": 0, "B": 1, "C": 2, "D": 3}
IDX_TO_LETTER = {0: "A", 1: "B", 2: "C", 3: "D"}

# Regex to find option boundaries: A/B/C/D followed by separator
OPTION_MARKER_RE = re.compile(
    r"(?=[A-D](?:[\.．、:：\)）]|\s))"
)

NORMAL_FONT = InlineFont(b=False, color="000000")
RED_BOLD_FONT = InlineFont(b=True, color="FF0000")


def parse_answer(answer_raw) -> set[int]:
    """Parse answer column into set of option indices (0=A, 1=B, ...)."""
    if answer_raw is None:
        return set()
    s = str(answer_raw).strip()
    if not s:
        return set()

    indices = set()

    # Pure digits like "13" -> 1 and 3 (not 13)
    if re.fullmatch(r"[1-4]+", s):
        for ch in s:
            indices.add(int(ch) - 1)
        return indices

    # Split by common separators
    parts = re.split(r"[,，;；\s|]+", s)
    for part in parts:
        part = part.strip().upper()
        if not part:
            continue
        if part in LETTER_TO_IDX:
            indices.add(LETTER_TO_IDX[part])
        elif re.fullmatch(r"[1-4]", part):
            indices.add(int(part) - 1)

    return indices


def find_option_spans(text: str) -> list[tuple[int, int, str]]:
    """Find spans for each A/B/C/D option in the options text."""
    if not text or not isinstance(text, str):
        return []

    markers = []
    for m in OPTION_MARKER_RE.finditer(text):
        letter = text[m.start()]
        markers.append((m.start(), letter))

    if not markers:
        return []

    spans = []
    for i, (start, letter) in enumerate(markers):
        end = markers[i + 1][0] if i + 1 < len(markers) else len(text)
        spans.append((start, end, letter))
    return spans


def build_rich_text(options_text: str, answer_indices: set[int]) -> CellRichText | str:
    """Build rich text with correct options in red bold."""
    spans = find_option_spans(options_text)
    if not spans:
        return options_text

    parts = []
    pos = 0
    for start, end, letter in spans:
        if start > pos:
            parts.append(TextBlock(NORMAL_FONT, options_text[pos:start]))
        idx = LETTER_TO_IDX.get(letter)
        font = RED_BOLD_FONT if idx is not None and idx in answer_indices else NORMAL_FONT
        parts.append(TextBlock(font, options_text[start:end]))
        pos = end

    if pos < len(options_text):
        parts.append(TextBlock(NORMAL_FONT, options_text[pos:]))

    return CellRichText(*parts)


def load_workbook_from_xls(xls_path: Path) -> openpyxl.Workbook:
    """Load .xls file, converting via pandas/xlrd if needed."""
    try:
        import pandas as pd

        xls_data = pd.read_excel(xls_path, sheet_name=None, header=None, dtype=str)
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for sheet_name, df in xls_data.items():
            ws = wb.create_sheet(title=str(sheet_name))
            for r_idx, row in enumerate(df.itertuples(index=False), start=1):
                for c_idx, val in enumerate(row, start=1):
                    if val is not None and str(val) != "nan":
                        ws.cell(row=r_idx, column=c_idx, value=str(val))
        return wb
    except Exception as e:
        raise RuntimeError(f"Failed to read .xls file: {e}") from e


def find_header_columns(ws) -> tuple[int, int, int] | None:
    """Find header row and column indices for 题目选项 and 题目答案."""
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        options_col = None
        answer_col = None
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            s = str(val).strip()
            if s == "题目选项":
                options_col = col_idx
            elif s == "题目答案":
                answer_col = col_idx
        if options_col is not None and answer_col is not None:
            return row_idx, options_col, answer_col
    return None


def process_sheet(ws, stats: dict) -> bool:
    """Process a single worksheet. Returns True if processed."""
    header_info = find_header_columns(ws)
    if header_info is None:
        return False

    header_row, options_col, answer_col = header_info
    stats["sheets"] += 1
    sheet_rows = 0
    sheet_matches = 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        options_cell = ws.cell(row=row_idx, column=options_col)
        answer_cell = ws.cell(row=row_idx, column=answer_col)

        options_text = options_cell.value
        if options_text is None:
            continue
        options_text = str(options_text).strip()
        if not options_text:
            continue

        answer_indices = parse_answer(answer_cell.value)
        if not answer_indices:
            continue

        sheet_rows += 1
        sheet_matches += len(answer_indices)

        rich = build_rich_text(options_text, answer_indices)
        options_cell.value = rich

    stats["rows"] += sheet_rows
    stats["matches"] += sheet_matches
    return True


def main():
    src = Path(r"D:\下载\安全知识竞赛题库(1).xls")
    dst = Path(r"D:\下载\安全知识竞赛题库_答案标红.xlsx")

    if not src.exists():
        print(f"ERROR: Source file not found: {src}")
        sys.exit(1)

    print(f"Reading: {src}")
    wb = load_workbook_from_xls(src)

    stats = {"sheets": 0, "rows": 0, "matches": 0}
    for ws in wb.worksheets:
        process_sheet(ws, stats)

    wb.save(dst)
    print(f"Saved: {dst}")
    print(f"处理结果: {stats['sheets']} 个工作表, {stats['rows']} 行, {stats['matches']} 个匹配选项")


if __name__ == "__main__":
    main()
