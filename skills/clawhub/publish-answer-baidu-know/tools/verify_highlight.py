#!/usr/bin/env python3
"""Verify highlighted answer output file."""

from pathlib import Path

import openpyxl

dst = Path(r"D:\下载\安全知识竞赛题库_答案标红.xlsx")
src = Path(r"D:\下载\安全知识竞赛题库(1).xls")

print(f"Source exists: {src.exists()}")
print(f"Output exists: {dst.exists()}")

wb = openpyxl.load_workbook(dst, rich_text=True)
ws = wb.active
print(f"Sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")

# Find header columns
header_row = None
options_col = None
answer_col = None
for row_idx in range(1, 20):
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val and str(val).strip() == "题目选项":
            options_col = col_idx
        if val and str(val).strip() == "题目答案":
            answer_col = col_idx
    if options_col and answer_col:
        header_row = row_idx
        break

print(f"Header row={header_row}, options_col={options_col}, answer_col={answer_col}")

# Sample verification for first 5 rows with answers 1,2,3
samples = []
for row_idx in range(header_row + 1, ws.max_row + 1):
    answer = ws.cell(row=row_idx, column=answer_col).value
    if answer in ("1", "2", "3", 1, 2, 3):
        cell = ws.cell(row=row_idx, column=options_col)
        val = cell.value
        runs = []
        if hasattr(val, "__iter__") and not isinstance(val, str):
            for block in val:
                font = block.font
                bold = font.b if font else None
                color = font.color if font else None
                runs.append((block.text, bold, color))
        samples.append((row_idx, answer, runs))
        if len(samples) >= 5:
            break

for row_idx, answer, runs in samples:
    print(f"\nRow {row_idx}, answer={answer}:")
    for text, bold, color in runs:
        marker = "RED_BOLD" if bold and color == "FF0000" else "normal"
        print(f"  [{marker}] {repr(text[:50])}")
