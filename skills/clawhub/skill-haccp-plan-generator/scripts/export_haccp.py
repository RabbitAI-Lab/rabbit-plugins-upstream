#!/usr/bin/env python3
"""
HACCP 计划表导出工具
将结构化的 HACCP 数据导出为 Excel 文件

用法：
    python export_haccp.py --product "产品名称" --data '{"ccps": [...]}'
    python export_haccp.py --input input.json --output output.xlsx
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("错误：缺少 openpyxl 库，请安装：pip install openpyxl==3.1.2")
    sys.exit(1)


def style_cell(cell, bold=False, size=11, color=None, bg_color=None, wrap=True, align="left"):
    """设置单元格样式"""
    cell.font = Font(bold=bold, size=size, color=color or "000000")
    cell.alignment = Alignment(vertical="top", wrap_text=wrap, horizontal=align)
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def export_haccp(product_name, ccps, output_path, info=None):
    """导出 HACCP 计划表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "HACCP计划表"
    
    date_str = datetime.now().strftime("%Y-%m-%d")
    
    # === 第1行：标题 ===
    ws.merge_cells("A1:L1")
    ws["A1"] = f"HACCP 计划表 - {product_name}"
    style_cell(ws["A1"], bold=True, size=14, align="center")
    
    # === 第2行：生成日期 ===
    ws.merge_cells("A2:L2")
    ws["A2"] = f"生成日期：{date_str}"
    style_cell(ws["A2"], size=10, align="left")
    
    # === 第3行：产品信息 ===
    if info:
        ws.merge_cells("A3:C3")
        ws["A3"] = f"规格：{info.get('spec', '[未提供]')}"
        ws.merge_cells("D3:F3")
        ws["D3"] = f"保质期：{info.get('shelf_life', '[未提供]')}"
        ws.merge_cells("G3:L3")
        ws["G3"] = f"储存条件：{info.get('storage', '[未提供]')}"
        for col in ['A3', 'D3', 'G3']:
            style_cell(ws[col], size=10)
    
    # === 第5行：表头 ===
    headers = [
        "NO", "过程名称", "危害类型", "关键限值", "控制措施",
        "监控对象", "监控方法", "监控频率", "监控人员",
        "纠偏行动", "验证", "记录"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        style_cell(cell, bold=True, color="FFFFFF", bg_color="4472C4", align="center")
    
    # === 数据行 ===
    for idx, ccp in enumerate(ccps):
        row = 6 + idx
        values = [
            ccp.get("no", ""),
            ccp.get("process_name", ""),
            ccp.get("hazard_type", ""),
            ccp.get("critical_limit", ""),
            ccp.get("control_measure", ""),
            ccp.get("monitor_object", ""),
            ccp.get("monitor_method", ""),
            ccp.get("monitor_frequency", ""),
            ccp.get("monitor_person", ""),
            ccp.get("corrective_action", ""),
            ccp.get("verification", ""),
            ccp.get("record", "")
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            style_cell(cell)
    
    # 保存文件
    wb.save(output_path)
    return output_path


def parse_arguments():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description="HACCP 计划表导出工具")
    parser.add_argument("--product", "-p", required=True, help="产品名称")
    parser.add_argument("--data", "-d", help="JSON 格式的 CCP 数据")
    parser.add_argument("--input", "-i", help="输入 JSON 文件路径")
    parser.add_argument("--output", "-o", help="输出 Excel 文件路径")
    parser.add_argument("--spec", "-s", help="产品规格")
    parser.add_argument("--shelf-life", help="保质期")
    parser.add_argument("--storage", help="储存条件")
    return parser.parse_args()


def main():
    args = parse_arguments()
    
    # 解析 CCP 数据
    if args.data:
        try:
            ccps = json.loads(args.data)
            if isinstance(ccps, dict):
                ccps = ccps.get("ccps", [])
        except json.JSONDecodeError:
            print("错误：JSON 格式无效")
            sys.exit(1)
    elif args.input:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"错误：输入文件不存在：{args.input}")
            sys.exit(1)
        with open(input_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            ccps = data.get("ccps", data) if isinstance(data, dict) else data
    else:
        print("错误：必须指定 --data 或 --input 参数")
        sys.exit(1)
    
    # 生成输出路径
    if args.output:
        output_path = args.output
    else:
        date_str = datetime.now().strftime("%Y%m%d")
        safe_name = args.product.replace("/", "_").replace("\\", "_")
        output_path = f"HACCP计划表_{safe_name}_{date_str}.xlsx"
    
    # 可选信息
    info = None
    if args.spec or args.shelf_life or args.storage:
        info = {
            "spec": args.spec or "[未提供]",
            "shelf_life": args.shelf_life or "[未提供]",
            "storage": args.storage or "[未提供]"
        }
    
    # 导出
    try:
        result = export_haccp(args.product, ccps, output_path, info)
        print(f"成功：HACCP 计划表已导出至 {result}")
        print(f"共 {len(ccps)} 个关键控制点")
    except Exception as e:
        print(f"错误：导出失败 - {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
