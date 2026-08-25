"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import json
import os
import platform
import subprocess
import sys
import tempfile
from pathlib import Path

from office.soffice import get_soffice_env

from openpyxl import load_workbook

MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

SCRIPT_XLC = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:libraries PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "libraries.dtd">
<library:libraries xmlns:library="http://openoffice.org/2000/library" xmlns:xlink="http://www.w3.org/1999/xlink">
 <library:library library:name="Standard" library:link="false"/>
</library:libraries>"""

SCRIPT_XBA_INDEX = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">
<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" library:readonly="false" library:passwordprotected="false">
 <library:element library:name="Module1"/>
</library:library>"""


def has_gtimeout():
    try:
        subprocess.run(
            ["gtimeout", "--version"], capture_output=True, timeout=1, check=False
        )
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def _write_macro_profile(profile_dir: Path) -> None:
    """Seed a throwaway LibreOffice user profile with the recalc macro.

    Uses a private, per-run profile directory (via -env:UserInstallation)
    instead of writing into the shared LibreOffice user config, so this
    never touches the caller's real LibreOffice installation.
    """
    basic_dir = profile_dir / "user" / "basic" / "Standard"
    basic_dir.mkdir(parents=True, exist_ok=True)
    (basic_dir / MACRO_FILENAME).write_text(RECALCULATE_MACRO)
    (basic_dir / "script.xlb").write_text(SCRIPT_XBA_INDEX)
    (profile_dir / "user" / "basic").joinpath("script.xlc").write_text(SCRIPT_XLC)


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    with tempfile.TemporaryDirectory(prefix="lo_recalc_profile_") as tmp_profile:
        profile_dir = Path(tmp_profile)
        try:
            _write_macro_profile(profile_dir)
        except Exception as e:
            return {"error": f"Failed to setup LibreOffice macro: {e}"}

        profile_url = f"file://{profile_dir.as_posix()}"

        cmd = [
            "soffice",
            "--headless",
            "--norestore",
            f"-env:UserInstallation={profile_url}",
            "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
            abs_path,
        ]

        if platform.system() == "Linux":
            cmd = ["timeout", str(timeout)] + cmd
        elif platform.system() == "Darwin" and has_gtimeout():
            cmd = ["gtimeout", str(timeout)] + cmd

        result = subprocess.run(
            cmd, capture_output=True, text=True, env=get_soffice_env()
        )

    if result.returncode != 0 and result.returncode != 124:
        error_msg = result.stderr or "Unknown error during recalculation"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            return {"error": "LibreOffice macro not configured properly"}
        return {"error": error_msg}

    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],  
                }

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_count += 1
        wb_formulas.close()

        result["total_formulas"] = formula_count

        return result

    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
