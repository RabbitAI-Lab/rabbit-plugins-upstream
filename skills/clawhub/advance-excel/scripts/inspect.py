"""
Excel file inspector — run before any read/write operation on an unknown file.

Returns JSON describing the file structure so the agent can choose the right
strategy before touching data.

Usage:
    python skills/xlsx/scripts/inspect.py <file.xlsx> [max_peek_rows]

Output:
    {
      "file": "report.xlsx",
      "has_vba": false,
      "sheets": ["Summary", "Raw Data"],
      "named_ranges": ["TotalRevenue", "DateRange"],
      "per_sheet": {
        "Raw Data": {
          "dimensions": "A1:Z5421",
          "row_count": 5421,
          "col_count": 26,
          "merged_ranges": ["A1:F1", "A2:A10"],
          "detected_header_row": 3,
          "sample_headers": ["SKU", "Description", "Units Sold"],
          "formula_count": 142,
          "has_images": false
        }
      },
      "recommendation": "merged_cells_detected"
    }
"""

import json
import sys
from pathlib import Path


def detect_header_row(ws, max_peek: int = 15) -> int:
    """Return 0-indexed row number most likely to be the header row.

    Heuristic: the row with the most non-null string values in the first
    max_peek rows.  Falls back to 0 if the sheet is empty.
    """
    best_row = 0
    best_score = -1
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_peek)):
        score = sum(
            1 for cell in row
            if cell.value is not None and isinstance(cell.value, str)
        )
        if score > best_score:
            best_score = score
            best_row = i
    return best_row


def inspect_sheet(ws, max_peek: int = 15) -> dict:
    result: dict = {}

    # Dimensions
    result["dimensions"] = ws.dimensions or "unknown"
    result["row_count"] = ws.max_row or 0
    result["col_count"] = ws.max_column or 0

    # Merged ranges
    merged = [str(r) for r in ws.merged_cells.ranges]
    result["merged_ranges"] = merged

    # Header row detection
    header_row_idx = detect_header_row(ws, max_peek)
    result["detected_header_row"] = header_row_idx  # 0-indexed, use as pandas header=N

    # Sample headers from detected row
    headers = []
    for cell in ws[header_row_idx + 1]:  # openpyxl rows are 1-indexed
        if cell.value is not None:
            headers.append(str(cell.value))
    result["sample_headers"] = headers[:20]

    # Formula count (scan up to 500 rows for speed)
    formula_count = 0
    scan_limit = min(ws.max_row or 0, 500)
    for row in ws.iter_rows(max_row=scan_limit):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1
    result["formula_count"] = formula_count

    # Images
    result["has_images"] = len(ws._images) > 0

    return result


def inspect(filepath: str, max_peek: int = 15) -> dict:
    from openpyxl import load_workbook

    path = Path(filepath)
    if not path.exists():
        return {"error": f"File not found: {filepath}"}

    suffix = path.suffix.lower()
    if suffix not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return {"error": f"Unsupported format: {suffix}. Use pandas for .csv/.tsv, pyxlsb for .xlsb."}

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return {"error": f"Failed to open file: {e}"}

    result: dict = {
        "file": path.name,
        "has_vba": suffix in (".xlsm", ".xltm"),
        "sheets": wb.sheetnames,
        "named_ranges": list(wb.defined_names.keys()) if wb.defined_names else [],
        "per_sheet": {},
    }

    for name in wb.sheetnames:
        try:
            result["per_sheet"][name] = inspect_sheet(wb[name], max_peek)
        except Exception as e:
            result["per_sheet"][name] = {"error": str(e)}

    wb.close()

    # Top-level recommendation
    recommendations = []
    for sheet_data in result["per_sheet"].values():
        if isinstance(sheet_data, dict):
            if sheet_data.get("merged_ranges"):
                recommendations.append("merged_cells_detected — unmerge_and_fill before pandas read")
            if sheet_data.get("detected_header_row", 0) > 0:
                recommendations.append(
                    f"header_not_at_row_0 — use pd.read_excel(..., header={sheet_data['detected_header_row']})"
                )
            if sheet_data.get("formula_count", 0) > 0:
                recommendations.append("formulas_present — run recalc.py after editing")
    result["recommendations"] = list(dict.fromkeys(recommendations))  # dedupe, preserve order

    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python inspect.py <file.xlsx> [max_peek_rows]")
        print("\nInspects an Excel file and returns structural JSON.")
        print("Run this before any read/write to choose the right strategy.")
        sys.exit(1)

    filepath = sys.argv[1]
    max_peek = int(sys.argv[2]) if len(sys.argv) > 2 else 15

    result = inspect(filepath, max_peek)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
