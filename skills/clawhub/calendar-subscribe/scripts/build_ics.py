#!/usr/bin/env python3
"""Timetable → subscription ICS.

Supports:
  --csv  date,start,end,title,location,notes
  --xlsx Chinese-university weekend grid (weeks as columns, slots as rows)
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

EXCEL_EPOCH = datetime(1899, 12, 30)

DEFAULT_SLOTS = {
    5: [("09:00", "10:30", "2-3节", "上午"), ("10:45", "12:15", "4-5节", "上午")],
    6: [("13:00", "14:30", "6-7节", "下午"), ("14:45", "16:15", "8-9节", "下午")],
    7: [("17:00", "18:30", "10-11节", "晚上"), ("18:45", "20:15", "12-13节", "晚上")],
}


def fold(line: str) -> str:
    if len(line.encode("utf-8")) <= 73:
        return line
    parts: list[str] = []
    buf = ""
    for ch in line:
        trial = buf + ch
        if buf and len(trial.encode("utf-8")) > 73:
            parts.append(buf)
            buf = ch
        else:
            buf = trial
    if buf:
        parts.append(buf)
    return "\r\n ".join(parts)


def ics_escape(text: str) -> str:
    return (
        text.replace("\\", "\\\\")
        .replace(";", "\\;")
        .replace(",", "\\,")
        .replace("\n", "\\n")
    )


def excel_date(serial) -> datetime | None:
    if serial is None:
        return None
    if isinstance(serial, datetime):
        return serial.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(serial, (int, float)):
        return EXCEL_EPOCH + timedelta(days=int(serial))
    return None


def slug(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "-", s).strip("-").lower()[:40] or "evt"


def load_csv(path: Path) -> list[dict]:
    events = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            date = (row.get("date") or "").strip()
            start = (row.get("start") or "").strip()
            end = (row.get("end") or "").strip()
            title = (row.get("title") or "").strip()
            if not (date and start and end and title):
                continue
            events.append(
                {
                    "summary": title,
                    "start": f"{date}T{start}:00" if len(start) == 5 else f"{date}T{start}",
                    "end": f"{date}T{end}:00" if len(end) == 5 else f"{date}T{end}",
                    "location": (row.get("location") or "").strip(),
                    "description": (row.get("notes") or "").strip(),
                    "date": date,
                    "travel_alarm": False,
                }
            )
    return events


def load_xlsx(path: Path, skip_re: re.Pattern, location: str) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    merged = {}
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        val = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged[(r, c)] = val

    def cell(r, c):
        v = ws.cell(r, c).value
        return v if v is not None else merged.get((r, c))

    events = []
    for col in range(2, (ws.max_column or 2) + 1):
        date = excel_date(cell(3, col))
        if date is None:
            continue
        weekday = str(cell(4, col) or "").strip()
        week = str(cell(2, col) or cell(2, col - 1) or "").strip()
        for row, slots in DEFAULT_SLOTS.items():
            raw = cell(row, col)
            if raw is None:
                continue
            text = re.sub(r"\s+", " ", str(raw)).strip()
            if not text or skip_re.search(text):
                continue
            special = None
            if "跨年" in text:
                special = ("17:30", "20:30", "活动", "晚上")
            used = [special] if special else slots
            for start_s, end_s, jie, period in used:
                sh, sm = map(int, start_s.split(":"))
                eh, em = map(int, end_s.split(":"))
                start = date.replace(hour=sh, minute=sm)
                end = date.replace(hour=eh, minute=em)
                events.append(
                    {
                        "summary": f"{text}（{jie}）" if not special else text[:40],
                        "start": start.isoformat(),
                        "end": end.isoformat(),
                        "location": location,
                        "description": f"{text}\n{week} {weekday} {jie} {period}",
                        "date": date.strftime("%Y-%m-%d"),
                        "travel_alarm": False,
                    }
                )
    return events


def mark_first_of_day(events: list[dict]) -> None:
    by: dict[str, list[dict]] = defaultdict(list)
    for ev in events:
        by[ev["date"]].append(ev)
    for items in by.values():
        first = min(items, key=lambda e: e["start"])
        first["travel_alarm"] = True


def to_ics(events: list[dict], name: str, tz: str, travel_minutes: int) -> str:
    now = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    lines = [
        "BEGIN:VCALENDAR",
        "PRODID:-//calendar-subscribe//EN",
        "VERSION:2.0",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:{ics_escape(name)}",
        f"X-WR-TIMEZONE:{tz}",
        "REFRESH-INTERVAL;VALUE=DURATION:P1D",
        "X-PUBLISHED-TTL:P1D",
        "BEGIN:VTIMEZONE",
        f"TZID:{tz}",
        "BEGIN:STANDARD",
        "TZOFFSETFROM:+0800",
        "TZOFFSETTO:+0800",
        "TZNAME:CST",
        "DTSTART:19700101T000000",
        "END:STANDARD",
        "END:VTIMEZONE",
    ]
    for ev in events:
        start = datetime.fromisoformat(ev["start"])
        end = datetime.fromisoformat(ev["end"])
        uid = f"{start.strftime('%Y%m%dT%H%M')}-{slug(ev['summary'])}@calendar-subscribe"
        lines += [
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{now}",
            f"DTSTART;TZID={tz}:{start.strftime('%Y%m%dT%H%M%S')}",
            f"DTEND;TZID={tz}:{end.strftime('%Y%m%dT%H%M%S')}",
            f"SUMMARY:{ics_escape(ev['summary'])}",
            f"LOCATION:{ics_escape(ev.get('location') or '')}",
            f"DESCRIPTION:{ics_escape(ev.get('description') or '')}",
            "STATUS:CONFIRMED",
            "TRANSP:OPAQUE",
        ]
        if ev.get("travel_alarm") and travel_minutes > 0:
            lines += [
                "BEGIN:VALARM",
                "ACTION:DISPLAY",
                "DESCRIPTION:Leave for class",
                f"TRIGGER:-PT{int(travel_minutes)}M",
                "END:VALARM",
            ]
        lines.append("END:VEVENT")
    lines.append("END:VCALENDAR")
    return "\r\n".join(fold(x) for x in lines) + "\r\n"


def main() -> None:
    ap = argparse.ArgumentParser()
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--xlsx")
    src.add_argument("--csv")
    ap.add_argument("--out", required=True)
    ap.add_argument("--name", default="Timetable")
    ap.add_argument("--location", default="")
    ap.add_argument("--tz", default="Asia/Shanghai")
    ap.add_argument("--travel-minutes", type=int, default=90)
    ap.add_argument("--skip", default=r"不上课|放假|补班|除夕")
    args = ap.parse_args()

    skip_re = re.compile(args.skip)
    if args.csv:
        events = load_csv(Path(args.csv))
        events = [e for e in events if not skip_re.search(e["summary"])]
    else:
        events = load_xlsx(Path(args.xlsx), skip_re, args.location)
    if not events:
        raise SystemExit("no events parsed")
    mark_first_of_day(events)
    events.sort(key=lambda e: e["start"])
    out = Path(args.out)
    out.write_bytes(to_ics(events, args.name, args.tz, args.travel_minutes).encode("utf-8"))
    days = {e["date"] for e in events}
    alarms = sum(1 for e in events if e["travel_alarm"])
    print(f"events={len(events)} days={len(days)} travel_alarms={alarms} ics={out}")


if __name__ == "__main__":
    main()
