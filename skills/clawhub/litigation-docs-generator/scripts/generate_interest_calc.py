#!/usr/bin/env python3
"""
利息损失计算表生成脚本 — v3（匹配 ☆利息计算表5.0 结构）

核心能力：
1. 内置完整利率数据库（2006-2026），涵盖央行基准利率 → LPR 全时段
2. 输出双重结构：利率数据库 + 利息计算表（与桌面 5.0 格式一致）
3. 未约定利率时自动使用 LPR 分段计算，利率变化处自动拆分
4. 所有计算列为 Excel 公式，可点击核验

用法：
  python3 generate_interest_calc.py <output.xlsx> \
    --principal 25583 --start 2025-01-22 --end 2026-07-10

  python3 generate_interest_calc.py <output.xlsx> \
    --principal 25583 --start 2025-01-22 --end 2026-07-10 --rate 3.1
"""

import sys
import argparse
import json
from datetime import datetime, date, timedelta
from dataclasses import dataclass
from typing import List, Optional, Tuple


# ======================== 利率数据库（完整内置） ========================

@dataclass
class RateRecord:
    effective_date: date
    rate_6m: float
    rate_6m_1y: float
    rate_1_3y: float
    rate_3_5y: float
    rate_5y_plus: float


RATE_DATABASE = [
    RateRecord(date(2006, 8, 19),  5.58, 6.12, 6.30, 6.48, 6.84),
    RateRecord(date(2007, 3, 18),  5.67, 6.39, 6.57, 6.75, 7.11),
    RateRecord(date(2007, 5, 19),  5.85, 6.57, 6.75, 6.93, 7.20),
    RateRecord(date(2007, 7, 21),  6.03, 6.84, 7.02, 7.20, 7.38),
    RateRecord(date(2007, 8, 22),  6.21, 7.02, 7.20, 7.38, 7.56),
    RateRecord(date(2007, 9, 15),  6.48, 7.29, 7.47, 7.65, 7.83),
    RateRecord(date(2007, 12, 21), 6.57, 7.47, 7.56, 7.74, 7.83),
    RateRecord(date(2008, 9, 16),  6.21, 7.20, 7.29, 7.56, 7.74),
    RateRecord(date(2008, 10, 9),  6.12, 6.93, 7.02, 7.29, 7.47),
    RateRecord(date(2008, 10, 30), 6.03, 6.66, 6.75, 7.02, 7.20),
    RateRecord(date(2008, 11, 27), 5.04, 5.58, 5.67, 5.94, 6.12),
    RateRecord(date(2008, 12, 23), 4.86, 5.31, 5.40, 5.76, 5.94),
    RateRecord(date(2010, 10, 20), 5.10, 5.56, 5.60, 5.96, 6.14),
    RateRecord(date(2010, 12, 26), 5.35, 5.81, 5.85, 6.22, 6.40),
    RateRecord(date(2011, 2, 9),   5.60, 6.06, 6.10, 6.45, 6.60),
    RateRecord(date(2011, 4, 6),   5.85, 6.31, 6.40, 6.65, 6.80),
    RateRecord(date(2011, 7, 7),   6.10, 6.56, 6.65, 6.90, 7.05),
    RateRecord(date(2012, 6, 8),   5.85, 6.31, 6.40, 6.65, 6.80),
    RateRecord(date(2012, 7, 6),   5.60, 6.00, 6.15, 6.40, 6.55),
    RateRecord(date(2014, 11, 22), 5.60, 5.60, 6.00, 6.00, 6.15),
    RateRecord(date(2015, 3, 1),   5.35, 5.35, 5.75, 5.75, 5.90),
    RateRecord(date(2015, 5, 11),  5.10, 5.10, 5.50, 5.50, 5.65),
    RateRecord(date(2015, 6, 28),  4.85, 4.85, 5.25, 5.25, 5.40),
    RateRecord(date(2015, 8, 26),  4.60, 4.60, 5.00, 5.00, 5.15),
    RateRecord(date(2015, 10, 24), 4.35, 4.35, 4.75, 4.75, 4.90),
    # LPR 时期
    RateRecord(date(2019, 8, 20),  4.25, 4.25, 4.25, 4.25, 4.85),
    RateRecord(date(2019, 9, 20),  4.20, 4.20, 4.20, 4.20, 4.85),
    RateRecord(date(2019, 10, 21), 4.20, 4.20, 4.20, 4.20, 4.85),
    RateRecord(date(2019, 11, 20), 4.15, 4.15, 4.15, 4.15, 4.80),
    RateRecord(date(2019, 12, 20), 4.15, 4.15, 4.15, 4.15, 4.80),
    RateRecord(date(2020, 1, 20),  4.15, 4.15, 4.15, 4.15, 4.80),
    RateRecord(date(2020, 2, 20),  4.05, 4.05, 4.05, 4.05, 4.75),
    RateRecord(date(2020, 3, 20),  4.05, 4.05, 4.05, 4.05, 4.75),
    RateRecord(date(2020, 4, 20),  3.85, 3.85, 3.85, 3.85, 4.65),
    RateRecord(date(2021, 12, 20), 3.80, 3.80, 3.80, 3.80, 4.65),
    RateRecord(date(2022, 1, 20),  3.70, 3.70, 3.70, 3.70, 4.60),
    RateRecord(date(2022, 5, 20),  3.70, 3.70, 3.70, 3.70, 4.45),
    RateRecord(date(2022, 8, 22),  3.65, 3.65, 3.65, 3.65, 4.30),
    RateRecord(date(2023, 6, 20),  3.55, 3.55, 3.55, 3.55, 4.20),
    RateRecord(date(2023, 8, 21),  3.45, 3.45, 3.45, 3.45, 4.20),
    RateRecord(date(2024, 2, 20),  3.45, 3.45, 3.45, 3.45, 3.95),
    RateRecord(date(2024, 7, 22),  3.35, 3.35, 3.35, 3.35, 3.85),
    RateRecord(date(2024, 10, 21), 3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2024, 11, 20), 3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2024, 12, 20), 3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2025, 1, 20),  3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2025, 2, 20),  3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2025, 3, 20),  3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2025, 4, 21),  3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2025, 5, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 6, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 7, 21),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 8, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 9, 22),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 10, 20), 3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 11, 20), 3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 12, 22), 3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2026, 1, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2026, 2, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2026, 3, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2026, 4, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2026, 5, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2026, 6, 22),  3.00, 3.00, 3.00, 3.00, 3.50),
]
RATE_DATABASE.sort(key=lambda r: r.effective_date)


# =============== 利率选择 / 分段逻辑 ===============

_RATE_ATTRS = ["rate_6m", "rate_6m_1y", "rate_1_3y", "rate_3_5y", "rate_5y_plus"]
_RATE_COL_NAMES = ["六个月以内", "六个月至一年", "一至三年", "三至五年", "五年以上"]
_RATE_TERM_MAP = {"6M": 0, "1Y": 1, "1-3Y": 2, "3-5Y": 3, "5Y": 4}


def get_rate(rec: RateRecord, term: str) -> float:
    idx = _RATE_TERM_MAP.get(term, 1)
    return getattr(rec, _RATE_ATTRS[idx])


def auto_select_term(start: date, end: date) -> str:
    years = (end - start).days / 365.25
    if years < 0.5:   return "6M"
    elif years < 1:   return "1Y"
    elif years < 3:   return "1-3Y"
    elif years < 5:   return "3-5Y"
    else:             return "5Y"


def applicable_rate(db: List[RateRecord], d: date, term: str) -> float:
    rec = db[0]
    for r in db:
        if r.effective_date <= d:
            rec = r
        else:
            break
    return get_rate(rec, term)


def split_periods(start: date, end: date, db: List[RateRecord],
                  term: str = "1Y", fixed_rate: float = None
                  ) -> List[Tuple[date, date, float]]:
    if fixed_rate is not None:
        return [(start, end, fixed_rate)]

    periods = []
    current = start
    for r in db:
        if r.effective_date <= current:
            continue
        if r.effective_date > end:
            rate = applicable_rate(db, current, term)
            periods.append((current, end, rate))
            current = end
            break
        period_end = r.effective_date - timedelta(days=1)
        if period_end >= current:
            rate = applicable_rate(db, current, term)
            periods.append((current, min(period_end, end), rate))
            current = r.effective_date
    if current < end:
        periods.append((current, end, applicable_rate(db, current, term)))

    # 合并相邻同利率段
    merged = []
    for ps, pe, pr in periods:
        if merged and merged[-1][2] == pr:
            merged[-1] = (merged[-1][0], pe, pr)
        else:
            merged.append((ps, pe, pr))
    return merged


# ======================== Excel 生成（双 sheet，匹配 5.0 结构） ========================

# openpyxl imports (shared)
try:
    from openpyxl import Workbook
    from openpyxl.workbook.properties import CalcProperties
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def generate_excel(output_path: str, items: list, submitter: str = "",
                   title: str = "利息自动计算表", rate_term: str = "1Y"):
    if not _HAS_OPENPYXL:
        print("错误：需要 openpyxl 库。pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    # 开启自动计算
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True)

    # ---- Sheet 1: 利息计算表（默认可见）----
    ws_calc = wb.active
    ws_calc.title = "利息计算表"
    _build_calc_sheet(ws_calc, items, submitter, title, rate_term)

    # ---- Sheet 2: 利率数据库 ----
    ws_db = wb.create_sheet("利率数据库")
    _build_rate_db_sheet(ws_db)

    wb.save(output_path)
    return output_path


def _build_rate_db_sheet(ws):
    """构建利率数据库 sheet（完整历史数据 + 公式）"""
    title_font = Font(name="宋体", size=11, bold=True)
    data_font = Font(name="宋体", size=10)
    header_font = Font(name="宋体", size=10, bold=True)
    thin = Border(left=Side('thin'), right=Side('thin'),
                  top=Side('thin'), bottom=Side('thin'))
    center = Alignment(horizontal="center", vertical="center")

    # 列宽
    widths = {"A": 14, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w

    # Row 1-2: 档位断点
    ws.cell(row=1, column=1, value="人民银行公布链接").font = Font(name="宋体", size=9, color="666666")
    for ci, (val, col) in enumerate([(0, 2), (0.5, 3), (1, 4), (3, 5), (5, 6)], 1):
        ws.cell(row=2, column=col, value=val).font = data_font

    # Row 3: 表头
    headers = ["调整时间", "六个月以内\n（含六个月）", "六个月至一年\n（含一年）",
               "一至三年\n（含三年）", "三至五年\n（含五年）", "五年以上"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = header_font; c.alignment = center; c.border = thin

    # Row 4+: 数据
    for i, rec in enumerate(RATE_DATABASE):
        r = i + 4
        ws.cell(row=r, column=1, value=rec.effective_date).font = data_font
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=1).number_format = 'yyyy-mm-dd'
        for ci, attr in enumerate(_RATE_ATTRS, 2):
            c = ws.cell(row=r, column=ci, value=getattr(rec, attr))
            c.font = data_font; c.alignment = center; c.number_format = '0.00'
            c.border = thin
        ws.cell(row=r, column=1).border = thin


def _build_calc_sheet(ws, items, submitter, title, rate_term):
    """构建利息计算表 sheet（匹配 5.0 结构）"""
    title_font = Font(name="宋体", size=14, bold=True)
    label_font = Font(name="宋体", size=11)
    header_font = Font(name="宋体", size=11, bold=True)
    data_font = Font(name="宋体", size=11)
    bold_font = Font(name="宋体", size=11, bold=True)
    thin = Border(left=Side('thin'), right=Side('thin'),
                  top=Side('thin'), bottom=Side('thin'))
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # 列宽
    for cw in [("A", 16), ("B", 16), ("C", 16), ("D", 16), ("E", 18)]:
        ws.column_dimensions[cw[0]].width = cw[1]

    row = 1
    # 标题
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=title)
    c.font = title_font; c.alignment = center
    row += 1

    for item_idx, item in enumerate(items):
        principal = item["principal"]
        start_str = item["start"]
        end_str = item["end"]
        fixed_rate = item.get("rate")
        item_term = item.get("term", rate_term)

        start_dt = datetime.strptime(start_str, "%Y-%m-%d").date()
        end_dt = datetime.strptime(end_str, "%Y-%m-%d").date()

        if fixed_rate is None and item.get("term") is None:
            item_term = auto_select_term(start_dt, end_dt)

        periods = split_periods(start_dt, end_dt, RATE_DATABASE, item_term, fixed_rate)

        # --- 输入区 ---
        if item_idx == 0:
            ws.cell(row=row, column=1, value="金额（元）：").font = label_font
            ws.cell(row=row, column=2, value=principal).font = label_font
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value="输入金额数字。").font = Font(name="宋体", size=9, color="999999")
            row += 1
            ws.cell(row=row, column=1, value="起始日：").font = label_font
            ws.cell(row=row, column=2, value=start_dt).font = label_font
            ws.cell(row=row, column=2).number_format = 'yyyy-mm-dd'
            ws.cell(row=row, column=3, value="输入格式如左。").font = Font(name="宋体", size=9, color="999999")
            row += 1
            ws.cell(row=row, column=1, value="终止日：").font = label_font
            ws.cell(row=row, column=2, value=end_dt).font = label_font
            ws.cell(row=row, column=2).number_format = 'yyyy-mm-dd'
            ws.cell(row=row, column=3, value="输入格式如左。").font = Font(name="宋体", size=9, color="999999")
            row += 1
        else:
            row += 3

        # 计算数据区起点
        # 当前 row=期间天数行,  +1=利息总计, +2=说明, +3=空行, +4=表头, +5=首条数据
        # 但利息总计行后有 row+=2（跳过了利率信息行），所以实际是 +6
        data_first = row + 6
        data_last = data_first + len(periods) - 1

        # 期间天数（row）
        ws.cell(row=row, column=1, value="期间天数（日）：").font = label_font
        ws.cell(row=row, column=2).value = f"=SUM(D{data_first}:D{data_last})"
        ws.cell(row=row, column=2).font = label_font
        ws.cell(row=row, column=3, value="自动计算，无需输入。").font = Font(name="宋体", size=9, color="999999")
        row += 1

        # 利息总计（row）
        ws.cell(row=row, column=1, value="利息总计（元）：").font = label_font
        ws.cell(row=row, column=2).value = f"=SUM(E{data_first}:E{data_last})"
        ws.cell(row=row, column=2).font = label_font
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value="自动计算，无需输入。").font = Font(name="宋体", size=9, color="999999")
        rate_note = f"约定利率：{fixed_rate}%（不分段）" if fixed_rate else f"利率档位：{item_term}（按 LPR 分段计算）"
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row=row, column=5, value=rate_note).font = Font(name="宋体", size=9, color="666666")
        row += 2

        # 说明行
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        note = ws.cell(row=row, column=1,
                       value="以下清单为自动生成，无需输入。说明：期间利息=金额×期间天数×期间利率÷360")
        note.font = Font(name="宋体", size=10, italic=True, color="666666")
        row += 1
        # 空行
        row += 1

        # 表头
        headers = ["期间起始日", "期间终止日", "期间利率（%）", "期间天数(日）", "期间利息（元）"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = header_font; c.alignment = center; c.border = thin
        row += 1

        # 数据行（全部使用公式）
        first_data_row = row
        for ps, pe, pr in periods:
            days = (pe - ps).days + 1
            if days <= 0:
                continue

            # A: 起始日
            ws.cell(row=row, column=1, value=ps).font = data_font
            ws.cell(row=row, column=1).alignment = center
            ws.cell(row=row, column=1).number_format = 'yyyy-mm-dd'

            # B: 终止日
            ws.cell(row=row, column=2, value=pe).font = data_font
            ws.cell(row=row, column=2).alignment = center
            ws.cell(row=row, column=2).number_format = 'yyyy-mm-dd'

            # C: 利率（数值，百分比格式）
            rc = ws.cell(row=row, column=3, value=pr / 100)
            rc.font = data_font; rc.alignment = center; rc.number_format = '0.00%'

            # D: 天数 = B-A+1（公式）
            df = f"=B{row}-A{row}+1"
            dc = ws.cell(row=row, column=4)
            dc.value = df; dc.font = data_font; dc.alignment = center

            # E: 利息 = 本金 * D * C / 360（公式）
            ef = f"={principal}*D{row}*C{row}/360"
            ec = ws.cell(row=row, column=5)
            ec.value = ef; ec.font = data_font; ec.alignment = right
            ec.number_format = '#,##0.00'

            for ci in range(1, 6):
                ws.cell(row=row, column=ci).border = thin
            row += 1

        last_data_row = row - 1

        # 合计行
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        tc = ws.cell(row=row, column=1, value="合计")
        tc.font = bold_font; tc.alignment = center; tc.border = thin
        for ci in range(1, 6):
            ws.cell(row=row, column=ci).border = thin

        sum_d = f"=SUM(D{first_data_row}:D{last_data_row})"
        ds = ws.cell(row=row, column=4)
        ds.value = sum_d; ds.font = bold_font; ds.alignment = center; ds.border = thin

        sum_e = f"=SUM(E{first_data_row}:E{last_data_row})"
        es = ws.cell(row=row, column=5)
        es.value = sum_e; es.font = bold_font; es.alignment = right
        es.number_format = '#,##0.00'; es.border = thin

        row += 3

    # 提交人
    row += 2
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3, value="提交人：").font = data_font
    ws.cell(row=row, column=3).alignment = right
    ws.cell(row=row, column=5, value=submitter).font = data_font
    row += 1

    # 日期
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    dc = ws.cell(row=row, column=3, value=f"{datetime.now().year}年  月  日")
    dc.font = data_font; dc.alignment = right


# ======================== CLI ========================

def main():
    parser = argparse.ArgumentParser(description="利息损失计算表 v3（匹配 5.0 结构）")
    parser.add_argument("output_path", help="输出 Excel 文件路径")
    parser.add_argument("--principal", type=float)
    parser.add_argument("--start")
    parser.add_argument("--end")
    parser.add_argument("--rate", type=float)
    parser.add_argument("--lpr-term", default=None, choices=["6M","1Y","1-3Y","3-5Y","5Y"])
    parser.add_argument("--items")
    parser.add_argument("--submitter", default="")
    parser.add_argument("--title", default="利息自动计算表")

    args = parser.parse_args()
    items = []
    if args.items:
        items = json.loads(args.items)
    elif args.principal and args.start and args.end:
        it = {"principal": args.principal, "start": args.start, "end": args.end}
        if args.rate:
            it["rate"] = args.rate
        if args.lpr_term:
            it["term"] = args.lpr_term
        items = [it]
    else:
        parser.print_help()
        sys.exit(1)

    generate_excel(args.output_path, items, args.submitter, args.title,
                   args.lpr_term or "1Y")

    for i, it in enumerate(items):
        p, s, e = it["principal"], it["start"], it["end"]
        r = it.get("rate")
        sd = datetime.strptime(s, "%Y-%m-%d").date()
        ed = datetime.strptime(e, "%Y-%m-%d").date()
        term = it.get("term", auto_select_term(sd, ed))
        per = split_periods(sd, ed, RATE_DATABASE, term, r)
        tag = f"固定利率 {r}%" if r else f"{term} LPR 分段"
        print(f"  {i+1}. 本金 {p:,.2f} 元，{s}~{e}，{tag}，共 {len(per)} 段")

    print(f"\n✅ 已生成：{args.output_path}")
    print(f"   含「利率数据库」+「利息计算表」两个 sheet")


if __name__ == "__main__":
    main()
