#!/usr/bin/env python3
"""
aje_extract.py — Extract adjusting journal entries from PDF/image/Excel files
using the Anthropic API (Claude).

Standalone script called by models.process_aje_file() for non-native formats.
Outputs JSON to stdout matching the parse_csw_aje() return format.

Usage:
    python aje_extract.py <file_path>

Requires:
    ANTHROPIC_API_KEY environment variable set.
    pip install anthropic

Output JSON format:
    {
      "entries": [
        {
          "num": "01",
          "date": "2024-12-31",
          "description": "Accrue wages",
          "lines": [
            {"csw_account": "Wages payable", "csw_number": "2100", "amount_cents": 150000},
            {"csw_account": "Wage expense", "csw_number": "5100", "amount_cents": -150000}
          ]
        }
      ],
      "csw_accounts": [
        {"name": "Wages payable", "number": "2100"},
        {"name": "Wage expense", "number": "5100"}
      ]
    }

Positive amount_cents = debit, negative = credit.
"""

import sys
import os
import json
import base64

SUPPORTED_EXTENSIONS = {
    '.pdf', '.png', '.jpg', '.jpeg', '.gif', '.webp',
    '.xlsx', '.xls', '.csv',
}

IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.webp'}
PDF_EXTENSIONS = {'.pdf'}

EXTRACTION_PROMPT = """You are an accounting data extraction tool. Extract ALL adjusting journal entries (AJEs) from this document.

For each AJE, extract:
- Entry number (e.g. "01", "02")
- Date (YYYY-MM-DD format)
- Description/memo
- Each debit and credit line with:
  - Account name (as written in the document)
  - Account number (if shown)
  - Amount in cents (positive = debit, negative = credit)

Return ONLY valid JSON in this exact format, no other text:
{
  "entries": [
    {
      "num": "01",
      "date": "2024-12-31",
      "description": "Accrue wages",
      "lines": [
        {"csw_account": "Wages payable", "csw_number": "2100", "amount_cents": 150000},
        {"csw_account": "Wage expense", "csw_number": "5100", "amount_cents": -150000}
      ]
    }
  ]
}

Rules:
- Positive amount_cents = DEBIT, negative = CREDIT
- Convert dollar amounts to cents (e.g. $1,500.00 = 150000)
- Each entry's lines must balance (sum to zero)
- Use account names exactly as shown in the document
- Include account numbers if visible
- If the date is not explicit for an entry, use the document date or fiscal year end
- If no entries are found, return {"entries": []}
"""


def extract_from_file(file_path):
    """Extract AJE data from a file using the Anthropic API."""
    import anthropic

    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY environment variable not set.")

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file type: {ext}")

    client = anthropic.Anthropic(api_key=api_key)

    # Build the message content based on file type
    content = []

    if ext in IMAGE_EXTENSIONS:
        media_map = {
            '.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
            '.gif': 'image/gif', '.webp': 'image/webp',
        }
        with open(file_path, 'rb') as f:
            data = base64.standard_b64encode(f.read()).decode('utf-8')
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": media_map[ext], "data": data},
        })

    elif ext in PDF_EXTENSIONS:
        with open(file_path, 'rb') as f:
            data = base64.standard_b64encode(f.read()).decode('utf-8')
        content.append({
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": data},
        })

    elif ext in ('.xlsx', '.xls'):
        # Convert spreadsheet to text representation
        text = _spreadsheet_to_text(file_path)
        content.append({"type": "text", "text": f"Spreadsheet contents:\n\n{text}"})

    elif ext == '.csv':
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            text = f.read()
        content.append({"type": "text", "text": f"CSV contents:\n\n{text}"})

    content.append({"type": "text", "text": EXTRACTION_PROMPT})

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        messages=[{"role": "user", "content": content}],
    )

    # Extract JSON from response
    response_text = response.content[0].text.strip()

    # Handle markdown code blocks
    if response_text.startswith('```'):
        lines = response_text.split('\n')
        # Remove first and last lines (```json and ```)
        lines = [l for l in lines if not l.strip().startswith('```')]
        response_text = '\n'.join(lines)

    parsed = json.loads(response_text)
    return parsed


def _spreadsheet_to_text(file_path):
    """Convert Excel file to text representation."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl required for Excel files: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    output = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else '' for c in row]
            if any(cells):
                output.append('\t'.join(cells))

    return '\n'.join(output)


def build_output(parsed):
    """Build the final output matching parse_csw_aje() format."""
    entries = parsed.get('entries', [])

    # Collect unique accounts
    seen = {}
    for entry in entries:
        for line in entry.get('lines', []):
            name = line.get('csw_account', '')
            number = line.get('csw_number', '')
            if name and name not in seen:
                seen[name] = {'name': name, 'number': number}

    return {
        'entries': entries,
        'csw_accounts': list(seen.values()),
    }


def main():
    if len(sys.argv) != 2:
        print("Usage: python aje_extract.py <file_path>", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    ext = os.path.splitext(file_path)[1].lower()

    # For native formats, delegate to models.parse_csw_aje
    if ext in ('.iif', '.txt', '.ven'):
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        import models
        result = models.parse_csw_aje(file_path)
        print(json.dumps(result))
        sys.exit(0)

    parsed = extract_from_file(file_path)
    output = build_output(parsed)
    print(json.dumps(output))


if __name__ == '__main__':
    main()
