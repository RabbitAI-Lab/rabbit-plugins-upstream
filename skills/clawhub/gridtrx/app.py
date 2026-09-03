"""
Grid — Browser-based double-entry accounting.
Simple. Analog. Robust.
"""
import os
import sys
import json
import re
import shutil
import time
import threading
import sqlite3
from datetime import datetime, date, timedelta
import io
from urllib.parse import quote
from flask import (Flask, render_template, request, redirect, url_for, 
                   flash, jsonify, send_file, session)
import models
import pdf_reports
from pdf_reports import (
    _setup_fonts, _fmt_money, _short_date,
    _get_bs_account_ids, _get_report_account_order,
    _build_account_detail, gl_pdf, aje_pdf, report_pdf, account_ledger_pdf,
)


app = Flask(__name__)
app.secret_key = 'grid-accounting-local-use-only'

# ─── Config: where client files live ────────────────────────────────

CONFIG_FILE = None  # Set at startup

def get_config_path():
    """Config file lives next to the program (GRID_CONFIG overrides it — for a
    second install, or a test run that must not touch the real one)."""
    return os.environ.get('GRID_CONFIG') or os.path.join(
        os.path.dirname(os.path.abspath(__file__)), 'grid.json')

def load_config():
    path = get_config_path()
    if os.path.exists(path):
        with open(path, 'r') as f:
            return json.load(f)
    return {'library_path': '', 'last_opened': ''}

def save_config(cfg):
    with open(get_config_path(), 'w') as f:
        json.dump(cfg, f, indent=2)

def get_library_path():
    cfg = load_config()
    return cfg.get('library_path', '')

def list_client_books():
    """Scan the library folder for client book files."""
    lib = get_library_path()
    if not lib or not os.path.isdir(lib):
        return []
    
    clients = []
    for entry in sorted(os.listdir(lib)):
        client_dir = os.path.join(lib, entry)
        if os.path.isdir(client_dir):
            db_path = os.path.join(client_dir, 'books.db')
            exists = os.path.exists(db_path)
            lock = models.books_lock_info(db_path) if exists else None
            clients.append({
                'name': entry,
                'path': db_path,
                'folder': client_dir,
                'exists': exists,
                'size': os.path.getsize(db_path) if exists else 0,
                'modified': datetime.fromtimestamp(os.path.getmtime(db_path)).strftime('%Y-%m-%d %H:%M') if exists else '',
                'locked_by': (f"{lock.get('prog','?')} on {lock.get('host','?')} since {lock.get('started','?')}"
                              if lock else ''),
            })
    return clients

# ─── Jinja2 Filters ────────────────────────────────────────────────

# Endpoints that render nothing and never read the books: the window heartbeat,
# the goodbye beacon, the quit call, static files. They are already exempt from
# the stale-tab file check (_FILE_CHECK_EXEMPT) so that a heartbeat can always
# get through — but load_company_info runs FIRST and queried the db regardless,
# so a database that could not be opened turned every heartbeat into a 500 and
# defeated the exemption exactly when it mattered most.
_NO_BOOKS_NEEDED = ('/api/alive', '/api/window-gone', '/api/quit', '/static/')


@app.before_request
def load_company_info():
    from flask import g
    if any((request.path or '').startswith(p) for p in _NO_BOOKS_NEEDED):
        g.grid_file = models.get_db_path() or ''
        return
    # v114 — the books file this page is being rendered from. base.html hands it
    # back on every write (X-Grid-File / _grid_file) so a stale second window
    # cannot post into whatever is open NOW. Empty string = no books open.
    g.grid_file = models.get_db_path() or ''
    if models.get_db_path():
      try:
        g.company_name = models.get_meta('company_name', '')
        fye = models.get_meta('fiscal_year_end', '')  # MM-DD
        fy_year = models.get_meta('fiscal_year', '')   # YYYY
        # Format as "31 Dec 2025"
        if fye and fy_year:
            months = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
                      7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
            try:
                parts = fye.split('-')
                mm, dd = int(parts[0]), int(parts[1])
                g.fiscal_display = f"{dd} {months.get(mm,'?')} {fy_year}"
            except:
                g.fiscal_display = f"{fye} {fy_year}"
        elif fye:
            g.fiscal_display = fye
        else:
            g.fiscal_display = ''
        # Core report IDs for nav links
        with models.get_db() as db:
            for rpt in db.execute("SELECT id, name FROM reports WHERE name IN ('BS','IS','AJE')"):
                g.__dict__[f'rpt_{rpt["name"].lower()}'] = rpt['id']
        # Unmapped-TB-accounts badge. Skip for /api/* (JSON; no template uses it) so
        # autocomplete/entry keystrokes don't trigger the scan on every request.
        if request.path.startswith('/api/'):
            g.unmapped_count = 0
            g.proof_cents = 0
        else:
            g.unmapped_count = _count_unmapped_accounts()
            # LAP PROOF line: the trial-balance tie as an always-visible zero.
            # Sum of ALL line amounts must be 0 — anything else is a red flag.
            # models.proof_cents() reuses the cached whole-book balances the
            # badge above just asked for, so the header costs one scan per
            # write instead of two on every single page.
            g.proof_cents = models.proof_cents()
      except sqlite3.OperationalError:
        # The header is decoration — company name, year-end, the PROOF chip. If
        # the books cannot be read RIGHT NOW (a bulk write holding them, a file
        # briefly unavailable) that must not be what fails the request. Render
        # blanks; whatever the page actually came to do will fail on its own,
        # with an error that says something useful about the real operation.
        g.company_name = ''
        g.fiscal_display = ''
        g.unmapped_count = 0
        g.proof_cents = 0

def _get_bs_is_mapped_ids():
    """Get account IDs that are mapped to BS or IS, either directly or via total-to chain."""
    with models.get_db() as db:
        # Direct: accounts on BS or IS
        direct = set()
        for r in db.execute("""SELECT DISTINCT ri.account_id FROM report_items ri
            JOIN reports rp ON ri.report_id = rp.id
            WHERE rp.name IN ('BS','IS') AND ri.account_id IS NOT NULL"""):
            direct.add(r['account_id'])

        # Build lookups
        acct_id_by_name = {}
        for a in db.execute("SELECT id, name FROM accounts"):
            acct_id_by_name[a['name']] = a['id']

        # Build total_to map from ALL reports: account_id -> set of total_to_1 targets
        acct_total_to = {}
        for ri in db.execute("""SELECT ri.account_id, ri.total_to_1 FROM report_items ri
            WHERE ri.account_id IS NOT NULL
            AND ri.total_to_1 IS NOT NULL AND ri.total_to_1 != ''"""):
            acct_total_to.setdefault(ri['account_id'], set()).add(ri['total_to_1'])

        # All accounts on any report (not just BS/IS)
        all_on_reports = set()
        for r in db.execute("""SELECT DISTINCT account_id FROM report_items
            WHERE account_id IS NOT NULL"""):
            all_on_reports.add(r['account_id'])

        # For each account not directly on BS/IS, walk total_to chain across all reports
        indirect = set()
        for acct_id in all_on_reports - direct:
            targets = acct_total_to.get(acct_id)
            if not targets:
                continue
            visited = set()
            frontier = set(targets)
            found = False
            for _ in range(10):
                if not frontier:
                    break
                next_frontier = set()
                for target_name in frontier:
                    target_id = acct_id_by_name.get(target_name)
                    if not target_id:
                        continue
                    if target_id in direct:
                        found = True
                        break
                    if target_id not in visited:
                        visited.add(target_id)
                        next_frontier |= acct_total_to.get(target_id, set())
                if found:
                    break
                frontier = next_frontier
            if found:
                indirect.add(acct_id)

        return direct | indirect


def _count_unmapped_accounts():
    """Count posting accounts with non-zero balances not mapped to BS/IS (directly or via chain)."""
    try:
        mapped_ids = _get_bs_is_mapped_ids()
        balances = models.get_all_account_balances()   # one bulk query, not N per-account
        with models.get_db() as db:
            posting_ids = [r['id'] for r in
                           db.execute("SELECT id FROM accounts WHERE account_type='posting'")]
        return sum(1 for aid in posting_ids
                   if aid not in mapped_ids and balances.get(aid, 0) != 0)
    except:
        return 0

@app.template_filter('money')
def money_filter(cents):
    return models.fmt_amount(cents)

@app.template_filter('money_plain')
def money_plain_filter(cents):
    return models.fmt_amount_plain(cents)

@app.template_filter('money_dr')
def money_dr_filter(cents):
    return models.fmt_amount(cents) if cents > 0 else ''

@app.template_filter('money_cr')
def money_cr_filter(cents):
    return models.fmt_amount(abs(cents)) if cents < 0 else ''

@app.template_filter('signoff')
def signoff_filter(stamp):
    """'CW 2026-07-30' → 'CW 30-Jul-26'. Stored ISO (sortable, unambiguous),
    shown short — who signed AND when, at a glance."""
    s = str(stamp or '').strip()
    if not s:
        return ''
    who, _, when = s.partition(' ')
    try:
        when = datetime.strptime(when.strip(), '%Y-%m-%d').strftime('%d-%b-%y')
    except ValueError:
        pass
    return f'{who} {when}'.strip()

# Make fmt_amount available as a global in templates
app.jinja_env.globals['fmt'] = models.fmt_amount

def pct_fmt(basis_points):
    """Format basis points as percentage string. 10000 bp = 100.0%"""
    if basis_points is None or basis_points == 0:
        return '—'
    pct = basis_points / 100.0
    return f'{pct:,.1f}%'

app.jinja_env.globals['pct_fmt'] = pct_fmt

# ─── File Picker / Library ──────────────────────────────────────────

@app.route('/library')
def library():
    """The starting screen — pick a client or create a new one."""
    cfg = load_config()
    lib_path = cfg.get('library_path', '')
    clients = list_client_books() if lib_path else []
    op_name, op_init = current_operator()
    return render_template('library.html', clients=clients, library_path=lib_path,
                           operator_name=op_name, operator_initials=op_init)

@app.route('/library/set-path', methods=['POST'])
def set_library_path():
    """Set the folder where client files live."""
    path = request.form.get('library_path', '').strip()
    if path and os.path.isdir(path):
        cfg = load_config()
        cfg['library_path'] = path
        save_config(cfg)
        flash(f'Library path set to: {path}', 'success')
    else:
        flash('That folder does not exist. Please check the path.', 'error')
    return redirect(url_for('library'))

@app.route('/library/new-client', methods=['POST'])
def new_client():
    """Create a new client folder with empty books. The operator keys the
    LEGAL name and the fiscal year; the folder name is derived (v142) —
    LegalNameCompact-YYYYgx — never typed."""
    # v114 — creating a client also switches the open books, so it honours the
    # same one-books rule as /open (and refuses BEFORE making a folder).
    if models.get_db_path():
        flash('Another client\'s books are already open — use ✕ Close to Library '
              'first, then create the new client.', 'error')
        return redirect(url_for('library'))

    name = request.form.get('client_name', '').strip()
    year = request.form.get('fiscal_year', '').strip()
    try:
        folder = models.client_folder_name(name, year)
    except ValueError as e:
        flash(str(e), 'error')
        return redirect(url_for('library'))
    
    lib = get_library_path()
    if not lib:
        flash('Set your library path first.', 'error')
        return redirect(url_for('library'))
    
    client_dir = os.path.join(lib, folder)
    db_path = os.path.join(client_dir, 'books.db')
    
    if os.path.exists(client_dir):
        flash(f'{folder} already exists in the library.', 'error')
        return redirect(url_for('library'))
    
    os.makedirs(client_dir)
    models.create_starter_books(db_path, name, working_year=int(year))
    
    # Save as last opened
    cfg = load_config()
    cfg['last_opened'] = db_path
    save_config(cfg)
    
    flash(f'Created new books for {name} — filed as {folder}', 'success')
    return redirect(url_for('open_client', path=db_path))

# ─── The locked-books screen (v130) ────────────────────────────────
# A refusal with no door is what sends staff to delete .lock files by hand.
# Every way of NOT getting into a set of books now lands on ONE screen that
# says who has them and offers ONE key — F3 — that does the truthful thing:
#   • held by a live session on this machine → close that session properly
#   • held by another machine (incl. a lock a file-sync copied in) → clear it
#   • a different client is open in THIS window        → close it, open this
# LAP programs behaved this way; the hard lock-out was ours, and it was wrong.

def _lock_screen(path, e=None, switch_from=None):
    """Render the one screen. `e` is a models.BooksLocked; `switch_from` is the
    other client's path when the block is our own open window."""
    if switch_from:
        view = {
            'state': 'switch',
            'headline': f"{_client_label(switch_from)} is open in this window.",
            'detail': "Grid keeps one set of books open at a time, so this one has to "
                      "be put away before the other opens.",
            'door': f"Close {_client_label(switch_from)} and open {_client_label(path)}.",
            'warn': '', 'holder': '',
        }
    else:
        view = {
            'state': e.state,
            'headline': e.headline,
            'detail': e.detail,
            'door': e.door,
            'warn': e.warn,
            'holder': f"{e.prog} · PID {e.pid} · {e.host} · since {e.started}",
        }
    return render_template('locked.html', path=path, client=_client_label(path),
                           lock_file=(e.lock_path if e else ''), **view)

def _client_label(path):
    """The client's name as the library shows it — the folder, not 'books.db'."""
    try:
        return os.path.basename(os.path.dirname(os.path.abspath(path))) or os.path.basename(path)
    except Exception:
        return os.path.basename(path or 'these books')

def _handoff(e):
    """Ask the session that holds the books to close them properly, and wait.

    A live Grid instance records its port in the lock file, so we can knock on
    it rather than kill it: /api/quit checkpoints the WAL, releases the lock and
    stops. That is a clean close, not a stolen file. Returns True if the lock
    went away. False just means we fall back to clearing it."""
    if not e.port:
        return False
    import urllib.request
    try:
        urllib.request.urlopen(
            urllib.request.Request(f'http://127.0.0.1:{e.port}/api/quit', data=b'',
                                   method='POST'), timeout=3).read()
    except Exception:
        return False   # nothing listening, or not Grid — the PID was recycled
    for _ in range(40):                       # up to 8s for it to let go
        if not os.path.exists(e.lock_path):
            return True
        time.sleep(0.2)
    return not os.path.exists(e.lock_path)

# ─── Procedures (v133) — LAP-style: one menu, whole jobs ──────
# The firm's own LAP procedure library was mostly this — a job
# you do enough times to get sick of assembling by hand, written down once.
#
# It is NOT a scripting layer and must not become one. Judgement does not run:
# deciding what a suspense row is stays a person's work. What lives here is the
# assembly — the same reports, the same order, every time.
#
# The ONE thing each procedure must ask is the date. the legacy print job printed the
# CURRENT year and nothing else, so reprinting an earlier year meant editing the
# procedure. Every procedure here is asked for its as-of date and offered the
# working year-end as the default, never forced to it.

_BALANCE_REPORTS = {'BS'}


def _ye_package(as_of):
    company = models.get_meta('company_name', 'My Books')
    pdf, included = pdf_reports.ye_package_pdf(company, as_of)
    return pdf, f'YE-Package-{as_of}.pdf', ' · '.join(included)


def _thirteen_column(as_of, report_name='IS'):
    """Twelve months + a total, ending at the date asked for.

    The MODE follows the report, because the accounting does: an income
    statement reports ACTIVITY, so each column is that month and the twelve add
    up to a total; a balance sheet is PERPETUAL, so each column is the balance
    as at that month-end and there is nothing to total."""
    rep = models.find_report_by_name(report_name)
    if not rep:
        raise ValueError(f"No '{report_name}' report in these books.")
    income_style = (report_name or '').upper() not in _BALANCE_REPORTS
    company = models.get_meta('company_name', 'My Books')
    pdf = pdf_reports.monthly_columns_pdf(
        company, rep['id'], rep['description'] or rep['name'], as_of,
        months=12, mode='periodic' if income_style else 'cumulative')
    return pdf, f'{report_name}-13col-{as_of}.pdf', (
        'monthly activity + total' if income_style else 'month-end balances')


PROCEDURES = [
    {'key': 'ye_package',
     'title': 'Print PDF of YE Pkg',
     'note': 'Balance sheet, income statement and the adjusting entries — '
             'one PDF. Each statement shows the year, the year before, '
             'the dollar change and the percentage change.',
     'prompt': 'Year-end to print',
     'run': _ye_package},
    {'key': 'thirteen_column',
     'title': 'Print 13 column report',
     'note': 'Twelve monthly columns ending at the date you give, plus a total. '
             'Counts back from that date, so it does not have to be a fiscal '
             'year. Pick the report it runs on.',
     'prompt': 'Last month to include (any date in it)',
     'report_choice': True,
     'run': _thirteen_column},
]


def _procedure(key):
    for p in PROCEDURES:
        if p['key'] == key:
            return p
    return None


@app.route('/procedures')
def procedures_menu():
    """The list behind \\ -> Run Procedure."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    return jsonify([{k: p[k] for k in ('key', 'title', 'note')} for p in PROCEDURES])


@app.route('/procedure/<key>', methods=['GET', 'POST'])
def run_procedure(key):
    proc = _procedure(key)
    if not proc or not models.get_db_path():
        return redirect(url_for('library'))
    anc = models.fiscal_anchor()
    default = (anc or {}).get('cy_end', '') or models.fiscal_ceiling()
    reports = [r for r in models.get_reports() if r['name'] in ('IS', 'BS')]

    if request.method == 'GET':
        return render_template('procedure.html', proc=proc, default=default,
                               reports=reports,
                               company=models.get_meta('company_name', 'My Books'))

    as_of = models.normalize_date(request.form.get('as_of', '').strip() or default)
    if not as_of:
        flash('That date could not be read. Use YYYY-MM-DD.', 'error')
        return redirect(url_for('run_procedure', key=key))
    try:
        if proc.get('report_choice'):
            pdf, filename, note = proc['run'](as_of, request.form.get('report_name', 'IS'))
        else:
            pdf, filename, note = proc['run'](as_of)
    except Exception as e:
        flash(f"{proc['title']} could not be produced: {e}", 'error')
        return redirect(url_for('run_procedure', key=key))
    return send_file(io.BytesIO(pdf), mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


@app.route('/tools')
def tools_page():
    """File-level utilities. Settings is what the CLIENT is; this is what you can
    DO to a file. Keeping them apart stops the client's own properties being
    buried among maintenance jobs."""
    return render_template('tools.html')



@app.route('/check')
def check_books_page():
    """Check Books. The PROOF chip in the topbar says the books tie; this says
    everything else — and says plainly which items are WRONG and which are just
    waiting for someone."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    return render_template('check.html', result=models.check_books(),
                           company=models.get_meta('company_name', 'My Books'))


@app.route('/locked')
def locked_screen():
    """The lock screen on its own URL, so a boot into a locked file lands here
    instead of on a library the operator then has to navigate back through."""
    path = request.args.get('path', '')
    if not path or not os.path.exists(path):
        return redirect(url_for('library'))
    blocked = models.books_lock_state(path)  # read-only: asking takes nothing
    if blocked is not None:
        return _lock_screen(path, blocked)
    return redirect(url_for('open_client', path=path))

# ─── Report columns: built ONCE, read by the screen and by print ───
# These two used to be separate copies of the same forty lines, which is how a
# column type ends up existing on screen and missing from the printed page.

COLUMN_TYPES = ('A', 'D', 'C', 'change', 'pct_change', 'pct_acct', 'spacer')
DERIVED_TYPES = ('change', 'pct_change', 'spacer')   # need no period of their own

# What each column type OWNS. A saved config keeps only these keys, so changing
# a column's type cannot leave the previous type's settings behind — switch
# "% of acct" back to "% chg" and the account code is gone, not lurking in the
# saved JSON waiting to be read by something. (The same trap already existed
# for the chg A/B numbers; this closes both.)
COLUMN_FIELDS = {
    'A':          ('begin', 'end', 'label'),
    'D':          ('begin', 'end', 'label'),
    'C':          ('begin', 'end', 'label'),
    'change':     ('a', 'b', 'label'),
    'pct_change': ('a', 'b', 'label'),
    'pct_acct':   ('begin', 'end', 'label', 'acct'),
    'spacer':     ('label',),
}


# A column with no explicit type is an ordinary balance column. ONE definition,
# used by both the cleaner and the builder — they used to disagree, and that is
# what silently emptied saved statements (see _clean_column_cfg).
COLUMN_TYPE_DEFAULT = 'A'


def _clean_column_cfg(cfg):
    """Keep only what each column's own type uses. Idempotent."""
    out = {}
    for i in range(1, 7):
        ctype = cfg.get(f'c{i}_type', '')
        if not ctype:
            # A plain balance column carries dates and a label but NO type —
            # _column_slots has always defaulted it to 'A'. Treating "no type"
            # as "no column" here deleted every base column, on save AND on
            # load, which left the derived $ chg / % chg columns pointing at
            # slots that no longer existed and a statement that rendered
            # completely empty. An operator saw a blank balance sheet on a
            # perfectly good set of books.
            if not any(cfg.get(f'c{i}_{k}', '')
                       for k in ('begin', 'end', 'label', 'acct')):
                continue                       # genuinely an unused slot
            ctype = COLUMN_TYPE_DEFAULT
        out[f'c{i}_type'] = ctype              # written out, so it heals itself
        for k in COLUMN_FIELDS.get(ctype, ()):
            v = cfg.get(f'c{i}_{k}', '')
            if v:
                out[f'c{i}_{k}'] = v
    return out


def _column_slots(cfg, report_id, items, all_items):
    """Turn a saved columns_<id> config into up to six built columns.

    'A' is the ordinary net balance for a period. 'D' and 'C' are FLOWS — one
    side of each account rather than the net (LAP doctrine: the debit side of AR is
    sales by customer, the credit side of AP is purchases by supplier). Both
    ripple up the total-to chain exactly as a net balance does."""
    def carg(k, d=''):
        return cfg.get(k, d)

    slots = [None] * 6
    for i in range(1, 7):
        end = carg(f'c{i}_end', '')
        begin = carg(f'c{i}_begin', '')
        ctype = carg(f'c{i}_type', COLUMN_TYPE_DEFAULT)
        label = carg(f'c{i}_label', '')
        if not end and ctype not in DERIVED_TYPES:
            continue
        if ctype in ('change', 'pct_change'):
            slots[i-1] = {'type': ctype, 'label': label or ('$ chg' if ctype == 'change' else '% chg'),
                          'a': int(carg(f'c{i}_a', '1')) - 1,
                          'b': int(carg(f'c{i}_b', '2')) - 1, 'data': None}
        elif ctype == 'spacer':
            slots[i-1] = {'type': 'spacer', 'label': label or '', 'data': None}
        elif ctype == 'pct_acct':
            # Common-size: every line as a percentage of ONE account for the same
            # period — expenses against revenue, the column that makes a big year
            # and a small year comparable. LAP's [% of Account].
            acct = (carg(f'c{i}_acct', '') or '').strip().upper()
            data = models.compute_report_column(report_id, date_from=begin or None,
                                                date_to=end or None,
                                                _display_items=items, _all_items=all_items)
            # A code that names nothing must SAY so in the heading. Silently
            # rendering a column of em-dashes leaves the operator staring at a
            # blank column with no idea they mistyped the account.
            known = bool(acct) and models.get_account_by_name(acct) is not None
            base = models.trace_account(acct, date_from=begin or None,
                                        date_to=end or None)['display'] if known else 0
            if not acct:
                head = '% of — name an account'
            elif not known:
                head = f'% of {acct} — no such account'
            else:
                head = label or f'% of {acct}'
            slots[i-1] = {'type': 'pct_acct', 'acct': acct, 'begin': begin, 'end': end,
                          'label': head, 'base': base, 'known': known,
                          'data': [(item, (round(v * 10000 / abs(base)) if base else 0))
                                   for item, v in data]}
        else:
            side = ctype if ctype in ('D', 'C') else None
            data = models.compute_report_column(report_id, date_from=begin or None,
                                                date_to=end or None, side=side,
                                                _display_items=items, _all_items=all_items)
            default_label = {'D': 'Debits', 'C': 'Credits'}.get(ctype) or (end[:4] if end else 'Current')
            slots[i-1] = {'type': 'actual', 'side': side, 'begin': begin, 'end': end,
                          'label': label or default_label, 'data': data}
    return slots        # padded to 6 — the column editor needs the empty slots


def _has_custom_columns(cfg):
    return any(cfg.get(f'c{i}_end') or cfg.get(f'c{i}_type') in DERIVED_TYPES
               for i in range(1, 7))


@app.route('/open')
def open_client():
    """Open a specific client's books."""
    path = request.args.get('path', '')
    if not path or not os.path.exists(path):
        flash('File not found.', 'error')
        return redirect(url_for('library'))
    # Set only by the F3 door on the lock screen — never by an ordinary open.
    force = request.args.get('force') == '1'

    # v114 — ONE CLIENT, ONE SET OF BOOKS. Refuse to switch out from under an
    # open window; switching is an explicit Close → Open, never an ambient side
    # effect (that ambient switch is the state the stale-tab guard above exists
    # to catch after the fact). No liveness test is needed: Grid is one process
    # serving one window, so a request arriving IS a live window — and the books
    # lock in models.acquire_books_lock is the cross-PROCESS belt.
    # Re-opening the SAME books is allowed — that is just a reload.
    # v130 — the refusal became a door: F3 does the Close → Open for you.
    cur = models.get_db_path()
    if cur and os.path.abspath(path) != os.path.abspath(cur):
        if not force:
            return _lock_screen(path, switch_from=cur)
        models.checkpoint_books()            # put the other books away properly
        models.set_db_path(None)

    try:
        # Single gated open (models.init_db): integrity check → daily snapshot →
        # migrate. A corrupt file is refused here, before anything writes to it.
        models.init_db(path)
    except models.BooksLocked as e:
        if not force:
            return _lock_screen(path, e)
        # F3, on a screen that named the holder. A live session is asked to
        # close itself first; only if it will not answer do we clear the lock.
        if not (e.state == models.LOCK_RUNNING and _handoff(e)):
            models.clear_books_lock(path)
        print(f"  !! lock cleared by the operator — was {e.prog} PID {e.pid} "
              f"on {e.host} since {e.started}")
        try:
            models.init_db(path, force_lock=True)
        except ValueError as e2:
            flash(str(e2), 'error')
            return redirect(url_for('library'))
    except ValueError as e:
        flash(str(e), 'error')
        return redirect(url_for('library'))
    for line in models.re_repair_note():
        flash(line, 'error')       # loud on purpose: this can move the balance sheet
    st = models.backup_status()
    if st['error']:
        flash(st['note'] + ' — the books opened, but you are NOT protected today. Fix this (disk full?).', 'error')
    cfg = load_config()
    cfg['last_opened'] = path
    save_config(cfg)

    return redirect(url_for('home'))

@app.route('/close')
def close_books():
    """Close current books and return to the library/startup screen."""
    models.checkpoint_books()   # fold WAL into the .db so the synced file is complete at rest
    models.set_db_path(None)
    cfg = load_config()
    cfg['last_opened'] = ''
    save_config(cfg)
    return redirect(url_for('library'))

# ─── Lifecycle: one formal close, the DOS way ───────────────────────
# LAP made you close the data file before leaving. A browser cannot
# refuse its own X — no web page can — so we invert it: the X performs the
# formal close instead of being blocked by it, and the window and the server
# live and die together. Nothing can be left half-open.

WINDOW_GRACE  = 6.0     # a window said goodbye — wait this long for another
BOOT_DEADLINE = 90.0    # a window never appeared at all — don't run forever

# v116 — a BACKGROUNDED window is not an absent one. At 5 minutes this closed
# Grid whenever the operator worked in another application for a while: EVERY
# modern browser (Firefox, Chromium-family, Safari alike) throttles timers in
# hidden tabs and windows, so the heartbeat stops arriving even though the window
# is sitting right there. Grid then did exactly what it was told and shut itself
# down — which read as "Grid-XB closed Grid". It did not: they listen on 5000 and
# 5001 and never meet. The authoritative "the window is gone" signal is the
# pagehide beacon, which fires on a real close. This is only the backstop for a
# browser that CRASHED and never got to say goodbye, so it can afford to be long
# — and relaunching just hands back a window on this same instance.
IDLE_QUIT     = 1800.0  # 30 min of total silence = the browser died

_server      = None     # the werkzeug server, set in main()
_client_seen = False    # a Grid window has checked in at least once
_last_seen   = 0.0
_boot_time   = 0.0      # set in main(); the BOOT_DEADLINE clock
_inflight    = 0
_grace_until = 0.0
_quitting    = False

# Routes that deliberately switch books, or don't target the open file at all.
# The lifecycle endpoints are exempt too: a stale window's heartbeat must still
# keep the server alive, or the watchdog closes Grid under the operator's feet.
_FILE_CHECK_EXEMPT = ('/open', '/close', '/library', '/api/alive',
                      '/api/window-gone', '/api/quit', '/static/')


@app.before_request
def _mark_alive():
    """Every request is proof a window is alive — and cancels any pending close
    (this is what makes ordinary page-to-page navigation safe). Then the v114
    stale-tab guard: a window rendered against a DIFFERENT set of books may not
    write into the ones open now."""
    global _client_seen, _last_seen, _inflight, _grace_until
    _client_seen = True
    _last_seen = time.time()
    _grace_until = 0.0
    _inflight += 1

    path = request.path or ''
    if any(path.startswith(p) for p in _FILE_CHECK_EXEMPT):
        return
    tab_file = request.headers.get('X-Grid-File')
    if tab_file is None and request.method == 'POST' and not request.is_json:
        tab_file = request.form.get('_grid_file')
    if not tab_file:          # no claim made (plain link, old page) → nothing to check
        return
    cur = models.get_db_path() or ''
    if cur and os.path.abspath(tab_file) == os.path.abspath(cur):
        return
    msg = ('This window was showing a different set of books — another window has '
           'since opened a different client. Reload this window before making '
           'changes (nothing was saved).')
    wants_json = (path.startswith('/api/') or request.is_json or
                  'application/json' in (request.headers.get('Accept') or ''))
    if wants_json:
        return jsonify({'ok': False, 'error': msg}), 409
    flash(msg, 'error')
    return redirect(url_for('library'))

@app.teardown_request
def _mark_done(exc=None):
    global _inflight, _last_seen
    _inflight = max(0, _inflight - 1)
    _last_seen = time.time()

@app.route('/api/alive')
def api_alive():
    """The window's heartbeat. The work is done in _mark_alive."""
    return jsonify({'ok': True})

# A page that had been hidden longer than this when it disappeared was almost
# certainly discarded by the browser, not closed by the operator.
DISCARD_HIDDEN_MS = 5000

@app.route('/api/window-gone', methods=['POST'])
def api_window_gone():
    """A Grid window went away. Start a short grace — if another Grid page
    checks in (an ordinary link click), the close is cancelled; if none does,
    the watchdog performs the formal close and stops Grid.

    v116 — the beacon carries how many milliseconds the page had been hidden.
    Browsers fire pagehide when they DISCARD a background tab to reclaim memory,
    and treating that as a goodbye shut Grid down while the operator was working
    in another application. A window that was in use when it went away reports
    ~0; a reclaimed background tab reports minutes. Only the first is a goodbye —
    the other is left to the idle backstop, which is half an hour of complete
    silence and cannot be reached by a window that is merely in the background."""
    global _grace_until
    try:
        hidden_ms = float((request.get_data(as_text=True) or '0').strip() or 0)
    except ValueError:
        hidden_ms = 0
    if hidden_ms > DISCARD_HIDDEN_MS:
        return jsonify({'ok': True, 'ignored': 'page was hidden — browser discard, not a close'})
    _grace_until = time.time() + WINDOW_GRACE
    return jsonify({'ok': True})

@app.route('/api/quit', methods=['POST'])
def api_quit():
    """The red Close link — the formal close, all the way out."""
    _begin_shutdown('closed by the operator')
    return jsonify({'ok': True})

def _formal_close(reason):
    """Everything a clean exit owes the books: fold the WAL into the .db (so
    the Box-synced file is complete at rest), release the single-instance lock,
    forget the file so the next launch starts at the library."""
    try:
        if models.DB_PATH:
            models.checkpoint_books()
            models.set_db_path(None)      # releases the v106 books lock
        cfg = load_config()
        cfg['last_opened'] = ''
        save_config(cfg)
    except Exception as e:
        print(f"  !! while closing the books: {e}")

def _begin_shutdown(reason):
    """Formal close, then stop the server. Idempotent."""
    global _quitting
    if _quitting:
        return
    _quitting = True
    _formal_close(reason)
    print(f"\n  Books closed ({reason}). Grid has stopped.")
    if _server is not None:
        # Let the in-flight response flush first; shutdown() must not run on
        # the serving thread.
        threading.Timer(0.5, _server.shutdown).start()
    else:
        threading.Timer(0.5, lambda: os._exit(0)).start()

def _watchdog():
    """Closing the window closes Grid. Runs every 2s; never fires while a
    request is in flight (a slow report render is not an absent window)."""
    global _last_seen
    last_tick = time.time()
    while not _quitting:
        time.sleep(2)
        now = time.time()
        # The machine slept: the watchdog was suspended too, so the silence
        # proves nothing. Reset rather than close books behind the operator.
        if now - last_tick > 60:
            _last_seen = now
            last_tick = now
            continue
        last_tick = now
        if _inflight:
            continue
        if not _client_seen:
            # v114 — no window EVER checked in: the browser failed to launch, or
            # was shut before it painted. Without this, Grid ran on forever
            # behind a console nobody could interpret, still holding the books
            # lock so no other window (or agent) could get in. Skipped when the
            # operator is driving it by hand (GRID_NO_WINDOW).
            if (not os.environ.get('GRID_NO_WINDOW')
                    and now - _boot_time > BOOT_DEADLINE):
                _begin_shutdown('no window ever appeared')
                return
            continue
        if _grace_until and now > _grace_until:
            _begin_shutdown('window closed')
            return
        if now - _last_seen > IDLE_QUIT:
            _begin_shutdown(f'no window for {int(IDLE_QUIT // 60)} min')
            return

# ─── Home ───────────────────────────────────────────────────────────

@app.route('/')
def home():
    if not models.get_db_path():
        return redirect(url_for('library'))
    reports = models.get_reports()
    company = models.get_meta('company_name', 'My Books')
    db_path = models.get_db_path()
    client_folder = os.path.basename(os.path.dirname(db_path))
    return render_template('home.html', reports=reports, company=company,
                         client_folder=client_folder, backup=models.backup_status(),
                         openings=models.openings_state())

# ─── Opening balances (conversion) ─────────────────────────────────
# The one door. Everything about a conversion — creating TRX.OPEN, the per-account
# entries, working out opening retained earnings, proving it ties — happens here
# or not at all. See models.post_opening_balances for why.

OPENING_GRID_ROWS = 25          # what the operator sees on a fresh grid


def _opening_confirmed():
    """Did the browser already put the "these are opening balances" question to
    the operator? A conversion entry is editable — but never by a stray click,
    an F5 mid-entry, or a sweep selection that happened to include it."""
    if request.is_json:
        return bool((request.get_json(silent=True) or {}).get('confirm_opening'))
    return request.form.get('confirm_opening', '') == '1'


def _opening_grid_from_form():
    """Read the grid back off the form, keeping exactly what was typed so a
    bounce-back never loses the operator's work."""
    descs = request.form.getlist('row_desc')
    accts = request.form.getlist('row_acct')
    amts  = request.form.getlist('row_amt')
    rows, raw = [], []
    for i in range(max(len(descs), len(accts), len(amts))):
        d = (descs[i] if i < len(descs) else '').strip()
        a = (accts[i] if i < len(accts) else '').strip()
        t = (amts[i] if i < len(amts) else '').strip()
        cents, bad = 0, ''
        if t:
            try:
                cents = models.parse_amount(t)
            except ValueError:
                bad = f"'{t}' is not an amount."
        rows.append({'account': a, 'description': d, 'amount': cents})
        # amount_error is kept apart from error so the model's per-row findings
        # can be merged in without either one hiding the other.
        raw.append({'account': a, 'description': d, 'amount_text': t,
                    'amount_error': bad, 'error': '', 'warning': ''})
    return rows, raw


def _opening_context(rows_raw, conversion_date, expected_text='', result=None):
    """Everything the screen renders, in one place — used by the first paint and
    by every bounce-back."""
    state = models.openings_state()
    conv = models.get_account_by_name(models.CONVERSION_ACCT)
    posted_re = ''
    if state['status'] == 'posted':
        ob = models.get_account_by_name(models.OPENING_RE_ACCT)
        c = -models.get_account_balance(ob['id']) if ob else 0
        posted_re = models.fmt_amount_plain(c) + (' CR' if c >= 0 else ' DR')
    return dict(state=state, grid=rows_raw, accounts=models.opening_account_choices(),
                conversion_date=conversion_date, default_date=state['default_date'],
                expected_re=expected_text, result=result,
                posted_re_text=posted_re, editing=False,
                conv_account_id=(conv['id'] if conv else 0),
                company=models.get_meta('company_name', 'My Books'))


@app.route('/opening-balances')
def opening_balances():
    if not models.get_db_path():
        return redirect(url_for('library'))
    state = models.openings_state()
    grid, date = [], state['default_date']
    # v116 — Edit loads the posted conversion back into the grid rather than
    # making anyone re-key it. Retained earnings and the TRX.OPEN contra are not
    # in there: Grid owns both, and RE is recomputed from whatever the edited
    # grid ends up saying.
    editing = request.args.get('edit') == '1' and state['status'] == 'posted'
    if editing:
        date = state['conversion_date'] or date
        grid = [{'account': r['account'], 'description': r['description'],
                 'amount_text': models.fmt_amount_plain(r['amount']),
                 'error': '', 'warning': ''} for r in models.opening_rows_from_batch()]
    while len(grid) < OPENING_GRID_ROWS:
        grid.append({'account': '', 'description': '', 'amount_text': '',
                     'error': '', 'warning': ''})
    ctx = _opening_context(grid, date)
    ctx['editing'] = editing
    return render_template('opening_balances.html', **ctx)


@app.route('/opening-balances', methods=['POST'])
def opening_balances_post():
    """Check, or post. Both re-validate in the model layer — the browser's running
    totals are a courtesy, never the authority."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    conversion_date = (request.form.get('conversion_date') or '').strip()
    expected_text = (request.form.get('expected_re') or '').strip()
    expected = None
    rows, raw = _opening_grid_from_form()

    ctx_err = None
    if expected_text:
        try:
            expected = models.parse_amount(expected_text)
        except ValueError:
            ctx_err = f"'{expected_text}' is not an amount — leave it blank if you don't have the figure."

    # Validate EVERYTHING every time. An amount that could not be read must not
    # hide a blocked account three lines further down — the operator should see
    # the whole list once, not discover it one bounce at a time.
    v = models.validate_opening_rows(conversion_date, rows, expected)
    v = dict(v)
    v['errors'] = list(v['errors'])
    # Hand the per-row findings back to the exact line they came from.
    for r in v['rows']:
        raw[r['index']]['error'] = r['error']
        raw[r['index']]['warning'] = r['warning']
    bad_amounts = 0
    for rr in raw:
        if rr['amount_error']:
            rr['error'] = rr['amount_error']       # an unreadable amount wins its own line
            bad_amounts += 1
    if bad_amounts:
        v['errors'].insert(0, f"{bad_amounts} amount(s) could not be read — a number, "
                              f"(1,000.00) for a credit, or leave the line blank.")
    if ctx_err:
        v['errors'].insert(0, ctx_err)
    v['ok'] = v['ok'] and not bad_amounts and not ctx_err

    editing = request.form.get('replace') == '1'
    if request.form.get('action') == 'post' and v['ok']:
        try:
            res = models.post_opening_balances(conversion_date, rows, expected,
                                               replace=editing)
        except ValueError as e:
            ctx = _opening_context(raw, conversion_date, expected_text, {'errors': [str(e)]})
            ctx['editing'] = editing
            return render_template('opening_balances.html', **ctx)
        re_txt = models.fmt_amount_plain(res['re_credit_cents'])
        verb = 'replaced' if res.get('replaced') else 'posted'
        flash(f"Opening balances {verb} — {len(res['txn_ids'])} entries dated {conversion_date}, "
              f"with opening retained earnings of {re_txt} CR. "
              f"TRX.OPEN is back to zero.", 'success')
        conv = models.get_account_by_name(models.CONVERSION_ACCT)
        return redirect(url_for('account_ledger', account_id=conv['id']))

    if v['ok']:
        flash('Everything ties — nothing posted yet. Check the retained-earnings figure, '
              'then Post opening balances.', 'success')
    ctx = _opening_context(raw, conversion_date, expected_text, v)
    ctx['editing'] = editing
    return render_template('opening_balances.html', **ctx)


@app.route('/opening-balances/decline', methods=['POST'])
def opening_balances_decline():
    """Brand-new client — no history to bring over. Remembered, so the prompt
    stops asking and day-one entry can just begin."""
    models.decline_openings()
    flash('Starting from zero — no opening balances. You can still enter them later '
          'from the TRX report if the client turns out to have history.', 'success')
    return redirect(url_for('home'))


@app.route('/opening-balances/delete', methods=['POST'])
def opening_balances_delete():
    try:
        n = models.delete_opening_balances()
        flash(f'Conversion deleted — {n} opening entries removed. Enter them again when ready.',
              'success')
    except ValueError as e:
        flash(str(e), 'error')
    return redirect(url_for('opening_balances'))


# ─── Report View (with total-to chains) ───────────────────────────

@app.route('/report/<int:report_id>')
def report_view(report_id):
    import time as _time
    _t0 = _time.time()
    
    if not models.get_db_path():
        return redirect(url_for('library'))
    report = models.get_report(report_id)
    if not report:
        flash('Report not found', 'error')
        return redirect(url_for('home'))
    
    company = models.get_meta('company_name', 'My Books')
    hide_zero = request.args.get('hide_zero', '0') == '1'
    show_setup = request.args.get('show_setup', '0') == '1'
    
    # ── Column Config: persist per report ──
    import json as _json
    columns = []
    items = models.get_report_items(report_id)
    all_items = models.get_all_report_items()  # fetch once, reuse

    # Which total accounts get the blue "subledger" link: ONLY those that drill to a
    # DIFFERENT report (e.g. BS "Detailed AR" → AR subledger). Plain subtotals whose
    # feeders live on this same report (e.g. "Total Bank Accounts", "Total AR") stay
    # black — everyone knows those are totals. Matches the click-to-jump behaviour.
    _feeder_reports = {}
    for _it in all_items:
        for _k in ('total_to_1', 'total_to_2', 'total_to_3', 'total_to_4', 'total_to_5'):
            _tgt = _it[_k]
            if _tgt:
                _feeder_reports.setdefault(_tgt, set()).add(_it['report_id'])
    sublink_ids = set()
    for _it in items:
        if _it['account_type'] == 'total' and _it['account_id']:
            _feeders = _feeder_reports.get(_it['acct_name'], set())
            if len(_feeders) == 1 and report_id not in _feeders:
                sublink_ids.add(_it['account_id'])

    # Check if user submitted new config via Apply button
    from_query = any(request.args.get(f'c{i}_end') or request.args.get(f'c{i}_type') in DERIVED_TYPES
                     for i in range(1, 7))

    if from_query:
        # Save to meta — _clean_column_cfg drops anything the chosen type does
        # not own, so a column that used to be "% of acct" does not keep its
        # account code after being switched to something else.
        cfg = {}
        for i in range(1, 7):
            for k in ('begin', 'end', 'type', 'label', 'a', 'b', 'acct'):
                v = request.args.get(f'c{i}_{k}', '')
                if v: cfg[f'c{i}_{k}'] = v
        cfg = _clean_column_cfg(cfg)
        models.set_meta(f'columns_{report_id}', _json.dumps(cfg))
    elif request.args.get('reset'):
        cfg = {}
        models.set_meta(f'columns_{report_id}', '')
    else:
        # Load from meta
        raw = models.get_meta(f'columns_{report_id}', '')
        cfg = _clean_column_cfg(_json.loads(raw)) if raw else {}
    
    def carg(k, d=''):
        return cfg.get(k, d)
    
    has_custom = _has_custom_columns(cfg)

    all_columns = _column_slots(cfg, report_id, items, all_items) if has_custom else [None] * 6
    columns = [c for c in all_columns if c is not None]
    
    if not columns:
        # Default columns derive from the fiscal-year SETTING: CY primary + PY
        # comparative. Income statements use periodic amounts (full-year activity);
        # balance reports (BS, AR/AP subledgers, etc.) use perpetual as-of year-end.
        # Journals/registers keep the simple single all-dates column.
        anc = models.fiscal_anchor()
        rname = (report['name'] or '').upper()
        JOURNALS = {'AJE', 'TRX'}
        if anc and rname not in JOURNALS:
            income_style = rname == 'IS'
            cy_begin = anc['cy_start'] if income_style else ''
            py_begin = anc['py_start'] if income_style else ''
            cy_data = models.compute_report_column(report_id,
                date_from=cy_begin or None, date_to=anc['cy_end'],
                _display_items=items, _all_items=all_items)
            py_data = models.compute_report_column(report_id,
                date_from=py_begin or None, date_to=anc['py_end'],
                _display_items=items, _all_items=all_items)
            columns.append({'type': 'actual', 'begin': cy_begin, 'end': anc['cy_end'],
                           'label': str(anc['fy']), 'data': cy_data})
            columns.append({'type': 'actual', 'begin': py_begin, 'end': anc['py_end'],
                           'label': str(anc['fy'] - 1), 'data': py_data})
        else:
            # Journals / no fiscal_year set: single column, all dates.
            col_data = models.compute_report_column(report_id,
                _display_items=items, _all_items=all_items)
            columns.append({'type': 'actual', 'begin': '', 'end': '',
                           'label': 'Balance', 'data': col_data})
    
    # For change/pct columns, compute deltas; for spacers, fill with None
    for col in columns:
        if col['type'] == 'change':
            a_idx, b_idx = col['a'], col['b']
            if a_idx < len(columns) and b_idx < len(columns):
                a_data = columns[a_idx].get('data', [])
                b_data = columns[b_idx].get('data', [])
                if a_data and b_data:
                    change_data = []
                    for j in range(len(a_data)):
                        item_a, bal_a = a_data[j]
                        _, bal_b = b_data[j] if j < len(b_data) else (None, 0)
                        change_data.append((item_a, bal_b - bal_a))
                    col['data'] = change_data
        elif col['type'] == 'pct_change':
            a_idx, b_idx = col['a'], col['b']
            if a_idx < len(columns) and b_idx < len(columns):
                a_data = columns[a_idx].get('data', [])
                b_data = columns[b_idx].get('data', [])
                if a_data and b_data:
                    pct_data = []
                    for j in range(len(a_data)):
                        item_a, bal_a = a_data[j]
                        _, bal_b = b_data[j] if j < len(b_data) else (None, 0)
                        if bal_a != 0:
                            pct = round((bal_b - bal_a) * 10000 / abs(bal_a))  # basis points → will format later
                        else:
                            pct = 0  # avoid div/0
                        pct_data.append((item_a, pct))
                    col['data'] = pct_data
        elif col['type'] == 'spacer':
            # Spacer needs data array matching row count for iteration
            pass  # handled below
    
    # Build unified row structure: items + balance per column
    rows = []
    base_items = columns[0]['data'] if columns[0].get('data') else []
    # If first column is a spacer, find first actual data column
    if not base_items:
        for col in columns:
            if col.get('data'):
                base_items = col['data']
                break
    for idx, (item, _) in enumerate(base_items):
        bals = []
        for col in columns:
            if col['type'] == 'spacer':
                bals.append(None)  # None = spacer cell
            elif col.get('data') and idx < len(col['data']):
                bals.append(col['data'][idx][1])
            else:
                bals.append(0)
        rows.append((item, bals))
    
    col_labels = [c['label'] for c in columns]
    col_types = [c['type'] for c in columns]
    
    _elapsed = (_time.time() - _t0) * 1000
    app.logger.info(f'Report {report_id} data computed in {_elapsed:.0f}ms')
    
    # GIFI mapping data for BS/IS reports
    gifi_map = {}
    if report['name'] in ('BS', 'IS'):
        with models.get_db() as db:
            for r in db.execute("SELECT id, gifi_code FROM accounts WHERE gifi_code != ''"):
                gifi_map[r['id']] = r['gifi_code']

    # Lead-sheet marks — on every report, because an account belongs to a lead
    # wherever it appears (the working-paper index, not a tax mapping).
    ls_map = {}
    with models.get_db() as db:
        for r in db.execute("SELECT id, leadsheet FROM accounts WHERE leadsheet != ''"):
            ls_map[r['id']] = r['leadsheet']

    return render_template('report.html', report=report, rows=rows,
                         col_labels=col_labels, col_types=col_types,
                         columns=columns, all_columns=all_columns, cfg=cfg,
                         company=company, hide_zero=hide_zero, show_setup=show_setup,
                         gifi_map=gifi_map, gifi_codes=models.GIFI_CODES,
                         sublink_ids=sublink_ids, ls_map=ls_map, show_ls=True,
                         ls_max=models.LEADSHEET_MAX,
                         show_gifi=(report['name'] in ('BS', 'IS')))

@app.route('/report/<int:report_id>/print')
def report_print(report_id):
    """Print-friendly report view. Options: begin, end, hide_zero, ledger, debit, credit."""
    report = models.get_report(report_id)
    if not report:
        return 'Report not found', 404
    company = models.get_meta('company_name', 'My Books')
    begin = request.args.get('begin', '')
    end = request.args.get('end', '')
    hide_zero = request.args.get('hide_zero', '0') == '1'
    mode = request.args.get('mode', 'report')  # report, ledger, debit, credit

    if mode == 'ledger':
        # Full ledger dump for all posting accounts in this report
        items = models.get_report_items(report_id)
        ledger_data = []
        for it in items:
            if it['account_id'] and it['account_type'] == 'posting':
                entries = models.get_ledger(it['account_id'],
                    date_from=begin or None, date_to=end or None)
                if entries or not hide_zero:
                    ledger_data.append({
                        'name': it['acct_name'],
                        'desc': it['description'] or it['acct_desc'] or it['acct_name'],
                        'entries': entries
                    })
        return render_template('print_ledger.html', report=report, company=company,
                             ledger_data=ledger_data, begin=begin, end=end, mode=mode,
                             now=__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M'))
    else:
        # Check for saved multi-column config
        raw = models.get_meta(f'columns_{report_id}', '')
        cfg = json.loads(raw) if raw else {}
        def carg(k, d=''):
            return cfg.get(k, d)

        has_custom = _has_custom_columns(cfg)

        if has_custom:
            # Same builder as the screen — one definition, so a column type can
            # never exist in one place and be missing from the other.
            items = models.get_report_items(report_id)
            all_items = models.get_all_report_items()
            columns = [c for c in _column_slots(cfg, report_id, items, all_items) if c is not None]

            if not columns:
                # Fallback to single column
                col_data = models.compute_report_column(report_id,
                    _display_items=items, _all_items=all_items)
                columns = [{'type': 'actual', 'begin': '', 'end': '',
                           'label': 'Balance', 'data': col_data}]

            # Compute change/pct columns
            for col in columns:
                if col['type'] == 'change':
                    a_idx, b_idx = col['a'], col['b']
                    if a_idx < len(columns) and b_idx < len(columns):
                        a_data = columns[a_idx].get('data', [])
                        b_data = columns[b_idx].get('data', [])
                        if a_data and b_data:
                            change_data = []
                            for j in range(len(a_data)):
                                item_a, bal_a = a_data[j]
                                _, bal_b = b_data[j] if j < len(b_data) else (None, 0)
                                change_data.append((item_a, bal_b - bal_a))
                            col['data'] = change_data
                elif col['type'] == 'pct_change':
                    a_idx, b_idx = col['a'], col['b']
                    if a_idx < len(columns) and b_idx < len(columns):
                        a_data = columns[a_idx].get('data', [])
                        b_data = columns[b_idx].get('data', [])
                        if a_data and b_data:
                            pct_data = []
                            for j in range(len(a_data)):
                                item_a, bal_a = a_data[j]
                                _, bal_b = b_data[j] if j < len(b_data) else (None, 0)
                                if bal_a != 0:
                                    pct = round((bal_b - bal_a) * 10000 / abs(bal_a))
                                else:
                                    pct = 0
                                pct_data.append((item_a, pct))
                            col['data'] = pct_data

            # Build rows
            base_items = columns[0]['data'] if columns[0].get('data') else []
            if not base_items:
                for col in columns:
                    if col.get('data'):
                        base_items = col['data']
                        break
            rows = []
            for idx, (item, _) in enumerate(base_items):
                bals = []
                for col in columns:
                    if col['type'] == 'spacer':
                        bals.append(None)
                    elif col.get('data') and idx < len(col['data']):
                        bals.append(col['data'][idx][1])
                    else:
                        bals.append(0)
                if hide_zero and item.get('item_type') in ('account',) and all((b is None or b == 0) for b in bals):
                    continue
                rows.append((item, bals))

            col_labels = [c['label'] for c in columns]
            col_types = [c['type'] for c in columns]

            return render_template('print_report.html', report=report, company=company,
                             has_marks=any((i.get('ref_mark') or '') for i, _ in rows) if rows else False,
                                 rows=rows, begin='', end='', hide_zero=hide_zero,
                                 col_labels=col_labels, col_types=col_types, multicol=True,
                                 now=__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M'))
        else:
            # Single column print (original behavior)
            col_data = models.compute_report_column(report_id,
                date_from=begin or None, date_to=end or None)
            rows = []
            for item, bal in col_data:
                if hide_zero and bal == 0 and item.get('item_type') in ('account',):
                    continue
                rows.append((item, bal))
            return render_template('print_report.html', report=report, company=company,
                             has_marks=any((i.get('ref_mark') or '') for i, _ in rows) if rows else False,
                                 rows=rows, begin=begin, end=end, hide_zero=hide_zero,
                                 col_labels=['Balance'], col_types=['actual'], multicol=False,
                                 now=__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M'))

@app.route('/report/<int:report_id>/csv')
def report_csv(report_id):
    """Export report as CSV."""
    import csv, io
    report = models.get_report(report_id)
    if not report:
        return 'Not found', 404
    begin = request.args.get('begin', '')
    end = request.args.get('end', '')
    col_data = models.compute_report_column(report_id,
        date_from=begin or None, date_to=end or None)
    
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['Description', 'Account', 'Type', 'Balance'])
    for item, bal in col_data:
        desc = item.get('description') or item.get('acct_desc') or item.get('acct_name', '')
        name = item.get('acct_name', '')
        itype = item.get('item_type', '')
        if itype in ('label', 'separator'):
            w.writerow([desc, '', itype, ''])
        else:
            w.writerow([desc, name, itype, f'{bal/100:.2f}' if bal else ''])
    
    from flask import Response
    return Response(buf.getvalue(), mimetype='text/csv',
                   headers={'Content-Disposition': f'attachment; filename={report["name"]}.csv'})

# ─── Multi-Column (13-Column) Report ─────────────────────────────

@app.route('/report/<int:report_id>/multicol')
def report_multicol(report_id):
    """Generate a multi-column monthly report as landscape PDF.
    Params: start (yyyy-mm), months (1-24), mode (periodic|cumulative), hide_zero."""
    from calendar import monthrange
    from datetime import date as _date
    
    report = models.get_report(report_id)
    if not report:
        return 'Report not found', 404
    company = models.get_meta('company_name', 'My Books')
    
    start = request.args.get('start', '')
    num_months = int(request.args.get('months', '12'))
    mode = request.args.get('mode', 'periodic')  # periodic or cumulative
    hide_zero = request.args.get('hide_zero', '0') == '1'
    
    if not start:
        # Show config form
        fye = models.get_meta('fiscal_year_end', '12-31')
        fy_year = models.get_meta('fiscal_year', str(_date.today().year))
        # Default start: beginning of fiscal year
        fye_mm = fye.split('-')[0] if '-' in fye else '01'
        default_start = f'{fy_year}-{fye_mm}'
        return f'''<!DOCTYPE html><html><head><title>Multi-Column Report</title>
        <style>body{{font-family:system-ui;max-width:500px;margin:40px auto;background:#1a1a2e;color:#e0e0e0}}
        label{{display:block;margin:10px 0 4px;font-size:13px;font-weight:600}}
        input,select{{padding:6px 10px;font-size:14px;border:1px solid #444;border-radius:4px;
        background:#2a2a4a;color:#e0e0e0;font-family:monospace}}
        .btn{{padding:8px 20px;background:#4a6fa5;color:white;border:none;border-radius:4px;
        cursor:pointer;font-size:14px;margin-top:14px}}
        h2{{color:#8ab4f8}}</style></head><body>
        <h2>{report["name"]} — Multi-Column Report</h2>
        <form>
        <label>Start Month (yyyy-mm)</label>
        <input name="start" value="{default_start}" placeholder="2025-01">
        <label>Number of Months</label>
        <input name="months" value="12" type="number" min="1" max="36" style="width:60px">
        <label>Mode</label>
        <select name="mode">
        <option value="periodic">Periodic (each month standalone — use for Income Statement)</option>
        <option value="cumulative">Cumulative (running balance — use for Balance Sheet)</option>
        </select>
        <label><input type="checkbox" name="hide_zero" value="1" style="width:auto"> Hide zero rows</label>
        <br><button type="submit" class="btn">Generate PDF</button>
        </form></body></html>'''
    
    # Build month ranges
    import re
    m = re.match(r'^(\d{4})-(\d{2})$', start)
    if not m:
        return 'Invalid start format. Use yyyy-mm', 400
    
    start_year, start_month = int(m.group(1)), int(m.group(2))
    
    month_ranges = []
    for i in range(num_months):
        y = start_year + (start_month - 1 + i) // 12
        mo = (start_month - 1 + i) % 12 + 1
        last_day = monthrange(y, mo)[1]
        d_from = f'{y:04d}-{mo:02d}-01'
        d_to = f'{y:04d}-{mo:02d}-{last_day:02d}'
        
        if mode == 'cumulative':
            # BS mode: from the beginning of time to end of this month
            month_ranges.append((None, d_to, f'{y:04d}-{mo:02d}'))
        else:
            # IS mode: just this month
            month_ranges.append((d_from, d_to, f'{y:04d}-{mo:02d}'))
    
    # Compute columns — prefetch items once
    items = models.get_report_items(report_id)
    all_items = models.get_all_report_items()
    
    columns = []
    for d_from, d_to, label in month_ranges:
        col_data = models.compute_report_column(report_id,
            date_from=d_from, date_to=d_to,
            _display_items=items, _all_items=all_items)
        columns.append((label, col_data))
    
    # Build rows: use first column for structure
    if not columns or not columns[0][1]:
        return 'No data', 404
    
    base = columns[0][1]
    row_data = []
    for idx, (item, _) in enumerate(base):
        vals = []
        for _, col in columns:
            if idx < len(col):
                vals.append(col[idx][1])
            else:
                vals.append(0)
        # Total column
        if mode == 'cumulative':
            total = vals[-1] if vals else 0  # last month IS the cumulative total
        else:
            total = sum(vals)
        
        if hide_zero and item.get('item_type') == 'account' and all(v == 0 for v in vals):
            continue
        row_data.append((item, vals, total))
    
    # Generate landscape PDF
    pdf_bytes = _multicol_pdf(report, company, columns, row_data, mode, num_months)
    
    from flask import Response
    resp = Response(pdf_bytes, mimetype='application/pdf')
    resp.headers['Content-Disposition'] = f'inline; filename="{report["name"]}_multicol.pdf"'
    return resp


def _multicol_pdf(report, company, columns, row_data, mode, num_months):
    """Generate a landscape multi-column PDF report."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except ImportError:
        from flask import abort
        abort(500, 'reportlab not installed')
    import io

    pw, ph = letter[1], letter[0]  # landscape
    margin = 36  # 0.5"

    font, font_b = _setup_fonts()

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(pw, ph))
    
    ncols = len(columns)
    show_total = (mode == 'periodic')  # periodic gets a total column; cumulative doesn't need one
    num_data_cols = ncols + (1 if show_total else 0)
    
    # Layout calculation
    usable = pw - 2 * margin
    # Account description gets leftover space after amount columns
    # Each amount column needs ~58pt for numbers like (999,999)
    col_w = 56
    if num_data_cols > 13:
        col_w = 50  # squeeze for >13 columns
    if num_data_cols > 18:
        col_w = 44
    
    desc_w = usable - (num_data_cols * col_w)
    if desc_w < 100:
        desc_w = 100
        col_w = max(40, (usable - desc_w) // num_data_cols)
    
    fs = 6.5
    if num_data_cols > 14:
        fs = 6
    if num_data_cols > 20:
        fs = 5.5
    line_h = fs + 2.5
    
    col_starts = []
    x = margin + desc_w
    for i in range(num_data_cols):
        col_starts.append(x)
        x += col_w
    
    y = ph - margin
    page_num = 1
    right_edge = pw - margin
    
    mode_label = 'Cumulative' if mode == 'cumulative' else 'Periodic'
    first_label = columns[0][0] if columns else ''
    last_label = columns[-1][0] if columns else ''
    
    def header():
        nonlocal y
        c.setFont(font_b, 8)
        c.drawString(margin, ph - margin + 5, f'{company} — {report["name"]} ({mode_label})')
        c.setFont(font, 6)
        c.drawString(margin, ph - margin - 4, f'{first_label} to {last_label}')
        c.drawRightString(right_edge, ph - margin + 5, f'Page {page_num}')
        y = ph - margin - 14
    
    def col_header():
        nonlocal y
        c.setFont(font_b, fs - 0.5)
        c.drawString(margin, y, 'Account')
        for i, (label, _) in enumerate(columns):
            # Show month abbreviation: yyyy-mm -> Mon
            try:
                yr, mo = label.split('-')
                from calendar import month_abbr
                short = month_abbr[int(mo)]
                if ncols <= 12:
                    short = f'{short} {yr[2:]}'
                c.drawRightString(col_starts[i] + col_w - 2, y, short)
            except:
                c.drawRightString(col_starts[i] + col_w - 2, y, label)
        if show_total:
            c.drawRightString(col_starts[-1] + col_w - 2, y, 'Total')
        y -= 2
        c.setLineWidth(0.4)
        c.line(margin, y, right_edge, y)
        y -= line_h
    
    def check_page():
        nonlocal y, page_num
        if y < margin + 2 * line_h:
            c.showPage()
            page_num += 1
            header()
            col_header()
    
    def fmt_val(v):
        """Format cents as string with parens for negative, em dash for zero."""
        if v == 0:
            return '\u2014'
        if v < 0:
            return f'({abs(v)/100:,.0f})'
        return f'{v/100:,.0f}'
    
    header()
    col_header()
    
    for item, vals, total in row_data:
        check_page()
        itype = item.get('item_type', 'account')
        indent = item.get('indent', 0) or 0
        
        if itype == 'separator':
            style = item.get('sep_style', 'single')
            if style == 'double':
                c.setLineWidth(0.5)
                c.line(col_starts[0], y + line_h * 0.4, right_edge, y + line_h * 0.4)
                c.line(col_starts[0], y + line_h * 0.4 - 2, right_edge, y + line_h * 0.4 - 2)
            elif style == 'blank':
                pass
            else:
                c.setLineWidth(0.3)
                c.line(col_starts[0], y + line_h * 0.4, right_edge, y + line_h * 0.4)
            y -= line_h
            continue
        
        desc = item.get('description') or item.get('acct_desc') or item.get('acct_name') or ''
        is_total = itype == 'total'
        fn = font_b if is_total else font
        
        c.setFont(fn, fs)
        # Truncate description to fit
        max_desc_chars = int(desc_w / (fs * 0.6))
        display_desc = '  ' * indent + desc
        c.drawString(margin, y, display_desc[:max_desc_chars])
        
        # Draw month values
        for i, v in enumerate(vals):
            c.drawRightString(col_starts[i] + col_w - 2, y, fmt_val(v))

        # Total column
        if show_total:
            c.setFont(font_b, fs)
            c.drawRightString(col_starts[-1] + col_w - 2, y, fmt_val(total))
        
        y -= line_h
    
    c.save()
    return buf.getvalue()


# ─── Report Item Management API ──────────────────────────────────

@app.route('/api/report/<int:report_id>/items')
def api_report_items(report_id):
    """Return report items as JSON for placement dialogs."""
    try:
        items = models.get_report_items(report_id)
        result = []
        for item in items:
            result.append({
                'id': item['id'], 'item_type': item['item_type'],
                'description': item['description'],
                'acct_name': item['acct_name'] or '',
                'acct_desc': item['acct_desc'] or '',
                'indent': item['indent'], 'position': item['position'],
            })
        return jsonify({'ok': True, 'items': result})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/report/<int:report_id>/add-item', methods=['POST'])
def api_add_report_item(report_id):
    """Add a new account/item to a report."""
    try:
        item_type = request.form.get('item_type', 'account')
        acct_name = request.form.get('account_name', '').strip().upper()
        description = request.form.get('description', '')
        total_to_1 = request.form.get('total_to_1', '').strip().upper()
        after_pos = int(request.form.get('after_position', '0'))
        indent = int(request.form.get('indent', '2'))
        sep_style = request.form.get('sep_style', '')
        nb = request.form.get('normal_balance', 'D')
        
        account_id = None
        if item_type in ('account', 'total') and acct_name:
            acct = models.get_account_by_name(acct_name)
            if not acct:
                atype = 'total' if item_type == 'total' else 'posting'
                account_id = models.add_account(acct_name, nb, description, atype)
            else:
                account_id = acct['id']
                # Update the existing account's normal_balance and description
                # in case the user is correcting a previous mistake
                with models.get_db() as db:
                    db.execute("UPDATE accounts SET normal_balance=? WHERE id=?", (nb, acct['id']))
                    if description:
                        db.execute("UPDATE accounts SET description=? WHERE id=?", (description, acct['id']))
        
        position = after_pos + 5 if after_pos else None
        
        models.add_report_item(report_id, item_type, description, account_id,
                              indent, position, total_to_1, sep_style=sep_style)
        
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/report-item/<int:item_id>/update', methods=['POST'])
def api_update_report_item(item_id):
    """Update fields on a report item (total-to's, description, indent, sep_style)."""
    try:
        models.check_trx_pinned(item_id)   # …and it is not renamed or re-wired
        data = request.get_json() if request.is_json else request.form
        kwargs = {}
        for field in ('description','indent','total_to_1','total_to_2','total_to_3',
                      'total_to_4','total_to_5','total_to_6','sep_style','position','item_type'):
            if field in data:
                val = data[field]
                if field == 'indent':
                    val = int(val)
                elif field == 'position':
                    val = int(val)
                elif field.startswith('total_to_'):
                    val = str(val).strip().upper()
                kwargs[field] = val
        models.update_report_item(item_id, **kwargs)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/report-item/<int:item_id>/delete', methods=['POST'])
def api_delete_report_item(item_id):
    """Delete a report item (with safety checks)."""
    try:
        models.delete_report_item(item_id)
        return jsonify({'ok': True})
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/report-item/<int:item_id>/move', methods=['POST'])
def api_move_report_item(item_id):
    """Move a report item up or down."""
    try:
        models.check_trx_pinned(item_id)   # the TRX head does not move
        data = request.get_json() if request.is_json else request.form
        direction = int(data.get('direction', 0))
        if direction not in (-1, 1):
            return jsonify({'ok': False, 'error': 'Direction must be -1 or 1'})
        moved = models.move_report_item(item_id, direction)
        return jsonify({'ok': True, 'moved': moved})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/account/<int:account_id>/rename', methods=['POST'])
def api_rename_account(account_id):
    """Rename an account's description (not the code name)."""
    try:
        data = request.get_json() if request.is_json else request.form
        # An ABSENT field means "leave it alone" — sending only account_number
        # used to blank the description, because '' is not None.
        desc = data.get('description')
        acct_num = data.get('account_number')
        if desc is None and acct_num is None:
            return jsonify({'ok': False, 'error': 'Nothing to change'})
        models.update_account(account_id, description=desc, account_number=acct_num)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/reports/reorder', methods=['POST'])
def api_reorder_reports():
    """Reorder reports. Expects JSON: {order: [id, id, id, ...]}"""
    try:
        data = request.get_json()
        order = data.get('order', [])
        with models.get_db() as db:
            for i, rid in enumerate(order):
                db.execute("UPDATE reports SET sort_order=? WHERE id=?", (i * 10, rid))
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/cascade-dates', methods=['POST'])
def api_cascade_dates():
    """One-click column setup: CY, PY, $ chg, % chg on both BS and IS."""
    try:
        import json as _json
        data = request.get_json()
        ye_date = data.get('ye_date', '').strip()
        if not ye_date or len(ye_date) != 10:
            return jsonify({'ok': False, 'error': 'Enter fiscal year-end as YYYY-MM-DD'})

        # Parse YE date and compute fiscal periods
        ye_y, ye_m, ye_d = int(ye_date[:4]), int(ye_date[5:7]), int(ye_date[8:10])
        ye = date(ye_y, ye_m, ye_d)

        # PY year-end = same month/day, one year back
        py_ye = date(ye_y - 1, ye_m, ye_d)

        # FY start = day after PY year-end
        from datetime import timedelta
        fy_start = py_ye + timedelta(days=1)
        py_fy_start = date(fy_start.year - 1, fy_start.month, fy_start.day)

        # Find BS and IS report IDs
        with models.get_db() as db:
            bs = db.execute("SELECT id FROM reports WHERE name='BS'").fetchone()
            is_ = db.execute("SELECT id FROM reports WHERE name='IS'").fetchone()
            if not bs or not is_:
                return jsonify({'ok': False, 'error': 'BS or IS report not found'})

            # BS columns: col1 = CY end, col2 = PY end, col3 = $ chg, col4 = % chg
            bs_cfg = {
                'c1_end': ye.isoformat(),
                'c1_label': str(ye_y),
                'c2_end': py_ye.isoformat(),
                'c2_label': str(ye_y - 1),
                'c3_type': 'change', 'c3_a': '2', 'c3_b': '1', 'c3_label': '$ chg',
                'c4_type': 'pct_change', 'c4_a': '2', 'c4_b': '1', 'c4_label': '% chg',
            }
            models.set_meta(f'columns_{bs["id"]}', _json.dumps(bs_cfg))

            # IS columns: col1 = CY period, col2 = PY period, col3 = $ chg, col4 = % chg
            is_cfg = {
                'c1_begin': fy_start.isoformat(),
                'c1_end': ye.isoformat(),
                'c1_label': str(ye_y),
                'c2_begin': py_fy_start.isoformat(),
                'c2_end': py_ye.isoformat(),
                'c2_label': str(ye_y - 1),
                'c3_type': 'change', 'c3_a': '2', 'c3_b': '1', 'c3_label': '$ chg',
                'c4_type': 'pct_change', 'c4_a': '2', 'c4_b': '1', 'c4_label': '% chg',
            }
            models.set_meta(f'columns_{is_["id"]}', _json.dumps(is_cfg))

        return jsonify({
            'ok': True,
            'bs': {'col1': ye.isoformat(), 'col2': py_ye.isoformat()},
            'is': {'col1': f'{fy_start.isoformat()} to {ye.isoformat()}',
                   'col2': f'{py_fy_start.isoformat()} to {py_ye.isoformat()}'},
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/account/flip-sign', methods=['POST'])
def api_flip_sign():
    """Flip an account's normal_balance between D and C."""
    try:
        data = request.get_json()
        account_id = int(data.get('account_id', 0))
        with models.get_db() as db:
            acct = db.execute("SELECT id, name, normal_balance FROM accounts WHERE id=?",
                              (account_id,)).fetchone()
            if not acct:
                return jsonify({'ok': False, 'error': 'Account not found'})
            new_nb = 'C' if acct['normal_balance'] == 'D' else 'D'
            db.execute("UPDATE accounts SET normal_balance=? WHERE id=?", (new_nb, account_id))
        return jsonify({'ok': True, 'name': acct['name'], 'old': acct['normal_balance'], 'new': new_nb})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/gifi/set', methods=['POST'])
def api_set_gifi():
    """Set GIFI code on an account via AJAX."""
    try:
        data = request.get_json()
        account_id = int(data.get('account_id', 0))
        gifi_code = data.get('gifi_code', '').strip()
        if gifi_code and gifi_code not in models.GIFI_CODES:
            return jsonify({'ok': False, 'error': f'Invalid GIFI code: {gifi_code}'})
        with models.get_db() as db:
            db.execute("UPDATE accounts SET gifi_code = ? WHERE id = ?", (gifi_code, account_id))
        desc = models.GIFI_CODES.get(gifi_code, '') if gifi_code else ''
        return jsonify({'ok': True, 'gifi_code': gifi_code, 'description': desc})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/gifi/codes')
def api_gifi_codes():
    """Return all GIFI codes as JSON for the dropdown."""
    return jsonify([{'code': k, 'desc': v} for k, v in sorted(models.GIFI_CODES.items())])


@app.route('/api/report/<int:report_id>/description', methods=['POST'])
def api_report_description(report_id):
    """Update a report's description."""
    try:
        data = request.get_json()
        desc = data.get('description', '').strip()
        if not desc:
            return jsonify({'ok': False, 'error': 'Description cannot be empty'})
        models.update_report(report_id, description=desc)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# ─── Account Ledger ─────────────────────────────────────────────────

@app.route('/account/<int:account_id>')
@app.route('/ledger/<int:account_id>')
def account_ledger(account_id):
    if not models.get_db_path():
        return redirect(url_for('library'))
    account = models.get_account(account_id)
    if not account:
        flash('Account not found', 'error')
        return redirect(url_for('home'))

    # Total (accumulator) accounts are NOT data-entry ledgers — they receive total-to
    # from a subledger. Jump straight to that subledger; if there's no single one
    # (e.g. a grand total), fall through to a read-only notice instead of a misleading
    # entry screen.
    is_total = account['account_type'] == 'total'
    subledger_report = None
    total_balance = 0
    if is_total:
        from_rid = request.args.get('from_report', '')
        target_rid = models.subledger_report_for_total(account['name'])
        if target_rid:
            if str(target_rid) != str(from_rid):
                return redirect(url_for('report_view', report_id=target_rid))
            subledger_report = models.get_report(target_rid)
        total_balance = models.trace_account(account['name']).get('display', 0)

    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')
    show_all = request.args.get('all', '')
    sort_by = request.args.get('sort', '')  # 'description' for AJE sort

    # Fiscal-year-based default window (browser only — models layer unchanged)
    opening_balance = 0
    windowed = False
    prev_from = ''

    # Determine parent report early — needed for AJE detection before windowing
    from_report_id = request.args.get('from_report', '')
    parent_report = None
    if from_report_id:
        parent_report = models.get_report(int(from_report_id))
    if not parent_report:
        parent_report = models.find_report_for_account(account_id)

    # AJE/TRX accounts: show ALL transactions, no windowing.
    # These reports have few entries — no performance concern.
    # BS/IS accounts keep the 2-year window to avoid loading decades of bank data.
    # Check parent_report first, then fall back to checking if the account lives
    # on ANY non-BS/IS report (so it works regardless of how the user navigated here).
    is_aje = parent_report and parent_report['name'] not in ('BS', 'IS')
    if not is_aje and not parent_report:
        with models.get_db() as db:
            rpt = db.execute("""SELECT rp.name FROM report_items ri
                JOIN reports rp ON ri.report_id = rp.id
                WHERE ri.account_id = ? AND rp.name NOT IN ('BS','IS')
                LIMIT 1""", (account_id,)).fetchone()
            if rpt:
                is_aje = True

    # Helper: compute FY start month/day from fiscal_year_end meta
    def _fy_start_md():
        fye_md = models.get_meta('fiscal_year_end', '')
        if not fye_md:
            fye_md = '01-01'
        try:
            fye_m, fye_d = int(fye_md.split('-')[0]), int(fye_md.split('-')[1])
        except (ValueError, IndexError):
            fye_m, fye_d = 1, 1
        if fye_m == 12 and fye_d == 31:
            return 1, 1
        from calendar import monthrange
        _, last = monthrange(2000, fye_m)  # 2000 is leap year — safe for monthrange
        if fye_d >= last:
            return (fye_m + 1 if fye_m < 12 else 1), 1
        return fye_m, fye_d + 1

    # Fiscal anchor: defaults derive from the BOOK's fiscal-year setting, not the
    # wall clock. cy_end also drives the post-YE divider in the template.
    anc = models.fiscal_anchor()
    cy_end = anc['cy_end'] if anc else ''

    if not show_all and not is_aje:
        if not date_from and not date_to:
            # Auto-window anchored to the fiscal-year SETTING. Default view = current
            # fiscal year (CY) onward, so post-YE activity stays visible below; PY is
            # one click back via "Load prior year".
            if anc:
                date_from = anc['cy_start']
                prev_from = anc['py_start']
            else:
                # Fallback (no fiscal_year set): wall-clock current FY onward.
                fy_start_m, fy_start_d = _fy_start_md()
                today = date.today()
                current_fy_start = date(today.year, fy_start_m, fy_start_d)
                if current_fy_start > today:
                    current_fy_start = date(today.year - 1, fy_start_m, fy_start_d)
                date_from = current_fy_start.isoformat()
                prev_from = date(current_fy_start.year - 1, fy_start_m, fy_start_d).isoformat()
            windowed = True
        elif date_from and not date_to:
            # User clicked "Load prior year" — keep the bar visible
            fy_start_m, fy_start_d = _fy_start_md()
            from_date = date(*(int(p) for p in date_from.split('-')))
            prev_from = date(from_date.year - 1, fy_start_m, fy_start_d).isoformat()
            windowed = True

    # Compute opening balance for windowed or user-filtered views
    if date_from:
        day_before = date(*(int(p) for p in date_from.split('-'))) - timedelta(days=1)
        raw_opening = models.get_account_balance(account_id, date_to=day_before.isoformat())
        sign = 1 if account['normal_balance'] == 'D' else -1
        opening_balance = raw_opening * sign

    entries = models.get_ledger(account_id, date_from or None, date_to or None,
                                opening_balance=opening_balance)

    # AJE sort by description (display-only — doesn't affect DB order)
    if is_aje and sort_by == 'description':
        entries = sorted(entries, key=lambda e: (e.get('description', '') or '').lower())

    # A year's AJE journal is an ordinary account that supports TWO VIEWS of the
    # same rows. Which one you land in follows HOW YOU ARRIVED: come down from the
    # AJE report and you meant to work on adjustments, so they are grouped; arrive
    # by F5 from another account and you are jumping through accounts, so it is
    # the standard ledger like everywhere else. `?view=` always wins.
    # F2 is DRILL UP, not "go back". It must land on the report this account
    # actually belongs to — `parent_report` is the referrer whenever a
    # `from_report` was carried in, so F5-ing from an AJE into a balance-sheet
    # account and hitting F2 used to return to the AJE report.
    home_report = models.find_report_for_account(account_id)
    is_aje_journal = models.is_aje_journal(account_id)
    aje_view = False
    aje_entries = []
    if is_aje_journal:
        want = (request.args.get('view', '') or '').strip().lower()
        if want in ('aje', 'ledger'):
            aje_view = (want == 'aje')
        else:
            aje_view = (request.args.get('from_report', '') == str(models.aje_report()['id']))
        if aje_view:
            aje_entries = models.aje_groups(account_id)
    balance = models.get_account_balance(account_id)
    sign = 1 if account['normal_balance'] == 'D' else -1
    balance = balance * sign
    if is_total:
        balance = total_balance  # total accounts have no direct lines; show accumulated value
    company = models.get_meta('company_name', 'My Books')

    # Default AJE ref prefix from fiscal year
    fy = models.get_meta('fiscal_year', '')
    default_prefix = f"{fy[-2:]}AJE" if fy else 'AJE'

    return render_template('ledger.html', account=account, entries=entries,
                         balance=balance, date_from=date_from, date_to=date_to,
                         company=company, today=date.today().isoformat(),
                         parent_report=parent_report, default_prefix=default_prefix,
                         opening_balance=opening_balance, windowed=windowed,
                         prev_from=prev_from, show_all=show_all, cy_end=cy_end,
                         is_aje=is_aje, sort_by=sort_by,
                         is_total=is_total, subledger_report=subledger_report,
                         # v116 — the conversion ledger carries the way back into
                         # the opening-balance grid (see templates/ledger.html).
                         is_conversion_account=(account['name'].upper()
                                                == models.CONVERSION_ACCT),
                         openings_posted=bool(models.opening_batch()),
                         # v125 — a year's AJE journal is the SAME ledger in a
                         # different MODE: read-only rows, and the reference is
                         # the door into the entry screen. Not a second screen.
                         is_aje_journal=is_aje_journal, home_report=home_report,
                         aje_view=aje_view, aje_entries=aje_entries,
                         aje_next_ref=(models.next_aje_ref(account_id)
                                       if is_aje_journal else ''))

# ─── Adjusting entries ─────────────────────────────────────────────
# One door for AJEs, the same way opening balances got one: the operator states
# the adjustment, Grid lays it down in the house shape. See models.post_aje.

@app.route('/aje/new', methods=['GET', 'POST'])
def aje_new():
    """Start a year of adjustments. The name and description are SUGGESTIONS —
    house style wins if it differs."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    if request.method == 'POST':
        try:
            acct = models.create_aje_batch(request.form.get('account_name', ''),
                                           request.form.get('description', ''))
        except ValueError as e:
            flash(str(e), 'error')
            return redirect(url_for('aje_new'))
        return redirect(url_for('account_ledger', account_id=acct['id']))
    company = models.get_meta('company_name', 'My Books')
    return render_template('aje_new.html', company=company,
                           suggest=models.suggest_aje_batch(),
                           journals=models.aje_journals())


@app.route('/aje/<int:journal_id>/entry', methods=['GET', 'POST'])
def aje_entry(journal_id):
    """The entry screen: reference, date, one description, and the legs."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    acct = models.get_account(journal_id)
    if not acct or not models.is_aje_journal(journal_id):
        flash('Not a year of adjusting entries', 'error')
        return redirect(url_for('home'))

    company = models.get_meta('company_name', 'My Books')
    accounts = [a for a in models.get_accounts()
                if a['account_type'] == 'posting' and not models.is_aje_journal(a['id'])]
    anchor = models.fiscal_anchor()

    if request.method == 'POST':
        rows = []
        for name, amt in zip(request.form.getlist('row_acct'), request.form.getlist('row_amt')):
            if not (name or '').strip() and not (amt or '').strip():
                continue
            try:
                cents = models.parse_amount(amt) if (amt or '').strip() else 0
            except ValueError:
                flash(f"'{amt}' is not an amount.", 'error')
                cents = None
            if cents is None:
                return redirect(url_for('aje_entry', journal_id=journal_id))
            rows.append({'account': name, 'amount': cents})
        replace = (request.form.get('replace_ref', '') or '').strip() or None
        try:
            v = models.post_aje(journal_id, request.form.get('ref', ''),
                                request.form.get('entry_date', ''),
                                request.form.get('description', ''), rows,
                                replace_ref=replace)
        except ValueError as e:
            flash(str(e), 'error')
            return redirect(url_for('aje_entry', journal_id=journal_id,
                                    ref=replace or None))
        flash(f"{v['ref']} posted — {len(v['txn_ids'])} lines.", 'success')
        return redirect(url_for('account_ledger', account_id=journal_id))

    edit_ref = (request.args.get('ref', '') or '').strip()
    entry = models.aje_rows_from_ref(journal_id, edit_ref) if edit_ref else None
    if edit_ref and not entry:
        flash(f'No entry {edit_ref} in this journal', 'error')
        return redirect(url_for('account_ledger', account_id=journal_id))
    if not entry:
        entry = {'ref': models.next_aje_ref(journal_id),
                 'date': anchor['cy_end'] if anchor else date.today().isoformat(),
                 'description': '', 'rows': []}
    return render_template('aje_entry.html', company=company, account=acct, entry=entry,
                           accounts=accounts, editing=bool(edit_ref), replace_ref=edit_ref,
                           desc_max=models.AJE_DESC_MAX,
                           next_ref=models.next_aje_ref(journal_id))


@app.route('/aje/<int:journal_id>/delete', methods=['POST'])
def aje_delete(journal_id):
    """Remove one adjustment — every leg of it, or none."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    try:
        n = models.delete_aje(journal_id, request.form.get('ref', ''))
        flash(f"{request.form.get('ref', '')} deleted ({n} lines).", 'success')
    except ValueError as e:
        flash(str(e), 'error')
    return redirect(url_for('account_ledger', account_id=journal_id))


# ─── Cross-Account Jump ────────────────────────────────────────────

@app.route('/ledger-by-name/<name>')
def ledger_by_name(name):
    """Jump to an account's ledger by name. Used from F5 in distribution view."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    acct = models.get_account_by_name(name.upper())
    if not acct:
        flash(f'Account "{name}" not found', 'error')
        return redirect(url_for('home'))
    focus_txn = request.args.get('focus_txn', '')
    from_report = request.args.get('from_report', '')
    view = request.args.get('view', '')
    url = f'/ledger/{acct["id"]}'
    params = []
    if from_report: params.append(f'from_report={from_report}')
    if view: params.append(f'view={view}')
    if params: url += '?' + '&'.join(params)
    if focus_txn: url += f'#txn-{focus_txn}'
    return redirect(url)

@app.route('/jump/<int:txn_id>/from/<int:from_account_id>')
def jump_to_cross(txn_id, from_account_id):
    """Jump to the cross-account of a transaction."""
    txn, lines = models.get_transaction(txn_id)
    if not txn:
        flash('Transaction not found', 'error')
        return redirect(url_for('home'))
    
    for line in lines:
        if line['account_id'] != from_account_id:
            return redirect(url_for('account_ledger', account_id=line['account_id'],
                                  _anchor=f'txn-{txn_id}'))
    return redirect(url_for('account_ledger', account_id=from_account_id))

# ─── New Transaction ────────────────────────────────────────────────

@app.route('/transaction/new', methods=['GET', 'POST'])
@app.route('/transaction/new/in/<int:account_id>', methods=['GET', 'POST'])
def new_transaction(account_id=None):
    if not models.get_db_path():
        return redirect(url_for('library'))
    
    if request.method == 'POST':
        try:
            date_str = request.form['date']
            reference = request.form.get('reference', '')
            description = request.form.get('description', '')
            mode = request.form.get('mode', 'simple')
            
            if mode == 'simple':
                debit_name = request.form['debit_account']
                credit_name = request.form['credit_account']
                amount_str = request.form['amount']
                
                debit_acct = models.get_account_by_name(debit_name)
                credit_acct = models.get_account_by_name(credit_name)
                
                if not debit_acct:
                    flash(f'Account not found: {debit_name}', 'error')
                    return redirect(request.url)
                if not credit_acct:
                    flash(f'Account not found: {credit_name}', 'error')
                    return redirect(request.url)
                
                amount = models.parse_amount(amount_str)
                txn_id = models.add_simple_transaction(
                    date_str, reference, description,
                    debit_acct['id'], credit_acct['id'], amount)
            else:
                lines = []
                acct_names = request.form.getlist('line_account[]')
                amounts = request.form.getlist('line_amount[]')
                descs = request.form.getlist('line_desc[]')
                
                for acct_name, amt_str, desc in zip(acct_names, amounts, descs):
                    if not acct_name.strip() or not amt_str.strip():
                        continue
                    acct = models.get_account_by_name(acct_name.strip())
                    if not acct:
                        flash(f'Account not found: {acct_name}', 'error')
                        return redirect(request.url)
                    amount = models.parse_amount(amt_str)
                    lines.append((acct['id'], amount, desc))
                
                txn_id = models.add_transaction(date_str, reference, description, lines)
            
            if not description or not description.strip():
                flash('Transaction posted — warning: blank description', 'warning')
            else:
                flash('Transaction posted', 'success')
            if account_id:
                return redirect(url_for('account_ledger', account_id=account_id))
            return redirect(url_for('home'))

        except ValueError as e:
            flash(str(e), 'error')
            return redirect(request.url)
    
    account = models.get_account(account_id) if account_id else None
    accounts = models.get_accounts()
    today = date.today().isoformat()
    company = models.get_meta('company_name', 'My Books')
    return render_template('transaction.html', account=account, accounts=accounts,
                         today=today, company=company)

# ─── Edit Transaction ───────────────────────────────────────────────

@app.route('/transaction/<int:txn_id>/edit', methods=['GET', 'POST'])
def edit_transaction(txn_id):
    if not models.get_db_path():
        return redirect(url_for('library'))
    
    if request.method == 'POST':
        try:
            date_str = request.form['date']
            reference = request.form.get('reference', '')
            description = request.form.get('description', '')
            return_to = request.form.get('return_to', '')
            
            lines = []
            acct_names = request.form.getlist('line_account[]')
            amounts = request.form.getlist('line_amount[]')
            descs = request.form.getlist('line_desc[]')
            reconcileds = request.form.getlist('line_reconciled[]')
            doc_flags = request.form.getlist('line_doc_on_file[]')
            
            for i, (acct_name, amt_str, desc) in enumerate(zip(acct_names, amounts, descs)):
                if not acct_name.strip() or not amt_str.strip():
                    continue
                acct = models.get_account_by_name(acct_name.strip())
                if not acct:
                    flash(f'Account not found: {acct_name}', 'error')
                    return redirect(request.url)
                amount = models.parse_amount(amt_str)
                # reconciled may hold a TEXT tag (statement date) — pass through as-is
                rec = reconcileds[i] if i < len(reconcileds) else 0
                if str(rec).strip() in ('', '0'):
                    rec = 0
                doc = int(doc_flags[i]) if i < len(doc_flags) else 0
                lines.append((acct['id'], amount, desc, rec, doc))
            
            models.update_transaction(txn_id, date_str, reference, description, lines,
                                      allow_opening=_opening_confirmed())
            flash('Transaction updated', 'success')
            
            if return_to:
                return redirect(return_to)
            return redirect(url_for('home'))
            
        except ValueError as e:
            flash(str(e), 'error')
            return redirect(request.url)
    
    txn, lines = models.get_transaction(txn_id)
    if not txn:
        flash('Transaction not found', 'error')
        return redirect(url_for('home'))
    accounts = models.get_accounts()
    company = models.get_meta('company_name', 'My Books')
    return_to = request.args.get('return_to', '')
    from_account_id = request.args.get('from_account', '')
    from_report_id = request.args.get('from_report', '')
    
    # Build breadcrumb context
    from_account = models.get_account(int(from_account_id)) if from_account_id else None
    from_report = models.get_report(int(from_report_id)) if from_report_id else None
    
    return render_template('edit_transaction.html', txn=txn, lines=lines,
                         accounts=accounts, company=company, return_to=return_to,
                         from_account=from_account, from_report=from_report,
                         is_opening=models.is_opening_txn(txn_id))

# ─── Delete Transaction ────────────────────────────────────────────

@app.route('/transaction/<int:txn_id>/delete', methods=['POST'])
def delete_transaction(txn_id):
    try:
        models.delete_transaction(txn_id, allow_opening=_opening_confirmed())
        flash('Transaction deleted', 'success')
    except ValueError as e:
        flash(str(e), 'error')
    return_to = request.form.get('return_to', url_for('home'))
    return redirect(return_to)

@app.route('/api/bulk-delete', methods=['POST'])
def api_bulk_delete():
    """Delete multiple transactions at once."""
    try:
        data = request.get_json()
        txn_ids = data.get('txn_ids', [])
        if not txn_ids:
            return jsonify({'ok': False, 'error': 'No transactions selected'})
        deleted, skipped = models.bulk_delete_transactions(
            txn_ids, allow_opening=_opening_confirmed())
        msg = f'Deleted {deleted} transaction{"s" if deleted != 1 else ""}'
        if skipped:
            msg += (f' ({skipped} skipped — locked, reconciled, or opening balances. '
                    f'To replace a conversion, use the opening-balances screen.)')
        return jsonify({'ok': True, 'deleted': deleted, 'skipped': skipped, 'message': msg})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/gst-split/<int:txn_id>', methods=['POST'])
def api_gst_split(txn_id):
    """Split an existing 2-line transaction into 3 lines with GST.
    F8 = purchase (GST.IN / ITC), F9 = sale (GST.OUT / collected).
    Rate and accounts are configurable in Options."""
    try:
        data = request.get_json()
        split_type = data.get('type', 'purchase')  # 'purchase' or 'sale'
        
        txn, lines = models.get_transaction(txn_id)
        if not txn:
            return jsonify({'ok': False, 'error': 'Transaction not found'})
        
        if len(lines) != 2:
            return jsonify({'ok': False, 'error': f'Can only split simple 2-line transactions (this has {len(lines)} lines)'})
        
        lock = models.get_meta('lock_date', '')
        if lock and txn['date'] <= lock:
            return jsonify({'ok': False, 'error': f'Transaction is locked (before {lock})'})
        
        # Load configurable rate and accounts
        rate_num = int(models.get_meta('gst_rate_num', '5'))
        rate_den = int(models.get_meta('gst_rate_den', '105'))
        
        if split_type == 'purchase':
            gst_acct_name = models.get_meta('f8_tax_acct', 'GST.IN')
            post_acct_name = models.get_meta('f8_post_acct', '')
        else:
            gst_acct_name = models.get_meta('f9_tax_acct', 'GST.OUT')
            post_acct_name = models.get_meta('f9_post_acct', '')
        
        # Figure out which line is the bank (the account we're viewing) and which is the cross
        from_account = int(data.get('from_account', 0))
        bank_line = None
        cross_line = None
        for ln in lines:
            if ln['account_id'] == from_account:
                bank_line = ln
            else:
                cross_line = ln
        
        if not bank_line or not cross_line:
            bank_line = lines[0]
            cross_line = lines[1]
        
        # Calculate tax from gross: tax = gross * rate_num / rate_den
        gross = abs(cross_line['amount'])  # in cents
        tax = round(gross * rate_num / rate_den)
        net = gross - tax
        
        # Resolve accounts
        gst_acct = models.get_account_by_name(gst_acct_name)
        if not gst_acct:
            return jsonify({'ok': False, 'error': f'Tax account "{gst_acct_name}" not found. Set it up in Options.'})
        
        # If a default posting account is configured, re-assign the cross line
        cross_acct_id = cross_line['account_id']
        if post_acct_name:
            post_acct = models.get_account_by_name(post_acct_name)
            if post_acct:
                cross_acct_id = post_acct['id']
        
        # Preserve the sign of the cross line
        sign = 1 if cross_line['amount'] > 0 else -1
        
        new_lines = [
            (bank_line['account_id'], bank_line['amount'], bank_line['description']),
            (cross_acct_id, sign * net, cross_line['description']),
            (gst_acct['id'], sign * tax, f'GST: {cross_line["description"][:40]}'),
        ]
        
        # Verify balance
        total = sum(l[1] for l in new_lines)
        if total != 0:
            return jsonify({'ok': False, 'error': f'Split does not balance (off by {total/100:.2f})'})
        
        models.update_transaction(txn_id, txn['date'], txn['reference'], txn['description'],
                                  new_lines, allow_opening=_opening_confirmed())
        
        return jsonify({'ok': True, 'txn_id': txn_id})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# ─── Reconcile Toggle ──────────────────────────────────────────────

@app.route('/transaction/<int:txn_id>/reconcile', methods=['POST'])
def reconcile(txn_id):
    line_id = request.form.get('line_id', 0, type=int)
    if line_id:
        models.toggle_reconcile(line_id)
    return_to = request.form.get('return_to', url_for('home'))
    return redirect(return_to)

# ─── Reconciliation View ─────────────────────────────────────────

@app.route('/reconcile/<int:account_id>')
def reconcile_view(account_id):
    """Bank reconciliation screen."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    account = models.get_account(account_id)
    if not account:
        flash('Account not found', 'error')
        return redirect(url_for('home'))
    
    def _amt(name):
        s = request.args.get(name, '')
        try:
            return s, (models.parse_amount(s) if s.strip() else None)
        except ValueError:
            return s, None

    stmt_bal_str, stmt_bal = _amt('stmt')          # statement CLOSING balance
    stmt_open_str, stmt_open = _amt('stmt_open')   # statement OPENING balance
    stmt_date = request.args.get('stmt_date', '').strip()  # tag stamped into [Rec]

    entries = models.get_ledger(account_id)
    sign = 1 if account['normal_balance'] == 'D' else -1
    summary = models.get_reconcile_summary(account_id)
    company = models.get_meta('company_name', 'My Books')

    # Separate cleared vs outstanding
    outstanding = [e for e in entries if not e['reconciled'] and e['amount'] != 0]
    cleared = [e for e in entries if e['reconciled'] and e['amount'] != 0]

    # LAP continuity checks:
    # (1) statement OPENING must equal the cleared balance BEFORE this session —
    #     otherwise the previous reconciliation is incomplete or a statement is
    #     missing. (2) when done, cleared balance equals the statement CLOSING.
    diff_open = None
    if stmt_open is not None:
        diff_open = (stmt_open * sign) - summary['cleared_balance']
    diff = None
    if stmt_bal is not None:
        diff = (stmt_bal * sign) - summary['cleared_balance']

    from_report = request.args.get('from_report', '')
    parent_report = models.get_report(int(from_report)) if from_report else None

    return render_template('reconcile.html', account=account,
        outstanding=outstanding, cleared=cleared, summary=summary,
        stmt_bal=stmt_bal_str, diff=diff, company=company,
        stmt_open=stmt_open_str, diff_open=diff_open, stmt_date=stmt_date,
        parent_report=parent_report, from_report=from_report)

@app.route('/account/<int:account_id>/history')
def account_history_view(account_id):
    """LAP-style account History view: month × debit/credit/net/balance grid."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    account = models.get_account(account_id)
    if not account:
        flash('Account not found', 'error')
        return redirect(url_for('home'))
    hist = models.account_history(account_id)
    max_net = max((abs(h['net']) for h in hist), default=0) or 1
    company = models.get_meta('company_name', 'My Books')
    from_report = request.args.get('from_report', '')
    parent_report = models.get_report(int(from_report)) if from_report else None
    totals = {'dr': sum(h['dr'] for h in hist), 'cr': sum(h['cr'] for h in hist),
              'net': sum(h['net'] for h in hist)}
    return render_template('account_history.html', account=account, hist=hist,
        max_net=max_net, totals=totals, company=company,
        parent_report=parent_report, from_report=from_report)

def current_operator():
    """Per-INSTALL identity (machine-local config, not the books): the human at
    this machine. 'Cam Ware' → CW; another machine set to 'Brian W' stamps BW."""
    name = (load_config().get('operator_name') or '').strip()
    initials = ''.join(w[0] for w in name.split() if w)[:3].upper()
    return name, initials

@app.route('/library/set-user', methods=['POST'])
def set_library_user():
    name = request.form.get('operator_name', '').strip()
    cfg = load_config()
    cfg['operator_name'] = name
    save_config(cfg)
    flash(f'This Grid install now signs as: {name or "(nobody)"}', 'success')
    return redirect(url_for('library'))

# ─── Working Papers (the weave) ────────────────────────────────────

@app.route('/api/wp-folder', methods=['POST'])
def api_wp_folder_add():
    try:
        fid = models.add_wp_folder(request.form.get('name', ''),
                                   request.form.get('parent_id', '') or None)
        return jsonify({'ok': True, 'id': fid})
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/wp-folder/<int:folder_id>/rename', methods=['POST'])
def api_wp_folder_rename(folder_id):
    try:
        models.rename_wp_folder(folder_id, request.form.get('name', ''))
        return jsonify({'ok': True})
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/wp-folder/<int:folder_id>/delete', methods=['POST'])
def api_wp_folder_delete(folder_id):
    try:
        models.delete_wp_folder(folder_id)
        return jsonify({'ok': True})
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/workpaper/open-by-ref', methods=['POST'])
def api_workpaper_open_by_ref():
    """The statement red mark, made live: exact ref match to a working paper
    with a present file → open it in the local viewer. Typed text that matches
    nothing is just typed text (opened=False; the GUI falls back to editing)."""
    ref = request.form.get('ref', '').strip()
    wp = models.find_workpaper_by_ref(ref) if ref else None
    if not wp or not wp['path']:
        return jsonify({'ok': True, 'opened': False})
    try:
        p = models.wp_resolve(wp['path'])
    except ValueError:
        return jsonify({'ok': True, 'opened': False})
    if not os.path.exists(p):
        return jsonify({'ok': True, 'opened': False,
                        'error': f"{ref}: file missing ({wp['path']})"})
    import subprocess, platform
    try:
        sysname = platform.system()
        if sysname == 'Windows':
            os.startfile(p)          # type: ignore[attr-defined]
        elif sysname == 'Darwin':
            subprocess.Popen(['open', p])
        else:
            subprocess.Popen(['xdg-open', p],
                             stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return jsonify({'ok': True, 'opened': True})
    except Exception as e:
        return jsonify({'ok': True, 'opened': False, 'error': str(e)})

@app.route('/workpapers')
def workpapers_page():
    if not models.get_db_path():
        return redirect(url_for('library'))
    company = models.get_meta('company_name', 'My Books')
    fy = request.args.get('fy', '') or models.wp_fy()
    fys = models.workpaper_fys()
    if models.wp_fy() not in fys:
        fys = [models.wp_fy()] + fys
    op_name, op_init = current_operator()
    papers = models.list_workpapers(fy)
    folders = models.list_wp_folders()
    linked_now = {os.path.normpath(p['path']).lower() for p in papers if p['path']}
    browse = [{'path': f, 'linked': os.path.normpath(f).lower() in linked_now}
              for f in models.scan_client_docs(include_linked=True)]
    return render_template('workpapers.html', company=company,
        papers=papers, folders=folders, root_id=models.wp_root_id(),
        fy=fy, fys=fys, browse=browse,
        unindexed=models.scan_client_docs(), verify=models.verify_workpapers(fy),
        op_name=op_name, op_init=op_init, hl=request.args.get('hl', ''),
        client_folder=os.path.basename(models.client_dir() or ''))

@app.route('/api/workpaper', methods=['POST'])
def api_workpaper_add():
    try:
        wp_id = models.add_workpaper(request.form.get('ref', ''),
                                     request.form.get('description', ''),
                                     request.form.get('path', ''),
                                     request.form.get('fy', '') or None,
                                     request.form.get('folder_id', '') or None)
        return jsonify({'ok': True, 'id': wp_id})
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/workpaper/<int:wp_id>/update', methods=['POST'])
def api_workpaper_update(wp_id):
    try:
        models.update_workpaper(wp_id, request.form.get('field', ''),
                                request.form.get('value', ''))
        return jsonify({'ok': True})
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/workpaper/<int:wp_id>/sign', methods=['POST'])
def api_workpaper_sign(wp_id):
    """Stamp prep/review with THIS install's operator initials + today.
    Clicking a filled cell clears it (typo/redo)."""
    role = request.form.get('role', 'prep')
    field = 'prep_by' if role == 'prep' else 'rev_by'
    wp = models.get_workpaper(wp_id)
    if not wp:
        return jsonify({'ok': False, 'error': 'Working paper not found'})
    if wp[field]:
        models.update_workpaper(wp_id, field, '')
        return jsonify({'ok': True, 'stamp': ''})
    name, initials = current_operator()
    if not initials:
        return jsonify({'ok': False, 'error':
            'No user set for this Grid install — set your name on the Library screen first.'})
    stamp = f"{initials} {datetime.now().strftime('%Y-%m-%d')}"
    models.update_workpaper(wp_id, field, stamp)
    return jsonify({'ok': True, 'stamp': stamp})

@app.route('/api/workpaper/<int:wp_id>/delete', methods=['POST'])
def api_workpaper_delete(wp_id):
    models.delete_workpaper(wp_id)
    return jsonify({'ok': True})

@app.route('/workpaper/<int:wp_id>/file')
def workpaper_file(wp_id):
    """Serve the linked file for in-browser viewing (containment-checked)."""
    wp = models.get_workpaper(wp_id)
    if not wp or not wp['path']:
        flash('No file linked to this working paper', 'error')
        return redirect(url_for('workpapers_page'))
    try:
        p = models.wp_resolve(wp['path'])
    except ValueError as e:
        flash(str(e), 'error')
        return redirect(url_for('workpapers_page'))
    if not os.path.exists(p):
        flash(f"File missing: {wp['path']}", 'error')
        return redirect(url_for('workpapers_page'))
    return send_file(p)

@app.route('/api/workpaper/browse-dialog', methods=['POST'])
def api_workpaper_browse_dialog():
    """Browse — pop the OS file navigator (Explorer/Finder/GTK) on the machine
    running Grid and hand back what the operator picked, as a client-folder-
    relative link ready for the Link box."""
    try:
        chosen = models.pick_file_dialog()
    except Exception as e:
        # No chooser on this machine: say so and let the page fall back to
        # its own in-window list rather than dead-ending.
        return jsonify({'ok': False, 'error': str(e), 'fallback': True})
    if not chosen:
        return jsonify({'ok': True, 'cancelled': True})
    try:
        rel = models.wp_relativize(chosen)
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})
    return jsonify({'ok': True, 'path': rel})

@app.route('/api/workpaper/<int:wp_id>/open', methods=['POST'])
def api_workpaper_open(wp_id):
    """Open the linked file in the OS default app (local editor) — the
    CaseWare double-click, minus the absolute paths."""
    wp = models.get_workpaper(wp_id)
    if not wp or not wp['path']:
        return jsonify({'ok': False, 'error': 'No file linked'})
    try:
        p = models.wp_resolve(wp['path'])
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})
    if not os.path.exists(p):
        return jsonify({'ok': False, 'error': f"File missing: {wp['path']}"})
    import subprocess, platform
    try:
        sysname = platform.system()
        if sysname == 'Windows':
            os.startfile(p)          # type: ignore[attr-defined]
        elif sysname == 'Darwin':
            subprocess.Popen(['open', p])
        else:
            subprocess.Popen(['xdg-open', p],
                             stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Could not launch viewer: {e}'})

@app.route('/api/report-item/<int:item_id>/refmark', methods=['POST'])
def api_set_ref_mark(item_id):
    """Set/clear the working-paper index mark on a report line (the red pencil)."""
    try:
        mark = (request.get_json(silent=True) or {}).get('mark', '') if request.is_json \
               else request.form.get('mark', '')
        models.set_ref_mark(item_id, mark)
        return jsonify({'ok': True, 'mark': str(mark or '').strip()})
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/account-leadsheet/<int:account_id>', methods=['POST'])
def api_set_leadsheet(account_id):
    """Assign this account to a lead sheet code (blank clears)."""
    acct = models.get_account(account_id)
    if not acct:
        return jsonify({'ok': False, 'error': 'Account not found'})
    try:
        models.set_leadsheet(acct['name'], request.form.get('value', ''))
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})
    return jsonify({'ok': True, 'leadsheet': models.get_account(account_id)['leadsheet']})

@app.route('/leadsheets')
def leadsheets_page():
    if not models.get_db_path():
        return redirect(url_for('library'))
    company = models.get_meta('company_name', 'My Books')
    return render_template('leadsheets.html', company=company,
                           sheets=models.leadsheet_index(), sheet=None)

@app.route('/leadsheet/<code>')
def leadsheet_page(code):
    if not models.get_db_path():
        return redirect(url_for('library'))
    company = models.get_meta('company_name', 'My Books')
    return render_template('leadsheets.html', company=company,
                           sheets=models.leadsheet_index(),
                           sheet=models.leadsheet_data(code))

@app.route('/api/account-nextref/<int:account_id>', methods=['POST'])
def api_account_nextref(account_id):
    """Set the account's [Next Ref#] auto-numbering counter (0 = off)."""
    try:
        value = int(str(request.form.get('value', '0')).strip() or '0')
        models.set_account_next_ref(account_id, value)
        return jsonify({'ok': True, 'next_ref': max(0, value)})
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': 'Enter a whole number (0 turns numbering off)'})

@app.route('/api/reconcile-toggle/<int:line_id>', methods=['POST'])
def api_reconcile_toggle(line_id):
    """AJAX toggle for reconciliation checkboxes. Optional `tag` (statement
    date/ref) is stamped into the [Rec] field when closing."""
    tag = (request.get_json(silent=True) or {}).get('tag', '') if request.is_json \
          else request.form.get('tag', '')
    new_val = models.toggle_reconcile(line_id, tag)
    # Return updated summary for the account
    with models.get_db() as db:
        row = db.execute("SELECT account_id FROM lines WHERE id=?", (line_id,)).fetchone()
    summary = models.get_reconcile_summary(row['account_id'])
    return jsonify({'ok': True, 'reconciled': new_val,
                   'cleared_balance': summary['cleared_balance'],
                   'uncleared': summary['uncleared'],
                   'book_balance': summary['book_balance']})

@app.route('/api/doc-toggle/<int:line_id>', methods=['POST'])
def api_doc_toggle(line_id):
    """AJAX toggle for document-on-file flag."""
    new_val = models.toggle_doc_on_file(line_id)
    return jsonify({'ok': True, 'doc_on_file': new_val})

# ─── Trial Balance ──────────────────────────────────────────────────

@app.route('/trial-balance')
def trial_balance():
    if not models.get_db_path():
        return redirect(url_for('library'))
    as_of = request.args.get('as_of', '')
    accounts, total_dr, total_cr = models.get_trial_balance(as_of or None)
    company = models.get_meta('company_name', 'My Books')
    reports = models.get_reports()
    # Find accounts mapped to BS or IS (directly or via total-to chain)
    mapped_ids = _get_bs_is_mapped_ids()
    return render_template('trial_balance.html', accounts=accounts,
                         total_dr=total_dr, total_cr=total_cr,
                         as_of=as_of, company=company, reports=reports,
                         mapped_ids=mapped_ids)

# ─── Search ─────────────────────────────────────────────────────────

@app.route('/search')
def search():
    if not models.get_db_path():
        return redirect(url_for('library'))
    query = request.args.get('q', '')
    results = models.search_transactions(query) if query else []
    company = models.get_meta('company_name', 'My Books')
    return render_template('search.html', query=query, results=results,
                         company=company)

# ─── Account Management ────────────────────────────────────────────

@app.route('/account/new', methods=['GET', 'POST'])
@app.route('/account/new/in/<int:report_id>', methods=['GET', 'POST'])
def new_account(report_id=None):
    if not models.get_db_path():
        return redirect(url_for('library'))
    if request.method == 'POST':
        name = request.form['name'].strip().upper()
        desc = request.form.get('description', '')
        nb = request.form.get('normal_balance', 'D')
        acct_num = request.form.get('account_number', '')
        # Check for duplicate before attempting insert
        existing = models.get_account_by_name(name)
        if existing:
            flash(f'Account "{name}" already exists. Use a different name.', 'error')
        else:
            try:
                acct_id = models.add_account(name, nb, desc, account_number=acct_num)
                flash(f'Account {name} created', 'success')
                return redirect(url_for('home'))
            except Exception as e:
                flash(str(e), 'error')
    reports = models.get_reports()
    company = models.get_meta('company_name', 'My Books')
    return render_template('new_account.html', reports=reports, 
                         report_id=report_id, company=company)

# ─── Report Management ─────────────────────────────────────────────

@app.route('/report/new', methods=['GET', 'POST'])
def new_report():
    if not models.get_db_path():
        return redirect(url_for('library'))
    if request.method == 'POST':
        name = request.form['name'].strip()
        desc = request.form.get('description', '')
        try:
            rid = models.add_report(name, desc)
            # Auto-seed with blank lines so the report is never empty
            models.add_report_item(rid, 'label', desc.upper() if desc else name.upper(), position=10, indent=0)
            models.add_report_item(rid, 'label', '', position=20)
            models.add_report_item(rid, 'separator', position=30, sep_style='double')
            flash(f'Report "{name}" created', 'success')
            return redirect(url_for('report_view', report_id=rid))
        except Exception as e:
            flash(str(e), 'error')
    company = models.get_meta('company_name', 'My Books')
    return render_template('new_report.html', company=company)

# ─── Settings ───────────────────────────────────────────────────────

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if not models.get_db_path():
        return redirect(url_for('library'))
    if request.method == 'POST':
        # Global Options + System Fiscal Year go through the guarded model
        # writer as ONE set — it refuses rather than half-writing a fiscal year.
        try:
            models.set_fiscal_settings(
                company_name=request.form.get('company_name', ''),
                working_ye=request.form.get('working_ye', ''),
                ceiling_mode=request.form.get('fy_ceiling_mode') or None,
                lock_date=request.form.get('lock_date', ''))
        except ValueError as e:
            flash(str(e), 'error')
            return redirect(url_for('settings'))
        # F8/F9 GST split settings
        models.set_meta('gst_rate_num', request.form.get('gst_rate_num', '5'))
        models.set_meta('gst_rate_den', request.form.get('gst_rate_den', '105'))
        models.set_meta('f8_tax_acct', request.form.get('f8_tax_acct', 'GST.IN'))
        models.set_meta('f8_post_acct', request.form.get('f8_post_acct', ''))
        models.set_meta('f9_tax_acct', request.form.get('f9_tax_acct', 'GST.OUT'))
        models.set_meta('f9_post_acct', request.form.get('f9_post_acct', ''))
        # Stripe procedure settings
        models.set_meta('stripe_fee_acct', request.form.get('stripe_fee_acct', 'EX.CC'))
        models.set_meta('stripe_ar_report', request.form.get('stripe_ar_report', ''))
        flash('Settings saved', 'success')
        return redirect(url_for('settings'))
    company = models.get_meta('company_name', 'My Books')
    fye = models.get_meta('fiscal_year_end', '12-31')
    fy_year = models.get_meta('fiscal_year', '')
    lock_date = models.get_meta('lock_date', '')
    db_path = models.get_db_path()
    # Fiscal dates: the year-end being worked on, and how far past it posting stays open.
    anchor = models.fiscal_anchor()
    ceiling_mode = models.get_meta('fy_ceiling_mode', 'cy') or 'cy'
    working_ye = anchor['cy_end'] if anchor else ''
    cy_ceiling = anchor['cy_end'] if anchor else ''
    next_ceiling = anchor['next_end'] if anchor else ''
    latest_txn = models.latest_transaction_date()
    # 'Same as year-end' is unavailable while postings sit past it — say so up front
    # rather than only on refusal.
    cy_blocked = bool(anchor and latest_txn and latest_txn > cy_ceiling)
    cy_blocked_n = models.transactions_after(cy_ceiling) if cy_blocked else 0
    # F8/F9 settings
    gst_rate_num = models.get_meta('gst_rate_num', '5')
    gst_rate_den = models.get_meta('gst_rate_den', '105')
    f8_tax_acct = models.get_meta('f8_tax_acct', 'GST.IN')
    f8_post_acct = models.get_meta('f8_post_acct', '')
    f9_tax_acct = models.get_meta('f9_tax_acct', 'GST.OUT')
    f9_post_acct = models.get_meta('f9_post_acct', '')
    # Stripe settings
    stripe_fee_acct = models.get_meta('stripe_fee_acct', 'EX.CC')
    stripe_ar_report = models.get_meta('stripe_ar_report', '')
    accounts = models.get_accounts()
    reports = models.get_reports()
    fy_ceiling = models.fiscal_ceiling()
    return render_template('settings.html', company=company, fye=fye, fy_year=fy_year,
                         lock_date=lock_date, db_path=db_path,
                         gst_rate_num=gst_rate_num, gst_rate_den=gst_rate_den,
                         f8_tax_acct=f8_tax_acct, f8_post_acct=f8_post_acct,
                         f9_tax_acct=f9_tax_acct, f9_post_acct=f9_post_acct,
                         stripe_fee_acct=stripe_fee_acct, stripe_ar_report=stripe_ar_report,
                         accounts=accounts, reports=reports, fy_ceiling=fy_ceiling,
                         working_ye=working_ye, ceiling_mode=ceiling_mode,
                         cy_ceiling=cy_ceiling, next_ceiling=next_ceiling,
                         latest_txn=latest_txn, cy_blocked=cy_blocked,
                         cy_blocked_n=cy_blocked_n, anchor=anchor,
                         backups=models.list_backups(), backup_keep=models.BACKUP_KEEP)

@app.route('/backup-now', methods=['POST'])
def backup_now():
    """Manual snapshot (the automatic one runs daily at open)."""
    try:
        path = models.backup_books(force=True)
        flash(f'Backed up: backups/{os.path.basename(path)}', 'success')
    except Exception as e:
        flash(f'Backup failed: {e}', 'error')
    return redirect(url_for('settings'))

# ─── CSV Import ──────────────────────────────────────────────────

def _fix_csv_rows(all_rows):
    """Fix rows where unquoted commas in fields cause extra columns.
    Strategy: header defines N columns. For rows with >N cols, we know the
    rightmost columns are numeric (amounts) and the excess splits happened
    in text fields. We anchor from the right (amounts) and merge the middle."""
    if not all_rows:
        return all_rows
    ncols = len(all_rows[0])
    fixed = [all_rows[0]]
    for row in all_rows[1:]:
        if len(row) <= ncols:
            fixed.append(row)
            continue
        excess = len(row) - ncols
        # Take the first few columns as-is (account type, number, date, ref)
        # These are never the ones that split. Then merge the description fields.
        # Take the last columns as-is (amounts, empty trailing fields).
        # Strategy: keep first 4 cols, keep last (ncols-5) cols, merge everything in between.
        # For an 8-col header: keep [0:4], merge [4:4+1+excess], keep [-3:]
        # This merges the Description fields (cols 4 and 5 in the original)
        left_keep = 4  # Account Type, Account Number, Date, Cheque#
        right_keep = ncols - left_keep - 2  # Amount cols + trailing (usually 2-3 cols)
        # But we need at least the desc cols (normally 2) to end up as ncols - left_keep - right_keep
        middle_start = left_keep
        middle_end = len(row) - right_keep
        middle = row[middle_start:middle_end]
        # middle should be exactly 2 fields (desc1, desc2) but has 2+excess
        # Merge all middle fields into exactly 2: first field stays, rest merge
        if len(middle) >= 2:
            desc1 = middle[0]
            desc2 = ', '.join(middle[1:])
            new_row = row[:left_keep] + [desc1, desc2] + row[middle_end:]
        else:
            new_row = row[:ncols]
        fixed.append(new_row[:ncols])
    return fixed

def _overlay_mapping(det, m):
    """Overlay a saved profile or the operator's just-submitted tagging onto the
    detection defaults — the human's tagging always outranks the sniffer."""
    ncols = det['num_cols']
    def _col(key, fallback):
        try:
            v = int(m.get(key, fallback))
        except (TypeError, ValueError):
            return fallback
        return v if -1 <= v < ncols else fallback
    det['best_date_col'] = _col('col_date', det['best_date_col'])
    det['best_desc_col'] = _col('col_desc', det['best_desc_col'])
    det['best_amount_col'] = _col('col_amount', det['best_amount_col'])
    det['sel_ref'] = _col('col_ref', -1)
    det['sel_desc2'] = _col('col_desc2', -1)
    det['sel_debit'] = _col('col_debit', -1)
    det['sel_credit'] = _col('col_credit', -1)
    det['has_header'] = bool(m.get('skip_header', det['has_header']))
    det['flip_sign_checked'] = bool(m.get('flip_sign', False))
    if m.get('date_format') in ('MDY', 'DMY', 'YMD', 'auto'):
        det['date_fmt'] = m['date_format']
    return det

def _file_signature(all_rows, has_header):
    """Stable fingerprint of a file SHAPE (not its data) — the key for
    remembering how the operator tagged this kind of file. Headered files key on
    their header names; headerless ones on their column count."""
    import hashlib as _hl
    if has_header and all_rows:
        key = '|'.join(str(c).strip().lower() for c in all_rows[0])
    else:
        key = f'headerless-{len(all_rows[0]) if all_rows else 0}cols'
    return _hl.md5(key.encode()).hexdigest()[:10]

def _detect_mapping(all_rows):
    """Auto-detect date/amount/description columns, the date format, and whether
    row 0 is a header. Returns the template vars for the column-mapping step.

    Every guess is a pre-selection the user can override — nothing here posts on
    its own. The point is that leaving a required column unmapped is what makes
    an import come back empty, so the file should arrive already mapped."""
    headers = all_rows[0]
    ncols = len(headers)
    sample_rows = all_rows[1:min(51, len(all_rows))]

    # ── Date column: most values that parse as a date ──
    best_date_col = 0
    best_date_score = 0
    for ci in range(ncols):
        samples = [str(row[ci]).strip() for row in sample_rows if ci < len(row) and str(row[ci]).strip()]
        score = sum(1 for s in samples if _parse_date(s) is not None)
        if score > best_date_score:
            best_date_score = score
            best_date_col = ci

    # ── Header row? Row 0 is data if it parses as a date in the date column ──
    row0 = str(headers[best_date_col]).strip() if best_date_col < ncols else ''
    has_header = _parse_date(row0) is None
    # When row 0 IS data it must be scored too, and the column labels are just
    # positions — the template shows "Col n" with no name.
    if not has_header:
        sample_rows = all_rows[:min(51, len(all_rows))]

    date_samples = [str(row[best_date_col]).strip() for row in sample_rows
                   if best_date_col < len(row) and str(row[best_date_col]).strip()]
    date_fmt, confidence, detail = _detect_date_format(date_samples)

    # ── Amount column ──
    # A column qualifies if most of its non-blank values are numeric. Constant
    # columns (account/card numbers) are excluded — an amount column varies.
    # Preference: header name, then a column carrying negatives (a running
    # balance rarely does), then the right-most candidate.
    AMT_WORDS = ('amount', 'amt', 'debit', 'credit', 'withdraw', 'deposit', 'value', 'total')
    BAL_WORDS = ('balance', 'bal')
    amount_candidates = []
    for ci in range(ncols):
        if ci == best_date_col:
            continue
        vals = [str(row[ci]).strip() for row in sample_rows if ci < len(row) and str(row[ci]).strip()]
        if not vals:
            continue
        nums = []
        for v in vals:
            try:
                nums.append(float(v.replace(',', '').replace('$', '')))
            except ValueError:
                pass
        if len(nums) < len(vals) * 0.8 or len(set(nums)) < 2:
            continue
        name = str(headers[ci]).strip().lower() if has_header else ''
        amount_candidates.append({
            'col': ci,
            'named': any(w in name for w in AMT_WORDS),
            'is_balance': any(w in name for w in BAL_WORDS),
            'has_neg': any(n < 0 for n in nums),
            'has_cents': any(abs(n) % 1 for n in nums),
        })
    best_amount_col = -1
    if amount_candidates:
        best_amount_col = max(amount_candidates, key=lambda c: (
            c['named'] and not c['is_balance'],
            not c['is_balance'],
            c['has_neg'],
            c['has_cents'],
            -c['col'],
        ))['col']

    # ── Description column: the wordiest non-numeric column ──
    DESC_WORDS = ('desc', 'memo', 'narrative', 'payee', 'detail', 'particular', 'reference')
    best_desc_col, best_desc_score = -1, 0.0
    for ci in range(ncols):
        if ci in (best_date_col, best_amount_col):
            continue
        vals = [str(row[ci]).strip() for row in sample_rows if ci < len(row) and str(row[ci]).strip()]
        if not vals:
            continue
        wordy = [v for v in vals if any(ch.isalpha() for ch in v)]
        if len(wordy) < len(vals) * 0.5:
            continue
        score = sum(len(v) for v in wordy) / len(vals)
        name = str(headers[ci]).strip().lower() if has_header else ''
        if any(w in name for w in DESC_WORDS):
            score += 1000
        if score > best_desc_score:
            best_desc_score, best_desc_col = score, ci

    # ── Date-format comparison table ──
    date_comparisons = []
    has_ambiguous = False  # only show MDY/DMY columns if there are ambiguous numeric dates
    for raw in date_samples[:10]:
        parsed_mdy = _parse_date(raw, 'MDY')
        parsed_dmy = _parse_date(raw, 'DMY')
        if parsed_mdy and parsed_dmy and parsed_mdy != parsed_dmy:
            has_ambiguous = True
        parsed = _parse_date(raw, date_fmt)
        date_comparisons.append({
            'raw': raw, 'parsed': parsed,
            'MDY': parsed_mdy, 'DMY': parsed_dmy,
            'ok': parsed is not None,
        })

    all_parse_ok = sum(1 for raw in date_samples if _parse_date(raw, date_fmt) is not None)
    all_parse_fail = len(date_samples) - all_parse_ok

    data_rows = all_rows[1:] if has_header else all_rows
    return dict(
        headers=headers, preview_rows=data_rows[:5], num_cols=ncols,
        date_fmt=date_fmt, date_confidence=confidence, date_detail=detail,
        date_comparisons=date_comparisons, best_date_col=best_date_col,
        best_amount_col=best_amount_col, best_desc_col=best_desc_col,
        has_header=has_header, total_rows=len(data_rows),
        has_ambiguous=has_ambiguous,
        all_parse_ok=all_parse_ok, all_parse_fail=all_parse_fail)

@app.route('/import', methods=['GET', 'POST'])
def csv_import():
    if not models.get_db_path():
        return redirect(url_for('library'))
    company = models.get_meta('company_name', 'My Books')
    accounts = models.get_accounts()
    import json as _json, tempfile, hashlib
    
    # Cache helpers — store import data in temp files to avoid cookie size limits
    cache_dir = os.path.join(tempfile.gettempdir(), 'grid_import')
    os.makedirs(cache_dir, exist_ok=True)
    
    def _cache_key():
        """Get or create a per-session cache key."""
        k = session.get('import_cache_key')
        if not k:
            k = hashlib.md5(os.urandom(16)).hexdigest()[:12]
            session['import_cache_key'] = k
        return k
    
    def _cache_put(name, data):
        path = os.path.join(cache_dir, f'{_cache_key()}_{name}.json')
        with open(path, 'w') as f:
            _json.dump(data, f)
    
    def _cache_get(name):
        path = os.path.join(cache_dir, f'{_cache_key()}_{name}.json')
        if os.path.exists(path):
            with open(path) as f:
                return _json.load(f)
        return None
    
    def _cache_clear():
        k = session.get('import_cache_key', '')
        if k:
            for fn in os.listdir(cache_dir):
                if fn.startswith(k):
                    os.remove(os.path.join(cache_dir, fn))
            session.pop('import_cache_key', None)
    
    step = request.form.get('_step', '')
    
    # ═══ STEP 3: CONFIRM IMPORT ═══
    if step == 'do_import':
        cached = _cache_get('parsed')
        if not cached:
            flash('Session expired — please re-upload the file', 'error')
            return redirect(url_for('csv_import'))
        
        account_id = int(session.get('import_account_id', 0))
        acct = models.get_account(account_id)
        if not acct:
            flash('Account not found', 'error')
            return redirect(url_for('csv_import'))
        
        susp = models.get_account_by_name('EX.SUSP')
        susp_id = susp['id'] if susp else models.add_account('EX.SUSP', 'D', 'Suspense')
        nb_flip = -1 if acct['normal_balance'] == 'C' else 1

        # Build every transaction FIRST, then post the whole file as ONE atomic
        # batch: it lands complete or not at all, and lands tagged so one click
        # deletes it. No more half-imported files, no re-run duplicates.
        try:
            txns = []
            for item in cached:
                d, ref, desc, amount = item['date'], item['ref'], item['desc'], item['amount']
                post_amount = amount * nb_flip
                rule_acct, rule_tax, tax_split = models.apply_rules(desc, amount)
                cross = models.get_account_by_name(rule_acct)
                cross_id = cross['id'] if cross else susp_id

                if tax_split:
                    tax_acct = models.get_account_by_name(tax_split['tax_acct'])
                    tax_acct_id = tax_acct['id'] if tax_acct else susp_id
                    net, tax = tax_split['net'], tax_split['tax']
                    if post_amount > 0:
                        lines = [(account_id, post_amount, desc),
                                 (cross_id, -net, desc),
                                 (tax_acct_id, -tax, f'Tax: {desc[:40]}')]
                    else:
                        lines = [(account_id, post_amount, desc),
                                 (cross_id, net, desc),
                                 (tax_acct_id, tax, f'Tax: {desc[:40]}')]
                elif post_amount > 0:
                    lines = [(account_id, post_amount, desc), (cross_id, -post_amount, desc)]
                else:
                    lines = [(cross_id, -post_amount, desc), (account_id, post_amount, desc)]
                txns.append((d, ref, desc, lines))

            batch_id, ids = models.post_import_batch(txns)

            # The tagging that just WORKED becomes the remembered profile for
            # this file shape — next upload arrives pre-tagged.
            sig = _cache_get('sig')
            mapping = _cache_get('mapping')
            if sig and mapping:
                models.set_meta('import_profile_' + sig, _json.dumps(mapping))

            _cache_clear()
            flash(f'Imported {len(ids)} transactions into {acct["name"]} — '
                  f'batch {batch_id} (undo on the Import page)', 'success')
            return redirect(url_for('account_ledger', account_id=account_id))
        except Exception as e:
            flash(f'Import failed — {e}. Nothing was imported.', 'error')
            return redirect(url_for('csv_import'))
    
    # ═══ STEP 2b: REFRESH PREVIEW (re-apply rules after adding new ones) ═══
    if step == 'refresh_preview':
        cached = _cache_get('parsed')
        if not cached:
            flash('Session expired — please re-upload the file', 'error')
            return redirect(url_for('csv_import'))
        
        account_id = int(session.get('import_account_id', 0))
        acct = models.get_account(account_id)
        if not acct:
            flash('Account not found', 'error')
            return redirect(url_for('csv_import'))
        
        # Re-apply rules to the already-parsed data
        preview = []
        for item in cached:
            rule_acct, rule_tax, tax_split = models.apply_rules(item['desc'], item['amount'])
            preview.append({**item,
                'rule_acct': rule_acct, 'rule_tax': rule_tax, 'tax_split': tax_split})
        
        matched = sum(1 for p in preview if p['rule_acct'] != 'EX.SUSP')
        return render_template('import.html', company=company, accounts=accounts,
            rules_preview=preview, account_id=account_id,
            matched_count=matched, acct_name=acct['name'], tax_codes=models.get_tax_codes())
    
    # ═══ STEP 2: PARSE WITH DATE FORMAT + SHOW RULES PREVIEW ═══
    if step == 'preview':
        cached_rows = _cache_get('rows')
        if not cached_rows:
            flash('Session expired — please re-upload the file', 'error')
            return redirect(url_for('csv_import'))
        
        account_id = int(request.form.get('account_id', session.get('import_account_id', 0)))
        session['import_account_id'] = account_id
        acct = models.get_account(account_id)
        if not acct:
            flash('Account not found', 'error')
            return redirect(url_for('csv_import'))
        
        date_fmt = request.form.get('date_format', 'MDY')
        col_date = int(request.form.get('col_date', 0))
        col_ref = int(request.form.get('col_ref', -1))
        col_desc = int(request.form.get('col_desc', 0))
        col_desc2 = int(request.form.get('col_desc2', -1))
        col_amount = int(request.form.get('col_amount', -1))
        col_debit = int(request.form.get('col_debit', -1))
        col_credit = int(request.form.get('col_credit', -1))
        skip_header = request.form.get('skip_header', '0') == '1'
        flip_sign = request.form.get('flip_sign', '0') == '1'

        # Remember the operator's tagging — becomes the stored profile for this
        # file shape once the import actually lands (saved in do_import), and is
        # re-shown as-submitted if this preview has to bounce back.
        mapping = {
            'col_date': col_date, 'col_ref': col_ref, 'col_desc': col_desc,
            'col_desc2': col_desc2, 'col_amount': col_amount,
            'col_debit': col_debit, 'col_credit': col_credit,
            'skip_header': skip_header, 'flip_sign': flip_sign,
            'date_format': date_fmt}
        _cache_put('mapping', mapping)

        def _bounce(errs):
            """Re-render the mapping step with the operator's own tagging and the
            reasons nothing parsed — never a dead-end redirect."""
            det = _overlay_mapping(_detect_mapping(cached_rows), mapping)
            if det['best_amount_col'] < 0 and det['sel_debit'] < 0 and det['sel_credit'] < 0:
                det['best_amount_col'] = _detect_mapping(cached_rows)['best_amount_col']
            return render_template('import.html', company=company, accounts=accounts,
                account_id=account_id, errors=errs, **det)

        # An unmapped amount makes every row skip silently — catch it here and
        # send the file back to the mapping step rather than reporting an empty
        # result the user can't explain.
        if col_amount < 0 and col_debit < 0 and col_credit < 0:
            flash('Map an Amount column (or a Debit/Credit pair) — no amount column was selected.', 'error')
            return _bounce(['No amount column selected, so every row was skipped.'])

        data_rows = cached_rows[1:] if skip_header else cached_rows
        lock = models.get_meta('lock_date', '')
        errors = []
        parsed = []
        skipped_blank_amt = 0
        skipped_zero_amt = 0

        for i, row in enumerate(data_rows):
            line_num = i + (2 if skip_header else 1)
            try:
                if not row or all(str(c).strip() == '' for c in row):
                    continue
                def col(idx, _row=row):
                    if idx < 0 or idx >= len(_row): return ''
                    return str(_row[idx]).strip()
                
                d = _parse_date(col(col_date), fmt=date_fmt)
                if not d:
                    errors.append(f"Row {line_num}: Invalid date '{col(col_date)}'")
                    continue
                
                # Validate year range
                try:
                    yr = int(d[:4])
                    if yr < 1950 or yr > 2099:
                        errors.append(f"Row {line_num}: Year {yr} out of range (1950-2099)")
                        continue
                except:
                    pass
                
                if lock and d <= lock:
                    errors.append(f"Row {line_num}: Date {d} is on or before lock date ({lock})")
                    continue
                
                ref = col(col_ref)
                desc = col(col_desc)
                if col_desc2 >= 0:
                    d2 = col(col_desc2)
                    if d2: desc = f"{desc} — {d2}" if desc else d2
                
                if col_amount >= 0:
                    amt_str = col(col_amount).replace(',', '').replace('$', '').replace('"', '')
                    if not amt_str:
                        skipped_blank_amt += 1
                        continue
                    amount = models.parse_amount(amt_str)
                else:
                    dr_str = col(col_debit).replace(',', '').replace('$', '').replace('"', '')
                    cr_str = col(col_credit).replace(',', '').replace('$', '').replace('"', '')
                    dr_val = models.parse_amount(dr_str) if dr_str else 0
                    cr_val = models.parse_amount(cr_str) if cr_str else 0
                    amount = dr_val - cr_val
                
                if amount == 0:
                    skipped_zero_amt += 1
                    continue
                if flip_sign: amount = -amount
                parsed.append({'date': d, 'ref': ref, 'desc': desc, 'amount': amount})
            except Exception as e:
                errors.append(f"Row {line_num}: {e}")

        # Rows dropped by the amount tests never raised an error, so say so
        # explicitly — otherwise a mis-mapped column reads as an empty file.
        if skipped_blank_amt:
            src = f"column {col_amount}" if col_amount >= 0 else 'the debit/credit columns'
            errors.append(f"{skipped_blank_amt} row(s) skipped: {src} was blank — "
                          f"check that the Amount column is mapped to the right column.")
        if skipped_zero_amt:
            errors.append(f"{skipped_zero_amt} row(s) skipped: amount was zero.")

        if not parsed:
            flash(f"{len(errors)} error(s) found, no valid transactions.", 'error')
            return _bounce(errors)

        # Cache parsed data for the confirm step
        _cache_put('parsed', parsed)
        
        # Build rules preview
        preview = []
        for item in parsed:
            rule_acct, rule_tax, tax_split = models.apply_rules(item['desc'], item['amount'])
            preview.append({**item,
                'rule_acct': rule_acct, 'rule_tax': rule_tax, 'tax_split': tax_split})
        
        matched = sum(1 for p in preview if p['rule_acct'] != 'EX.SUSP')
        return render_template('import.html', company=company, accounts=accounts,
            rules_preview=preview, account_id=account_id,
            matched_count=matched, errors=errors if errors else None,
            acct_name=acct['name'], tax_codes=models.get_tax_codes())
    
    # ═══ STEP 1: UPLOAD FILE → SHOW COLUMN MAPPING + DATE DETECTION ═══
    if request.method == 'POST':
        account_id = int(request.form.get('account_id', 0))
        session['import_account_id'] = account_id
        
        file = request.files.get('csv_file')
        if not file or not file.filename:
            flash('No file selected', 'error')
            return redirect(url_for('csv_import'))
        
        fname = file.filename.lower()
        if not any(fname.endswith(ext) for ext in ('.csv', '.xlsx', '.xls', '.txt', '.tsv')):
            flash('Unsupported file type. Use .csv, .xlsx, or .xls', 'error')
            return redirect(url_for('csv_import'))
        
        try:
            all_rows = _read_upload_to_rows(file)
            all_rows = _fix_csv_rows(all_rows)
        except Exception as e:
            flash(f'Error reading file: {e}', 'error')
            return redirect(url_for('csv_import'))
        
        if len(all_rows) < 2:
            flash('File has no data rows', 'error')
            return redirect(url_for('csv_import'))
        
        # Cache rows in temp file
        _cache_put('rows', all_rows)

        det = _detect_mapping(all_rows)

        # Import profile: the operator's tagging from the LAST import of this
        # file shape wins over fresh guesses — teach once, remember after,
        # exactly like the rules engine.
        sig = _file_signature(all_rows, det['has_header'])
        _cache_put('sig', sig)
        prof_raw = models.get_meta('import_profile_' + sig, '')
        if prof_raw:
            try:
                det = _overlay_mapping(det, _json.loads(prof_raw))
                det['profile_used'] = True
            except Exception:
                pass

        return render_template('import.html', company=company, accounts=accounts,
            account_id=account_id, errors=None, **det)

    # ═══ INITIAL: Upload form ═══
    _cache_clear()
    return render_template('import.html', company=company, accounts=accounts,
        batches=models.list_import_batches())

@app.route('/import/delete-batch', methods=['POST'])
def delete_import_batch_route():
    if not models.get_db_path():
        return redirect(url_for('library'))
    batch_id = request.form.get('batch_id', '')
    try:
        n = models.delete_import_batch(batch_id)
        flash(f'Deleted import batch {batch_id} — {n} transaction(s) removed.', 'success')
    except ValueError as e:
        flash(f'Could not delete batch: {e}', 'error')
    return redirect(url_for('csv_import'))

def _parse_date(s, fmt='auto'):
    """Parse a date string into yyyy-mm-dd format.
    fmt: 'auto' (guess), 'MDY', 'DMY', 'YMD'.
    Handles: ISO, slashes, dashes, dots, month names (short/long), spaces, 2/4-digit years."""
    import re
    s = str(s).strip().strip('"').strip("'")
    if not s or s.lower() in ('', 'none', 'nat', 'null'):
        return None

    # Drop a trailing time component — xlsx/xls date cells stringify as
    # "2026-07-28 00:00:00" (and some exports use the ISO "T" separator).
    m = re.match(r'^(.*?)[ T]\d{1,2}:\d{2}(?::\d{2})?(?:\.\d+)?\s*(?:[AaPp]\.?[Mm]\.?)?$', s)
    if m and m.group(1).strip():
        s = m.group(1).strip()

    months = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
              'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12,
              'january':1,'february':2,'march':3,'april':4,'june':6,
              'july':7,'august':8,'september':9,'october':10,'november':11,'december':12}
    
    def _fix_year(y):
        if y < 100:
            return y + 2000 if y < 50 else y + 1900
        return y
    
    def _valid(y, m, d):
        """Check if date components are valid."""
        if m < 1 or m > 12 or d < 1 or d > 31:
            return None
        try:
            from datetime import datetime
            datetime(y, m, d)
            return f"{y:04d}-{m:02d}-{d:02d}"
        except ValueError:
            return None
    
    # ── Unambiguous formats (always try first regardless of fmt) ──
    
    # ISO: yyyy-mm-dd or yyyy/mm/dd or yyyy.mm.dd
    m = re.match(r'^(\d{4})[\-/.](\d{1,2})[\-/.](\d{1,2})$', s)
    if m:
        return _valid(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    
    # yyyymmdd (compact ISO)
    m = re.match(r'^(\d{4})(\d{2})(\d{2})$', s)
    if m:
        return _valid(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    
    # ── Month name formats (unambiguous — month is spelled out) ──
    
    # dd Mon yyyy or dd-Mon-yyyy or dd/Mon/yyyy or dd.Mon.yyyy (e.g. "31 May 2025", "15-Jan-2025")
    m = re.match(r'^(\d{1,2})[\s\-/.](\w+)[\s\-/.,]+(\d{2,4})$', s, re.I)
    if m and m.group(2).lower() in months:
        return _valid(_fix_year(int(m.group(3))), months[m.group(2).lower()], int(m.group(1)))
    
    # Mon dd, yyyy or Mon dd yyyy (e.g. "May 31, 2025", "Jan 15 2025")
    m = re.match(r'^(\w+)[\s\-/.](\d{1,2})[,\s]+(\d{2,4})$', s, re.I)
    if m and m.group(1).lower() in months:
        return _valid(_fix_year(int(m.group(3))), months[m.group(1).lower()], int(m.group(2)))
    
    # yyyy Mon dd or yyyy-Mon-dd (e.g. "2025 May 31")
    m = re.match(r'^(\d{4})[\s\-/.](\w+)[\s\-/.](\d{1,2})$', s, re.I)
    if m and m.group(2).lower() in months:
        return _valid(int(m.group(1)), months[m.group(2).lower()], int(m.group(3)))
    
    # ── Numeric formats (ambiguous — depends on fmt) ──
    
    # nn/nn/nnnn or nn-nn-nnnn or nn.nn.nnnn
    m = re.match(r'^(\d{1,2})[\-/.](\d{1,2})[\-/.](\d{4})$', s)
    if m:
        a, b, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if fmt == 'DMY':
            return _valid(yr, b, a)
        else:  # MDY or auto
            return _valid(yr, a, b)
    
    # nn/nn/nn (2-digit year)
    m = re.match(r'^(\d{1,2})[\-/.](\d{1,2})[\-/.](\d{2})$', s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        yr = _fix_year(int(m.group(3)))
        if fmt == 'DMY':
            return _valid(yr, b, a)
        else:
            return _valid(yr, a, b)
    
    return None

def _detect_date_format(date_strings):
    """Auto-detect date format from sample strings. Returns (fmt, confidence, detail).
    fmt is only meaningful for ambiguous numeric dates (MDY vs DMY).
    Month-name and ISO dates are always unambiguous."""
    import re
    if not date_strings:
        return 'MDY', 'low', 'no dates found'
    
    # Count how many parse successfully with auto
    auto_ok = sum(1 for d in date_strings if _parse_date(d, 'auto') is not None)
    
    # YMD: year comes first
    ymd_count = sum(1 for d in date_strings if re.match(r'^\d{4}[/\-.]', str(d)))
    if ymd_count > len(date_strings) * 0.8:
        return 'YMD', 'high', f'{ymd_count}/{len(date_strings)} start with 4-digit year'
    
    # Month names are unambiguous
    month_count = sum(1 for d in date_strings if re.search(r'[A-Za-z]{3,}', str(d)))
    if month_count > len(date_strings) * 0.8:
        return 'auto', 'high', f'month names detected — dates are unambiguous ({auto_ok}/{len(date_strings)} parsed OK)'
    
    # MDY vs DMY: check if first or second component > 12
    firsts, seconds = [], []
    for d in date_strings:
        m = re.match(r'^(\d{1,2})[/\-.](\d{1,2})[/\-.]', str(d))
        if m:
            firsts.append(int(m.group(1)))
            seconds.append(int(m.group(2)))
    
    if not firsts:
        return 'MDY', 'low', 'could not parse date components'
    
    first_over_12 = any(v > 12 for v in firsts)
    second_over_12 = any(v > 12 for v in seconds)
    
    if first_over_12 and not second_over_12:
        return 'DMY', 'high', f'first component goes up to {max(firsts)} (must be day)'
    if second_over_12 and not first_over_12:
        return 'MDY', 'high', f'second component goes up to {max(seconds)} (must be day)'
    
    return 'MDY', 'ambiguous', 'all values ≤ 12 — could be MDY or DMY. Defaulting to MDY (North American).'

def _read_upload_to_rows(fileobj):
    """Read uploaded file (csv, xlsx, xls) into list of lists of strings."""
    import csv, io
    fname = fileobj.filename.lower()
    
    if fname.endswith('.xlsx'):
        import openpyxl, datetime as _dt
        wb = openpyxl.load_workbook(fileobj, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = []
            for v in row:
                if v is None:
                    cells.append('')
                elif isinstance(v, _dt.datetime):
                    # Date cells come back as datetime objects — emit the ISO date
                    # directly instead of str()'s "2026-07-28 00:00:00", which no
                    # date parser should ever have to see.
                    cells.append(v.strftime('%Y-%m-%d') if (v.hour, v.minute, v.second) == (0, 0, 0)
                                 else v.strftime('%Y-%m-%d %H:%M:%S'))
                elif isinstance(v, _dt.date):
                    cells.append(v.strftime('%Y-%m-%d'))
                elif isinstance(v, float) and v == int(v):
                    # Keep numbers clean (no trailing .0), matching the .xls branch
                    cells.append(str(int(v)))
                else:
                    cells.append(str(v))
            rows.append(cells)
        return rows
    elif fname.endswith('.xls'):
        # Old Excel 97-2003 binary format — needs xlrd
        try:
            import xlrd
        except ImportError:
            raise ImportError('xlrd is required for .xls files. Run: pip install xlrd')
        raw = fileobj.read()
        wb = xlrd.open_workbook(file_contents=raw)
        ws = wb.sheet_by_index(0)
        rows = []
        for rx in range(ws.nrows):
            row = []
            for cx in range(ws.ncols):
                cell = ws.cell(rx, cx)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    # Convert Excel date to string
                    try:
                        dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        row.append(dt.strftime('%Y-%m-%d'))
                    except:
                        row.append(str(cell.value))
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    # Keep numbers clean (no trailing .0 for integers)
                    v = cell.value
                    row.append(str(int(v)) if v == int(v) else str(v))
                else:
                    row.append(str(cell.value) if cell.value is not None else '')
            rows.append(row)
        return rows
    else:
        # CSV — handle various encodings
        raw = fileobj.read()
        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                content = raw.decode(enc)
                break
            except (UnicodeDecodeError, AttributeError):
                continue
        else:
            content = raw.decode('utf-8', errors='replace')
        
        reader = csv.reader(io.StringIO(content))
        return [row for row in reader]

# ─── Quick Entry API (inline ledger posting) ───────────────────────

@app.route('/api/quick-entry/<int:account_id>', methods=['POST'])
def api_quick_entry(account_id):
    """Post a transaction from the inline ledger entry row."""
    try:
        date_str = request.form.get('date', '')
        reference = request.form.get('reference', '')
        description = request.form.get('description', '')
        amount_str = request.form.get('amount', '0')
        cross_name = request.form.get('cross_account', '').strip().upper()
        gst_split = request.form.get('gst_split', '0') == '1'
        
        if not date_str:
            return jsonify({'ok': False, 'error': 'Date required'})
        
        # Validate and normalize date
        parsed_date = _parse_date(date_str)
        if not parsed_date:
            return jsonify({'ok': False, 'error': 'Invalid date format. Use yyyy-mm-dd (e.g. 2025-01-15)'})
        date_str = parsed_date
        
        # AJE reference format enforcement
        parent_rpt = models.find_report_for_account(account_id)
        if parent_rpt and parent_rpt['name'] == 'AJE' and reference.strip():
            if not re.match(r'^\d{2}AJE\d{2,}$', reference.strip(), re.IGNORECASE):
                return jsonify({'ok': False, 'error': 'AJE reference must be in format xxAJEyy (e.g. 25AJE03). xx=year, yy=sequence number.'})

        amount = models.parse_amount(amount_str) if amount_str.strip() else 0
        this_acct = models.get_account(account_id)

        # LAP balance-field entry: type the running BALANCE instead of the
        # amount and the amount is derived (target − current closing balance,
        # both in the ledger's display sign).
        balance_str = request.form.get('balance', '')
        if not amount_str.strip() and balance_str.strip():
            dsign = -1 if this_acct['normal_balance'] == 'C' else 1
            target = models.parse_amount(balance_str)
            amount = target - models.get_account_balance(account_id) * dsign

        # LAP [Next Ref#]: blank reference on an auto-numbered account takes the
        # counter value; bumped only after the post succeeds (manual refs never bump).
        auto_ref = False
        next_ref = this_acct['next_ref'] if 'next_ref' in this_acct.keys() else 0
        if not reference.strip() and next_ref and next_ref > 0:
            reference = str(next_ref)
            auto_ref = True

        # Flip sign for credit-normal accounts: user types positive to mean
        # "increase this account" which for a credit-normal account means a credit (negative internally)
        if this_acct['normal_balance'] == 'C' and amount != 0:
            amount = -amount
        
        # Blank/placeholder entry: date only, no amount, no cross-account
        # Creates a single zero-amount line (memo/separator)
        if amount == 0 and not cross_name:
            # Memo lines obey the same date fences as real postings
            lock = models.get_meta('lock_date', '')
            if lock and date_str <= lock:
                return jsonify({'ok': False, 'error': f'Date {date_str} is on or before the lock date ({lock}).'})
            ceiling = models.fiscal_ceiling()
            if ceiling and date_str > ceiling:
                return jsonify({'ok': False, 'error': f'Date {date_str} is after the fiscal year end ({ceiling}).'})
            if not reference or not reference.strip():
                reference = models.generate_ref()
            with models.get_db() as db:
                cur = db.execute("INSERT INTO transactions(date, reference, description) VALUES(?,?,?)",
                    (date_str, reference, description))
                txn_id = cur.lastrowid
                db.execute("INSERT INTO lines(transaction_id, account_id, amount, description, sort_order) VALUES(?,?,0,?,0)",
                    (txn_id, account_id, description))
            if auto_ref: models.bump_next_ref(account_id)
            return jsonify({'ok': True})
        
        if gst_split and amount != 0:
            gst_type = request.form.get('gst_type', 'purchase')
            
            # Load configurable rate and accounts from settings
            rate_num = int(models.get_meta('gst_rate_num', '5'))
            rate_den = int(models.get_meta('gst_rate_den', '105'))
            
            if gst_type == 'sale':
                gst_acct_name = models.get_meta('f9_tax_acct', 'GST.OUT')
                default_post = models.get_meta('f9_post_acct', '')
            else:
                gst_acct_name = models.get_meta('f8_tax_acct', 'GST.IN')
                default_post = models.get_meta('f8_post_acct', '')
            
            # Use the cross-account if provided, otherwise the configured default, otherwise EX.SUSP
            if cross_name:
                cross_acct = models.get_account_by_name(cross_name)
                if not cross_acct:
                    return jsonify({'ok': False, 'error': f'Account not found: {cross_name}'})
            elif default_post:
                cross_acct = models.get_account_by_name(default_post)
                if not cross_acct:
                    return jsonify({'ok': False, 'error': f'Default posting account "{default_post}" not found. Check Options.'})
            else:
                cross_acct = models.get_account_by_name('EX.SUSP')
                if not cross_acct:
                    return jsonify({'ok': False, 'error': 'EX.SUSP account not found'})
            
            gst_acct = models.get_account_by_name(gst_acct_name)
            if not gst_acct:
                return jsonify({'ok': False, 'error': f'Tax account "{gst_acct_name}" not found. Check Options.'})
            
            abs_amt = abs(amount)
            gst_cents = round(abs_amt * rate_num / rate_den)
            net_cents = abs_amt - gst_cents
            gst_desc = 'GST collected' if gst_type == 'sale' else 'ITCs paid'
            
            if amount < 0:
                lines = [
                    (account_id, amount, description),
                    (cross_acct['id'], net_cents, description),
                    (gst_acct['id'], gst_cents, gst_desc),
                ]
            else:
                lines = [
                    (account_id, amount, description),
                    (cross_acct['id'], -net_cents, description),
                    (gst_acct['id'], -gst_cents, gst_desc),
                ]
            
            models.add_transaction(date_str, reference, description, lines)
            if auto_ref: models.bump_next_ref(account_id)
            return jsonify({'ok': True})

        if cross_name:
            cross_acct = models.get_account_by_name(cross_name)
            if not cross_acct:
                return jsonify({'ok': False, 'error': f'Account not found: {cross_name}'})
            
            if amount > 0:
                models.add_simple_transaction(date_str, reference, description,
                    account_id, cross_acct['id'], amount)
            elif amount < 0:
                models.add_simple_transaction(date_str, reference, description,
                    cross_acct['id'], account_id, abs(amount))
            else:
                return jsonify({'ok': False, 'error': 'Amount cannot be zero'})
        else:
            return jsonify({'ok': False, 'error': 'Cross-account required'})

        if auto_ref: models.bump_next_ref(account_id)
        return jsonify({'ok': True})

    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/account-search')
def api_account_search():
    q = request.args.get('q', '')
    posting_only = request.args.get('posting', '0') == '1'
    if len(q) < 1:
        return jsonify([])
    accounts = models.search_accounts(q)
    on_report = models.accounts_on_any_report() if posting_only else None
    results = []
    for a in accounts:
        if posting_only:
            if a['account_type'] == 'total':
                continue
            if a['id'] not in on_report:   # orphan: removed from all reports — not postable
                continue
        results.append({'name': a['name'], 'description': a['description'], 'id': a['id']})
    return jsonify(results)

@app.route('/api/inline-edit', methods=['POST'])
def api_inline_edit():
    """Edit a single field on an existing transaction/line — inline editing."""
    try:
        data = request.get_json()
        txn_id = data.get('txn_id')
        line_id = data.get('line_id')
        field = data.get('field')
        value = data.get('value', '')
        
        with models.get_db() as db:
            lock = models.get_meta('lock_date', '')
            txn = db.execute("SELECT date FROM transactions WHERE id=?", (txn_id,)).fetchone()

            # Opening balances: editable, but only after the operator has been
            # asked. The browser turns this into "are you sure?" and retries.
            if not _opening_confirmed() and models.is_opening_txn(txn_id, db):
                return jsonify({'ok': False, 'error': models.OPENING_EDIT_WARNING})

            # Reconciled lines are settled against a statement — no field of the
            # transaction may change until they are unreconciled (LAP rule).
            rec = models.reconciled_count(txn_id, db)
            if rec:
                return jsonify({'ok': False, 'error':
                    f'{rec} line(s) reconciled — unreconcile first (✓ flag or Reconcile screen)'})

            # For non-date fields: reject if transaction is already locked
            if field != 'date' and lock and txn and txn['date'] <= lock:
                return jsonify({'ok': False, 'error': f'Transaction is locked (before {lock})'})
            
            if field == 'date':
                parsed = _parse_date(value)
                if not parsed:
                    return jsonify({'ok': False, 'error': 'Invalid date format. Use yyyy-mm-dd.'})
                # Validate year is reasonable
                try:
                    from datetime import datetime as dt
                    d = dt.strptime(parsed, '%Y-%m-%d')
                    if d.year < 1950 or d.year > 2099:
                        return jsonify({'ok': False, 'error': f'Year {d.year} is out of range (1950-2099).'})
                except:
                    return jsonify({'ok': False, 'error': 'Invalid date.'})
                # Check lock: don't allow moving INTO a locked period
                if lock and parsed <= lock:
                    return jsonify({'ok': False, 'error': f'Cannot set date to {parsed} — on or before lock date ({lock}). Change lock date first.'})
                # Check lock: don't allow editing a transaction that's already locked
                if lock and txn and txn['date'] <= lock:
                    return jsonify({'ok': False, 'error': f'Transaction is locked (date {txn["date"]} is before lock date {lock}).'})
                db.execute("UPDATE transactions SET date=? WHERE id=?", (parsed, txn_id))
            elif field == 'reference':
                db.execute("UPDATE transactions SET reference=? WHERE id=?", (value, txn_id))
            elif field == 'description':
                db.execute("UPDATE transactions SET description=? WHERE id=?", (value, txn_id))
                db.execute("UPDATE lines SET description=? WHERE id=?", (value, line_id))
            elif field == 'amount':
                # Check if this is a split (multi-line) transaction
                line_count = db.execute(
                    "SELECT COUNT(*) as cnt FROM lines WHERE transaction_id=?",
                    (txn_id,)).fetchone()['cnt']
                if line_count > 2:
                    return jsonify({'ok': False, 'error': 'Split transaction — edit in transaction detail view'})
                new_amt = models.parse_amount(value) if value.strip() else 0
                # The user sees amounts in normal-balance terms, so flip for credit-normal accounts
                acct_row = db.execute(
                    "SELECT a.normal_balance FROM lines l JOIN accounts a ON l.account_id=a.id WHERE l.id=?",
                    (line_id,)).fetchone()
                if acct_row and acct_row['normal_balance'] == 'C':
                    new_amt = -new_amt
                old_line = db.execute("SELECT amount, account_id FROM lines WHERE id=?", (line_id,)).fetchone()
                old_amt = old_line['amount']
                diff = new_amt - old_amt
                # Update this line
                db.execute("UPDATE lines SET amount=? WHERE id=?", (new_amt, line_id))
                # Find the cross-account line and adjust it by the opposite
                cross = db.execute(
                    "SELECT id, amount FROM lines WHERE transaction_id=? AND id!=? LIMIT 1",
                    (txn_id, line_id)).fetchone()
                if cross:
                    db.execute("UPDATE lines SET amount=? WHERE id=?", (cross['amount'] - diff, cross['id']))
            elif field == 'account':
                # Change the cross-account on a 2-line transaction
                line_count = db.execute(
                    "SELECT COUNT(*) as cnt FROM lines WHERE transaction_id=?",
                    (txn_id,)).fetchone()['cnt']
                if line_count > 2:
                    return jsonify({'ok': False, 'error': 'Split transaction — edit in transaction detail view'})
                acct_name = value.strip().upper()
                if not acct_name:
                    return jsonify({'ok': False, 'error': 'Account name required'})
                new_acct = models.get_account_by_name(acct_name)
                if not new_acct:
                    return jsonify({'ok': False, 'error': f'Account "{acct_name}" not found'})
                if new_acct['account_type'] == 'total':
                    return jsonify({'ok': False, 'error': f'"{acct_name}" is a total/computed account — you can\'t post to it. Pick a detail account.'})
                # Find the cross-account line (the OTHER line, not the one we're viewing)
                cross = db.execute(
                    "SELECT id, account_id FROM lines WHERE transaction_id=? AND id!=? LIMIT 1",
                    (txn_id, line_id)).fetchone()
                if cross:
                    db.execute("UPDATE lines SET account_id=? WHERE id=?", (new_acct['id'], cross['id']))
                else:
                    return jsonify({'ok': False, 'error': 'No cross-account line found'})
            else:
                return jsonify({'ok': False, 'error': f'Unknown field: {field}'})
        
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/report/<int:report_id>/sort-accounts', methods=['POST'])
def api_sort_accounts(report_id):
    """Sort account rows alphabetically within sections (between labels/totals)."""
    try:
        items = models.get_report_items(report_id)
        # Group items into sections separated by non-account rows
        sections = []
        current_section = []
        non_account_buffer = []
        
        for item in items:
            itype = item['item_type']
            if itype == 'account':
                current_section.append(item)
            else:
                if current_section:
                    sections.append(('accounts', current_section))
                    current_section = []
                sections.append(('other', [item]))
        if current_section:
            sections.append(('accounts', current_section))
        
        # Sort account sections alphabetically by name, reassign positions
        pos = 10
        with models.get_db() as db:
            for stype, sitems in sections:
                if stype == 'accounts':
                    sitems.sort(key=lambda x: (x['acct_desc'] or x['acct_name'] or '').upper())
                for item in sitems:
                    db.execute("UPDATE report_items SET position=? WHERE id=?", (pos, item['id']))
                    pos += 10

        # An alphabetical sort would put RE.OB above TRX.OPEN and lift both above
        # the heading. The head is not the operator's to arrange, so put it back —
        # the healer is the single source of truth for that layout.
        rpt = models.get_report(report_id)
        if rpt and rpt['name'] == 'TRX':
            models.ensure_trx_layout()

        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/description-suggest')
def api_description_suggest():
    """Suggest descriptions based on previous entries."""
    q = request.args.get('q', '')
    if len(q) < 2:
        return jsonify([])
    with models.get_db() as db:
        rows = db.execute(
            "SELECT DISTINCT description FROM transactions WHERE description LIKE ? ORDER BY description LIMIT 15",
            (f'%{q}%',)).fetchall()
    return jsonify([r['description'] for r in rows if r['description']])

@app.route('/api/block-move', methods=['POST'])
def api_block_move():
    """Move a block of transaction lines from one account to another.
    Expects: {line_ids: [...], from_account_id: int, to_account_name: str}"""
    try:
        data = request.get_json()
        line_ids = data.get('line_ids', [])
        to_name = data.get('to_account_name', '').strip().upper()
        
        if not line_ids:
            return jsonify({'ok': False, 'error': 'No lines selected'})
        if not to_name:
            return jsonify({'ok': False, 'error': 'No target account specified'})
        
        to_acct = models.get_account_by_name(to_name)
        if not to_acct:
            return jsonify({'ok': False, 'error': f'Account "{to_name}" not found'})
        if to_acct['account_type'] == 'total':
            return jsonify({'ok': False, 'error': f'"{to_name}" is a total/computed account — you can\'t move lines onto it. Pick a detail account.'})

        lock = models.get_meta('lock_date', '')
        
        allow_opening = _opening_confirmed()
        with models.get_db() as db:
            moved = 0
            skipped_rec = 0
            skipped_open = 0
            for lid in line_ids:
                row = db.execute(
                    "SELECT t.id AS tid, t.date FROM lines l JOIN transactions t ON l.transaction_id=t.id WHERE l.id=?",
                    (lid,)).fetchone()
                if not row:
                    continue
                if lock and row['date'] <= lock:
                    continue  # skip locked
                if models.reconciled_count(row['tid'], db):
                    skipped_rec += 1   # settled against a statement — don't move silently
                    continue
                # A block move is a sweep — it must not repoint the conversion
                # unless that was actually the intent.
                if not allow_opening and models.is_opening_txn(row['tid'], db):
                    skipped_open += 1
                    continue
                db.execute("UPDATE lines SET account_id=? WHERE id=?", (to_acct['id'], lid))
                moved += 1

        return jsonify({'ok': True, 'moved': moved, 'skipped_reconciled': skipped_rec,
                        'skipped_opening': skipped_open})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/account-balance/<int:account_id>')
def api_account_balance(account_id):
    bal = models.get_account_balance(account_id)
    return jsonify({'balance': bal, 'formatted': models.fmt_amount(bal)})

# ─── Stripe Deposit Procedure ──────────────────────────────────────

@app.route('/api/stripe-config')
def api_stripe_config():
    """Return Stripe procedure configuration and AR customer list."""
    fee_acct = models.get_meta('stripe_fee_acct', 'EX.CC')
    ar_report_name = models.get_meta('stripe_ar_report', '')
    
    # Find AR customer accounts
    customers = []
    ar_report = None
    if ar_report_name:
        ar_report = models.find_report_by_name(ar_report_name)
    if not ar_report:
        # Try common names
        for name in ['AR', 'Accounts Receivable', 'Receivable']:
            ar_report = models.find_report_by_name(name)
            if ar_report:
                break
    if ar_report:
        accts = models.get_report_accounts(ar_report['id'])
        customers = [{'id': a['id'], 'name': a['name'], 'description': a['description']} for a in accts]
    
    # Verify fee account exists
    fee_acct_obj = models.get_account_by_name(fee_acct)
    
    return jsonify({
        'fee_account': fee_acct,
        'fee_account_ok': fee_acct_obj is not None,
        'ar_report': ar_report['name'] if ar_report else '',
        'ar_report_id': ar_report['id'] if ar_report else None,
        'customers': customers
    })

@app.route('/api/stripe-post', methods=['POST'])
def api_stripe_post():
    """Post ALL Stripe deposit charges as ONE compound distribution entry.
    
    Creates 1 transaction structured as:
      Line 0:  DR clearing_account  (total net — this is what shows in the clearing ledger)
      Then for each charge:
        CR customer_account  (gross amount)
        DR fee_account       (fee amount)  — omitted if fee is 0
    
    The clearing ledger shows ONE row for the total net deposit.
    F3/double-click opens the distribution showing all the customer + fee detail.
    
    REF format: S.MMMdd (e.g. S.Jun01)
    Description: Stripe deposits, net
    """
    try:
        from datetime import datetime as dt
        data = request.get_json()
        deposit_date = data.get('date', '')
        items = data.get('items', [])
        clearing_account_id = data.get('clearing_account_id')
        fee_acct_name = data.get('fee_account', 'EX.CC')
        
        if not deposit_date:
            return jsonify({'ok': False, 'error': 'Deposit date is required'})
        parsed_date = _parse_date(deposit_date)
        if not parsed_date:
            return jsonify({'ok': False, 'error': 'Invalid date format'})
        
        if not items:
            return jsonify({'ok': False, 'error': 'No line items to post'})
        
        # Build REF as S.MMMdd
        try:
            d = dt.strptime(parsed_date, '%Y-%m-%d')
            ref = 'S.' + d.strftime('%b%d')  # e.g. S.Jun01
        except:
            ref = 'S.Stripe'
        
        # Resolve fee account
        fee_acct = models.get_account_by_name(fee_acct_name)
        if not fee_acct:
            return jsonify({'ok': False, 'error': f'Fee account "{fee_acct_name}" not found'})
        
        # Resolve clearing account
        if not clearing_account_id:
            return jsonify({'ok': False, 'error': 'Clearing account not specified'})
        clearing_acct = models.get_account(clearing_account_id)
        if not clearing_acct:
            return jsonify({'ok': False, 'error': 'Clearing account not found'})
        
        # Validate all lines first, accumulate totals and build detail lines
        detail_lines = []  # customer CRs and fee DRs
        total_net = 0
        total_fees = 0
        total_gross = 0
        errors = []
        
        for i, item in enumerate(items):
            customer_name = item.get('customer', '').strip().upper()
            gross_str = item.get('gross', '0')
            fee_str = item.get('fee', '0')
            description = item.get('description', '').strip()
            
            if not customer_name:
                errors.append(f'Line {i+1}: Client AR account required')
                continue
            
            gross_cents = models.parse_amount(gross_str)
            fee_cents = models.parse_amount(fee_str)
            
            if gross_cents <= 0:
                errors.append(f'Line {i+1}: Gross must be positive')
                continue
            if fee_cents < 0:
                errors.append(f'Line {i+1}: Fee cannot be negative')
                continue
            
            net_cents = gross_cents - fee_cents
            
            cust_acct = models.get_account_by_name(customer_name)
            if not cust_acct:
                errors.append(f'Line {i+1}: Account "{customer_name}" not found')
                continue
            
            line_desc = description or 'Stripe pmt'
            
            # CR customer AR (gross) — goes into distribution detail
            detail_lines.append((cust_acct['id'], -gross_cents, line_desc))
            # DR fee account — goes into distribution detail
            if fee_cents > 0:
                detail_lines.append((fee_acct['id'], fee_cents, line_desc))
            
            total_net += net_cents
            total_fees += fee_cents
            total_gross += gross_cents
        
        if errors:
            return jsonify({'ok': False, 'error': 'Validation errors', 'errors': errors})
        
        if not detail_lines:
            return jsonify({'ok': False, 'error': 'No valid lines to post'})
        
        # Build the full transaction:
        # Line 0: ONE DR to clearing for the total net (this is the single ledger row)
        # Lines 1+: all the customer CRs and fee DRs (distribution detail via F3)
        txn_desc = 'Stripe deposits, net'
        all_lines = [(clearing_acct['id'], total_net, txn_desc)] + detail_lines
        
        models.add_transaction(parsed_date, ref, txn_desc, all_lines)
        
        return jsonify({
            'ok': True,
            'posted': len(items),
            'total_items': len(items),
            'total_net': models.fmt_amount(total_net),
            'total_fees': models.fmt_amount(total_fees),
            'total_gross': models.fmt_amount(total_gross),
            'errors': []
        })
        
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ─── CaseWare AJE Import ──────────────────────────────────────────

@app.route('/api/parse-csw-aje', methods=['POST'])
def api_parse_csw_aje():
    """Parse a CaseWare AJE export file (IIF or Venice)."""
    if not models.get_db_path():
        return jsonify({'ok': False, 'error': 'No books open'})

    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file uploaded'})

    f = request.files['file']
    if not f.filename:
        return jsonify({'ok': False, 'error': 'No file selected'})

    # Save to temp
    import tempfile, uuid
    tmp_dir = os.path.join(tempfile.gettempdir(), 'grid_import')
    os.makedirs(tmp_dir, exist_ok=True)
    cache_key = str(uuid.uuid4())[:8]
    tmp_path = os.path.join(tmp_dir, f'csw_{cache_key}_{f.filename}')
    f.save(tmp_path)

    try:
        parsed = models.parse_csw_aje(tmp_path)
    except ValueError as e:
        os.remove(tmp_path)
        return jsonify({'ok': False, 'error': str(e)})

    if not parsed['entries']:
        os.remove(tmp_path)
        return jsonify({'ok': False, 'error': 'No entries found in file'})

    # Auto-match accounts
    suggestions = models.auto_match_accounts(parsed['csw_accounts'])

    # Cache parsed data for import step
    import json as _json
    cache_path = os.path.join(tmp_dir, f'csw_{cache_key}.json')
    with open(cache_path, 'w') as cf:
        _json.dump(parsed, cf)
    os.remove(tmp_path)

    # Build suggestions dict for JSON
    sugg_json = {}
    for csw_name, match in suggestions.items():
        sugg_json[csw_name] = match  # {'id': int, 'name': str} or None

    # Get fiscal year for default prefix
    fy = models.get_meta('fiscal_year', '')
    default_prefix = f"{fy[-2:]}AJE" if fy else 'AJE'

    return jsonify({
        'ok': True,
        'format': parsed['format'],
        'entry_count': len(parsed['entries']),
        'entries': parsed['entries'],
        'csw_accounts': parsed['csw_accounts'],
        'suggestions': sugg_json,
        'cache_key': cache_key,
        'default_prefix': default_prefix,
    })


@app.route('/api/import-csw-aje', methods=['POST'])
def api_import_csw_aje():
    """Import parsed CaseWare AJE entries."""
    if not models.get_db_path():
        return jsonify({'ok': False, 'error': 'No books open'})

    try:
        import json as _json
        data = request.get_json()
        cache_key = data.get('cache_key', '')
        account_map = data.get('account_map', {})  # {csw_name: grid_acct_id}
        ref_prefix = data.get('ref_prefix', 'AJE')

        # Load cached parsed data
        import tempfile
        tmp_dir = os.path.join(tempfile.gettempdir(), 'grid_import')
        cache_path = os.path.join(tmp_dir, f'csw_{cache_key}.json')

        if not os.path.exists(cache_path):
            return jsonify({'ok': False, 'error': 'Parsed data expired. Please re-upload the file.'})

        with open(cache_path, 'r') as cf:
            parsed = _json.load(cf)

        # Convert account_map values to int
        int_map = {}
        for csw_name, acct_id in account_map.items():
            int_map[csw_name] = int(acct_id)

        result = models.import_aje_entries(parsed['entries'], int_map, ref_prefix, journal_account=ref_prefix)

        # Clean up cache
        try:
            os.remove(cache_path)
        except OSError:
            pass

        return jsonify({
            'ok': True,
            'posted': result['posted'],
            'skipped': result['skipped'],
            'errors': result['errors'],
        })

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ─── CSV Export ─────────────────────────────────────────────────────

@app.route('/export/ledger/<int:account_id>')
def export_ledger(account_id):
    import csv, io
    account = models.get_account(account_id)
    entries = models.get_ledger(account_id)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Reference', 'Description', 'Cross Account', 
                    'Debit', 'Credit', 'Balance'])
    for e in entries:
        dr = models.fmt_amount_plain(e['amount']) if e['amount'] > 0 else ''
        cr = models.fmt_amount_plain(abs(e['amount'])) if e['amount'] < 0 else ''
        writer.writerow([e['date'], e['reference'], e['description'],
                        e['cross_accounts'], dr, cr,
                        models.fmt_amount_plain(e['running_balance'])])
    output.seek(0)
    return output.getvalue(), 200, {
        'Content-Type': 'text/csv',
        'Content-Disposition': f'attachment; filename={account["name"]}_ledger.csv'
    }

@app.route('/export/trial-balance')
def export_trial_balance():
    import csv, io
    accounts, total_dr, total_cr = models.get_trial_balance()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Account', 'Number', 'Description', 'Debit', 'Credit'])
    for a in accounts:
        writer.writerow([a['name'], a['account_number'], a['description'],
                        models.fmt_amount_plain(a['debit']) if a['debit'] else '',
                        models.fmt_amount_plain(a['credit']) if a['credit'] else ''])
    writer.writerow(['', '', 'TOTALS', models.fmt_amount_plain(total_dr),
                    models.fmt_amount_plain(total_cr)])
    output.seek(0)
    return output.getvalue(), 200, {
        'Content-Type': 'text/csv',
        'Content-Disposition': 'attachment; filename=trial_balance.csv'
    }

# ─── Books Export / Import (Full Backup & Restore) ───────────────

@app.route('/export/structure')
def export_structure():
    """Export the full account framework: meta, accounts, reports+items, tax codes, import rules."""
    if not models.get_db_path():
        return redirect(url_for('library'))

    company = models.get_meta('company_name', 'Books')

    with models.get_db() as db:
        # Meta (all settings)
        meta = {}
        for row in db.execute("SELECT key, value FROM meta").fetchall():
            meta[row['key']] = row['value']

        # Accounts — keyed by name for portability
        accounts = []
        for a in db.execute("SELECT * FROM accounts ORDER BY id").fetchall():
            accounts.append({
                'name': a['name'],
                'description': a['description'],
                'normal_balance': a['normal_balance'],
                'account_type': a['account_type'],
                'account_number': a['account_number'] or '',
                'notes': a['notes'] or '',
            })

        # Reports with their items
        reports = []
        for r in db.execute("SELECT * FROM reports ORDER BY sort_order, id").fetchall():
            items = []
            for it in db.execute(
                "SELECT ri.*, a.name as acct_name FROM report_items ri "
                "LEFT JOIN accounts a ON ri.account_id = a.id "
                "WHERE ri.report_id=? ORDER BY ri.position", (r['id'],)).fetchall():
                items.append({
                    'position': it['position'],
                    'item_type': it['item_type'],
                    'description': it['description'] or '',
                    'account_name': it['acct_name'] or '',
                    'indent': it['indent'],
                    'total_to_1': it['total_to_1'] or '',
                    'total_to_2': it['total_to_2'] or '',
                    'total_to_3': it['total_to_3'] or '',
                    'total_to_4': it['total_to_4'] or '',
                    'total_to_5': it['total_to_5'] or '',
                    'total_to_6': it['total_to_6'] or '',
                    'sep_style': it['sep_style'] or '',
                })
            reports.append({
                'name': r['name'],
                'description': r['description'] or '',
                'sort_order': r['sort_order'],
                'period_begin': r['period_begin'] or '',
                'period_end': r['period_end'] or '',
                'items': items,
            })

        # Tax codes
        tax_codes = []
        for tc in db.execute("SELECT * FROM tax_codes ORDER BY id").fetchall():
            tax_codes.append({
                'id': tc['id'],
                'description': tc['description'] or '',
                'rate_percent': tc['rate_percent'],
                'collected_account': tc['collected_account'] or '',
                'paid_account': tc['paid_account'] or '',
            })

        # Import rules
        import_rules = []
        for ir in db.execute("SELECT * FROM import_rules ORDER BY priority DESC, keyword").fetchall():
            import_rules.append({
                'keyword': ir['keyword'],
                'account_name': ir['account_name'],
                'tax_code': ir['tax_code'] or '',
                'priority': ir['priority'],
                'notes': ir['notes'] or '',
            })

    payload = {
        '_grid_export': 'structure',
        '_version': 1,
        '_exported': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        '_company': company,
        'meta': meta,
        'accounts': accounts,
        'reports': reports,
        'tax_codes': tax_codes,
        'import_rules': import_rules,
    }

    data = json.dumps(payload, indent=2)
    safe_name = company.replace(' ', '_').replace('/', '-')[:40]
    return data, 200, {
        'Content-Type': 'application/json',
        'Content-Disposition': f'attachment; filename={safe_name}_structure.json'
    }


@app.route('/export/data')
def export_data():
    """Export all transactions and lines. Accounts referenced by name."""
    if not models.get_db_path():
        return redirect(url_for('library'))

    company = models.get_meta('company_name', 'Books')

    with models.get_db() as db:
        # Build account ID → name map
        acct_map = {}
        for a in db.execute("SELECT id, name FROM accounts").fetchall():
            acct_map[a['id']] = a['name']

        transactions = []
        for t in db.execute("SELECT * FROM transactions ORDER BY date, id").fetchall():
            lines = []
            for ln in db.execute(
                "SELECT * FROM lines WHERE transaction_id=? ORDER BY sort_order", (t['id'],)).fetchall():
                lines.append({
                    'account_name': acct_map.get(ln['account_id'], f'UNKNOWN_{ln["account_id"]}'),
                    'amount': ln['amount'],
                    'description': ln['description'] or '',
                    'reconciled': ln['reconciled'],
                    'sort_order': ln['sort_order'],
                })
            transactions.append({
                'date': t['date'],
                'reference': t['reference'] or '',
                'description': t['description'] or '',
                'created_at': t['created_at'] or '',
                'lines': lines,
            })

    payload = {
        '_grid_export': 'data',
        '_version': 1,
        '_exported': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        '_company': company,
        '_transaction_count': len(transactions),
        'transactions': transactions,
    }

    data = json.dumps(payload, indent=2)
    safe_name = company.replace(' ', '_').replace('/', '-')[:40]
    return data, 200, {
        'Content-Type': 'application/json',
        'Content-Disposition': f'attachment; filename={safe_name}_data.json'
    }


@app.route('/export/writeup')
def export_writeup():
    """The write-up handoff (Willy): CY+PY itemized, 5-year comparatives,
    chart, layouts, check_books. One definition — models.export_writeup()."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    try:
        payload = models.export_writeup()
    except ValueError as e:
        flash(str(e))
        return redirect(url_for('tools'))
    company = payload['meta']['company_name']
    safe_name = company.replace(' ', '_').replace('/', '-')[:40]
    fy = payload['meta']['fiscal_year']
    return json.dumps(payload, indent=1), 200, {
        'Content-Type': 'application/json',
        'Content-Disposition':
            f'attachment; filename={safe_name}_{fy}_writeup.json'
    }


def _apply_structure(payload, txn_count=0):
    """Write a structure payload (accounts, reports, tax codes, rules) into the
    OPEN books. One definition — every structure import
    comes through here, so they can never disagree about what an import means.

    Returns (accounts, reports, rules) actually written."""
    with models.get_db() as db:
        # Clear existing structure (order matters for FK constraints)
        db.execute("DELETE FROM report_items")
        db.execute("DELETE FROM reports")
        # Only delete accounts that have no transaction lines
        # For a fresh file this deletes all; for a file with data, preserve referenced accounts
        if txn_count == 0:
            db.execute("DELETE FROM accounts")
        db.execute("DELETE FROM tax_codes")
        db.execute("DELETE FROM import_rules")

        # Meta
        for key, value in payload.get('meta', {}).items():
            db.execute("INSERT OR REPLACE INTO meta(key, value) VALUES(?,?)", (key, value))

        # Accounts
        acct_name_to_id = {}
        for a in payload.get('accounts', []):
            # Check if account already exists (e.g. in a file with data)
            existing = db.execute("SELECT id FROM accounts WHERE name=? COLLATE NOCASE",
                                 (a['name'],)).fetchone()
            if existing:
                acct_name_to_id[a['name']] = existing['id']
                # Update metadata
                db.execute("UPDATE accounts SET description=?, normal_balance=?, account_type=?, "
                          "account_number=?, notes=? WHERE id=?",
                          (a.get('description', ''), a['normal_balance'], a['account_type'],
                           a.get('account_number', ''), a.get('notes', ''), existing['id']))
            else:
                cur = db.execute(
                    "INSERT INTO accounts(name, description, normal_balance, account_type, account_number, notes) "
                    "VALUES(?,?,?,?,?,?)",
                    (a['name'], a.get('description', ''), a['normal_balance'],
                     a['account_type'], a.get('account_number', ''), a.get('notes', '')))
                acct_name_to_id[a['name']] = cur.lastrowid

        # Reports + items
        for r in payload.get('reports', []):
            cur = db.execute(
                "INSERT INTO reports(name, description, sort_order, period_begin, period_end) VALUES(?,?,?,?,?)",
                (r['name'], r.get('description', ''), r.get('sort_order', 0),
                 r.get('period_begin', ''), r.get('period_end', '')))
            report_id = cur.lastrowid
            for it in r.get('items', []):
                acct_id = acct_name_to_id.get(it.get('account_name')) if it.get('account_name') else None
                db.execute(
                    "INSERT INTO report_items(report_id, position, item_type, description, account_id, "
                    "indent, total_to_1, total_to_2, total_to_3, total_to_4, total_to_5, total_to_6, sep_style) "
                    "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (report_id, it.get('position', 0), it['item_type'], it.get('description', ''),
                     acct_id, it.get('indent', 0),
                     it.get('total_to_1', ''), it.get('total_to_2', ''), it.get('total_to_3', ''),
                     it.get('total_to_4', ''), it.get('total_to_5', ''), it.get('total_to_6', ''),
                     it.get('sep_style', '')))

        # Tax codes
        for tc in payload.get('tax_codes', []):
            db.execute(
                "INSERT OR REPLACE INTO tax_codes(id, description, rate_percent, collected_account, paid_account) "
                "VALUES(?,?,?,?,?)",
                (tc['id'], tc.get('description', ''), tc.get('rate_percent', 0),
                 tc.get('collected_account', ''), tc.get('paid_account', '')))

        # Import rules
        for ir in payload.get('import_rules', []):
            db.execute(
                "INSERT INTO import_rules(keyword, account_name, tax_code, priority, notes) VALUES(?,?,?,?,?)",
                (ir['keyword'], ir['account_name'], ir.get('tax_code', ''),
                 ir.get('priority', 0), ir.get('notes', '')))
    return (len(payload.get('accounts', [])), len(payload.get('reports', [])),
            len(payload.get('import_rules', [])))


@app.route('/import/structure', methods=['POST'])
def import_structure():
    """Import structure JSON into the CURRENT open books file.
    WARNING: This replaces all accounts, reports, tax codes, and rules."""
    if not models.get_db_path():
        flash('No books file open', 'error')
        return redirect(url_for('library'))

    f = request.files.get('file')
    if not f:
        flash('No file selected', 'error')
        return redirect(url_for('settings'))

    try:
        raw = f.read().decode('utf-8')
        payload = json.loads(raw)
    except Exception as e:
        flash(f'Invalid JSON file: {e}', 'error')
        return redirect(url_for('settings'))

    if payload.get('_grid_export') != 'structure':
        flash('This is not a Grid structure export file', 'error')
        return redirect(url_for('settings'))

    # Check if there are existing transactions — warn user
    with models.get_db() as db:
        txn_count = db.execute("SELECT COUNT(*) as cnt FROM transactions").fetchone()['cnt']
        if txn_count > 0 and not request.form.get('confirm_overwrite'):
            flash(f'This file has {txn_count} transactions. Import structure into a FRESH file '
                  '(create one first via Library → New). Importing structure wipes accounts/reports.', 'error')
            return redirect(url_for('settings'))

    try:
        acct_count, rpt_count, rule_count = _apply_structure(payload, txn_count)
        flash(f'Structure imported: {acct_count} accounts, {rpt_count} reports, {rule_count} rules', 'success')

    except Exception as e:
        flash(f'Import failed: {e}', 'error')

    return redirect(url_for('settings'))


@app.route('/import/data', methods=['POST'])
def import_data():
    """Import transactions from a data JSON file into the current books."""
    if not models.get_db_path():
        flash('No books file open', 'error')
        return redirect(url_for('library'))

    f = request.files.get('file')
    if not f:
        flash('No file selected', 'error')
        return redirect(url_for('settings'))

    try:
        raw = f.read().decode('utf-8')
        payload = json.loads(raw)
    except Exception as e:
        flash(f'Invalid JSON file: {e}', 'error')
        return redirect(url_for('settings'))

    if payload.get('_grid_export') != 'data':
        flash('This is not a Grid data export file', 'error')
        return redirect(url_for('settings'))

    # Build account name → id map from current database (total accounts are
    # not postable — a restore must not smuggle lines onto them)
    with models.get_db() as db:
        acct_map = {}
        total_names = set()
        for a in db.execute("SELECT id, name, account_type FROM accounts").fetchall():
            acct_map[a['name'].upper()] = a['id']
            if a['account_type'] == 'total':
                total_names.add(a['name'].upper())

    # Temporarily disable lock date for import
    saved_lock = models.get_meta('lock_date', '')
    if saved_lock:
        models.set_meta('lock_date', '')

    imported = 0
    skipped = 0
    errors = []

    try:
        with models.get_db() as db:
            for txn in payload.get('transactions', []):
                # Resolve account names to IDs
                lines_resolved = []
                missing = []
                for ln in txn.get('lines', []):
                    acct_name = ln['account_name']
                    acct_id = acct_map.get(acct_name.upper())
                    if not acct_id:
                        missing.append(acct_name)
                    elif acct_name.upper() in total_names:
                        missing.append(f"{acct_name} (total account — not postable)")
                    else:
                        lines_resolved.append((acct_id, ln['amount'], ln.get('description', ''),
                                              ln.get('reconciled', 0), ln.get('sort_order', 0)))

                if missing:
                    errors.append(f"Txn {txn['date']} '{txn['description']}': missing accounts: {', '.join(missing)}")
                    skipped += 1
                    continue

                # Check balance
                total = sum(lr[1] for lr in lines_resolved)
                if total != 0:
                    errors.append(f"Txn {txn['date']} '{txn['description']}': does not balance (off by {total})")
                    skipped += 1
                    continue

                # Insert transaction
                cur = db.execute(
                    "INSERT INTO transactions(date, reference, description, created_at) VALUES(?,?,?,?)",
                    (txn['date'], txn.get('reference', ''), txn.get('description', ''),
                     txn.get('created_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))))
                txn_id = cur.lastrowid

                # Insert lines with reconciled flag
                for acct_id, amount, desc, reconciled, sort_order in lines_resolved:
                    db.execute(
                        "INSERT INTO lines(transaction_id, account_id, amount, description, reconciled, sort_order) "
                        "VALUES(?,?,?,?,?,?)",
                        (txn_id, acct_id, amount, desc, reconciled, sort_order))

                imported += 1

    except Exception as e:
        flash(f'Import failed at transaction {imported + 1}: {e}', 'error')
        # Restore lock date
        if saved_lock:
            models.set_meta('lock_date', saved_lock)
        return redirect(url_for('settings'))

    # Restore lock date
    if saved_lock:
        models.set_meta('lock_date', saved_lock)

    if errors:
        for err in errors[:10]:  # Show first 10 errors
            flash(err, 'error')
        if len(errors) > 10:
            flash(f'... and {len(errors) - 10} more errors', 'error')

    flash(f'Data imported: {imported} transactions ({skipped} skipped)', 'success')
    return redirect(url_for('settings'))


# ─── Import Rules Management ──────────────────────────────────────

@app.route('/rules', methods=['GET'])
def import_rules_page():
    if not models.get_db_path():
        return redirect(url_for('library'))
    company = models.get_meta('company_name', 'My Books')
    rules = models.get_import_rules()
    tax_codes = models.get_tax_codes()
    accounts = models.get_accounts()
    return render_template('rules.html', company=company, rules=rules,
        tax_codes=tax_codes, accounts=accounts)

@app.route('/api/rule-add', methods=['POST'])
def api_rule_add():
    """AJAX endpoint: add an import rule from the import preview screen."""
    try:
        data = request.get_json()
        keyword = (data.get('keyword', '') or '').strip()
        account_name = (data.get('account_name', '') or '').strip().upper()
        tax_code = (data.get('tax_code', '') or '').strip()
        priority = int(data.get('priority', 5))
        if not keyword or not account_name:
            return jsonify({'ok': False, 'error': 'Keyword and account are required'})
        # Verify account exists
        acct = models.get_account_by_name(account_name)
        if not acct:
            return jsonify({'ok': False, 'error': f'Account "{account_name}" not found'})
        models.save_import_rule(None, keyword, account_name, tax_code, priority, '')
        return jsonify({'ok': True, 'keyword': keyword, 'account': account_name, 'tax': tax_code})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/rules/save', methods=['POST'])
def save_rule():
    rid = request.form.get('rule_id', '')
    keyword = request.form.get('keyword', '').strip()
    account_name = request.form.get('account_name', '').strip().upper()
    tax_code = request.form.get('tax_code', '').strip()
    priority = int(request.form.get('priority', '0'))
    notes = request.form.get('notes', '').strip()
    if keyword and account_name:
        models.save_import_rule(int(rid) if rid else None, keyword, account_name, tax_code, priority, notes)
        flash(f'Rule saved: "{keyword}" → {account_name}', 'success')
    else:
        flash('Keyword and account are required', 'error')
    return redirect(url_for('import_rules_page'))

@app.route('/rules/delete/<int:rule_id>', methods=['POST'])
def delete_rule(rule_id):
    models.delete_import_rule(rule_id)
    flash('Rule deleted', 'success')
    return redirect(url_for('import_rules_page'))

@app.route('/tax/save', methods=['POST'])
def save_tax():
    code_id = request.form.get('code_id', '').strip().upper()
    description = request.form.get('description', '').strip()
    rate = float(request.form.get('rate_percent', '0'))
    collected = request.form.get('collected_account', '').strip()
    paid = request.form.get('paid_account', '').strip()
    if code_id:
        models.save_tax_code(code_id, description, rate, collected, paid)
        flash(f'Tax code {code_id} saved', 'success')
    return redirect(url_for('import_rules_page'))

@app.route('/tax/delete/<code_id>', methods=['POST'])
def delete_tax(code_id):
    models.delete_tax_code(code_id)
    flash(f'Tax code {code_id} deleted', 'success')
    return redirect(url_for('import_rules_page'))

# ─── Rules Export/Import ─────────────────────────────────────────

@app.route('/rules/export')
def export_rules():
    """Export import rules as CSV."""
    import csv, io
    rules = models.get_import_rules() if hasattr(models, 'get_import_rules') else []
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['keyword', 'account', 'description', 'tax_code', 'priority'])
    for r in rules:
        w.writerow([r['keyword'], r['account_name'], r['notes'] or '',
                    r['tax_code'] or '', r['priority'] or 0])
    from flask import Response
    return Response(buf.getvalue(), mimetype='text/csv',
                   headers={'Content-Disposition': 'attachment; filename=rules.csv'})

@app.route('/rules/import', methods=['POST'])
def import_rules_csv():
    """Import rules from CSV file."""
    import csv, io
    f = request.files.get('file')
    if not f:
        flash('No file selected', 'error')
        return redirect(url_for('import_rules_page'))
    text = f.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    count = 0
    for row in reader:
        kw = row.get('keyword', '').strip()
        acct = row.get('account', '').strip()
        desc = row.get('description', '').strip()
        tax = row.get('tax_code', '').strip()
        pri = int(row.get('priority', '0') or '0')
        if kw and acct:
            models.save_import_rule(None, kw, acct, tax, pri, desc)
            count += 1
    flash(f'Imported {count} rules', 'success')
    return redirect(url_for('import_rules_page'))

# ─── Template Clone (New File from Existing) ─────────────────────

@app.route('/clone', methods=['GET', 'POST'])
def clone_file():
    """Create a new client file from an existing one — keeps COA, reports, rules; purges transactions.
    The folder name is derived from the legal name + fiscal year (v142)."""
    if request.method == 'POST':
        source = request.form.get('source', '')
        new_company = request.form.get('company', '').strip()
        year = request.form.get('fiscal_year', '').strip()
        if not source:
            flash('All fields required', 'error')
            return redirect(url_for('clone_file'))
        try:
            new_folder = models.client_folder_name(new_company, year)
        except ValueError as e:
            flash(str(e), 'error')
            return redirect(url_for('clone_file'))
        
        import shutil
        base = os.path.dirname(os.path.dirname(source))
        new_dir = os.path.join(base, new_folder)
        os.makedirs(new_dir, exist_ok=True)
        new_path = os.path.join(new_dir, 'books.db')
        if os.path.exists(new_path):
            flash('File already exists in that folder', 'error')
            return redirect(url_for('clone_file'))
        
        shutil.copy2(source, new_path)
        
        # Purge transactions from the new file
        import sqlite3
        conn = sqlite3.connect(new_path)
        conn.execute("DELETE FROM lines")
        conn.execute("DELETE FROM transactions")
        conn.execute("UPDATE meta SET value=? WHERE key='company_name'", (new_company,))
        conn.execute("UPDATE meta SET value=? WHERE key='fiscal_year'", (year,))
        # Reset any column configs
        conn.execute("DELETE FROM meta WHERE key LIKE 'columns_%'")
        conn.commit()
        conn.close()
        
        flash(f'Created new file for "{new_company}" in {new_folder}/', 'success')
        return redirect(url_for('library'))
    
    # GET: show form
    clients = list_client_books()
    company = models.get_meta('company_name', '') if models.get_db_path() else ''
    return render_template('clone.html', clients=clients, company=company)

# ─── Reports Section ─────────────────────────────────────────────

@app.route('/reports')
def reports_page():
    """Reports landing page."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    company = models.get_meta('company_name', 'My Books')
    fiscal_ye = models.get_meta('fiscal_year_end', '12-31')
    # The year being reported is the one in SETTINGS — the year-end being worked
    # on — not whichever one the wall clock happens to be nearest.
    from datetime import date, timedelta
    anchor = models.fiscal_anchor()
    if anchor:
        fy_begin, fy_end = anchor['cy_start'], anchor['cy_end']
    else:
        today = date.today()
        fy_month, fy_day = [int(x) for x in fiscal_ye.split('-')]
        fye = models.year_end_on(today.year, fy_month, fy_day)
        if fye > today:
            fye = models.year_end_on(today.year - 1, fy_month, fy_day)
        fy_begin = (models.year_end_on(fye.year - 1, fy_month, fy_day) + timedelta(days=1)).isoformat()
        fy_end = fye.isoformat()

    return render_template('reports.html', company=company,
        fiscal_display=date(*(int(x) for x in fy_end.split('-'))).strftime('%d %b %Y'),
        fy_begin=fy_begin, fy_end=fy_end)



# _get_bs_account_ids, _get_report_account_order, _build_account_detail, _fmt_money
# imported from pdf_reports


@app.route('/reports/gl')
def report_gl():
    """Generate General Ledger report."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    
    begin = request.args.get('begin', '')
    end = request.args.get('end', '')
    fmt = request.args.get('fmt', 'pdf')
    dr_cr_filter = request.args.get('filter', 'all')
    company = models.get_meta('company_name', 'My Books')
    
    bs_ids = _get_bs_account_ids()
    
    # Get accounts in BS order then IS order
    accounts = _get_report_account_order('BS') + _get_report_account_order('IS')
    
    if fmt == 'csv':
        return _gl_csv(accounts, bs_ids, begin, end, dr_cr_filter, company)
    else:
        try:
            return _gl_pdf(accounts, bs_ids, begin, end, dr_cr_filter, company)
        except Exception as e:
            flash(f'PDF error: {e}. Install reportlab: pip install reportlab', 'error')
            return redirect(url_for('reports_page'))


def _gl_csv(accounts, bs_ids, begin, end, dr_cr_filter, company):
    """Generate GL as CSV download."""
    import csv, io
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(['Account', 'Description', 'Date', 'Reference', 'Detail', 'Debit', 'Credit', 'Balance', 'Cross Account'])
    
    for aid, aname, adesc in accounts:
        is_bs = aid in bs_ids
        opening, rows, closing = _build_account_detail(aid, aname, adesc, begin, end, is_bs, dr_cr_filter)
        
        if not rows and opening == 0:
            continue  # skip empty accounts
        
        # Opening balance row
        w.writerow([aname, adesc, begin or '', '', 'Opening Balance',
                    _fmt_money(opening) if opening > 0 else '',
                    _fmt_money(-opening) if opening < 0 else '',
                    _fmt_money(opening), ''])
        
        for r in rows:
            w.writerow([aname, adesc, r['date'], r['ref'], r['desc'],
                        _fmt_money(r['debit']) if r['debit'] else '',
                        _fmt_money(r['credit']) if r['credit'] else '',
                        _fmt_money(r['balance']), r['cross']])
        
        # Closing balance row
        w.writerow([aname, adesc, end or '', '', 'Closing Balance', '', '', _fmt_money(closing), ''])
        w.writerow([])  # blank row between accounts
    
    resp = app.make_response(output.getvalue())
    resp.headers['Content-Type'] = 'text/csv'
    resp.headers['Content-Disposition'] = f'attachment; filename=GL_{begin}_{end}.csv'
    return resp


def _gl_pdf(accounts, bs_ids, begin, end, dr_cr_filter, company):
    """Generate GL as monospaced PDF — delegates to pdf_reports.gl_pdf."""
    pdf_bytes = gl_pdf(company, accounts, bs_ids, begin, end, dr_cr_filter)
    resp = app.make_response(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename=GL_{begin}_{end}.pdf'
    return resp


@app.route('/reports/account')
def report_account_detail():
    """Generate single account detail report."""
    if not models.get_db_path():
        return redirect(url_for('library'))
    
    acct_name = request.args.get('account', '').strip().upper()
    begin = request.args.get('begin', '')
    end = request.args.get('end', '')
    fmt = request.args.get('fmt', 'pdf')
    dr_cr_filter = request.args.get('filter', 'all')
    company = models.get_meta('company_name', 'My Books')
    
    acct = models.get_account_by_name(acct_name)
    if not acct:
        flash(f'Account "{acct_name}" not found', 'error')
        return redirect(url_for('reports_page'))
    
    bs_ids = _get_bs_account_ids()
    accounts = [(acct['id'], acct['name'], acct['description'])]
    
    if fmt == 'csv':
        return _gl_csv(accounts, bs_ids, begin, end, dr_cr_filter, company)
    else:
        try:
            return _gl_pdf(accounts, bs_ids, begin, end, dr_cr_filter, company)
        except Exception as e:
            flash(f'PDF error: {e}. Install reportlab: pip install reportlab', 'error')
            return redirect(url_for('reports_page'))


@app.route('/reports/formatted')
def report_formatted():
    """Generate formatted BS or IS report. Placeholder — redirects to print view."""
    report_name = request.args.get('report', 'BS')
    end = request.args.get('end', '')
    begin = request.args.get('begin', '')
    fmt = request.args.get('fmt', 'pdf')
    
    reports = models.get_reports()
    rpt = next((r for r in reports if r['name'] == report_name), None)
    if not rpt:
        flash(f'Report "{report_name}" not found', 'error')
        return redirect(url_for('reports_page'))
    
    # Redirect to existing print view with parameters
    params = f'begin={begin}&end={end}' if begin else f'end={end}'
    return redirect(f'/report/{rpt["id"]}/print?{params}')


# ─── AJE Print Report ─────────────────────────────────────────────

@app.route('/reports/aje/<int:account_id>')
def report_aje(account_id):
    """Generate AJE print report for a single account, grouped by reference."""
    if not models.get_db_path():
        return redirect(url_for('library'))

    acct = models.get_account(account_id)
    if not acct:
        flash('Account not found', 'error')
        return redirect(url_for('home'))

    begin = request.args.get('begin', '')
    end = request.args.get('end', '')
    company = models.get_meta('company_name', 'My Books')

    try:
        return _aje_pdf(account_id, acct, begin, end, company)
    except Exception as e:
        flash(f'PDF error: {e}. Install reportlab: pip install reportlab', 'error')
        return redirect(url_for('account_ledger', account_id=account_id))


def _aje_pdf(account_id, acct, begin, end, company):
    """Generate AJE report as PDF — delegates to pdf_reports.aje_pdf."""
    acct_name = acct['name']
    acct_desc = acct['description'] or ''
    pdf_bytes = aje_pdf(company, account_id, acct_name, acct_desc, begin, end)
    resp = app.make_response(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    fname = f'{acct_name}_{begin}_{end}.pdf' if begin and end else f'{acct_name}.pdf'
    resp.headers['Content-Disposition'] = f'inline; filename={fname}'
    return resp


@app.route('/reports/ledger/<int:account_id>')
def report_ledger_pdf(account_id):
    """Generate single-account ledger PDF. Defaults to current fiscal year."""
    if not models.get_db_path():
        return redirect(url_for('library'))

    acct = models.get_account(account_id)
    if not acct:
        flash('Account not found', 'error')
        return redirect(url_for('home'))

    begin = request.args.get('begin', '')
    end = request.args.get('end', '')

    # Default window = the fiscal year being worked on. NOT the ceiling (which
    # may run a year past it) and not the wall clock.
    if not begin and not end:
        anchor = models.fiscal_anchor()
        if anchor:
            begin, end = anchor['cy_start'], anchor['cy_end']
        else:
            from datetime import timedelta
            fye = models.get_meta('fiscal_year_end', '12-31')
            try:
                fye_m, fye_d = int(fye.split('-')[0]), int(fye.split('-')[1])
            except (ValueError, IndexError):
                fye_m, fye_d = 12, 31
            today = date.today()
            fy_end = models.year_end_on(today.year, fye_m, fye_d)
            if fy_end > today:
                fy_end = models.year_end_on(today.year - 1, fye_m, fye_d)
            begin = (models.year_end_on(fy_end.year - 1, fye_m, fye_d) + timedelta(days=1)).isoformat()
            end = fy_end.isoformat()

    company = models.get_meta('company_name', 'My Books')
    bs_ids = _get_bs_account_ids()
    is_bs = acct['id'] in bs_ids

    try:
        pdf_bytes = account_ledger_pdf(company, account_id, acct['name'],
                                        acct['description'] or '', begin, end, is_bs)
        resp = app.make_response(pdf_bytes)
        resp.headers['Content-Type'] = 'application/pdf'
        fname = f"{acct['name']}_{begin}_{end}.pdf"
        resp.headers['Content-Disposition'] = f'inline; filename={fname}'
        return resp
    except Exception as e:
        flash(f'PDF error: {e}. Install reportlab: pip install reportlab', 'error')
        return redirect(url_for('account_ledger', account_id=account_id))


# ─── Setup Subledgers ──────────────────────────────────────────────

@app.route('/api/setup-detailed-ar', methods=['POST'])
def api_setup_detailed_ar():
    """One-click scaffold for Detailed AR subledger report."""
    if not models.get_db_path():
        return jsonify({'ok': False, 'error': 'No books open'})
    try:
        result = models.setup_detailed_ar()
        return jsonify({'ok': True, 'message': result})
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/setup-detailed-ap', methods=['POST'])
def api_setup_detailed_ap():
    """One-click scaffold for Detailed AP subledger report."""
    if not models.get_db_path():
        return jsonify({'ok': False, 'error': 'No books open'})
    try:
        result = models.setup_detailed_ap()
        return jsonify({'ok': True, 'message': result})
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ─── Entry Point ────────────────────────────────────────────────────

def launch_window(url):
    """Open Grid's window.

    Grid itself is browser-agnostic — it is a localhost web app and ANY browser
    can open it. This function only decides HOW the window is raised: a
    Chromium-family browser (Chrome/Edge/Brave/Chromium/Vivaldi) can give a
    DEDICATED APPLICATION WINDOW with no tabs and no address bar via --app=,
    which is the vDOS feel. Firefox has no equivalent (site-specific browsers
    were removed), so on a Firefox machine this falls through to the OS default
    browser and Grid opens as an ordinary window. Everything works either way;
    only the chrome around it differs. GRID_NO_WINDOW=1 prints the URL instead."""
    import subprocess, platform, webbrowser
    if os.environ.get('GRID_NO_WINDOW'):
        print(f"  Open {url} in your browser")
        return False
    sysname = platform.system()
    quiet = {'stdout': subprocess.DEVNULL, 'stderr': subprocess.DEVNULL}
    try:
        if sysname == 'Darwin':
            for name in ('Google Chrome', 'Microsoft Edge', 'Brave Browser', 'Chromium'):
                if os.path.isdir(f'/Applications/{name}.app'):
                    subprocess.Popen(['open', '-na', name, '--args', f'--app={url}'], **quiet)
                    return True
        elif sysname == 'Windows':
            pf = [os.environ.get('PROGRAMFILES', r'C:\Program Files'),
                  os.environ.get('PROGRAMFILES(X86)', r'C:\Program Files (x86)'),
                  os.environ.get('LOCALAPPDATA', '')]
            rel = [r'Google\Chrome\Application\chrome.exe',
                   r'Microsoft\Edge\Application\msedge.exe',
                   r'BraveSoftware\Brave-Browser\Application\brave.exe']
            for base in pf:
                for r in rel:
                    exe = os.path.join(base, r)
                    if base and os.path.exists(exe):
                        subprocess.Popen([exe, f'--app={url}'], **quiet)
                        return True
        else:
            for exe in ('google-chrome', 'google-chrome-stable', 'chromium',
                        'chromium-browser', 'microsoft-edge', 'brave-browser', 'vivaldi'):
                p = shutil.which(exe)
                if p:
                    subprocess.Popen([p, f'--app={url}'], **quiet)
                    return True
    except Exception:
        pass
    # open_new, not open: open() may re-use the CURRENT tab, which would unload
    # whatever is already sitting in it — including another app's page, whose
    # pagehide beacon would then shut that app down. Ask for our own window.
    webbrowser.open_new(url)
    return False

def main():
    global _server, _last_seen, _boot_time
    import socket
    port = int(os.environ.get('GRID_PORT', '5000'))
    url = f'http://localhost:{port}'

    # One instance at a time — settled BEFORE the books are touched, so a second
    # launch never opens, snapshots, or complains about the lock on a file the
    # first instance is holding. Probe the port rather than letting werkzeug bind
    # it: werkzeug catches its own bind failure, prints "Port N is in use" and
    # exits, so no handler of ours would ever run.
    _probe = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        _probe.bind(('127.0.0.1', port))
    except OSError:
        # v114 — a second double-click means "give me my window back", not "tell
        # me off". Hand the operator a window on the instance that IS running and
        # leave quietly (exit 0, so the launcher closes this console too — a
        # second dead console is what makes people double-click a third time).
        _probe.close()
        print(f"\n  Grid is already running — opening its window.\n")
        launch_window(url)
        sys.exit(0)
    finally:
        try: _probe.close()
        except OSError: pass

    cfg = load_config()

    # This instance records the port it is listening on, so that if it ever
    # crashes or is forgotten, the NEXT launch can knock on it and ask it to
    # close the books properly instead of clearing its lock out from under it.
    models.LOCK_EXTRA['port'] = port

    # If we have a last-opened file, open it automatically
    last = cfg.get('last_opened', '')
    locked = ''
    if last and not os.path.exists(last):
        # A remembered file that is gone (moved library, new machine, a zip
        # extracted over the install) is not an open file. Say so and start
        # at the library — falling through used to reach get_meta() with no
        # books open, which was a TypeError before the window ever appeared.
        print(f"\n  !! Last file not found: {last}")
        print("  Starting at the library instead.")
        last = ''
    if last:
        try:
            models.init_db(last)   # gated open: integrity check → snapshot → migrate
        except models.BooksLocked as e:
            # v130 — a locked file is not a dead end and not a console error.
            # Boot anyway and land the window on the lock screen for THIS file,
            # where F3 gets the operator in. (He was fixing one of these from a
            # phone; a traceback and a .lock file to delete is not a fix.)
            print(f"\n  !! {e.headline} {e.detail}")
            print(f"  Grid will open on the lock screen — F3 there gets you in.")
            locked, last = last, ''
        except ValueError as e:
            print(f"\n  !! {e}")
            print("  Starting at the library instead.")
            last = ''
    if last:
        for line in models.re_repair_note():
            print(f"\n  !! {line}")
        st = models.backup_status()
        if st['error']:
            print(f"  !! {st['note']}")
        company = models.get_meta('company_name', 'My Books')
        print(f"\n  Grid — {company}")
        print(f"  File: {last}")
    elif locked:
        # Don't tell him to pick a client — the window is landing on the file
        # he asked for, with the one key that gets him into it.
        print(f"\n  Grid — {os.path.basename(os.path.dirname(locked))}")
        print(f"  File: {locked}  (locked)")
    else:
        print(f"\n  Grid — Accounting")
        print(f"  No books open. Select a client from the library.")
    
    from werkzeug.serving import make_server
    _server = make_server('127.0.0.1', port, app, threaded=True)

    # A hard kill of the console still owes the books a clean close.
    def _on_signal(signum, frame):
        _formal_close('console closed')
        os._exit(0)
    import signal
    for s in ('SIGINT', 'SIGTERM', 'SIGBREAK', 'SIGHUP'):
        if hasattr(signal, s):
            try:
                signal.signal(getattr(signal, s), _on_signal)
            except (ValueError, OSError):
                pass

    _last_seen = _boot_time = time.time()
    threading.Thread(target=_watchdog, daemon=True).start()

    print(f"  ✕ Close to Library puts the books away — shut the window to stop Grid.\n")
    launch_window((url + '/locked?path=' + quote(locked, safe='')) if locked else url)

    _server.serve_forever()          # returns when the operator closes Grid
    print("  Your books are closed and safe.\n")

if __name__ == '__main__':
    main()
