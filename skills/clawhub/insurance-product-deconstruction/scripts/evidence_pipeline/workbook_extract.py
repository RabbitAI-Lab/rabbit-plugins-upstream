"""Extract complete workbook structure while preserving provenance."""

from __future__ import annotations

import re
from pathlib import Path

UNIT_PATTERN = re.compile(r"(?:单位[：:]?\s*)?([^\n]*(?:元/|人民币|百分比|%)[^\n]*)")


def _column_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(index)


def _filled_merged_values(ws) -> dict[tuple[int, int], object]:
    values: dict[tuple[int, int], object] = {}
    for merged in ws.merged_cells.ranges:
        value = ws.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                values[(row, col)] = value
    return values


def _header_paths(ws, depth: int) -> dict[str, list[str]]:
    merged = _filled_merged_values(ws)
    result = {}
    for col in range(1, ws.max_column + 1):
        path = []
        for row in range(1, min(depth, ws.max_row) + 1):
            value = merged.get((row, col), ws.cell(row, col).value)
            if value not in (None, ""):
                text = str(value).strip()
                if not path or path[-1] != text:
                    path.append(text)
        result[_column_letter(col)] = path
    return result


def _extract_xlsx(path: Path, header_depth: int, max_rows: int = 500) -> dict:
    import openpyxl
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    sheets = []
    for ws in workbook.worksheets:
        total_rows = ws.max_row
        total_cols = ws.max_column
        read_rows = min(total_rows, max_rows)
        rows = [[ws.cell(row, col).value for col in range(1, total_cols + 1)] for row in range(1, read_rows + 1)]
        unit_hints = []
        for row in rows[:20]:
            for value in row:
                if isinstance(value, str) and ("单位" in value or "元/" in value or "%" in value):
                    unit_hints.append(value.strip())
        sheets.append({
            "name": ws.title,
            "hidden": ws.sheet_state != "visible",
            "max_row": total_rows,
            "max_column": total_cols,
            "rows_read": read_rows,
            "hidden_rows": [index for index, dim in ws.row_dimensions.items() if dim.hidden],
            "hidden_columns": [key for key, dim in ws.column_dimensions.items() if dim.hidden],
            "merged_ranges": [str(item) for item in ws.merged_cells.ranges],
            "header_paths": _header_paths(ws, header_depth),
            "unit_hints": unit_hints,
            "rows": rows,
        })
    workbook.close()
    return {"path": str(path), "format": "xlsx", "sheets": sheets}


def _extract_xls(path: Path, header_depth: int, max_rows: int = 500) -> dict:
    import xlrd
    workbook = xlrd.open_workbook(str(path), formatting_info=True)
    sheets = []
    for ws in workbook.sheets():
        total_rows = ws.nrows
        read_rows = min(total_rows, max_rows)
        rows = [ws.row_values(index) for index in range(read_rows)]
        unit_hints = [
            str(value).strip()
            for row in rows[:20]
            for value in row
            if isinstance(value, str) and ("单位" in value or "元/" in value or "%" in value)
        ]
        headers = {}
        for col in range(ws.ncols):
            path_values = []
            for row in range(min(header_depth, read_rows)):
                value = ws.cell_value(row, col)
                if value not in (None, ""):
                    text = str(value).strip()
                    if not path_values or path_values[-1] != text:
                        path_values.append(text)
            headers[_column_letter(col + 1)] = path_values
        sheets.append({
            "name": ws.name,
            "hidden": ws.visibility != 0,
            "max_row": total_rows,
            "max_column": ws.ncols,
            "rows_read": read_rows,
            "hidden_rows": [index + 1 for index, info in enumerate(ws.rowinfo_map.values()) if getattr(info, "hidden", False)],
            "hidden_columns": [],
            "merged_ranges": [f"{rlo + 1}:{rhi},{clo + 1}:{chi}" for rlo, rhi, clo, chi in ws.merged_cells],
            "header_paths": headers,
            "unit_hints": unit_hints,
            "rows": rows,
        })
    return {"path": str(path), "format": "xls", "sheets": sheets}


def extract_workbook(path: Path, *, header_depth: int = 5, max_rows: int = 500) -> dict:
    path = Path(path).resolve()
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _extract_xlsx(path, header_depth, max_rows)
    if suffix == ".xls":
        return _extract_xls(path, header_depth, max_rows)
    raise ValueError(f"Unsupported workbook format: {suffix}")
