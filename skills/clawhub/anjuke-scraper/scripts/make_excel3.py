#!/usr/bin/env python3
"""Step 5: 合并数据生成 Excel 对比表（过滤租金、按 距离→房租 排序、14 列）
用法: python3 make_excel3.py [--config path/to/config.json]
"""
import json, re, math, os, pathlib, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = pathlib.Path(__file__).resolve().parent

def load_config():
    """查找 config.json：--config 参数 > 当前目录 > 脚本目录 > skill 根目录"""
    cfg_path = None
    if len(sys.argv) >= 3 and sys.argv[1] == "--config":
        cfg_path = pathlib.Path(sys.argv[2])
    else:
        for cand in [pathlib.Path.cwd() / "config.json",
                     BASE / "config.json",
                     BASE.parent / "config.json"]:
            if cand.exists():
                cfg_path = cand
                break
    if cfg_path is None:
        sys.exit("未找到 config.json：请复制 config.example.json 为 config.json 后重试，或用 --config 指定路径")
    cfg = json.load(open(cfg_path, encoding="utf-8"))
    cfg.setdefault("output_dir", str(cfg_path.parent / "output"))
    os.makedirs(cfg["output_dir"], exist_ok=True)
    return cfg

CFG = load_config()
OUT = CFG["output_dir"].rstrip("/") + "/"
MAX_RENT = CFG.get("max_rent", 2000)  # 预算上限

records = json.load(open(OUT + "final_records.json"))
details = json.load(open(OUT + "detail_results.json"))
coords_data = json.load(open(OUT + "coords.json"))
company = coords_data['company']
ccomms = coords_data['communities']

COMP = (float(company['lat']), float(company['lng']))

def hav(lat1, lon1, lat2, lon2):
    """haversine 直线距离（km）"""
    R = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return R * 2 * math.asin(math.sqrt(a))

def dist_to_company(comm):
    c = ccomms.get(comm) if comm else None
    if not c or not c.get('lng'):
        return None
    try:
        return hav(COMP[0], COMP[1], float(c['lat']), float(c['lng']))
    except Exception:
        return None

def clean_title(t):
    return re.sub(r'[\ue000-\uf8ff]', '', t or '').strip()

rows = []
removed_over_budget = 0
for r in records:
    det = details.get(r['id'], {})
    if det.get('status') != 'ok':
        det = {}
    rm = r.get('rent_min'); rx = r.get('rent_max')
    # 租金过滤：固定价 >= 上限 删；区间价上限 >= 上限 删
    if rx is not None and rx >= MAX_RENT:
        removed_over_budget += 1
        continue
    if rm is None:
        continue
    community = r.get('community') or det.get('community_exact', '')
    d = dist_to_company(community)
    phone = r.get('phone') or det.get('phone_detail') or ''
    if not phone:
        phone = '微信扫码联系'
    pay = det.get('pay_way') or ' | '.join(r.get('pay', []))
    addr_parts = [community, r.get('district', '')]
    if det.get('address'):
        addr_parts.append(det['address'])
    address = ' '.join([x for x in addr_parts if x])
    model = det.get('model_full') or r.get('model', '')
    area = det.get('area_exact') or r.get('area', '')
    room = f"{model} {area}".strip()
    rent = f"{int(rm) if rm else ''}" + (f"-{int(rx)}" if rx and rx != rm else '')
    rows.append({
        '_dist': d if d is not None else 999,
        '_rent': float(rm) if rm else 99999,
        '地址': address,
        '距目标(km)': round(d, 1) if d is not None else '',
        '房租(元/月)': rent,
        '交付类型': pay if pay else '',
        '房型': room,
        '联系人': r.get('contact', ''),
        '电话': phone,
        '标题': clean_title(r.get('title', '')),
        '朝向': r.get('orientation', ''),
        '楼层': det.get('floor_full') or r.get('floor', ''),
        '装修': det.get('decoration', ''),
        '小区': community,
        '区域': det.get('district_exact') or r.get('district', ''),
        '链接': r['url'],
    })

# 排序：距离(升) → 房租(升)
rows.sort(key=lambda x: (x['_dist'], x['_rent']))
for r in rows:
    r.pop('_dist'); r.pop('_rent')

wb = Workbook()
ws = wb.active
ws.title = f"房源({MAX_RENT}以内)"

headers = ['地址', '距目标(km)', '房租(元/月)', '交付类型', '房型', '联系人', '电话', '标题', '朝向', '楼层', '装修', '小区', '区域', '链接']
ws.append(headers)

head_fill = PatternFill('solid', fgColor='4472C4')
head_font = Font(bold=True, color='FFFFFF', size=11)
thin = Border(*[Side(style='thin', color='BBBBBB')]*4)
for ci in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=ci)
    cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin

for r in rows:
    ws.append([r[h] for h in headers])

widths = [38, 12, 14, 12, 18, 10, 16, 40, 8, 12, 10, 14, 16, 50]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.border = thin
        cell.alignment = Alignment(vertical='center', wrap_text=(cell.column in (1, 8)))

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

for i in range(2, ws.max_row + 1):
    v = ws.cell(row=i, column=2).value
    if isinstance(v, (int, float)):
        if v <= 2.0:
            ws.cell(row=i, column=2).fill = PatternFill('solid', fgColor='C6EFCE')
        elif v <= 5.0:
            ws.cell(row=i, column=2).fill = PatternFill('solid', fgColor='FFEB9C')

out_path = OUT + f"{CFG['keyword']}附近房源.xlsx"
wb.save(out_path)
print(f"已生成: {out_path}")
print(f"过滤掉 >= {MAX_RENT}元: {removed_over_budget} 套")
print(f"剩余房源: {len(rows)} 套")
print(f"有距离: {sum(1 for r in rows if isinstance(r['距目标(km)'], (int,float)))} | 明文电话: {sum(1 for r in rows if r['电话'].startswith('1'))} | 有交付类型: {sum(1 for r in rows if r['交付类型'])}")
print("\n=== 前15条(距离→房租) ===")
for r in rows[:15]:
    print(f"{r['距目标(km)']}km | {r['小区']} | {r['房租(元/月)']}元 | {r['交付类型']} | {r['房型'][:14]} | {r['联系人']} | {r['电话']}")
