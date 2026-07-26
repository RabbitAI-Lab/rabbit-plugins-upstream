#!/usr/bin/env python3
"""Excel表格解析脚本 - 提取生产周报数据"""
import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(json.dumps({"status": "error", "message": "openpyxl未安装"}))
    sys.exit(1)


def parse_xlsx(file_path: str, sheet_name: str = None) -> dict:
    """解析Excel表格，提取生产周报相关数据"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return {"status": "error", "message": f"无法打开文件: {str(e)}"}

    result = {
        "status": "success",
        "source": "excel",
        "file": file_path,
        "data": {
            "overview": "",
            "production_metrics": [],
            "completed_items": [],
            "issues": [],
            "next_week_plan": [],
            "pending_matters": []
        }
    }

    # 使用指定sheet或第一个sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # 扫描所有单元格，识别数据结构
    metrics_keywords = ["产量", "良率", "OEE", "交付", "效率", "不良", "报废", "停机"]
    section_keywords = {
        "completed": ["完成", "已完"],
        "issues": ["异常", "问题", "故障", "不良"],
        "plan": ["计划", "下周"],
        "pending": ["协调", "支持", "需"]
    }

    # 收集指标数据
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100)):
        cells = [str(cell.value).strip() if cell.value else "" for cell in row]
        
        # 检测是否为指标行
        if any(keyword in cells[0] for keyword in metrics_keywords) if cells[0] else False:
            if len(cells) >= 2:
                result["data"]["production_metrics"].append({
                    "name": cells[0],
                    "value": cells[1],
                    "unit": cells[2] if len(cells) > 2 else ""
                })
        
        # 检测章节行
        for section, keywords in section_keywords.items():
            if cells[0] and any(keyword in cells[0] for keyword in keywords):
                # 收集该章节后续内容
                for r in ws.iter_rows(min_row=row[0].row + 1, max_row=row[0].row + 10):
                    row_data = [str(cell.value).strip() if cell.value else "" for cell in r]
                    if row_data[0] and not any(k in row_data[0] for k in ["产量", "良率", "OEE", "完成", "问题", "计划", "协调"]):
                        if isinstance(result["data"][section], list):
                            result["data"][section].append(row_data[0])
                    else:
                        break

    # 尝试从表头识别标准格式
    for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row)):
        cells = [str(cell.value).strip() if cell.value else "" for cell in row]
        # 识别周次、日期等元信息
        for cell in cells:
            if "W" in cell and any(c.isdigit() for c in cell):
                result["data"]["week_id"] = cell
            if "-" in cell and len(cell) == 10:
                result["data"]["date"] = cell

    return result


def main():
    parser = argparse.ArgumentParser(description="解析Excel表格提取生产周报数据")
    parser.add_argument("--file", required=True, help="Excel文件路径")
    parser.add_argument("--sheet", required=False, help="指定Sheet名称（可选）")
    args = parser.parse_args()

    if not Path(args.file).exists():
        print(json.dumps({"status": "error", "message": f"文件不存在: {args.file}"}))
        sys.exit(1)

    result = parse_xlsx(args.file, args.sheet)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
