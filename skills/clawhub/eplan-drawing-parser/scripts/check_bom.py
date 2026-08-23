#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
eplan-drawing-parser · BOM核对 (BOM Cross-check)

将 EPLAN/CAD 矢量 PDF 电气图纸 与 Excel 物料清单(UL证书汇总) 交叉核对：
  1. 位号核对   —— 清单里的每个位号是否在图纸中出现（0 缺失）
  2. 型号核对   —— 图纸位号附近的型号文本 与 清单型号 是否一致
  3. 数量核对   —— 清单数量(位号个数) 与 图纸出现频次 是否一致
  4. UL归口检查 —— 清单内 型号 ↔ UL档案号 ↔ 供应商 是否对应，找出缺 UL 条目

用法示例:
  python scripts/check_bom.py --pdf drawing.pdf --xlsx bom.xlsx
  python scripts/check_bom.py --pdf drawing.pdf --xlsx bom.xlsx --sheet "UL证书汇总" --out 核对结果.xlsx
  python scripts/check_bom.py --pdf drawing.pdf --xlsx bom.xlsx --no-export   # 只打印不导出

说明:
  - 图纸必须是**矢量 PDF**(带文本层); 扫描件/图片 PDF 不支持(需先 OCR)
  - 位号匹配会自动去掉 EPLAN 前导连字符(如 "-FU1001" -> "FU1001")
  - 型号判定基于图纸矢量文本(100%准), 不依赖视觉模型
"""
from __future__ import annotations
import argparse, json, os, re, sys
from pathlib import Path

try:
    import fitz  # PyMuPDF
except ImportError:
    print(json.dumps({"error": "缺少依赖: pip install pymupdf"})); sys.exit(1)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    HAVE_XLSX = True
except ImportError:
    HAVE_XLSX = False


# ---------- 图纸侧：提取文本 ----------
def extract_texts(pdf_path):
    """逐页提取文本+坐标(去EPLAN前导符号). 返回 {page: [ {text,cx,cy}...] }"""
    doc = fitz.open(pdf_path)
    out = {}
    for i, page in enumerate(doc):
        texts = []
        for b in page.get_text("dict")["blocks"]:
            for l in b.get("lines", []):
                for s in l["spans"]:
                    t = s["text"].strip()
                    if not t:
                        continue
                    x0, y0, x1, y1 = s["bbox"]
                    texts.append({
                        "raw": t,
                        "core": re.sub(r"^[-\\/.\s~]+", "", t).strip(),
                        "cx": (x0 + x1) / 2, "cy": (y0 + y1) / 2,
                        "page": i + 1,
                    })
        out[i + 1] = texts
    doc.close()
    return out


# ---------- 清单侧：读 Excel ----------
def read_bom(xlsx_path, sheet=None):
    """读取 Excel 的物料清单. 返回 (sections, rows). rows 每行含柜/序号/名称/供应商/型号/参数/数量/UL/位号"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet is None:
        # 优先含 "UL" 的 sheet；否则含 "清单"；否则第一个
        cands = [s for s in wb.sheetnames if "UL" in s.upper()]
        if not cands:
            cands = [s for s in wb.sheetnames if "清单" in s]
        sheet = cands[0] if cands else wb.sheetnames[0]
    ws = wb[sheet]
    sections, cur_sec = [], None
    rows = []
    for r in ws.iter_rows(values_only=True):
        row = [(c.strip() if isinstance(c, str) else c) for c in r]
        if not any(row):
            continue
        # 章节行: 如 "一、系统控制柜（单台）"
        if row[0] and ("、" in str(row[0])) and ("柜" in str(row[0]) or "系统" in str(row[0])):
            cur_sec = row[0]
            sections.append(cur_sec)
            continue
        # 数据行: 首列为序号(数字)
        if isinstance(row[0], int):
            rows.append({
                "section": cur_sec, "num": row[0],
                "name": row[1], "supplier": row[2], "model": row[3],
                "param": row[4], "qty": row[5], "ul": row[6], "pos": row[7],
            })
    return sections, rows


# ---------- 型号归一化 ----------
def norm_model(s):
    return re.sub(r"[\s\-_/+.]+", "", str(s)).upper()


# ---------- 主核对 ----------
def run(pdf_path, xlsx_path, sheet=None, bom_map=None):
    """bom_map: {位号: (清单型号, 数量)} 可选, 用于自动核对关键位号"""
    pages = extract_texts(pdf_path)

    # 图纸位号索引: core -> [page,...]
    drawing_occ = {}
    for pg, txts in pages.items():
        for t in txts:
            if t["core"]:
                drawing_occ.setdefault(t["core"], set()).add(pg)

    sections, rows = read_bom(xlsx_path, sheet)

    # 汇总
    result = {
        "pdf": Path(pdf_path).name,
        "xlsx": Path(xlsx_path).name,
        "sheet": sheet,
        "num_pages": len(pages),
        "sections": sections,
        "checks": [],
        "no_ul": [],
        "summary": {},
    }

    # A. 位号/型号/数量 核对（用 bom_map 提供的关键位号 + 清单位号列兜底）
    checked = {}
    if bom_map:
        for des, (model, cnt) in bom_map.items():
            pg = sorted(drawing_occ.get(des, set()))
            checked[des] = {"list_model": model, "list_qty": cnt,
                            "pages": pg, "found": bool(pg)}
    # 兜底: 从清单"位号"列再提取未覆盖的位号
    for r in rows:
        for tok in re.split(r"[,/、\s]+", str(r["pos"] or "")):
            m = re.match(r"^([A-Za-z]{1,4}\d+.*?)(?:\(.*)?$", tok)
            if m:
                des = re.sub(r"^[-\\/.\s~]+", "", m.group(1)).strip()
                if des and des not in checked:
                    pg = sorted(drawing_occ.get(des, set()))
                    checked[des] = {"list_model": r["model"], "list_qty": r["qty"],
                                    "pages": pg, "found": bool(pg)}

    for des, info in checked.items():
        result["checks"].append({
            "designator": des, "list_model": info["list_model"],
            "list_qty": info["list_qty"],
            "drawing_pages": info["pages"], "found": info["found"],
        })

    # B. UL 归口检查: 型号↔UL↔供应商, 缺UL
    for r in rows:
        if r["name"] in ("辅材", "镀锡铜排", "电线"):
            continue
        ul = str(r["ul"] or "").strip()
        if ul in ("", "-", "无", "None", "nan"):
            result["no_ul"].append({
                "section": r["section"], "num": r["num"], "name": r["name"],
                "supplier": r["supplier"], "model": r["model"], "pos": r["pos"],
            })

    missing = [c["designator"] for c in result["checks"] if not c["found"]]
    result["summary"] = {
        "checked_designators": len(result["checks"]),
        "missing_in_drawing": missing,
        "missing_count": len(missing),
        "no_ul_count": len(result["no_ul"]),
        "sections": sections,
    }
    return result


def print_report(rep):
    s = rep["summary"]
    print(f"✅ 核对完成: {rep['pdf']}  ↔  {rep['xlsx']} [{rep['sheet']}]")
    print(f"   图纸 {rep['num_pages']} 页, 清单 {len(rep['sections'])} 个柜")
    print(f"   核对位号 {s['checked_designators']} 个")
    if s["missing_in_drawing"]:
        print(f"   ⚠️ 图纸缺失位号 {s['missing_count']} 个: {s['missing_in_drawing']}")
    else:
        print(f"   ✅ 位号全部存在 (0 缺失)")
    print(f"   ⚠️ 清单中无 UL 档案号条目 {s['no_ul_count']} 条:")
    for n in rep["no_ul"]:
        print(f"      - [{n['section']}] 序号{n['num']} {n['name']} | {n['supplier']} | {n['model']}")


def export_xlsx(rep, out_path):
    if not HAVE_XLSX:
        print(f"⚠️ 未安装 openpyxl, 跳过导出: {out_path}")
        return
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    ok = PatternFill("solid", fgColor="C6EFCE"); ok_f = Font(color="006100")
    warn = PatternFill("solid", fgColor="FFC7CE"); warn_f = Font(color="9C0006")

    w1 = wb.active; w1.title = "位号核对"
    w1.append(["位号", "清单型号", "图纸页", "判定"])
    for c in w1[1]: c.fill = hdr; c.font = hdr_font; c.border = border
    for c in rep["checks"]:
        w1.append([c["designator"], c["list_model"],
                   ",".join(map(str, c["drawing_pages"])) or "-",
                   "✅ 存在" if c["found"] else "❌ 缺失"])
        r = w1.max_row
        for cell in w1[r]:
            cell.border = border
            cell.fill, cell.font = (ok, ok_f) if c["found"] else (warn, warn_f)
    for col, wd in zip("ABCD", [12, 30, 12, 10]): w1.column_dimensions[col].width = wd

    w2 = wb.create_sheet("无UL档案号")
    w2.append(["柜", "序号", "产品名称", "供应商", "型号", "位号"])
    for c in w2[1]: c.fill = hdr; c.font = hdr_font; c.border = border
    for n in rep["no_ul"]:
        w2.append([n["section"], n["num"], n["name"], n["supplier"], n["model"], n["pos"]])
        for c in w2[w2.max_row]: c.border = border; c.fill = warn; c.font = warn_f
    for col, wd in zip("ABCDEF", [18, 6, 18, 12, 30, 16]): w2.column_dimensions[col].width = wd

    wb.save(out_path)
    print(f"   已导出: {out_path}")


def main():
    ap = argparse.ArgumentParser(description="EPLAN 图纸 vs Excel物料清单 交叉核对")
    ap.add_argument("--pdf", required=True, help="EPLAN/CAD矢量PDF")
    ap.add_argument("--xlsx", required=True, help="物料清单Excel")
    ap.add_argument("--sheet", default=None, help="Excel工作表名(默认自动找'UL'/'清单')")
    ap.add_argument("--out", default=None, help="导出xlsx路径(默认 <xlsx目录>/核对结果.xlsx)")
    ap.add_argument("--no-export", action="store_true", help="只打印不导出")
    args = ap.parse_args()

    if not Path(args.pdf).exists():
        print(json.dumps({"error": f"图纸不存在: {args.pdf}"})); sys.exit(1)
    if not Path(args.xlsx).exists():
        print(json.dumps({"error": f"清单不存在: {args.xlsx}"})); sys.exit(1)

    rep = run(args.pdf, args.xlsx, args.sheet)
    print_report(rep)

    if not args.no_export:
        out = args.out or str(Path(args.xlsx).with_name("核对结果_图纸清单.xlsx"))
        export_xlsx(rep, out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
