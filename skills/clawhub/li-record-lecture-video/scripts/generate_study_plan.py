# -*- coding: utf-8 -*-
"""
录播课学习计划生成器 - v2.0.2 通用化描述+三国语言同步
支持三种模式:
  1. 用户自定义模板 — --template my_template.xlsx
  2. 内置 11 列默认模板 — 无 --template 时自动使用
  3. 命令行快捷模式 — --input videos.xlsx --course-name "课程名"

安全约束:
  - 输出路径不允许路径遍历（../ 等）
  - 禁止写入系统关键目录
  - 视频标题自动脱敏（防 Excel 公式注入）
  - 输入参数范围校验（每日上限 0.25-16h, 最多 500 个视频）
  - 默认文案已通用化，不含任何平台/机构特定信息

用法:
  python generate_study_plan.py --config config.json
  python generate_study_plan.py --config config.json --template my_template.xlsx
  python generate_study_plan.py --input videos.xlsx --course-name "课程名" --output out.xlsx

config.json 完整结构见 references/template-spec.md
"""

import argparse, json, os, re, sys
from datetime import timedelta

try:
    import openpyxl
except ImportError:
    print("[ERROR] 需要 openpyxl。安装: pip install openpyxl")
    sys.exit(1)

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# v2.0: 模板引擎 — 确保同目录模块可导入
_script_dir = os.path.dirname(os.path.abspath(__file__))
if _script_dir not in sys.path:
    sys.path.insert(0, _script_dir)
import template_engine


# ============================================================
# 安全常量
# ============================================================
MAX_VIDEOS = 500            # 单次最多处理视频数
MAX_TITLE_LEN = 300         # 视频标题最大字符数
MIN_DAILY_HOURS = 0.25      # 每日最少学习时长
MAX_DAILY_HOURS = 16.0      # 每日最多学习时长

# 禁止写入的系统关键目录（Windows 和 Unix）
FORBIDDEN_OUTPUT_PREFIXES = [
    # Windows
    'C:\\Windows', 'C:\\Windows\\System32', 'C:\\Program Files',
    'C:\\Program Files (x86)',
    # Unix
    '/etc', '/bin', '/sbin', '/usr/bin', '/usr/sbin',
    '/boot', '/dev', '/proc', '/sys', '/root',
    # macOS
    '/System', '/Library/System',
]

# Excel 公式注入危险前缀
FORMULA_INJECTION_PREFIXES = ('=', '+', '-', '@', '\t', '\r', '\n')


# ============================================================
# 默认文案模板 — 已通用化，不包含任何平台/机构特定信息
# ============================================================
DEFAULT_REGISTRATION = (
    "1、注册学习平台账号，观看复习视频（学习网址由班主任提供）；\n"
    "2、开课当周班主任添加联系方式并邀请进入班级群，获取学习计划和电子学习资料；\n"
    "3、表中「学习内容」列编号为考试大纲编号。"
)

DEFAULT_EXAM_INFO = (
    "1，考试信息：\n"
    "培训结束后由班主任在班级群内同步考试券申请链接，按官方流程申请；\n"
    "2，预约考试：\n"
    "拿到考试券后，登录考试平台自主预约线上或考点考试；\n"
    "3，取消考试规则：\n"
    "按官方平台规则，考前48小时外可免费改期，48小时内不可取消；\n"
    "4，备考资料：\n"
    "①课程讲义 ②章节练习题 ③课程视频 ④官方样题。"
)

DEFAULT_EXAM_PHASE = (
    "1，考试时间：\n"
    "在自己预约的考试日期当天，至少提前15分钟到达考试中心；\n"
    "2，考试当天需携带两个有效证件（原件）：一类＋一类/一类＋二类\n"
    "一类（身份证、护照、驾驶证）；\n"
    "二类（社保卡、居住证、信用卡、工作证、学生证；此类证件上需有姓名加照片）。"
)

DEFAULT_SCORE_INFO = (
    "1，成绩公布时间：考完一周左右公布成绩（可关注官方机构网站）；\n"
    "2，证书下载：电子证书可在官方网站下载；\n"
    "3，证书有效期：以官方机构规定为准。"
)

DEFAULT_LEARNING_METHOD = (
    "1、严格执行每日学习计划，提高备考效率；\n"
    "2、视频学习时长可根据个人习惯调整播放速度；\n"
    "3、重点关注老师讲解的重难点及典型题型；\n"
    "4、回顾重点知识，对照章节课后题检查掌握情况；\n"
    "5、遇到课程相关疑问及时向老师寻求指导。"
)

DEFAULT_LEARNING_MATERIALS = (
    "1、课程讲义\n"
    "2、章节练习题\n"
    "3、课程资料包"
)

DEFAULT_HOMEWORK = "完成本章节课后练习题"

# ============================================================
# 样式常量
# ============================================================
FONT_TITLE  = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
FONT_STAGE  = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
FONT_HEADER = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
FONT_BODY   = Font(name='微软雅黑', size=10)
FONT_BOLD   = Font(name='微软雅黑', size=10, bold=True)
FONT_NOTE   = Font(name='微软雅黑', size=10, color='333333')

FILL_TITLE   = PatternFill('solid', fgColor='1F4E78')
FILL_STAGE   = PatternFill('solid', fgColor='4472C4')
FILL_HEADER  = PatternFill('solid', fgColor='5B9BD5')
FILL_SECTION = PatternFill('solid', fgColor='D6DCE4')
FILL_NOTE    = PatternFill('solid', fgColor='F2F2F2')

ALIGN_C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LC = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_LT = Alignment(horizontal='left',   vertical='top',    wrap_text=True)

THIN  = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = ['时间安排', '学习时长（H）', '学习章节', '学习目标',
           '目标验证方式', '学习内容', '项目实战', '综合项目实战',
           '课后作业', '学习要求及方法', '学习资料']

LAST_COL = 12  # A=1 留白, B-L 共11列
COL_WIDTHS = [3, 14, 13, 44, 38, 34, 44, 16, 16, 16, 18, 18]

# ============================================================
# 安全校验函数
# ============================================================
def sanitize_title(title: str) -> str:
    """防 Excel 公式注入：对以 = + - @ 开头的标题加单引号前缀"""
    if not title:
        return title
    t = title.strip()
    if t and t[0] in FORMULA_INJECTION_PREFIXES:
        return "'" + t
    return t


def validate_output_path(path: str) -> str:
    """校验输出路径安全：禁止路径遍历和系统目录写入"""
    # 规范化路径
    abs_path = os.path.abspath(path)

    # 禁止路径遍历
    if '..' in path.replace('\\', '/'):
        raise ValueError(f"输出路径不允许包含 '..': {path}")

    # 禁止写入系统目录
    for prefix in FORBIDDEN_OUTPUT_PREFIXES:
        if abs_path.lower().startswith(prefix.lower()):
            raise ValueError(f"禁止写入系统目录: {abs_path} (匹配 {prefix})")

    # 必须是 .xlsx 扩展名
    if not abs_path.lower().endswith('.xlsx'):
        raise ValueError(f"输出文件必须是 .xlsx 格式: {path}")

    return abs_path


def validate_config(config: dict):
    """校验配置参数的合法性和安全性"""
    errors = []

    # course_name 必填且非空
    course_name = config.get('course_name', '')
    if not course_name or not str(course_name).strip():
        errors.append("course_name 不能为空")
    elif len(str(course_name)) > 200:
        errors.append("course_name 不能超过200字符")

    # daily_hours 范围
    dh = config.get('daily_hours', 2)
    if not isinstance(dh, (int, float)) or dh < MIN_DAILY_HOURS or dh > MAX_DAILY_HOURS:
        errors.append(f"daily_hours 必须在 {MIN_DAILY_HOURS}-{MAX_DAILY_HOURS} 之间，当前值: {dh}")

    # 视频数据
    videos = []
    if config.get('videos'):
        videos = config['videos']
    elif config.get('input_excel'):
        inp = config['input_excel']
        inp_path = inp.get('path', '')
        if not inp_path or not os.path.exists(inp_path):
            errors.append(f"input_excel.path 文件不存在: {inp_path}")
    else:
        errors.append("必须提供 videos 或 input_excel")

    if videos and len(videos) > MAX_VIDEOS:
        errors.append(f"视频数量超过上限 {MAX_VIDEOS}，当前: {len(videos)}")

    for i, v in enumerate(videos):
        if len(v) != 2:
            errors.append(f"videos[{i}] 格式错误，应为 [标题, 分钟数]")
            continue
        title, minutes = v
        title_str = str(title) if title else ''
        if len(title_str) > MAX_TITLE_LEN:
            errors.append(f"视频标题过长 ({len(title_str)}/{MAX_TITLE_LEN}): {title_str[:50]}...")
        if not isinstance(minutes, (int, float)) or minutes <= 0:
            errors.append(f"videos[{i}] 时长无效: {minutes}")

    # output 路径
    output = config.get('output', 'output/学习计划.xlsx')
    try:
        validate_output_path(output)
    except ValueError as e:
        errors.append(str(e))

    if errors:
        raise ValueError("配置校验失败:\n  - " + "\n  - ".join(errors))


# ============================================================
# 工具函数
# ============================================================
def parse_duration(val):
    """将 openpyxl 单元格值转为分钟数（安全版：限制范围）"""
    if isinstance(val, timedelta):
        mins = val.total_seconds() / 60.0
    elif isinstance(val, (int, float)):
        mins = float(val) * 1440 if 0 < val < 1 else float(val) / 60.0
    else:
        return 0.0
    # 单视频时长上限 24h
    return min(max(mins, 0), 1440)


def load_videos_from_excel(path, sheet_name=None, title_col=1, dur_col=2, header_row=1):
    """从 Excel 安全读取视频清单（限制列号和行数）"""
    # 校验列范围
    if not (1 <= title_col <= 100) or not (1 <= dur_col <= 100):
        raise ValueError(f"列号超出范围: title_col={title_col}, dur_col={dur_col}")
    if header_row < 1 or header_row > 100:
        raise ValueError(f"表头行号超出范围: {header_row}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"工作表不存在: {sheet_name}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    videos = []
    count = 0
    for r in range(header_row + 1, ws.max_row + 1):
        count += 1
        if count > MAX_VIDEOS:
            print(f"[WARN] 视频数量超过上限 {MAX_VIDEOS}，已截断")
            break
        title = ws.cell(row=r, column=title_col).value
        dur   = ws.cell(row=r, column=dur_col).value
        if title:
            title_str = sanitize_title(str(title).strip()[:MAX_TITLE_LEN])
            minutes = parse_duration(dur)
            videos.append((title_str, minutes))
    wb.close()
    return videos


def group_by_daily_limit(videos, daily_hours):
    """将视频按每天上限（小时）分组，返回 [[(title, min), ...], ...]"""
    limit_min = daily_hours * 60
    groups, current, current_min = [], [], 0
    for title, m in videos:
        if current and current_min + m > limit_min:
            groups.append(current)
            current, current_min = [(title, m)], m
        else:
            current.append((title, m))
            current_min += m
    if current:
        groups.append(current)
    return groups


def merge_and_write(ws, row, col_s, col_e, value, font, fill, alignment, height):
    """合并单元格并安全写入（value 脱敏处理）"""
    safe_val = value
    if isinstance(safe_val, str):
        safe_val = sanitize_title(safe_val)
    ws.merge_cells(start_row=row, start_column=col_s, end_row=row, end_column=col_e)
    cell = ws.cell(row=row, column=col_s, value=safe_val)
    cell.font = font
    if fill is not None:
        cell.fill = fill
    cell.alignment = alignment
    ws.row_dimensions[row].height = height


def write_row(ws, row, values, font=None, fill=None, alignment=None, height=60, borders=True):
    """安全写入一行数据（values 从 B 列开始，自动脱敏标题）"""
    f = font or FONT_BODY
    a = alignment or ALIGN_LT
    for ci, v in enumerate(values, 2):
        safe_v = v
        if ci == 4 and isinstance(safe_v, str):
            # 学习章节列：防公式注入
            safe_v = sanitize_title(safe_v)
        cell = ws.cell(row=row, column=ci, value=safe_v)
        cell.font = f
        cell.alignment = a if ci >= 4 else ALIGN_C
        if borders:
            cell.border = BORDER
    ws.row_dimensions[row].height = height


# ============================================================
# 主生成逻辑
# ============================================================
def generate(config):
    """主入口：根据 config 安全生成 Excel"""
    # ---------- 校验 ----------
    validate_config(config)

    # ---------- 解析视频 ----------
    videos = []
    if config.get('videos'):
        videos_raw = config['videos']
        for v in videos_raw:
            title = sanitize_title(str(v[0]).strip()[:MAX_TITLE_LEN])
            mins = float(v[1])
            if mins <= 0 or mins > 1440:
                raise ValueError(f"视频时长无效 ({mins}分钟): {title}")
            videos.append((title, mins))
    elif config.get('input_excel'):
        inp = config['input_excel']
        videos = load_videos_from_excel(
            inp['path'],
            sheet_name=inp.get('sheet_name'),
            title_col=inp.get('title_col', 1),
            dur_col=inp.get('dur_col', 2),
            header_row=inp.get('header_row', 1)
        )
    else:
        raise ValueError("config 中必须提供 videos 或 input_excel")

    if not videos:
        raise ValueError("视频清单为空")

    total_min = sum(m for _, m in videos)
    total_h = total_min / 60
    print(f"共 {len(videos)} 个视频，总时长 {total_h:.2f}h ({total_min:.0f}min)")

    # ---------- 分组 ----------
    daily_hours = config.get('daily_hours', 2)
    groups = group_by_daily_limit(videos, daily_hours)
    print(f"按每天 ≤{daily_hours}h 分为 {len(groups)} 天")
    for i, grp in enumerate(groups, 1):
        gm = sum(m for _, m in grp)
        print(f"  第{i}天: {len(grp)}个视频, {gm:.0f}min ({gm/60:.2f}h)")

    # ---------- 参数 ----------
    course_name  = str(config.get('course_name', '课程录播课学习计划'))[:200]
    phase_name   = str(config.get('phase_name', '精讲阶段'))[:100]
    phase_dates  = str(config.get('phase_dates', 'X月XX日-X月XX日'))[:100]
    reg_info     = str(config.get('registration_info', DEFAULT_REGISTRATION))
    exam_info    = str(config.get('exam_info', DEFAULT_EXAM_INFO))
    exam_phase   = str(config.get('exam_phase', DEFAULT_EXAM_PHASE))
    score_info   = str(config.get('score_info', DEFAULT_SCORE_INFO))
    lm_method    = str(config.get('learning_method', DEFAULT_LEARNING_METHOD))
    lm_materials = str(config.get('learning_materials', DEFAULT_LEARNING_MATERIALS))
    homework     = str(config.get('homework', DEFAULT_HOMEWORK))[:200]
    guide_map    = config.get('guide', {})

    # --- 学习指引覆盖率检查 ---
    if guide_map:
        missing = [t for t, _ in videos if t not in guide_map]
        if missing:
            print(f"[WARN] 以下 {len(missing)} 个视频缺少学习指引，对应列将留空:")
            for t in missing[:5]:
                print(f"  - {t}")
            if len(missing) > 5:
                print(f"  ... 还有 {len(missing)-5} 个")
    else:
        print("[INFO] 未提供学习指引（guide），学习目标/验证方式/学习内容列将留空")

    # ---------- v2.0: 日期天数校验 ----------
    days_info = template_engine.get_days_info(groups, daily_hours, phase_dates)
    if days_info['mismatch_warning']:
        print(f"[WARN] {days_info['mismatch_warning']}")

    # ---------- v2.0: 模板模式 ----------
    template_path = config.get('template')
    if template_path:
        print(f"[INFO] 使用模板模式: ...{os.path.basename(template_path)}")
        output = template_engine.generate_from_template(config, groups, guide_map)
        print(f'\n已保存学习计划（{len(videos)}个视频, {len(groups)}天, {total_h:.1f}h，模板模式）')
        return output

    # ---------- 内置模式（向后兼容） ----------
    output_raw   = config.get('output', 'output/学习计划.xlsx')
    output       = validate_output_path(output_raw)

    full_title   = f'{course_name} 课程录播课学习计划（学习+备考周期为1个月）'
    phase_title  = f'{phase_name}【{phase_dates} 共约{round(total_h)}课时】'

    # ---------- 创建 Excel ----------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '学习计划'

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    # R1 大标题
    merge_and_write(ws, row, 2, LAST_COL, full_title, FONT_TITLE, FILL_TITLE, ALIGN_C, 36); row += 1
    # R2 报名阶段
    merge_and_write(ws, row, 2, LAST_COL, '学员培训报名阶段【开课前完成】', FONT_STAGE, FILL_STAGE, ALIGN_C, 26); row += 1
    # R3 报名说明
    merge_and_write(ws, row, 2, LAST_COL, reg_info, FONT_NOTE, FILL_NOTE, ALIGN_LT, 66); row += 1
    # R4 精讲阶段
    merge_and_write(ws, row, 2, LAST_COL, phase_title, FONT_STAGE, FILL_STAGE, ALIGN_C, 26); row += 1
    # R5 空行
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=LAST_COL)
    ws.row_dimensions[row].height = 8; row += 1
    # R6 表头
    for ci, h in enumerate(HEADERS, 2):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font, cell.fill, cell.alignment = FONT_HEADER, FILL_HEADER, ALIGN_C
        cell.border = BORDER
    ws.row_dimensions[row].height = 28; row += 1

    # ---------- R7+ 数据 ----------
    for day_idx, grp in enumerate(groups, 1):
        start_row = row
        day_min = 0
        for ti, (title, minutes) in enumerate(grp):
            day_min += minutes
            g = guide_map.get(title, ('', '', ''))
            if isinstance(g, str):
                g = (g, '', '')
            if not isinstance(g, (list, tuple)):
                g = ('', '', '')
            hours = round(minutes / 60, 2)

            vals = [
                f'第{day_idx}天' if ti == 0 else '',
                hours,
                sanitize_title(title),
                str(g[0])[:500] if len(g) > 0 else '',
                str(g[1])[:500] if len(g) > 1 else '',
                str(g[2])[:500] if len(g) > 2 else '',
                '',
                '',
                homework,
                lm_method,
                lm_materials,
            ]
            write_row(ws, row, vals); row += 1

        # 合并同天的时间安排列
        if row - 1 > start_row:
            ws.merge_cells(start_row=start_row, start_column=2, end_row=row - 1, end_column=2)
            ws.cell(row=start_row, column=2).alignment = ALIGN_C

        # 合并同天的学习要求/资料列
        if row - 1 > start_row:
            ws.merge_cells(start_row=start_row, start_column=11, end_row=row - 1, end_column=11)
            ws.merge_cells(start_row=start_row, start_column=12, end_row=row - 1, end_column=12)

        # 小计行
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=LAST_COL)
        ws.cell(row=row, column=2, value='')
        ws.cell(row=row, column=3, value=round(day_min / 60, 2))
        ws.cell(row=row, column=4, value=f'△ 第{day_idx}天累计 {int(day_min)}分钟')
        for c in range(2, LAST_COL + 1):
            cell = ws.cell(row=row, column=c)
            cell.font, cell.fill = FONT_BOLD, FILL_SECTION
            cell.alignment = ALIGN_C if c <= 3 else ALIGN_LC
            cell.border = BORDER
        ws.row_dimensions[row].height = 24; row += 1

    # ---------- 底部信息 ----------
    sections = [
        ('考试介绍和备考资料', exam_info, 180),
        ('考试阶段',           exam_phase, 110),
        ('成绩查询和证书维持', score_info, 72),
    ]
    for title, content, h in sections:
        merge_and_write(ws, row, 2, LAST_COL, title, FONT_STAGE, FILL_STAGE, ALIGN_C, 26); row += 1
        merge_and_write(ws, row, 2, LAST_COL, content, FONT_BODY, FILL_NOTE, ALIGN_LT, h); row += 1

    # 冻结
    ws.freeze_panes = 'B7'

    # 确保输出目录存在（仅限非系统目录）
    out_dir = os.path.dirname(output)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output)
    # 安全日志：仅打印文件名，不打印完整路径
    print(f'\n已保存学习计划（{len(videos)}个视频, {len(groups)}天, {total_h:.1f}h）')
    return output


# ============================================================
# CLI
# ============================================================
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='录播课学习计划生成器 v2.0.2 (通用模板引擎)')
    parser.add_argument('--config', help='JSON 配置文件路径')
    parser.add_argument('--template', help='Excel 模板路径（用户自定义模板）')
    parser.add_argument('--input', help='Excel 视频清单路径（快捷模式）')
    parser.add_argument('--course-name', help='课程名称（快捷模式）')
    parser.add_argument('--daily-hours', type=float, default=2, help=f'每日学习上限小时 ({MIN_DAILY_HOURS}-{MAX_DAILY_HOURS})')
    parser.add_argument('--output', default='output/学习计划.xlsx', help='输出 .xlsx 路径')
    parser.add_argument('--sheet', help='Excel 工作表名')
    parser.add_argument('--title-col', type=int, default=1, help='标题所在列号')
    parser.add_argument('--dur-col', type=int, default=2, help='时长所在列号')
    args = parser.parse_args()

    # 构建 config
    if args.config:
        config_path = os.path.abspath(args.config)
        if not os.path.exists(config_path):
            print(f"[ERROR] 配置文件不存在: {args.config}")
            sys.exit(1)
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        # 命令行 --template 可覆盖 config 中的 template
        if args.template:
            config['template'] = args.template
    elif args.input and args.course_name:
        input_path = os.path.abspath(args.input)
        if not os.path.exists(input_path):
            print(f"[ERROR] 视频清单文件不存在: {args.input}")
            sys.exit(1)
        config = {
            'course_name': args.course_name,
            'daily_hours': args.daily_hours,
            'input_excel': {
                'path': input_path,
                'sheet_name': args.sheet,
                'title_col': args.title_col,
                'dur_col': args.dur_col,
            },
            'output': args.output,
        }
    else:
        parser.print_help()
        sys.exit(1)

    try:
        generate(config)
    except ValueError as e:
        print(f"[ERROR] {e}")
        sys.exit(1)
