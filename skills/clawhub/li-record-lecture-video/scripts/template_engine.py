# -*- coding: utf-8 -*-
"""
录播课学习计划 — 模板引擎 v2.0.2
支持三种模式:
  1. 占位符模式 — 模板中含 {{DAILY_DATA}} 标记
  2. 自动检测模式 — 无占位符时扫描表头关键字
  3. 内置默认模式 — 无模板时使用 11 列 TOGAF 模板

安全约束: 同 generate_study_plan.py 的安全常量
"""

import os, re, copy
from datetime import timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("需要 openpyxl。安装: pip install openpyxl")

# ── 列名关键词映射（中文优先） ──
COLUMN_KEYWORDS = {
    'day':           ['时间安排', '日期', '天次', '第X天', '第.*天', 'Day', 'Date', 'Schedule'],
    'duration':      ['学习时长', '时长', '课时', '时间.*H', 'Duration', 'Hours', 'Time'],
    'title':         ['学习章节', '章节', '课程内容', '视频名称', 'Chapter', 'Title', 'Content', 'Topic'],
    'goal':          ['学习目标', '目标', 'Objective', 'Goal', 'Aim'],
    'verify':        ['验证方式', '目标验证', '验证', '检验', 'Verification', 'Verify', 'Check'],
    'content':       ['学习内容', '内容', '知识点', '大纲', 'Content', 'Topics', 'Syllabus'],
    'project':       ['项目实战', '实战', 'Project', 'Practice'],
    'comprehensive': ['综合项目', '综合实战', '综合', 'Comprehensive'],
    'homework':      ['课后作业', '作业', 'Homework', 'Assignment', 'Task'],
    'method':        ['学习要求', '学习方法', '要求', '要求及方法', 'Method', 'Requirement', 'Approach'],
    'materials':     ['学习资料', '资料', 'Materials', 'Resources', 'Reference'],
}

PLACEHOLDER = '{{DAILY_DATA}}'
MAX_COL = 50  # 最多扫描列数

# ── 安全常量（从 generate_study_plan 继承） ──
FORMULA_INJECTION_PREFIXES = ('=', '+', '-', '@', '\t', '\r', '\n')
MAX_TITLE_LEN = 300

def _sanitize(val):
    """防 Excel 公式注入"""
    if isinstance(val, str) and val and val[0] in FORMULA_INJECTION_PREFIXES:
        return "'" + val
    return val


# ═══════════════════════════════════════════════════════════
# 模板加载与检测
# ═══════════════════════════════════════════════════════════

def load_template(path=None):
    """
    加载模板。返回 (workbook, worksheet, template_info)
    template_info = {
        'mode': 'placeholder' | 'autodetect' | 'builtin',
        'insert_row': int,           # 数据插入起始行
        'header_row': int | None,    # 表头行号
        'col_map': {field: col_idx}, # 列映射
        'merge_cols': [col_idx, ...],# 同天合并的列
        'protected_top': int,        # 顶部保护区结束行
        'protected_bottom': int,     # 底部保护区起始行
        'template_ws': worksheet,    # 模板工作表引用
    }
    """
    if path and os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        info = _analyze_template(ws)
        info['mode'] = info.get('mode', 'autodetect')
        return wb, ws, info
    else:
        # 无模板 → 内置模式
        return _create_builtin_template()


def _analyze_template(ws):
    """分析模板：查找占位符或自动检测表头"""
    max_row = ws.max_row or 100
    max_col = ws.max_column or MAX_COL

    # 1) 查找占位符
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str) and PLACEHOLDER in val:
                return {
                    'mode': 'placeholder',
                    'insert_row': r,
                    'header_row': _find_header_above(ws, r),
                    'col_map': {},
                    'merge_cols': [],
                    'protected_top': r - 1,
                    'protected_bottom': r + 1,
                }

    # 2) 自动检测表头
    header_row = _detect_header_row(ws, max_row, max_col)
    if header_row:
        col_map = _detect_column_mapping(ws, header_row, max_col)
        return {
            'mode': 'autodetect',
            'insert_row': header_row + 1,
            'header_row': header_row,
            'col_map': col_map,
            'merge_cols': _detect_merge_hints(ws, header_row, col_map),
            'protected_top': header_row,
            'protected_bottom': _find_bottom_boundary(ws, header_row, max_row),
        }

    # 3) 无法检测 → 当占位符在最后处理
    return {
        'mode': 'autodetect',
        'insert_row': max_row + 1,
        'header_row': None,
        'col_map': {},
        'merge_cols': [],
        'protected_top': max_row,
        'protected_bottom': max_row + 1,
    }


def _find_header_above(ws, placeholder_row):
    """从占位符行往上找最近的表头行"""
    for r in range(placeholder_row - 1, 0, -1):
        text = ' '.join(str(ws.cell(row=r, column=c).value or '')
                       for c in range(1, min(ws.max_column or MAX_COL, MAX_COL) + 1))
        if any(kw in text for kw in ['学习目标', '学习章节', '时间安排', 'Objective', 'Chapter']):
            return r
    return None


def _is_header_cell(val):
    """判断单元格值是否是表头"""
    if not val or not isinstance(val, str):
        return False
    v = val.strip()
    for keywords in COLUMN_KEYWORDS.values():
        for kw in keywords:
            if re.search(kw, v):
                return True
    return False


def _detect_header_row(ws, max_row, max_col):
    """自动检测表头行（最多 2 个匹配即确认）"""
    for r in range(1, max_row + 1):
        matches = 0
        for c in range(1, min(max_col, MAX_COL) + 1):
            val = ws.cell(row=r, column=c).value
            if _is_header_cell(val):
                matches += 1
        if matches >= 2:
            return r
    return None


def _detect_column_mapping(ws, header_row, max_col):
    """根据表头文字建立列映射"""
    col_map = {}
    for c in range(1, min(max_col, MAX_COL) + 1):
        val = ws.cell(row=header_row, column=c).value
        if not val or not isinstance(val, str):
            continue
        v = val.strip()
        for field, keywords in COLUMN_KEYWORDS.items():
            for kw in keywords:
                if re.search(kw, v):
                    col_map[field] = c
                    break
            if field in col_map:
                break
    return col_map


def _detect_merge_hints(ws, header_row, col_map):
    """检测模板中哪些列是同天合并的（通过检查合并单元格）"""
    merge_cols = []
    for merged_range in ws.merged_cells.ranges:
        # 检查合并是否在数据区且跨行
        if merged_range.min_row > header_row and merged_range.min_row == merged_range.max_row:
            continue  # 单行合并不算
        if merged_range.min_col == merged_range.max_col and merged_range.min_row > header_row:
            # 跨行合并的单列 → 很可能是同天合并列
            col_idx = merged_range.min_col
            if col_idx not in merge_cols:
                merge_cols.append(col_idx)
    return merge_cols


def _find_bottom_boundary(ws, header_row, max_row):
    """找到数据区底部（考试/成绩等章节标记）"""
    section_keywords = ['考试', '成绩', '证书', 'Exam', 'Score', 'Certificate', '报名']
    for r in range(header_row + 2, max_row + 1):
        text = ' '.join(str(ws.cell(row=r, column=c).value or '')
                       for c in range(1, min(ws.max_column or MAX_COL, 10)))
        if any(kw in text for kw in section_keywords):
            return r
    return max_row + 1


# ═══════════════════════════════════════════════════════════
# 内置默认模板
# ═══════════════════════════════════════════════════════════

def _create_builtin_template():
    """创建内置 11 列默认模板（向后兼容）"""
    from openpyxl.styles import Font as F, PatternFill as PF, Alignment as A, Border as B, Side as S

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '学习计划'

    thin = S(style='thin', color='BFBFBF')
    border = B(left=thin, right=thin, top=thin, bottom=thin)

    font_title  = F(name='微软雅黑', size=14, bold=True, color='FFFFFF')
    font_stage  = F(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    font_header = F(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    font_body   = F(name='微软雅黑', size=10)
    font_bold   = F(name='微软雅黑', size=10, bold=True)
    font_note   = F(name='微软雅黑', size=10, color='333333')

    fill_title   = PF('solid', fgColor='1F4E78')
    fill_stage   = PF('solid', fgColor='4472C4')
    fill_header  = PF('solid', fgColor='5B9BD5')
    fill_section = PF('solid', fgColor='D6DCE4')
    fill_note    = PF('solid', fgColor='F2F2F2')

    al_c  = A(horizontal='center', vertical='center', wrap_text=True)
    al_lt = A(horizontal='left',   vertical='top',    wrap_text=True)
    al_lc = A(horizontal='left',   vertical='center', wrap_text=True)

    headers = ['时间安排', '学习时长（H）', '学习章节', '学习目标',
               '目标验证方式', '学习内容', '项目实战', '综合项目实战',
               '课后作业', '学习要求及方法', '学习资料']
    col_widths = [3, 14, 13, 44, 38, 34, 44, 16, 16, 16, 18, 18]
    LAST_COL = 12

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    def _merge_write(r, cs, ce, val, font, fill, al, h):
        ws.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
        cell = ws.cell(row=r, column=cs, value=val)
        cell.font, cell.fill, cell.alignment = font, fill, al
        ws.row_dimensions[r].height = h

    _merge_write(row, 2, LAST_COL, '{{COURSE_TITLE}}', font_title, fill_title, al_c, 36); row += 1  # R1
    _merge_write(row, 2, LAST_COL, '学员培训报名阶段【开课前完成】', font_stage, fill_stage, al_c, 26); row += 1  # R2
    _merge_write(row, 2, LAST_COL, '{{REGISTRATION_INFO}}', font_note, fill_note, al_lt, 66); row += 1  # R3
    _merge_write(row, 2, LAST_COL, '{{PHASE_TITLE}}', font_stage, fill_stage, al_c, 26); row += 1  # R4
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=LAST_COL)
    ws.row_dimensions[row].height = 8; row += 1  # R5

    for ci, h in enumerate(headers, 2):  # R6
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font, cell.fill, cell.alignment = font_header, fill_header, al_c
        cell.border = border
    ws.row_dimensions[row].height = 28; row += 1

    # 占位符行
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=LAST_COL)
    cell = ws.cell(row=row, column=2, value=PLACEHOLDER)
    cell.font, cell.fill, cell.alignment = font_body, PF('solid', fgColor='FFFFFF'), al_c
    ws.row_dimensions[row].height = 24; row += 1

    # 底部章节
    for title, h in [('考试介绍和备考资料', 180), ('考试阶段', 110), ('成绩查询和证书维持', 72)]:
        _merge_write(row, 2, LAST_COL, title, font_stage, fill_stage, al_c, 26); row += 1
        _merge_write(row, 2, LAST_COL, '', font_body, fill_note, al_lt, h); row += 1

    ws.freeze_panes = 'B7'

    info = _analyze_template(ws)
    info['mode'] = 'builtin'
    return wb, ws, info


# ═══════════════════════════════════════════════════════════
# 数据插入核心
# ═══════════════════════════════════════════════════════════

def insert_daily_data(ws, groups, guide_map, homework, method_text, materials_text, info):
    """
    在模板的数据插入位置，按天分组写入视频数据。
    groups: [[(title, mins), ...], ...]
    info:   模板分析结果
    """
    insert_row = info['insert_row']
    col_map = info.get('col_map', {})
    merge_cols = info.get('merge_cols', [])

    # 获取模板中插入行的样式作为默认样式
    template_styles = _capture_row_styles(ws, insert_row)

    # 清除占位符行（如果有）
    if info.get('mode') == 'placeholder':
        _clear_placeholder_row(ws, insert_row)

    # 从底部往上逐行下移，腾出空间
    content_end = _get_content_end(ws, insert_row)
    total_new_rows = _count_needed_rows(groups)  # 视频行 + 小计行
    rows_to_shift = content_end - insert_row
    if rows_to_shift > 0:
        _shift_rows_down(ws, insert_row, rows_to_shift, total_new_rows)

    # 逐天写入
    curr_row = insert_row
    for day_idx, grp in enumerate(groups, 1):
        start_row = curr_row
        day_min = 0
        for vi, (title, minutes) in enumerate(grp):
            day_min += minutes
            g = _resolve_guide(guide_map, title)
            hours = round(minutes / 60, 2)

            row_data = {
                'day':       f'第{day_idx}天' if vi == 0 else '',
                'duration':  hours,
                'title':     _sanitize(title),
                'goal':      str(g[0])[:500] if len(g) > 0 else '',
                'verify':    str(g[1])[:500] if len(g) > 1 else '',
                'content':   str(g[2])[:500] if len(g) > 2 else '',
                'project':   '',
                'comprehensive': '',
                'homework':  homework,
                'method':    method_text,
                'materials': materials_text,
            }
            _write_data_row(ws, curr_row, row_data, col_map, template_styles)
            curr_row += 1

        # 同天合并
        if curr_row - 1 > start_row:
            _merge_day_columns(ws, start_row, curr_row - 1, col_map, merge_cols)

        # 小计行
        _write_subtotal_row(ws, curr_row, day_idx, day_min, col_map, template_styles)
        curr_row += 1

    return curr_row


def _resolve_guide(guide_map, title):
    """解析学习指引"""
    g = guide_map.get(title, ('', '', ''))
    if isinstance(g, str):
        g = (g, '', '')
    if not isinstance(g, (list, tuple)):
        g = ('', '', '')
    return g


def _capture_row_styles(ws, row):
    """捕获模板行的样式信息"""
    styles = {}
    for c in range(1, min(ws.max_column or MAX_COL, MAX_COL) + 1):
        cell = ws.cell(row=row, column=c)
        styles[c] = {
            'font': copy.copy(cell.font),
            'fill': copy.copy(cell.fill),
            'alignment': copy.copy(cell.alignment),
            'border': copy.copy(cell.border),
        }
    return styles


def _clear_placeholder_row(ws, row):
    """清除占位符行的内容和合并"""
    # 拆解该行的所有合并
    merged_to_remove = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row:
            merged_to_remove.append(str(mr))
    for mr_str in merged_to_remove:
        ws.unmerge_cells(mr_str)
    # 清空内容
    for c in range(1, min(ws.max_column or MAX_COL, MAX_COL) + 1):
        ws.cell(row=row, column=c).value = None


def _get_content_end(ws, insert_row):
    """找到数据插入点之后的内容区结束行"""
    max_row = ws.max_row or insert_row
    return max_row


def _count_needed_rows(groups):
    """计算共需多少行（视频行 + 小计行）"""
    total = 0
    for grp in groups:
        total += len(grp) + 1  # 视频行 + 小计行
    return total


def _shift_rows_down(ws, start_row, existing_rows, new_row_count):
    """下移现有行以腾出空间"""
    if new_row_count <= existing_rows:
        return  # 新行数不大于已有行数，原地覆盖
    extra = new_row_count - existing_rows
    ws.insert_rows(start_row + existing_rows, extra)


def _write_data_row(ws, row, data, col_map, styles):
    """写入单行视频数据"""
    field_to_col = {
        'day': 2, 'duration': 3, 'title': 4, 'goal': 5, 'verify': 6,
        'content': 7, 'project': 8, 'comprehensive': 9, 'homework': 10,
        'method': 11, 'materials': 12,
    }
    # 如果有关键列映射，优先使用
    for field, col in col_map.items():
        if field in field_to_col:
            field_to_col[field] = col

    for field, col in field_to_col.items():
        val = data.get(field, '')
        cell = ws.cell(row=row, column=col, value=val)
        if col in styles:
            s = styles[col]
            cell.font = copy.copy(s['font'])
            cell.alignment = copy.copy(s['alignment'])
            cell.border = copy.copy(s['border'])
            # 不复制 fill，让数据行用白色背景（看得更清楚）
        ws.row_dimensions[row].height = 60


def _merge_day_columns(ws, start_row, end_row, col_map, merge_cols):
    """合并同天行的指定列"""
    # 内置列索引（B=2 为时间安排列）
    day_col = col_map.get('day', 2)

    # 合并时间安排列
    if start_row < end_row:
        ws.merge_cells(start_row=start_row, start_column=day_col,
                      end_row=end_row, end_column=day_col)

    # 合并学习方法列（K=11）
    method_col = col_map.get('method', 11)
    if start_row < end_row and method_col != day_col:
        try:
            ws.merge_cells(start_row=start_row, start_column=method_col,
                          end_row=end_row, end_column=method_col)
        except:
            pass

    # 合并学习资料列（L=12）
    mat_col = col_map.get('materials', 12)
    if start_row < end_row and mat_col != day_col and mat_col != method_col:
        try:
            ws.merge_cells(start_row=start_row, start_column=mat_col,
                          end_row=end_row, end_column=mat_col)
        except:
            pass


def _write_subtotal_row(ws, row, day_idx, day_min, col_map, styles):
    """写入小计行"""
    day_col = col_map.get('day', 2)
    dur_col = col_map.get('duration', 3)
    title_col = col_map.get('title', 4)
    last_content_col = max(col_map.values()) if col_map else 12

    # 默认小计样式
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    font_sub = Font(name='微软雅黑', size=10, bold=True)
    fill_sub = PatternFill('solid', fgColor='D6DCE4')
    al_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_lc = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 合并学习章节到最后一列
    ws.merge_cells(start_row=row, start_column=title_col, end_row=row, end_column=last_content_col)

    ws.cell(row=row, column=day_col, value='')
    dur_cell = ws.cell(row=row, column=dur_col, value=round(day_min / 60, 2))
    ws.cell(row=row, column=title_col, value=f'△ 第{day_idx}天累计 {int(day_min)}分钟')

    for c in range(1, last_content_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font_sub
        cell.fill = fill_sub
        cell.alignment = al_c if c <= 3 else al_lc
        cell.border = border
    ws.row_dimensions[row].height = 24


# ═══════════════════════════════════════════════════════════
# 顶层接口
# ═══════════════════════════════════════════════════════════

def generate_from_template(config, groups, guide_map):
    """
    使用模板（或内置默认）生成学习计划 Excel。
    
    参数:
      config  — 完整配置 dict（含 template_path, course_name, output 等）
      groups  — 分组后的视频数据 [[(title, min), ...], ...]
      guide_map — 学习指引 {title: (goal, verify, content)}
    
    返回: 输出文件路径
    """
    template_path = config.get('template')
    output_path = config.get('output', 'output/学习计划.xlsx')

    # 加载模板
    wb, ws, info = load_template(template_path)

    # 处理默认文本
    course_name = config.get('course_name', '课程录播课学习计划')
    phase_name = config.get('phase_name', '精讲阶段')
    phase_dates = config.get('phase_dates', 'X月XX日-X月XX日')
    homework = config.get('homework', '完成本章节课后练习题')
    method = config.get('learning_method', '')
    materials = config.get('learning_materials', '')
    reg_info = config.get('registration_info', '')
    exam_info = config.get('exam_info', '')
    exam_phase = config.get('exam_phase', '')
    score_info = config.get('score_info', '')

    total_h = sum(m for _, m in sum(groups, [])) / 60

    # 占位符替换
    _replace_placeholder_text(ws, '{{COURSE_TITLE}}',
        f'{course_name} 课程录播课学习计划（学习+备考周期为1个月）')
    _replace_placeholder_text(ws, '{{REGISTRATION_INFO}}', reg_info)
    _replace_placeholder_text(ws, '{{PHASE_TITLE}}',
        f'{phase_name}【{phase_dates} 共约{round(total_h)}课时】')

    # 插入数据
    insert_daily_data(ws, groups, guide_map, homework, method, materials, info)

    # 处理底部文本（仅占位符模式）
    _replace_placeholder_text(ws, '{{EXAM_INFO}}', exam_info)
    _replace_placeholder_text(ws, '{{EXAM_PHASE}}', exam_phase)
    _replace_placeholder_text(ws, '{{SCORE_INFO}}', score_info)

    # 保存
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    return output_path


def _replace_placeholder_text(ws, placeholder, text):
    """在整个工作表中查找并替换占位符文本"""
    if not text:
        return
    max_row = ws.max_row or 100
    max_col = ws.max_column or MAX_COL
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value and isinstance(cell.value, str) and placeholder in cell.value:
                cell.value = cell.value.replace(placeholder, text)
                return  # 只替换第一处


def get_days_info(groups, daily_hours, phase_dates=None):
    """获取分组后的天数信息和日期校验提示

    校验维度：
      1) 数学估算：ceil(total_hours / daily_hours) vs actual_days
      2) 用户日期范围：解析"8月5日-8月10日"中的指定天数 vs actual_days
    """
    import re
    actual_days = len(groups)
    total_min = sum(m for grp in groups for _, m in grp)
    total_h = round(total_min / 60, 2)

    warnings = []

    # 维度1：数学估算校验
    expected_min = round(total_min / 60 / daily_hours + 0.5)
    if actual_days != expected_min:
        warnings.append(
            f'数学估算约 {expected_min} 天，实际分组 {actual_days} 天（贪婪算法产生）'
        )

    # 维度2：用户日期范围校验
    specified_days = 0
    if phase_dates:
        # 解析 "8月5日-8月10日" → 提取日期差值
        m = re.search(r'(\d+)月(\d+)日\s*[-–—至到]\s*(\d+)月(\d+)日', phase_dates)
        if m:
            m1, d1, m2, d2 = map(int, m.groups())
            specified_days = (int(m2) - int(m1)) * 31 + (int(d2) - int(d1)) + 1
            # 同月简化
            if m1 == m2:
                specified_days = d2 - d1 + 1
        else:
            # 尝试 "8月5日-10日" 格式
            m2 = re.search(r'(\d+)月(\d+)日\s*[-–—至到]\s*(\d+)日', phase_dates)
            if m2:
                m, d1, d2 = map(int, m2.groups())
                specified_days = d2 - d1 + 1

    if specified_days > 0 and specified_days != actual_days:
        warnings.append(
            f'日期范围指定 {specified_days} 天，实际分组 {actual_days} 天 — 请调整日期范围或每日学时'
        )

    return {
        'actual_days': actual_days,
        'total_hours': total_h,
        'daily_hours': daily_hours,
        'specified_days': specified_days,
        'mismatch_warning': '；'.join(warnings) if warnings else None
    }
