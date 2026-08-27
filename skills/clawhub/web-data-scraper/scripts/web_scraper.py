#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
浏览器DOM数据抓取插件 - Web Scraper via Chrome DevTools Protocol (CDP)
核心原理：通过 127.0.0.1:9222 CDP 连接用户已打开的浏览器，直接读取 DOM 绕过反爬与风控。
"""

import os
import sys
import json
import time
import re
import argparse
import urllib.request
from typing import Dict, Any, List, Optional
import websocket
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==========================================
# 1. 平台选择器适配层
# ==========================================
PLATFORM_CONFIGS = {
    "xiaohongshu": {
        "domain": "xiaohongshu.com",
        "selectors": {
            "title": "h1.title, .title",
            "author": ".author-wrapper .username, .name",
            "publish_time": ".date, .time",
            "desc": "#detail-desc, .desc, .note-text",
            "likes": ".like-wrapper .count, [class*='like'] .count",
            "collects": "[class*='collect'] .count",
            "comments_count": ".comments-el, [class*='chat'] .count",
            "comment_container": ".comments-container",
            "comment_item": ".comment-item",
            "comment_sub_item": ".comment-item-sub, .sub-comment-item",
            "comment_user": ".user-name, .name",
            "comment_content": ".content",
            "comment_like": ".like-wrapper .count",
            "comment_time": ".date, .time"
        }
    },
    "douyin": {
        "domain": "douyin.com",
        "selectors": {
            "title": "[data-e2e='video-desc'], h1",
            "author": "[data-e2e='user-info'] .account-name, .account-name",
            "publish_time": ".publish-time",
            "desc": "[data-e2e='video-desc']",
            "likes": "[data-e2e='video-player-share-container'] [data-e2e='feed-comment-icon'] + .count", # Placeholder fallback
            "collects": "",
            "comments_count": "",
            "comment_container": "[data-e2e='comment-list']",
            "comment_item": "[data-e2e='comment-item']",
            "comment_sub_item": "[data-e2e='sub-comment-item']",
            "comment_user": "[data-e2e='comment-user-name']",
            "comment_content": "[data-e2e='comment-text']",
            "comment_like": "[data-e2e='comment-like-count']",
            "comment_time": ".comment-time"
        }
    },
    "bilibili": {
        "domain": "bilibili.com",
        "selectors": {
            "title": "h1.video-title, .video-info-title-inner",
            "author": ".username, .up-name",
            "publish_time": ".pubdate-ip-text, .pubdate",
            "desc": ".desc-info-text, .video-desc",
            "likes": ".video-like-info, .like",
            "collects": ".video-fav-info, .collect",
            "comments_count": ".comment-title .count-text",
            "comment_container": ".reply-list, .comment-list",
            "comment_item": ".reply-item, .comment-item",
            "comment_sub_item": ".sub-reply-item",
            "comment_user": ".user-name, .name",
            "comment_content": ".reply-content, .text",
            "comment_like": ".reply-like, .like-info",
            "comment_time": ".reply-time, .time"
        }
    }
}

# ==========================================
# 2. CDP 通信管理类
# ==========================================
class CDPBrowserScraper:
    def __init__(self, port: int = 9222):
        self.port = port
        self.base_url = f"http://127.0.0.1:{port}"

    def get_target_pages(self) -> List[Dict[str, Any]]:
        """获取浏览器当前打开的所有页面"""
        try:
            url = f"{self.base_url}/json"
            with urllib.request.urlopen(url, timeout=5) as response:
                return json.loads(response.read().decode())
        except Exception as e:
            print(f"[-] 无法连接到 Chrome 调试端口 {self.port}。请确保启动了 Chrome 调试模式！错误信息: {e}")
            sys.exit(1)

    def find_page_by_platform(self, platform: str) -> Optional[Dict[str, Any]]:
        """根据域名或名称匹配目标页面"""
        pages = self.get_target_pages()
        config = PLATFORM_CONFIGS.get(platform)
        domain = config["domain"] if config else platform
        
        for p in pages:
            if p.get("type") == "page" and domain in p.get("url", ""):
                return p
        return None

    def execute_js(self, ws_url: str, js_code: str) -> Any:
        """连接 WebSocket 并执行 JS，获取返回值"""
        ws = websocket.create_connection(ws_url)
        payload = {
            "id": 1,
            "method": "Runtime.evaluate",
            "params": {
                "expression": js_code,
                "returnByValue": True,
                "awaitPromise": True
            }
        }
        ws.send(json.dumps(payload))
        result = json.loads(ws.recv())
        ws.close()
        
        if "result" in result and "result" in result["result"]:
            val = result["result"]["result"].get("value")
            return val
        return None

# ==========================================
# 3. 浏览器内执行的高级 JS 抓取脚本
# ==========================================
JS_SCRAPER_TEMPLATE = """
(() => {
    const selectors = %s;
    
    // 1. 点赞等数值转换函数
    const parseNum = (str) => {
        if (!str) return 0;
        str = str.trim().toLowerCase();
        let match = str.match(/([0-9.]+)\\s*([kw万千]?)/i);
        if (!match) return 0;
        let num = parseFloat(match[1]);
        let unit = match[2];
        if (unit === 'w' || unit === '万') return Math.round(num * 10000);
        if (unit === 'k' || unit === '千') return Math.round(num * 1000);
        return Math.round(num);
    };

    // 2. 自动点击“展开/查看更多”按钮（文本过滤）
    const expandButtons = () => {
        const btnSelector = '[class*="expand"],[class*="more"],[class*="reply"],[class*="btn"]';
        const btns = Array.from(document.querySelectorAll(btnSelector));
        const keywords = /展开|查看更多|加载更多|更多回复|回复/;
        let count = 0;
        for (let btn of btns) {
            if (count >= 10) break; // 限量防误点
            const txt = btn.innerText || "";
            if (txt.length < 15 && keywords.test(txt)) {
                btn.click();
                count++;
            }
        }
        return count;
    };

    // 执行一次展开动作
    expandButtons();

    // 3. 提取主页面数据
    const getElText = (sel) => {
        if (!sel) return "";
        const el = document.querySelector(sel);
        return el ? el.innerText.trim() : "";
    };

    const metadata = {
        title: getElText(selectors.title) || document.title,
        author: getElText(selectors.author),
        publish_time: getElText(selectors.publish_time),
        desc: getElText(selectors.desc),
        likes: parseNum(getElText(selectors.likes)),
        collects: parseNum(getElText(selectors.collects)),
        comments_count: parseNum(getElText(selectors.comments_count)),
        url: window.location.href
    };

    // 4. 提取评论（包含子评论关系层级挂载）
    const commentsList = [];
    const container = document.querySelector(selectors.comment_container);
    if (container) {
        // 获取所有评论元素
        const items = Array.from(container.querySelectorAll(`${selectors.comment_item}, ${selectors.comment_sub_item}`));
        let currentParent = null;

        items.forEach(el => {
            const isSub = el.matches(selectors.comment_sub_item) || (selectors.comment_item !== selectors.comment_sub_item && !el.matches(selectors.comment_item));
            
            const userEl = el.querySelector(selectors.comment_user);
            const contentEl = el.querySelector(selectors.comment_content);
            const likeEl = el.querySelector(selectors.comment_like);
            const timeEl = el.querySelector(selectors.comment_time);

            const cData = {
                user: userEl ? userEl.innerText.trim() : "未知用户",
                content: contentEl ? contentEl.innerText.trim() : "",
                likes: parseNum(likeEl ? likeEl.innerText : "0"),
                time: timeEl ? timeEl.innerText.trim() : "",
                sub_replies: []
            };

            if (!isSub) {
                // 主评论
                currentParent = cData;
                commentsList.push(cData);
            } else {
                // 子评论
                if (currentParent) {
                    currentParent.sub_replies.push(`${cData.user}：${cData.content}（👍${cData.likes} ${cData.time}）`);
                } else {
                    // 若没有前置父评论则作为独立主评论
                    commentsList.push(cData);
                }
            }
        });
    }

    return { metadata, comments: commentsList };
})();
"""

# ==========================================
# 4. 数据清洗与数值归一化（Python层兜底）
# ==========================================
def clean_and_normalize_likes(val) -> int:
    """中文或特殊格式点赞数归一化，如 1.2w -> 12000, 1.2k -> 1200"""
    if not val:
        return 0
    if isinstance(val, int):
        return val
    val_str = str(val).strip().lower()
    if val_str in ["赞", "点赞", "回复", ""]:
        return 0
    try:
        match = re.search(r"([0-9.]+)\s*([kw万千]?)", val_str)
        if not match:
            return 0
        num = float(match.group(1))
        unit = match.group(2)
        if unit in ["w", "万"]:
            return int(num * 10000)
        elif unit in ["k", "千"]:
            return int(num * 1000)
        return int(num)
    except Exception:
        return 0

# ==========================================
# 5. Excel 生成器 (openpyxl)
# ==========================================
def save_to_excel(metadata: Dict[str, Any], comments: List[Dict[str, Any]], filename: str):
    wb = openpyxl.Workbook()
    
    # 样式定义
    font_family = "Microsoft YaHei"
    title_font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=11, bold=True, color="333333")
    data_font = Font(name=font_family, size=10, color="000000")
    
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    title_fill = PatternFill(start_color="741923", end_color="741923", fill_type="solid") # Elegant Deep Red
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ------------------ Sheet 1: 概览页 ------------------
    ws_meta = wb.active
    ws_meta.title = "概览"
    
    # 标题栏
    ws_meta.merge_cells("A1:H1")
    title_cell = ws_meta["A1"]
    title_cell.value = "网页数据抓取概览"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = align_center
    ws_meta.row_dimensions[1].height = 40
    
    headers_meta = ["标题", "作者", "发布时间", "点赞数", "收藏数", "评论数", "正文/描述", "URL"]
    ws_meta.append([]) # 空一行
    ws_meta.append(headers_meta)
    ws_meta.row_dimensions[3].height = 25
    
    # 填充表头样式
    for col_idx in range(1, len(headers_meta) + 1):
        cell = ws_meta.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = grid_border

    # 数据行
    row_data = [
        metadata.get("title", ""),
        metadata.get("author", ""),
        metadata.get("publish_time", ""),
        metadata.get("likes", 0),
        metadata.get("collects", 0),
        metadata.get("comments_count", 0),
        metadata.get("desc", ""),
        metadata.get("url", "")
    ]
    ws_meta.append(row_data)
    ws_meta.row_dimensions[4].height = 50
    
    for col_idx in range(1, len(row_data) + 1):
        cell = ws_meta.cell(row=4, column=col_idx)
        cell.font = data_font
        cell.alignment = align_left if col_idx in [1, 7, 8] else align_center
        cell.border = grid_border
        
    # 自适应列宽
    for col in ws_meta.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_meta.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    # ------------------ Sheet 2: 评论页 ------------------
    ws_comments = wb.create_sheet(title="评论详情")
    
    ws_comments.merge_cells("A1:F1")
    c_title_cell = ws_comments["A1"]
    c_title_cell.value = f"【{metadata.get('title', '网页')}】 评论区数据"
    c_title_cell.font = title_font
    c_title_cell.fill = title_fill
    c_title_cell.alignment = align_center
    ws_comments.row_dimensions[1].height = 40
    
    headers_comm = ["序号", "用户", "评论内容", "点赞数", "时间", "折叠回复(子评论)"]
    ws_comments.append([]) # 空一行
    ws_comments.append(headers_comm)
    ws_comments.row_dimensions[3].height = 25
    
    for col_idx in range(1, len(headers_comm) + 1):
        cell = ws_comments.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = grid_border

    # 写入评论数据
    row_idx = 4
    for idx, c in enumerate(comments, 1):
        sub_replies_str = "\n".join(c.get("sub_replies", []))
        row_data = [
            idx,
            c.get("user", ""),
            c.get("content", ""),
            c.get("likes", 0),
            c.get("time", ""),
            sub_replies_str
        ]
        ws_comments.append(row_data)
        ws_comments.row_dimensions[row_idx].height = 45 if sub_replies_str else 30
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws_comments.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = align_left if col_idx in [2, 3, 6] else align_center
            cell.border = grid_border
        row_idx += 1

    # 自适应列宽
    for col in ws_comments.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_comments.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    wb.save(filename)
    print(f"[+] 数据成功导出至 Excel 文件: {filename}")

# ==========================================
# 6. 命令行解析与主执行入口
# ==========================================
def main():
    parser = argparse.ArgumentParser(description="浏览器DOM无代理、抗风控数据抓取工具（基于 CDP 协议）")
    parser.add_argument("--platform", type=str, default="xiaohongshu", choices=["xiaohongshu", "douyin", "bilibili", "generic"],
                        help="目标平台 (默认: xiaohongshu)")
    parser.add_argument("--port", type=int, default=9222, help="Chrome 调试端口 (默认: 9222)")
    parser.add_argument("--out", type=str, default="result.xlsx", help="导出 Excel 的路径 (默认: result.xlsx)")
    
    # 通用模式自定义选择器
    parser.add_argument("--sel-title", type=str, help="自定义标题 CSS 选择器")
    parser.add_argument("--sel-author", type=str, help="自定义作者 CSS 选择器")
    parser.add_argument("--sel-desc", type=str, help="自定义正文 CSS 选择器")
    parser.add_argument("--sel-likes", type=str, help="自定义点赞 CSS 选择器")
    parser.add_argument("--sel-comments-count", type=str, help="自定义评论总数 CSS 选择器")
    parser.add_argument("--sel-comment-container", type=str, help="自定义评论区容器 CSS 选择器")
    parser.add_argument("--sel-comment-item", type=str, help="自定义主评论项 CSS 选择器")
    parser.add_argument("--sel-comment-sub", type=str, help="自定义子评论项 CSS 选择器")
    parser.add_argument("--sel-comment-user", type=str, help="自定义评论用户名 CSS 选择器")
    parser.add_argument("--sel-comment-content", type=str, help="自定义评论内容 CSS 选择器")
    
    args = parser.parse_args()

    scraper = CDPBrowserScraper(args.port)
    
    # 自动定位匹配页面
    print(f"[*] 正在连接 127.0.0.1:{args.port} 检测浏览器页面...")
    target_page = scraper.find_page_by_platform(args.platform)
    
    if not target_page:
        print(f"[-] 未能在已打开的浏览器页面中找到与 {args.platform} 相关的标签页。")
        print("[*] 请确保您的浏览器已经打开了相应的目标网页，例如小红书的帖子详情页！")
        sys.exit(1)
        
    print(f"[+] 找到目标标签页: {target_page.get('title')} ({target_page.get('url')})")
    
    # 配置选择器
    if args.platform == "generic":
        selectors = {
            "title": args.sel_title or "h1",
            "author": args.sel_author or "",
            "publish_time": "",
            "desc": args.sel_desc or "article",
            "likes": args.sel_likes or "",
            "collects": "",
            "comments_count": args.sel_comments_count or "",
            "comment_container": args.sel_comment_container or ".comments",
            "comment_item": args.sel_comment_item or ".comment",
            "comment_sub_item": args.sel_comment_sub or ".reply",
            "comment_user": args.sel_comment_user or ".user",
            "comment_content": args.sel_comment_content or ".text",
            "comment_like": "",
            "comment_time": ""
        }
    else:
        selectors = PLATFORM_CONFIGS[args.platform]["selectors"]

    # 渲染 JS 并发送执行
    ws_url = target_page.get("webSocketDebuggerUrl")
    if not ws_url:
        print("[-] 未能获取到页面的 websocket 调试地址。")
        sys.exit(1)
        
    print("[*] 正在执行 DOM 树抓取与数据分析...")
    js_to_run = JS_SCRAPER_TEMPLATE % json.dumps(selectors)
    raw_data = scraper.execute_js(ws_url, js_to_run)
    
    if not raw_data:
        print("[-] 抓取失败或返回了空数据。")
        sys.exit(1)
        
    metadata = raw_data.get("metadata", {})
    comments = raw_data.get("comments", [])
    
    # 清洗数值
    metadata["likes"] = clean_and_normalize_likes(metadata.get("likes"))
    metadata["collects"] = clean_and_normalize_likes(metadata.get("collects"))
    metadata["comments_count"] = clean_and_normalize_likes(metadata.get("comments_count"))
    
    print("\n" + "="*30 + " 抓取成功 " + "="*30)
    print(f" 标题: {metadata.get('title')}")
    print(f" 作者: {metadata.get('author')}")
    print(f" 点赞: {metadata.get('likes')} | 收藏: {metadata.get('collects')} | 评论总数: {metadata.get('comments_count')}")
    print(f" 正文摘要: {str(metadata.get('desc'))[:80]}...")
    print(f" 获取到评论条数 (包含挂载子回复): {len(comments)}")
    print("="*68 + "\n")
    
    # 写入 Excel
    save_to_excel(metadata, comments, args.out)

if __name__ == "__main__":
    main()
