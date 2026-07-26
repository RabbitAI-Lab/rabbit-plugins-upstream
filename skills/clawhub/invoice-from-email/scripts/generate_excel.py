#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 JSON 文件自动生成差旅报销 Excel（双 Sheet）

读取工作目录所有 .json 文件（parse_invoice.py 输出），自动按类型分组：
  Sheet1「费用清单」：发票 + 火车票
  Sheet2「行程单汇总」：行程单 + 火车票行程

用法：
  python generate_excel.py <json_dir> [output.xlsx]

若省略 output，默认输出到 json_dir 同级的 费用清单.xlsx。
"""

import sys
import os
import json
import re
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def load_json_files(json_dir: str) -> dict:
    """加载目录中所有 JSON 文件，按类型分组。"""
    result = {"invoice": [], "itinerary": [], "train": [], "unknown": []}
    for fname in sorted(os.listdir(json_dir)):
        if not fname.endswith(".json"):
            continue
        with open(os.path.join(json_dir, fname), "r", encoding="utf-8") as f:
            data = json.load(f)
        t = data.get("type", "unknown")
        if t in result:
            result[t].append(data)
        else:
            result["unknown"].append(data)
    return result


def pair_invoice_itinerary(invoices: list, itineraries: list) -> list:
    """
    按文件名【】前缀配对发票和行程单，检测金额差异。
    返回 [(invoice_dict, itinerary_dict_or_None), ...]
    """
    # 提取分组 key：文件名中【】内的文字，或去掉类型后缀
    def group_key(fname):
        m = re.search(r'【([^】]+)】', fname)
        if m:
            return m.group(1).strip()
        return re.sub(r'(电子发票|电子普通发票|电子行程单|行程报销单)\.pdf$', '', fname, flags=re.IGNORECASE).rstrip('-_ ')

    inv_map = {}
    for inv in invoices:
        k = group_key(inv.get("file", ""))
        if k not in inv_map:
            inv_map[k] = []
        inv_map[k].append(inv)

    itin_map = {}
    for itin in itineraries:
        k = group_key(itin.get("file", ""))
        if k not in itin_map:
            itin_map[k] = []
        itin_map[k].append(itin)

    paired = []
    # 配对：同 key 下发票和行程单一一对应
    for key in sorted(set(list(inv_map.keys()) + list(itin_map.keys()))):
        invs = inv_map.get(key, [])
        itins = itin_map.get(key, [])
        for i in range(max(len(invs), len(itins))):
            inv = invs[i] if i < len(invs) else None
            itin = itins[i] if i < len(itins) else None
            if inv:
                # 标记金额差异
                if itin and inv.get("total_amount", 0) > 0 and itin.get("amount", 0) > 0:
                    diff = abs(inv["total_amount"] - itin["amount"])
                    if diff > 0.02:
                        inv["_amount_mismatch"] = f"发票 ¥{inv['total_amount']:.2f} vs 行程单 ¥{itin['amount']:.2f} 差 ¥{diff:.2f}"
            paired.append((inv, itin))
    return paired


def generate_excel(json_dir: str, output_path: str):
    data = load_json_files(json_dir)

    # 配对待处理
    pairs = pair_invoice_itinerary(data["invoice"], data["itinerary"])

    wb = openpyxl.Workbook()
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))

    # === Sheet1: 费用清单 ===
    ws1 = wb.active
    ws1.title = "费用清单"
    ws1.merge_cells("A1:I1")
    c = ws1["A1"]
    c.value = "差旅报销费用清单"
    c.font = Font(name="微软雅黑", bold=True, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 35

    headers1 = ["序号", "发票号码", "开票日期", "销售方", "项目名称", "金额(不含税)", "税额", "价税合计", "备注"]
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=2, column=col, value=h)
        c.font = hfont; c.fill = hf
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = tb
    ws1.row_dimensions[2].height = 22

    row_num = 3
    seq = 0
    te = tt = ta = 0

    # 发票配对行
    for inv, itin in pairs:
        if inv is None:
            continue
        seq += 1
        note = inv.get("_amount_mismatch", "")
        warnings = inv.get("warnings", [])
        if warnings:
            note += ("；" if note else "") + "⚠️ " + "; ".join(warnings)

        vals = [seq, inv.get("invoice_number", ""), inv.get("invoice_date", ""),
                inv.get("seller_name", ""), inv.get("file", "").replace(".pdf", ""),
                inv.get("amount_excluding_tax", 0), inv.get("tax_amount", 0),
                inv.get("total_amount", 0), note]
        for j, v in enumerate(vals):
            c = ws1.cell(row=row_num, column=j+1, value=v)
            c.border = tb; c.font = Font(name="微软雅黑", size=10)
            if j in [0, 1, 2]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif j in [5, 6, 7]:
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = "¥#,##0.00"
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws1.row_dimensions[row_num].height = 28
        te += inv.get("amount_excluding_tax", 0)
        tt += inv.get("tax_amount", 0)
        ta += inv.get("total_amount", 0)
        row_num += 1

    # 火车票行
    for train in data.get("train", []):
        seq += 1
        vals = [seq, train.get("file", "").replace(".pdf", ""),
                train.get("departure_date", ""),
                "中国铁路",
                f"{train.get('train_number', '')} {train.get('departure_station', '')}→{train.get('arrival_station', '')} {train.get('seat_type', '')}",
                0, 0, train.get("amount", 0),
                "票价已含税" if train.get("amount", 0) > 0 else ""]
        for j, v in enumerate(vals):
            c = ws1.cell(row=row_num, column=j+1, value=v)
            c.border = tb; c.font = Font(name="微软雅黑", size=10)
            if j in [0, 1, 2]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif j in [5, 6, 7]:
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = "¥#,##0.00"
            elif j == 8:
                c.font = Font(name="微软雅黑", size=9, color="888888")
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws1.row_dimensions[row_num].height = 28
        ta += train.get("amount", 0)
        row_num += 1

    # 合计行
    for col in range(1, 10):
        c = ws1.cell(row=row_num, column=col)
        c.border = tb
        c.font = Font(name="微软雅黑", bold=True, size=10, color="FF0000")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=row_num, column=1, value="合计")
    ws1.cell(row=row_num, column=6, value=te).number_format = "¥#,##0.00"
    ws1.cell(row=row_num, column=7, value=tt).number_format = "¥#,##0.00"
    ws1.cell(row=row_num, column=8, value=ta).number_format = "¥#,##0.00"
    ws1.row_dimensions[row_num].height = 25

    for i, w in enumerate([6, 22, 12, 30, 40, 14, 10, 14, 32], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A3"

    # === Sheet2: 行程单汇总 ===
    ws2 = wb.create_sheet("行程单汇总")
    ws2.merge_cells("A1:I1")
    c = ws2["A1"]
    c.value = "行程单汇总"
    c.font = Font(name="微软雅黑", bold=True, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35

    headers2 = ["序号", "日期", "时间", "交通工具", "车次/服务商", "出发地", "目的地", "座位/车型", "金额"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = hfont; c.fill = hf
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = tb
    ws2.row_dimensions[2].height = 22

    row_num = 3
    seq = 0
    total_trip = 0

    # 行程单行
    for _, itin in pairs:
        if itin is None:
            continue
        seq += 1
        vals = [seq, itin.get("date", ""), itin.get("time", ""),
                "打车", itin.get("file", "").replace(".pdf", "")[:20],
                itin.get("departure", ""), itin.get("destination", ""),
                itin.get("vehicle", ""), itin.get("amount", 0)]
        for j, v in enumerate(vals):
            c = ws2.cell(row=row_num, column=j+1, value=v)
            c.border = tb; c.font = Font(name="微软雅黑", size=10)
            if j in [0, 3, 4, 7]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif j == 8:
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = "¥#,##0.00"
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[row_num].height = 28
        total_trip += itin.get("amount", 0)
        row_num += 1

    # 火车票行（行程角度）
    for train in data.get("train", []):
        seq += 1
        vals = [seq, train.get("departure_date", ""), train.get("departure_time", ""),
                "火车", train.get("train_number", ""),
                train.get("departure_station", ""), train.get("arrival_station", ""),
                train.get("seat_type", ""), train.get("amount", 0)]
        for j, v in enumerate(vals):
            c = ws2.cell(row=row_num, column=j+1, value=v)
            c.border = tb; c.font = Font(name="微软雅黑", size=10)
            if j in [0, 3, 4, 7]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif j == 8:
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = "¥#,##0.00"
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[row_num].height = 28
        total_trip += train.get("amount", 0)
        row_num += 1

    # 飞机票行（从发票备注提取）
    for inv, _ in pairs:
        if inv is None:
            continue
        note = inv.get("_amount_mismatch", "")
        # 从文件名中提取航班信息
        fname = inv.get("file", "")
        flight_m = re.search(r'(ZH\d+)', fname)
        route_m = re.search(r'([\u4e00-\u9fa5]+)[→→]([\u4e00-\u9fa5]+)', fname)
        if flight_m and "机票" in (inv.get("_amount_mismatch", "") + inv.get("file", "")):
            continue  # 已处理
        if flight_m and inv.get("total_amount", 0) > 100:
            seq += 1
            vals = [seq, inv.get("invoice_date", ""), "", "飞机",
                    flight_m.group(1) if flight_m else "",
                    route_m.group(1) if route_m else "",
                    route_m.group(2) if route_m else "",
                    "经济舱", inv.get("total_amount", 0)]
            for j, v in enumerate(vals):
                c = ws2.cell(row=row_num, column=j+1, value=v)
                c.border = tb; c.font = Font(name="微软雅黑", size=10)
                if j in [0, 3, 4, 7]:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif j == 8:
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.number_format = "¥#,##0.00"
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")
            ws2.row_dimensions[row_num].height = 28
            total_trip += inv.get("total_amount", 0)
            row_num += 1

    # 合计行
    for col in range(1, 10):
        c = ws2.cell(row=row_num, column=col)
        c.border = tb
        c.font = Font(name="微软雅黑", bold=True, size=10, color="FF0000")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.cell(row=row_num, column=1, value="合计")
    ws2.cell(row=row_num, column=9, value=total_trip).number_format = "¥#,##0.00"
    ws2.row_dimensions[row_num].height = 25

    for i, w in enumerate([6, 12, 8, 10, 14, 22, 22, 10, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"

    wb.save(output_path)
    return len(pairs), len(data.get("train", 0))


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    json_dir = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) >= 3 else os.path.join(os.path.dirname(json_dir), "费用清单.xlsx")

    if not os.path.isdir(json_dir):
        print(f"❌ 目录不存在: {json_dir}")
        sys.exit(1)

    invoice_count, train_count = generate_excel(json_dir, output_path)
    print(f"✅ Excel 已生成: {output_path}")
    print(f"   Sheet1: {invoice_count + train_count} 项")
    print(f"   Sheet2: 行程汇总")


if __name__ == "__main__":
    sys.exit(main())
