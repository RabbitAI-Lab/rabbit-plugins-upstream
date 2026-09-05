from __future__ import annotations

import json
import os
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from uuid import UUID, uuid4

import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq
import pytest

from conftest import sign_payload
import sql_data_analyst_local.ingest as ingest_module
from sql_data_analyst_local.contracts import ExpectedTicket
from sql_data_analyst_local.datasets import DatasetError, DatasetRepository
from sql_data_analyst_local.ingest import IngestService
from sql_data_analyst_local.settings import RUNNER_VERSION
from sql_data_analyst_local.tickets import TicketError, TicketVerifier


INSTALLATION_ID = UUID("018f47a2-7b2b-7e47-8794-c11316f5023c")
INPUT_FINGERPRINT = "a" * 64
NOW = datetime(2026, 8, 24, 0, 1, tzinfo=timezone.utc)
MAX_SOURCE_BYTES = 100 * 1024 * 1024


def valid_ticket(operation: str = "dataset.ingest"):
    return sign_payload(
        {
            "schema_version": 1,
            "execution_id": str(uuid4()),
            "operation": operation,
            "installation_id": str(INSTALLATION_ID),
            "input_fingerprint": INPUT_FINGERPRINT,
            "billing_units": 1,
            "charged_amount": "0.050000",
            "currency": "CNY",
            "runner_min_version": RUNNER_VERSION,
            "issued_at": "2026-08-24T00:00:00Z",
            "expires_at": "2026-08-24T00:05:00Z",
        }
    )


@pytest.fixture
def repository(tmp_path, ticket_fixture):
    repo = DatasetRepository(tmp_path / "workspace")
    service = IngestService(
        repo,
        TicketVerifier(ticket_fixture.public_keys),
        ExpectedTicket(
            operation="dataset.ingest",
            installation_id=INSTALLATION_ID,
            input_fingerprint=INPUT_FINGERPRINT,
            runner_version=RUNNER_VERSION,
        ),
        now=lambda: NOW,
    )
    return repo, service


@pytest.fixture
def fixtures(tmp_path):
    root = tmp_path / "fixtures"
    root.mkdir()
    (root / "sales.csv").write_text("category,amount\nbooks,12\ngames,15\n")
    (root / "sales.jsonl").write_text(
        '{"category":"books","amount":12}\n'
        '{"category":"games","amount":15}\n'
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["category", "amount"])
    sheet.append(["books", 12])
    sheet.append(["games", 15])
    workbook.save(root / "sales.xlsx")
    workbook.close()

    pq.write_table(
        pa.table({"category": ["books", "games"], "amount": [12, 15]}),
        root / "sales.parquet",
    )
    return root


@pytest.mark.parametrize(
    "fixture_name", ["sales.csv", "sales.jsonl", "sales.xlsx", "sales.parquet"]
)
def test_ingest_normalizes_supported_files_without_persisting_samples(
    repository, fixtures, fixture_name
):
    repo, service = repository
    dataset_id = uuid4()
    ticket = valid_ticket()

    manifest = service.ingest(fixtures / fixture_name, dataset_id, ticket)

    assert manifest.tables
    assert all(table.parquet_path.endswith(".parquet") for table in manifest.tables)
    assert all(not Path(table.parquet_path).is_absolute() for table in manifest.tables)
    serialized = manifest.model_dump_json()
    assert "rows" not in serialized.lower()
    assert "books" not in serialized
    assert str((fixtures / fixture_name).resolve()) not in serialized
    assert ticket.signature not in serialized
    assert ticket.signed_payload not in serialized
    assert repo.inspect(dataset_id) == manifest
    for table in manifest.tables:
        normalized = repo.table_path(dataset_id, table)
        assert normalized.is_file()
        assert pq.read_table(normalized).num_rows == 2


@pytest.mark.parametrize(
    ("label", "content"),
    [
        ("utf8-bom", "姓名,金额\n甲,12\n".encode("utf-8-sig")),
        ("gb18030", "姓名,金额\n甲,12\n".encode("gb18030")),
    ],
)
def test_ingest_accepts_csv_bom_and_gb18030(repository, tmp_path, label, content):
    _, service = repository
    source = tmp_path / f"{label}.csv"
    source.write_bytes(content)

    manifest = service.ingest(source, uuid4(), valid_ticket())

    assert manifest.tables[0].row_count == 1
    assert [column.display_name for column in manifest.tables[0].columns] == [
        "姓名",
        "金额",
    ]


def test_ingest_accepts_flat_json_array(repository, tmp_path):
    _, service = repository
    source = tmp_path / "sales.json"
    source.write_text('[{"name":"a","amount":1},{"name":"b","amount":2}]')

    manifest = service.ingest(source, uuid4(), valid_ticket())

    assert manifest.source_format == "json"
    assert manifest.tables[0].row_count == 2


def test_ingest_rejects_changed_source_digest_before_parsing(
    repository, tmp_path, monkeypatch
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "changed.csv"
    source.write_text("value\nchanged\n", encoding="utf-8")
    parsed = False

    def observe_parse(*args, **kwargs):
        nonlocal parsed
        parsed = True
        raise AssertionError("mismatched source must not be parsed")

    monkeypatch.setattr(ingest_module, "parse_source", observe_parse)

    with pytest.raises(DatasetError, match="^dataset_invalid$"):
        service.ingest(
            source,
            dataset_id,
            valid_ticket(),
            expected_source_fingerprint="0" * 64,
        )

    assert parsed is False
    assert not repo.dataset_path(dataset_id).exists()
    assert list(repo.datasets_root.glob(".*.tmp")) == []


def test_ingest_rejects_non_string_expected_source_digest(repository, tmp_path):
    _, service = repository
    source = tmp_path / "source.csv"
    source.write_text("value\n1\n", encoding="utf-8")

    with pytest.raises(DatasetError, match="^dataset_invalid$"):
        service.ingest(
            source,
            uuid4(),
            valid_ticket(),
            expected_source_fingerprint=123,
        )


@pytest.mark.parametrize("extension", ["json", "jsonl"])
def test_ingest_losslessly_unions_flat_json_keys(repository, tmp_path, extension):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / f"union.{extension}"
    if extension == "json":
        source.write_text('[{}, {"a":1}, {"b":2,"a":3}]')
    else:
        source.write_text('{}\n{"a":1}\n{"b":2,"a":3}\n')

    manifest = service.ingest(source, dataset_id, valid_ticket())
    table = pq.read_table(repo.table_path(dataset_id, manifest.tables[0]))

    assert table.column_names == ["a", "b"]
    assert table.to_pylist() == [
        {"a": None, "b": None},
        {"a": 1, "b": None},
        {"a": 3, "b": 2},
    ]


def test_ingest_rejects_nested_json_without_echoing_content(repository, tmp_path):
    repo, service = repository
    dataset_id = uuid4()
    secret = "nested-secret-must-not-escape"
    source = tmp_path / "nested.json"
    source.write_text(json.dumps([{"name": "a", "nested": {"secret": secret}}]))

    with pytest.raises(DatasetError, match="^dataset_invalid$") as caught:
        service.ingest(source, dataset_id, valid_ticket())

    assert secret not in str(caught.value)
    assert not repo.dataset_path(dataset_id).exists()
    assert list(repo.datasets_root.glob(".*.tmp")) == []


def test_ingest_accepts_exactly_twenty_xlsx_sheets(repository, tmp_path):
    _, service = repository
    source = tmp_path / "twenty.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "sheet-01"
    for number in range(2, 21):
        workbook.create_sheet(f"sheet-{number:02d}")
    for sheet in workbook.worksheets:
        sheet.append(["value"])
        sheet.append([1])
    workbook.save(source)
    workbook.close()

    manifest = service.ingest(source, uuid4(), valid_ticket())

    assert len(manifest.tables) == 20


def test_ingest_rejects_twenty_one_xlsx_sheets_and_cleans_staging(
    repository, tmp_path
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "twenty-one.xlsx"
    workbook = openpyxl.Workbook()
    for number in range(20):
        workbook.create_sheet(f"sheet-{number:02d}")
    workbook.save(source)
    workbook.close()

    with pytest.raises(DatasetError, match="^dataset_too_large$"):
        service.ingest(source, dataset_id, valid_ticket())

    assert not repo.dataset_path(dataset_id).exists()
    assert list(repo.datasets_root.glob(".*.tmp")) == []


def test_ingest_normalizes_duplicate_and_blank_column_names(repository, tmp_path):
    _, service = repository
    source = tmp_path / "columns.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.append(["Amount", "amount", None, "amount"])
    workbook.active.append([1, 2, 3, 4])
    workbook.save(source)
    workbook.close()

    manifest = service.ingest(source, uuid4(), valid_ticket())

    assert [column.name for column in manifest.tables[0].columns] == [
        "Amount",
        "amount_2",
        "column_3",
        "amount_3",
    ]
    assert [column.display_name for column in manifest.tables[0].columns] == [
        "Amount",
        "amount",
        "",
        "amount",
    ]


def test_ingest_preserves_duplicate_csv_headers_as_display_names(repository, tmp_path):
    _, service = repository
    source = tmp_path / "duplicate.csv"
    source.write_text("Amount,amount,Amount\n1,2,3\n")

    manifest = service.ingest(source, uuid4(), valid_ticket())

    assert [column.name for column in manifest.tables[0].columns] == [
        "Amount",
        "amount_2",
        "Amount_3",
    ]
    assert [column.display_name for column in manifest.tables[0].columns] == [
        "Amount",
        "amount",
        "Amount",
    ]


def test_ingest_accepts_exactly_two_hundred_columns(repository, tmp_path):
    _, service = repository
    source = tmp_path / "wide-boundary.csv"
    source.write_text(
        ",".join(f"c{number}" for number in range(200))
        + "\n"
        + ",".join("1" for _ in range(200))
        + "\n"
    )

    manifest = service.ingest(source, uuid4(), valid_ticket())

    assert len(manifest.tables[0].columns) == 200


def test_ingest_rejects_more_than_two_hundred_columns(repository, tmp_path):
    _, service = repository
    source = tmp_path / "wide.csv"
    source.write_text(",".join(f"c{number}" for number in range(201)) + "\n")

    with pytest.raises(DatasetError, match="^dataset_too_large$"):
        service.ingest(source, uuid4(), valid_ticket())


def test_ingest_rejects_xlsx_expansion_bomb(repository, tmp_path):
    _, service = repository
    source = tmp_path / "bomb.xlsx"
    with zipfile.ZipFile(source, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "A" * 1_000_000)

    with pytest.raises(DatasetError, match="^dataset_too_large$"):
        service.ingest(source, uuid4(), valid_ticket())


def test_ingest_rejects_xlsx_archive_traversal(repository, tmp_path):
    _, service = repository
    source = tmp_path / "traversal.xlsx"
    with zipfile.ZipFile(source, "w") as archive:
        archive.writestr("../outside.xml", "safe-looking")

    with pytest.raises(DatasetError, match="^dataset_invalid$"):
        service.ingest(source, uuid4(), valid_ticket())


def test_ingest_rejects_source_larger_than_one_hundred_mib_before_copying(
    repository, tmp_path
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "large.csv"
    with source.open("wb") as stream:
        stream.truncate(MAX_SOURCE_BYTES + 1)

    with pytest.raises(DatasetError, match="^dataset_too_large$"):
        service.ingest(source, dataset_id, valid_ticket())

    assert not repo.dataset_path(dataset_id).exists()
    assert list(repo.datasets_root.glob(".*.tmp")) == []


@pytest.mark.parametrize(
    ("name", "content"),
    [
        ("spoofed.parquet", b"name,amount\na,1\n"),
        ("spoofed.xlsx", b"name,amount\na,1\n"),
        ("spoofed.csv", b"PAR1not-a-real-parquetPAR1"),
    ],
)
def test_ingest_rejects_format_spoofing(repository, tmp_path, name, content):
    _, service = repository
    source = tmp_path / name
    source.write_bytes(content)

    with pytest.raises(DatasetError, match="^dataset_invalid$"):
        service.ingest(source, uuid4(), valid_ticket())


def test_ingest_verifies_ticket_before_touching_source(repository, tmp_path):
    _, service = repository
    missing_source = tmp_path / "must-not-be-opened.csv"

    with pytest.raises(TicketError, match="^authorization_invalid$"):
        service.ingest(missing_source, uuid4(), valid_ticket("query.execute"))


def test_ingest_rejects_symlink_source(repository, tmp_path):
    _, service = repository
    target = tmp_path / "target.csv"
    target.write_text("name\na\n")
    source = tmp_path / "linked.csv"
    source.symlink_to(target)

    with pytest.raises(DatasetError, match="^dataset_invalid$"):
        service.ingest(source, uuid4(), valid_ticket())


def test_ingest_rejects_symlinked_source_ancestor(repository, tmp_path):
    _, service = repository
    actual = tmp_path / "actual"
    actual.mkdir()
    (actual / "source.csv").write_text("name\na\n")
    alias = tmp_path / "alias"
    alias.symlink_to(actual, target_is_directory=True)

    with pytest.raises(DatasetError, match="^dataset_invalid$"):
        service.ingest(alias / "source.csv", uuid4(), valid_ticket())


def test_workspace_root_swap_after_staging_cannot_redirect_or_orphan_source(
    repository, tmp_path, monkeypatch
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "source.csv"
    secret = "root-swap-secret"
    source.write_text(f"name\n{secret}\n")
    workspace = repo.workspace_root
    original = tmp_path / "original-workspace"
    outside = tmp_path / "outside"
    real_copy = ingest_module._copy_source

    def swap_then_copy(selected_source, stage, destination_name):
        temporary_name = stage.temporary_name
        workspace.rename(original)
        (outside / "datasets" / temporary_name).mkdir(parents=True)
        workspace.symlink_to(outside, target_is_directory=True)
        return real_copy(selected_source, stage, destination_name)

    monkeypatch.setattr(ingest_module, "_copy_source", swap_then_copy)

    with pytest.raises(DatasetError, match="^dataset_invalid$"):
        service.ingest(source, dataset_id, valid_ticket())

    assert secret not in "".join(
        path.read_text(errors="ignore")
        for path in outside.rglob("*")
        if path.is_file()
    )
    external_stages = list((outside / "datasets").iterdir())
    assert len(external_stages) == 1
    assert list(external_stages[0].iterdir()) == []
    assert not any(path.name.endswith(".tmp") for path in original.rglob("*"))


def test_workspace_swap_between_precheck_and_rename_rolls_back_published_uuid(
    repository, tmp_path, monkeypatch
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "source.csv"
    source.write_text("name\nrace-secret\n")
    workspace = repo.workspace_root
    original = tmp_path / "original-workspace"
    outside = tmp_path / "outside"
    outside.mkdir()
    real_rename = os.rename
    swapped = False

    def swap_at_rename(*args, **kwargs):
        nonlocal swapped
        if not swapped:
            real_rename(os.fspath(workspace), os.fspath(original))
            workspace.symlink_to(outside, target_is_directory=True)
            swapped = True
        return real_rename(*args, **kwargs)

    monkeypatch.setattr(os, "rename", swap_at_rename)

    with pytest.raises(DatasetError, match="^dataset_invalid$"):
        service.ingest(source, dataset_id, valid_ticket())

    assert not (original / "datasets" / str(dataset_id)).exists()
    assert not any(path.name.endswith(".tmp") for path in original.rglob("*"))
    assert list(outside.iterdir()) == []


def test_staging_creation_fsync_failure_leaves_no_orphan(
    repository, tmp_path, monkeypatch
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "source.csv"
    source.write_text("name\na\n")
    real_fsync = os.fsync
    calls = 0

    def fail_first_fsync(descriptor):
        nonlocal calls
        calls += 1
        if calls == 1:
            raise OSError("staging directory durability failure")
        return real_fsync(descriptor)

    monkeypatch.setattr(os, "fsync", fail_first_fsync)

    with pytest.raises(DatasetError, match="^dataset_invalid$"):
        service.ingest(source, dataset_id, valid_ticket())

    assert not repo.dataset_path(dataset_id).exists()
    assert list(repo.datasets_root.glob(".*.tmp")) == []


def test_parquet_file_is_fsynced_before_dataset_publication(
    repository, tmp_path, monkeypatch
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "source.csv"
    source.write_text("name\na\n")
    synced_inodes: set[int] = set()
    real_fsync = os.fsync

    def record_fsync(descriptor):
        metadata = os.fstat(descriptor)
        if metadata.st_mode:
            synced_inodes.add(metadata.st_ino)
        return real_fsync(descriptor)

    monkeypatch.setattr(os, "fsync", record_fsync)

    manifest = service.ingest(source, dataset_id, valid_ticket())
    parquet = repo.table_path(dataset_id, manifest.tables[0])

    assert parquet.stat().st_ino in synced_inodes


def test_pre_rename_directory_fsync_failure_aborts_staging(
    repository, tmp_path, monkeypatch
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "source.csv"
    source.write_text("name\na\n")
    manifest_written = False
    real_write_manifest = ingest_module._write_manifest
    real_fsync = os.fsync

    def observe_manifest(*args, **kwargs):
        nonlocal manifest_written
        result = real_write_manifest(*args, **kwargs)
        manifest_written = True
        return result

    def fail_stage_sync(descriptor):
        if manifest_written:
            raise OSError("pre-rename stage durability failure")
        return real_fsync(descriptor)

    monkeypatch.setattr(ingest_module, "_write_manifest", observe_manifest)
    monkeypatch.setattr(os, "fsync", fail_stage_sync)

    with pytest.raises(DatasetError, match="^dataset_invalid$"):
        service.ingest(source, dataset_id, valid_ticket())

    assert not repo.dataset_path(dataset_id).exists()
    assert list(repo.datasets_root.glob(".*.tmp")) == []


def test_rename_failure_aborts_and_removes_complete_staging_tree(
    repository, tmp_path, monkeypatch
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "source.csv"
    source.write_text("name\na\n")

    def fail_rename(*args, **kwargs):
        raise OSError("rename failure")

    monkeypatch.setattr(os, "rename", fail_rename)

    with pytest.raises(DatasetError, match="^dataset_invalid$"):
        service.ingest(source, dataset_id, valid_ticket())

    assert not repo.dataset_path(dataset_id).exists()
    assert list(repo.datasets_root.glob(".*.tmp")) == []


def test_post_rename_fsync_failure_uses_published_success_semantics(
    repository, tmp_path, monkeypatch
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "source.csv"
    source.write_text("name\na\n")
    renamed = False
    real_rename = os.rename
    real_fsync = os.fsync

    def observe_rename(*args, **kwargs):
        nonlocal renamed
        result = real_rename(*args, **kwargs)
        renamed = True
        return result

    def fail_after_rename(descriptor):
        if renamed:
            raise OSError("post-rename durability failure")
        return real_fsync(descriptor)

    monkeypatch.setattr(os, "rename", observe_rename)
    monkeypatch.setattr(os, "fsync", fail_after_rename)

    manifest = service.ingest(source, dataset_id, valid_ticket())

    assert repo.inspect(dataset_id) == manifest
    assert list(repo.datasets_root.glob(".*.tmp")) == []


def test_post_rename_identity_match_close_failure_keeps_published_success(
    repository, tmp_path, monkeypatch
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "source.csv"
    source.write_text("name\na\n")
    renamed = False
    postcheck_descriptor: int | None = None
    close_failed = False
    real_rename = os.rename
    real_open_datasets = repo._open_datasets
    real_close = os.close

    def observe_rename(*args, **kwargs):
        nonlocal renamed
        result = real_rename(*args, **kwargs)
        renamed = True
        return result

    def capture_postcheck_datasets(*, create):
        nonlocal postcheck_descriptor
        descriptor = real_open_datasets(create=create)
        if renamed:
            postcheck_descriptor = descriptor
        return descriptor

    def fail_matched_descriptor_close(descriptor):
        nonlocal close_failed
        if descriptor == postcheck_descriptor and not close_failed:
            close_failed = True
            real_close(descriptor)
            raise OSError("matched post-rename descriptor close failure")
        return real_close(descriptor)

    monkeypatch.setattr(os, "rename", observe_rename)
    monkeypatch.setattr(repo, "_open_datasets", capture_postcheck_datasets)
    monkeypatch.setattr(os, "close", fail_matched_descriptor_close)

    manifest = service.ingest(source, dataset_id, valid_ticket())

    assert close_failed
    assert repo.inspect(dataset_id) == manifest
    assert list(repo.datasets_root.glob(".*.tmp")) == []


def test_rollback_child_close_failure_cannot_leave_published_dataset(
    repository, tmp_path, monkeypatch
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "source.csv"
    source.write_text("name\na\n")
    identity_checks = 0
    rollback_started = False
    rollback_child: int | None = None
    close_failed = False
    real_assert_same = repo._assert_same_datasets
    real_open = os.open
    real_close = os.close

    def fail_post_rename_identity(*args, **kwargs):
        nonlocal identity_checks, rollback_started
        identity_checks += 1
        if identity_checks == 2:
            rollback_started = True
            raise DatasetError()
        return real_assert_same(*args, **kwargs)

    def capture_rollback_child(path, flags, *args, **kwargs):
        nonlocal rollback_child
        descriptor = real_open(path, flags, *args, **kwargs)
        if rollback_started and path == "normalized":
            rollback_child = descriptor
        return descriptor

    def fail_rollback_child_close(descriptor):
        nonlocal close_failed
        if descriptor == rollback_child and not close_failed:
            close_failed = True
            real_close(descriptor)
            raise OSError("rollback child close failure")
        return real_close(descriptor)

    monkeypatch.setattr(repo, "_assert_same_datasets", fail_post_rename_identity)
    monkeypatch.setattr(os, "open", capture_rollback_child)
    monkeypatch.setattr(os, "close", fail_rollback_child_close)

    with pytest.raises(DatasetError, match="^dataset_invalid$"):
        service.ingest(source, dataset_id, valid_ticket())

    assert close_failed
    assert not repo.dataset_path(dataset_id).exists()
    assert list(repo.datasets_root.glob(".*.tmp")) == []


@pytest.mark.parametrize("close_target", ["stage", "datasets"])
def test_post_rename_close_failure_cannot_turn_success_into_error(
    repository, tmp_path, monkeypatch, close_target
):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "source.csv"
    source.write_text("name\na\n")
    stage_descriptor: int | None = None
    datasets_descriptor: int | None = None
    renamed = False
    close_failed = False
    real_begin = repo._begin_staging
    real_rename = os.rename
    real_close = os.close

    def capture_stage(selected_dataset_id):
        nonlocal stage_descriptor, datasets_descriptor
        stage = real_begin(selected_dataset_id)
        stage_descriptor = stage.descriptor
        datasets_descriptor = stage._datasets_descriptor
        return stage

    def observe_rename(*args, **kwargs):
        nonlocal renamed
        result = real_rename(*args, **kwargs)
        renamed = True
        return result

    def fail_published_stage_close(descriptor):
        nonlocal close_failed
        target_descriptor = (
            stage_descriptor if close_target == "stage" else datasets_descriptor
        )
        if renamed and descriptor == target_descriptor and not close_failed:
            close_failed = True
            real_close(descriptor)
            raise OSError("published stage close failure")
        return real_close(descriptor)

    monkeypatch.setattr(repo, "_begin_staging", capture_stage)
    monkeypatch.setattr(os, "rename", observe_rename)
    monkeypatch.setattr(os, "close", fail_published_stage_close)

    manifest = service.ingest(source, dataset_id, valid_ticket())

    assert close_failed
    assert repo.inspect(dataset_id) == manifest
    assert list(repo.datasets_root.glob(".*.tmp")) == []


def test_ingest_writes_private_dataset_directories_and_files(repository, tmp_path):
    repo, service = repository
    dataset_id = uuid4()
    source = tmp_path / "private.csv"
    source.write_text("name\na\n")

    manifest = service.ingest(source, dataset_id, valid_ticket())
    dataset = repo.dataset_path(dataset_id)

    assert dataset.stat().st_mode & 0o777 == 0o700
    assert (dataset / "normalized").stat().st_mode & 0o777 == 0o700
    assert (dataset / "manifest.json").stat().st_mode & 0o777 == 0o600
    assert repo.table_path(dataset_id, manifest.tables[0]).stat().st_mode & 0o777 == 0o600


def test_ingest_failure_does_not_replace_existing_dataset(repository, tmp_path):
    repo, service = repository
    dataset_id = uuid4()
    valid = tmp_path / "valid.csv"
    valid.write_text("name\na\n")
    original = service.ingest(valid, dataset_id, valid_ticket())
    invalid = tmp_path / "invalid.json"
    invalid.write_text('[{"nested":{"secret":true}}]')

    with pytest.raises(DatasetError, match="^dataset_exists$"):
        service.ingest(invalid, dataset_id, valid_ticket())

    assert repo.inspect(dataset_id) == original
