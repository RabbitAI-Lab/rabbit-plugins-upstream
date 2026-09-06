from __future__ import annotations

import os
import re
import stat
from pathlib import Path
from uuid import UUID

import openpyxl
import pytest
from pydantic import ValidationError

from sql_data_analyst_local.report import (
    ReportError,
    ReportSummary,
    ReportWriter,
    parse_summary,
)


EXECUTION_ID = UUID("018f47a2-7b2b-7e47-8794-c11316f5023b")


def summary(**changes: object) -> ReportSummary:
    value: dict[str, object] = {
        "schema_version": 1,
        "title": "Quarterly review",
        "findings": ["Revenue increased"],
        "limitations": ["Partial quarter"],
        "tables": [
            {
                "title": "Results",
                "columns": ["label", "amount"],
                "rows": [["Safe", 10], ["=2+2", 20], ["bad\ud800text", 30]],
            }
        ],
        "charts": [
            {
                "title": "Amounts",
                "kind": "bar",
                "table": 0,
                "x": "label",
                "y": "amount",
            }
        ],
    }
    value.update(changes)
    return ReportSummary.model_validate(value)


def test_report_writes_private_xlsx_and_escapes_formula_cells(tmp_path: Path) -> None:
    artifacts = ReportWriter(tmp_path).create(EXECUTION_ID, summary())

    workbook = openpyxl.load_workbook(artifacts.xlsx_path, data_only=False)
    worksheet = workbook["Results"]
    assert worksheet["A3"].value == "'=2+2"
    assert worksheet["A4"].value == "bad\ufffdtext"
    assert stat.S_IMODE(artifacts.xlsx_path.stat().st_mode) == 0o600
    assert artifacts.xlsx_path.parent == tmp_path / "reports" / str(EXECUTION_ID)


def test_html_is_self_contained_and_denies_network_access(tmp_path: Path) -> None:
    artifacts = ReportWriter(tmp_path).create(EXECUTION_ID, summary())
    html = artifacts.html_path.read_text(encoding="utf-8")

    assert "Content-Security-Policy" in html
    assert "default-src 'none'" in html
    assert "connect-src 'none'" in html
    assert "plotly.js" in html.casefold()
    assert re.search(r"<(?:script|img|link)[^>]+(?:src|href)\s*=", html, re.I) is None
    assert "fetch(" not in html
    assert "XMLHttpRequest" not in html
    assert "WebSocket" not in html
    assert "bad\ufffdtext" in html
    assert stat.S_IMODE(artifacts.html_path.stat().st_mode) == 0o600


def test_html_disables_worker_and_network_loaders_but_keeps_basic_chart(
    tmp_path: Path,
) -> None:
    artifacts = ReportWriter(tmp_path).create(EXECUTION_ID, summary())
    document = artifacts.html_path.read_text(encoding="utf-8")

    forbidden = (
        r"\bfetch\s*\(",
        r"\bXMLHttpRequest\b",
        r"\bWebSocket\b",
        r"\bimportScripts\s*\(",
        r"\b(?:Shared)?Worker\s*\(",
        r"\bEventSource\s*\(",
        r"\bsendBeacon\s*\(",
    )
    for pattern in forbidden:
        assert re.search(pattern, document, re.I) is None
    assert re.search(r"\.(?:src|href)\s*=\s*['\"]https?", document, re.I) is None
    assert "http://www.w3.org/2000/svg" in document
    assert "http://www.w3.org/1999/xhtml" in document
    assert "about:blank#www.w3.org" not in document
    assert '<div id="chart-0" class="chart"></div>' in document
    assert "Plotly.newPlot('chart-0'" in document


def test_summary_rejects_unknown_fields_and_unbounded_charts() -> None:
    with pytest.raises(ValidationError):
        summary(secret="not allowed")

    chart = {
        "title": "Amounts",
        "kind": "bar",
        "table": 0,
        "x": "label",
        "y": "amount",
    }
    with pytest.raises(ValidationError):
        summary(charts=[chart] * 9)


@pytest.mark.parametrize(
    "cell",
    [
        {"nested": "object"},
        ["nested", "array"],
        float("nan"),
        float("inf"),
        float("-inf"),
        "x" * 32768,
    ],
)
def test_summary_rejects_cells_excel_cannot_safely_represent(cell: object) -> None:
    value = summary().model_dump()
    value["tables"][0]["rows"] = [[cell, 1]]

    with pytest.raises(ReportError, match="^report_failed$"):
        parse_summary(value)


def test_summary_accepts_finite_json_scalars(tmp_path: Path) -> None:
    value = summary().model_dump()
    value["tables"] = [
        {
            "title": "Scalars",
            "columns": ["null", "boolean", "integer", "number", "text"],
            "rows": [[None, True, 42, 1.25, "safe"]],
        }
    ]
    value["charts"] = []

    artifacts = ReportWriter(tmp_path).create(EXECUTION_ID, parse_summary(value))

    workbook = openpyxl.load_workbook(artifacts.xlsx_path, data_only=False)
    assert [workbook["Scalars"].cell(2, column).value for column in range(1, 6)] == [
        None,
        True,
        42,
        1.25,
        "safe",
    ]


def test_report_rejects_preexisting_symlinked_report_root(tmp_path: Path) -> None:
    outside = tmp_path / "outside"
    outside.mkdir()
    os.symlink(outside, tmp_path / "reports")

    with pytest.raises(ReportError, match="^report_failed$"):
        ReportWriter(tmp_path).create(EXECUTION_ID, summary())
    assert list(outside.iterdir()) == []


def test_report_assigns_unique_sheet_names_for_duplicate_titles(tmp_path: Path) -> None:
    duplicate = {
        "title": "Results",
        "columns": ["label"],
        "rows": [["second"]],
    }
    value = summary(tables=[summary().tables[0].model_dump(), duplicate])

    artifacts = ReportWriter(tmp_path).create(EXECUTION_ID, value)

    workbook = openpyxl.load_workbook(artifacts.xlsx_path, read_only=True)
    assert len(workbook.sheetnames) == 3
    assert len({name.casefold() for name in workbook.sheetnames}) == 3
