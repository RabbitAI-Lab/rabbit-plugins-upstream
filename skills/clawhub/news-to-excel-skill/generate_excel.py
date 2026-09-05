
---

### 📄 2. scripts/generate_excel.py (辅助脚本)

```python
#!/usr/bin/env python3
"""
新闻信息转Excel辅助脚本
功能：将提取的结构化数据（CSV格式）输出为标准Excel文件
用法：python generate_excel.py --input data.csv --output output.xlsx
"""

import argparse
import csv
import sys
import os

def parse_arguments():
    parser = argparse.ArgumentParser(description="新闻信息CSV转Excel工具")
    parser.add_argument("--input", "-i", required=True, help="输入的CSV文件路径（UTF-8编码）")
    parser.add_argument("--output", "-o", required=True, help="输出的Excel文件路径（.xlsx）")
    return parser.parse_args()

def csv_to_excel(csv_filepath, excel_filepath):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ 错误：请先安装 openpyxl 库：pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "新闻汇总"

    # 读取CSV并写入Excel
    with open(csv_filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = next(reader)  # 读取表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row_idx, row in enumerate(reader, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    # 设置列宽
    ws.column_dimensions['A'].width = 18  # 时间列
    ws.column_dimensions['B'].width = 60  # 事件简介列

    # 冻结首行
    ws.freeze_panes = 'A2'

    wb.save(excel_filepath)
    print(f"✅ Excel文件已成功生成: {excel_filepath}")

def main():
    args = parse_arguments()
    csv_to_excel(args.input, args.output)

if __name__ == "__main__":
    main()