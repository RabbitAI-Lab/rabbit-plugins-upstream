#!/usr/bin/env python3
"""
统一入口：截图定位/搜索 → 按屏截图(+contact sheet) → Agent 视觉门控
        → 只下载勾选原图 → 分辨率/清晰度排序 → final/ + XLSX

工作流：
  0. python3 run.py check-env
  1. python3 run.py discover --query "<词>" --workdir ./out
       主通道：浏览器打开百度图片页 → DOM提取(URL+标题+坐标) → 按屏截图
       补充：Bing 结果拼成分页 contact sheet
       logo 类关键词自动触发权威源直取提示（见 source_router）
  2. Agent 看 screens/page_XX.png（按屏截图）或 contact_sheet_XX.jpg，勾选候选 id
  3. python3 run.py select --workdir ./out --ids 3,7,12 --prefix <前缀>
       只下载勾选编号的原图；文件名含候选 ID 防冲突；分辨率/清晰度排序
  4. python3 run.py verify --workdir ./out   （网页截屏复核，可选）
  5. python3 run.py report --workdir ./out   （重生成 XLSX）

所有路径通过 --workdir 传入，无工作目录硬编码。
"""
import sys, json, re, time, uuid
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import argparse

sys.path.insert(0, str(Path(__file__).parent))
import search_engines as se
from vision_pipeline import _download_image, build_sheet, sharpness_of

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


# ========== 场景识别 ==========

def classify_scene(query):
    """识别场景类别。logo 类自动触发权威源逻辑与纯净 logo 筛选。"""
    q = query.lower()
    if re.search(r'logo|标志|图标|icon|品牌标识|商标', q):
        return "logo"
    if re.search(r'发布会|大会|联名|营销|活动|名场面|官宣|代言|签约|揭牌|开幕', query):
        return "event"
    if re.search(r'大楼|大厦|总部|园区|建筑', query):
        return "building"
    if re.search(r'山|湖|河|公园|景区|瀑布|草原|沙漠|风景|古城|寺', query):
        return "scenery"
    if re.search(r'菜|席|宴|美食|小吃|餐厅|餐馆|招牌菜|味道', query):
        return "food"
    return "person"


def cmd_check_env(args):
    se.ensure_environment(verbose=True)
    ok_pw, _ = se.playwright_available()
    print()
    if not ok_pw:
        print("提示：主通道（截图定位）需要 playwright + Chromium：")
        print("  pip install playwright")
        print("  python3 -m playwright install chromium")
        print("  （Linux 沙箱如遇权限问题可加 --break-system-packages）")


# ========== discover ==========

def cmd_discover(args):
    workdir = Path(args.workdir)
    workdir.mkdir(parents=True, exist_ok=True)
    (workdir / "thumbs").mkdir(exist_ok=True)

    scene = classify_scene(args.query)
    print(f"[discover] 场景识别: {scene}", flush=True)
    if scene == "logo":
        print("[discover] logo 类关键词，执行 logo 专项流程：")
        print("  ① 建议先用 source_router 直取官网/App Store/自媒体头像等权威源")
        print("     （python3 source_router.py '<品牌名>' <权威源目录>）")
        print("     将得到的 URL 通过 --extra-file 注入，与搜索结果合并")
        print("  ② 视觉门控只选纯色/透明背景的标准 logo，")
        print("     挂 logo 的门店/大楼/产品/招牌一律不选")

    # 截图定位主通道 + Bing 补充
    items, screenshots = se.search_candidates(
        args.query, workdir,
        use_baidu_shot=not args.no_baidu,
        use_bing=not args.no_bing,
        bing_pages=args.bing_pages,
    )

    # 外部补充 URL（权威源直取 / 网页上下文验证结果）
    extra_urls = []
    if args.extra_urls:
        extra_urls.extend(args.extra_urls)
    if args.extra_file and Path(args.extra_file).exists():
        txt = Path(args.extra_file).read_text(encoding="utf-8").strip()
        try:
            data = json.loads(txt)
            if isinstance(data, list):
                for x in data:
                    if isinstance(x, str):
                        extra_urls.append(x)
                    elif isinstance(x, dict):
                        if x.get("url"):
                            extra_urls.append(x["url"])
                        elif x.get("objurl"):
                            extra_urls.append(x["objurl"])
        except json.JSONDecodeError:
            extra_urls.extend([u.strip() for u in txt.splitlines() if u.strip()])
    for u in extra_urls:
        items.append({
            "id": len(items), "url": u, "thumb_src": u,
            "fromurl": "", "title": "(权威源/补充)",
            "x": 0, "y": 0, "w": 0, "h": 0,
            "screen": -1, "idx_in_screen": -1,
            "engine": "extra", "page": "",
        })
    print(f"[discover] 含补充候选共 {len(items)} 条", flush=True)

    # 去重（按去参数基础 URL）
    seen, uniq = set(), []
    for it in items:
        key = re.sub(r'[?#].*$', '', it.get("url", ""))
        if not key or key in seen:
            continue
        seen.add(key)
        uniq.append(it)
    for i, it in enumerate(uniq):
        it["id"] = i
    print(f"[discover] 去重后 {len(uniq)} 条", flush=True)

    print(f"[discover] 生成 contact sheet（下载缩略图）...", flush=True)

    def fetch_thumb(idx_item):
        idx, it = idx_item
        url = it.get("thumb_src") or it.get("url")
        if not url:
            return None
        img, content = _download_image(url, referer=it.get("page") or it.get("fromurl"))
        if not img:
            return None
        w, h = img.size
        if w < 40 or h < 40:
            return None
        thumb = img.copy()
        thumb.thumbnail((240, 240))
        tp = workdir / "thumbs" / f"{idx:03d}.jpg"
        thumb.convert("RGB").save(tp, "JPEG", quality=85)
        it["_thumb"] = thumb
        it["thumb_path"] = str(tp)
        it["actual_w"], it["actual_h"] = w, h
        return it

    valid = []
    with ThreadPoolExecutor(max_workers=8) as ex:
        futs = [ex.submit(fetch_thumb, (i, it)) for i, it in enumerate(uniq)]
        for f in as_completed(futs):
            r = f.result()
            if r:
                valid.append(r)
    valid.sort(key=lambda x: x["id"])
    print(f"[discover] 缩略图成功 {len(valid)}/{len(uniq)}", flush=True)

    serializable = [{k: v for k, v in it.items() if not k.startswith("_")} for it in uniq]
    (workdir / "candidates.json").write_text(
        json.dumps(serializable, ensure_ascii=False, indent=2), encoding="utf-8")

    # 分页 contact sheet（每张最多 sheet-size 格，避免单图过大被内容过滤）
    sheet_paths = build_paged_sheets(valid, workdir, per_page=args.sheet_size,
                                     cols=args.cols, cell=260)

    print()
    print("[discover] 产物：", flush=True)
    if screenshots:
        print(f"  按屏截图（主）：{workdir}/screens/page_XX.png  共 {len(screenshots)} 屏", flush=True)
    for sp in sheet_paths:
        print(f"  contact sheet：{sp}", flush=True)
    print(f"  候选元数据：{workdir}/candidates.json", flush=True)
    print()
    print("下一步：Agent 用多模态视觉看截图/sheet，记录勾选的候选 id，然后 select --ids ...")
    print("（logo 类只选纯色/透明底标准 logo；事件/活动类存疑时用 verify 网页截屏复核）")
    print()
    print("[视觉能力探测] 视觉门控前，Agent 必须先做一次探测：")
    print("  用 fetch(type=file_path) 读取上面任一截图（如 screens/page_00.png），")
    print("  并请其描述图片内容。若能返回图片的画面描述（而非仅文件元数据），")
    print("  说明平台已桥接视觉能力，可继续视觉门控；")
    print("  若 fetch 无法获得图片内容描述，则当前环境不具备视觉理解能力，")
    print("  应停止任务并提示用户更换支持视觉的模型。")


def build_paged_sheets(valid, workdir, per_page=36, cols=6, cell=260):
    """分页生成 contact sheet，每张最多 per_page 个，避免单图过大。"""
    paths = []
    for start in range(0, len(valid), per_page):
        chunk = valid[start:start + per_page]
        page_no = start // per_page
        out = workdir / f"contact_sheet_{page_no+1:02d}.jpg"
        build_sheet(chunk, out, cols=cols, cell=cell)
        paths.append(str(out))
    return paths


# ========== select（只下载勾选原图；两步重命名防冲突丢图） ==========

def cmd_select(args):
    workdir = Path(args.workdir)
    candidates = json.loads((workdir / "candidates.json").read_text(encoding="utf-8"))
    ids = [int(x) for x in re.split(r"[,\s]+", args.ids) if x.strip().lstrip("-").isdigit()]
    print(f"[select] 勾选候选 id: {ids}", flush=True)

    final_dir = workdir / "final"
    final_dir.mkdir(exist_ok=True)
    raw_results = []

    for sid in ids:
        if sid < 0 or sid >= len(candidates):
            print(f"  [{sid:03d}] id 越界，跳过", flush=True)
            continue
        c = candidates[sid]
        origin_url = c.get("url") or c.get("objurl") or c.get("thumb_src")
        img, content = _download_image(origin_url, referer=c.get("page") or c.get("fromurl"))
        used_thumb_fallback = False
        if not img or len(content or b"") < 5000:
            tp = c.get("thumb_path")
            if tp and Path(tp).exists():
                from PIL import Image
                img = Image.open(tp)
                content = Path(tp).read_bytes()
                origin_url = "thumb_fallback"
                used_thumb_fallback = True
            else:
                print(f"  [{sid:03d}] 原图下载失败: {str(origin_url)[:60]}", flush=True)
                continue
        w, h = img.size
        sharpness = sharpness_of(img)
        ext = ".png" if img.mode in ("RGBA", "LA", "P") else ".jpg"
        raw_results.append({
            "id": sid,
            "img": img, "ext": ext,
            "url": origin_url, "thumb_fallback": used_thumb_fallback,
            "w": w, "h": h, "size_bytes": len(content),
            "sharpness": round(sharpness, 2) if sharpness else None,
            "title": c.get("title", ""), "page": c.get("page") or c.get("fromurl", ""),
            "engine": c.get("engine", ""),
        })
        print(f"  [{sid:03d}] {w}x{h} sharp={sharpness and round(sharpness,1)} "
              f"{'(缩略图回退)' if used_thumb_fallback else ''}", flush=True)

    def score(r):
        """客观质量排序：以短边分辨率为主，清晰度作为同分辨率下的加分项。

        注意：清晰度（Laplacian 全图梯度）对浅景深背景虚化、雪景/雾景/天空等
        低纹理场景会系统性偏低，这些恰恰常是好照片的特征，因此不能作为乘法
        惩罚因子，只做小幅 tie-break 加分。分辨率才是主排序键。
        """
        short_side = min(r["w"], r["h"])
        sharp = r.get("sharpness") or 0
        # 清晰度加分：sharp=10 给满 5% 分辨率加成；避免一张糊图靠分辨率登顶
        sharp_bonus = min(short_side * 0.05, short_side * (sharp / 200.0))
        # 缩略图回退的（只有 240px 小图）压到后面
        if r.get("thumb_fallback"):
            sharp_bonus -= 100000
        return short_side + sharp_bonus
    raw_results.sort(key=score, reverse=True)

    # 两步重命名，彻底避免同名冲突丢图：
    # 第一步全部存为临时唯一名（含候选 id + uuid），第二步再改为最终名
    temp_paths = []
    for rank, r in enumerate(raw_results, 1):
        tmp = final_dir / f".tmp_{r['id']}_{uuid.uuid4().hex[:8]}{r['ext']}"
        if r["ext"] == ".jpg":
            r["img"].convert("RGB").save(tmp, "JPEG", quality=92)
        else:
            r["img"].save(tmp)
        temp_paths.append((rank, r, tmp))

    final_records = []
    used_names = set()
    for rank, r, tmp in temp_paths:
        short = min(r["w"], r["h"])
        name = f"{args.prefix}_{rank:02d}_{short}px{r['ext']}"
        if name in used_names or (final_dir / name).exists():
            name = f"{args.prefix}_{rank:02d}_id{r['id']}_{short}px{r['ext']}"
        used_names.add(name)
        target = final_dir / name
        tmp.rename(target)
        r["filename"] = name
        r["path"] = str(target)
        final_records.append(r)
        print(f"  → {name}", flush=True)

    out_records = [{k: v for k, v in r.items() if k != "img"} for r in final_records]
    (workdir / "selected.json").write_text(
        json.dumps(out_records, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[select] 共 {len(final_records)} 张 → {workdir/'selected.json'}", flush=True)

    if HAS_XLSX and not args.no_xlsx:
        write_xlsx(workdir / "report.xlsx", args.query or "", candidates, out_records, ids,
                   scene=classify_scene(args.query))
        print(f"[select] XLSX → {workdir/'report.xlsx'}", flush=True)
    elif not HAS_XLSX:
        print("[select] openpyxl 未安装，pip install openpyxl 后可用", flush=True)


# ========== verify（网页截屏复核） ==========

def cmd_verify(args):
    """对选中图片的来源页做整页截屏 + 图注/上下文提取，确认是否准确配图。"""
    import webctx_verify
    workdir = Path(args.workdir)
    sel_path = workdir / "selected.json"
    if not sel_path.exists():
        print("[verify] 未找到 selected.json，请先 select")
        return
    selected = json.loads(sel_path.read_text(encoding="utf-8"))
    webctx_verify.verify_sources(selected, workdir / "verify", max_pages=args.max_pages)


# ========== XLSX ==========

def write_xlsx(path, query, all_candidates, selected, selected_ids, scene=""):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "选中图片"
    headers = ["排名", "候选id", "文件名", "宽", "高", "短边", "大小KB", "清晰度",
               "引擎", "标题", "来源页", "原图URL", "缩略图回退"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2F5496")
    for rank, r in enumerate(selected, 1):
        ws.append([
            rank, r.get("id"), r.get("filename", ""), r.get("w"), r.get("h"),
            min(r.get("w", 0), r.get("h", 0)),
            round(r.get("size_bytes", 0) / 1024, 1),
            r.get("sharpness") or "",
            r.get("engine", ""),
            r.get("title", ""),
            r.get("page", ""),
            r.get("url", ""),
            "是" if r.get("thumb_fallback") else "",
        ])
    for col, width in zip("ABCDEFGHIJKLM",
                          [6, 8, 34, 8, 8, 8, 10, 8, 12, 36, 46, 56, 10]):
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet("全部候选")
    ws2.append(["候选id", "选中", "引擎", "屏号", "屏内序", "实际宽", "实际高",
                "标题", "来源页", "URL"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="548235")
    sel_set = set(selected_ids)
    for c in all_candidates:
        ws2.append([
            c.get("id"), "Y" if c.get("id") in sel_set else "",
            c.get("engine", ""),
            c.get("screen", ""), c.get("idx_in_screen", ""),
            c.get("actual_w", ""), c.get("actual_h", ""),
            c.get("title", ""),
            c.get("page") or c.get("fromurl", ""),
            c.get("url", ""),
        ])
    for col, width in zip("ABCDEFGHIJ", [8, 6, 12, 6, 6, 8, 8, 36, 46, 56]):
        ws2.column_dimensions[col].width = width

    ws3 = wb.create_sheet("任务信息")
    ws3.append(["搜索词", query])
    ws3.append(["场景类别", scene])
    ws3.append(["候选总数", len(all_candidates)])
    ws3.append(["选中数量", len(selected)])
    ws3.append(["生成时间", time.strftime("%Y-%m-%d %H:%M:%S")])
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 60
    wb.save(path)


def cmd_report(args):
    workdir = Path(args.workdir)
    candidates = json.loads((workdir / "candidates.json").read_text(encoding="utf-8"))
    sel_path = workdir / "selected.json"
    selected = json.loads(sel_path.read_text(encoding="utf-8")) if sel_path.exists() else []
    ids = [r.get("id") for r in selected]
    if HAS_XLSX:
        write_xlsx(workdir / "report.xlsx", args.query or "unknown",
                   candidates, selected, ids, scene=classify_scene(args.query or ""))
        print(f"报告: {workdir / 'report.xlsx'}")
    else:
        print("openpyxl 未安装")


def main():
    ap = argparse.ArgumentParser(description="多模态视觉门控图片下载器")
    ap.add_argument("--workdir", default="./out", help="工作目录（所有产物在此）")
    sub = ap.add_subparsers(dest="cmd", required=True)

    sub.add_parser("check-env", help="环境自检").set_defaults(func=cmd_check_env)

    d = sub.add_parser("discover", help="截图定位/搜索 + 生成 contact sheet")
    d.add_argument("--query", required=True)
    d.add_argument("--cols", type=int, default=6)
    d.add_argument("--sheet-size", type=int, default=36, help="每张 contact sheet 最多格子数")
    d.add_argument("--bing-pages", type=int, default=1)
    d.add_argument("--no-bing", action="store_true")
    d.add_argument("--no-baidu", action="store_true")
    d.add_argument("--extra-urls", nargs="*")
    d.add_argument("--extra-file")
    d.set_defaults(func=cmd_discover)

    s = sub.add_parser("select", help="按候选 id 下载原图 + 生成报告")
    s.add_argument("--ids", required=True, help="逗号分隔候选 id，如 3,7,12")
    s.add_argument("--prefix", default="photo")
    s.add_argument("--query", default="")
    s.add_argument("--no-xlsx", action="store_true")
    s.set_defaults(func=cmd_select)

    v = sub.add_parser("verify", help="网页截屏复核来源页")
    v.add_argument("--max-pages", type=int, default=6)
    v.set_defaults(func=cmd_verify)

    r = sub.add_parser("report", help="重新生成 XLSX 报告")
    r.add_argument("--query", default="")
    r.set_defaults(func=cmd_report)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
