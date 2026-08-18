#!/usr/bin/env python3
"""
PDF 增值税发票提取 v2.0.0
- 修复：通行费 PDF 文本跨行 bug（合 计 / 金额分两行）
- 新增：命令行参数支持（python3 invoice_extractor_v2.py <PDF> [输出.xlsx]）
- 新增：康熙字典 ** → 生/海/消 等修复
- 新增：水印页发票号/日期/税号兜底
- 新增：多税率留空
- 布局感知（基于 pdfplumber 坐标分析）
- 支持增值税电子发票（普通发票）+ 通行费 + 其他 PDF
- 输出 15 列 Excel：序号/发票号/日期/购方/购方税号/销售方/销售方税号/项目/金额/税率/税额/价税合计/备注/报销人/报销日期
"""
import sys
import pdfplumber, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

PDF_PATH = sys.argv[1] if len(sys.argv) >= 2 else "invoices.pdf"
OUTPUT_PATH = sys.argv[2] if len(sys.argv) >= 3 else (PDF_PATH.rsplit(".", 1)[0] + "_发票统计.xlsx")

COL_SPLIT = 295; COL_SPLIT_RIGHT = 302

KANGXI = {'\u2f63':'\u751f','\u2f53':'\u6d77','\u2f49':'\u6708','\u2f47':'\u65e5',
          '\u2f43':'\u4ea7','\u2f55':'\u6d88','\u2f41':'\u7518','\u2f4b':'\u76ee'}

def norm(t):
    if not t: return t
    for k,v in KANGXI.items(): t = t.replace(k, v)
    return t

def parse_float(s):
    if not s: return 0.0
    s = str(s).replace(",","").replace("¥","").strip()
    m = re.search(r'[-+]?\d*\.\d+|\d+', s); return float(m.group()) if m else 0.0

def normalize_date(text):
    if not text: return ""
    m = re.search(r'开票日期[：:\s]*([^\n]{6,30})', text)
    if m:
        raw = re.sub(r'\s+', '', m.group(1).strip())
        dm = re.search(r'(\d{4})年(\d{2})[\u6708\u2F49\u4e00-\u9fff]*(\d{2})[\u65e5\u2F47\u4e00-\u9fff]*', raw)
        if dm: return f"{dm.group(1)}-{dm.group(2)}-{dm.group(3)}"
    wm = re.search(r'国家税务总局\s*(\d{4})年(\d{2})月(\d{2})日', text)
    if wm: return f"{wm.group(1)}-{wm.group(2)}-{wm.group(3)}"
    return ""

def _extract_num(s):
    '''从含¥的字符串中提取数字（小数优先，其次整数）'''
    m = re.search(r'¥\s*([0-9,]+\.\d+)', s)
    if m: return m.group(1)
    m = re.search(r'¥\s*([0-9,]+)', s)
    return m.group(1) if m else None

def extract_amounts(text):
    # 先在合汁行提取
    heji_match = re.search(r'(?<!价税)合\s*计[ \t]+([^\n]+)', text)
    if heji_match:
        heji_line = heji_match.group(1)
        nums = re.findall(r'¥\s*([0-9,]+\.\d+|\*+|\d+)', heji_line)
        if nums:
            amount_raw = nums[0].replace(',','')
            # 免税/不征税：¥后直接是***，或只有一项且为整数
            if amount_raw == '***' or re.match(r'^\d+$', amount_raw) and len(nums) == 1:
                # 检查是否为免税/不征税（¥后直接是***）
                if '***' in heji_line or '免税' in text or '不征税' in text:
                    jshj_m = re.search(r'价税合计[^¥]*¥\s*([0-9,]+\.\d+)', text)
                    total = parse_float(jshj_m.group(1).replace(',','')) if jshj_m else 0.0
                    return 0.0, 0.0, total
            amount = float(amount_raw) if re.search(r'\.', amount_raw) else float(amount_raw)
            tax_raw = nums[1].replace(',','') if len(nums) > 1 else '0'
            tax = float(tax_raw) if tax_raw != '***' else 0.0
            jshj_m = re.search(r'价税合计[^¥]*¥\s*([0-9,]+\.\d+)', text)
            total = parse_float(jshj_m.group(1).replace(',','')) if jshj_m else 0.0
            return amount, tax, total
    # 回退：逐行扫描
    lines = text.split('\n')
    for i, line in enumerate(lines):
        if line.strip().startswith('¥'):
            nums = re.findall(r'¥\s*([0-9,]+\.\d+|\*+|\d+)', line)
            if nums:
                amount_raw = nums[0].replace(',',''); amount = float(amount_raw) if '.' in amount_raw else float(amount_raw)
                tax_raw = nums[1].replace(',','') if len(nums) > 1 else '0'
                tax = float(tax_raw) if tax_raw != '***' else 0.0
                for j in range(i+1, len(lines)):
                    jm = re.search(r'¥\s*([0-9,]+\.\d+)', lines[j])
                    if jm: return amount, tax, float(jm.group(1).replace(',',''))
                return amount, tax, 0.0
    return 0.0, 0.0, 0.0

def extract_tax_rate(text):
    # 先统计税率数量，用于判断是否多税率
    rates = re.findall(r'(\d+)%', text)
    unique = sorted(set(rates))
    if len(unique) == 0:
        # 无数值税率，再检查特殊标签
        if '不征税' in text: return '不征税'
        if '免税' in text: return '免税'
        return ""
    if len(unique) > 1:
        return ""  # 多税率，无论是否含免税/不征税字样，均留空
    # 单一数值税率
    return f"{unique[0]}%" 

def extract_item_name(text, page_words):
    """修复：预处理`**`合并 + Kangxi标准化 + top分组合并"""
    header_y = col_end_x = None
    for w in page_words:
        if '项目名称' in w['text']: header_y = w['top']; break
    for w in page_words:
        if header_y and abs(w['top']-header_y)<3 and '规格' in w['text']:
            col_end_x = w['x0']; break
    if header_y is None: return ""
    ITEM_END_X = col_end_x or 120
    x0_min = 200
    for w in page_words:
        if w['top'] > header_y and w['text'].startswith('*'):
            x0_min = min(x0_min, w['x0'])
    top_groups = defaultdict(list)
    for w in page_words:
        if w['top'] > header_y and x0_min <= w['x0'] <= ITEM_END_X:
            top_groups[int(w['top'])].append((w['x0'], norm(w['text'])))
    sorted_keys = sorted(top_groups.keys())
    merged = []
    for k in sorted_keys:
        if merged and (k - merged[-1][0]) <= 20:
            merged[-1] = (merged[-1][0], merged[-1][1] + top_groups[k])
        else:
            merged.append((k, top_groups[k]))

    def parse_star_word(txt):
        stars = [i for i, c in enumerate(txt) if c == '*']
        if len(stars) < 2: return txt
        double = any(stars[i]+1 == stars[i+1] for i in range(len(stars)-1))
        if double:
            sl, last = stars[-2], stars[-1]
            cat, name = txt[1:sl], txt[last+1:]
        else:
            second = stars[1]; cat, name = txt[1:second], txt[second+1:]
        return f"*{cat}*{name}" if cat and name else txt

    seen = set(); results = []
    for k, words_list in merged:
        line_text = ''.join(w[1] for w in sorted(words_list))
        if '*' not in line_text: continue
        # 修复：预处理`**`→`*`，避免regex在`**`后无法匹配
        line_text_normalized = re.sub(r'\*\*', '*', line_text)
        for m in re.finditer(r'\*([^\*]+)\*([^\s\*]+)', line_text_normalized):
            full = parse_star_word(m.group())
            if full and full not in seen: seen.add(full); results.append(full)
    return '\n'.join(results) if results else ""

def concat_name(words, start_word, max_x):
    if not start_word: return ""
    top = start_word['top']
    relevant = [w for w in words if abs(w['top']-top)<4 and start_word['x0']<=w['x0']<=max_x]
    relevant.sort(key=lambda w: w['x0'])
    raw = ''.join(norm(w['text']) for w in relevant)
    return re.sub(r'^名称[：:]\s*','',raw).strip()

def find_invoice_number(text):
    m = re.search(r'发票号码[：:\s]*(\d{10,})', text)
    if m: return m.group(1)
    m = re.search(r'(?:监\s*制\s*)(\d{10,})', text)
    return m.group(1) if m else None

def find_tax_code_in_zone(words, zone_x_min, zone_x_max, top_hint=None, exclude_watermark=False):
    """在指定x范围内查找税号，支持水印页兜底"""
    for w in words:
        if not (zone_x_min <= w['x0'] <= zone_x_max): continue
        if top_hint is not None and abs(w['top'] - top_hint) > 30: continue
        if exclude_watermark and w["top"] < 55: continue
        cm = re.search(r'([A-Z0-9]{10,})', w['text'])
        if cm: return cm.group(1)
    return None

def extract_train_ticket_info(text, page_words=None):
    """v2.4.0 新增：从 pdfplumber 文本抽取火车票字段
    返回字典（包含新规则的字段）：
      item_name: "火车票" / "火车票-退票费"
      total: 价税合计
      rate: "9%" / "6%"
      tax: 税额（反算）
      amount: 金额（=total-tax）
      buyer: 购买方名称
      buyer_code: 购买方税号
      train_no, start, end: 车次/起点/终点（用于备注）
    """
    result = {
        'item_name': '火车票',
        'total': 0.0,
        'rate': '9%',
        'tax': 0.0,
        'amount': 0.0,
        'buyer': None,
        'buyer_code': None,
        'train_no': '',
        'start': '',
        'end': '',
    }
    
    # 1. 区分票价 vs 退票费
    is_refund = '退票费' in text
    if is_refund:
        # 退票费场景：税率 6%
        m = re.search(r'退票费[:：]\s*[￥¥]\s*(\d+\.\d{2})', text)
        if m:
            result['total'] = float(m.group(1))
            result['rate'] = '6%'
            result['item_name'] = '火车票-退票费'
    else:
        # 票价场景：税率 9%
        m = re.search(r'票价[:：]\s*[￥¥]\s*(\d+\.\d{2})', text)
        if m:
            result['total'] = float(m.group(1))
            result['rate'] = '9%'
            result['item_name'] = '火车票'
    
    # fallback: 退票费标签在但金额正则没匹配上（兼容 PDF 上"退票费:"没金额的情况）
    if result['total'] == 0:
        m = re.search(r'[￥¥]\s*(\d+\.\d{2})', text)
        if m:
            result['total'] = float(m.group(1))
            if is_refund:
                # 退票费场景，rate 和 item_name 也要更新
                result['rate'] = '6%'
                result['item_name'] = '火车票-退票费'
            # 按 is_refund 决定税率
    
    # 2. 计算税额和金额
    if result['total'] > 0:
        rate_decimal = 0.06 if is_refund else 0.09
        result['tax'] = round(result['total'] / (1 + rate_decimal) * rate_decimal, 2)
        result['amount'] = round(result['total'] - result['tax'], 2)
    
    # 3. 提取购方名称
    # 例: "购买方名称:伊登优格(上海)咨询管理有限公司 统一社会信用代码:91310000MAK7RM1U9D"
    m = re.search(r'购买方名称[:：]\s*([^统一\n]+?)(?=\s*统一|$)', text, re.MULTILINE)
    if m:
        result['buyer'] = m.group(1).strip()
    
    # 4. 提取购方税号
    m = re.search(r'统一社会信用代码[:：]\s*([A-Z0-9]{18,20})', text)
    if m:
        result['buyer_code'] = m.group(1)
    
    # 5. 提取车次 + 起点终点
    # 例: "蚌埠南 G2667 南京南" 或 "上海虹桥 G1235 杭州东"
    m = re.search(r'([\u4e00-\u9fa5]{2,5}站?)\s+([GDC]\d{1,4})\s+([\u4e00-\u9fa5]{2,5}站?)', text)
    if m:
        result['start'] = m.group(1)
        result['train_no'] = m.group(2)
        result['end'] = m.group(3)
    
    # 6. 构建备注（起终点缺失时省略，自动补"站"字，不显示 "XX站" 占位符）
    def add_station_suffix(name):
        if not name: return ''
        return name if name.endswith('站') else name + '站'
    
    note_parts = []
    start_disp = add_station_suffix(result['start'])
    end_disp = add_station_suffix(result['end'])
    if start_disp and end_disp:
        note_parts.append(f"{start_disp}-{end_disp}")
    elif start_disp:
        note_parts.append(start_disp)
    if result['train_no']:
        note_parts.append(f"车次{result['train_no']}")
    result['note'] = '，'.join(note_parts)
    
    return result


def get_page_info(page, text, page_num):
    page_words = page.extract_words() or []
    inv = find_invoice_number(text)
    if not inv: return None
    date_str = normalize_date(text)
    
    # 火车票识别（v3.0 新增，v2.2 集成 PaddleOCR）
    # 特征：含 "电子客票"、"中国铁路" 或 "国铁" 字样
    is_train_ticket = bool(re.search(r'电子客票|中国铁路|国铁|12306', text))
    
    # v2.3 新增：PP-StructureV3 兜底（适合复杂版面发票）
    # v2.2 PaddleOCR 兜底（适合火车票）
    if not text.strip() or len(text.strip()) < 20:
        try:
            from pp_structure import extract_vat_invoice_pp_structure
            import pymupdf
            pdf_path = page.parent.filename if hasattr(page, 'parent') else None
            if pdf_path and pdf_path.lower().endswith('.pdf'):
                # 转图片
                doc = pymupdf.open(pdf_path)
                pix = doc[page_num].get_pixmap(matrix=pymupdf.Matrix(2.5, 2.5))
                tmp_path = f"/tmp/inv_pp_{os.getpid()}.jpg"
                pix.save(tmp_path)
                doc.close()
                pp_result = extract_vat_invoice_pp_structure(tmp_path)
                os.remove(tmp_path)
                if pp_result.get('发票号码'):
                    return {
                        'invoice_id': pp_result['发票号码'],
                        'date': pp_result['开票日期'],
                        'buyer': pp_result.get('购买方信息-名称', ''),
                        'buyer_code': pp_result.get('购买方信息-统一社会信用代码', ''),
                        'seller': pp_result.get('销售方信息-名称', ''),
                        'seller_code': pp_result.get('销售方信息-统一社会信用代码', ''),
                        'amount': pp_result.get('金额', 0),
                        'tax': pp_result.get('税额', 0),
                        'total': pp_result.get('价税合计', 0),
                        'rate': pp_result.get('税率/征收率', ''),
                        'item_name': ' '.join(pp_result.get('项目名称', [])),
                        'page_num': page_num, 'is_last': True,
                    }
        except Exception:
            pass
        
        try:
            from paddle_ocr import extract_train_ticket
            train_result = extract_train_ticket(page.parent.filename if hasattr(page, 'parent') else None)
            if train_result.get('is_train_ticket'):
                fields = train_result['fields']
                raw_text = train_result.get('raw_text', '')
                is_refund = '退票费' in raw_text
                total = fields.get('票价', 0) or 0
                rate_decimal = 0.06 if is_refund else 0.09
                tax = round(total / (1 + rate_decimal) * rate_decimal, 2) if total else 0
                amount = round(total - tax, 2) if total else 0
                
                # 备注：起点-终点，车次GXXXX
                note_parts = []
                if fields.get('起点') and fields.get('终点'):
                    note_parts.append(f"{fields['起点']}-{fields['终点']}")
                elif fields.get('起点'):
                    note_parts.append(fields['起点'])
                if fields.get('车次'):
                    note_parts.append(f"车次{fields['车次']}")
                
                return {
                    'invoice_id': fields.get('发票号码', ''),
                    'date': fields.get('开票日期', ''),
                    'buyer': fields.get('购方', ''),
                    'buyer_code': fields.get('购方税号', ''),
                    'seller': None, 'seller_code': None,
                    'amount': amount, 'tax': tax, 'total': total,
                    'rate': f'{int(rate_decimal*100)}%',
                    'item_name': '火车票-退票费' if is_refund else '火车票',
                    'note': '，'.join(note_parts),
                    'paddle_ocr': True,
                    'page_num': page_num, 'is_last': True,
                }
        except Exception as e:
            pass  # PaddleOCR 失败，回退到常规路径
    
    if is_train_ticket:
        # v2.4.0 新规则：购方/税号/金额/税额/税率/起终点车次全填
        info = extract_train_ticket_info(text, page_words)
        
        # 如果购方/税号未识别到，fallback 到 PaddleOCR
        if (not info['buyer'] or not info['buyer_code']) and page.parent:
            try:
                from paddle_ocr import ocr_pdf_page, is_train_ticket_ocr, extract_train_ticket_fields
                import pymupdf
                pdf_path = page.parent.filename
                if pdf_path and pdf_path.lower().endswith('.pdf'):
                    doc = pymupdf.open(pdf_path)
                    pix = doc[page_num].get_pixmap(matrix=pymupdf.Matrix(2.5, 2.5))
                    tmp_path = f"/tmp/inv_ocr_{os.getpid()}_{page_num}.jpg"
                    pix.save(tmp_path)
                    doc.close()
                    from paddle_ocr import ocr_image
                    ocr_items = ocr_image(tmp_path)
                    os.remove(tmp_path)
                    if is_train_ticket_ocr(ocr_items):
                        ocr_fields = extract_train_ticket_fields(ocr_items)
                        if not info['buyer'] and ocr_fields.get('购方'):
                            info['buyer'] = ocr_fields['购方']
                        if not info['buyer_code'] and ocr_fields.get('购方税号'):
                            info['buyer_code'] = ocr_fields['购方税号']
                        if not info['train_no'] and ocr_fields.get('车次'):
                            info['train_no'] = ocr_fields['车次']
                            info['note'] = (info['start'] + '-' + info['end'] if info['start'] and info['end'] else '') + ('，车次' + ocr_fields['车次'] if (info['start'] or info['end']) else '车次' + ocr_fields['车次'])
                            info['note'] = info['note'].strip('，')
            except Exception:
                pass
        
        return {
            'invoice_id': inv, 'date': date_str,
            'buyer': info['buyer'], 'buyer_code': info['buyer_code'],
            'seller': None, 'seller_code': None,
            'amount': info['amount'], 'tax': info['tax'], 'total': info['total'],
            'rate': info['rate'], 'item_name': info['item_name'],
            'note': info['note'],
            'page_num': page_num, 'is_last': True,
        }
    
    # 购买方/销售方名称
    name_w_buyer = name_w_seller = None
    for w in page_words:
        if ('名称' in w['text'] and '统一' not in w['text']
            and not w['text'].startswith('项目名称')):
            if w['x0'] < COL_SPLIT and not name_w_buyer: name_w_buyer = w
            elif w['x0'] > COL_SPLIT_RIGHT and not name_w_seller: name_w_seller = w
    buyer_name = concat_name(page_words, name_w_buyer, COL_SPLIT-1)
    seller_name = concat_name(page_words, name_w_seller, 600)
    
    # 税号：先在标签词内找，再在水印页兜底
    buyer_code = seller_code = ""
    # 标准逻辑：在标签词内找税号
    for w in page_words:
        if '统一社会信用代码' in w['text'] or '纳税人识别号' in w['text']:
            cm = re.search(r'([A-Z0-9]{10,})', w['text'])
            if cm:
                if w['x0'] < COL_SPLIT and not buyer_code: buyer_code = cm.group(1)
                elif w['x0'] > COL_SPLIT_RIGHT and not seller_code: seller_code = cm.group(1)
    # 水印页兜底：在zone内找18位字母数字
    if not buyer_code:
        bc = find_tax_code_in_zone(page_words, 0, COL_SPLIT-1)
        if bc: buyer_code = bc
    if not seller_code:
        sc = find_tax_code_in_zone(page_words, COL_SPLIT_RIGHT+1, 700, exclude_watermark=True)
        if sc: seller_code = sc
    
    amount, tax, total = extract_amounts(text)
    rate_str = extract_tax_rate(text)
    item_name = extract_item_name(text, page_words)
    is_last = bool(re.search(r'价税合计[^\n]*¥\s*\d', text))
    return {
        'invoice_id': inv, 'date': date_str,
        'buyer': buyer_name, 'buyer_code': buyer_code,
        'seller': seller_name, 'seller_code': seller_code,
        'amount': amount, 'tax': tax, 'total': total,
        'rate': rate_str, 'item_name': item_name,
        'page_num': page_num, 'is_last': is_last,
    }

def extract_all_invoices(pdf_path):
    inv_pages = defaultdict(list)
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            try: text = page.extract_text()
            except: continue
            if not text: continue
            info = get_page_info(page, text, page_num)
            if info: inv_pages[info['invoice_id']].append(info)
    invoices = []
    for inv, pages in inv_pages.items():
        if len(pages) == 1:
            p = pages[0]
            total = p['total'] if p['total'] > 0 else p['amount']
            invoices.append({
                '发票号码': inv, '开票日期': p['date'],
                '购买方信息-名称': p['buyer'],
                '购买方信息-统一社会信用代码': p['buyer_code'],
                '销售方信息-名称': p['seller'],
                '销售方信息-统一社会信用代码': p['seller_code'],
                '项目名称': p['item_name'],
                '金额': p['amount'], '税率/征收率': p['rate'],
                '税额': p['tax'], '价税合计': total,
                '备注': p.get('note', ''),
            })
        else:
            pages.sort(key=lambda x: x['page_num'])
            last = pages[-1]; first = pages[0]
            total = last['total'] if last['total'] > 0 else last['amount']
            all_rates = [p['rate'] for p in pages if p['rate']]
            unique = sorted(set(all_rates))
            final_rate = unique[0] if len(unique)==1 else ('' if len(unique)>1 else '')
            invoices.append({
                '发票号码': inv, '开票日期': first['date'],
                '购买方信息-名称': first['buyer'],
                '购买方信息-统一社会信用代码': first['buyer_code'],
                '销售方信息-名称': first['seller'],
                '销售方信息-统一社会信用代码': first['seller_code'],
                '项目名称': first['item_name'],
                '金额': last['amount'], '税率/征收率': final_rate,
                '税额': last['tax'], '价税合计': total,
                '备注': first.get('note', ''),
            })
    return invoices

def detect_duplicates(invoices):
    dupes = []
    for i, inv in enumerate(invoices):
        for j, other in enumerate(invoices):
            if i >= j: continue
            if (inv['发票号码']==other['发票号码'] and inv['开票日期']==other['开票日期']
                and inv['购买方信息-名称']==other['购买方信息-名称']
                and inv['销售方信息-名称']==other['销售方信息-名称']
                and abs(inv['金额']-other['金额'])<0.01
                and abs(inv['价税合计']-other['价税合计'])<0.01):
                dupes.extend([i,j])
    return sorted(set(dupes))

def format_rate(rate_str):
    if not rate_str: return ""
    if rate_str in ('不征税','免税'): return rate_str
    if '%' in rate_str:
        try: return float(rate_str.replace('%',''))/100
        except: return rate_str
    return rate_str

def create_excel(invoices, duplicate_rows, output_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "发票统计"
    headers = ['序号','发票号码','开票日期','购买方信息-名称',
               '购买方信息-统一社会信用代码','销售方信息-名称',
               '销售方信息-统一社会信用代码','项目名称',
               '金额','税率/征收率','税额','价税合计','备注','报销人','报销日期']
    hdr_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
    hdr_font = Font(bold=True)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = hdr_align
    yellow_fill = PatternFill(fill_type='solid', fgColor='FFFF00')
    for row_idx, inv in enumerate(invoices, 1):
        row_num = row_idx + 1; is_dup = (row_idx-1) in duplicate_rows
        rate_val = format_rate(inv['税率/征收率'])
        row_data = [row_idx, inv['发票号码'], inv['开票日期'],
                    inv['购买方信息-名称'], inv['购买方信息-统一社会信用代码'],
                    inv['销售方信息-名称'], inv['销售方信息-统一社会信用代码'],
                    inv['项目名称'], inv['金额'], rate_val, inv['税额'],
                    inv['价税合计'], inv.get('备注', ''), '', '']
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            if is_dup: c.fill = yellow_fill
            if col==2: c.number_format='@'
            elif col==3: c.number_format='YYYY-MM-DD'
            elif col in [5,7]: c.number_format='@'
            elif col in [9,11,12]: c.number_format='#,##0.00'
            elif col==10 and isinstance(val,float): c.number_format='0.00%'
            elif col==8: c.alignment=Alignment(wrap_text=True,vertical='top')
    ws.freeze_panes='A2'
    for col in range(1,len(headers)+1):
        letter=get_column_letter(col)
        max_len=len(headers[col-1])
        for row in range(2,len(invoices)+2):
            v=ws.cell(row=row,column=col).value
            if v: max_len=min(max(max_len,len(str(v))),50)
        ws.column_dimensions[letter].width=max_len+3
    wb.save(output_path)

if __name__=='__main__':
    invoices = extract_all_invoices(PDF_PATH)
    empty_date = [i+1 for i,inv in enumerate(invoices) if not inv['开票日期']]
    empty_buyer = [i+1 for i,inv in enumerate(invoices) if not inv['购买方信息-名称']]
    empty_buyer_code = [i+1 for i,inv in enumerate(invoices) if not inv['购买方信息-统一社会信用代码']]
    zero_amt = [i+1 for i,inv in enumerate(invoices) if inv['金额']==0]
    empty_item = [i+1 for i,inv in enumerate(invoices) if not inv['项目名称']]
    print(f"总数:{len(invoices)} | 日期空:{len(empty_date)} | 购方空:{len(empty_buyer)} | 购方税号空:{len(empty_buyer_code)} | 金额0:{len(zero_amt)} | 项目空:{len(empty_item)}")
    if empty_date: print(f"  日期空: {[invoices[i-1]['发票号码'] for i in empty_date]}")
    if empty_buyer_code: print(f"  购方税号空: {[invoices[i-1]['发票号码'] for i in empty_buyer_code]}")
    if empty_item: print(f"  项目空: {[invoices[i-1]['发票号码'] for i in empty_item]}")
    # 验证问题发票
    check_invs = {'26312000004018164286','26352000001516076731','26512000002667610411',
                  '26127000000297200980','26317000002248685472'}
    for inv_check in check_invs:
        for inv in invoices:
            if inv['发票号码'] == inv_check:
                print(f"  {inv_check}: 项目={inv['项目名称'][:30] if inv['项目名称'] else '(空)'!r}, 购方税号={inv['购买方信息-统一社会信用代码']!r}, 金额={inv['金额']}, 税率={inv['税率/征收率']!r}, 价税合计={inv['价税合计']}")
    dupes = detect_duplicates(invoices)
    print(f"重复:{len(dupes)}行")
    create_excel(invoices, dupes, OUTPUT_PATH)
    from collections import Counter
    rates = Counter(inv['税率/征收率'] for inv in invoices)
    print(f"税率:{dict(sorted(rates.items(),key=lambda x:-x[1]))}")

