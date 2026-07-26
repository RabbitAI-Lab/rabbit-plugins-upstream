#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
med-doc-preprocess — 医疗文档预处理公共库

将 pdf/doc/docx/xls/xlsx/csv/txt/json/图片 统一加载为标准中间产物（artifact），
供其他 skill 的 run.py 导入使用。

公开接口：
    PreprocessError
    TEXT_FILE_TYPES, IMAGE_FILE_TYPES, TABLE_FILE_TYPES, SUPPORTED_FILE_TYPES
    detect_input_type(path, explicit) -> str
    normalize_header(value) -> str
    load_input_artifact(path, input_type, encoding, sheet_name, *, pdf_as_single_text) -> dict
    extract_pdf_pages(path) -> List[str]
    extract_pdf_text(path) -> str
    extract_image_text(path) -> str
"""

import csv
import json
import os
import re
import subprocess
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook as _load_workbook
except Exception:
    _load_workbook = None  # type: ignore[assignment]

try:
    from pypdf import PdfReader as _PdfReader
except Exception:
    _PdfReader = None  # type: ignore[assignment]


# ---------------------------------------------------------------------------
# 公开常量
# ---------------------------------------------------------------------------

TEXT_FILE_TYPES: frozenset = frozenset({"txt", "md"})
IMAGE_FILE_TYPES: frozenset = frozenset({"png", "jpg", "jpeg", "bmp", "tif", "tiff"})
TABLE_FILE_TYPES: frozenset = frozenset({"csv", "xlsx", "xls"})
SUPPORTED_FILE_TYPES: frozenset = (
    TEXT_FILE_TYPES | IMAGE_FILE_TYPES | TABLE_FILE_TYPES | frozenset({"json", "pdf", "doc", "docx"})
)


# ---------------------------------------------------------------------------
# 异常
# ---------------------------------------------------------------------------

class PreprocessError(ValueError):
    """文件读取或预处理失败时抛出。"""


# ---------------------------------------------------------------------------
# 辅助工具
# ---------------------------------------------------------------------------

def detect_input_type(path: Path, explicit: str) -> str:
    """从文件后缀或显式参数推断输入类型字符串。"""
    if explicit != "auto":
        return explicit
    suffix = path.suffix.lower().lstrip(".")
    if suffix in SUPPORTED_FILE_TYPES:
        return suffix
    raise PreprocessError(f"Unsupported input file type: {path.suffix or '(none)'}")


def normalize_header(value: str) -> str:
    """去除空白/连字符并转小写，用于列名模糊匹配。"""
    return re.sub(r"[\s_\-]+", "", value.strip().lower())


def _shutil_which(name: str) -> Optional[str]:
    for directory in os.environ.get("PATH", "").split(os.pathsep):
        candidate = Path(directory) / name
        if candidate.exists() and os.access(candidate, os.X_OK):
            return str(candidate)
    return None


def _detect_tesseract_langs(tesseract_bin: str) -> Sequence[str]:
    proc = subprocess.run(
        [tesseract_bin, "--list-langs"],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        return []
    return [line.strip() for line in proc.stdout.splitlines()[1:] if line.strip()]


# ---------------------------------------------------------------------------
# 文件读取函数
# ---------------------------------------------------------------------------

def _read_text_file(path: Path, encoding: str) -> str:
    try:
        return path.read_text(encoding=encoding)
    except UnicodeDecodeError:
        return path.read_text(encoding="utf-8-sig")


def _read_json_file(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _read_csv_rows(path: Path, encoding: str) -> List[List[str]]:
    with path.open("r", encoding=encoding, newline="") as fh:
        reader = csv.reader(fh)
        return [[str(cell).strip() for cell in row] for row in reader]


def _read_xlsx_tables(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    if _load_workbook is None:
        raise PreprocessError("openpyxl is required to parse xlsx inputs. Run: pip install openpyxl")
    workbook = _load_workbook(filename=str(path), read_only=True, data_only=True)
    selected_names = [sheet_name] if sheet_name else list(workbook.sheetnames)
    tables: List[Dict[str, Any]] = []
    for name in selected_names:
        if name not in workbook.sheetnames:
            raise PreprocessError(f"Sheet not found: {name}")
        sheet = workbook[name]
        rows: List[List[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if any(values):
                rows.append(values)
        if rows:
            tables.append({"name": name, "rows": rows})
    if not tables:
        raise PreprocessError("No non-empty rows found in workbook.")
    return tables


def _extract_docx_text(path: Path) -> str:
    paragraphs: List[str] = []
    with zipfile.ZipFile(path) as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    for para in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in para.findall(".//w:t", ns)]
        line = "".join(texts).strip()
        if line:
            paragraphs.append(line)
    if not paragraphs:
        raise PreprocessError("No text found in docx document.")
    return "\n".join(paragraphs)


def _extract_with_soffice(path: Path) -> str:
    office_bin = _shutil_which("soffice") or _shutil_which("libreoffice")
    if not office_bin:
        raise PreprocessError("libreoffice/soffice not found. Install: sudo apt-get install libreoffice")
    with tempfile.TemporaryDirectory(prefix="med-doc-preprocess-") as tmp_dir:
        proc = subprocess.run(
            [office_bin, "--headless", "--convert-to", "txt:Text", "--outdir", tmp_dir, str(path)],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        out_path = Path(tmp_dir) / f"{path.stem}.txt"
        if proc.returncode != 0 or not out_path.exists():
            raise PreprocessError(
                f"Failed to convert office document to text: {proc.stderr.strip() or proc.stdout.strip()}"
            )
        return out_path.read_text(encoding="utf-8", errors="replace")


def extract_pdf_pages(path: Path) -> List[str]:
    """提取 PDF 各页文本，返回非空页面列表。适合需要逐页处理的场景。"""
    if _PdfReader is not None:
        try:
            reader = _PdfReader(str(path))
            pages = [(page.extract_text() or "").strip() for page in reader.pages]
            pages = [p for p in pages if p]
            if pages:
                return pages
        except Exception:
            pass

    pdf_to_text = _shutil_which("pdftotext")
    if pdf_to_text:
        proc = subprocess.run(
            [pdf_to_text, "-layout", str(path), "-"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        if proc.returncode == 0 and proc.stdout.strip():
            return [proc.stdout]
    raise PreprocessError(
        "Unable to extract text from pdf. Install pypdf (`pip install pypdf`) or pdftotext (`apt-get install poppler-utils`)."
    )


def extract_pdf_text(path: Path) -> str:
    """提取 PDF 全文，所有页合并为一段字符串。适合把 PDF 当整体处理的场景。"""
    if _PdfReader is not None:
        try:
            reader = _PdfReader(str(path))
            pages = [(page.extract_text() or "").strip() for page in reader.pages]
            text = "\n\n".join(p for p in pages if p)
            if text.strip():
                return text
        except Exception:
            pass

    pdf_to_text = _shutil_which("pdftotext")
    if pdf_to_text:
        proc = subprocess.run(
            [pdf_to_text, "-layout", str(path), "-"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        if proc.returncode == 0 and proc.stdout.strip():
            return proc.stdout
    raise PreprocessError(
        "Unable to extract text from pdf. Install pypdf (`pip install pypdf`) or pdftotext (`apt-get install poppler-utils`)."
    )


def extract_image_text(path: Path) -> str:
    """用 tesseract OCR 识别图片文本。"""
    tesseract_bin = _shutil_which("tesseract")
    if not tesseract_bin:
        raise PreprocessError(
            "tesseract not found. Install: sudo apt-get install tesseract-ocr tesseract-ocr-chi-sim"
        )
    langs = _detect_tesseract_langs(tesseract_bin)
    lang_arg = "chi_sim+eng" if "chi_sim" in langs and "eng" in langs else None
    cmd = [tesseract_bin, str(path), "stdout"]
    if lang_arg:
        cmd.extend(["-l", lang_arg])
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=False)
    if proc.returncode != 0 or not proc.stdout.strip():
        raise PreprocessError(f"Image OCR failed: {proc.stderr.strip() or 'no text returned'}")
    return proc.stdout


# ---------------------------------------------------------------------------
# 主入口
# ---------------------------------------------------------------------------

def load_input_artifact(
    path: Path,
    input_type: str,
    encoding: str = "utf-8",
    sheet_name: str = "",
    *,
    pdf_as_single_text: bool = False,
) -> Dict[str, Any]:
    """
    将任意格式文件加载为标准中间产物 artifact。

    返回值格式之一：
        {"kind": "text",   "text": str}
        {"kind": "pages",  "pages": List[str]}
        {"kind": "json",   "data": Any}
        {"kind": "tables", "tables": List[{"name": str, "rows": List[List[str]]}]}

    参数：
        path              文件路径
        input_type        文件类型字符串（来自 detect_input_type）
        encoding          txt/csv 文件编码，默认 utf-8
        sheet_name        xlsx 指定 sheet，为空则读取全部
        pdf_as_single_text
            False（默认）：PDF 返回 pages 格式，每页独立 —— 适合慢病/大病审核
            True：PDF 返回 text 格式，所有页合并 —— 适合复诊结构化/病历生成
    """
    if input_type in TEXT_FILE_TYPES:
        return {"kind": "text", "text": _read_text_file(path, encoding)}
    if input_type == "json":
        return {"kind": "json", "data": _read_json_file(path)}
    if input_type == "csv":
        return {"kind": "tables", "tables": [{"name": path.stem, "rows": _read_csv_rows(path, encoding)}]}
    if input_type == "xlsx":
        return {"kind": "tables", "tables": _read_xlsx_tables(path, sheet_name)}
    if input_type == "xls":
        text = _extract_with_soffice(path)
        return {"kind": "text", "text": text} if pdf_as_single_text else {"kind": "pages", "pages": [text]}
    if input_type == "docx":
        return {"kind": "text", "text": _extract_docx_text(path)}
    if input_type == "doc":
        return {"kind": "text", "text": _extract_with_soffice(path)}
    if input_type == "pdf":
        if pdf_as_single_text:
            return {"kind": "text", "text": extract_pdf_text(path)}
        return {"kind": "pages", "pages": extract_pdf_pages(path)}
    if input_type in IMAGE_FILE_TYPES:
        text = extract_image_text(path)
        return {"kind": "text", "text": text} if pdf_as_single_text else {"kind": "pages", "pages": [text]}
    raise PreprocessError(f"Unsupported input type: {input_type}")
