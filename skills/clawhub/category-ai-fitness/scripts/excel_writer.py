"""
Excel 输出模块（含主图缩略图嵌入）
"""
import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Microsoft YaHei", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


COLUMNS = [
    ("类目原始输入", 30, "raw_input"),
    ("识别平台", 12, "platform"),
    ("搜索关键词", 25, "keyword"),
    ("样本数", 8, "sample_count"),
    ("📊 场景化适配度", 14, "scene_fitness"),
    ("AI改图难度", 12, "ai_difficulty"),
    ("推荐改图策略", 18, "strategy"),
    ("🎨 IP侵权风险", 12, "infringement_risk"),
    ("图片同质化", 12, "uniqueness_risk"),
    ("必须改图", 10, "must_modify_image"),
    ("场景图占比", 12, "lifestyle_ratio"),
    ("白底图占比", 12, "white_ratio"),
    ("主导商品形态", 18, "dominant_form"),
    ("中位价格 ($)", 12, "median_price"),
    ("中位月销", 12, "median_sales"),
    ("⭐ 最终决策", 16, "decision"),
    ("决策理由", 40, "reason"),
    ("⚠️ IP风险预警", 35, "ip_warning"),
    ("代表图1", 12, "img_1"),
    ("代表图2", 12, "img_2"),
    ("代表图3", 12, "img_3"),
    ("抓取时间", 18, "fetched_at"),
]


def write_excel(rows: list, output_path: str, image_cache_dir: Path):
    """
    rows: list of dict, 每行一个类目结果
    image_cache_dir: 图片缓存目录
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "类目改图适配分析"

    for col_idx, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    for row_idx, row_data in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = 90
        for col_idx, (_, _, key) in enumerate(COLUMNS, start=1):
            value = row_data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = CELL_FONT
            cell.alignment = LEFT if key in ("reason", "raw_input") else CENTER
            cell.border = BORDER

            if key.startswith("img_"):
                img_idx = int(key.split("_")[1]) - 1
                images = row_data.get("representative_images", [])
                if img_idx < len(images):
                    _embed_image(ws, row_idx, col_idx, images[img_idx], image_cache_dir)
                continue

            if isinstance(value, bool):
                cell.value = "✅" if value else ""
            elif isinstance(value, float):
                if key == "scene_fitness" or key == "infringement_score":
                    cell.value = round(value, 1)
                elif key in ("lifestyle_ratio", "white_ratio", "has_human_ratio"):
                    cell.value = f"{value*100:.0f}%"
                else:
                    cell.value = round(value, 2)
            elif value is None:
                cell.value = ""
            else:
                cell.value = value

            if key == "decision":
                cell.font = Font(name="Microsoft YaHei", size=11, bold=True)
                sv = str(value)
                if "AI改图搬" in sv and sv.startswith("✅"):
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif "可改图搬" in sv and sv.startswith("✅"):
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif "白底直搬" in sv and sv.startswith("✅"):
                    cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                elif "⚠" in sv:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif "❌" in sv:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif key in ("infringement_risk", "uniqueness_risk"):
                sv = str(value)
                if "高" in sv:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif "中" in sv:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif "低" in sv:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


    ws.freeze_panes = "A2"

    raw_ws = wb.create_sheet("原始样本明细")
    raw_headers = ["类目", "ASIN", "标题", "品牌", "价格", "评分", "评论数", "月销", "图片URL",
                   "背景类型", "AI难度", "形态", "品牌物", "通用图", "水印", "场景适配分"]
    for col_idx, h in enumerate(raw_headers, start=1):
        c = raw_ws.cell(row=1, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER

    rrow = 2
    for row_data in rows:
        cat = row_data.get("raw_input", "")
        samples = row_data.get("raw_samples", [])
        for s in samples:
            p = s.get("product", {})
            v = s.get("vision", {})
            row_values = [
                cat,
                p.get("asin", ""),
                p.get("title", "")[:120],
                p.get("brand", ""),
                p.get("price", ""),
                p.get("rating", ""),
                p.get("reviews", ""),
                p.get("monthly_sales", ""),
                p.get("image", ""),
                v.get("background_type", ""),
                v.get("ai_difficulty", ""),
                v.get("product_form", ""),
                "✅" if v.get("has_brand_elements") else "",
                "✅" if v.get("is_generic_supplier_image") else "",
                "✅" if v.get("has_text_watermark") else "",
                v.get("scene_fit_score", ""),
            ]
            for c_idx, val in enumerate(row_values, start=1):
                raw_ws.cell(row=rrow, column=c_idx, value=val)
            rrow += 1

    raw_ws.freeze_panes = "A2"
    raw_ws.column_dimensions["A"].width = 30
    raw_ws.column_dimensions["B"].width = 14
    raw_ws.column_dimensions["C"].width = 60
    raw_ws.column_dimensions["I"].width = 50
    for col_letter in "DEFGHJKLMNOP":
        raw_ws.column_dimensions[col_letter].width = 14

    wb.save(output_path)
    print(f"[Excel] 已生成报告: {output_path}", flush=True)


def _embed_image(ws, row, col, image_url: str, cache_dir: Path):
    if not image_url:
        return
    import hashlib
    h = hashlib.md5(image_url.encode()).hexdigest()
    img_path = cache_dir / f"{h}.jpg"
    if not img_path.exists():
        return
    try:
        pil_img = PILImage.open(img_path)
        pil_img.thumbnail((100, 100))
        if pil_img.mode != "RGB":
            pil_img = pil_img.convert("RGB")
        buf = io.BytesIO()
        pil_img.save(buf, format="PNG")
        buf.seek(0)
        xl_img = XLImage(buf)
        cell_ref = f"{get_column_letter(col)}{row}"
        ws.add_image(xl_img, cell_ref)
    except Exception as e:
        print(f"[Excel] 嵌入图片失败 {image_url}: {e}", flush=True)
