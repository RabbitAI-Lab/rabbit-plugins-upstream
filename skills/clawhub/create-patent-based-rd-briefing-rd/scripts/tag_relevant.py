#!/usr/bin/env python3
"""Add auditable discovery signals to an authorized patent workbook.

Keyword matches are candidates for human review. They are never automatic
novelty, legal-status, infringement, FTO, technical-quality, or market findings.
The workbook is edited through openpyxl so existing links and embedded images
are retained as far as the library and workbook format permit.
"""

from __future__ import annotations

import argparse
import importlib.util
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from types import ModuleType
from typing import Iterable

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill


SAFE_TOPIC = re.compile(r"^[A-Za-z0-9_-]+$")
OUTPUT_HEADERS = (
    "Discovery disposition",
    "Inclusive terms matched",
    "Exclusion terms matched",
    "Review status",
    "Reviewer",
    "Review date",
)


class TaggingError(ValueError):
    """Raised when the workbook or configuration is not release-safe."""


@dataclass(frozen=True)
class MatchResult:
    disposition: str
    inclusive_terms: tuple[str, ...]
    exclusion_terms: tuple[str, ...]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add reviewable patent-discovery signals to an Excel workbook"
    )
    parser.add_argument("input_xlsx", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    parser.add_argument("topic_key")
    parser.add_argument("--reviewer", default="")
    parser.add_argument("--review-date", default="")
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Allow replacement of an existing output workbook",
    )
    return parser.parse_args()


def skill_root() -> Path:
    return Path(__file__).resolve().parent.parent


def validate_topic_key(topic_key: str) -> str:
    if not SAFE_TOPIC.fullmatch(topic_key):
        raise TaggingError(
            "Topic key must contain only ASCII letters, numbers, hyphens, and underscores"
        )
    return topic_key


def resolve_config(topic_key: str) -> Path:
    config_dir = (skill_root() / "config").resolve()
    validated = validate_topic_key(topic_key)
    legacy_topology = {
        "coffee-machine": "\u5496\u5561\u673a_keywords.py",
    }
    filename = legacy_topology.get(validated, f"{validated}_keywords.py")
    path = (config_dir / filename).resolve()
    if path.parent != config_dir:
        raise TaggingError("Configuration path escaped the package config directory")
    if not path.is_file():
        available = sorted(p.stem.removesuffix("_keywords") for p in config_dir.glob("*_keywords.py"))
        raise TaggingError(
            f"Keyword configuration not found for {topic_key!r}. Available keys: {available}"
        )
    return path


def load_module(path: Path, module_name: str) -> ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise TaggingError(f"Cannot load configuration: {path.name}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def require_string_list(module: ModuleType, name: str) -> list[str]:
    value = getattr(module, name, None)
    if not isinstance(value, list) or not value:
        raise TaggingError(f"{module.__name__}.{name} must be a non-empty list")
    if any(not isinstance(item, str) or not item.strip() for item in value):
        raise TaggingError(f"{module.__name__}.{name} must contain non-empty strings")
    return [item.strip() for item in value]


def normalize(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return " ".join(str(value).casefold().split())


def find_column(columns: Iterable[object], aliases: list[str]) -> object | None:
    normalized = {normalize(column): column for column in columns}
    for alias in aliases:
        found = normalized.get(normalize(alias))
        if found is not None:
            return found
    return None


def canonical_columns(df: pd.DataFrame, config: ModuleType) -> dict[str, object]:
    aliases = getattr(config, "COLUMN_ALIASES", None)
    if not isinstance(aliases, dict):
        raise TaggingError("COLUMN_ALIASES must be a dictionary")
    mapped: dict[str, object] = {}
    for canonical, candidates in aliases.items():
        if not isinstance(candidates, list):
            raise TaggingError(f"Aliases for {canonical!r} must be a list")
        column = find_column(df.columns, [canonical, *candidates])
        if column is not None:
            mapped[canonical] = column
    for required in ("publication_number", "title"):
        if required not in mapped:
            raise TaggingError(
                f"Required field {required!r} not found. Configure an explicit workbook alias."
            )
    return mapped


def record_text(row: pd.Series, mapped: dict[str, object], fields: list[str]) -> str:
    values = []
    for field in fields:
        source_column = mapped.get(field)
        if source_column is not None:
            value = normalize(row[source_column])
            if value:
                values.append(value)
    return "\n".join(values)


def match_terms(text: str, terms: list[str]) -> tuple[str, ...]:
    # Preserve configuration order and remove case-insensitive duplicates.
    found: list[str] = []
    seen: set[str] = set()
    for term in terms:
        key = normalize(term)
        if key and key in text and key not in seen:
            found.append(term)
            seen.add(key)
    return tuple(found)


def assess_record(
    row: pd.Series,
    mapped: dict[str, object],
    fields: list[str],
    include_terms: list[str],
    exclude_terms: list[str],
) -> MatchResult:
    text = record_text(row, mapped, fields)
    inclusive = match_terms(text, include_terms)
    exclusions = match_terms(text, exclude_terms)
    if inclusive and exclusions:
        disposition = "Candidate — inclusive and exclusion signals"
    elif inclusive:
        disposition = "Candidate — inclusive signal"
    elif exclusions:
        disposition = "Likely out of scope — exclusion signal"
    else:
        disposition = "No configured signal"
    return MatchResult(disposition, inclusive, exclusions)


def validate_paths(input_path: Path, output_path: Path, overwrite: bool) -> tuple[Path, Path]:
    input_path = input_path.expanduser().resolve()
    output_path = output_path.expanduser().resolve()
    if not input_path.is_file():
        raise TaggingError(f"Input workbook does not exist: {input_path}")
    if input_path.suffix.casefold() != ".xlsx":
        raise TaggingError("Input must be an .xlsx workbook")
    if output_path.suffix.casefold() != ".xlsx":
        raise TaggingError("Output must use the .xlsx extension")
    if input_path == output_path:
        raise TaggingError("Input and output paths must differ")
    if output_path.exists() and not overwrite:
        raise TaggingError("Output exists; pass --overwrite to replace it")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    return input_path, output_path


def validate_review_date(value: str) -> str:
    if not value:
        return ""
    try:
        return date.fromisoformat(value).isoformat()
    except ValueError as error:
        raise TaggingError("--review-date must be a valid ISO date (YYYY-MM-DD)") from error


def write_signals(
    input_path: Path,
    output_path: Path,
    df: pd.DataFrame,
    results: list[MatchResult],
    reviewer: str,
    review_date: str,
) -> None:
    # Work on a copy so failure cannot partially modify the source workbook.
    shutil.copy2(input_path, output_path)
    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active
    if worksheet.max_row - 1 < len(df):
        output_path.unlink(missing_ok=True)
        raise TaggingError("Worksheet row count is smaller than the parsed data table")

    existing_headers = {
        normalize(worksheet.cell(1, column).value): column
        for column in range(1, worksheet.max_column + 1)
    }
    header_columns: dict[str, int] = {}
    next_column = worksheet.max_column + 1
    for header in OUTPUT_HEADERS:
        column = existing_headers.get(normalize(header))
        if column is None:
            column = next_column
            next_column += 1
        worksheet.cell(1, column).value = header
        header_columns[header] = column

    candidate_fill = PatternFill(fill_type="solid", fgColor="EAF2F9")
    candidate_font = Font(color="163A63", bold=True)
    for row_number, result in enumerate(results, start=2):
        values = {
            "Discovery disposition": result.disposition,
            "Inclusive terms matched": " | ".join(result.inclusive_terms),
            "Exclusion terms matched": " | ".join(result.exclusion_terms),
            "Review status": "Requires human review" if "Candidate" in result.disposition else "Unreviewed",
            "Reviewer": reviewer,
            "Review date": review_date,
        }
        for header, value in values.items():
            worksheet.cell(row_number, header_columns[header]).value = value
        if "Candidate" in result.disposition:
            cell = worksheet.cell(row_number, header_columns["Discovery disposition"])
            cell.fill = candidate_fill
            cell.font = candidate_font

    try:
        workbook.save(output_path)
    except Exception:
        output_path.unlink(missing_ok=True)
        raise


def main() -> int:
    args = parse_args()
    try:
        input_path, output_path = validate_paths(args.input_xlsx, args.output_xlsx, args.overwrite)
        review_date = validate_review_date(args.review_date)
        config_path = resolve_config(args.topic_key)
        config = load_module(config_path, f"localized_keywords_{args.topic_key}")
        include_terms = require_string_list(config, "INCLUDE_KEYWORDS")
        exclude_terms = require_string_list(config, "EXCLUDE_KEYWORDS")
        search_fields = require_string_list(config, "SEARCH_FIELDS")
        df = pd.read_excel(input_path, engine="openpyxl")
        if df.empty:
            raise TaggingError("Input workbook contains no patent records")
        mapped = canonical_columns(df, config)
        results = [
            assess_record(row, mapped, search_fields, include_terms, exclude_terms)
            for _, row in df.iterrows()
        ]
        write_signals(input_path, output_path, df, results, args.reviewer.strip(), review_date)
        candidates = sum("Candidate" in result.disposition for result in results)
        conflicts = sum(bool(result.inclusive_terms and result.exclusion_terms) for result in results)
        print(f"Records processed: {len(results)}")
        print(f"Candidates requiring review: {candidates}")
        print(f"Candidates with conflicting signals: {conflicts}")
        print("Keyword matches are discovery signals, not final relevance or legal conclusions.")
        print(f"Output workbook: {output_path}")
        return 0
    except (OSError, ValueError, ImportError) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
