#!/usr/bin/env python3
"""Render a reviewed patent-based R&D briefing as self-contained HTML.

The renderer consumes a tagged Excel workbook and the source-provided topic
configurations. It publishes only reviewer-confirmed records, derives every
count from the included dataset, escapes workbook/config text, allowlists URLs,
and performs no network access. The output is technical intelligence, not a
legal opinion or a representation of exhaustive patent coverage.
"""

from __future__ import annotations

import argparse
import base64
import html
import importlib.util
import io
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from types import ModuleType
from typing import Iterable
from urllib.parse import urlparse

import openpyxl
import pandas as pd


SAFE_TOPIC = re.compile(r"^[A-Za-z0-9_-]+$")
INCLUDED_DISPOSITIONS = {
    "included",
    "included — reviewer confirmed",
    "include",
}
REVIEWED_STATUSES = {
    "reviewed",
    "approved for briefing",
    "complete",
}
MAX_EMBEDDED_IMAGE_BYTES = 5_000_000
MAX_IMAGES = 200


class ReportError(ValueError):
    """Raised when input cannot support a reviewed briefing."""


@dataclass(frozen=True)
class PatentRecord:
    publication_number: str
    title: str
    normalized_title: str
    applicant: str
    legal_status: str
    application_date: str
    publication_date: str
    technical_problem: str
    technical_solution: str
    technical_effect: str
    family_id: str
    source_url: str
    reviewer: str
    review_date: str
    inclusive_terms: tuple[str, ...]
    category_ids: tuple[str, ...]


@dataclass(frozen=True)
class ReportContext:
    topic_key: str
    topic_label: str
    start_date: str
    end_date: str
    report_date: str
    evidence_cutoff: str
    records: tuple[PatentRecord, ...]
    images: dict[str, str]
    search_scope: dict[str, object]
    categories: tuple[dict[str, object], ...]
    organization_leads: tuple[dict[str, object], ...]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a reviewed patent-based R&D briefing")
    parser.add_argument("tagged_xlsx", type=Path)
    parser.add_argument("output_html", type=Path)
    parser.add_argument("topic_key")
    parser.add_argument("start_date")
    parser.add_argument("end_date")
    parser.add_argument("--report-date", default=date.today().isoformat())
    parser.add_argument("--evidence-cutoff", default="")
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def skill_root() -> Path:
    return Path(__file__).resolve().parent.parent


def validate_iso_date(label: str, value: str) -> str:
    try:
        return date.fromisoformat(value).isoformat()
    except ValueError as error:
        raise ReportError(f"{label} must be a valid ISO date (YYYY-MM-DD): {value!r}") from error


def validate_topic_key(value: str) -> str:
    if not SAFE_TOPIC.fullmatch(value):
        raise ReportError("Topic key may contain only ASCII letters, numbers, hyphens, and underscores")
    return value


def config_filename(topic_key: str, suffix: str) -> str:
    legacy_stem = "\u5496\u5561\u673a"
    stem = legacy_stem if topic_key == "coffee-machine" else topic_key
    return f"{stem}_{suffix}.py"


def resolve_config(topic_key: str, suffix: str) -> Path:
    config_dir = (skill_root() / "config").resolve()
    path = (config_dir / config_filename(validate_topic_key(topic_key), suffix)).resolve()
    if path.parent != config_dir:
        raise ReportError("Configuration path escaped the package")
    if not path.is_file():
        raise ReportError(f"Missing {suffix} configuration for topic {topic_key!r}")
    return path


def load_module(path: Path, name: str) -> ModuleType:
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise ReportError(f"Cannot load configuration: {path.name}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def normalize(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return " ".join(str(value).strip().split())


def casefold(value: object) -> str:
    return normalize(value).casefold()


def find_column(columns: Iterable[object], aliases: Iterable[str]) -> object | None:
    normalized = {casefold(column): column for column in columns}
    for alias in aliases:
        column = normalized.get(casefold(alias))
        if column is not None:
            return column
    return None


def canonical_columns(df: pd.DataFrame, keywords: ModuleType) -> dict[str, object]:
    aliases = getattr(keywords, "COLUMN_ALIASES", None)
    if not isinstance(aliases, dict):
        raise ReportError("Keyword configuration must define COLUMN_ALIASES")
    mapped: dict[str, object] = {}
    for canonical, candidates in aliases.items():
        if not isinstance(candidates, list):
            raise ReportError(f"Aliases for {canonical!r} must be a list")
        column = find_column(df.columns, [canonical, *candidates])
        if column is not None:
            mapped[canonical] = column
    for required in ("publication_number", "title", "applicant"):
        if required not in mapped:
            raise ReportError(f"Required workbook field not found: {required}")
    workflow_aliases = {
        "disposition": ["Discovery disposition"],
        "inclusive_terms": ["Inclusive terms matched"],
        "review_status": ["Review status"],
        "reviewer": ["Reviewer"],
        "review_date": ["Review date"],
        "category_ids": ["Reviewed category IDs", "Category IDs"],
    }
    for canonical, candidates in workflow_aliases.items():
        column = find_column(df.columns, candidates)
        if column is not None:
            mapped[canonical] = column
    for required in ("disposition", "review_status", "reviewer", "review_date"):
        if required not in mapped:
            raise ReportError(f"Tagged workbook lacks review field: {required}")
    return mapped


def safe_url(value: object) -> str:
    text = normalize(value)
    if not text:
        return ""
    parsed = urlparse(text)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return ""
    return text


def split_tokens(value: object) -> tuple[str, ...]:
    text = normalize(value)
    if not text:
        return ()
    items = [item.strip() for item in re.split(r"[|,;]", text) if item.strip()]
    return tuple(dict.fromkeys(items))


def cell(row: pd.Series, mapped: dict[str, object], key: str) -> str:
    column = mapped.get(key)
    return normalize(row[column]) if column is not None else ""


def is_included(row: pd.Series, mapped: dict[str, object]) -> bool:
    disposition = casefold(row[mapped["disposition"]])
    review_status = casefold(row[mapped["review_status"]])
    reviewer = normalize(row[mapped["reviewer"]])
    review_date = normalize(row[mapped["review_date"]])
    if disposition not in INCLUDED_DISPOSITIONS:
        return False
    if review_status not in REVIEWED_STATUSES:
        return False
    if not reviewer:
        raise ReportError("An included record has no named reviewer")
    validate_iso_date("Review date", review_date)
    return True


def configured_categories(content: ModuleType) -> tuple[dict[str, object], ...]:
    categories = getattr(content, "TECHNOLOGY_CATEGORIES", None)
    if not isinstance(categories, list) or not categories:
        raise ReportError("Content configuration must define TECHNOLOGY_CATEGORIES")
    seen: set[str] = set()
    validated = []
    for item in categories:
        if not isinstance(item, dict):
            raise ReportError("Every technology category must be a dictionary")
        category_id = normalize(item.get("category_id"))
        label = normalize(item.get("label"))
        if not category_id or not label or category_id in seen:
            raise ReportError("Category IDs and labels must be non-empty and IDs unique")
        seen.add(category_id)
        validated.append(item)
    return tuple(validated)


def keyword_categories(keywords: ModuleType) -> dict[str, list[str]]:
    value = getattr(keywords, "TECHNOLOGY_CATEGORIES", None)
    if not isinstance(value, dict):
        raise ReportError("Keyword configuration must define TECHNOLOGY_CATEGORIES")
    output: dict[str, list[str]] = {}
    for category_id, terms in value.items():
        if not isinstance(category_id, str) or not isinstance(terms, list):
            raise ReportError("Technology-category keywords are malformed")
        output[category_id] = [normalize(term) for term in terms if normalize(term)]
    return output


def infer_candidate_categories(row: pd.Series, mapped: dict[str, object], keywords: dict[str, list[str]]) -> tuple[str, ...]:
    explicit = split_tokens(cell(row, mapped, "category_ids"))
    if explicit:
        return explicit
    searchable = "\n".join(
        cell(row, mapped, key).casefold()
        for key in (
            "title",
            "normalized_title",
            "technical_problem",
            "technical_solution",
            "technical_effect",
            "abstract",
            "independent_claims",
        )
    )
    inferred = []
    for category_id, terms in keywords.items():
        if any(term.casefold() in searchable for term in terms):
            inferred.append(category_id)
    return tuple(inferred)


def records_from_frame(
    df: pd.DataFrame,
    mapped: dict[str, object],
    category_keywords: dict[str, list[str]],
    valid_category_ids: set[str],
) -> tuple[PatentRecord, ...]:
    records = []
    seen_publications: set[str] = set()
    for _, row in df.iterrows():
        if not is_included(row, mapped):
            continue
        publication = cell(row, mapped, "publication_number")
        if not publication:
            raise ReportError("An included record has no publication number")
        if publication.casefold() in seen_publications:
            raise ReportError(f"Duplicate included publication number: {publication}")
        seen_publications.add(publication.casefold())
        categories = infer_candidate_categories(row, mapped, category_keywords)
        unknown = sorted(set(categories) - valid_category_ids)
        if unknown:
            raise ReportError(f"Record {publication} uses unknown category IDs: {unknown}")
        records.append(
            PatentRecord(
                publication_number=publication,
                title=cell(row, mapped, "title"),
                normalized_title=cell(row, mapped, "normalized_title"),
                applicant=cell(row, mapped, "applicant"),
                legal_status=cell(row, mapped, "legal_status"),
                application_date=cell(row, mapped, "application_date"),
                publication_date=cell(row, mapped, "publication_date"),
                technical_problem=cell(row, mapped, "technical_problem"),
                technical_solution=cell(row, mapped, "technical_solution"),
                technical_effect=cell(row, mapped, "technical_effect"),
                family_id=cell(row, mapped, "family_id"),
                source_url=safe_url(cell(row, mapped, "source_url")),
                reviewer=cell(row, mapped, "reviewer"),
                review_date=validate_iso_date("Review date", cell(row, mapped, "review_date")),
                inclusive_terms=split_tokens(cell(row, mapped, "inclusive_terms")),
                category_ids=categories,
            )
        )
    if not records:
        raise ReportError("No reviewer-confirmed included records are available")
    return tuple(records)


def publication_column_index(worksheet: openpyxl.worksheet.worksheet.Worksheet, aliases: list[str]) -> int | None:
    headers = {
        casefold(worksheet.cell(1, column).value): column
        for column in range(1, worksheet.max_column + 1)
    }
    for alias in aliases:
        column = headers.get(casefold(alias))
        if column is not None:
            return column
    return None


def embedded_images(path: Path, aliases: list[str], publications: set[str]) -> dict[str, str]:
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    pn_column = publication_column_index(worksheet, aliases)
    if pn_column is None:
        return {}
    row_to_publication = {
        row: normalize(worksheet.cell(row, pn_column).value)
        for row in range(2, worksheet.max_row + 1)
    }
    output: dict[str, str] = {}
    for image in worksheet._images[:MAX_IMAGES]:
        anchor = getattr(image, "anchor", None)
        origin = getattr(anchor, "_from", None)
        if origin is None:
            continue
        row_number = origin.row + 1
        publication = row_to_publication.get(row_number, "")
        if publication not in publications:
            continue
        try:
            data = image._data()
        except Exception:
            continue
        if len(data) > MAX_EMBEDDED_IMAGE_BYTES:
            continue
        image_format = normalize(getattr(image, "format", "png")).casefold()
        mime = "image/jpeg" if image_format in {"jpg", "jpeg"} else "image/png"
        output[publication] = f"data:{mime};base64,{base64.b64encode(data).decode('ascii')}"
    return output


def e(value: object) -> str:
    return html.escape(normalize(value), quote=True)


def link(record: PatentRecord) -> str:
    label = e(record.publication_number)
    if not record.source_url:
        return label
    return f'<a href="{e(record.source_url)}" rel="noopener noreferrer">{label}</a>'


def image_markup(record: PatentRecord, images: dict[str, str]) -> str:
    source = images.get(record.publication_number)
    if not source:
        return '<div class="image-empty">No reviewed embedded figure</div>'
    return (
        f'<img class="patent-image" src="{source}" '
        f'alt="Embedded figure for {e(record.publication_number)}">'
    )


def record_card(record: PatentRecord, images: dict[str, str]) -> str:
    title = record.normalized_title or record.title
    details = []
    if record.technical_problem:
        details.append(f"<dt>Problem</dt><dd>{e(record.technical_problem)}</dd>")
    if record.technical_solution:
        details.append(f"<dt>Approach</dt><dd>{e(record.technical_solution)}</dd>")
    if record.technical_effect:
        details.append(f"<dt>Reported effect</dt><dd>{e(record.technical_effect)}</dd>")
    return "".join(
        [
            '<article class="patent-card">',
            image_markup(record, images),
            '<div class="patent-body">',
            f'<p class="publication">{link(record)}</p>',
            f'<h4>{e(title)}</h4>',
            f'<p class="record-meta">{e(record.applicant)} · Status: {e(record.legal_status) or "Not recorded"}</p>',
            f'<dl>{"".join(details)}</dl>' if details else "",
            f'<p class="review-note">Reviewed by {e(record.reviewer)} on {e(record.review_date)}</p>',
            '</div></article>',
        ]
    )


def css() -> str:
    return """
:root{color-scheme:light;--ink:#172033;--muted:#5a6678;--navy:#173b63;--blue:#286aa6;--soft:#edf4fa;--line:#ccd6e1;--paper:#fff;--canvas:#f3f6f9;--warn:#8a4b08;--warn-bg:#fff7e8;--radius:6px;font-family:Inter,ui-sans-serif,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif}
*{box-sizing:border-box}body{margin:0;background:var(--canvas);color:var(--ink);font-size:15px;line-height:1.58}a{color:var(--blue);text-underline-offset:2px}.page{width:min(1180px,calc(100% - 40px));margin:24px auto 48px;background:var(--paper);border:1px solid var(--line);box-shadow:0 2px 10px rgba(23,59,99,.08)}
header{padding:36px 44px 28px;border-top:6px solid var(--navy);border-bottom:1px solid var(--line)}.eyebrow{margin:0 0 8px;color:var(--blue);font-size:12px;font-weight:750;letter-spacing:.09em;text-transform:uppercase}h1,h2,h3,h4{color:var(--navy);line-height:1.22}h1{margin:0 0 10px;font-size:clamp(28px,4vw,44px)}h2{margin:0 0 14px;font-size:25px}h3{margin:26px 0 10px;font-size:18px}h4{margin:0 0 7px;font-size:15px}.subtitle{max-width:850px;margin:0;color:var(--muted);font-size:17px}
.metadata{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin-top:22px}.meta{padding:11px 13px;border:1px solid #caddec;background:var(--soft);border-radius:var(--radius)}.meta span{display:block;color:var(--muted);font-size:11px;font-weight:700;text-transform:uppercase}.meta strong{display:block;margin-top:3px}
nav{padding:17px 44px;background:#f8fafc;border-bottom:1px solid var(--line)}nav ul{display:flex;gap:18px;flex-wrap:wrap;margin:0;padding:0;list-style:none}nav a{font-size:13px;font-weight:650}main{padding:0 44px}section{padding:36px 0;border-bottom:1px solid var(--line);scroll-margin-top:16px}.lead{max-width:900px;color:#334155}.callout{margin:17px 0;padding:15px 17px;border-left:4px solid var(--blue);background:var(--soft)}.callout.warning{border-left-color:var(--warn);background:var(--warn-bg)}
.metrics{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px;margin:18px 0}.metric{padding:16px;border:1px solid var(--line);border-radius:var(--radius)}.metric strong{display:block;color:var(--navy);font-size:26px;font-variant-numeric:tabular-nums}.metric span{color:var(--muted);font-size:12px}.table-wrap{overflow-x:auto;margin:16px 0 22px}table{width:100%;border-collapse:collapse;border:1px solid #a9b7c7;font-size:13px}caption{padding:0 0 7px;color:var(--muted);text-align:left;font-size:12px}th,td{padding:9px 10px;border:1px solid var(--line);text-align:left;vertical-align:top}th{color:var(--navy);background:#edf3f8}tbody tr:nth-child(even){background:#fbfcfe}
.patent-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:16px}.patent-card{display:grid;grid-template-columns:150px 1fr;min-height:170px;border:1px solid var(--line);border-radius:var(--radius);overflow:hidden}.patent-image,.image-empty{width:150px;height:100%;min-height:170px;object-fit:contain;background:#f5f7fa}.image-empty{display:flex;align-items:center;justify-content:center;padding:15px;color:var(--muted);font-size:11px;text-align:center}.patent-body{padding:15px}.publication{margin:0 0 5px;font-size:12px;font-weight:750}.record-meta,.review-note{color:var(--muted);font-size:12px}.patent-body dl{display:grid;grid-template-columns:70px 1fr;gap:4px 8px;margin:11px 0}.patent-body dt{font-size:12px;font-weight:700}.patent-body dd{margin:0;font-size:12px}.category{margin:24px 0}.category-header{padding:14px 16px;background:var(--navy);color:#fff;border-radius:var(--radius) var(--radius) 0 0}.category-header h3{margin:0;color:#fff}.category-body{padding:16px;border:1px solid var(--line);border-top:0}.empty{padding:18px;border:1px dashed #aab8c6;color:var(--muted)}footer{padding:26px 44px;color:var(--muted);background:#f8fafc}
@media(max-width:850px){.page{width:100%;margin:0;border:0}.metadata,.metrics,.patent-grid{grid-template-columns:1fr}.patent-card{grid-template-columns:120px 1fr}.patent-image,.image-empty{width:120px}header,nav,main,footer{padding-left:22px;padding-right:22px}}
@media print{@page{margin:14mm}body{background:#fff;font-size:10pt}.page{width:100%;margin:0;border:0;box-shadow:none}nav{display:none}section,.patent-card,table{break-inside:avoid}a{color:inherit;text-decoration:none}}
"""


def build_html(context: ReportContext) -> str:
    records = context.records
    organization_counts = Counter(record.applicant for record in records if record.applicant)
    family_values = {record.family_id for record in records if record.family_id}
    category_records: dict[str, list[PatentRecord]] = defaultdict(list)
    for record in records:
        for category_id in record.category_ids:
            category_records[category_id].append(record)
    title = f"Patent-Based R&D Briefing — {context.topic_label}"
    parts = [
        "<!doctype html>",
        '<html lang="en"><head><meta charset="utf-8">',
        '<meta name="viewport" content="width=device-width,initial-scale=1">',
        '<meta name="report-version" content="V1.1.0-localized">',
        f'<meta name="report-date" content="{e(context.report_date)}">',
        f'<meta name="evidence-cutoff" content="{e(context.evidence_cutoff)}">',
        '<meta name="review-status" content="Reviewer-confirmed records; specialist review pending">',
        f"<title>{e(title)} — V1.1.0-localized</title><style>{css()}</style></head><body>",
        '<article class="page"><header>',
        '<p class="eyebrow">Patent-based technical intelligence</p>',
        f"<h1>{e(title)}</h1>",
        f'<p class="subtitle">Reviewed evidence window: {e(context.start_date)} to {e(context.end_date)}. Counts describe the included workbook records, not the global patent universe.</p>',
        '<div class="metadata">',
        '<div class="meta"><span>Version</span><strong>V1.1.0-localized</strong></div>',
        f'<div class="meta"><span>Report date</span><strong>{e(context.report_date)}</strong></div>',
        f'<div class="meta"><span>Evidence cutoff</span><strong>{e(context.evidence_cutoff)}</strong></div>',
        '<div class="meta"><span>Review status</span><strong>Technical review complete; specialist review pending</strong></div>',
        '</div></header>',
        '<nav aria-label="Report sections"><ul>',
        '<li><a href="#scope">Scope and method</a></li>',
        '<li><a href="#portfolio">Reviewed dataset</a></li>',
        '<li><a href="#organizations">Organizations</a></li>',
        '<li><a href="#categories">Technology categories</a></li>',
        '<li><a href="#register">Evidence register</a></li>',
        '</ul></nav><main>',
        '<section id="scope"><h2>Scope, method, and boundaries</h2>',
        '<p class="lead">The workflow used configured discovery terms, retained match provenance, and published only records explicitly included by a named reviewer. Keyword signals do not establish novelty, technical merit, legal status, infringement, or freedom to operate.</p>',
        '<div class="callout warning"><strong>Legal boundary:</strong> This report is not legal advice. Legal status, claim scope, validity, enforceability, infringement, and freedom-to-operate questions require a qualified patent professional and current jurisdiction-specific review.</div>',
        '<div class="table-wrap"><table><caption>Declared reviewed universe</caption><thead><tr><th>Dimension</th><th>Definition</th></tr></thead><tbody>',
        f'<tr><td>Topic</td><td>{e(context.topic_label)}</td></tr>',
        f'<tr><td>Evidence window</td><td>{e(context.start_date)} to {e(context.end_date)}</td></tr>',
        f'<tr><td>Geography</td><td>{e(context.search_scope.get("geography", "Not specified"))}</td></tr>',
        f'<tr><td>Count unit</td><td>{e(context.search_scope.get("count_unit", "Publication records"))}</td></tr>',
        f'<tr><td>Exclusions</td><td>{e("; ".join(context.search_scope.get("excluded", [])))}</td></tr>',
        '</tbody></table></div></section>',
        '<section id="portfolio"><h2>Reviewed dataset</h2>',
        '<div class="metrics">',
        f'<div class="metric"><strong>{len(records)}</strong><span>included publication records</span></div>',
        f'<div class="metric"><strong>{len(family_values) if family_values else "—"}</strong><span>distinct recorded family IDs</span></div>',
        f'<div class="metric"><strong>{len(organization_counts)}</strong><span>normalized applicant strings</span></div>',
        f'<div class="metric"><strong>{sum(bool(r.category_ids) for r in records)}</strong><span>records with category evidence</span></div>',
        '</div>',
        '<p class="lead">Family totals appear only when the workbook supplies family IDs. Missing IDs are not inferred. Organization counts use workbook applicant strings and may require corporate-family normalization.</p>',
        '</section>',
        '<section id="organizations"><h2>Organization activity in the reviewed dataset</h2>',
        '<div class="table-wrap"><table><caption>Activity volume is not a leadership, quality, or FTO ranking.</caption><thead><tr><th>Applicant string</th><th>Included publications</th><th>Interpretation boundary</th></tr></thead><tbody>',
    ]
    for organization, count in sorted(organization_counts.items(), key=lambda item: (-item[1], item[0].casefold())):
        parts.append(
            f'<tr><td>{e(organization)}</td><td>{count}</td><td>Observed activity in the reviewed workbook; entity and family normalization may be incomplete.</td></tr>'
        )
    parts.extend(['</tbody></table></div></section>', '<section id="categories"><h2>Technology categories and reviewed records</h2>'])
    for category in context.categories:
        category_id = normalize(category.get("category_id"))
        label = normalize(category.get("label"))
        questions = category.get("decision_questions", [])
        selected = category_records.get(category_id, [])
        parts.extend([
            f'<div class="category" id="category-{e(category_id)}">',
            f'<div class="category-header"><h3>{e(label)} · {len(selected)} reviewed records</h3></div>',
            '<div class="category-body">',
            f'<p><strong>Decision questions:</strong> {e("; ".join(questions if isinstance(questions, list) else []))}</p>',
        ])
        if selected:
            parts.append('<div class="patent-grid">')
            parts.extend(record_card(record, context.images) for record in selected)
            parts.append('</div>')
        else:
            parts.append('<div class="empty">No included record was assigned to this category in the reviewed workbook. This is not evidence that no relevant patents exist.</div>')
        parts.append('</div></div>')
    unclassified = [record for record in records if not record.category_ids]
    if unclassified:
        parts.extend([
            '<div class="category"><div class="category-header"><h3>Unclassified reviewed records</h3></div><div class="category-body">',
            '<p>These records passed inclusion review but require explicit category review.</p><div class="patent-grid">',
        ])
        parts.extend(record_card(record, context.images) for record in unclassified)
        parts.append('</div></div></div>')
    parts.extend([
        '</section>',
        '<section id="register"><h2>Evidence register</h2>',
        '<div class="table-wrap"><table><caption>One row per reviewer-confirmed publication record.</caption><thead><tr><th>Publication</th><th>Applicant</th><th>Dates</th><th>Family ID</th><th>Categories</th><th>Reviewer</th></tr></thead><tbody>',
    ])
    for record in records:
        dates = "; ".join(
            value for value in (
                f"Application: {record.application_date}" if record.application_date else "",
                f"Publication: {record.publication_date}" if record.publication_date else "",
            ) if value
        )
        parts.append(
            "".join([
                '<tr>',
                f'<td>{link(record)}<br>{e(record.title)}</td>',
                f'<td>{e(record.applicant)}</td>',
                f'<td>{e(dates) or "Not recorded"}</td>',
                f'<td>{e(record.family_id) or "Not recorded"}</td>',
                f'<td>{e(", ".join(record.category_ids)) or "Unclassified"}</td>',
                f'<td>{e(record.reviewer)}<br>{e(record.review_date)}</td>',
                '</tr>',
            ])
        )
    parts.extend([
        '</tbody></table></div>',
        '<div class="callout"><strong>Next review:</strong> Confirm claim relevance for material findings, refresh legal status as of the decision date, normalize families and corporate entities, document search coverage, and route legal questions to a patent professional.</div>',
        '</section></main><footer>',
        '<p><strong>Patent-Based R&D Briefing — V1.1.0-localized</strong></p>',
        f'<p>Report date: {e(context.report_date)} · Evidence cutoff: {e(context.evidence_cutoff)} · Review status: technical review complete; specialist review pending.</p>',
        '<p>This report is not legal advice and is limited to the documented reviewed workbook. It does not confirm exhaustive coverage, validity, enforceability, infringement, non-infringement, or freedom to operate. Consult a qualified patent professional.</p>',
        '</footer></article></body></html>',
    ])
    return "".join(parts)


def validate_output(html_text: str, context: ReportContext) -> None:
    required = (
        '<html lang="en">',
        'id="scope"',
        'id="portfolio"',
        'id="organizations"',
        'id="categories"',
        'id="register"',
        "not legal advice",
        "patent professional",
        context.report_date,
        context.evidence_cutoff,
    )
    missing = [item for item in required if item not in html_text]
    if missing:
        raise ReportError(f"Generated report is missing required content: {missing}")
    forbidden = (
        "<" + "script",
        "linear" + "-gradient",
        "radial" + "-gradient",
        "inner" + "HTML",
        "open." + "zhihuiya.com",
        "analytics." + "zhihuiya.com",
    )
    found = [item for item in forbidden if item.casefold() in html_text.casefold()]
    if found:
        raise ReportError(f"Generated report contains forbidden content: {found}")


def main() -> int:
    args = parse_args()
    try:
        topic_key = validate_topic_key(args.topic_key)
        start_date = validate_iso_date("Start date", args.start_date)
        end_date = validate_iso_date("End date", args.end_date)
        if start_date > end_date:
            raise ReportError("Start date cannot be later than end date")
        report_date = validate_iso_date("Report date", args.report_date)
        evidence_cutoff = validate_iso_date("Evidence cutoff", args.evidence_cutoff or end_date)
        workbook_path = args.tagged_xlsx.expanduser().resolve()
        output_path = args.output_html.expanduser().resolve()
        if not workbook_path.is_file() or workbook_path.suffix.casefold() != ".xlsx":
            raise ReportError("Tagged input must be an existing .xlsx workbook")
        if output_path.suffix.casefold() != ".html":
            raise ReportError("Output must use the .html extension")
        if output_path.exists() and not args.overwrite:
            raise ReportError("Output exists; pass --overwrite to replace it")
        keywords = load_module(resolve_config(topic_key, "keywords"), f"briefing_keywords_{topic_key}")
        content = load_module(resolve_config(topic_key, "content"), f"briefing_content_{topic_key}")
        configured_topic = normalize(getattr(content, "TOPIC_KEY", ""))
        if configured_topic != topic_key:
            raise ReportError(f"Content configuration topic {configured_topic!r} does not match {topic_key!r}")
        categories = configured_categories(content)
        valid_category_ids = {normalize(item["category_id"]) for item in categories}
        category_keywords = keyword_categories(keywords)
        unknown_category_configs = sorted(set(category_keywords) - valid_category_ids)
        if unknown_category_configs:
            raise ReportError(f"Keyword configuration has unknown categories: {unknown_category_configs}")
        frame = pd.read_excel(workbook_path, engine="openpyxl")
        if frame.empty:
            raise ReportError("Tagged workbook has no records")
        mapped = canonical_columns(frame, keywords)
        records = records_from_frame(frame, mapped, category_keywords, valid_category_ids)
        aliases = getattr(keywords, "COLUMN_ALIASES")["publication_number"]
        images = embedded_images(workbook_path, ["publication_number", *aliases], {r.publication_number for r in records})
        search_scope = getattr(content, "REPORT_SCOPE", {})
        if not isinstance(search_scope, dict):
            raise ReportError("REPORT_SCOPE must be a dictionary")
        organization_leads = getattr(content, "ORGANIZATION_DISCOVERY_LEADS", [])
        if not isinstance(organization_leads, list):
            raise ReportError("ORGANIZATION_DISCOVERY_LEADS must be a list")
        context = ReportContext(
            topic_key=topic_key,
            topic_label=normalize(getattr(content, "TOPIC_LABEL", topic_key)),
            start_date=start_date,
            end_date=end_date,
            report_date=report_date,
            evidence_cutoff=evidence_cutoff,
            records=records,
            images=images,
            search_scope=search_scope,
            categories=categories,
            organization_leads=tuple(organization_leads),
        )
        html_text = build_html(context)
        validate_output(html_text, context)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        temporary = output_path.with_suffix(output_path.suffix + ".tmp")
        temporary.write_text(html_text, encoding="utf-8")
        temporary.replace(output_path)
        print(f"Included publication records: {len(records)}")
        print(f"Applicants in reviewed dataset: {len({record.applicant for record in records})}")
        print(f"Embedded reviewed figures: {len(images)}")
        print(f"Report written: {output_path}")
        return 0
    except (OSError, ValueError, ImportError) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
