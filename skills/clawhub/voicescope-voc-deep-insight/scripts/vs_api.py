#!/usr/bin/env python3
"""VoiceScope Skill CLI.

Default auth is browser-based device login. Developer API Key remains available
as a fallback through VOICESCOPE_API_KEY.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import platform
import sys
import time
import urllib.error
import urllib.request
import webbrowser
from pathlib import Path
from urllib.parse import urlparse

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

BASE = os.environ.get("VOICESCOPE_API_BASE", "https://voiceaiscope.com").rstrip("/")
PUB = f"{BASE}/api/public/v1"
AUTH_DIR = Path.home() / ".voicescope"
AUTH_FILE = AUTH_DIR / "auth.json"
DEV_API_KEY = os.environ.get("VOICESCOPE_API_KEY", "")

if "127.0.0.1" in BASE or "localhost" in BASE:
    _opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
else:
    _opener = urllib.request.build_opener()

EXIT_BY_CODE = {
    "auth_required": 2,
    "invalid_api_key": 2,
    "token_expired": 2,
    "token_revoked": 2,
    "quota_exceeded": 3,
    "task_not_ready": 4,
}


class ApiError(Exception):
    def __init__(self, status: int, code: str, message: str, payload: dict | None = None):
        super().__init__(message)
        self.status = status
        self.code = code
        self.message = message
        self.payload = payload or {}


def _print(data: dict) -> None:
    print(json.dumps(data, ensure_ascii=False, indent=2))


def _load_auth() -> dict:
    try:
        return json.loads(AUTH_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_auth(data: dict) -> None:
    AUTH_DIR.mkdir(parents=True, exist_ok=True)
    AUTH_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _clear_auth() -> None:
    try:
        AUTH_FILE.unlink()
    except FileNotFoundError:
        pass


def _request(method: str, path: str, body: dict | None = None, token: str | None = None) -> dict:
    headers = {"Content-Type": "application/json", "Accept": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(
        f"{PUB}{path}",
        data=json.dumps(body, ensure_ascii=False).encode("utf-8") if body is not None else None,
        method=method,
        headers=headers,
    )
    try:
        with _opener.open(req, timeout=120) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        try:
            payload = json.loads(e.read().decode("utf-8"))
        except Exception:
            payload = {}
        err = payload.get("error") if isinstance(payload, dict) else None
        if isinstance(err, dict):
            raise ApiError(e.code, err.get("code", "http_error"), err.get("message", f"HTTP {e.code}"), payload)
        raise ApiError(e.code, "http_error", str(err or payload or f"HTTP {e.code}"), payload)
    except urllib.error.URLError as e:
        raise ApiError(0, "network_error", f"Cannot connect to {BASE}: {e.reason}")


def _credential() -> tuple[str, str]:
    if DEV_API_KEY:
        return DEV_API_KEY, "api_key"
    auth = _load_auth()
    access = auth.get("access_token")
    if access:
        auth_base = str(auth.get("base_url") or "").rstrip("/")
        if auth_base and auth_base != BASE:
            raise ApiError(
                0,
                "auth_required",
                f"Saved login belongs to {auth_base}, but current API base is {BASE}. "
                "Run login again for the current environment.",
                {"login_command": "python scripts/vs_api.py login"},
            )
        return access, "access_token"
    raise ApiError(
        0,
        "auth_required",
        "Not logged in. Run: python scripts/vs_api.py login",
        {"login_command": "python scripts/vs_api.py login"},
    )


def _refresh_tokens() -> bool:
    auth = _load_auth()
    refresh = auth.get("refresh_token")
    if not refresh:
        return False
    try:
        data = _request("POST", "/auth/refresh", {"refresh_token": refresh})
    except ApiError:
        _clear_auth()
        return False
    _save_auth(
        {
            "access_token": data["access_token"],
            "refresh_token": data["refresh_token"],
            "token_type": data.get("token_type", "Bearer"),
            "created_at": int(time.time()),
            "base_url": BASE,
        }
    )
    return True


def call(method: str, path: str, body: dict | None = None, *, retry_refresh: bool = True) -> dict:
    token, source = _credential()
    try:
        return _request(method, path, body, token=token)
    except ApiError as e:
        if source == "access_token" and retry_refresh and e.code == "token_expired" and _refresh_tokens():
            return call(method, path, body, retry_refresh=False)
        if source == "access_token" and e.code == "token_revoked":
            _clear_auth()
        raise


def _request_bytes(method: str, path: str, token: str | None = None) -> bytes:
    """二进制下载（如 xlsx 导出）。出错时仍解析 JSON 错误体抛 ApiError。"""
    headers = {"Accept": "application/octet-stream"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(f"{PUB}{path}", method=method, headers=headers)
    try:
        with _opener.open(req, timeout=120) as resp:
            return resp.read()
    except urllib.error.HTTPError as e:
        try:
            payload = json.loads(e.read().decode("utf-8"))
        except Exception:
            payload = {}
        err = payload.get("error") if isinstance(payload, dict) else None
        if isinstance(err, dict):
            raise ApiError(e.code, err.get("code", "http_error"), err.get("message", f"HTTP {e.code}"), payload)
        raise ApiError(e.code, "http_error", str(err or payload or f"HTTP {e.code}"), payload)
    except urllib.error.URLError as e:
        raise ApiError(0, "network_error", f"Cannot connect to {BASE}: {e.reason}")


def call_bytes(method: str, path: str, *, retry_refresh: bool = True) -> bytes:
    token, source = _credential()
    try:
        return _request_bytes(method, path, token=token)
    except ApiError as e:
        if source == "access_token" and retry_refresh and e.code == "token_expired" and _refresh_tokens():
            return call_bytes(method, path, retry_refresh=False)
        if source == "access_token" and e.code == "token_revoked":
            _clear_auth()
        raise


def cmd_login(args) -> int:
    start = _request(
        "POST",
        "/auth/device/start",
        {
            "client_name": "VoiceScope Skill",
            "device_name": args.device_name or platform.node() or "Local machine",
            "scopes": "*",
        },
    )
    api_host = (urlparse(BASE).hostname or "").lower()
    verify_host = (urlparse(start.get("verification_uri", "")).hostname or "").lower()
    if verify_host and verify_host != api_host:
        print(
            f"⚠ 警告：服务端返回的授权地址 host（{verify_host}）与当前 API host（{api_host}）不一致。",
            file=sys.stderr, flush=True,
        )
        if verify_host in ("localhost", "127.0.0.1", "::1"):
            print(
                "  授权页指向本机 localhost，多半是后端 PUBLIC_CONSOLE_BASE_URL 没配成线上域名，"
                "浏览器会打开打不开的页面、登录无法完成。请让服务端修正后重试。",
                file=sys.stderr, flush=True,
            )
    print("Open this URL to authorize VoiceScope:", flush=True)
    print(start["verification_uri_complete"], flush=True)
    print(f"User code: {start['user_code']}", flush=True)
    if not args.no_browser:
        try:
            webbrowser.open(start["verification_uri_complete"])
        except Exception:
            pass

    deadline = time.time() + int(args.timeout)
    interval = max(int(start.get("interval") or 5), 2)
    while time.time() < deadline:
        try:
            data = _request("POST", "/auth/device/token", {"device_code": start["device_code"]})
            _save_auth(
                {
                    "access_token": data["access_token"],
                    "refresh_token": data["refresh_token"],
                    "token_type": data.get("token_type", "Bearer"),
                    "created_at": int(time.time()),
                    "base_url": BASE,
                }
            )
            who = call("GET", "/whoami")
            print(f"Logged in as {who.get('email') or who.get('user_id')}", flush=True)
            return 0
        except ApiError as e:
            if e.code in ("authorization_pending", "network_error"):
                time.sleep(interval)
                continue
            print(f"[{e.code}] {e.message}", file=sys.stderr)
            return EXIT_BY_CODE.get(e.code, 1)
    print("Login timed out. Run login again to get a new user code.", file=sys.stderr)
    return 1


def cmd_logout(_args) -> int:
    auth = _load_auth()
    access = auth.get("access_token")
    refresh = auth.get("refresh_token")
    if access or refresh:
        try:
            _request("POST", "/auth/logout", {"refresh_token": refresh}, token=access)
        except ApiError:
            pass
    _clear_auth()
    print("Logged out.")
    return 0


def cmd_auth_status(_args) -> int:
    if DEV_API_KEY:
        print("Using developer API Key from VOICESCOPE_API_KEY.")
    elif not _load_auth().get("access_token"):
        print("Not logged in. Run: python scripts/vs_api.py login")
        return 2
    try:
        return cmd_whoami(_args)
    except ApiError as e:
        if e.code == "token_expired" and _refresh_tokens():
            return cmd_whoami(_args)
        print(f"[{e.code}] {e.message}", file=sys.stderr)
        return EXIT_BY_CODE.get(e.code, 1)


def cmd_whoami(_args) -> int:
    data = call("GET", "/whoami")
    q = data.get("quota", {})
    print(f"User: {data.get('email') or data.get('user_id')}  plan: {data.get('plan')}")
    # 计量口径：月度 U 额度（行 × AI 列）= 公共 API 实际计量轴；已不再用每日行数门。
    included = q.get("monthly_included_units", 0)
    remaining = q.get("remaining", 0)
    used = q.get("used_units", 0)
    print(
        f"Credential: {data.get('credential_type')}  本月额度: 已用 {used} / 共 {included} U"
        f"（剩余 {remaining}）"
    )
    print(f"Console: {data.get('links', {}).get('console_url')}")
    return 0


def _read_lines(path: str) -> list[str]:
    with open(path, encoding="utf-8-sig") as f:
        return [ln.strip() for ln in f if ln.strip()]


def _read_xlsx(path: str) -> tuple[list[str], list[list]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ApiError(0, "missing_dependency", "Reading .xlsx requires openpyxl: pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            raise ApiError(0, "validation_failed", "Excel sheet is empty.")
        columns = [str(c).strip() if c is not None and str(c).strip() else f"col{i + 1}"
                   for i, c in enumerate(header)]
        rows: list[list] = []
        for r in rows_iter:
            vals = ["" if v is None else (v if isinstance(v, str) else str(v)) for v in r]
            vals = vals[: len(columns)]
            if any(str(v).strip() for v in vals):
                vals += [""] * (len(columns) - len(vals))
                rows.append(vals)
    finally:
        wb.close()
    if not rows:
        raise ApiError(0, "validation_failed", "Excel requires a header row and at least one data row.")
    return columns, rows


def _read_table(path: str) -> tuple[list[str], list[list]]:
    """读取 CSV/Excel/TXT 为 (columns, rows)。"""
    lower = path.lower()
    if lower.endswith(".csv"):
        with open(path, encoding="utf-8-sig", newline="") as f:
            reader = list(csv.reader(f))
        if len(reader) < 2:
            raise ApiError(0, "validation_failed", "CSV requires a header row and at least one data row.")
        return reader[0], reader[1:]
    if lower.endswith((".xlsx", ".xlsm")):
        return _read_xlsx(path)
    if lower.endswith(".xls"):
        raise ApiError(0, "validation_failed", "Legacy .xls is not supported; save as .xlsx and retry.")
    if lower.endswith(".txt"):
        rows = [[line] for line in _read_lines(path)]
        if not rows:
            raise ApiError(0, "validation_failed", "TXT file must contain at least one non-empty line.")
        return ["feedback"], rows
    raise ApiError(
        0,
        "validation_failed",
        "Unsupported file type. Use .csv, .xlsx, .xlsm, or .txt.",
    )


def _require_tabular_text_column(path: str, text_column: str | None) -> None:
    if path.lower().endswith((".csv", ".xlsx", ".xlsm")) and not text_column:
        raise ApiError(
            0,
            "validation_failed",
            "CSV/Excel requires --text-column after previewing and confirming the analysis column.",
        )


def _pick_text_rows(path: str, text_column: str | None) -> list[str]:
    """从文件抽出主文本列的非空值（CSV/Excel 按列名取，缺省第一列；TXT 按行）。"""
    columns, rows = _read_table(path)
    if text_column:
        if text_column not in columns:
            raise ApiError(0, "validation_failed", f"Column '{text_column}' not found. Available: {', '.join(columns)}")
        idx = columns.index(text_column)
    else:
        idx = 0
    out = []
    for r in rows:
        v = str(r[idx]).strip() if idx < len(r) and r[idx] is not None else ""
        if v:
            out.append(v)
    return out


def cmd_validate(args) -> int:
    with open(args.file, encoding="utf-8") as f:
        items = json.load(f)
    _print(call("POST", "/taxonomies/validate", {"items": items, "skip_llm": not args.llm}))
    return 0


def cmd_generate(args) -> int:
    _require_tabular_text_column(args.file, args.text_column)
    body = {"source_rows": _pick_text_rows(args.file, args.text_column)}
    if args.project:
        body["project_id"] = args.project
    data = call("POST", "/taxonomies/generate", body)
    print(f"Task submitted: task_id={data['task_id']} rows={len(body['source_rows'])}")
    print(f"Poll: python scripts/vs_api.py wait {data['task_id']}")
    return 0


def cmd_preview(args) -> int:
    """本地预览文件的列名和样例值，不上传。供向用户确认分析哪一列。"""
    columns, rows = _read_table(args.file)
    print(f"File: {args.file}")
    print(f"rows={len(rows)} columns={len(columns)}")
    for i, name in enumerate(columns):
        samples = []
        for r in rows[:20]:
            v = str(r[i]).strip() if i < len(r) and r[i] is not None else ""
            if v:
                samples.append(v[:40])
            if len(samples) >= 2:
                break
        print(f"  [{i}] {name}  e.g. {' | '.join(samples)}")
    return 0


def cmd_upload(args) -> int:
    _require_tabular_text_column(args.file, args.text_column)
    columns, rows = _read_table(args.file)

    body = {
        "name": args.name or os.path.splitext(os.path.basename(args.file))[0],
        "columns": columns,
        "rows": rows,
        "result_column_name": args.result_column,
    }
    if args.text_column:
        body["text_column"] = args.text_column
    if args.project:
        body["project_id"] = args.project
    data = call("POST", "/tables/import", body)
    print(f"Imported {data['row_count']} rows into {data['name']}")
    print(f"table_id={data['table_id']}")
    print(f"text_column_id={data['text_column_id']}")
    print(f"result_column_id={data.get('result_column_id')}")
    for c in data.get("columns", []):
        print(f"column {c['id']}  {c['name']}")
    return 0


def cmd_taxonomies(args) -> int:
    data = call("GET", "/taxonomies" + (f"?project_id={args.project}" if args.project else ""))
    taxes = data.get("taxonomies", [])
    print(f"Project {data.get('project_id')} taxonomies: {len(taxes)}")
    for t in taxes:
        preset = " preset" if t.get("is_preset") else ""
        print(f"{t['id']}  {t['name']}{preset}  tags={t.get('tag_count', 0)}")
    return 0


def cmd_draft(args) -> int:
    data = call("GET", f"/taxonomies/drafts/{args.task_id}")
    _print(data)
    _, project, open_link = _resolve_links(data)
    print("草案查看/入库：")
    if open_link != project:
        print(f"  • 直达链接（点开即定位到这份草案，未登录先登录再跳回）：{open_link}")
    else:
        print(f"  • 在线（需登录）：{project}")
    print("  展示草案给用户确认后，两种入库方式二选一：")
    print(f"    ① 直接确认入库：python scripts/vs_api.py confirm-draft {args.task_id} --name \"标签体系名\"")
    print("    ② 到控制台微调后入库（需人工精修时用）。")
    print("  入库后拿到 taxonomy_id 即可走 tag-batch 打标。")
    return 0


def cmd_confirm_draft(args) -> int:
    """把 generate 的草案确认入库为正式标签体系，返回 taxonomy_id。

    默认发布草案原 items；--items-file 指向一个 JSON 数组时，用编辑后的行入库
    （后端会按生成时 depth 重新校验）。入库前应已把草案展示给用户并得到确认。
    """
    body: dict = {"name": args.name}
    if args.description:
        body["description"] = args.description
    if args.items_file:
        with open(args.items_file, "r", encoding="utf-8") as f:
            edited = json.load(f)
        if not isinstance(edited, list):
            print("--items-file 必须是一个 JSON 数组（每行含 l1/l2，可选 l3/def/evidence）", file=sys.stderr)
            return 1
        body["edited_items"] = edited
    data = call("POST", f"/taxonomies/drafts/{args.task_id}/confirm", body)
    print(
        f"草案已入库：taxonomy_id={data['taxonomy_id']} name={data.get('name')} "
        f"tags={data.get('tag_count', 0)} project={data.get('project_id')}"
    )
    print(f"下一步打标：python scripts/vs_api.py tag-batch --table <table_id> --column <text_column_id> "
          f"--taxonomy {data['taxonomy_id']} --result-column <result_column_id>")
    return 0


def cmd_tag_batch(args) -> int:
    data = call(
        "POST",
        "/tag-batches",
        {
            "table_id": args.table,
            "column_id": args.column,
            "taxonomy_id": args.taxonomy,
            "result_column_id": args.result_column,
            "mode": args.mode,
        },
    )
    print(f"Tagging task submitted: task_id={data['task_id']}")
    print(f"Poll: python scripts/vs_api.py wait {data['task_id']}")
    return 0


def cmd_extract_viewpoints(args) -> int:
    body = {"table_id": args.table, "column_id": args.column}
    if args.result_column_id:
        body["result_column_id"] = args.result_column_id
    if args.background:
        body["background_knowledge"] = args.background
    data = call("POST", "/viewpoints/extract", body)
    result_col = data.get("result_column_id")
    print(f"Viewpoint extraction submitted: task_id={data['task_id']}")
    print(f"source_text_column_id={args.column}")
    print(f"viewpoint_column_id={result_col}")
    print("Note: cluster --column must use viewpoint_column_id, not the original text column.")
    print(f"Poll: python scripts/vs_api.py wait {data['task_id']}")
    if result_col:
        print(f"Then: python scripts/vs_api.py cluster --table {args.table} --column {result_col}")
    else:
        print("Then: fetch the task/table result column id, then run cluster with that viewpoint column.")
    return 0


def cmd_cluster(args) -> int:
    body = {"table_id": args.table, "column_id": args.column, "top_k": args.top_k}
    if args.force_refresh:
        body["force_refresh"] = True
    data = call("POST", "/viewpoints/cluster", body)
    print(f"Clustering task submitted: task_id={data['task_id']}")
    print(f"viewpoint_column_id={args.column}")
    print(f"Poll: python scripts/vs_api.py wait {data['task_id']}")
    print(f"Then: python scripts/vs_api.py cluster-results {data['task_id']}")
    return 0


BUCKET_NAMES = {"pos": "正面", "neg": "负面", "neu": "中性"}
BUCKET_TITLES = {"neg": "负面痛点", "pos": "正面亮点", "neu": "中性观察"}
BUCKET_ORDER = ("neg", "pos", "neu")


def _sample_text(sample) -> str:
    if isinstance(sample, dict):
        return str(sample.get("text") or sample.get("content") or sample.get("raw_text") or "").strip()
    return str(sample or "").strip()


def _resolve_links(data: dict) -> tuple[str, str, str]:
    """从响应里取控制台/工作台/结果深链；后端没给时回落。

    open：UI 深链（?openTaskId=xxx），直接打开这个任务的结果，跨刷新可分享、
    未登录会先提示登录再跳回。后端有 task_id 时才回，没有就回落到 project 首页。
    """
    links = data.get("links") or {}
    console = (links.get("console_url") or BASE).rstrip("/")
    project = links.get("project_url") or f"{console}/workbench"
    open_link = links.get("open_url") or project
    return console, project, open_link


def _print_where_to_view(
    data: dict,
    out_path: str | None,
    what: str,
    download_hint: str = "--out 文件名.csv 下载全部结果；也支持 .json。",
) -> None:
    """统一告诉用户结果在哪看：在线（需登录）+ 本地下载两个出口。"""
    _, project, open_link = _resolve_links(data)
    print("结果查看方式：")
    print(f"  • 在线（需登录 VoiceScope 控制台）：{open_link}")
    if open_link != project:
        print(f"    这是直达链接，点开即定位到本次结果（{what}）；未登录会先提示登录再跳回。正式结果写在你自己的项目里。")
    else:
        print(f"    登录后在工作台打开对应数据表，{what}。正式结果写在你自己的项目里。")
    if out_path:
        print(f"  • 本地文件（已下载）：{os.path.abspath(out_path)}")
    else:
        print(f"  • 本地文件：{download_hint}")


# 平台「强后续」引导：只在平台确有更丰富后续的分析类型上提示一句，把用户领去 Studio
# 继续做（词云 / 标签洞察 / 情感看板 / 报告）。摘要/场景/行动等无强后续，不提示，避免噪音。
# 当前为 project 级链接（用响应里的 project_url）；表/视图级深链待前端路由规则后再加。
PLATFORM_NEXT_STEPS = {
    "keyword": "平台 Studio 可把这列关键词一键生成**词云**，直观看高频词分布。",
    "sentiment": "平台有**情感分布看板**，可按时间 / 标签交叉看正负面占比。",
    "tag": "标签 **Studio** 分析很丰富——标签洞察、趋势、下钻、交叉透视，比终端强得多，建议去用起来。",
    "cluster": "平台可把聚类结果**保存为观点洞察报告**，并与标签做交叉分析。",
}


def _print_platform_next_step(data: dict, kind: str) -> None:
    tip = PLATFORM_NEXT_STEPS.get(kind)
    if not tip:
        return
    _, project, open_link = _resolve_links(data)
    print(f"下一步（平台更强）：{tip}")
    if open_link != project:
        print(f"  直达链接（点开即定位到本次结果，未登录先登录再跳回）：{open_link}")
    else:
        print(f"  登录工作台打开这张数据表继续：{project}")


def _bucket_pct(items: list, count: int) -> float:
    """桶内占比（与控制台报告导出同口径，保留 1 位小数）。"""
    total = sum(int((it or {}).get("count") or 0) for it in (items or []))
    return round((count / total) * 1000) / 10 if total > 0 else 0.0


def cmd_cluster_results(args) -> int:
    # --out *.xlsx：直接下载官方观点洞察 Excel（与 Studio「保存报告→导出」同一序列化，
    # 列为 观点/数量/占比/对应原声，按正负向分 sheet；不含内部缓存指纹）
    if args.out and args.out.lower().endswith(".xlsx"):
        os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
        content = call_bytes("GET", f"/viewpoints/clusters/{args.task_id}/export.xlsx")
        with open(args.out, "wb") as f:
            f.write(content)
        print(f"Exported viewpoint insight Excel to {os.path.abspath(args.out)}")

    data = call("GET", f"/viewpoints/clusters/{args.task_id}/results")
    if args.out and not args.out.lower().endswith(".xlsx"):
        os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
        if args.out.lower().endswith(".json"):
            with open(args.out, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        else:
            # CSV：列与官方导出对齐（观点/数量/占比/对应原声），额外带极性桶；无关键词
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(["polarity", "cluster", "count", "pct", "samples"])
            for bucket in BUCKET_ORDER:
                items = data.get(bucket) or []
                for it in items:
                    count = int(it.get("count") or 0)
                    writer.writerow([
                        BUCKET_NAMES[bucket],
                        it.get("label", ""),
                        count,
                        _bucket_pct(items, count),
                        " | ".join(_sample_text(s) for s in (it.get("samples") or [])[:3]),
                    ])
            with open(args.out, "w", encoding="utf-8-sig", newline="") as f:
                f.write(buf.getvalue())
        print(f"Exported clusters to {os.path.abspath(args.out)}")

    print(f"观点聚类摘要：total_fragments={data.get('total', 0)}")
    for bucket in BUCKET_ORDER:
        items = data.get(bucket) or []
        n = sum(int(it.get("count") or 0) for it in items)
        print(f"-- {BUCKET_TITLES[bucket]}（{bucket}）: {len(items)} clusters / {n} fragments")
        for it in items:
            print(f"  {int(it.get('count') or 0):>4}  {it.get('label')}")
            for sample in (it.get("samples") or [])[:2]:
                text = _sample_text(sample)
                if text:
                    print(f"      - {text[:120]}")
    _print_where_to_view(
        data,
        args.out,
        "查看观点聚类结果",
        "--out 文件名.xlsx 下载官方观点洞察 Excel（与控制台导出一致，推荐）；也支持 .csv / .json。",
    )
    _print_platform_next_step(data, "cluster")
    return 0


# ── 通用分析列（工作流 D）──────────────────────────────────────
# 复用后端 ai_analysis 引擎的 per-row 分析类型（情感判断/关键词/摘要/场景/行动），
# 与「添加自定义分析列」对应。tag / viewpoint 有独立工作流，不在此列。
# 端点契约：
#   POST /analysis  body={table_id, column_id, ai_type, result_column_id?, background_knowledge?, row_ids?}
#        -> {task_id, state, duplicate, result_column_id, links}（结果列缺省自动创建，按月度 U 计量）
#   GET  /analysis/{task_id}/results?offset&limit
#        -> {task_id, state, ai_type, total, results:[{row_id, source_text, value}], stats?, links}
ANALYSIS_TYPES = ("sentiment", "keyword", "summary", "scenario", "action")
ANALYSIS_TYPE_LABELS = {
    "sentiment": "情感判断",
    "keyword": "关键词提取",
    "summary": "摘要",
    "scenario": "场景识别",
    "action": "行动建议",
}


def cmd_analyze(args) -> int:
    body = {"table_id": args.table, "column_id": args.column, "ai_type": args.type}
    if args.result_column_id:
        body["result_column_id"] = args.result_column_id
    if args.background:
        body["background_knowledge"] = args.background
    try:
        data = call("POST", "/analysis", body)
    except ApiError as exc:
        if exc.status == 404:
            print(
                "当前后端不支持通用分析列端点 POST /analysis，请先确认服务端版本已部署。",
                file=sys.stderr,
            )
        raise
    label = ANALYSIS_TYPE_LABELS.get(args.type, args.type)
    result_col = data.get("result_column_id")
    print(f"{label}（{args.type}）分析已提交: task_id={data['task_id']}")
    print(f"source_column_id={args.column}")
    print(f"result_column_id={result_col}")
    print(f"Poll: python scripts/vs_api.py wait {data['task_id']}")
    print(f"Then: python scripts/vs_api.py analysis-results {data['task_id']}")
    return 0


def _analysis_value_str(value) -> str:
    if isinstance(value, list):
        return " | ".join(str(v).strip() for v in value if str(v).strip())
    return str(value if value is not None else "").strip()


def cmd_analysis_results(args) -> int:
    first = call("GET", f"/analysis/{args.task_id}/results?offset={args.offset}&limit={args.limit}")
    ai_type = first.get("ai_type") or ""
    label = ANALYSIS_TYPE_LABELS.get(ai_type, ai_type or "分析")

    if not args.out:
        preview = [
            {"row_id": r.get("row_id"), "value": _analysis_value_str(r.get("value"))}
            for r in (first.get("results") or [])[:5]
        ]
        _print({"task_id": first.get("task_id"), "ai_type": ai_type, "total": first.get("total", 0), "results": preview})
        print(f"total={first.get('total', 0)}（仅预览前 5 行；加 --out 导出全部）")
        _print_where_to_view(first, None, f"查看{label}结果列")
        _print_platform_next_step(first, ai_type)
        return 0

    all_rows = list(first.get("results") or [])
    total = int(first.get("total", 0))
    offset = args.offset + len(all_rows)
    while offset < total:
        page = call("GET", f"/analysis/{args.task_id}/results?offset={offset}&limit={args.limit}")
        rows = page.get("results") or []
        if not rows:
            raise ApiError(
                0,
                "incomplete_results",
                f"Analysis result pagination stopped at offset {offset} before total {total}; retry later.",
            )
        all_rows.extend(rows)
        offset += len(rows)

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    if args.out.lower().endswith(".json"):
        with open(args.out, "w", encoding="utf-8") as f:
            json.dump(all_rows, f, ensure_ascii=False, indent=2)
    else:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["row_id", "source_text", "value"])
        for r in all_rows:
            writer.writerow([r.get("row_id"), r.get("source_text", ""), _analysis_value_str(r.get("value"))])
        with open(args.out, "w", encoding="utf-8-sig", newline="") as f:
            f.write(buf.getvalue())

    filled = sum(1 for r in all_rows if _analysis_value_str(r.get("value")))
    nonempty_sources = sum(1 for r in all_rows if str(r.get("source_text") or "").strip())
    print(f"Exported {len(all_rows)} rows to {args.out}")
    print(f"{label}：有结果 {filled} / 空 {len(all_rows) - filled}（共 {len(all_rows)} 行）")
    if ai_type == "sentiment":
        dist: dict[str, int] = {}
        for r in all_rows:
            v = _analysis_value_str(r.get("value"))
            if v:
                dist[v] = dist.get(v, 0) + 1
        if dist:
            print("  情感分布：" + "，".join(f"{k} {v}" for k, v in sorted(dist.items(), key=lambda kv: -kv[1])))
    _print_where_to_view(first, args.out, f"查看{label}结果列")
    _print_platform_next_step(first, ai_type)
    if nonempty_sources and filled == 0:
        print(
            f"{label}任务显示完成，但 {nonempty_sources} 条非空源文本全部没有结果；"
            "请勿将本次任务视为成功，需检查后端写回或并发覆盖问题。",
            file=sys.stderr,
        )
        return 4
    return 0


def cmd_task(args) -> int:
    _print(call("GET", f"/tasks/{args.task_id}"))
    return 0


def cmd_task_report(args) -> int:
    _print(call("GET", f"/tasks/{args.task_id}/report"))
    return 0


def cmd_report(args) -> int:
    _print(call("GET", f"/reports/{args.report_id}"))
    return 0


def cmd_wait(args) -> int:
    deadline = time.time() + args.timeout
    delay = 2.0
    while True:
        data = call("GET", f"/tasks/{args.task_id}")
        state = data.get("state")
        total = data.get("total_rows")
        done = data.get("done_rows")
        progress = f"{done}/{total} rows" if total else f"{data.get('progress', 0)}%"
        print(f"[{time.strftime('%H:%M:%S')}] {state} {progress}")
        if state in ("done", "error", "cancelled"):
            if state == "error":
                print(data.get("error") or "Task failed", file=sys.stderr)
                return 4
            if state == "cancelled":
                print("Task was cancelled.", file=sys.stderr)
                return 4
            return 0
        if time.time() > deadline:
            print("Wait timed out.", file=sys.stderr)
            return 4
        time.sleep(delay)
        delay = min(delay * 1.5, 5.0)


def cmd_results(args) -> int:
    data = call("GET", f"/tag-batches/{args.task_id}/results?offset={args.offset}&limit={args.limit}")
    if not args.out:
        _print({**data, "results": data.get("results", [])[:5]})
        print(f"total={data.get('total', 0)}（仅预览前 5 行；加 --out 导出全部并查看命中统计）")
        _print_where_to_view(data, None, "查看 AI 标签结果列")
        _print_platform_next_step(data, "tag")
        return 0
    all_rows = list(data.get("results", []))
    total = int(data.get("total", 0))
    offset = args.offset + len(all_rows)
    while offset < total:
        page = call("GET", f"/tag-batches/{args.task_id}/results?offset={offset}&limit={args.limit}")
        rows = page.get("results", [])
        if not rows:
            raise ApiError(
                0,
                "incomplete_results",
                f"Tag result pagination stopped at offset {offset} before total {total}; retry later.",
            )
        all_rows.extend(rows)
        offset += len(rows)
    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    if args.out.lower().endswith(".json"):
        with open(args.out, "w", encoding="utf-8") as f:
            json.dump(all_rows, f, ensure_ascii=False, indent=2)
    else:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["row_id", "labels", "sentiments", "confidences"])
        for row in all_rows:
            tags = row.get("tags", [])
            writer.writerow(
                [
                    row.get("row_id"),
                    ";".join(t.get("label", "") for t in tags),
                    ";".join(t.get("sentiment", "") for t in tags),
                    ";".join(str(t.get("confidence", "")) for t in tags),
                ]
            )
        with open(args.out, "w", encoding="utf-8-sig", newline="") as f:
            f.write(buf.getvalue())
    hit = sum(1 for r in all_rows if r.get("tags"))
    miss = len(all_rows) - hit
    print(f"Exported {len(all_rows)} rows to {args.out}")
    print(f"命中 {hit} / 未命中 {miss}（共 {len(all_rows)} 行）")
    if miss:
        miss_ids = [str(r.get("row_id")) for r in all_rows if not r.get("tags")][:10]
        print(f"  未命中行 row_id 示例：{', '.join(miss_ids)}" + (" …" if miss > 10 else ""))
        print("  未命中＝没匹配到任何标签：可换更贴合的标签体系，或用 generate 从数据现生成一套再打标。")
    _print_where_to_view(data, args.out, "查看 AI 标签结果列")
    _print_platform_next_step(data, "tag")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="VoiceScope Skill CLI")
    sub = parser.add_subparsers(dest="cmd", required=True)

    login = sub.add_parser("login")
    login.add_argument("--no-browser", action="store_true")
    login.add_argument("--timeout", type=int, default=600)
    login.add_argument("--device-name", default=None)
    sub.add_parser("logout")
    sub.add_parser("auth-status")
    sub.add_parser("whoami")

    validate = sub.add_parser("validate")
    validate.add_argument("--file", required=True)
    validate.add_argument("--llm", action="store_true")
    generate = sub.add_parser("generate")
    generate.add_argument("--file", required=True)
    generate.add_argument("--text-column", default=None, dest="text_column")
    generate.add_argument("--project", default=None)
    preview = sub.add_parser("preview")
    preview.add_argument("--file", required=True)
    upload = sub.add_parser("upload")
    upload.add_argument("--file", required=True)
    upload.add_argument("--name", default=None)
    upload.add_argument("--text-column", default=None, dest="text_column")
    upload.add_argument("--result-column", default="AI 标签", dest="result_column")
    upload.add_argument("--project", default=None)
    taxonomies = sub.add_parser("taxonomies")
    taxonomies.add_argument("--project", default=None)
    draft = sub.add_parser("draft")
    draft.add_argument("task_id")
    confirm_draft = sub.add_parser("confirm-draft")
    confirm_draft.add_argument("task_id")
    confirm_draft.add_argument("--name", required=True)
    confirm_draft.add_argument("--description", default=None)
    confirm_draft.add_argument("--items-file", default=None, dest="items_file",
                               help="可选：编辑后草案的 JSON 数组文件（后端重新校验）")
    tag_batch = sub.add_parser("tag-batch")
    tag_batch.add_argument("--table", required=True)
    tag_batch.add_argument("--column", required=True)
    tag_batch.add_argument("--taxonomy", required=True)
    tag_batch.add_argument("--result-column", required=True, dest="result_column")
    tag_batch.add_argument("--mode", default="multi", choices=["single", "multi"])
    extract = sub.add_parser("extract-viewpoints")
    extract.add_argument("--table", required=True)
    extract.add_argument("--column", required=True)
    extract.add_argument("--result-column-id", default=None, dest="result_column_id")
    extract.add_argument("--background", default=None)
    cluster = sub.add_parser("cluster")
    cluster.add_argument("--table", required=True)
    cluster.add_argument("--column", required=True, help="观点结果列 id（extract-viewpoints 返回的 result_column_id）")
    cluster.add_argument("--top-k", type=int, default=10, dest="top_k")
    cluster.add_argument("--force-refresh", action="store_true", dest="force_refresh")
    cluster_results = sub.add_parser("cluster-results")
    cluster_results.add_argument("task_id")
    cluster_results.add_argument("--out", default=None)
    analyze = sub.add_parser("analyze")
    analyze.add_argument("--table", required=True)
    analyze.add_argument("--column", required=True)
    analyze.add_argument("--type", required=True, choices=list(ANALYSIS_TYPES), dest="type")
    analyze.add_argument("--result-column-id", default=None, dest="result_column_id")
    analyze.add_argument("--background", default=None)
    analysis_results = sub.add_parser("analysis-results")
    analysis_results.add_argument("task_id")
    analysis_results.add_argument("--offset", type=int, default=0)
    analysis_results.add_argument("--limit", type=int, default=200)
    analysis_results.add_argument("--out", default=None)
    task = sub.add_parser("task")
    task.add_argument("task_id")
    task_report = sub.add_parser("task-report")
    task_report.add_argument("task_id")
    report = sub.add_parser("report")
    report.add_argument("report_id")
    wait = sub.add_parser("wait")
    wait.add_argument("task_id")
    wait.add_argument("--timeout", type=int, default=600)
    results = sub.add_parser("results")
    results.add_argument("task_id")
    results.add_argument("--offset", type=int, default=0)
    results.add_argument("--limit", type=int, default=200)
    results.add_argument("--out", default=None)

    handlers = {
        "login": cmd_login,
        "logout": cmd_logout,
        "auth-status": cmd_auth_status,
        "whoami": cmd_whoami,
        "validate": cmd_validate,
        "generate": cmd_generate,
        "preview": cmd_preview,
        "upload": cmd_upload,
        "taxonomies": cmd_taxonomies,
        "draft": cmd_draft,
        "confirm-draft": cmd_confirm_draft,
        "tag-batch": cmd_tag_batch,
        "extract-viewpoints": cmd_extract_viewpoints,
        "cluster": cmd_cluster,
        "cluster-results": cmd_cluster_results,
        "analyze": cmd_analyze,
        "analysis-results": cmd_analysis_results,
        "task": cmd_task,
        "task-report": cmd_task_report,
        "report": cmd_report,
        "wait": cmd_wait,
        "results": cmd_results,
    }
    args = parser.parse_args()
    try:
        return handlers[args.cmd](args)
    except ApiError as exc:
        print(f"[{exc.code}] {exc.message}", file=sys.stderr)
        return EXIT_BY_CODE.get(exc.code, 1)
    except FileNotFoundError as exc:
        print(f"File not found: {exc.filename}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
