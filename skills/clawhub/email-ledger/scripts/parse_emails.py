#!/usr/bin/env python3
"""
报备邮件台账登记工具 - 核心解析脚本
功能：解析 .eml 邮件文件（含zip压缩包），提取关键信息，生成报备邮件台账 Excel
支持基于现有台账模板追加记录（按邮件主题+发送时间去重）

用法：
  python3 parse_emails.py --input <zip文件或eml目录> --output <输出xlsx路径> [--template <已有台账xlsx>]
  
示例：
  python3 parse_emails.py --input ./邮件.zip --output ./报备邮件台账_已登记.xlsx
  python3 parse_emails.py --input ./邮件.zip --output ./台账.xlsx --template ./报备邮件台账模板.xlsx
"""

import os
import sys
import email
import zipfile
import argparse
import re
from email.header import decode_header
from email.utils import parseaddr, parsedate_to_datetime
from datetime import datetime

# ============================================================
# 邮件解析工具函数
# ============================================================

def decode_str(s):
    """解码邮件头部字段"""
    if s is None:
        return ''
    parts = decode_header(s)
    result = []
    for data, charset in parts:
        if isinstance(data, bytes):
            result.append(data.decode(charset or 'utf-8', errors='replace'))
        else:
            result.append(data)
    return ''.join(result)


def parse_date(date_str):
    """解析邮件日期为统一格式 YYYY-MM-DD HH:MM"""
    if not date_str:
        return ''
    try:
        dt = parsedate_to_datetime(date_str)
        return dt.strftime('%Y-%m-%d %H:%M')
    except Exception:
        return date_str


def get_email_body(msg, max_chars=3000):
    """提取邮件正文（优先纯文本，回退HTML去标签）"""
    body = ''
    html_body = ''
    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            if ct == 'text/plain' and not body:
                payload = part.get_payload(decode=True)
                charset = part.get_content_charset() or 'utf-8'
                body = payload.decode(charset, errors='replace')
            elif ct == 'text/html' and not html_body:
                payload = part.get_payload(decode=True)
                charset = part.get_content_charset() or 'utf-8'
                html_body = payload.decode(charset, errors='replace')
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            charset = msg.get_content_charset() or 'utf-8'
            text = payload.decode(charset, errors='replace')
            if msg.get_content_type() == 'text/html':
                html_body = text
            else:
                body = text

    # 如果纯文本为空，使用HTML去标签
    if not body and html_body:
        body = re.sub(r'<[^>]+>', ' ', html_body)
        body = re.sub(r'\s+', ' ', body)

    return body.strip()[:max_chars]


def format_addr_list(raw):
    """格式化收件人/抄送人列表，提取姓名和邮箱"""
    if not raw:
        return ''
    addrs = []
    for item in raw.split(','):
        item = item.strip()
        if not item:
            continue
        name, addr = parseaddr(item)
        name = decode_str(name)
        # 优先显示姓名，没有姓名显示邮箱前缀
        if name and addr:
            display = f"{name}({addr.split('@')[0]})"
        elif addr:
            display = addr.split('@')[0]
        elif name:
            display = name
        else:
            display = item
        addrs.append(display)
    return '、'.join(addrs)


def extract_sender(raw):
    """提取发件人显示名"""
    if not raw:
        return ''
    name, addr = parseaddr(raw)
    name = decode_str(name)
    if name and addr:
        return f"{name}({addr.split('@')[0]})"
    elif name:
        return name
    elif addr:
        return addr
    return raw


# ============================================================
# 报备类型智能分类
# ============================================================

CATEGORY_PATTERNS = [
    # (模式关键词, 报备类型) — 按优先级排序，越具体的越靠前
    (r'商务结果|中标结果|采购结果', '商务结果报备'),
    (r'阳光承诺书.*免签|免签署.*阳光承诺书|免签.*承诺书|豁免.*阳光承诺书|免签.*阳光承诺书', '阳光承诺书免签申请'),
    (r'指定供应商|指定.*供应商', '指定供应商申请'),
    (r'新增简易供应商|简易供应商.*特批', '新增简易供应商/阳光承诺书特批'),
    (r'白名单', '白名单申请'),
    (r'授权自采|自购申请|申请授权自采', '授权自采申请'),
    (r'费用变更|变更申请|视频制作费用变更', '费用变更报备'),
    (r'特殊入库|特殊.*入库', '特殊入库报备'),
    (r'审批申请.*供应商新增|供应商新增', '供应商新增审批'),
    (r'审批类|审批申请', '审批类报备'),
    (r'采购方式申请', '采购方式报备'),
    (r'申请.*外采|外采.*申请|采购申请', '采购申请'),
    (r'报备', '一般报备'),
]

VENDOR_CHANNEL_PATTERNS = [
    # (模式, 供应商/渠道 分类)
    (r'场地|会议室|酒店|租赁|租房|园区|数码港|CyberArena', '渠道'),
]


def classify_report_type(subject, body=''):
    """根据邮件主题和正文分类报备类型"""
    text = subject + ' ' + body[:500]
    for pattern, category in CATEGORY_PATTERNS:
        if re.search(pattern, text):
            return category
    return '其他报备'


def classify_vendor_or_channel(subject, body=''):
    """判断是供应商还是渠道"""
    text = subject + ' ' + body[:500]
    for pattern, vc_type in VENDOR_CHANNEL_PATTERNS:
        if re.search(pattern, text):
            return vc_type
    return '供应商'


# ============================================================
# 供应商名称提取
# ============================================================

def extract_vendor_name(subject, body):
    """从邮件主题和正文中提取供应商/渠道名称
    
    核心思路：先从主题中按已知模式提取，再从正文中用"新增/指定/供应商名称"等语义线索提取，
    最后回退到正文中找公司名。每步都做清洗，避免提取到无关内容。
    """
    # ---------- 从主题提取 ----------
    
    # 【新增简易供应商】广东雷曦文化有限公司-申请特批
    m = re.search(r'[】\]]\s*([\u4e00-\u9fa5\w（）\(\)·\s]+?(?:有限公司|股份公司|集团公司|科技公司|公司))', subject)
    if m:
        return m.group(1).strip()

    # 【新增简易供应商】XXX、XXX-申请特批（多供应商，含外文名）
    m = re.search(r'[】\]]\s*(.+?)(?:\s*[-—]\s*申请|\s*[-—]\s*特批)', subject)
    if m and len(m.group(1).strip()) > 2:
        candidate = m.group(1).strip()
        # 排除明显非供应商名的关键词
        if not any(kw in candidate for kw in ['免签', '阳光', '承诺书', '入库', '报备']):
            return candidate

    # 白名单申请-禾赛科技
    m = re.search(r'白名单申请[-—]\s*(.+)', subject)
    if m:
        return m.group(1).strip()

    # 供应商嘉立创特殊入库 → 提取供应商后面的名称词
    m = re.search(r'供应商\s*([\u4e00-\u9fa5\w]+?)(?:特殊|入库|报备|免签|新增)', subject)
    if m and len(m.group(1).strip()) > 1:
        return m.group(1).strip()

    # ---------- 从正文提取 ----------
    
    # "对方（建信住房服务有限责任公司福建分公司）无法签署"
    m = re.search(r'对方[（(]([^)）]+?)[)）]', body[:800])
    if m:
        return m.group(1).strip()

    # "拟新增 杭州冠寓投资管理有限公司 为我司供应商"
    m = re.search(r'拟?\s*新增\s+(.{2,40}?)\s*(?:为.*?供应商|进入.*?供应商|加入)', body[:800])
    if m:
        return m.group(1).strip()

    # "指定当地供应商（枣庄市云鑫广告有限公司）制作"
    m = re.search(r'供应商[（(]([^)）]+?)[)）]', body[:800])
    if m:
        return m.group(1).strip()

    # "供应商名称：XXX"
    m = re.search(r'供应商名称\s*[:：]\s*([^\n,，。；;]{2,40})', body[:800])
    if m:
        return m.group(1).strip()

    # "预中标供应商为：思达"
    m = re.search(r'预?中标供应商[为：:\s]*([^，,。\n；;]{1,30})', body[:800])
    if m:
        candidate = m.group(1).strip()
        if len(candidate) > 0:
            return candidate

    # "与XXX合作"
    m = re.search(r'与\s*([\u4e00-\u9fa5\w（）\(\)·]+?(?:有限公司|股份公司|集团公司|科技公司|公司))\s*(?:沟通)?合作', body[:800])
    if m:
        return m.group(1).strip()

    # 渠道类：场地名称
    m = re.search(r'举办场地为\s*(.{2,40}?)(?:。|，|费用|因)', body[:800])
    if m:
        return m.group(1).strip()
    m = re.search(r'使用\s*(.{2,30}?)(?:作为|场地)', body[:800])
    if m:
        return m.group(1).strip()
    m = re.search(r'(?:场地|园区|会议)[为在]\s*(.{4,30}?(?:基地|中心|厅|港|场))', body[:800])
    if m:
        return m.group(1).strip()

    # ---------- 回退：正文中找公司名 ----------
    clean_body = body.split('-----原始邮件-----')[0]
    clean_body = clean_body.split('-----Original Message-----')[0]
    
    # 找所有公司名，排除"科大讯飞"（我方公司）
    companies = re.findall(r'([\u4e00-\u9fa5]{2,8}(?:有限公司|股份公司|集团公司|科技公司))', clean_body[:1000])
    companies = [c for c in companies if '科大讯飞' not in c]
    if companies:
        return companies[0]
    
    # 极端回退：外文公司名
    m = re.search(r'([A-Z][A-Za-z\s]+(?:CO\.,?\s*LTD\.?|INC\.?|CORP\.?|GMBH))', body[:800])
    if m:
        return m.group(1).strip()

    return ''


# ============================================================
# 审批人提取
# ============================================================

def extract_approver(subject, body, to_addr):
    """提取审批人
    
    优先从主题/正文中找明确的审批人称呼（如"余总"、"章总"、"徐总"），
    再从正文开头称呼推断，最后从收件人推断。避免提取到"您"等代词。
    """
    # 策略1：主题中 "请余总审批"、"请章总审批"、"请XX审阅"
    m = re.search(r'请\s*([^\s,，。；;]{1,8}总)\s*(?:审批|审阅)', subject)
    if m:
        return m.group(1).strip()

    # 策略2：主题中 "XXX审批" 且前面有称谓
    m = re.search(r'([^\s,，。；;]{1,8}总)\s*审批', subject)
    if m:
        return m.group(1).strip()

    # 策略3：正文中的 "请XXX审批" 或 "请@XXX审批"
    m = re.search(r'请\s*(@?[^\s@,，。；;]{1,10}?)\s*(?:依次\s*)?审批', body[:500])
    if m:
        candidate = m.group(1).strip().lstrip('@')
        if candidate and candidate not in ('您', '领导', '烦'):
            return candidate

    # 策略4：正文开头的称呼 "徐总：" / "韩总：" / "吴总好"
    m = re.search(r'^\s*([^\s,，。；;\n]{1,10}总)[：:，,\s您好]', body[:300])
    if m:
        return m.group(1).strip()

    # 策略5：正文中间的称呼 "XXX总好，" 
    m = re.search(r'([^\s,，。；;\n]{1,8}总)[好，,\s]', body[:100])
    if m:
        return m.group(1).strip()

    # 策略6：收件人推断
    if to_addr:
        name, addr = parseaddr(to_addr)
        name = decode_str(name)
        if name:
            return name

    return ''


# ============================================================
# 主要事由生成
# ============================================================

def generate_summary(subject, body, report_type):
    """生成主要事由摘要"""
    # 去除原始邮件引用部分
    main_body = body.split('-----原始邮件-----')[0]
    main_body = main_body.split('-----Original Message-----')[0]
    main_body = main_body.strip()

    # 去除签名部分
    main_body = re.sub(r'\n-{3,}.*', '', main_body, flags=re.DOTALL)

    if main_body:
        summary = main_body.replace('\n', ' ').replace('\r', ' ')
        summary = re.sub(r'\s+', ' ', summary).strip()
        if len(summary) > 200:
            summary = summary[:197] + '...'
        return summary

    return subject


# ============================================================
# 核心：解析单封邮件
# ============================================================

def parse_single_email(filepath):
    """解析单封 .eml 文件，返回台账记录字典"""
    with open(filepath, 'rb') as f:
        msg = email.message_from_bytes(f.read())

    subject = decode_str(msg.get('Subject', ''))
    from_raw = msg.get('From', '')
    to_raw = msg.get('To', '')
    cc_raw = msg.get('Cc', '')
    date_raw = msg.get('Date', '')

    body = get_email_body(msg)
    report_type = classify_report_type(subject, body)
    vendor_or_channel = classify_vendor_or_channel(subject, body)
    vendor_name = extract_vendor_name(subject, body)
    approver = extract_approver(subject, body, to_raw)
    summary = generate_summary(subject, body, report_type)

    record = {
        '发送时间': parse_date(date_raw),
        '发件人': extract_sender(from_raw),
        '邮件主题': subject,
        '收件人': format_addr_list(to_raw),
        '抄送人': format_addr_list(cc_raw),
        '报备类型': report_type,
        '供应商/渠道': vendor_or_channel,
        '供应商/渠道名称': vendor_name if vendor_name else '（需补充）',
        '主要事由': summary,
        '审批人': approver if approver else '（需补充）',
    }
    return record


# ============================================================
# 批量解析入口
# ============================================================

def parse_emails(input_path):
    """解析zip或目录中的所有eml文件，返回记录列表"""
    eml_files = []

    if zipfile.is_zipfile(input_path):
        with zipfile.ZipFile(input_path, 'r') as zf:
            import tempfile
            tmp_dir = tempfile.mkdtemp(prefix='email_ledger_')
            zf.extractall(tmp_dir)
            for root, dirs, files in os.walk(tmp_dir):
                for fname in files:
                    if fname.lower().endswith('.eml'):
                        eml_files.append(os.path.join(root, fname))
    elif os.path.isdir(input_path):
        for root, dirs, files in os.walk(input_path):
            for fname in files:
                if fname.lower().endswith('.eml'):
                    eml_files.append(os.path.join(root, fname))
    else:
        print(f"错误：输入路径既不是zip文件也不是目录：{input_path}", file=sys.stderr)
        return []

    eml_files.sort()
    records = []
    for fpath in eml_files:
        try:
            rec = parse_single_email(fpath)
            records.append(rec)
        except Exception as e:
            print(f"警告：解析 {os.path.basename(fpath)} 失败：{e}", file=sys.stderr)

    # 按发送时间排序
    records.sort(key=lambda r: r['发送时间'] or '9999')
    return records


# ============================================================
# 读取已有台账（模板合并 & 去重）
# ============================================================

def read_existing_ledger(template_path):
    """读取已有台账 xlsx 中的记录列表，返回 list[dict]"""
    import openpyxl
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue  # 跳过空行
        rec = {}
        for i, h in enumerate(headers):
            if i < len(row):
                rec[h] = row[i] if row[i] is not None else ''
        records.append(rec)
    return records, headers


def deduplicate_records(existing_records, new_records):
    """按 邮件主题+发送时间 去重，返回合并后的记录列表
    
    已有记录保持不变，新记录中与已有记录重复的跳过。
    """
    existing_keys = set()
    for r in existing_records:
        key = (r.get('邮件主题', ''), r.get('发送时间', ''))
        existing_keys.add(key)
    
    truly_new = []
    dup_count = 0
    for r in new_records:
        key = (r.get('邮件主题', ''), r.get('发送时间', ''))
        if key in existing_keys:
            dup_count += 1
        else:
            truly_new.append(r)
            existing_keys.add(key)
    
    return existing_records + truly_new, len(truly_new), dup_count


# ============================================================
# Excel 生成
# ============================================================

def generate_excel(records, output_path):
    """生成台账Excel文件"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    HEADERS = ['发送时间', '发件人', '邮件主题', '收件人', '抄送人', '报备类型',
               '供应商/渠道', '供应商/渠道名称', '主要事由', '审批人']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '报备邮件台账'

    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_align = Alignment(wrap_text=True, vertical='center')
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 写表头
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 写数据
    for row_idx, rec in enumerate(records, 2):
        for col_idx, key in enumerate(HEADERS, 1):
            val = rec.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = wrap_align
            # 待补充项标黄
            if '需补充' in str(val):
                cell.fill = yellow_fill

    # 列宽
    col_widths = [18, 28, 50, 32, 40, 24, 10, 42, 65, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 冻结首行 + 自动筛选
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:J{len(records)+1}"

    wb.save(output_path)
    return output_path


# ============================================================
# main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='报备邮件台账登记工具')
    parser.add_argument('--input', '-i', required=True, help='输入：邮件zip压缩包或eml文件目录')
    parser.add_argument('--output', '-o', required=True, help='输出：台账Excel文件路径')
    parser.add_argument('--template', '-t', default=None, help='可选：已有台账xlsx（追加记录+去重）')
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"错误：输入路径不存在：{args.input}", file=sys.stderr)
        sys.exit(1)

    # 解析新邮件
    print(f"正在解析邮件：{args.input}")
    new_records = parse_emails(args.input)

    if not new_records:
        print("未找到任何 .eml 邮件文件", file=sys.stderr)
        sys.exit(1)

    print(f"共解析 {len(new_records)} 封新邮件")

    # 如果有模板，合并去重
    final_records = new_records
    new_count = len(new_records)
    dup_count = 0

    if args.template and os.path.exists(args.template):
        print(f"正在读取已有台账：{args.template}")
        existing_records, existing_headers = read_existing_ledger(args.template)
        print(f"已有台账 {len(existing_records)} 条记录")
        
        final_records, new_count, dup_count = deduplicate_records(existing_records, new_records)
        print(f"合并结果：新增 {new_count} 条，重复跳过 {dup_count} 条，总计 {len(final_records)} 条")

    # 统计
    type_count = {}
    for r in final_records:
        t = r.get('报备类型', '未知')
        type_count[t] = type_count.get(t, 0) + 1
    print("\n报备类型分布：")
    for t, c in sorted(type_count.items(), key=lambda x: -x[1]):
        print(f"  {t}: {c}封")

    # 生成Excel
    output = generate_excel(final_records, args.output)
    print(f"\n✅ 台账已生成：{output}")
    print(f"   共 {len(final_records)} 条记录" + (f"（其中新增 {new_count} 条）" if args.template else ""))

    # 待补充项提醒
    need_vendor = sum(1 for r in final_records if '需补充' in str(r.get('供应商/渠道名称', '')))
    need_approver = sum(1 for r in final_records if '需补充' in str(r.get('审批人', '')))
    if need_vendor or need_approver:
        print(f"\n⚠️  需人工补充：供应商名称 {need_vendor} 条、审批人 {need_approver} 条（已标黄）")


if __name__ == '__main__':
    main()
