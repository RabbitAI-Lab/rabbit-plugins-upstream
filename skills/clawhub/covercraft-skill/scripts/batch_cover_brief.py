#!/usr/bin/env python3
"""Batch cover brief builder.

This helper converts a CSV/XLSX title table into a Markdown + JSON batch brief.
It does not call any LLM or image API. It only structures inputs so the Skill can
process many covers consistently.

Expected columns, all optional except title:
- title
- platform
- field
- ratio
- goal
- style
- reference_degree
- has_portrait
- notes

Usage:
  python scripts/batch_cover_brief.py titles.csv --out outputs/batch.md --json-out outputs/batch.json
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, Iterable, List


TITLE_TYPES = [
    ("避坑型", ["别", "不要", "千万", "避坑", "踩坑", "错误", "后悔", "注意"]),
    ("反常识型", ["原来", "没想到", "其实", "90%", "大多数", "都错", "真相"]),
    ("结果展示型", ["我用", "做出", "1分钟", "30天", "从0", "提升", "搞定"]),
    ("教程型", ["教程", "方法", "步骤", "指南", "手把手", "怎么", "如何"]),
    ("工具测评型", ["工具", "软件", "插件", "AI", "测评", "神器"]),
    ("对比型", ["对比", "vs", "VS", "区别", "以前", "现在", "小白", "高手"]),
    ("警告型", ["危险", "立刻", "停止", "警告", "别碰"]),
    ("情绪冲突型", ["离谱", "崩溃", "震惊", "惊了", "炸了", "太狠"]),
]

PLATFORM_DEFAULTS = {
    "B站": "16:9",
    "bilibili": "16:9",
    "YouTube": "16:9",
    "youtube": "16:9",
    "小红书": "3:4",
    "抖音": "9:16",
    "douyin": "9:16",
    "视频号": "9:16",
    "公众号": "16:9",
}


@dataclass
class CoverBrief:
    index: int
    title: str
    platform: str = "短视频平台"
    field: str = "未指定"
    ratio: str = "9:16"
    goal: str = "涨点击"
    style: str = "强点击但不低俗"
    reference_degree: str = "中度参考"
    has_portrait: str = "未指定"
    notes: str = ""
    inferred_title_type: str = "待模型判断"
    suggested_keyword: str = "待模型提炼"
    suggested_emotion: str = "待模型匹配"
    suggested_gesture: str = "待模型匹配"


def read_csv(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_xlsx(path: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("Reading XLSX requires openpyxl. Install it or use CSV.") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    data: List[Dict[str, Any]] = []
    for row in rows[1:]:
        item: Dict[str, Any] = {}
        for i, header in enumerate(headers):
            if header:
                item[header] = row[i] if i < len(row) else ""
        data.append(item)
    return data


def infer_title_type(title: str) -> str:
    for label, keys in TITLE_TYPES:
        if any(k.lower() in title.lower() for k in keys):
            return label
    return "干货型/待模型细分"


def suggest_by_type(title_type: str) -> tuple[str, str]:
    if "避坑" in title_type or "警告" in title_type:
        return "严肃警告", "单手停止 / 摆手拒绝"
    if "反常识" in title_type:
        return "疑惑或震惊", "托下巴 / 捂脸"
    if "结果" in title_type:
        return "兴奋展示", "双手张开 / 展示成果"
    if "教程" in title_type or "工具" in title_type:
        return "自信讲解", "指向侧边标题区 / 拿手机展示"
    if "对比" in title_type:
        return "惊讶对照", "指向左右对比"
    if "情绪" in title_type:
        return "强情绪", "捂脸 / 指向镜头"
    return "自信、亲和", "指向镜头 / 指向标题区"


def normalize_row(index: int, row: Dict[str, Any]) -> CoverBrief:
    title = str(row.get("title") or row.get("标题") or "").strip()
    if not title:
        raise ValueError(f"Row {index} has no title/标题 column value.")

    platform = str(row.get("platform") or row.get("平台") or "短视频平台").strip()
    ratio = str(row.get("ratio") or row.get("比例") or PLATFORM_DEFAULTS.get(platform, "9:16")).strip()
    title_type = infer_title_type(title)
    emotion, gesture = suggest_by_type(title_type)

    return CoverBrief(
        index=index,
        title=title,
        platform=platform,
        field=str(row.get("field") or row.get("领域") or "未指定").strip(),
        ratio=ratio,
        goal=str(row.get("goal") or row.get("目标") or "涨点击").strip(),
        style=str(row.get("style") or row.get("风格") or "强点击但不低俗").strip(),
        reference_degree=str(row.get("reference_degree") or row.get("参考程度") or "中度参考").strip(),
        has_portrait=str(row.get("has_portrait") or row.get("是否有人像") or "未指定").strip(),
        notes=str(row.get("notes") or row.get("备注") or "").strip(),
        inferred_title_type=title_type,
        suggested_emotion=emotion,
        suggested_gesture=gesture,
    )


def render_markdown(briefs: Iterable[CoverBrief]) -> str:
    lines = [
        "# 批量封面生产简报",
        "",
        "> 本文件由 batch_cover_brief.py 生成。它只做结构化整理，不替代 Skill 的策略判断。",
        "",
        "| 编号 | 标题 | 平台 | 比例 | 领域 | 初判类型 | 建议情绪 | 建议手势 | 目标 |",
        "|---:|---|---|---|---|---|---|---|---|",
    ]
    for b in briefs:
        lines.append(
            f"| {b.index} | {b.title} | {b.platform} | {b.ratio} | {b.field} | "
            f"{b.inferred_title_type} | {b.suggested_emotion} | {b.suggested_gesture} | {b.goal} |"
        )
    lines.append("")
    lines.append("## 给 Skill 的批量处理提示")
    lines.append("")
    lines.append("请基于上表逐条输出：标题类型、最大关键词、人物动作、画面结构、主色/强调色、无字底图提示词摘要、后期排版摘要、预估评分。")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build batch cover briefs from CSV/XLSX.")
    parser.add_argument("input", type=Path, help="Input CSV or XLSX with a title/标题 column.")
    parser.add_argument("--out", type=Path, default=Path("batch_cover_briefs.md"), help="Markdown output path.")
    parser.add_argument("--json-out", type=Path, default=None, help="Optional JSON output path.")
    args = parser.parse_args()

    if not args.input.exists():
        print(f"Input not found: {args.input}", file=sys.stderr)
        return 2

    try:
        if args.input.suffix.lower() in {".xlsx", ".xlsm"}:
            rows = read_xlsx(args.input)
        else:
            rows = read_csv(args.input)
        briefs = [normalize_row(i + 1, row) for i, row in enumerate(rows)]
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(render_markdown(briefs), encoding="utf-8")

    if args.json_out:
        args.json_out.parent.mkdir(parents=True, exist_ok=True)
        args.json_out.write_text(json.dumps([asdict(b) for b in briefs], ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote {args.out}")
    if args.json_out:
        print(f"Wrote {args.json_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
