#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""xlsx_tool.py — 基于 openpyxl 的表格清洗与交付前检查辅助工具。

子命令：
    dedup <文件> [--sheet 名] [--cols A,B]   按指定列去重，输出到新文件
    check <文件> [--sheet 名]                交付前检查（sheet 清单、尺寸、空行、公式、冻结窗格）

用法示例：
    python scripts/xlsx_tool.py dedup 数据.xlsx --cols A,B
    python scripts/xlsx_tool.py dedup 数据.xlsx --sheet 明细 --cols A -o 输出.xlsx
    python scripts/xlsx_tool.py check 数据.xlsx

依赖：openpyxl（pip install openpyxl）。
"""

import argparse
import os
import shutil
import sys

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string, get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


def load(path, data_only=False):
    """加载工作簿，附统一的中文错误提示。"""
    if not os.path.isfile(path):
        print(f"[错误] 找不到文件：{path}")
        sys.exit(1)
    try:
        return load_workbook(path, data_only=data_only)
    except Exception as exc:  # openpyxl 会抛出多种异常，统一兜底
        print(f"[错误] 无法打开 {path}：{exc}")
        print("       请确认文件是有效的 .xlsx（.xls/.et 旧格式请先在 WPS/Excel 中另存为 .xlsx）。")
        sys.exit(1)


def pick_sheet(wb, sheet_name):
    """按名称选择工作表；未指定时返回活动工作表。"""
    if sheet_name is None:
        return wb.active
    if sheet_name not in wb.sheetnames:
        print(f"[错误] 工作簿中不存在工作表：{sheet_name}")
        print("       可用的工作表：" + ", ".join(wb.sheetnames))
        sys.exit(1)
    return wb[sheet_name]


def cmd_dedup(args):
    """按指定列去重：保留首次出现的行，输出到新文件。操作前自动备份原文件。"""
    wb = load(args.file)
    ws = pick_sheet(wb, args.sheet)

    # 解析去重列：默认第 A 列
    col_letters = [c.strip().upper() for c in (args.cols or "A").split(",") if c.strip()]
    try:
        col_indexes = [column_index_from_string(c) for c in col_letters]
    except ValueError:
        print(f"[错误] 列名无效：{args.cols}（应为 A、B、C 这类字母列名）")
        sys.exit(1)

    # 备份原文件（不覆盖已有备份）
    backup_path = os.path.splitext(args.file)[0] + ".bak.xlsx"
    if os.path.exists(backup_path):
        print(f"[信息] 备份文件已存在，跳过备份：{backup_path}")
    else:
        shutil.copy2(args.file, backup_path)
        print(f"[信息] 已备份原文件：{backup_path}")

    # 保留表头（第 1 行），从第 2 行开始去重
    header = [cell.value for cell in ws[1]]
    seen = set()
    kept_rows = []
    removed = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = tuple(row[i - 1] if i - 1 < len(row) else None for i in col_indexes)
        if key in seen:
            removed += 1
            continue
        seen.add(key)
        kept_rows.append(row)

    # 清空原数据区并重写
    ws.delete_rows(2, ws.max_row)
    for row in kept_rows:
        ws.append(row)

    output = args.output or os.path.splitext(args.file)[0] + "-dedup.xlsx"
    wb.save(output)
    cols_desc = ",".join(get_column_letter(i) for i in col_indexes)
    print(f"[完成] 按列 {cols_desc} 去重：删除 {removed} 行，保留 {len(kept_rows)} 行（不含表头）。")
    print(f"[完成] 结果已写入：{output}")


def count_empty_rows(ws):
    """统计完全为空的行数（在已使用区域内）。"""
    empty = 0
    for row in ws.iter_rows(values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            empty += 1
    return empty


def count_formula_cells(ws):
    """统计公式单元格数量（值以 = 开头的字符串）。"""
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                count += 1
    return count


def cmd_check(args):
    """交付前检查：打印工作表清单、各表尺寸、空行数、公式数量、冻结窗格状态。"""
    wb = load(args.file)
    print(f"[文件] {args.file}")
    print(f"[工作表清单] 共 {len(wb.sheetnames)} 个：" + ", ".join(wb.sheetnames))

    targets = [pick_sheet(wb, args.sheet)] if args.sheet else [wb[name] for name in wb.sheetnames]
    for ws in targets:
        empty_rows = count_empty_rows(ws)
        formulas = count_formula_cells(ws)
        freeze = ws.freeze_panes or "无"
        print(f"  - {ws.title}：{ws.max_row} 行 × {ws.max_column} 列，"
              f"空行 {empty_rows} 个，公式单元格 {formulas} 个，冻结窗格：{freeze}")

    if args.sheet is None and wb.sheetnames:
        print("[提示] 可用 --sheet 名 只检查指定工作表。")


def main():
    parser = argparse.ArgumentParser(
        description="基于 openpyxl 的 .xlsx 清洗与交付前检查工具。",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "示例：\n"
            "  python scripts/xlsx_tool.py dedup 数据.xlsx --cols A,B\n"
            "  python scripts/xlsx_tool.py dedup 数据.xlsx --sheet 明细 -o 清洗后.xlsx\n"
            "  python scripts/xlsx_tool.py check 数据.xlsx"
        ),
    )
    subparsers = parser.add_subparsers(dest="command", required=True, metavar="子命令")

    p_dedup = subparsers.add_parser("dedup", help="按指定列去重（自动备份原文件）")
    p_dedup.add_argument("file", help="待处理的 .xlsx 文件")
    p_dedup.add_argument("--sheet", default=None, help="工作表名称（默认活动工作表）")
    p_dedup.add_argument("--cols", default="A", help="去重依据列，如 A 或 A,B（默认 A）")
    p_dedup.add_argument("-o", "--output", default=None, help="输出文件（默认 原名-dedup.xlsx）")
    p_dedup.set_defaults(func=cmd_dedup)

    p_check = subparsers.add_parser("check", help="交付前检查（sheet、尺寸、空行、公式、冻结窗格）")
    p_check.add_argument("file", help="待检查的 .xlsx 文件")
    p_check.add_argument("--sheet", default=None, help="只检查指定工作表（默认全部）")
    p_check.set_defaults(func=cmd_check)

    args = parser.parse_args()

    if not _OPENPYXL_OK:
        print("[错误] 缺少第三方库 openpyxl，请先安装：")
        print("       pip install openpyxl")
        sys.exit(2)

    args.func(args)


if __name__ == "__main__":
    main()
