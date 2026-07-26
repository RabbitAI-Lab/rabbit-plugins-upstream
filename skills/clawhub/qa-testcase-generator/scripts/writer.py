#!/usr/bin/env python3
"""
Test Case Excel Writer
Reads structured JSON from stdin/file, writes formatted Excel.
No business logic - formatting (output) only.

Supports both Chinese and English field names.

Usage:
    python scripts/writer.py input.json                              # 输出路径从 JSON 内部取
    python scripts/writer.py input.json --output output/report.xlsx  # 命令行覆盖输出路径
    echo '<JSON>' | python scripts/writer.py                         # 从标准输入
    python scripts/writer.py -                                       # 从标准输入
"""

import json
import sys
import argparse
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === Styles ===
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
MODULE_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
MODULE_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)

PRIORITY_STYLES = {
    "P0": {"fill": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"), "font": Font(name="微软雅黑", size=9, bold=True, color="FFFFFF")},
    "P1": {"fill": PatternFill(start_color="FFA726", end_color="FFA726", fill_type="solid"), "font": Font(name="微软雅黑", size=9, bold=True)},
    "P2": {"fill": PatternFill(start_color="81C784", end_color="81C784", fill_type="solid"), "font": Font(name="微软雅黑", size=9)},
    "P3": {"fill": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"), "font": Font(name="微软雅黑", size=9)},
}

BASE_FONT = Font(name="微软雅黑", size=9)
WRAP = Alignment(wrap_text=True, vertical="top")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Columns: 序号, 用例编号, 优先级, 测试维度, 用例类型, 业务域, 设计方法,
#           测试场景, 测试点, 操作步骤, 预期结果, 测试数据, 前置条件, 需求来源, 测试结果
HEADERS = [
    "序号", "用例编号", "优先级", "测试维度", "用例类型", "业务域", "设计方法",
    "测试场景", "测试点", "操作步骤", "预期结果", "测试数据",
    "前置条件", "需求来源", "测试结果",
]

COL_WIDTHS = [6, 12, 7, 9, 10, 12, 10, 25, 20, 30, 25, 20, 20, 10, 10]


def write_excel(test_cases, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # Header row
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = BORDER
    ws.freeze_panes = "A2"

    for col, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 2
    current_module = None
    seq_num = 1

    for tc in test_cases:
        domain = tc.get("业务域", tc.get("business_domain", tc.get("module", "未分类")))

        # Module separator
        if domain != current_module:
            current_module = domain
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
            cell = ws.cell(row=row, column=1, value=f"【{domain}】")
            cell.font = MODULE_FONT
            cell.fill = MODULE_FILL
            cell.alignment = WRAP
            cell.border = BORDER
            ws.row_dimensions[row].height = 25
            row += 1

        # Build steps and expected text (paired by step number)
        steps = tc.get("操作步骤", tc.get("steps", []))
        if isinstance(steps, list):
            steps_lines = []
            expected_lines = []
            for s in steps:
                step_num = s.get("步骤", s.get("step", ""))
                action = s.get("操作", s.get("action", ""))
                expected = s.get("预期", s.get("expected", ""))
                steps_lines.append(f"步骤{step_num}: {action}")
                expected_lines.append(f"步骤{step_num}: {expected}")
            steps_text = "\n".join(steps_lines)
            expected_text = "\n".join(expected_lines)
        else:
            steps_text = str(steps)
            expected_text = ""

        # Map fields - support both Chinese and English
        case_id = tc.get("用例编号", tc.get("id", ""))
        priority = tc.get("优先级", tc.get("priority", ""))
        test_dim = tc.get("测试维度", tc.get("test_dimension", ""))
        case_type = tc.get("用例类型", tc.get("case_type", ""))
        design = tc.get("设计方法", tc.get("design_method", ""))
        scenario = tc.get("测试场景", tc.get("scenario", ""))
        test_point = tc.get("测试点", tc.get("test_point", ""))
        test_data = tc.get("测试数据", tc.get("test_data", ""))
        precondition = tc.get("前置条件", tc.get("precondition", ""))
        req_source = tc.get("需求来源", tc.get("requirement_source", ""))

        row_data = [
            seq_num, case_id, priority, test_dim, case_type, domain,
            design, scenario, test_point, steps_text, expected_text,
            test_data, precondition, req_source, "",
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = BASE_FONT
            cell.alignment = WRAP
            cell.border = BORDER

            # Priority column styling (column 3)
            if col == 3:
                pstyle = PRIORITY_STYLES.get(str(value))
                if pstyle:
                    cell.fill = pstyle["fill"]
                    cell.font = pstyle["font"]

        # Row height
        n = len(steps) if isinstance(steps, list) else 1
        ws.row_dimensions[row].height = max(20, n * 25)
        row += 1
        seq_num += 1

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="测试用例 Excel 生成器")
    parser.add_argument("input", nargs="?", default="-",
                        help="输入 JSON 文件路径（默认: 标准输入, 使用 - 显式指定标准输入）")
    parser.add_argument("-o", "--output", help="输出 Excel 文件路径（覆盖 JSON 中的输出路径）")
    args = parser.parse_args()

    # Read input JSON
    data = None
    if args.input == "-":
        try:
            data = json.load(sys.stdin)
        except json.JSONDecodeError as e:
            print(f"错误: 标准输入 JSON 解析失败 - {e}", file=sys.stderr)
            print("提示: 请检查 JSON 是否格式正确", file=sys.stderr)
            sys.exit(1)
    else:
        src = args.input
        src_path = Path(src)
        if not src_path.exists():
            print(f"错误: 文件不存在 - {src}", file=sys.stderr)
            sys.exit(1)
        try:
            with open(src, "r", encoding="utf-8-sig") as f:
                data = json.load(f)
        except json.JSONDecodeError as e:
            print(f"错误: JSON 解析失败 - {e}", file=sys.stderr)
            print("提示: 请检查 JSON 是否格式正确（常见问题：中英文引号混用、尾部多余逗号）", file=sys.stderr)
            sys.exit(1)

    # Extract test cases
    test_cases = data.get("测试用例", data.get("test_cases", []))

    if not isinstance(test_cases, list):
        print("错误: '测试用例' / 'test_cases' 字段应为数组", file=sys.stderr)
        sys.exit(1)

    if len(test_cases) == 0:
        print("错误: '测试用例' 列表为空，无法生成 Excel", file=sys.stderr)
        sys.exit(1)

    # Determine output path: CLI > JSON > default
    output = args.output or data.get("输出路径", data.get("output", ""))
    if not output:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = f"output/testcases_{ts}.xlsx"

    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    path = write_excel(test_cases, output)
    print(f"测试用例已生成: {path}")


if __name__ == "__main__":
    main()
