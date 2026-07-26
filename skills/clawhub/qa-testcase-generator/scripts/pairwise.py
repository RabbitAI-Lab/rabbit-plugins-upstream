#!/usr/bin/env python3
"""
Pairwise (All-Pairs) Test Combination Generator
================================================

Generates minimal test case combinations that cover all possible pairs
of parameter values. Useful when testing config combinations, browser/OS
matrices, permission roles, or any multi-parameter scenario where
exhaustive combination (Cartesian product) would be prohibitively large.

Input: JSON structure defining parameters and their values.
Output: JSON array of minimal pair-covering combinations + optional Excel.

Example input (stdin or file):
{
  "parameters": [
    {"name": "浏览器", "values": ["Chrome", "Firefox", "Edge", "Safari"]},
    {"name": "操作系统", "values": ["Windows", "macOS", "Linux"]},
    {"name": "用户角色", "values": ["管理员", "普通用户", "访客"]}
  ]
}

Usage:
  # From file
  python scripts/pairwise.py params.json

  # From stdin
  cat params.json | python scripts/pairwise.py

  # With output file
  python scripts/pairwise.py params.json -o combinations.xlsx

Algorithm: Uses a greedy in-parameter-order strategy with horizontal
extension to achieve pair coverage. While not guaranteed optimal, it
produces near-minimal test suites for practical use.
"""

import json
import sys
import itertools
import random
from pathlib import Path
from datetime import datetime
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


def read_input():
    """Read JSON input from file or stdin."""
    if len(sys.argv) > 1 and sys.argv[1] != "-":
        src = sys.argv[1]
        src_path = Path(src)
        if not src_path.exists():
            print(f"错误: 文件不存在 - {src}", file=sys.stderr)
            sys.exit(1)
        with open(src, "r", encoding="utf-8-sig") as f:
            data = json.load(f)
    else:
        data = json.load(sys.stdin)
    return data


def parse_output_path(data):
    """Extract output path from data or command-line args."""
    output = data.get("output", "")
    if output:
        return output

    for i, arg in enumerate(sys.argv[1:]):
        if arg in ("-o", "--output") and i + 2 < len(sys.argv):
            return sys.argv[i + 2]
        if arg == "--excel":
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            return f"output/pairwise_{ts}.xlsx"

    return ""


def generate_pairwise(params):
    """
    Greedy in-parameter-order (IPO-like) pairwise generation.

    Args:
        params: list of {"name": str, "values": list}

    Returns:
        list of dicts, each dict mapping name -> value
    """
    if not params:
        return []

    names = [p["name"] for p in params]
    values = [p["values"] for p in params]

    # Track uncovered pairs: {(param_i, value, param_j, value)}
    uncovered = set()
    for i in range(len(params)):
        for j in range(i + 1, len(params)):
            for vi in values[i]:
                for vj in values[j]:
                    uncovered.add((i, vi, j, vj))

    # Collect all value options per parameter for horizontal filling
    all_value_sets = [set(v) for v in values]

    result = []
    # Use combinations of first two params as seeds
    for vi in values[0]:
        for vj in values[1] if len(values) > 1 else [None]:
            if len(values) == 1:
                row = {names[0]: vi}
            else:
                row = {names[0]: vi, names[1]: vj}
            result.append(row)

    # Fill remaining params for each row, greedily covering most pairs
    for row in result:
        for pi in range(2, len(params)):
            best_val = None
            best_count = -1
            pname = names[pi]
            for candidate in values[pi]:
                count = 0
                for pj in range(pi):
                    pj_name = names[pj]
                    pj_val = row.get(pj_name)
                    if pj_val is not None and (pj, pj_val, pi, candidate) in uncovered:
                        count += 1
                if count > best_count:
                    best_count = count
                    best_val = candidate
            if best_val is not None:
                row[pname] = best_val

    # Remove covered pairs
    for row in result:
        for i in range(len(params)):
            for j in range(i + 1, len(params)):
                vi = row.get(names[i])
                vj = row.get(names[j])
                if vi is not None and vj is not None:
                    uncovered.discard((i, vi, j, vj))

    # Add rows for remaining uncovered pairs (horizontal extension)
    max_attempts = 10000
    attempt = 0
    while uncovered and attempt < max_attempts:
        attempt += 1
        row = {}
        # Pick the first uncovered pair as seed
        pi, vi, pj, vj = next(iter(uncovered))
        # Start with all values unassigned
        assigned = {pi: vi, pj: vj}

        # Fill remaining params with best coverage
        for pk in range(len(params)):
            if pk in assigned:
                continue
            best_val = None
            best_count = -1
            for candidate in values[pk]:
                count = 0
                for pk2 in range(len(params)):
                    if pk2 == pk:
                        continue
                    pk2_val = assigned.get(pk2)
                    if pk2_val is not None:
                        if (pk2, pk2_val, pk, candidate) in uncovered:
                            count += 1
                if count > best_count:
                    best_count = count
                    best_val = candidate
            if best_val is not None:
                assigned[pk] = best_val

        row = {names[k]: assigned.get(k, random.choice(values[k])) for k in range(len(params))}

        # Remove newly covered pairs
        for i in range(len(params)):
            for j in range(i + 1, len(params)):
                vi = row.get(names[i])
                vj = row.get(names[j])
                if vi is not None and vj is not None:
                    uncovered.discard((i, vi, j, vj))

        result.append(row)

    return result


def format_table(result):
    """Format pairwise result as a markdown table."""
    if not result:
        return "(空)"
    names = list(result[0].keys())
    lines = []
    header = "| 序号 | " + " | ".join(names) + " |"
    sep = "|------|" + "|".join("------" for _ in names) + "|"
    lines.append(header)
    lines.append(sep)
    for i, row in enumerate(result, 1):
        vals = [str(row.get(n, "")) for n in names]
        lines.append(f"| {i} | " + " | ".join(vals) + " |")
    return "\n".join(lines)


def write_excel(result, output_path):
    """Write pairwise combinations to Excel."""
    if not HAS_EXCEL:
        print("警告: 未安装 openpyxl，跳过 Excel 输出", file=sys.stderr)
        return None

    if not result:
        print("警告: 组合为空，跳过 Excel 输出", file=sys.stderr)
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Pairwise组合"

    names = list(result[0].keys())
    headers = ["序号"] + names

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    BASE_FONT = Font(name="微软雅黑", size=9)
    WRAP = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = BORDER

    ws.freeze_panes = "A2"

    for col, h in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = max(12, len(h) * 2)

    for i, row in enumerate(result, 2):
        ws.cell(row=i, column=1, value=i - 1).font = BASE_FONT
        ws.cell(row=i, column=1).alignment = WRAP
        ws.cell(row=i, column=1).border = BORDER
        for j, name in enumerate(names):
            cell = ws.cell(row=i, column=j + 2, value=str(row.get(name, "")))
            cell.font = BASE_FONT
            cell.alignment = WRAP
            cell.border = BORDER
        ws.row_dimensions[i].height = 20

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def print_statistics(result, params, full_count):
    """Print comparison stats."""
    if not result:
        return
    pair_count = len(result)
    print(f"\n=== 统计信息 ===")
    print(f"参数总数: {len(params)}")
    print(f"全组合（笛卡尔积）: {full_count:,} 条")
    print(f"Pairwise 组合: {pair_count} 条")
    print(f"缩减率: {100 - (pair_count / full_count * 100):.1f}%")
    print(f"节省测试量: {full_count - pair_count:,} 条")


def main():
    data = read_input()

    params = data.get("parameters", data.get("params", []))
    if not params:
        print("错误: 缺少 'parameters' 字段。输入格式应为 {\"parameters\": [...]}", file=sys.stderr)
        sys.exit(1)

    # Validate
    for p in params:
        if "name" not in p or "values" not in p:
            print(f"错误: 每个参数需要 'name' 和 'values' 字段。出错: {p}", file=sys.stderr)
            sys.exit(1)
        if not isinstance(p["values"], list) or len(p["values"]) < 1:
            print(f"错误: 参数 '{p['name']}' 的 values 至少需要 1 个值", file=sys.stderr)
            sys.exit(1)

    # Calculate full Cartesian product count
    full_count = 1
    for p in params:
        full_count *= len(p["values"])

    print(f"正在生成 pairwise 组合...")
    print(f"参数: {', '.join(p['name'] + '(' + str(len(p['values'])) + ')' for p in params)}")

    # Generate
    result = generate_pairwise(params)

    # Output JSON
    output = {
        "parameters": params,
        "total_combinations": len(result),
        "full_cartesian_product": full_count,
        "combinations": result,
    }

    # Print markdown table
    print("\n=== Pairwise 组合表 ===")
    print(format_table(result))

    # Stats
    print_statistics(result, params, full_count)

    # Write JSON output if requested
    json_output = data.get("output_json", "")
    if json_output:
        json_path = Path(json_output)
        json_path.parent.mkdir(parents=True, exist_ok=True)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        print(f"\nJSON 已保存: {json_path}")

    # Try Excel output
    xlsx_path = parse_output_path(data)
    if xlsx_path:
        written = write_excel(result, xlsx_path)
        if written:
            print(f"\nExcel 已保存: {written}")

    # For pipeline usage, write full output to stdout as JSON
    if parse_output_path(data) == "" and json_output == "":
        print("\n--- JSON 输出 ---")
        print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
