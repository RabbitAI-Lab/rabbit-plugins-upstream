"""
ahkb_extract_xlsx.py — XLSX 解析器
提取：各 sheet 表格数据 + 嵌入图片/媒体
输出：chunks 嵌套结构，按 sheet 分块
"""
from pathlib import Path
import zipfile, hashlib

# ─── 媒体扩展名 ───

IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.wmf', '.emf', '.svg'}
VIDEO_EXTS = {'.mp4', '.avi', '.mov', '.wmv', '.m4v'}
AUDIO_EXTS = {'.mp3', '.wav', '.wma', '.aac'}
MEDIA_EXTS = IMAGE_EXTS | VIDEO_EXTS | AUDIO_EXTS


def _get_media_type(ext):
    ext = ext.lower()
    if ext in IMAGE_EXTS:
        return "image"
    elif ext in VIDEO_EXTS:
        return "video"
    elif ext in AUDIO_EXTS:
        return "audio"
    return "other"


def extract_xlsx(filepath, workspace):
    """Extract data and media from XLSX. Returns structured dict with chunks."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    base = Path(filepath).stem
    safe_base = "".join(c if c.isalnum() or c in '-_ ' else '_' for c in base)

    # 目录结构
    img_dir = Path(workspace) / "图片及其他资源" / "images"
    video_dir = Path(workspace) / "图片及其他资源" / "videos"
    audio_dir = Path(workspace) / "图片及其他资源" / "audios"
    other_dir = Path(workspace) / "图片及其他资源" / "others"
    for d in [img_dir, video_dir, audio_dir, other_dir]:
        d.mkdir(parents=True, exist_ok=True)

    result = {
        "file": str(filepath),
        "type": "xlsx",
        "metadata": {
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
        },
        "chunks": [],
        "full_text": "",
        "resources_flat": [],
    }

    # ── 第一步：提取每个 sheet 的文本数据 ──
    sheets_data = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = []
        for row in ws.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in row_data):
                rows_data.append(row_data)

        sheet_text = f"## Sheet: {sheet_name}\n"
        for row in rows_data:
            sheet_text += " | ".join(row) + "\n"

        sheets_data.append({
            "name": sheet_name,
            "row_count": len(rows_data),
            "header": rows_data[0] if rows_data else [],
            "sample_rows": rows_data[:10],
            "full_text": sheet_text,
        })
        result["full_text"] += f"\n\n--- Sheet: {sheet_name} ---\n\n{sheet_text}"

    wb.close()

    # ── 第二步：从 ZIP 中提取媒体文件（xl/media/）──
    # XLSX 中的图片/媒体通常放在 xl/media/ 目录
    # 它们在哪个 sheet 需要通过 XML 分析定位，这里简化处理：
    # 找出所有媒体文件，将上下文设为所有 sheet 文本的拼接

    media_files = []  # [{filename, ext, bytes, type}]
    seen_blobs = set()

    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            for name in z.namelist():
                if name.startswith("xl/media/"):
                    ext = Path(name).suffix.lower()
                    if ext not in MEDIA_EXTS:
                        continue
                    fname = Path(name).name
                    img_bytes = z.read(name)
                    blob_hash = hashlib.md5(img_bytes).hexdigest()[:16]
                    if blob_hash in seen_blobs:
                        continue
                    seen_blobs.add(blob_hash)

                    mtype = _get_media_type(ext)
                    target_dir = {"image": img_dir, "video": video_dir, "audio": audio_dir, "other": other_dir}[mtype]

                    count = len(media_files) + 1
                    out_name = f"{safe_base}-{mtype}{count:02d}{ext}"
                    with open(target_dir / out_name, "wb") as f:
                        f.write(img_bytes)

                    media_files.append({
                        "type": mtype,
                        "filename": out_name,
                        "ext": ext[1:],
                    })

            # 尝试建立 sheet 与图片的关联（通过 XML 分析）
            # 对于 XLSX，图片在 xl/drawings/drawingN.xml 中通过
            # xdr:twoCellAnchor 指定锚定的单元格
            # 我们简化处理：将所有媒体归入第一个有数据的 sheet
            try:
                # 读取 xl/_rels/workbook.xml.rels 找绘图关系
                drawing_map = {}  # sheet_idx → media filenames
                # 简化：所有媒体的上下文文本是全部 sheet 文本
                all_sheet_text = "\n\n".join([
                    f"--- {s['name']} ---\n{s['full_text']}"
                    for s in sheets_data
                ])
                for m in media_files:
                    m["context_text"] = all_sheet_text[:2000]  # 限长
            except Exception:
                for m in media_files:
                    m["context_text"] = ""
    except Exception as e:
        pass

    # ── 第三步：构建 chunks ──
    # 按 sheet 分块，每个 sheet 为一个 chunk
    # 如果 sheet 有媒体，关联到该 sheet

    # 简单策略：所有媒体归入第一个有内容的 sheet's chunk
    # 更好的策略是通过 XML 解析来确定，但 XLSX 的绘图定位较复杂
    # 这里用简化方式
    for s_idx, sheet in enumerate(sheets_data):
        # 为该 sheet 分配媒体（第一 sheet 拿全部，或者平均分配）
        if s_idx == 0:
            chunk_resources = []
            for m in media_files:
                chunk_resources.append({
                    "type": m["type"],
                    "filename": m["filename"],
                    "ext": m["ext"],
                    "context_text": m.get("context_text", sheet["full_text"]),
                    "source_ref": f"sheet '{sheet['name']}' - xl/media/",
                })
        else:
            chunk_resources = []

        chunk = {
            "id": f"sheet-{s_idx+1:03d}",
            "heading": sheet["name"],
            "source_position": f"sheet '{sheet['name']}'",
            "type": "sheet",
            "text": sheet["full_text"],
            "resources": chunk_resources,
        }
        result["chunks"].append(chunk)

    # ── 扁平资源列表 ──
    flat = []
    for chunk in result["chunks"]:
        for r in chunk["resources"]:
            r_copy = dict(r)
            r_copy["belongs_to_chunk"] = chunk["id"]
            r_copy["chunk_heading"] = chunk["heading"]
            r_copy["chunk_text"] = chunk["text"]
            flat.append(r_copy)
    result["resources_flat"] = flat

    return result
