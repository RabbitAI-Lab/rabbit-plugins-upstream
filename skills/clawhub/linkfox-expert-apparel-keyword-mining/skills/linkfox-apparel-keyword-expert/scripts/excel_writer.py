#!/usr/bin/env python3
"""
Excel multi-sheet writer for Apparel Keyword Expert.
Sheets: 摘要 / 打标明细 / 肯定词库 / 否定词库 / 待确认词库 / 完整属性短语 / 原始挖掘结果
"""

from datetime import datetime
from typing import List, Dict, Any, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _style_header(ws, col_count):
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.freeze_panes = 'A2'


def _auto_width(ws, col_count, min_w=10, max_w=55):
    for col_idx in range(1, col_count + 1):
        max_len = min_w
        for row_idx in range(1, min(ws.max_row + 1, 200)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                display_len = sum(2 if ord(c) > 127 else 1 for c in str(val))
                if display_len > max_len:
                    max_len = display_len
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_w)


def _write_tagged_sheet(ws, items: List[Dict], title_headers: Optional[List[str]] = None):
    headers = title_headers or [
        "keyword", "primary_type", "secondary_types", "attribute_categories",
        "is_complete_attribute_phrase", "relevance", "relevance_reason",
        "library", "suggested_positions", "confidence", "notes"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, len(headers))

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.get("keyword", ""))
        ws.cell(row=row_idx, column=2, value=item.get("primary_type", ""))
        ws.cell(row=row_idx, column=3, value=", ".join(item.get("secondary_types") or []))
        ws.cell(row=row_idx, column=4, value=", ".join(item.get("attribute_categories") or []))
        ws.cell(row=row_idx, column=5, value=bool(item.get("is_complete_attribute_phrase", False)))
        ws.cell(row=row_idx, column=6, value=item.get("relevance", ""))
        ws.cell(row=row_idx, column=7, value=item.get("relevance_reason", ""))
        ws.cell(row=row_idx, column=8, value=item.get("library", ""))
        ws.cell(row=row_idx, column=9, value=", ".join(str(x) for x in (item.get("suggested_positions") or [])))
        ws.cell(row=row_idx, column=10, value=item.get("confidence", ""))
        ws.cell(row=row_idx, column=11, value=item.get("notes", ""))

    _auto_width(ws, len(headers))


def write_apparel_excel(
    output_path: str,
    seed: str,
    product_context: str,
    tagged_results: List[Dict],
    positive: List[Dict],
    negative: List[Dict],
    review: List[Dict],
    raw_keywords: List[Any],
    summary: Dict,
    modes: List[str],
    market: str = "US",
) -> str:
    """
    Write the final multi-sheet Excel for the apparel expert.
    Returns the output path.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ---------- 1. 摘要 ----------
    ws = wb.active
    ws.title = "摘要"
    summary_rows = [
        ["种子词", seed],
        ["产品上下文", product_context or "（未提供）"],
        ["站点", market],
        ["挖掘模式", ", ".join(modes)],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [],
        ["--- 统计 ---", ""],
        ["关键词总数（打标后）", len(tagged_results)],
        ["肯定词库", summary.get("positive_count", len(positive))],
        ["否定词库", summary.get("negative_count", len(negative))],
        ["待确认词库", summary.get("review_count", len(review))],
        ["高相关词数量", summary.get("high_relevance_count", 0)],
        ["完整属性短语数量", summary.get("complete_attribute_phrase_count", 0)],
        ["原始挖掘词数量", len(raw_keywords)],
    ]
    for r_idx, row in enumerate(summary_rows, 1):
        if len(row) == 2:
            ws.cell(row=r_idx, column=1, value=row[0]).font = Font(bold=True)
            ws.cell(row=r_idx, column=2, value=row[1])
        elif len(row) == 1:
            ws.cell(row=r_idx, column=1, value=row[0]).font = Font(bold=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 70

    # ---------- 2. 打标明细 ----------
    ws_detail = wb.create_sheet("打标明细")
    _write_tagged_sheet(ws_detail, tagged_results)

    # ---------- 3. 肯定词库 ----------
    # sort: high relevance first, then confidence desc
    positive_sorted = sorted(
        positive,
        key=lambda x: (
            0 if x.get("relevance") == "high" else 1 if x.get("relevance") == "medium" else 2,
            -float(x.get("confidence") or 0),
        ),
    )
    ws_pos = wb.create_sheet("肯定词库")
    _write_tagged_sheet(ws_pos, positive_sorted)

    # ---------- 4. 否定词库 ----------
    ws_neg = wb.create_sheet("否定词库")
    _write_tagged_sheet(ws_neg, negative)

    # ---------- 5. 待确认词库 ----------
    ws_rev = wb.create_sheet("待确认词库")
    _write_tagged_sheet(ws_rev, review)

    # ---------- 6. 完整属性短语 ----------
    complete_phrases = [r for r in tagged_results if r.get("is_complete_attribute_phrase")]
    # group by primary_type
    from collections import defaultdict
    groups = defaultdict(list)
    for item in complete_phrases:
        key = item.get("primary_type") or "Other"
        groups[key].append(item)

    ws_attr = wb.create_sheet("完整属性短语")
    headers = ["primary_type", "keyword", "attribute_categories", "relevance", "suggested_positions", "confidence"]
    for col, h in enumerate(headers, 1):
        ws_attr.cell(row=1, column=col, value=h)
    _style_header(ws_attr, len(headers))

    row_idx = 2
    for ptype in sorted(groups.keys()):
        for item in groups[ptype]:
            ws_attr.cell(row=row_idx, column=1, value=ptype)
            ws_attr.cell(row=row_idx, column=2, value=item.get("keyword", ""))
            ws_attr.cell(row=row_idx, column=3, value=", ".join(item.get("attribute_categories") or []))
            ws_attr.cell(row=row_idx, column=4, value=item.get("relevance", ""))
            ws_attr.cell(row=row_idx, column=5, value=", ".join(item.get("suggested_positions") or []))
            ws_attr.cell(row=row_idx, column=6, value=item.get("confidence", ""))
            row_idx += 1
    _auto_width(ws_attr, len(headers))

    # ---------- 7. 原始挖掘结果 ----------
    ws_raw = wb.create_sheet("原始挖掘结果")
    ws_raw.cell(row=1, column=1, value="keyword")
    _style_header(ws_raw, 1)
    for r_idx, kw in enumerate(raw_keywords, 2):
        if isinstance(kw, dict):
            ws_raw.cell(row=r_idx, column=1, value=kw.get("keyword", str(kw)))
        else:
            ws_raw.cell(row=r_idx, column=1, value=str(kw))
    _auto_width(ws_raw, 1)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # minimal smoke
    demo_results = [
        {
            "keyword": "above the knee midi dress",
            "primary_type": "Core Product",
            "secondary_types": ["Dress Length"],
            "attribute_categories": ["Dress Length"],
            "is_complete_attribute_phrase": True,
            "relevance": "high",
            "relevance_reason": "demo",
            "library": "positive",
            "suggested_positions": ["title", "bullet"],
            "confidence": 0.93,
            "notes": "",
        }
    ]
    path = "/tmp/apparel_demo.xlsx"
    write_apparel_excel(
        output_path=path,
        seed="summer dress",
        product_context="Women's casual summer dresses",
        tagged_results=demo_results,
        positive=demo_results,
        negative=[],
        review=[],
        raw_keywords=["summer dress", "above the knee midi dress"],
        summary={"positive_count": 1, "negative_count": 0, "review_count": 0,
                 "high_relevance_count": 1, "complete_attribute_phrase_count": 1},
        modes=["expand", "az", "numbers", "reverse", "gap"],
        market="US",
    )
    print(f"Demo written to {path}")
