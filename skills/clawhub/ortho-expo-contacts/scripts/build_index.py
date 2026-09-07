# -*- coding: utf-8 -*-
"""
build_index.py — 骨科展会名录索引构建器

从「展会源目录」读取展会名录 Excel（只读，绝不改动源文件），
归一化去重后写入本地 SQLite + FTS5 全文索引。

源目录定位（三选一，优先级从高到低）：
  1. --src <目录>       命令行指定
  2. ORTHO_EXPO_SRC     环境变量
  3. <技能目录>/sources  默认相对目录（把展会 Excel 按子文件夹放进来即可）
子目录约定：01-AAOS / 03-SOFCOT / 04-AOSSM / 05-OMTEC / 06-AAHKS / 08-DKOU
（未提供的展会自动跳过，不影响其余构建）

数据分级：
  L1 公开级 — 展会官方参展商名录（公司/官网/展位/公开邮箱）
  L2 受限级 — OMTEC 个人参会者名单（姓名/职务/工作邮箱）
  L3 私密级 — 自有客户表与供应商台账（手机/WhatsApp），默认不入库

用法：
  python build_index.py                        # 构建（默认不含 L3）
  python build_index.py --src <展会Excel目录>   # 指定源目录
  python build_index.py --include-l3           # 含 L3（仅本人使用，勿外传）
  python build_index.py --stats                # 只看统计，不重建

依赖：仅本脚本需要 openpyxl（解析展会 Excel 用）。
  缺省环境无该库时会打印安装指引；日常查询 query.py / 闸门 gate.py 零依赖（纯标准库）。
"""
import argparse
import hashlib
import os
import re
import sqlite3
import sys
import warnings
from datetime import datetime

try:
    import openpyxl
except ModuleNotFoundError:
    openpyxl = None  # 懒加载：仅构建路径需要，缺失时 main() 打印指引退出

warnings.filterwarnings("ignore")

SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_ROOT = os.environ.get("ORTHO_EXPO_SRC",
                          os.path.join(SKILL_DIR, "sources"))
DB_PATH = os.path.join(SKILL_DIR, "data", "contacts.db")

# ---------------------------------------------------------------- 归一化工具

_SUFFIX = re.compile(
    r"\b(inc|llc|ltd|limited|corp|corporation|co|company|gmbh|ag|sa|sas|srl|spa|bv|nv|ab|as|oy|plc|pte|pty|kk|k\.k\.|co\.|inc\.|ltd\.|llc\.|gmbh\.|ag\.|s\.a\.|b\.v\.|group|holding)\b[.,]?\s*",
    re.I,
)
_PUNCT = re.compile(r"[^\w\s]", re.U)


def norm_company(name: str) -> str:
    """公司名归一化：小写、去法律后缀、去标点空格 —— 用于跨展会去重与撞车检测"""
    if not name:
        return ""
    s = str(name).lower().strip()
    s = _SUFFIX.sub(" ", s)
    s = _PUNCT.sub(" ", s)
    s = re.sub(r"\s+", "", s)
    return s


def clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none", "#value!", "n/a", "-", "--"):
        return ""
    return s


def split_emails(s: str):
    if not s:
        return []
    return re.findall(r"[\w.+-]+@[\w-]+\.[\w.-]+", s)


_PHONE_RE = re.compile(r"(\+?\d[\d\s\-().]{6,}\d)")


def looks_like_phone(s: str) -> bool:
    if not s or "@" in s or re.search(r"http|www\.", s, re.I):
        return False
    digits = re.sub(r"\D", "", s)
    return 7 <= len(digits) <= 20


def split_contact(s: str):
    """从"联系方式"混合列里拆出 (邮箱, 电话)

    实测坑：AAOS 名录该列 479 行里 7 个真邮箱 + 65 个电话 + 407 空，
    只按 @ 提取会把 65 个电话全丢掉。
    """
    if not s:
        return "", ""
    mails = split_emails(s)
    email = mails[0] if mails else ""
    phone = ""
    m = _PHONE_RE.search(s)
    if m and looks_like_phone(m.group(1)):
        phone = m.group(1).strip()
    if not email and not phone and looks_like_phone(s):
        phone = s.strip()
    return email, phone


# ---------------------------------------------------------------- 表结构

SCHEMA = """
DROP TABLE IF EXISTS contacts;
DROP TABLE IF EXISTS contacts_fts;
DROP TABLE IF EXISTS meta;

CREATE TABLE contacts (
    id           INTEGER PRIMARY KEY,
    source       TEXT NOT NULL,
    tier         TEXT NOT NULL,
    kind         TEXT NOT NULL,
    company      TEXT,
    company_norm TEXT,
    person       TEXT,
    title        TEXT,
    country      TEXT,
    city         TEXT,
    category     TEXT,
    website      TEXT,
    booth        TEXT,
    email        TEXT,
    phone        TEXT,
    note         TEXT,
    src_file     TEXT,
    src_sheet    TEXT,
    src_row      INTEGER
);

CREATE VIRTUAL TABLE contacts_fts USING fts5(
    company, person, title, country, city, category,
    content='contacts', content_rowid='id', tokenize='unicode61'
);

CREATE TABLE meta (k TEXT PRIMARY KEY, v TEXT);
CREATE INDEX idx_norm  ON contacts(company_norm);
CREATE INDEX idx_tier  ON contacts(tier);
CREATE INDEX idx_src   ON contacts(source);
"""


# ---------------------------------------------------------------- 各展会解析器

def add(rows, *, source, tier, kind, company, person="", title="", country="",
        city="", category="", website="", booth="", email="", phone="",
        note="", src_file="", src_sheet="", src_row=0):
    company = clean(company)
    if not company and not person:
        return
    rows.append(dict(
        source=source, tier=tier, kind=kind,
        company=company, company_norm=norm_company(company),
        person=clean(person), title=clean(title),
        country=clean(country), city=clean(city),
        category=clean(category), website=clean(website),
        booth=clean(booth), email=clean(email), phone=clean(phone),
        note=clean(note), src_file=src_file, src_sheet=src_sheet, src_row=src_row,
    ))


def parse_aaos(rows):
    """AAOS 2026 参展商名录（含公开邮箱）— L1"""
    for fn, sheets in [
        ("AAOS 2026 参展商名录.xlsx", ["AAOS 2026 参展商", "AAOS 2026 参展商 (2)"]),
        ("AAOS 2026 参展商名录-1.xlsx", ["AAOS 2026 参展商"]),
    ]:
        p = os.path.join(SRC_ROOT, "01-AAOS", fn)
        if not os.path.exists(p):
            continue
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        for ws in wb.worksheets:
            if sheets and ws.title not in sheets:
                continue
            hdr = None
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                vals = [clean(v) for v in row]
                if not any(vals):
                    continue
                if hdr is None:
                    if "公司名称" in vals:
                        hdr = {h: j for j, h in enumerate(vals) if h}
                        continue
                    continue

                def g(key):
                    j = hdr.get(key)
                    return vals[j] if (j is not None and j < len(vals)) else ""

                email, phone = split_contact(g("联系方式/邮箱"))
                add(rows, source="AAOS 2026", tier="L1", kind="exhibitor",
                    company=g("公司名称"), country=g("国家"), city=g("城市"),
                    category=" / ".join(x for x in [g("主营业务"), g("产品分类")] if x),
                    website=g("公司官方网站"), booth=g("展台位置"),
                    email=email, phone=phone,
                    note=g("备注") or g("参展历史"),
                    src_file=fn, src_sheet=ws.title, src_row=i)
        wb.close()


def parse_aaos_customers(rows):
    """AAOS 展会客户信息表（自有客户，含手机/WhatsApp）— L3"""
    p = os.path.join(SRC_ROOT, "01-AAOS", "AAOS展会客户信息表-修改.xlsx")
    if not os.path.exists(p):
        return
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [clean(v) for v in row]
        if not any(vals):
            continue
        if hdr is None:
            if "公司名称" in vals:
                hdr = {h: j for j, h in enumerate(vals) if h}
            continue

        def g(key):
            j = hdr.get(key)
            return vals[j] if (j is not None and j < len(vals)) else ""

        add(rows, source="AAOS 2026", tier="L3", kind="person",
            company=g("公司名称"), person=g("联系人名称"), title=g("职位"),
            country=g("国家"), category=g("key product"),
            website=g("website"), email=g("邮箱"), phone=g("电话/whatsapp"),
            note="感兴趣产品: " + g("感兴趣产品") if g("感兴趣产品") else "",
            src_file="AAOS展会客户信息表-修改.xlsx", src_sheet=ws.title, src_row=i)
    wb.close()


def parse_sofot(rows):
    """SOFCOT 2025 Exhibitor list — L1（无邮箱，仅官网展位）"""
    p = os.path.join(SRC_ROOT, "03-SOFCOT", "SOFCOT 2025 Exhibitor list.xlsx")
    if not os.path.exists(p):
        return
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [clean(v) for v in row]
        if not any(vals):
            continue
        if hdr is None:
            if "公司名称" in vals:
                hdr = {h: j for j, h in enumerate(vals) if h}
            continue

        def g(key):
            j = hdr.get(key)
            return vals[j] if (j is not None and j < len(vals)) else ""

        add(rows, source="SOFCOT 2025", tier="L1", kind="exhibitor",
            company=g("公司名称"), category=g("主营产品"),
            website=g("网站"), booth=g("展台"),
            src_file="SOFCOT 2025 Exhibitor list.xlsx", src_sheet=ws.title, src_row=i)
    wb.close()


def parse_aossm(rows):
    """AOSSM 2025 Exhibitor List — L1

    坑记录：Sheet1 的 Company 列大量单元格公式无缓存值（data_only 读出来是 None），
    实测 136 行里只有 8 行有公司名。真正的完整名录在 Sheet2（105 家，Name + Booth），
    Sheet1 剩下 8 家带 Key Product / Website。因此以 Sheet2 为底，用 Sheet1 补齐明细。
    """
    p = os.path.join(SRC_ROOT, "04-AOSSM", "AOSSM 2025 Exhibitor List.xlsx")
    if not os.path.exists(p):
        return
    fn = "AOSSM 2025 Exhibitor List.xlsx"
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    wss = wb.worksheets

    # --- Sheet1：公司名 -> (原名, 产品, 官网) 明细补充表
    detail = {}
    hdr = None
    for i, row in enumerate(wss[0].iter_rows(values_only=True), 1):
        vals = [clean(v) for v in row]
        if not any(vals):
            continue
        if hdr is None:
            if "Company" in vals:
                hdr = {h: j for j, h in enumerate(vals) if h}
            continue

        def g(key, vals=vals, hdr=hdr):
            j = hdr.get(key)
            return vals[j] if (j is not None and j < len(vals)) else ""

        comp = g("Company")
        if comp:
            detail[norm_company(comp)] = (comp, g("Key Product"), g("Website"))

    # --- Sheet2：完整名录（Name + Booth）
    hdr2 = None
    for i, row in enumerate(wss[1].iter_rows(values_only=True), 1):
        vals = [clean(v) for v in row]
        if not any(vals):
            continue
        if hdr2 is None:
            if "Name" in vals:
                hdr2 = {h: j for j, h in enumerate(vals) if h}
            continue

        def g2(key, vals=vals, hdr=hdr2):
            j = hdr.get(key)
            return vals[j] if (j is not None and j < len(vals)) else ""

        comp = g2("Name")
        if not comp:
            continue
        _, prod, web = detail.get(norm_company(comp), (comp, "", ""))
        booth = g2("Booth").replace("Booth #", "").replace("Booth", "").strip()
        add(rows, source="AOSSM 2025", tier="L1", kind="exhibitor",
            company=comp, category=prod, website=web, booth=booth,
            src_file=fn, src_sheet="Sheet2", src_row=i)

    # --- Sheet1 里有明细但 Sheet2 漏掉的公司，补进来（保证不丢数据）
    have = {r["company_norm"] for r in rows if r["source"] == "AOSSM 2025"}
    for k, (raw, prod, web) in detail.items():
        if not k or k in have:
            continue
        add(rows, source="AOSSM 2025", tier="L1", kind="exhibitor",
            company=raw, category=prod, website=web,
            src_file=fn, src_sheet="Sheet1", src_row=0)
    wb.close()


def parse_omtec_attendees(rows):
    """OMTEC 2025 Attendee List — L2 个人参会者（表头在第 12 行）"""
    p = os.path.join(SRC_ROOT, "05-OMTEC", "Final_OMTEC_Attendee_List_6_25.xlsx")
    if not os.path.exists(p):
        return
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    for ws in wb.worksheets:
        hdr = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [clean(v) for v in row]
            if not any(vals):
                continue
            if hdr is None:
                if "First Name" in vals and "Email Address" in vals:
                    hdr = {h: j for j, h in enumerate(vals) if h}
                continue

            def g(key):
                j = hdr.get(key)
                return vals[j] if (j is not None and j < len(vals)) else ""

            person = " ".join(x for x in [g("First Name"), g("Last Name")] if x)
            kind = "person" if ws.title != "Registrant No Shows" else "person"
            add(rows, source="OMTEC 2025", tier="L2", kind=kind,
                company=g("Company Name"), person=person, title=g("Job Title"),
                country=g("Country (Work Address)"),
                city=g("City (Work Address)"),
                category=g("Registrant Type") if "Registrant Type" in hdr else ws.title,
                email=g("Email Address"),
                note="参会人类别: " + ws.title,
                src_file="Final_OMTEC_Attendee_List_6_25.xlsx",
                src_sheet=ws.title, src_row=i)
    wb.close()


def parse_omtec_suppliers(rows):
    """OMTEC Supplier list（自有供应商台账，含国内联系人电话）— L3"""
    p = os.path.join(SRC_ROOT, "05-OMTEC", "OMTEC Supplier list.xlsx")
    if not os.path.exists(p):
        return
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [clean(v) for v in row]
        if not any(vals):
            continue
        if hdr is None:
            if "公司名称" in vals:
                hdr = {h: j for j, h in enumerate(vals) if h}
            continue
        # 表有两个"邮箱"列，openpyxl 字典会覆盖，这里按位置取
        cols = list(hdr.keys())

        def g(key):
            j = hdr.get(key)
            return vals[j] if (j is not None and j < len(vals)) else ""

        # 国内联系人 / 电话 / 邮箱 / 国外联系人 / 邮箱(2)
        email_idx = [j for j, h in enumerate(cols) if h == "邮箱"]
        email1 = vals[hdr["邮箱"]] if "邮箱" in hdr and hdr["邮箱"] < len(vals) else ""
        email2 = ""
        if len(email_idx) > 1:
            j2 = hdr["邮箱"]
            # 第二个"邮箱"列位置：取表头里最后一个邮箱下标
            last = None
            for j, h in enumerate(vals):
                pass
            # 用原始行表头重新定位
        # 简化：把整行非空且含 @ 的值都收集
        mails = [v for v in vals if "@" in v]
        email1 = mails[0] if mails else ""
        email2 = mails[1] if len(mails) > 1 else ""

        add(rows, source="OMTEC 2025", tier="L3", kind="person",
            company=g("公司名称"), person=g("国内联系人") or g("国外联系人"),
            country=g("国家"), category=g("主营业务"), website=g("公司网站"),
            phone=g("电话"), email=email1,
            note=" / ".join(x for x in [("国外联系人:" + g("国外联系人")) if g("国外联系人") else "",
                                        ("备用邮箱:" + email2) if email2 else "",
                                        ("沟通方式:" + g("目前沟通方式")) if g("目前沟通方式") else ""] if x),
            src_file="OMTEC Supplier list.xlsx", src_sheet=ws.title, src_row=i)
    wb.close()


def parse_aahks(rows):
    """AAHKS 2025 Exhibitors list — L1"""
    fn = "AAHKS 2025 Exhibitors list - 2025.09.04.xlsx"
    if not os.path.exists(os.path.join(SRC_ROOT, "06-AAHKS", fn)):
        fn = "AAHKS 2025 Exhibitors list.xlsx"
    p = os.path.join(SRC_ROOT, "06-AAHKS", fn)
    if not os.path.exists(p):
        return
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    seen_sheets = []
    for ws in wb.worksheets:
        hdr = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [clean(v) for v in row]
            if not any(vals):
                continue
            if hdr is None:
                if "Company Name" in vals:
                    hdr = {h: j for j, h in enumerate(vals) if h}
                continue

            def g(key):
                j = hdr.get(key)
                return vals[j] if (j is not None and j < len(vals)) else ""

            add(rows, source="AAHKS 2025", tier="L1", kind="exhibitor",
                company=g("Company Name"), country=g("Country"), city=g("City"),
                category=g("Key product"), website=g("Website"), booth=g("Booth"),
                note=g("备注"), src_file=fn, src_sheet=ws.title, src_row=i)
        seen_sheets.append(ws.title)
    wb.close()


def parse_dkou(rows):
    """DKOU 2026 参展商清单 — L1（含联系方式）"""
    p = os.path.join(SRC_ROOT, "08-DKOU", "DKOU2026_参展商清单.xlsx")
    if not os.path.exists(p):
        return
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [clean(v) for v in row]
        if not any(vals):
            continue
        if hdr is None:
            if "公司名称（中文）" in vals or "公司名称（英文）" in vals:
                hdr = {h: j for j, h in enumerate(vals) if h}
            continue

        def g(key):
            j = hdr.get(key)
            return vals[j] if (j is not None and j < len(vals)) else ""

        contact = g("联系方式")
        email, phone = split_contact(contact)
        add(rows, source="DKOU 2026", tier="L1", kind="exhibitor",
            company=g("公司名称（英文）") or g("公司名称（中文）"),
            country=g("国家/地区"),
            category=" / ".join(x for x in [g("类别"), g("类型")] if x),
            website=g("官网"), booth=g("展位号"), email=email, phone=phone,
            note=" / ".join(x for x in [("中文名:" + g("公司名称（中文）")) if g("公司名称（中文）") else "",
                                        g("备注")] if x),
            src_file="DKOU2026_参展商清单.xlsx", src_sheet=ws.title, src_row=i)
    wb.close()


# ---------------------------------------------------------------- 主流程

def main():
    global SRC_ROOT
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", help="展会源目录（含各展会 Excel 的根目录）",
                    default=None)
    ap.add_argument("--include-l3", action="store_true",
                    help="同时索引 L3 私密数据（自有客户/供应商，仅本人使用）")
    ap.add_argument("--stats", action="store_true", help="只打印统计")
    args = ap.parse_args()

    if args.stats and not os.path.exists(DB_PATH):
        print("索引尚未构建，请先运行 build_index.py")
        return 1
    if not args.stats and openpyxl is None:
        print("本脚本解析展会 Excel 需要 openpyxl，当前环境未安装。")
        print("两种解决方式任选其一：")
        print("  1) 安装：  pip install openpyxl")
        print("  2) 若系统 Python 已装该库，用 py 启动：  py -3 scripts/build_index.py --stats")
        return 1

    if args.src:
        SRC_ROOT = os.path.abspath(args.src)
        if not os.path.isdir(SRC_ROOT):
            print(f"  源目录不存在：{SRC_ROOT}")
            return 1

    if args.stats:
        con = sqlite3.connect(DB_PATH)
        for row in con.execute("SELECT tier, source, COUNT(*) FROM contacts GROUP BY tier, source ORDER BY tier, source"):
            print(f"  {row[0]}  {row[1]:<14} {row[2]:>6}")
        total = con.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
        print(f"  合计 {total} 条")
        con.close()
        return 0

    rows = []
    if not os.path.isdir(SRC_ROOT):
        print(f"  未找到展会源目录：{SRC_ROOT}")
        print("  请任选其一：")
        print("    · python build_index.py --src <展会Excel目录>")
        print("    · 设置环境变量 ORTHO_EXPO_SRC=<展会Excel目录>")
        print("    · 把展会 Excel 放进 <技能目录>/sources/ 对应子文件夹")
        return 1
    parse_aaos(rows)
    parse_sofot(rows)
    parse_aossm(rows)
    parse_omtec_attendees(rows)
    parse_aahks(rows)
    parse_dkou(rows)
    if args.include_l3:
        parse_aaos_customers(rows)
        parse_omtec_suppliers(rows)

    # 去重：同一 (source, tier, 公司归一化, 人, 邮箱) 视为重复
    seen = set()
    uniq = []
    for r in rows:
        key = (r["source"], r["tier"], r["company_norm"], r["person"].lower(), r["email"].lower())
        if key in seen:
            continue
        seen.add(key)
        uniq.append(r)

    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    con = sqlite3.connect(DB_PATH)
    con.executescript(SCHEMA)
    con.executemany("""
        INSERT INTO contacts (source, tier, kind, company, company_norm, person, title,
                              country, city, category, website, booth, email, phone, note,
                              src_file, src_sheet, src_row)
        VALUES (:source, :tier, :kind, :company, :company_norm, :person, :title,
                :country, :city, :category, :website, :booth, :email, :phone, :note,
                :src_file, :src_sheet, :src_row)
    """, uniq)
    con.execute("INSERT INTO contacts_fts(contacts_fts) VALUES('rebuild')")
    built = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    src_hash = hashlib.sha256(
        "|".join(sorted({r["src_file"] for r in uniq})).encode("utf-8")
    ).hexdigest()[:16]
    con.executemany("INSERT OR REPLACE INTO meta(k,v) VALUES(?,?)", [
        ("built_at", built),
        ("total", str(len(uniq))),
        ("l3_included", "1" if args.include_l3 else "0"),
        ("src_files_hash", src_hash),
        ("src_root", SRC_ROOT),
    ])
    con.commit()

    print("=" * 72)
    print("索引构建完成")
    print("=" * 72)
    print(f"  {'级别':<4}{'展会':<14}{'条数':>7}{'有官网':>8}{'有邮箱':>8}{'有电话':>8}   覆盖")
    print("  " + "-" * 78)
    warn_any = False
    # 源表本身就没有联系方式列的展会（已人工核对，非解析问题）
    NO_CONTACT_COL = {"AOSSM 2025": "源表仅公司名+展位，无联系方式列"}
    for tier, src, n in con.execute(
            "SELECT tier, source, COUNT(*) FROM contacts GROUP BY tier, source ORDER BY tier, source"):
        web = con.execute(
            "SELECT COUNT(*) FROM contacts WHERE source=? AND tier=? AND website<>''", (src, tier)).fetchone()[0]
        mail = con.execute(
            "SELECT COUNT(*) FROM contacts WHERE source=? AND tier=? AND email<>''", (src, tier)).fetchone()[0]
        tel = con.execute(
            "SELECT COUNT(*) FROM contacts WHERE source=? AND tier=? AND phone<>''", (src, tier)).fetchone()[0]
        # 空值率自检：联系信息全空的占比过高说明解析器可能踩坑
        cov = (web + mail + tel) / (3 * n) if n else 0
        flag = ""
        if src in NO_CONTACT_COL:
            flag = f"  （{NO_CONTACT_COL[src]}）"
        elif cov < 0.15:
            flag = "  <== 警告：联系信息覆盖率过低，请核对源表"
            warn_any = True
        print(f"  {tier:<4}{src:<14}{n:>7}{web:>8}{mail:>8}{tel:>8}   {cov * 100:5.1f}%{flag}")
    print("  " + "-" * 68)
    print(f"  {'':4}{'合计':<14}{len(uniq):>7} 条（去重前 {len(rows)} 条）")
    print(f"\n  构建时间: {built}")
    print(f"  L3 私密数据: {'已包含' if args.include_l3 else '未包含（默认）'}")
    print(f"  数据库: {DB_PATH}")
    if warn_any:
        print("\n  [!] 存在覆盖率告警，建议先用 --debug-null 定位是哪几列空了，再入库。")
    con.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
