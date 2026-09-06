# -*- coding: utf-8 -*-
"""房源数据 交互式 HTML 看板生成器（通用版，2026-08-11）
用法：python build_house_dashboard.py <Excel路径> [项目名]
整合三层：可视化图表 + 数据质量校验 + 统计分析 + 决策模块（剩余货值/预售证/对比/告警/外部验证/总结）
特性：自动探测表头行、动态列映射、动态状态配色、多项目支持
"""
import json, re, math, sys
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(__file__).parent
OUT = BASE / "output"

SRC = Path(sys.argv[1]) if len(sys.argv) > 1 else OUT / "晟悦家园_房源数据_20260811_094830.xlsx"
PROJ_NAME = sys.argv[2] if len(sys.argv) > 2 else SRC.name.split("_")[0]
print(f"数据源: {SRC.name} | 项目: {PROJ_NAME}")

wb = load_workbook(SRC, read_only=True)

# ---------- 通用：房源明细读取（自动探测表头行 + 动态列映射） ----------
ws = wb["房源明细"]
rows_all = [r for r in ws.iter_rows(values_only=True)]

# 找表头行（含"房号"的那一行）
header_idx = None
for i, row in enumerate(rows_all[:5]):
    if any("房号" in str(v) for v in row if v):
        header_idx = i
        break
assert header_idx is not None, "未找到表头行（含'房号'）"
HDR = [str(v).strip() if v else "" for v in rows_all[header_idx]]

# 列名 → 标准字段 映射
COL_MAP = {
    "项目": "项目", "预售证": "预售证", "楼栋": "楼栋", "单元": "单元", "楼层": "楼层",
    "房号": "房号", "用途": "用途", "建筑面积(㎡)": "面积", "建筑面积": "面积",
    "套内面积(㎡)": "套内", "套内": "套内", "分摊面积(㎡)": "分摊", "分摊": "分摊",
    "备案总价(元)": "总价", "备案总价": "总价", "总价": "总价",
    "备案单价(元/㎡)": "单价", "备案单价": "单价", "单价": "单价",
    "状态": "状态", "合同编号": "合同", "来源": "来源",
}
IDX = {}
for ci, h in enumerate(HDR):
    if h in COL_MAP:
        IDX[COL_MAP[h]] = ci

def cell(row, key):
    i = IDX.get(key)
    if i is None or i >= len(row):
        return None
    return row[i]

houses = []
for row in rows_all[header_idx + 1:]:
    if row[0] is None or all(v is None for v in row):
        continue
    h = {
        "项目": str(cell(row, "项目")) if cell(row, "项目") is not None else PROJ_NAME,
        "预售证": str(cell(row, "预售证")) if cell(row, "预售证") is not None else "",
        "楼栋": str(cell(row, "楼栋")) if cell(row, "楼栋") is not None else "",
        "单元": str(cell(row, "单元")) if cell(row, "单元") is not None else "",
        "楼层": cell(row, "楼层"),
        "房号": str(cell(row, "房号")) if cell(row, "房号") is not None else "",
        "用途": str(cell(row, "用途")) if cell(row, "用途") is not None else "",
        "面积": float(cell(row, "面积")) if isinstance(cell(row, "面积"), (int, float)) else None,
        "套内": float(cell(row, "套内")) if isinstance(cell(row, "套内"), (int, float)) else None,
        "分摊": float(cell(row, "分摊")) if isinstance(cell(row, "分摊"), (int, float)) else None,
        "总价": float(cell(row, "总价")) if isinstance(cell(row, "总价"), (int, float)) else None,
        "单价": float(cell(row, "单价")) if isinstance(cell(row, "单价"), (int, float)) else None,
        "状态": str(cell(row, "状态")) if cell(row, "状态") is not None else "",
        "合同": str(cell(row, "合同")) if cell(row, "合同") else "",
    }
    # 楼层 str→int
    fl = h["楼层"]
    if isinstance(fl, str) and fl.strip().isdigit():
        h["楼层"] = int(fl)
    houses.append(h)
print(f"房源: {len(houses)} 套 | 项目数: {len(set(h['项目'] for h in houses))}")

# ---------- 数据质量校验 ----------
quality = {}
keys = [f"{h['楼栋']}|{h['单元']}|{h['楼层']}|{h['房号']}" for h in houses]
quality["重复房源"] = f"{len(keys) - len(set(keys))} 套"
fields = ["面积", "套内", "分摊", "总价", "单价", "状态"]
null_rates = {}
for f in fields:
    n = sum(1 for h in houses if h[f] in (None, "", 0))
    null_rates[f] = f"{n}/{len(houses)} ({n/len(houses)*100:.1f}%)"
areas = [h["面积"] for h in houses if h["面积"]]
med = sorted(areas)[len(areas)//2] if areas else 0
area_abn = [h for h in houses if h["面积"] and (h["面积"] > med*2 or h["面积"] < med*0.5)]
quality["面积异常"] = f"{len(area_abn)} 套 (中位数 {med:.1f}㎡)"
priced = [h for h in houses if h["单价"] and h["单价"] > 0]
if priced:
    vals = [h["单价"] for h in priced]
    mu = sum(vals)/len(vals)
    sd = math.sqrt(sum((v-mu)**2 for v in vals)/len(vals)) or 1
    price_abn = [h for h in priced if abs(h["单价"]-mu) > 3*sd]
    quality["单价异常(z>3)"] = f"{len(price_abn)} 套 / 可售{len(priced)}套"
    quality["可售均价"] = f"{mu:.0f} 元/㎡ (σ={sd:.0f})"
mismatch = [h for h in priced if h["面积"] and h["单价"] and h["总价"]
            and abs(h["总价"] - h["单价"]*h["面积"]) / (h["单价"]*h["面积"]) > 0.01]
quality["总价单价一致性"] = f"{len(mismatch)} 套偏差>1%"

# ---------- 统计分析：楼层-价格梯度 ----------
buildings = sorted(set(h["楼栋"] for h in houses if h["楼栋"]))
gradients = []
for b in buildings:
    pts = [(h["楼层"], h["单价"]) for h in houses if h["楼栋"] == b and h["单价"] and h["单价"] > 0 and isinstance(h["楼层"], int)]
    if len(pts) >= 4:
        xs = [p[0] for p in pts]; ys = [p[1] for p in pts]
        n = len(xs); mx = sum(xs)/n; my = sum(ys)/n
        sxy = sum((x-mx)*(y-my) for x, y in pts)
        sxx = sum((x-mx)**2 for x in xs)
        slope = sxy/sxx if sxx else 0
        denom = sxx * sum((y-my)**2 for y in ys)
        r2 = (sxy**2)/denom if denom else 0
        gradients.append({"楼栋": b, "n": n, "斜率": round(slope, 1), "R²": round(r2, 3)})
from collections import Counter
status_cnt = Counter(h["状态"] for h in houses if h["状态"])
bld_status = {}
for b in buildings:
    bld_status[b] = dict(Counter(h["状态"] for h in houses if h["楼栋"] == b and h["状态"]))

# ---------- 决策模块（通用解析，各 Sheet 存在才读） ----------
DEC = {"permits": [], "value": {}, "compare": {}, "alerts": [], "ext": {"对比": [], "结论": []}, "summary": []}
SHEETS = wb.sheetnames

if "预售许可证清单" in SHEETS:
    ws = wb["预售许可证清单"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        k = str(row[0]).strip()
        if k in ("序号",) or k.startswith("晟悦") or k.startswith("深润川"):
            continue
        if len(row) >= 2 and row[1] and str(row[1]) not in ("—", "None"):
            DEC["permits"].append({
                "证号": str(row[1]), "项目": str(row[2]) if len(row) > 2 and row[2] else "",
                "用途": str(row[5]) if len(row) > 5 and row[5] else "",
                "面积": str(row[7]) if len(row) > 7 and row[7] else "",
                "日期": str(row[8]) if len(row) > 8 and row[8] else "",
                "备注": str(row[11]) if len(row) > 11 and row[11] else "",
            })
    # 去重
    seen = set(); dd = []
    for p in DEC["permits"]:
        if p["证号"] not in seen:
            seen.add(p["证号"]); dd.append(p)
    DEC["permits"] = dd

if "剩余货值测算" in SHEETS:
    ws = wb["剩余货值测算"]
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        k = str(row[0]).strip()
        v = row[5] if len(row) > 5 and row[5] is not None else (row[4] if len(row) > 4 and row[4] is not None else None)
        if "剩余房源合计套数" in k: DEC["value"]["剩余套数"] = v
        elif "可售(期房待售)套数" in k: DEC["value"]["可售套数"] = v
        elif "签约中(已录入合同)套数" in k: DEC["value"]["签约套数"] = v
        elif "合计可售面积" in k: DEC["value"]["可售面积"] = v
        elif "合计剩余面积" in k: DEC["value"]["剩余面积"] = v
        elif "合计可售挂牌货值" in k: DEC["value"]["可售货值"] = v
        elif "合计签约中测算货值" in k: DEC["value"]["签约货值"] = v
        elif "总剩余货值·主口径" in k: DEC["value"]["主口径货值"] = v
        elif "总剩余货值·参考口径" in k: DEC["value"]["参考口径货值"] = v
        elif "主口径基准来源" in k: DEC["value"]["基准来源"] = v
        elif "主口径基准明细" in k: DEC["value"]["基准明细"] = str(v)[:120] if v else ""
    # fallback：表格式结构（深润川：行=项目/证，列=主口径货值(万元)等）
    if not DEC["value"].get("主口径货值"):
        ws = wb["剩余货值测算"]
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            # 表头行：找含"主口径货值"的列
            if any("主口径货值" in str(c) for c in row if c):
                hdr = [str(c).strip() if c else "" for c in row]
                try:
                    i_area = hdr.index("可售面积(㎡)")
                    i_val = hdr.index("主口径货值(万元)")
                    i_total = hdr.index("总套数")
                    i_first = hdr.index("首次登记(可售存量)")
                    i_sold = hdr.index("已备案(已售)")
                except ValueError:
                    continue
                for r2 in ws.iter_rows(values_only=True):
                    if not r2 or not r2[0] or str(r2[0]).strip() == "合计":
                        if r2 and r2[0] and str(r2[0]).strip() == "合计" and len(r2) > max(i_val, i_total):
                            DEC["value"]["剩余套数"] = r2[i_total] if i_total < len(r2) else None
                            DEC["value"]["可售套数"] = r2[i_first] if i_first < len(r2) else None
                            DEC["value"]["已售套数"] = r2[i_sold] if i_sold < len(r2) else None
                            DEC["value"]["可售面积"] = r2[i_area] if i_area < len(r2) else None
                            # 万元→元
                            if i_val < len(r2) and isinstance(r2[i_val], (int, float)):
                                DEC["value"]["主口径货值"] = r2[i_val] * 10000
                        continue
                break
        if DEC["value"].get("主口径货值"):
            DEC["value"]["基准来源"] = "证级备案均价（表格式口径）"

if "数据变化对比" in SHEETS:
    ws = wb["数据变化对比"]
    cur = None
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        k = str(row[0]).strip()
        if "日度对比" in k: cur = "日度"; continue
        if "周度对比" in k: cur = "周度"; continue
        if "月度对比" in k or "季度对比" in k: cur = None; continue
        if cur and k not in ("指标",) and len(row) > 1 and row[1] is not None:
            DEC["compare"].setdefault(cur, {})[k] = {"当前": row[1], "上期": row[2] if len(row) > 2 else None,
                "变化": row[3] if len(row) > 3 else None, "变化率": row[4] if len(row) > 4 else None}

if "异常与告警" in SHEETS:
    ws = wb["异常与告警"]
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        k = str(row[0]).strip()
        if k in ("级别", "指标", "一、价格缺失统计", "二、数据突变检测（对比上一快照）"):
            continue
        DEC["alerts"].append({"级别": k, "检测项": str(row[1]),
            "当前": str(row[2]) if len(row) > 2 and row[2] else "",
            "对比": str(row[3]) if len(row) > 3 and row[3] else "",
            "说明": str(row[5]) if len(row) > 5 and row[5] else ""})

if "外部渠道价格交叉验证" in SHEETS:
    ws = wb["外部渠道价格交叉验证"]
    section = None
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        k = str(row[0]).strip()
        if "价格口径对比" in k: section = "对比"; continue
        if "交叉验证结论" in k: section = "结论"; continue
        if "采集明细" in k: section = None; continue
        if section == "对比" and k not in ("口径",):
            DEC["ext"]["对比"].append({"口径": k, "备案价": str(row[1]) if len(row) > 1 else "",
                "市场价": str(row[2]) if len(row) > 2 else "", "折让": str(row[3]) if len(row) > 3 else ""})
        elif section == "结论" and k:
            t = " ".join(str(x) for x in row if x).strip()
            if t and not t.startswith(("四、", "1.", "2.", "3.", "4.", "5.")):
                DEC["ext"]["结论"].append(t[:140])

if "分析总结" in SHEETS:
    ws = wb["分析总结"]
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        k = str(row[0]).strip()
        if k.startswith(("1.", "2.", "3.", "4.", "5.", "6.")) and "整体概况" not in k:
            DEC["summary"].append(k)
    DEC["summary"] = DEC["summary"][:8]

print(f"决策模块: permits={len(DEC['permits'])} value_主口径={DEC['value'].get('主口径货值')} alerts={len(DEC['alerts'])} ext对比={len(DEC['ext']['对比'])}")

# ---------- JSON ----------
data_json = json.dumps(houses, ensure_ascii=False)
bld_status_json = json.dumps(bld_status, ensure_ascii=False)
status_cnt_json = json.dumps(status_cnt, ensure_ascii=False)
grad_json = json.dumps(gradients, ensure_ascii=False)
quality_json = json.dumps(quality, ensure_ascii=False)
null_json = json.dumps(null_rates, ensure_ascii=False)
permits_json = json.dumps(DEC["permits"], ensure_ascii=False)
value_json = json.dumps(DEC["value"], ensure_ascii=False)
compare_json = json.dumps(DEC["compare"], ensure_ascii=False)
alerts_json = json.dumps(DEC["alerts"], ensure_ascii=False)
ext_json = json.dumps(DEC["ext"], ensure_ascii=False)
summary_json = json.dumps(DEC["summary"], ensure_ascii=False)
# 动态状态配色（前3个固定，其余自动生成）
STATUS_COLORS = ["#E6F1FB", "#E1F5EE", "#FAEEDA", "#FBEAF0", "#F1EFE8", "#FAECE7", "#EEEDFE", "#FCEBEB"]
ST_CLS_JS = {}
for i, st in enumerate(sorted(status_cnt.keys())):
    c = STATUS_COLORS[i % len(STATUS_COLORS)]
    ST_CLS_JS[st] = c
st_cls_json = json.dumps(ST_CLS_JS, ensure_ascii=False)
# 项目列表（多项目时显示）
proj_list = sorted(set(h["项目"] for h in houses))
proj_list_json = json.dumps(proj_list, ensure_ascii=False)

html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{PROJ_NAME} · 房源数据交互看板</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.5.1"></script>
<style>
:root {{
    --bg: #f5f6f8; --card: #ffffff; --text: #1f2937; --muted: #6b7280;
    --pos: #1D9E75; --neg: #D85A30; --radius: 10px; --gap: 14px;
}}
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family: -apple-system,'Segoe UI','Microsoft YaHei',sans-serif; background:var(--bg); color:var(--text); line-height:1.5; }}
.wrap {{ max-width:1440px; margin:0 auto; padding:16px; }}
.header {{ background:#10345e; color:#fff; padding:18px 24px; border-radius:var(--radius); margin-bottom:var(--gap); display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:10px; }}
.header h1 {{ font-size:19px; font-weight:600; }}
.header .sub {{ font-size:12px; opacity:.75; }}
.filters {{ display:flex; gap:10px; align-items:center; flex-wrap:wrap; }}
.filters select {{ padding:6px 10px; border:1px solid #cbd5e1; border-radius:6px; background:#fff; color:var(--text); font-size:13px; }}
.filters label {{ font-size:12px; color:#e2e8f0; }}
.kpi-row {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(150px,1fr)); gap:var(--gap); margin-bottom:var(--gap); }}
.kpi {{ background:var(--card); border-radius:var(--radius); padding:14px 18px; box-shadow:0 1px 2px rgba(0,0,0,.06); }}
.kpi .l {{ font-size:12px; color:var(--muted); }}
.kpi .v {{ font-size:22px; font-weight:700; margin-top:2px; }}
.kpi .d {{ font-size:12px; color:var(--muted); margin-top:2px; }}
.grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(400px,1fr)); gap:var(--gap); margin-bottom:var(--gap); }}
.card {{ background:var(--card); border-radius:var(--radius); padding:16px 20px; box-shadow:0 1px 2px rgba(0,0,0,.06); }}
.card h3 {{ font-size:14px; font-weight:600; margin-bottom:12px; color:var(--text); }}
.card canvas {{ max-height:280px; }}
.qcard {{ background:var(--card); border-radius:var(--radius); padding:16px 20px; box-shadow:0 1px 2px rgba(0,0,0,.06); }}
.qcard h3 {{ font-size:14px; font-weight:600; margin-bottom:12px; }}
.qtable {{ width:100%; border-collapse:collapse; font-size:13px; }}
.qtable td, .qtable th {{ padding:7px 10px; border-bottom:1px solid #f0f1f3; text-align:left; }}
.qtable th {{ color:var(--muted); font-weight:500; font-size:12px; }}
.ok {{ color:var(--pos); font-weight:600; }} .warn {{ color:#854F0B; font-weight:600; }} .bad {{ color:var(--neg); font-weight:600; }}
.table-sec {{ background:var(--card); border-radius:var(--radius); padding:16px 20px; box-shadow:0 1px 2px rgba(0,0,0,.06); overflow-x:auto; }}
.table-sec h3 {{ font-size:14px; font-weight:600; margin-bottom:12px; }}
table.data {{ width:100%; border-collapse:collapse; font-size:12.5px; }}
table.data th {{ text-align:left; padding:8px 10px; border-bottom:2px solid #e2e8f0; color:var(--muted); font-weight:600; font-size:12px; white-space:nowrap; cursor:pointer; user-select:none; }}
table.data td {{ padding:7px 10px; border-bottom:1px solid #f1f2f4; white-space:nowrap; }}
table.data tr:hover td {{ background:#f8fafc; }}
.st-badge {{ display:inline-block; padding:2px 8px; border-radius:10px; font-size:11.5px; }}
.pager {{ display:flex; gap:10px; align-items:center; margin-top:10px; font-size:13px; color:var(--muted); }}
.pager button {{ padding:4px 12px; border:1px solid #cbd5e1; border-radius:6px; background:#fff; cursor:pointer; font-size:13px; }}
.footer {{ font-size:12px; color:var(--muted); text-align:center; padding:14px 0 6px; }}
@media(max-width:800px){{ .grid{{grid-template-columns:1fr;}} .kpi-row{{grid-template-columns:repeat(2,1fr);}} }}
</style>
</head>
<body>
<div class="wrap">
  <div class="header">
    <div>
      <h1>{PROJ_NAME} · 房源数据交互看板</h1>
      <div class="sub">数据来源：深圳房地产信息平台 · {datetime.now().strftime('%Y-%m-%d')} ｜ 已备案/已售备案价不公开（平台规则）｜ 备案价非成交价</div>
    </div>
    <div class="filters">
      <label>项目</label><select id="f-proj" onchange="dash.apply()"><option value="all">全部</option></select>
      <label>楼栋</label><select id="f-bld" onchange="dash.apply()"><option value="all">全部</option></select>
      <label>状态</label><select id="f-status" onchange="dash.apply()"><option value="all">全部</option></select>
      <label>用途</label><select id="f-use" onchange="dash.apply()"><option value="all">全部</option></select>
    </div>
  </div>

  <div class="kpi-row" id="kpi-row"></div>
  <div class="kpi-row" id="value-row"></div>

  <div class="grid">
    <div class="card"><h3>房源状态分布</h3><canvas id="c-status"></canvas></div>
    <div class="card"><h3>楼栋 × 状态</h3><canvas id="c-bld"></canvas></div>
  </div>
  <div class="grid">
    <div class="card"><h3>可售备案单价分布</h3><canvas id="c-price"></canvas></div>
    <div class="card"><h3>面积分布</h3><canvas id="c-area"></canvas></div>
  </div>
  <div class="grid">
    <div class="card"><h3>楼栋平均备案单价（可售）</h3><canvas id="c-bldprice"></canvas></div>
    <div class="card"><h3>楼层-单价关系（可售）</h3><canvas id="c-floor"></canvas></div>
  </div>
  <div class="grid">
    <div class="qcard"><h3>数据质量校验</h3><table class="qtable" id="q-table"></table></div>
    <div class="qcard"><h3>楼层-价格梯度分析（线性回归）</h3><table class="qtable" id="g-table"></table><div style="font-size:12px;color:var(--muted);margin-top:8px;">斜率=每层价格变化(元/㎡)，R²=拟合优度；样本≥4层楼栋</div></div>
  </div>

  <div class="grid">
    <div class="qcard"><h3>预售许可证清单</h3><table class="qtable" id="p-table"></table></div>
    <div class="qcard"><h3>异常与告警</h3><div id="a-box" style="font-size:13px;"></div></div>
  </div>
  <div class="grid">
    <div class="qcard"><h3>数据变化对比（周度）</h3><table class="qtable" id="c-table"></table><div style="font-size:12px;color:var(--muted);margin-top:8px;">周度 = 当前 vs 上一周快照</div></div>
    <div class="qcard"><h3>外部渠道价格交叉验证</h3><table class="qtable" id="e-table"></table><div style="font-size:12px;color:var(--muted);margin-top:8px;" id="e-concl"></div></div>
  </div>
  <div class="grid">
    <div class="qcard"><h3>分析总结</h3><div id="s-box" style="font-size:13px;line-height:1.8;"></div></div>
  </div>

  <div class="table-sec">
    <h3>房源明细（点击表头排序）</h3>
    <table class="data" id="t-main"><thead></thead><tbody></tbody></table>
    <div class="pager"><button onclick="dash.prev()">‹ 上一页</button><span id="pg-info"></span><button onclick="dash.next()">下一页 ›</button></div>
  </div>

  <div class="footer">{PROJ_NAME} 房源数据看板 ｜ 生成于 {datetime.now().strftime('%Y-%m-%d %H:%M')} ｜ 备案价为政府备案口径非成交价 ｜ 来源标注：[深圳房地产信息平台 {datetime.now().strftime('%Y-%m-%d')}]</div>
</div>

<script>
const RAW = {data_json};
const BLD_STATUS = {bld_status_json};
const STATUS_CNT = {status_cnt_json};
const GRADS = {grad_json};
const QUALITY = {quality_json};
const NULL_RATES = {null_json};
const PERMITS = {permits_json};
const VALUE = {value_json};
const COMPARE = {compare_json};
const ALERTS = {alerts_json};
const EXT = {ext_json};
const SUMMARY = {summary_json};
const ST_COLORS = {st_cls_json};
const PROJS = {proj_list_json};
const C = ['#185FA5','#1D9E75','#D85A30','#854F0B','#7F77DD','#D4537E','#378ADD','#97C459'];

function fmt(v, dec=0) {{ return Number(v).toLocaleString('zh-CN', {{maximumFractionDigits: dec}}); }}
function fmtMoney(v) {{ if(!v) return '—'; return '¥' + fmt(v); }}

class Dashboard {{
  constructor() {{
    this.raw = RAW; this.f = {{proj:'all', bld:'all', status:'all', use:'all'}};
    this.page = 0; this.pageSize = 100; this.sort = null; this.sortDir = 'asc';
    this.populateFilters();
    this.initCharts();
    this.apply();
  }}
  populateFilters() {{
    this.fill('f-proj', PROJS);
    this.fill('f-bld', [...new Set(this.raw.map(h=>h['楼栋']))].sort());
    this.fill('f-status', Object.keys(STATUS_CNT));
    this.fill('f-use', [...new Set(this.raw.map(h=>h['用途']))]);
  }}
  fill(id, vals) {{ const s = document.getElementById(id); vals.forEach(v=>{{ const o=document.createElement('option'); o.value=v; o.textContent=v; s.appendChild(o); }}); }}
  get filtered() {{
    return this.raw.filter(h =>
      (this.f.proj==='all'||h['项目']===this.f.proj) &&
      (this.f.bld==='all'||h['楼栋']===this.f.bld) &&
      (this.f.status==='all'||h['状态']===this.f.status) &&
      (this.f.use==='all'||h['用途']===this.f.use));
  }}
  apply() {{
    this.f.proj = document.getElementById('f-proj').value;
    this.f.bld = document.getElementById('f-bld').value;
    this.f.status = document.getElementById('f-status').value;
    this.f.use = document.getElementById('f-use').value;
    this.page = 0;
    this.renderKPI(); this.updateAll(); this.renderTable();
  }}
  renderKPI() {{
    const d = this.filtered;
    const priced = d.filter(h=>h['单价']&&h['单价']>0);
    const areas = d.filter(h=>h['面积']).map(h=>h['面积']);
    const avgP = priced.length ? priced.reduce((a,h)=>a+h['单价'],0)/priced.length : 0;
    const avgA = areas.length ? areas.reduce((a,b)=>a+b,0)/areas.length : 0;
    const st = {{}}; d.forEach(h=>st[h['状态']]=(st[h['状态']]||0)+1);
    const top3 = Object.entries(st).sort((a,b)=>b[1]-a[1]).slice(0,3);
    const kpis = [
      ['总房源', d.length+' 套', ''],
      ...top3.map(([k,v])=>[k, v+' 套', ((v/d.length)*100).toFixed(1)+'%']),
      ['可售备案均价', fmt(avgP)+' 元/㎡', priced.length+' 套样本'],
      ['面积均值', fmt(avgA,1)+' ㎡', areas.length+' 套'],
    ];
    document.getElementById('kpi-row').innerHTML = kpis.map(k=>
      `<div class="kpi"><div class="l">${{k[0]}}</div><div class="v">${{k[1]}}</div><div class="d">${{k[2]}}</div></div>`).join('');
    const V = VALUE;
    const yiji = n => n>=1e8 ? (n/1e8).toFixed(2)+' 亿' : (n/1e4).toFixed(0)+' 万';
    const cards = [
      ['剩余房源', (V['剩余套数']||0)+' 套', '可售'+(V['可售套数']||0)+' + 签约'+(V['签约套数']||0)],
      ['合计剩余面积', V['剩余面积']?fmt(V['剩余面积'])+' ㎡':'—', ''],
      ['剩余货值 · 主口径', V['主口径货值']?yiji(V['主口径货值'])+' 元':'—', V['基准来源']||''],
      ['剩余货值 · 参考口径', V['参考口径货值']?yiji(V['参考口径货值'])+' 元':'—', '可售+签约'],
      ['可售挂牌货值', V['可售货值']?yiji(V['可售货值'])+' 元':'—', ''],
    ];
    document.getElementById('value-row').innerHTML = cards.map(k=>
      `<div class="kpi"><div class="l">${{k[0]}}</div><div class="v">${{k[1]}}</div><div class="d">${{k[2]}}</div></div>`).join('');
  }}
  initCharts() {{
    this.statusChart = new Chart(document.getElementById('c-status'), {{type:'doughnut', data:{{labels:[],datasets:[{{data:[],backgroundColor:C.map(c=>c+'CC'),borderColor:'#fff',borderWidth:2}}]}}, options:{{responsive:true,maintainAspectRatio:false,cutout:'58%',plugins:{{legend:{{position:'right'}}}}}}}});
    this.bldChart = new Chart(document.getElementById('c-bld'), {{type:'bar', data:{{labels:[],datasets:[]}}, options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'top'}}}},scales:{{x:{{stacked:true}},y:{{stacked:true}}}}}}}});
    this.priceChart = new Chart(document.getElementById('c-price'), {{type:'bar', data:{{labels:[],datasets:[{{data:[],backgroundColor:'#B5D4F4',borderColor:'#185FA5',borderWidth:1}}]}}, options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:{{y:{{beginAtZero:true}}}}}}}});
    this.areaChart = new Chart(document.getElementById('c-area'), {{type:'bar', data:{{labels:[],datasets:[{{data:[],backgroundColor:'#9FE1CB',borderColor:'#1D9E75',borderWidth:1}}]}}, options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:{{y:{{beginAtZero:true}}}}}}}});
    this.bldPriceChart = new Chart(document.getElementById('c-bldprice'), {{type:'bar', data:{{labels:[],datasets:[{{data:[],backgroundColor:C,borderWidth:1}}]}}, options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:{{y:{{beginAtZero:true,ticks:{{callback:v=>fmt(v)}}}}}}}}}});
    this.floorChart = new Chart(document.getElementById('c-floor'), {{type:'scatter', data:{{datasets:[]}}, options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'top'}}}},scales:{{x:{{title:{{display:true,text:'楼层'}}}},y:{{title:{{display:true,text:'备案单价(元/㎡)'}},ticks:{{callback:v=>fmt(v)}}}}}}}}}});
    this.renderQuality(); this.renderGrad(); this.renderDecision();
  }}
  updateAll() {{
    const d = this.filtered;
    const st = {{}}; d.forEach(h=>st[h['状态']]=(st[h['状态']]||0)+1);
    const stKeys = Object.keys(st);
    this.statusChart.data.labels = stKeys;
    this.statusChart.data.datasets[0].data = stKeys.map(k=>st[k]);
    this.statusChart.data.datasets[0].backgroundColor = stKeys.map(k=>ST_COLORS[k]||'#ddd');
    this.statusChart.update('none');
    const blds = [...new Set(d.map(h=>h['楼栋']))].sort();
    const allSt = Object.keys(STATUS_CNT);
    this.bldChart.data.labels = blds;
    this.bldChart.data.datasets = allSt.map((s,i)=>({{label:s,data:blds.map(b=>{{const m=BLD_STATUS[b]||{{}};return m[s]||0;}}),backgroundColor:C[i%8]+'CC'}}));
    this.bldChart.update('none');
    const priced = d.filter(h=>h['单价']&&h['单价']>0);
    if (priced.length) {{
      const pmin = Math.min(...priced.map(h=>h['单价'])), pmax = Math.max(...priced.map(h=>h['单价']));
      const bins = 10, binW = (pmax-pmin)/bins||1;
      const hist = Array(bins).fill(0), hlabels = [];
      priced.forEach(h=>{{const i=Math.min(bins-1,Math.floor((h['单价']-pmin)/binW));hist[i]++;}});
      for(let i=0;i<bins;i++) hlabels.push(fmt(pmin+i*binW)+'~'+fmt(pmin+(i+1)*binW));
      this.priceChart.data.labels = hlabels;
      this.priceChart.data.datasets[0].data = hist;
      this.priceChart.update('none');
    }}
    const areas = d.filter(h=>h['面积']).map(h=>h['面积']);
    if (areas.length) {{
      const amin = Math.floor(Math.min(...areas)), amax = Math.ceil(Math.max(...areas));
      const abins = Math.min(10, Math.max(4, Math.ceil((amax-amin)/4)));
      const aw = (amax-amin)/abins||1;
      const ahist = Array(abins).fill(0), alabels=[];
      areas.forEach(a=>{{const i=Math.min(abins-1,Math.floor((a-amin)/aw));ahist[i]++;}});
      for(let i=0;i<abins;i++) alabels.push(fmt(amin+i*aw)+'~'+fmt(amin+(i+1)*aw));
      this.areaChart.data.labels = alabels;
      this.areaChart.data.datasets[0].data = ahist;
      this.areaChart.update('none');
    }}
    const bp = blds.map(b=>{{const ps=priced.filter(h=>h['楼栋']===b);return ps.length?ps.reduce((a,h)=>a+h['单价'],0)/ps.length:0;}});
    this.bldPriceChart.data.labels = blds;
    this.bldPriceChart.data.datasets[0].data = bp;
    this.bldPriceChart.update('none');
    this.floorChart.data.datasets = blds.map((b,i)=>({{
      label:b, data:priced.filter(h=>h['楼栋']===b&&typeof h['楼层']==='number').map(h=>({{x:h['楼层'],y:h['单价']}})),
      backgroundColor:C[i%8]+'88', pointRadius:4
    }})).filter(ds=>ds.data.length>0);
    this.floorChart.update('none');
  }}
  renderQuality() {{
    const rows = [['重复房源','≤0','ok'],['面积空值',NULL_RATES['面积']||'—','warn'],['套内空值',NULL_RATES['套内']||'—','warn'],['分摊空值',NULL_RATES['分摊']||'—','warn'],['总价空值',NULL_RATES['总价']||'—','warn'],['单价空值',NULL_RATES['单价']||'—','warn'],['面积异常(>2x或<0.5x中位数)',QUALITY['面积异常']||'—','warn'],['单价异常(z>3)',QUALITY['单价异常(z>3)']||'—','warn'],['总价单价一致性(偏差>1%)',QUALITY['总价单价一致性']||'—','ok'],['可售均价',QUALITY['可售均价']||'—','ok']];
    document.getElementById('q-table').innerHTML = rows.map(r=>`<tr><td>${{r[0]}}</td><td class="${{r[2]}}">${{r[1]}}</td></tr>`).join('');
  }}
  renderGrad() {{
    const rows = GRADS.map(g=>[g['楼栋'], g['n']+'层', g['斜率']>0?'+':'', g['斜率'], g['R2']]);
    document.getElementById('g-table').innerHTML =
      '<tr><th>楼栋</th><th>样本</th><th colspan=2>斜率(元/㎡/层)</th><th>R²</th></tr>' +
      rows.map(r=>`<tr><td>${{r[0]}}</td><td>${{r[1]}}</td><td>${{r[2]}}</td><td>${{fmt(r[3],1)}}</td><td>${{r[4]}}</td></tr>`).join('');
  }}
  renderDecision() {{
    document.getElementById('p-table').innerHTML = '<tr><th>预售证号</th><th>用途</th><th>面积(㎡)</th><th>日期</th><th>备注</th></tr>' +
      PERMITS.map(p=>`<tr><td>${{p['证号']}}</td><td>${{p['用途']}}</td><td>${{p['面积']}}</td><td>${{p['日期']}}</td><td>${{(p['备注']||'').slice(0,60)}}</td></tr>`).join('');
    document.getElementById('a-box').innerHTML = ALERTS.length ? ALERTS.map(a=>
      `<div style="margin-bottom:8px;"><span class="warn">[${{a['级别']}}]</span> ${{a['检测项']}}：当前 ${{a['当前']}} 对比 ${{a['对比']}}<br><span style="color:var(--muted);font-size:12px;">${{a['说明']}}</span></div>`).join('')
      : '<div class="ok">暂无异常</div>';
    const wk = COMPARE['周度']||{{}};
    const cRows = Object.keys(wk).map(k=>{{
      const r = wk[k];
      const chg = r['变化']==null?0:r['变化'];
      const cls = chg>0?'warn':(chg<0?'ok':'');
      return `<tr><td>${{k}}</td><td>${{r['当前']}}</td><td>${{r['上期']}}</td><td class="${{cls}}">${{chg>0?'+':''}}${{chg}}</td><td>${{r['变化率']}}%</td></tr>`;
    }});
    document.getElementById('c-table').innerHTML = '<tr><th>指标</th><th>当前</th><th>上期</th><th>变化</th><th>变化率</th></tr>' + cRows.join('');
    const eRows = (EXT['对比']||[]).map(r=>`<tr><td>${{r['口径']}}</td><td>${{r['备案价']}}</td><td>${{r['市场价']}}</td><td>${{r['折让']}}</td></tr>`).join('');
    document.getElementById('e-table').innerHTML = '<tr><th>口径</th><th>政府备案价</th><th>外网市场价</th><th>折让幅度</th></tr>' + eRows;
    document.getElementById('e-concl').textContent = (EXT['结论']||[]).slice(0,3).join('；');
    document.getElementById('s-box').innerHTML = SUMMARY.map(s=>`<div>• ${{s}}</div>`).join('');
  }}
  renderTable() {{
    let d = [...this.filtered];
    if (this.sort) {{
      d.sort((a,b)=>{{const av=a[this.sort], bv=b[this.sort]; if(av==null)return 1; if(bv==null)return -1;
        return this.sortDir==='asc'?(av<bv?-1:av>bv?1:0):(av<bv?1:av>bv?-1:0);}});
    }}
    const cols = ['项目','预售证','楼栋','单元','楼层','房号','用途','面积','套内','分摊','总价','单价','状态','合同'];
    const total = d.length, pages = Math.max(1, Math.ceil(total/this.pageSize));
    this.page = Math.min(this.page, pages-1);
    const start = this.page*this.pageSize, end = Math.min(start+this.pageSize, total);
    const pageData = d.slice(start, end);
    const ths = cols.map(c=>`<th onclick="dash.sortBy('${{c}}')">${{c}}${{this.sort===c?(this.sortDir==='asc'?' ▲':' ▼'):''}}</th>`).join('');
    document.querySelector('#t-main thead').innerHTML = `<tr>${{ths}}</tr>`;
    document.querySelector('#t-main tbody').innerHTML = pageData.map(h=>
      `<tr><td>${{h['项目']}}</td><td>${{(h['预售证']||'').slice(0,14)}}</td><td>${{h['楼栋']}}</td><td>${{h['单元']}}</td><td>${{h['楼层']}}</td><td>${{h['房号']}}</td><td>${{h['用途']}}</td><td>${{h['面积']?h['面积'].toFixed(2):'—'}}</td><td>${{h['套内']?h['套内'].toFixed(2):'—'}}</td><td>${{h['分摊']?h['分摊'].toFixed(2):'—'}}</td><td>${{fmtMoney(h['总价'])}}</td><td>${{h['单价']?fmt(h['单价'])+' 元/㎡':'—'}}</td><td><span class="st-badge" style="background:${{ST_COLORS[h['状态']]||'#eee'}};color:#333;">${{h['状态']}}</span></td><td>${{h['合同']}}</td></tr>`).join('');
    document.getElementById('pg-info').textContent = `第 ${{this.page+1}}/${{pages}} 页 · 共 ${{total}} 套 (显示 ${{start+1}}-${{end}})`;
  }}
  sortBy(c) {{ if(this.sort===c) this.sortDir=this.sortDir==='asc'?'desc':'asc'; else {{this.sort=c;this.sortDir='asc';}} this.renderTable(); }}
  prev() {{ if(this.page>0){{this.page--;this.renderTable();}} }}
  next() {{ const pages=Math.max(1,Math.ceil(this.filtered.length/this.pageSize)); if(this.page<pages-1){{this.page++;this.renderTable();}} }}
}}
const dash = new Dashboard();
</script>
</body>
</html>"""

out = OUT / f"{PROJ_NAME}_交互看板_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
out.write_text(html, encoding="utf-8")
print(f"\n已生成: {out.name} ({out.stat().st_size/1024:.0f} KB)")
print(f"状态: {dict(status_cnt)}")
print(f"质量: {json.dumps(quality, ensure_ascii=False)}")
print(f"梯度: {json.dumps(gradients, ensure_ascii=False)[:300]}")
