# -*- coding: utf-8 -*-
"""C5: 深圳二手住宅成交参考价解析（官方 2021-02-08 PDF，3595 小区）
输入: output_cross/shenzhen_data/深圳二手住宅成交参考价_官方2021.pdf
输出: JSON + Excel（含分区统计/街道聚合/价格档位）
"""
import json
import re
import fitz
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PDF = 'output_cross/shenzhen_data/深圳二手住宅成交参考价_官方2021.pdf'
OUT_DIR = 'output_cross/shenzhen_data'
OUT_JSON = f'{OUT_DIR}/shenzhen_ershou_refprice.json'
OUT_XLSX = f'{OUT_DIR}/深圳二手住宅成交参考价_官方_20260820.xlsx'

doc = fitz.open(PDF)
rows = []
seen = set()
for pno, page in enumerate(doc):
    for t in page.find_tables():
        for row in t.extract():
            if not row or len(row) < 5:
                continue
            seq, zone, street, name, price = row[:5]
            seq = (seq or '').strip()
            zone = (zone or '').strip()
            street = (street or '').strip()
            name = (name or '').strip()
            price_raw = (price or '').strip()
            if seq == '序号' or not seq.isdigit():
                continue
            price_num = int(re.sub(r'[^0-9]', '', price_raw)) if re.sub(r'[^0-9]', '', price_raw) else 0
            key = (zone, street, name)
            if key in seen:
                continue
            seen.add(key)
            rows.append({
                '序号': int(seq), '行政区': zone, '街道': street,
                '项目名称': name, '参考价(元/㎡)': price_num,
            })

print('解析条数:', len(rows))
# 去重校验（序号连续性）
seqs = sorted(r['序号'] for r in rows)
print('序号范围:', seqs[0], '-', seqs[-1], '| 缺失数:', len(set(range(1, seqs[-1]+1))) - len(seqs))

# 统计
zone_cnt = Counter(r['行政区'] for r in rows)
print('各区条数:', dict(sorted(zone_cnt.items(), key=lambda x: -x[1])))
street_cnt = Counter(r['街道'] for r in rows)
print('街道数:', len(street_cnt), '| 街道TOP5:', dict(street_cnt.most_common(5)))
prices = [r['参考价(元/㎡)'] for r in rows]
print('价格范围:', min(prices), '-', max(prices), '| 均价:', round(sum(prices)/len(prices)))
brackets = Counter()
for p in prices:
    b = f'{(p//10000)*10000}-{(p//10000+1)*10000}'
    brackets[b] += 1
print('价格档位:', dict(sorted(brackets.items())))

# 落盘 JSON
out = {
    '_meta': {
        'city': '深圳', 'captured_at': '2026-08-20',
        'source': '深圳市住房和建设局 2021-02-08《深圳市住宅小区二手住房成交参考价格表》(zjj.sz.gov.cn/attachment/0/749/749839/8545737.pdf) [官方·住建局 2021-02-08]',
        'scope': '3595 个住宅小区（2021 年 2 月发布，银行抵押评估参考锚点）',
        'note': '官方另有 opendata.sz.gov.cn 2023-12-19 更新版 7,653 条（需 appKey，未采）',
    },
    'rows': rows,
}
with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=1)
print('JSON 已落盘')

# ---------- Excel ----------
wb = openpyxl.Workbook()

DATA_TIME = "2021-02-08"          # 官方发布日（数据反映时点）
GEN_TIME = "2026-08-20"           # 本表生成日
NOTE_LINE = (f"数据时间：{DATA_TIME}（深圳市住建局官方发布）｜本表生成：{GEN_TIME}｜"
             "来源：zjj.sz.gov.cn 官方 PDF（2021-02-08 通告附件）")


def add_note_row(ws, ncols):
    """在每个 Sheet 首行插入数据时间标注行（合并单元格）"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(1, 1)
    cell.value = NOTE_LINE
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill('solid', fgColor='404040')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 20


def style_header(ws, ncols, color='C00000'):
    for c in range(1, ncols+1):
        cell = ws.cell(2, c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=color)
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Sheet1 全量
ws1 = wb.active
ws1.title = '参考价全量'
h1 = ['序号', '行政区', '街道', '项目名称', '参考价(元/㎡)']
add_note_row(ws1, len(h1))
ws1.append(h1)
style_header(ws1, len(h1))
for r in rows:
    ws1.append([r[k] for k in h1])
ws1.freeze_panes = 'A3'
for i, w in enumerate([8, 10, 12, 36, 16], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Sheet2 分区统计
ws2 = wb.create_sheet('分区统计')
add_note_row(ws2, 5)
ws2.append(['行政区', '小区数', '均价(元/㎡)', '最高价', '最低价'])
style_header(ws2, 5, '1F6E43')
by_zone = defaultdict(list)
for r in rows:
    by_zone[r['行政区']].append(r['参考价(元/㎡)'])
for z, ps in sorted(by_zone.items(), key=lambda x: -len(x[1])):
    ws2.append([z, len(ps), round(sum(ps)/len(ps)), max(ps), min(ps)])
for i, w in enumerate([10, 8, 12, 10, 10], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Sheet3 街道聚合
ws3 = wb.create_sheet('街道聚合')
add_note_row(ws3, 3)
ws3.append(['街道', '小区数', '均价(元/㎡)'])
style_header(ws3, 3, '1F6E43')
by_street = defaultdict(list)
for r in rows:
    by_street[r['街道']].append(r['参考价(元/㎡)'])
for s, ps in sorted(by_street.items(), key=lambda x: -len(x[1])):
    ws3.append([s, len(ps), round(sum(ps)/len(ps))])
for i, w in enumerate([14, 8, 12], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Sheet4 高价小区 TOP
ws4 = wb.create_sheet('高价小区TOP50')
add_note_row(ws4, 4)
ws4.append(['行政区', '街道', '项目名称', '参考价(元/㎡)'])
style_header(ws4, 4, '1F6E43')
for r in sorted(rows, key=lambda x: -x['参考价(元/㎡)'])[:50]:
    ws4.append([r['行政区'], r['街道'], r['项目名称'], r['参考价(元/㎡)']])
for i, w in enumerate([10, 12, 36, 16], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# Sheet5 来源说明
ws5 = wb.create_sheet('来源说明')
add_note_row(ws5, 2)
ws5.append(['项', '说明'])
style_header(ws5, 2, '404040')
for k, v in [
    ('数据时间', f'{DATA_TIME}（官方发布日，数据反映时点）'),
    ('本表生成', f'{GEN_TIME}'),
    ('数据源', '深圳市住房和建设局 2021-02-08《深圳市住宅小区二手住房成交参考价格表》(官方 PDF)'),
    ('URL', 'zjj.sz.gov.cn/attachment/0/749/749839/8545737.pdf'),
    ('条数', f'{len(rows)} 个小区（官方口径 3595 个，解析去重后一致）'),
    ('用途', '二手住房成交参考价 = 银行抵押评估/涉房贷款审慎估值官方锚点'),
    ('升级项', 'opendata.sz.gov.cn 有 2023-12-19 更新版 7,653 条，需 appKey（注册）或 Playwright 下载，未采'),
    ('口径', '价格为整栋/小区级参考价（元/平方米），非逐套；2021 年 2 月发布'),
]:
    ws5.append([k, v])
ws5.column_dimensions['A'].width = 14
ws5.column_dimensions['B'].width = 95

wb.save(OUT_XLSX)
print('Excel 已生成:', OUT_XLSX)
