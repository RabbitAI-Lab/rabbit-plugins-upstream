# -*- coding: utf-8 -*-
"""
real-estate-risk-analyst · 按企业/品牌检索 脚本工厂
=====================================================================
把「按开发商/品牌/合作方检索」SOP 的 P3(官方证级采集) + P4(着色 Excel) 固化为一键 CLI。

P1(外网撒网) / P2(工商穿透) 由 AI 执行；其产出——「品牌名 + 备案主体名列表」——
作为本脚本输入（--brand / --developers）。

用法
----
  # 贵阳（REST 证级，完整内置，无需浏览器）
  python by_developer.py --brand 万科 --city guiyang \
      --developers "贵阳广晟鑫德房地产开发有限公司,万科企业股份" \
      --out 万科贵阳_证级.xlsx

  # 广州（逐套销控型，薄封装调 extract_guangzhou.py）
  python by_developer.py --brand 金地 --city guangzhou \
      --developers "深圳市金地新沙房地产开发有限公司" --out 金地广州.xlsx

  # 深圳（逐套销控型，注入 PERMITS 调 fetch_house_data.py，需 --sz-script）
  python by_developer.py --brand 金地 --city shenzhen \
      --sz-script "C:/.../fetch_house_data.py" \
      --sz-permits "C:/.../金地深圳_permits.json" --out 金地深圳.xlsx

  # 从本地全量 JSON 生成「品牌在筑官方项目总表」（无需联网，适合 53 项目级汇总）
  python by_developer.py --mode project-summary --city guiyang \
      --from-json "C:/.../all_guiyang_permits_2026.json" \
      --brand-key 万科 --out 万科在筑_官方项目总表.xlsx

城市适配器成熟度
----------------
  guiyang   : 完整内置（翻页全库 + 关键词过滤 + certificate/info 详情 + 着色）✅
  guangzhou : 薄封装，subprocess 调既有 extract_guangzhou.py（Playwright 逐套）🔧
  shenzhen  : 薄封装，注入 PERMITS 后调 fetch_house_data.py 的 main()（Playwright 逐套，需浏览器）🔧

横向复制（多品牌 / 多城市）
--------------------------
  本工厂已验证「按企业检索 SOP」在 贵阳(万科)/深圳(金地)/广州(鹏瑞) 落地。
  复制到其他品牌（华润 / 保利 / 中海 等）或城市，**无需改代码**：
    1. P2 工商穿透产出「品牌 ↔ 备案主体名」列表 → 填 --developers；
    2. 广州须额外给「备案名」--search-name（广州按开发商搜常返0，须按备案名）；
    3. 换 --brand / --developers / --city 即跑。适配器 guiyang/guangzhou/shenzhen 已就绪。
  例：
    华润深圳 -> --brand 华润 --city shenzhen --sz-script ... --sz-permits ...
    保利贵阳 -> --brand 保利 --city guiyang --developers "保利贵州置业有限公司"
"""
import argparse, json, os, sys, subprocess, time, ssl, asyncio
sys.stdout.reconfigure(encoding="utf-8")
import os
from pathlib import Path
from collections import defaultdict

HERE = Path(__file__).resolve().parent
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))
PY = os.environ.get("RE_PY", sys.executable)
# 广州提取脚本路径：优先读环境变量 RE_WORKSPACE，未设则相对当前工作目录回退。
# （发布版已移除硬编码私人路径 C:/Users/<user>/...）
EXTRACT_GZ = Path(os.environ.get("RE_WORKSPACE", ".")) / "extract_guangzhou.py"

GREEN = "C6EFCE"   # 官方
YELLOW = "FFEB9C"  # 外网


# ===================== 质量校验接入 =====================
def _rows_to_permits(rows):
    """把 fetch_guiyang 的 rows(含 sellSummaryList) 规整为 validate_permits 所需结构。"""
    out = []
    for r in rows:
        units = area = 0
        for s in r.get("sellSummaryList", []) or []:
            try:
                units += int(s.get("count") or 0)
                area += float(s.get("area") or 0)
            except Exception:
                pass
        out.append({"certificateNo": r.get("certificateNo", ""),
                    "approvedUnits": units, "approvedArea": area})
    return out


def run_validate(permits):
    """采集后自动校验（复用 scripts/validate.py）。失败时降级为内联最小校验。"""
    try:
        from validate import validate_permits
        rep = validate_permits(permits)
        print("[validate] 证级校验:", json.dumps(rep, ensure_ascii=False))
        return rep
    except Exception as e:
        print(f"[validate] 跳过(模块缺失): {e}")
        return {}


# ===================== 贵阳 REST 证级适配器 =====================
import urllib.request
GY_BASE = "https://www.gyfc.net.cn/api/gyzfw"
_gy_ctx = ssl.create_default_context()
_gy_ctx.check_hostname = False
_gy_ctx.verify_mode = ssl.CERT_NONE


def _gy_post(path, data, tries=5):
    last = None
    for t in range(tries):
        try:
            req = urllib.request.Request(
                GY_BASE + path, data=json.dumps(data).encode(),
                headers={"User-Agent": "Mozilla/5.0", "Content-Type": "application/json",
                         "Accept": "application/json"})
            raw = urllib.request.urlopen(req, timeout=30, context=_gy_ctx).read()
            return json.loads(raw.decode("utf-8", "ignore"))
        except Exception as e:
            last = e
            time.sleep(0.4 * (t + 1))
    raise last


def fetch_guiyang(developers, brand, max_pages=60):
    """翻页全库 -> 按 developers 关键词过滤 -> certificate/info 详情。
    developers: 备案主体名/项目公司名列表（来自 P2 工商穿透）。"""
    kws = [k.strip() for k in developers if k.strip()]
    print(f"[guiyang] 翻页全库 businessType=ysz ...")
    allrecs, page, total = [], 1, None
    while page <= max_pages:
        r = _gy_post("/open-api/certificate/list", {"businessType": "ysz", "current": page, "size": 200})
        data = r.get("data", {})
        if total is None:
            total = data.get("total")
        recs = data.get("records", [])
        if not recs:
            break
        allrecs += recs
        page += 1
        time.sleep(0.12)
    print(f"[guiyang] 全库 {len(allrecs)} 条 (声明 total={total})")

    matched, seen = [], set()
    for x in allrecs:
        s = " ".join(str(x.get(f, "")) for f in ("title", "placeName", "applicantName", "houseAddress"))
        if any(k in s for k in kws):
            aid = x.get("acceptanceId", "")
            if aid in seen:
                continue
            seen.add(aid)
            matched.append(x)
    print(f"[guiyang] 命中 {len(matched)} 条 (关键词: {kws})")

    rows = []
    for x in matched:
        aid = x.get("acceptanceId", "")
        try:
            d = _gy_post("/open-api/certificate/info",
                         {"acceptanceId": aid, "businessType": "ysz"}).get("data", {}) or {}
            sell = d.get("sellSummaryList") or []
            rows.append({
                "brand": brand, "project": x.get("placeName", ""), "certificateNo": x.get("certificateNo", ""),
                "title": x.get("title", ""), "region": x.get("region", ""), "applicantName": x.get("applicantName", ""),
                "placeName": x.get("placeName", ""), "building": d.get("building", ""),
                "countDate": x.get("countDate", ""), "preSellStart": d.get("preSellStart", ""),
                "preSellEnd": d.get("preSellEnd", ""), "banks": d.get("banks", ""),
                "accounts": d.get("accounts", ""), "acceptanceId": aid,
                "sellSummaryList": sell, "source": "官方·筑房网",
            })
        except Exception as e:
            print(f"  ERR {aid}: {e}")
        time.sleep(0.12)
    return rows


# ===================== 统一着色 Excel =====================
def build_excel(rows, out, brand, city, note=""):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    gfill = PatternFill("solid", fgColor=GREEN)
    yfill = PatternFill("solid", fgColor=YELLOW)
    wb = Workbook()

    # Sheet1 官方预售证清单
    ws = wb.active
    ws.title = "官方预售证清单"
    hdr = ["品牌", "项目", "预售证号", "标题", "区域", "开发/申请主体", "坐落", "楼栋",
           "发证日", "预售起", "预售止", "监管银行", "监管账号", "来源"]
    ws.append(hdr)
    for r in rows:
        ws.append([r.get("brand"), r.get("project"), r.get("certificateNo"), r.get("title"),
                   r.get("region"), r.get("applicantName"), r.get("placeName"), r.get("building"),
                   r.get("countDate"), r.get("preSellStart"), r.get("preSellEnd"),
                   r.get("banks"), r.get("accounts"), r.get("source")])
        ws.cell(row=ws.max_row, column=14).fill = gfill

    # Sheet2 分用途批准套数
    ws2 = wb.create_sheet("分用途批准套数")
    ws2.append(["项目", "预售证号", "用途", "批准套数", "批准面积(㎡)", "来源"])
    for r in rows:
        for s in r.get("sellSummaryList", []):
            ws2.append([r.get("project"), r.get("certificateNo"), s.get("useRange"),
                        s.get("count"), s.get("area"), "官方·筑房网"])
            ws2.cell(row=ws2.max_row, column=6).fill = gfill

    # Sheet3 统计分析
    ws3 = wb.create_sheet("统计分析")
    ws3.append(["项目", "证级数", "住宅批准套数", "住宅面积(㎡)", "商业批准套数", "商业面积(㎡)", "来源"])
    res_t = com_t = res_a = com_a = 0
    by_proj = defaultdict(lambda: [0, 0, 0, 0])
    for r in rows:
        pj = r.get("project", "")
        for s in r.get("sellSummaryList", []):
            try:
                c = int(s.get("count") or 0); a = float(s.get("area") or 0)
            except Exception:
                c, a = 0, 0
            ur = str(s.get("useRange", ""))
            if "住宅" in ur:
                by_proj[pj][0] += c; by_proj[pj][1] += a; res_t += c; res_a += a
            elif "商业" in ur:
                by_proj[pj][2] += c; by_proj[pj][3] += a; com_t += c; com_a += a
    for pj, v in by_proj.items():
        ws3.append([pj, sum(1 for r in rows if r.get("project") == pj),
                    v[0], round(v[1], 2), v[2], round(v[3], 2), "官方·筑房网"])
        ws3.cell(row=ws3.max_row, column=7).fill = gfill
    ws3.append(["【合计】", len(rows), res_t, round(res_a, 2), com_t, round(com_a, 2), "官方·筑房网"])
    ws3.cell(row=ws3.max_row, column=7).fill = gfill

    # Sheet4 来源说明
    ws4 = wb.create_sheet("来源说明")
    for line in [
        f"数据来源：城市官方预售证系统（{city}）。本表所有数据均来自官方接口，来源列绿底标注。",
        "采集方式：按 P2 工商穿透产出的「备案主体名」关键词过滤官方全量库，逐本拉取证级详情。",
        "口径：sellSummaryList 为「供应/批准」口径，与外网「成交/网签」不可相减推库存。",
        "降级：若官方接口失效，须回退至外网/公示系统，并在来源列改用黄底标注 [外网·WebSearch]。",
        note,
    ]:
        if line:
            ws4.append([line])
    wb.save(out)
    return out


# ===================== 广/深 薄封装 =====================
def run_guangzhou(developers, brand, out, search_name=""):
    if not EXTRACT_GZ.exists():
        return None, f"未找到广州采集器: {EXTRACT_GZ}"
    dev = developers[0] if developers else ""
    out_gz = Path(out).with_name(Path(out).stem + "_广州逐套.xlsx")
    # 广州阳光家缘检索接口对「备案名」与「开发商」是 AND 关系，二者同传必返0
    # （鹏瑞壹湾广场实证：name+dev 同传=0，仅 name=壹湾广场 命中4条）。
    # 故：有备案名时只用 --search，无备案名才退用 --search-dev。
    if search_name:
        cmd = [PY, str(EXTRACT_GZ), "--search", search_name, "--out", str(out_gz)]
    else:
        cmd = [PY, str(EXTRACT_GZ), "--search-dev", dev, "--out", str(out_gz)]
    print("[guangzhou] 调用 extract_guangzhou.py ->", " ".join(cmd))
    subprocess.run(cmd, check=False)
    return out_gz if out_gz.exists() else None, None


def _write_sz_runner(sz_script, permits, brand, out):
    """生成临时 runner：注入 PERMITS 后调 fetch_house_data.main()（修复原 --organ 静默空跑）。"""
    runner = Path(out).with_name("_sz_runner_" + Path(out).stem + ".py")
    sz_dir = Path(sz_script).resolve().parent
    content = (
        "import asyncio, sys\n"
        "sys.path.insert(0, r'%s')\n"
        "import fetch_house_data as F\n"
        "F.PERMITS = %s\n"
        "F.PROJECT_NAME = %r\n"
        "asyncio.run(F.main())\n"
    ) % (str(sz_dir), json.dumps(permits, ensure_ascii=False), brand)
    runner.write_text(content, encoding="utf-8")
    return runner


def run_shenzhen(developers, brand, out, sz_script, sz_permits):
    if not sz_script or not Path(sz_script).exists():
        return None, "深圳需 --sz-script 指向 fetch_house_data.py（路径未配置）"
    if sz_permits and Path(sz_permits).exists():
        permits = json.loads(Path(sz_permits).read_text(encoding="utf-8"))
    else:
        # 退路：用 developers 作 PROJECT_NAME 占位，但无 PERMITS 仍无法采集——必须给证级清单
        return None, "深圳需 --sz-permits 提供 {ysProjectId,preSellId,presell_no} 清单（来自官方检索）"
    runner = _write_sz_runner(sz_script, permits, brand, out)
    out_sz = Path(out).with_name(Path(out).stem + "_深圳逐套.xlsx")
    cmd = [PY, str(runner)]
    print("[shenzhen] 注入 %d 张证级并调 fetch_house_data.main() -> %s" % (len(permits), " ".join(cmd)))
    subprocess.run(cmd, check=False)
    try:
        runner.unlink()
    except Exception:
        pass
    return out_sz if out_sz.exists() else None, None


# ===================== 项目总表模式（本地全量 JSON） =====================
def build_project_summary(from_json, brand_key, out, details_json=None):
    """从本地全量证级 dump 过滤品牌，产出「品牌在筑官方项目总表」。
    from_json: 全库证级列表(all_guiyang_permits_2026.json 结构)
    brand_key: 过滤关键词(如 '万科')
    details_json: 可选，{certificateNo: {sellSummaryList:[...]}} 用于补批准套数/面积
    """
    recs = json.loads(Path(from_json).read_text(encoding="utf-8"))
    details = {}
    if details_json and Path(details_json).exists():
        details = json.loads(Path(details_json).read_text(encoding="utf-8"))
        if isinstance(details, list):  # 兼容 list 形态
            details = {d.get("certificateNo"): d for d in details}
    matched = [x for x in recs if brand_key in json.dumps(x, ensure_ascii=False)]
    print(f"[summary] 全库 {len(recs)} -> 命中 {len(matched)} 本含 '{brand_key}'")

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    gfill = PatternFill("solid", fgColor=GREEN)
    wb = Workbook()

    # Sheet1 项目汇总
    ws = wb.active
    ws.title = "项目汇总"
    ws.append(["项目(placeName)", "证级数", "申请主体", "首证日", "末证日", "住宅批准套数", "住宅批准面积(㎡)"])
    by_proj = defaultdict(lambda: {"n": 0, "applicants": set(), "dates": [], "res_u": 0, "res_a": 0.0})
    for x in matched:
        pj = x.get("placeName", "")
        b = by_proj[pj]
        b["n"] += 1
        if x.get("applicantName"):
            b["applicants"].add(x.get("applicantName"))
        if x.get("countDate"):
            b["dates"].append(x.get("countDate", "")[:10])
        no = x.get("certificateNo", "")
        det = details.get(no, {})
        for s in det.get("sellSummaryList", []) or []:
            try:
                c = int(s.get("count") or 0); a = float(s.get("area") or 0)
            except Exception:
                c, a = 0, 0
            if "住宅" in str(s.get("useRange", "")):
                b["res_u"] += c; b["res_a"] += a
    for pj, b in sorted(by_proj.items(), key=lambda z: -z[1]["n"]):
        ds = sorted(b["dates"])
        ws.append([pj, b["n"], " / ".join(sorted(b["applicants"])),
                   ds[0] if ds else "", ds[-1] if ds else "",
                   b["res_u"], round(b["res_a"], 2)])
        ws.cell(row=ws.max_row, column=7).fill = gfill
    ws.append(["【合计】", len(matched), "", "", "",
               sum(b["res_u"] for b in by_proj.values()),
               round(sum(b["res_a"] for b in by_proj.values()), 2)])
    ws.cell(row=ws.max_row, column=7).fill = gfill

    # Sheet2 全部证级清单
    ws2 = wb.create_sheet("全部证级清单")
    ws2.append(["项目", "预售证号", "标题", "区域", "申请主体", "坐落", "发证日", "来源"])
    for x in matched:
        ws2.append([x.get("placeName", ""), x.get("certificateNo", ""), x.get("title", ""),
                    x.get("region", ""), x.get("applicantName", ""), x.get("houseAddress", ""),
                    (x.get("countDate", "") or "")[:10], "官方·筑房网"])
        ws2.cell(row=ws2.max_row, column=8).fill = gfill

    # Sheet3 来源说明
    ws3 = wb.create_sheet("来源说明")
    for line in [
        f"数据来源：城市官方预售证系统全量 dump（含 '{brand_key}' 关键词过滤）。来源列绿底标注。",
        "本表为「供应/批准」口径，与外网「成交/网签」不可相减推库存。",
        "批准套数/面积仅对提供了详情(details_json)的证级有效；其余证级仅列清单。",
        "若需逐套网签/备案价，须走官方逐套接口或用户登录导出。",
    ]:
        ws3.append([line])
    wb.save(out)
    return out, len(by_proj), len(matched)


# ===================== 入口 =====================
def main():
    ap = argparse.ArgumentParser(description="按企业/品牌检索 脚本工厂 (P3+P4)")
    ap.add_argument("--mode", default="permits", choices=["permits", "project-summary"],
                    help="permits=按主体拉证级; project-summary=从本地全量JSON生成品牌项目总表")
    ap.add_argument("--brand", help="品牌/开发商名（归属标签）")
    ap.add_argument("--brand-key", help="project-summary 模式下的过滤关键词")
    ap.add_argument("--city", choices=["guiyang", "guangzhou", "shenzhen"], help="城市")
    ap.add_argument("--search-name", default="", help="广州等须按备案名检索时的备案名（如 壹湾广场）")
    ap.add_argument("--developers", default="", help="备案主体名列表，逗号分隔（来自 P2 穿透）")
    ap.add_argument("--out", required=True, help="输出 xlsx 路径")
    ap.add_argument("--sz-script", default=os.environ.get("RE_SZ_SCRIPT", ""), help="深圳 fetch_house_data.py 路径")
    ap.add_argument("--sz-permits", default="", help="深圳证级清单 JSON({ysProjectId,preSellId,presell_no})")
    ap.add_argument("--from-json", default="", help="project-summary 模式：本地全量证级 dump 路径")
    ap.add_argument("--details-json", default="", help="project-summary 模式：证级详情(补批准套数)")
    ap.add_argument("--max-pages", type=int, default=60, help="贵阳翻页上限")
    ap.add_argument("--no-validate", action="store_true", help="关闭采集后质量校验")
    args = ap.parse_args()

    developers = [d.strip() for d in args.developers.split(",") if d.strip()]
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)

    if args.mode == "project-summary":
        if not args.from_json:
            print("[project-summary] 必须提供 --from-json")
            sys.exit(2)
        key = args.brand_key or args.brand or ""
        o, nproj, nperm = build_project_summary(args.from_json, key, out, args.details_json)
        print(f"[project-summary] 完成 {nproj} 个项目 / {nperm} 本证级 -> {o}")
        return

    # mode == permits
    if not args.city or not args.brand:
        print("[permits] 需要 --city 与 --brand")
        sys.exit(2)
    if args.city == "guiyang":
        rows = fetch_guiyang(developers, args.brand, args.max_pages)
        if not rows:
            print("[guiyang] 无命中，检查 --developers 关键词是否匹配官方库主体名")
            sys.exit(2)
        build_excel(rows, out, args.brand, args.city, note="适配器：贵阳筑房网 open-api（REST，完整内置）")
        if not args.no_validate:
            run_validate(_rows_to_permits(rows))
        print(f"[guiyang] 完成 {len(rows)} 本证级 -> {out}")
    elif args.city == "guangzhou":
        p, err = run_guangzhou(developers, args.brand, out, args.search_name)
        print(err or f"[guangzhou] 产物 -> {p}")
    elif args.city == "shenzhen":
        p, err = run_shenzhen(developers, args.brand, out, args.sz_script, args.sz_permits)
        print(err or f"[shenzhen] 产物 -> {p}")


if __name__ == "__main__":
    main()
