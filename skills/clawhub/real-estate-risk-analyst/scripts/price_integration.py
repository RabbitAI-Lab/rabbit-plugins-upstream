#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
price_integration.py —— 外部价格整合与货值测算（D/E 类缺价城市闭环）

问题背景：
- A 类城市（深/佛/珠等）Playwright 实拉逐套含官方备案价，货值/折扣率/去化可直接算。
- D/E 类城市（合肥/东莞/惠州/肇庆/中山/江门等）官方多只公开证级、逐套/备案价锁登录或根本无制度，
  平台不公开价。为补全数据完整性，须用外网（贝壳/安居客/房天下/乐有家/吉屋等）交叉验证价格，
  形成"官方证级（绿底）+ 外网补价（黄底）"双色整合，支撑去化/折扣率/货值量级分析。

输入：
- 官方证级 JSON（projects[]，结构同 by_developer 的 project-summary，或筑房网/政务服务网 dump）
- 外部价格 JSON（由 Agent 用 WebSearch 获取后落盘，结构见 external_prices 示例）
  {
    "search_date": "2026-08-12",
    "city": "合肥",
    "tag": "WebSearch",
    "records": [
      {"channel":"贝壳","type":"一手","desc":"万科某某盘 89㎡ 三房","area":89,"unit_price":18000,"total_price":1602000,"date":"2026-08-10","url":"https://...","note":"挂牌"},
      ...
    ]
  }

输出：多 Sheet Excel（官方证级/外网补价/价格口径对比/货值测算/数据来源说明），来源着色。

用法：
  python price_integration.py --prices ext.json --permits permits.json --out out.xlsx
  python price_integration.py --demo          # 内置小样本跑通，验证逻辑
"""

import json
import argparse
import os
import statistics
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ---- 来源着色规范（与全 skill 一致）----
GREEN = PatternFill("solid", fgColor="C6EFCE")  # 官方
YELLOW = PatternFill("solid", fgColor="FFEB9C")  # 外网
RED = PatternFill("solid", fgColor="FFC7CE")     # 尽调红旗
HEADER = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True)


def load_external(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_permits(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def norm(s):
    """归一用于模糊匹配：去空格/小写/去常见后缀"""
    if not s:
        return ""
    s = str(s).strip().lower()
    for suf in ["（", "）", "(", ")", "有限公司", "公司", "置业", "地产", "房地产"]:
        s = s.replace(suf, "")
    return s


def match_projects(permits, external):
    """把外网价格记录按项目名模糊匹配到官方证级项目，返回 enriched 列表。"""
    pj = permits.get("projects", [])
    recs = external.get("records", [])
    # 建立官方项目名索引（归一）
    pindex = {}
    for p in pj:
        key = norm(p.get("project", ""))
        if key:
            pindex.setdefault(key, []).append(p)

    matched = []
    for r in recs:
        rk = norm(r.get("desc", ""))  # 子串匹配：项目级外网描述可命中楼栋级官方项目
        hit = None
        for key, plist in pindex.items():
            if key and (key in rk or rk in key):
                hit = plist[0]
                break
        enriched = dict(r)
        enriched["_matched_project"] = hit.get("project") if hit else None
        enriched["_matched_cert"] = hit.get("certificate_no") if hit else None
        matched.append(enriched)
    return matched


def compute_metrics(matched, permits):
    """计算折扣率/去化/货值量级。"""
    pj = permits.get("projects", [])
    rows = []
    for p in pj:
        approved = p.get("approved_units") or 0
        approved_area = p.get("approved_area") or 0
        official_avg = p.get("avg_price")  # 官方备案均价（A 类有，D/E 多为 None）
        # 该项目匹配到的外网价
        ext = [m for m in matched if m.get("_matched_project") == p.get("project")]
        ext_prices = [m["unit_price"] for m in ext if m.get("unit_price")]
        # 用中位数抗极端值（如楼王 56万/㎡ 会拉偏均值）
        ext_avg = round(statistics.median(ext_prices), 2) if ext_prices else None
        # 折扣率：外网市场价 / 官方备案均价（仅官方有价时算）
        discount = None
        if official_avg and ext_avg:
            discount = round(ext_avg / official_avg, 4)
        # 货值（外网口径，仅供参考）：外网均价 × 批准面积
        value_ext = round(ext_avg * approved_area, 2) if (ext_avg and approved_area) else None
        rows.append({
            "project": p.get("project"),
            "certificate_no": p.get("certificate_no"),
            "approved_units": approved,
            "approved_area": approved_area,
            "official_avg": official_avg,
            "ext_avg": ext_avg,
            "ext_count": len(ext),
            "discount_vs_official": discount,
            "value_ext_est": value_ext,
        })
    return rows


def build_excel(permits, external, matched, metrics, out, external_path):
    wb = Workbook()
    # Sheet1 官方证级（绿底）
    ws1 = wb.active
    ws1.title = "官方证级"
    cols1 = ["项目", "预售证号", "批准套数", "批准面积(㎡)", "官方备案均价", "来源"]
    ws1.append(cols1)
    for c in range(1, len(cols1) + 1):
        ws1.cell(1, c).fill = HEADER
        ws1.cell(1, c).font = BOLD
    for p in permits.get("projects", []):
        ws1.append([p.get("project"), p.get("certificate_no"),
                    p.get("approved_units"), p.get("approved_area"),
                    p.get("avg_price"), f"[官方·{permits.get('city','')} {external.get('search_date','')}]"])
        for c in range(1, len(cols1) + 1):
            ws1.cell(ws1.max_row, c).fill = GREEN

    # Sheet2 外网补价（黄底）
    ws2 = wb.create_sheet("外网补价")
    cols2 = ["渠道", "类型", "描述", "面积", "单价", "总价", "日期", "匹配项目", "来源"]
    ws2.append(cols2)
    for c in range(1, len(cols2) + 1):
        ws2.cell(1, c).fill = HEADER
        ws2.cell(1, c).font = BOLD
    for m in matched:
        ws2.append([m.get("channel"), m.get("type"), m.get("desc"), m.get("area"),
                    m.get("unit_price"), m.get("total_price"), m.get("date"),
                    m.get("_matched_project"), f"[外网·{external.get('tag','WebSearch')} {external.get('search_date','')}]"])
        for c in range(1, len(cols2) + 1):
            ws2.cell(ws2.max_row, c).fill = YELLOW

    # Sheet3 价格口径对比 + 货值测算
    ws3 = wb.create_sheet("价格对比与货值")
    cols3 = ["项目", "批准套数", "批准面积", "官方均价", "外网均价", "匹配外网条数",
             "折扣率(外网/官方)", "货值估算(外网口径)"]
    ws3.append(cols3)
    for c in range(1, len(cols3) + 1):
        ws3.cell(1, c).fill = HEADER
        ws3.cell(1, c).font = BOLD
    for r in metrics:
        ws3.append([r["project"], r["approved_units"], r["approved_area"], r["official_avg"],
                    r["ext_avg"], r["ext_count"], r["discount_vs_official"], r["value_ext_est"]])

    # Sheet4 数据来源说明
    ws4 = wb.create_sheet("数据来源说明")
    ws4.append(["渠道", "URL/用途", "可信度", "口径声明"])
    for c in range(1, 5):
        ws4.cell(1, c).fill = HEADER
        ws4.cell(1, c).font = BOLD
    ws4.append([f"[官方·{permits.get('city','')}]", "政府预售证系统", "高", "证级批准量=供应口径"])
    ws4.append([f"[外网·{external.get('tag','WebSearch')} {external.get('search_date','')}]",
                external_path, "中（挂牌/参考）", "市场价非政府备案价，仅作量级参考"])
    ws4.append(["", "", "", "供应(批准)≠成交(网签)，二者不可相减推库存"])

    # 列宽
    for ws in [ws1, ws2, ws3, ws4]:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    wb.save(out)
    return out


def demo():
    """内置小样本，验证脚本逻辑闭环（不依赖真实外网数据）。"""
    permits = {
        "city": "合肥(示例)",
        "projects": [
            {"project": "万科某某盘", "certificate_no": "合房预售证2024001", "approved_units": 500,
             "approved_area": 50000, "avg_price": None},
            {"project": "保利某盘", "certificate_no": "合房预售证2024002", "approved_units": 300,
             "approved_area": 30000, "avg_price": 20000},
        ],
    }
    external = {
        "search_date": "2026-08-12",
        "city": "合肥(示例)",
        "tag": "WebSearch",
        "records": [
            {"channel": "贝壳", "type": "一手", "desc": "万科某某盘 89㎡ 三房", "area": 89,
             "unit_price": 18000, "total_price": 1602000, "date": "2026-08-10", "url": "https://x", "note": "挂牌"},
            {"channel": "安居客", "type": "一手", "desc": "保利某盘 99㎡", "area": 99,
             "unit_price": 19000, "total_price": 1881000, "date": "2026-08-09", "url": "https://y", "note": "挂牌"},
        ],
    }
    matched = match_projects(permits, external)
    metrics = compute_metrics(matched, permits)
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "demo_price_integration.xlsx")
    build_excel(permits, external, matched, metrics, out, "<demo>")
    print(f"[demo] 已生成 {out}")
    print(f"[demo] 匹配外网记录 {len(matched)} 条，货值测算 {len(metrics)} 行")
    for r in metrics:
        print(f"  - {r['project']}: 外网均价={r['ext_avg']} 折扣率={r['discount_vs_official']} 货值估算={r['value_ext_est']}")


def main():
    ap = argparse.ArgumentParser(description="外部价格整合与货值测算")
    ap.add_argument("--prices", help="外部价格 JSON（WebSearch 落盘）")
    ap.add_argument("--permits", help="官方证级 JSON")
    ap.add_argument("--out", default="price_integration.xlsx")
    ap.add_argument("--demo", action="store_true", help="内置小样本跑通")
    args = ap.parse_args()

    if args.demo:
        demo()
        return
    if not (args.prices and args.permits):
        ap.error("--prices 与 --permits 必须同时提供（或用 --demo）")

    external = load_external(args.prices)
    permits = load_permits(args.permits)
    matched = match_projects(permits, external)
    metrics = compute_metrics(matched, permits)
    out = build_excel(permits, external, matched, metrics, args.out, args.prices)
    print(f"已生成 {out}（官方证级 {len(permits.get('projects', []))} 项，外网匹配 {len(matched)} 条）")


if __name__ == "__main__":
    main()
