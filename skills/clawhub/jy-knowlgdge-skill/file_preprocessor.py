"""
文件预处理引擎
将任意格式文件转换为高质量 Markdown，供 EasyDataset 使用。
采用"占位符替换法"处理图片和表格：
  1. 解析文档结构 → 提取纯文本 + 占位符标记图片/表格位置
  2. 异步处理非文本元素（视觉LLM 处理图片/复杂表格）
  3. 占位符替换 → 回填到 Markdown 对应位置
"""

import os
import re
import hashlib
import json
from typing import Optional, Dict, Any, List, Tuple
from pathlib import Path
from dataclasses import dataclass, field


@dataclass
class Placeholder:
    """占位符"""
    tag: str           # 如 {{IMG_001}}, {{TBL_001}}
    type: str          # "image" | "table" | "pdf_page"
    position: int      # 在文本中的位置
    file_path: str     # 图片/截图文件路径
    metadata: dict     # 额外元数据（如表格检测结果）
    result: str = ""   # 处理后的结果


class FilePreprocessor:
    """
    文件预处理引擎
    支持格式：.md, .txt, .docx, .pdf, .png, .jpg, .jpeg, .bmp, .webp, .xlsx, .csv
    """

    # 支持的文件类型
    TEXT_FORMATS = {".md", ".txt"}
    DOCX_FORMATS = {".docx"}
    PDF_FORMATS = {".pdf"}
    IMAGE_FORMATS = {".png", ".jpg", ".jpeg", ".bmp", ".webp"}
    TABLE_FORMATS = {".xlsx", ".csv"}

    def __init__(self, vision_llm_client=None, temp_dir: str = None, max_vision_concurrency: int = 10):
        """
        vision_llm_client: 视觉大模型客户端（需实现 analyze_image(image_path, prompt) -> str）
        temp_dir: 临时文件目录
        max_vision_concurrency: 视觉LLM最大并发数（≤20）
        """
        self.vision_llm = vision_llm_client
        self.temp_dir = temp_dir or "D:/knowledge_skill/processed"
        self.max_vision_concurrency = min(max_vision_concurrency, 20)
        try:
            os.makedirs(self.temp_dir, exist_ok=True)
        except PermissionError:
            import tempfile
            self.temp_dir = os.path.join(tempfile.gettempdir(), "knowledge_skill_processed")
            os.makedirs(self.temp_dir, exist_ok=True)

        # 测试写入权限
        test_file = os.path.join(self.temp_dir, ".write_test")
        try:
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
        except (PermissionError, OSError):
            import tempfile
            self.temp_dir = os.path.join(tempfile.gettempdir(), "knowledge_skill_processed")
            os.makedirs(self.temp_dir, exist_ok=True)

    def _ensure_writable(self, output_dir: str) -> str:
        """确保输出目录可写，否则回退到临时目录"""
        try:
            os.makedirs(output_dir, exist_ok=True)
            test_file = os.path.join(output_dir, ".write_test")
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
            return output_dir
        except (PermissionError, OSError):
            import tempfile
            fallback = os.path.join(tempfile.gettempdir(), "knowledge_skill_processed")
            os.makedirs(fallback, exist_ok=True)
            return fallback

    def process(self, file_path: str, output_dir: str = None) -> str:
        """
        主入口：处理任意文件，返回 Markdown 文件路径

        file_path: 输入文件路径
        output_dir: 输出目录（默认为 self.temp_dir）
        返回：生成的 .md 文件路径
        """
        file_path = os.path.normpath(file_path)
        ext = os.path.splitext(file_path)[1].lower()
        output_dir = self._ensure_writable(output_dir or self.temp_dir)

        # 生成输出文件名（基于原始文件名 + MD5 短哈希）
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        file_hash = self._file_md5(file_path)[:8]
        safe_name = self._safe_filename(f"{base_name}_{file_hash}")
        output_path = os.path.join(output_dir, f"{safe_name}.md")

        # 如果已处理过，直接返回
        if os.path.exists(output_path):
            return output_path

        # 根据文件类型路由处理
        if ext in self.TEXT_FORMATS:
            markdown = self._process_text(file_path)
        elif ext in self.DOCX_FORMATS:
            markdown = self._process_docx(file_path)
        elif ext in self.PDF_FORMATS:
            markdown = self._process_pdf(file_path)
        elif ext in self.IMAGE_FORMATS:
            markdown = self._process_image(file_path)
        elif ext in self.TABLE_FORMATS:
            markdown = self._process_table_file(file_path)
        else:
            raise ValueError(f"Unsupported file type: {ext}")

        # 写入输出文件
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(markdown)

        return output_path

    # ==================== 各格式处理器 ====================

    def _process_text(self, file_path: str) -> str:
        """处理 .md / .txt 文件"""
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()

        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".txt":
            # TXT 文件添加基本标题
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            content = f"# {base_name}\n\n{content}"

        # 清理多余空行
        content = re.sub(r'\n{4,}', '\n\n\n', content)
        return content.strip() + "\n"

    def _process_docx(self, file_path: str) -> str:
        """处理 .docx 文件：提取文本 + 图片(视觉LLM) + 表格 → Markdown"""
        try:
            import mammoth
        except ImportError:
            raise ImportError("请安装 mammoth: pip install mammoth")

        # Step 1: 使用 mammoth 提取文档内容（得到包含 base64 图片的 markdown）
        with open(file_path, "rb") as f:
            result = mammoth.convert_to_markdown(f)
        markdown = result.value

        # Step 2: 如果有视觉LLM，并行处理内嵌的 base64 图片
        if self.vision_llm:
            markdown = self._replace_base64_images(markdown, max_concurrency=self.max_vision_concurrency)

        # Step 3: 清理 mammoth 转换产物
        markdown = self._clean_markdown(markdown)
        return markdown.strip() + "\n"

    def _replace_base64_images(self, markdown: str, max_concurrency: int = 10) -> str:
        """检测 markdown 中的 base64 图片，并行调用视觉LLM分析后替换为文字描述"""
        import base64
        from concurrent.futures import ThreadPoolExecutor, as_completed

        pattern = r'!\[.*?\]\((data:image/[^;]+;base64,([^\)]+))\)'

        # Step 1: 找出所有匹配
        matches = list(re.finditer(pattern, markdown, flags=re.DOTALL))
        if not matches:
            return markdown

        # Step 2: 准备图片信息（解码 + 写临时文件）
        img_infos = []
        for i, match in enumerate(matches):
            try:
                mime = match.group(1).split(';')[0].split(':')[1]
                b64_data = match.group(2)
                img_bytes = base64.b64decode(b64_data)
                ext_map = {'image/png': '.png', 'image/jpeg': '.jpg', 'image/gif': '.gif',
                           'image/webp': '.webp', 'image/bmp': '.bmp'}
                ext = ext_map.get(mime, '.png')
                import tempfile
                img_path = os.path.join(tempfile.gettempdir(), f"docx_img_{i}_{hash(match.group(0))}{ext}")
                with open(img_path, 'wb') as f:
                    f.write(img_bytes)
                img_infos.append((i, img_path, match))
            except Exception:
                img_infos.append((i, None, match))  # 标记为失败

        # Step 3: 并行调用视觉LLM（并发数 = min(max_concurrency, 图片数)）
        result_map = {}  # i -> description
        llm = self.vision_llm
        prompt = (
            "请详细描述此图片的全部内容，包括："
            "1. 图中人物、物体、场景；"
            "2. 图中的文字内容（完整提取）；"
            "3. 图表/数据的内容和含义（如是图表）；"
            "4. 图片主题和传达的信息。"
        )

        valid_infos = [(i, path) for i, path, _ in img_infos if path]
        workers = min(max_concurrency, max(len(valid_infos), 1))

        def analyze_one(i, img_path):
            try:
                desc = llm.analyze_image(img_path, prompt)
                return (i, desc.strip() if desc else "")
            except Exception as e:
                return (i, None, str(e))
            finally:
                try:
                    os.remove(img_path)
                except Exception:
                    pass

        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(analyze_one, i, path): i for i, path in valid_infos}
            for future in as_completed(futures):
                result = future.result()
                i = result[0]
                if len(result) == 2:
                    result_map[i] = result[1]
                else:
                    result_map[i] = f"[图片 - 处理失败: {result[2]}]"

        # 处理失败的图片
        for i, path, _ in img_infos:
            if i not in result_map:
                result_map[i] = "[图片 - 处理失败]"

        # Step 4: 从后往前替换（避免索引偏移，使用 enumerate 避免 O(n²) 查找）
        # 先构建 (原始索引, match) 对，然后反转顺序替换
        indexed_matches = list(enumerate(matches))
        parts = list(markdown)
        for i, match in reversed(indexed_matches):
            desc = result_map.get(i, "[图片 - 处理失败]")
            replacement = f"\n\n[图片描述]\n\n{desc}\n\n"
            start, end = match.span()
            parts[start:end] = list(replacement)

        return ''.join(parts)

    def _process_pdf(self, file_path: str) -> str:
        """
        处理 PDF 文件
        策略：
        1. 先尝试直接提取文本（pdfplumber）
        2. 如果文本质量差/含图片，使用视觉LLM逐页处理
        """
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        texts = []

        # 策略1：文本提取
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                for i, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if text and text.strip():
                        texts.append(f"## 第 {i+1} 页\n\n{text.strip()}")
                    else:
                        texts.append(None)  # 标记需要视觉处理
        except ImportError:
            # pdfplumber 不可用，全部使用视觉处理
            if self.vision_llm:
                texts = [None]  # 简化处理，使用视觉LLM
            else:
                raise ImportError("请安装 pdfplumber: pip install pdfplumber，或提供视觉LLM客户端")

        # 策略2：对空页使用视觉LLM（并行处理，并发数≤max_vision_concurrency）
        if self.vision_llm:
            needs_vision = [i for i, t in enumerate(texts) if t is None]
            if needs_vision:
                page_images = self._pdf_to_images(file_path, pages=needs_vision)

                from concurrent.futures import ThreadPoolExecutor, as_completed
                workers = min(self.max_vision_concurrency, len(needs_vision))
                prompt = "请将此PDF页面内容转为Markdown格式，保留标题层级、段落结构、列表、表格数据"

                def analyze_page(i, img_path):
                    try:
                        desc = self.vision_llm.analyze_image(img_path, prompt)
                        return (i, desc.strip() if desc else None)
                    except Exception:
                        return (i, None)

                with ThreadPoolExecutor(max_workers=workers) as executor:
                    futures = {executor.submit(analyze_page, pg, path): pg
                               for pg, path in zip(needs_vision, page_images)}
                    for future in as_completed(futures):
                        pg, desc = future.result()
                        if desc:
                            texts[pg] = f"## 第 {pg+1} 页\n\n{desc}"

        # 合成最终 Markdown
        content = f"# {base_name}\n\n"
        for i, text in enumerate(texts):
            if text:
                content += text + "\n\n"

        return content.strip() + "\n"

    def _process_image(self, file_path: str) -> str:
        """处理纯图片文件：视觉LLM分析 → 文字描述"""
        if not self.vision_llm:
            raise RuntimeError("处理图片需要视觉LLM客户端")

        base_name = os.path.splitext(os.path.basename(file_path))[0]
        prompt = (
            "请详细描述此图片的全部内容，包括："
            "1. 图中的人物、物体、场景"
            "2. 图中的文字内容（如有）"
            "3. 图表/数据的内容和含义（如是图表）"
            "4. 图片的主题和传达的信息"
        )
        description = self.vision_llm.analyze_image(file_path, prompt)

        content = f"# {base_name}\n\n[图片描述]\n\n{description.strip()}\n"
        return content

    def _process_table_file(self, file_path: str) -> str:
        """
        处理表格文件 (.xlsx/.csv/.xls)
        优先使用脚本转换，合并单元格等复杂情况使用视觉LLM
        """
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        ext = os.path.splitext(file_path)[1].lower()

        markdown = f"# {base_name}\n\n"

        if ext == ".csv":
            # CSV 直接用 pandas
            try:
                import pandas as pd
                df = pd.read_csv(file_path)
                markdown += df.to_markdown(index=False)
                return markdown.strip() + "\n"
            except ImportError:
                raise ImportError("请安装 pandas: pip install pandas")

        elif ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path)

                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    has_merged = bool(ws.merged_cells.ranges)

                    if has_merged and self.vision_llm:
                        # 复杂表格：截图后用视觉LLM处理
                        markdown += f"## {sheet_name}\n\n"
                        markdown += self._handle_complex_table(file_path, sheet_name)

                    else:
                        # 简单表格：直接转换
                        markdown += f"## {sheet_name}\n\n"
                        markdown += self._sheet_to_markdown(ws)
                        markdown += "\n"

                wb.close()
                return markdown.strip() + "\n"

            except ImportError:
                raise ImportError("请安装 openpyxl: pip install openpyxl")

        return markdown.strip() + "\n"

    # ==================== 辅助方法 ====================

    def _clean_markdown(self, text: str) -> str:
        """清理 Markdown 文本"""
        # 移除多余空行
        text = re.sub(r'\n{4,}', '\n\n\n', text)
        # 标准化标题格式
        text = re.sub(r'^#{1,6}\s*$', '', text, flags=re.MULTILINE)
        return text

    def _safe_filename(self, name: str) -> str:
        """安全文件名（移除非法字符）"""
        return re.sub(r'[<>:"/\\|?*]', '_', name)

    def _file_md5(self, file_path: str) -> str:
        """计算文件 MD5"""
        h = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                h.update(chunk)
        return h.hexdigest()

    def _pdf_to_images(self, pdf_path: str, pages: list = None) -> list:
        """PDF 转图片列表"""
        try:
            import fitz  # pymupdf
        except ImportError:
            raise ImportError("请安装 pymupdf: pip install pymupdf")

        images = []
        doc = fitz.open(pdf_path)
        page_indices = pages or range(len(doc))

        for i in page_indices:
            if i >= len(doc):
                break
            page = doc[i]
            pix = page.get_pixmap(dpi=150)
            img_path = os.path.join(self.temp_dir, f"pdf_page_{i+1}_{self._file_md5(pdf_path)[:8]}.png")
            pix.save(img_path)
            images.append(img_path)

        doc.close()
        return images

    def _sheet_to_markdown(self, ws) -> str:
        """openpyxl worksheet 转 Markdown 表格"""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return ""

        lines = []
        for i, row in enumerate(rows):
            cells = [str(cell) if cell is not None else "" for cell in row]
            lines.append("| " + " | ".join(cells) + " |")
            if i == 0:
                lines.append("|" + "|".join([" --- "] * len(cells)) + "|")

        return "\n".join(lines)

    def _handle_complex_table(self, file_path: str, sheet_name: str) -> str:
        """
        处理复杂表格（合并单元格等）
        将表格区域渲染为截图 → 视觉LLM识别 → 输出文字描述
        """
        if not self.vision_llm:
            return "[复杂表格 - 请手动处理]\n"

        img_path = None
        try:
            import openpyxl
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            from PIL import Image

            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]

            # 获取有数据的行列范围
            rows_data = []
            max_cols = 0
            for row in ws.iter_rows(min_row=1, values_only=True):
                # 跳过完全空行
                row_vals = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in row_vals):
                    rows_data.append(row_vals)
                    max_cols = max(max_cols, len(row_vals))

            if not rows_data:
                wb.close()
                return "[空表格]\n"

            # 填充到统一列数
            rows_data = [r + [""] * (max_cols - len(r)) for r in rows_data]

            # 用 matplotlib 渲染表格截图
            fig, ax = plt.subplots(figsize=(max(12, max_cols * 2.5), max(3, len(rows_data) * 0.6)))
            ax.axis('off')
            ax.axis('tight')

            table = ax.table(
                cellText=rows_data,
                loc='center',
                cellLoc='left',
                colWidths=[0.15] * max_cols
            )
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1.0, 1.2)

            # 设置表头行样式（深色背景）
            for j in range(max_cols):
                cell = table[0, j]
                cell.set_facecolor('#4472C4')
                cell.set_text_props(color='white', fontweight='bold')

            # 保存为临时图片
            import tempfile
            img_path = os.path.join(tempfile.gettempdir(), f"complex_table_{hash(sheet_name)}_{os.getpid()}.png")
            plt.tight_layout(pad=0.5)
            plt.savefig(img_path, dpi=150, bbox_inches='tight', facecolor='white')
            plt.close(fig)
            wb.close()

            # 发送给视觉LLM
            prompt = (
                f"请将此Excel表格'{sheet_name}'的内容转为Markdown表格格式。"
                "保留所有行列结构，正确处理合并单元格信息。"
                "如果是合并单元格，请将合并单元格的内容填充到对应位置。"
                "只输出Markdown表格，不要任何额外解释。"
            )
            description = self.vision_llm.analyze_image(img_path, prompt)

            # 清理临时图片
            try:
                os.remove(img_path)
            except Exception:
                pass

            return (description or "[复杂表格 - 视觉LLM返回空]") + "\n"

        except Exception as e:
            # 清理临时文件
            if img_path:
                try:
                    os.remove(img_path)
                except Exception:
                    pass
            return f"[复杂表格 - 转换失败: {e}]\n"


class VisionLLMClient:
    """
    通用视觉大模型客户端
    支持 OpenAI 兼容 API（可对接 Ollama/DeepSeek/本地模型等）
    """

    def __init__(self, base_url: str, api_key: str, model: str,
                 temperature: float = 0.3, max_tokens: int = 4096):
        self.base_url = base_url.rstrip("/")
        self.api_key = api_key
        self.model = model
        self.temperature = temperature
        self.max_tokens = max_tokens

    def analyze_image(self, image_path: str, prompt: str) -> str:
        """
        使用视觉模型分析图片
        image_path: 图片文件路径
        prompt: 分析提示词
        返回模型回复文本
        """
        import base64

        # 读取并编码图片
        with open(image_path, "rb") as f:
            image_data = base64.b64encode(f.read()).decode("utf-8")

        # 推断 MIME 类型
        ext = os.path.splitext(image_path)[1].lower()
        mime_map = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                     ".webp": "image/webp", ".bmp": "image/bmp"}
        mime_type = mime_map.get(ext, "image/png")

        import requests
        payload = {
            "model": self.model,
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {
                        "url": f"data:{mime_type};base64,{image_data}"
                    }}
                ]
            }],
            "temperature": self.temperature,
            "max_tokens": self.max_tokens
        }

        resp = requests.post(
            f"{self.base_url}/chat/completions",
            json=payload,
            headers={
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json"
            },
            timeout=120
        )
        resp.raise_for_status()
        result = resp.json()
        msg = result["choices"][0]["message"]
        # 兼容思考模型：content 优先，为空时用 reasoning_content
        content = msg.get("content", "")
        if not content or not content.strip():
            content = msg.get("reasoning_content", "")
        # 清理思考模型标记
        import re
        content = re.sub(r'<\|?im_end\|?>', '', content)
        content = re.sub(r'<\|?im_start\|?>[^\n]*\n?', '', content)
        content = re.sub(r'^\s* response\s*', '', content.strip())
        return content.strip()
