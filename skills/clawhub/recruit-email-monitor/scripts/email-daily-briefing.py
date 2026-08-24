#!/usr/bin/env python3
"""
每日招聘邮件简报
每天早上 9:00 运行，汇总表格中的信息并发送
"""

import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import json
import os
import sys

# 本地配置文件（含 feishu_target，不随 Skill 发布）
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
from excel_styles import EXCEL_PATH, SHEET_MAIL  # noqa: E402
CONFIG_FILE = os.path.join(SCRIPT_DIR, 'config.json')


def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


CFG = load_config()
FEISHU_TARGET = CFG.get('feishu_target', 'user:YOUR_FEISHU_USER_ID')

# 表格路径（合并总表，读邮件 sheet）
EXCEL_PATH = EXCEL_PATH

# 简报输出路径
BRIEFING_PATH = '/home/erhao/shared/招聘邮件每日简报.txt'

# 超期自动归档阈值（超过该天数的待处理邮件自动标记为已完成）
STALE_DAYS = 30

def load_pending_emails():
    """加载待处理的邮件（状态不是已完成的）"""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[SHEET_MAIL] if SHEET_MAIL in wb.sheetnames else wb.active
        
        emails = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 日期列
                status = row[4] if len(row) > 4 else ''
                # 待处理：状态不是"已完成"
                if status and '✅' not in status and '完成' not in status:
                    emails.append({
                        'date': row[0],
                        'account': row[1],
                        'subject': row[2],
                        'from': row[3],
                        'status': status,
                        'type': row[5] if len(row) > 5 else '',
                        'link': row[6] if len(row) > 6 else '',
                        'deadline': row[7] if len(row) > 7 else ''
                    })
        
        return emails
    
    except Exception as e:
        print(f"❌ 读取表格失败：{e}")
        return []

def parse_date(value):
    """把表格中的日期值解析为 datetime，兼容 datetime 对象和字符串。"""
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def auto_expire_stale_emails():
    """自动归档超期待处理邮件：收到时间超过 STALE_DAYS 天的待处理邮件标记为已完成。

    返回本次归档数量。归档后的邮件不再出现在简报中。
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[SHEET_MAIL] if SHEET_MAIL in wb.sheetnames else wb.active
        now = datetime.now()
        archived = 0

        for row in ws.iter_rows(min_row=2):
            date_val = row[0].value if len(row) > 0 else None
            status_cell = row[4] if len(row) > 4 else None
            if date_val is None or status_cell is None or not status_cell.value:
                continue
            status = str(status_cell.value)
            # 已完成的跳过
            if '✅' in status or '完成' in status:
                continue
            dt = parse_date(date_val)
            if dt is None:
                continue
            if (now - dt).days > STALE_DAYS:
                status_cell.value = '✅ 已完成（超期自动归档）'
                archived += 1

        if archived:
            wb.save(EXCEL_PATH)
            print(f"🗂️ 自动归档 {archived} 封超过 {STALE_DAYS} 天的待处理邮件")
        else:
            print("🗂️ 无超期待处理邮件需要归档")
        return archived
    except Exception as e:
        print(f"❌ 自动归档失败：{e}")
        return 0


def generate_briefing(emails, archived=0):
    """生成简报内容"""
    if not emails:
        archived_note = f"\n🗂️ 本次自动归档 {archived} 封超期邮件（> {STALE_DAYS} 天）\n" if archived else ""
        return f"""
═══════════════════════════════════════════════════
📧 招聘邮件每日简报
日期：{datetime.now().strftime('%Y年%m月%d日 %A')}
═══════════════════════════════════════════════════
{archived_note}
✅ 所有邮件都已处理完毕！

祝你有愉快的一天！✨
"""
    
    # 按类型分组
    by_type = defaultdict(list)
    for email in emails:
        by_type[email['type']].append(email)
    
    # 检查是否有即将截止的
    today = datetime.now()
    urgent = []
    for email in emails:
        if email.get('deadline'):
            try:
                deadline_str = str(email['deadline'])[:10]
                deadline = datetime.strptime(deadline_str, '%Y-%m-%d')
                days_left = (deadline - today).days
                if days_left <= 3:
                    urgent.append((email, days_left))
            except:
                pass
    
    # 生成简报
    briefing = f"""
═══════════════════════════════════════════════════
📧 招聘邮件每日简报
日期：{datetime.now().strftime('%Y年%m月%d日 %A')}
═══════════════════════════════════════════════════

📊 待处理概览
───────────────────────────────────────────────────
待处理邮件：{len(emails)} 封
"""
    
    if archived:
        briefing += f"🗂️ 本次自动归档 {archived} 封超期邮件（> {STALE_DAYS} 天）\n"
    
    # 类型统计
    for email_type, type_emails in sorted(by_type.items()):
        briefing += f"  • {email_type}: {len(type_emails)} 封\n"
    
    if urgent:
        briefing += f"\n⚠️ 即将截止：{len(urgent)} 封（3 天内）\n"
    
    briefing += f"""
───────────────────────────────────────────────────
📋 待处理详情
───────────────────────────────────────────────────
"""
    
    # 先列出紧急的
    if urgent:
        briefing += "\n🔴 即将截止（3 天内）\n"
        briefing += "─" * 40 + "\n"
        for email, days_left in urgent:
            deadline_str = str(email.get('deadline', '未知'))[:10]
            briefing += f"""
• {email['subject']}
  状态：{email['status']}
  截止：{deadline_str}（还剩{days_left}天）
  邮箱：{email['account']}
"""
            if email.get('link'):
                briefing += f"  链接：{email['link'][:100]}...\n"
    
    # 按类型列出详情
    for email_type, type_emails in sorted(by_type.items()):
        emoji = {
            '笔试/测评': '✍️',
            '面试': '🎤',
            'Offer/录用': '🎉',
            '宣讲会': '📢',
            '投递确认': '✅',
            '其他': '📧'
        }.get(email_type, '📧')
        
        briefing += f"\n{emoji} {email_type} ({len(type_emails)}封)\n"
        briefing += "─" * 40 + "\n"
        
        for i, email in enumerate(type_emails, 1):
            # 跳过已在紧急列表中的
            if any(e is email for e, _ in urgent):
                continue
            
            briefing += f"""
{i}. {email['subject']}
   状态：{email['status']}
   邮箱：{email['account']}
   发件人：{email['from']}
   时间：{email['date']}
"""
            if email.get('deadline'):
                briefing += f"   截止：{str(email['deadline'])[:10]}\n"
            if email.get('link'):
                briefing += f"   链接：{email['link']}\n"
    
    briefing += f"""
═══════════════════════════════════════════════════
💡 提醒
───────────────────────────────────────────────────
"""
    
    # 智能提醒
    reminders = []
    
    if urgent:
        reminders.append("• ⚠️ 有邮件即将截止，请优先处理！")
    
    if '笔试/测评' in by_type:
        reminders.append("• 注意查看笔试/测评邮件，通常有截止时间")
    
    if '面试' in by_type:
        reminders.append("• 面试邮件请及时回复确认")
    
    if '宣讲会' in by_type:
        reminders.append("• 宣讲会通常需要提前报名")
    
    if 'Offer/录用' in by_type:
        reminders.append("• 🎉 恭喜！Offer 邮件请仔细阅读条款")
    
    if reminders:
        briefing += "\n".join(reminders) + "\n"
    else:
        briefing += "• 保持关注，继续加油！💪\n"
    
    briefing += f"""
═══════════════════════════════════════════════════
生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
═══════════════════════════════════════════════════
"""
    
    return briefing

def send_via_feishu_api(briefing):
    """通过飞书开放平台 API 直接发送简报（不经过 Agent/LLM）。

    使用 config.json 中的 feishu_app_id / feishu_app_secret 获取 tenant_access_token，
    然后调用 im/v1/messages 接口发送文本消息。
    成功返回 True，失败返回 False。
    """
    import json
    import urllib.request
    import urllib.error

    app_id = CFG.get('feishu_app_id', '')
    app_secret = CFG.get('feishu_app_secret', '')
    if not app_id or not app_secret:
        print("⚠️ 未配置 feishu_app_id/feishu_app_secret，跳过 API 直发")
        return False

    receive_id = FEISHU_TARGET.split(':', 1)[-1] if ':' in FEISHU_TARGET else FEISHU_TARGET

    try:
        # 1. 获取 tenant_access_token
        token_url = 'https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal'
        token_body = json.dumps({'app_id': app_id, 'app_secret': app_secret}).encode('utf-8')
        req = urllib.request.Request(token_url, data=token_body, headers={'Content-Type': 'application/json'}, method='POST')
        with urllib.request.urlopen(req, timeout=30) as resp:
            token_data = json.loads(resp.read().decode('utf-8'))
        if token_data.get('code') != 0:
            print(f"❌ 获取 tenant_access_token 失败：{token_data}")
            return False
        token = token_data['tenant_access_token']

        # 2. 发送文本消息
        msg_url = 'https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=open_id'
        content = json.dumps({'text': briefing.strip()}, ensure_ascii=False)
        msg_body = json.dumps({
            'receive_id': receive_id,
            'msg_type': 'text',
            'content': content
        }, ensure_ascii=False).encode('utf-8')
        req = urllib.request.Request(msg_url, data=msg_body, headers={
            'Content-Type': 'application/json; charset=utf-8',
            'Authorization': f'Bearer {token}'
        }, method='POST')
        with urllib.request.urlopen(req, timeout=30) as resp:
            msg_data = json.loads(resp.read().decode('utf-8'))
        if msg_data.get('code') == 0:
            print(f"✅ 简报已通过飞书 API 直发成功（msg_id={msg_data.get('data', {}).get('message_id', '?')}）")
            return True
        else:
            print(f"❌ 飞书发送失败：{msg_data}")
            return False
    except urllib.error.HTTPError as e:
        print(f"❌ 飞书 API HTTP 错误 {e.code}：{e.read().decode('utf-8', errors='replace')}")
        return False
    except Exception as e:
        print(f"❌ 飞书 API 直发异常：{e}")
        return False


def send_briefing(briefing):
    """保存并输出简报，默认通过飞书 API 直发。

    注意：不要在 Agent 会话运行期间调用 `openclaw message send` CLI，
    否则会因会话文件锁（SessionWriteLockTimeoutError）而失败。
    默认走飞书开放平台 API 直发（不经过 Agent/LLM），
    可通过环境变量 BRIEFING_SEND_API=0 关闭直发、BRIEFING_SEND_CLI=1 启用 CLI 发送。
    """
    import os
    import subprocess

    # 保存简报到文件
    try:
        with open(BRIEFING_PATH, 'w', encoding='utf-8') as f:
            f.write(briefing)
        print(f"✅ 简报已保存到：{BRIEFING_PATH}")
    except Exception as e:
        print(f"❌ 保存简报失败：{e}")
    
    # 打印简报内容（cron Agent 需要把完整内容转发给用户）
    print("\n" + briefing)

    # 默认：飞书 API 直发（不经过 LLM）
    if os.environ.get('BRIEFING_SEND_API', '1') != '0':
        print("\n📤 正在通过飞书 API 直发...")
        ok = send_via_feishu_api(briefing)
        if ok:
            return

    # 可选：独立 CLI 发送（仅 BRIEFING_SEND_CLI=1 时启用）
    if os.environ.get('BRIEFING_SEND_CLI') == '1':
        print("\n📤 正在通过 CLI 发送 Feishu 消息...")
        try:
            cmd = [
                'openclaw', 'message', 'send',
                '--channel', 'feishu',
                '--target', FEISHU_TARGET,
                '--message', briefing.strip()
            ]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=90)
            if result.returncode == 0:
                print("✅ Feishu 消息发送成功！")
            else:
                print(f"❌ 发送失败：{result.stderr}")
        except subprocess.TimeoutExpired:
            print("❌ 发送超时")
        except Exception as e:
            print(f"❌ 发送异常：{e}")

def main():
    print('🌅 每日招聘邮件简报\n')
    print(f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print()
    
    # 先自动归档超期待处理邮件（>30 天标记为已完成）
    archived = auto_expire_stale_emails()
    print()

    # 加载待处理的邮件
    emails = load_pending_emails()
    
    # 生成简报
    briefing = generate_briefing(emails, archived)
    
    # 发送简报
    send_briefing(briefing)
    
    print("\n✅ 每日简报完成")

if __name__ == '__main__':
    main()
