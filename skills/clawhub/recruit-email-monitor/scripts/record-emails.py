#!/usr/bin/env python3
"""
根据 Agent 判定结果（pending_judged.json）记录招聘邮件到 Excel，
并把所有已判定邮件标记为已处理（避免重复拉取）。

判定文件格式（由 Agent 生成）:
[
  {
    "id": 1,
    "verdict": "yes" | "no",
    "type": "笔试/测评|面试|Offer/录用|宣讲会|投递确认|其他招聘相关",  # verdict=yes 时必填
    "deadline": "8月10日 23:59",   # 可选，从邮件内容提取的截止时间
    "reason": "简短判断理由"
  }
]

用法:
    python3 scripts/record-emails.py
"""

import json
import os
import subprocess
import sys
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from excel_styles import (ensure_headers, style_header, style_body, style_status_cell, style_type_cell,
                          refresh_filter, setup_status_column, EXCEL_PATH, SHEET_MAIL, MAIL_HEADERS)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROCESSED_FILE = os.path.join(SCRIPT_DIR, 'processed_emails.json')
CANDIDATES_FILE = os.path.join(SCRIPT_DIR, 'pending_candidates.json')
JUDGED_FILE = os.path.join(SCRIPT_DIR, 'pending_judged.json')

VALID_TYPES = ['笔试/测评', '面试', 'Offer/录用', '宣讲会', '投递确认', '其他招聘相关']


def clean_from_address(from_addr):
    """清理发件人地址，格式与旧数据保持一致"""
    if from_addr.startswith('"') and '<' in from_addr:
        return from_addr
    if '<' in from_addr and not from_addr.startswith('"'):
        parts = from_addr.split('<')
        name = parts[0].strip()
        email = '<' + parts[1]
        if name and not name.startswith('"'):
            return f'"{name}" {email}'
        return from_addr
    return from_addr


def main():
    if not os.path.exists(CANDIDATES_FILE) or not os.path.exists(JUDGED_FILE):
        print("缺少 pending_candidates.json 或 pending_judged.json，请先运行 fetch-emails.py 并由 Agent 完成判定")
        return 1

    with open(CANDIDATES_FILE, 'r', encoding='utf-8') as f:
        candidates = json.load(f)
    with open(JUDGED_FILE, 'r', encoding='utf-8') as f:
        judged = json.load(f)

    cand_by_id = {c['id']: c for c in candidates}
    processed = []
    if os.path.exists(PROCESSED_FILE):
        with open(PROCESSED_FILE, 'r', encoding='utf-8') as f:
            processed = json.load(f)

    # 表格不存在时自动创建（含表头）；存在则确保表头样式统一（显式取邮件 sheet）
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_MAIL
        ensure_headers(ws, MAIL_HEADERS)
    else:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[SHEET_MAIL] if SHEET_MAIL in wb.sheetnames else wb.active
        ensure_headers(ws, MAIL_HEADERS)
        style_header(ws)

    added = 0
    rejected = 0
    skipped = 0

    for j in judged:
        c = cand_by_id.get(j.get('id'))
        if not c:
            print(f"⚠️ 判定引用了不存在的候选 id={j.get('id')}，跳过")
            skipped += 1
            continue
        uid = c['uid']
        if uid in processed:
            continue

        verdict = str(j.get('verdict', 'no')).strip().lower()
        if verdict in ('yes', 'true', '1', 'y'):
            email_type = j.get('type') or '其他招聘相关'
            if email_type not in VALID_TYPES:
                print(f"⚠️ 未知类型 '{email_type}'（候选 #{c['id']}），回退为 '其他招聘相关'")
                email_type = '其他招聘相关'
            deadline = j.get('deadline') or None
            ws.append([
                c['date'],
                c['account'],
                c['subject'],
                clean_from_address(c['from']),
                '⏳ 待处理',
                email_type,
                c.get('link'),
                deadline,
            ])
            added += 1
            print(f"✅ 记录招聘邮件: {c['subject'][:60]} [{email_type}]")
        else:
            rejected += 1
            print(f"⏭️  非招聘邮件: {c['subject'][:60]}")

        processed.append(uid)

    # 美化：整表统一样式（斑马纹/边框/行高/超链接）+ 状态/类型彩色标签
    style_body(ws, ncols=len(MAIL_HEADERS), left_cols=(3, 4, 7), link_col=7)
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
        style_status_cell(row[0])
        style_type_cell(row[1])
    refresh_filter(ws)
    # 状态列下拉列表 + 条件格式（用户可直接在表格里切换状态，颜色自动跟随）
    setup_status_column(ws)

    wb.save(EXCEL_PATH)
    with open(PROCESSED_FILE, 'w', encoding='utf-8') as f:
        json.dump(processed, f, ensure_ascii=False, indent=2)

    # 根据 Agent 判定中的 progress 指令，增量更新投递记录进度表
    # （不再全量重建：Agent 已智能判断轮次/是否更新，规则重建会覆盖这些判断）
    try:
        subprocess.run(
            [sys.executable, os.path.join(SCRIPT_DIR, 'apply-progress-updates.py')],
            check=False, timeout=120,
        )
    except Exception as e:
        print(f"⚠️ 更新投递记录进度表失败：{e}")

    # 清理临时文件
    for p in (CANDIDATES_FILE, JUDGED_FILE):
        if os.path.exists(p):
            os.remove(p)

    print(f"\n完成: 记录 {added} 封招聘邮件，标记 {rejected} 封为非招聘，跳过 {skipped} 条异常")
    return 0


if __name__ == '__main__':
    sys.exit(main())
