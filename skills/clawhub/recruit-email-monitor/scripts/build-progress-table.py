#!/usr/bin/env python3
"""
投递记录进度表生成器（全量重建版，手动运行/初始化用）

从总表「招聘邮件汇总」sheet 聚合生成「投递记录进度表」sheet：
- 只统计 2026-08-01 以来的邮件（用户 8 月开始投秋招正式批）
- 按公司分组邮件，自动提取：投递时间、测评时间、一面/二面/三面/HR面时间、结果、投递链接、最近动态
- 保留用户手动维护的字段：投递岗位、备注、自定义结果（不在自动结果候选集内的旧值视为手动填写）
- 宣讲会邮件不构成投递进度，自动排除；用户标记"不需要"的邮件自动排除

⚠️ 注意：日常更新请走 Agent 判定流程（record-emails.py → apply-progress-updates.py），
本脚本只用于初始化/修复进度表，不要频繁全量重建（会覆盖 Agent 的智能轮次判断）。

用法:
    python3 scripts/build-progress-table.py
"""

import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from excel_styles import (ensure_headers, style_header, style_body, style_result_cell, refresh_filter,
                          setup_result_column, EXCEL_PATH, SHEET_MAIL, SHEET_PROGRESS, PROGRESS_HEADERS)
from company_extract import extract_company, extract_position

# 只统计该日期（含）之后的邮件
SINCE_DATE = '2026-08-01'

# 自动推断结果候选值（不在其中的旧值视为用户手动填写，重建时保留）
AUTO_RESULT_VALUES = {
    '✅ Offer', '⏳ 进行中', '🎤 流程结束', '✍️ 流程结束', '📮 已投递',
}

# 面试邮件中不计入面试轮次的主题关键词（反馈/问卷/体验类）
NON_INTERVIEW_KEYWORDS = ['反馈', '问卷', '体验']


def first_time_on_day(mails, day, etype):
    """返回该公司某日期第一封指定类型邮件的时间"""
    for m in mails:
        if str(m['date'])[:10] == day and m['type'] == etype:
            return str(m['date'])
    return day


def aggregate_company(mails):
    """按公司聚合：提取进度表各字段"""
    mails = sorted(mails, key=lambda m: str(m['date']))
    dates = [str(m['date']) for m in mails]
    types = [m['type'] for m in mails]

    # 投递时间：优先投递确认，否则最早邮件
    confirms = [d for m, d in zip(mails, dates) if m['type'] == '投递确认']
    apply_time = confirms[0] if confirms else dates[0]

    # 测评时间：最早的笔试/测评邮件
    assesses = [d for m, d in zip(mails, dates) if m['type'] == '笔试/测评']
    assess_time = assesses[0] if assesses else None

    # 面试轮次：排除反馈/问卷/体验类，按日期去重
    interview_days = []
    for m in mails:
        if m['type'] != '面试':
            continue
        subj = m['subject'] or ''
        if any(k in subj for k in NON_INTERVIEW_KEYWORDS):
            continue
        day = str(m['date'])[:10]
        if day not in interview_days:
            interview_days.append(day)

    # HR面：主题明确含 HR面/HR面试/人力资源 的面试邮件
    hr_time = None
    for m in mails:
        if m['type'] == '面试':
            subj = m['subject'] or ''
            if any(k in subj for k in ['HR面', 'HR面试', '人力资源', 'hr面']):
                hr_time = str(m['date'])
                break
    if hr_time is None and len(interview_days) >= 4:
        hr_time = first_time_on_day(mails, interview_days[3], '面试')

    def iv_time(idx):
        return first_time_on_day(mails, interview_days[idx], '面试') if len(interview_days) > idx else None

    # 结果推断
    if 'Offer/录用' in types:
        result = '✅ Offer'
    elif any('待处理' in (m['status'] or '') or '⏳' in (m['status'] or '') for m in mails):
        result = '⏳ 进行中'
    elif '面试' in types:
        result = '🎤 流程结束'
    elif '笔试/测评' in types:
        result = '✍️ 流程结束'
    else:
        result = '📮 已投递'

    # 投递链接：最早出现的有意义链接
    link = next((m['link'] for m in mails if m.get('link')), None)

    # 最近动态：最新一封邮件
    latest = mails[-1]
    latest_dynamic = f"{latest['date']} {latest['type']}：{(latest['subject'] or '')[:30]}"

    # 岗位提取：只用流程类邮件主题（投递确认/笔试测评/面试/Offer），排除宣传类
    flow_subjects = ' '.join(
        m['subject'] or '' for m in mails
        if m['type'] in ('投递确认', '笔试/测评', '面试', 'Offer/录用')
    )
    return {
        'company': mails[0]['company'],
        'position': extract_position(flow_subjects),
        'apply_time': apply_time,
        'assess_time': assess_time,
        'i1': iv_time(0),
        'i2': iv_time(1),
        'i3': iv_time(2),
        'hr': hr_time,
        'result': result,
        'dynamic': latest_dynamic,
        'link': link,
        'remark': None,
        'mail_count': len(mails),
    }


HEADERS = PROGRESS_HEADERS


def load_old_rows(wb):
    """从总表的进度 sheet 读旧行，返回 {公司名: [行dict, ...]}（保留用户手动字段）"""
    old = defaultdict(list)
    if SHEET_PROGRESS not in wb.sheetnames:
        return old
    try:
        ws = wb[SHEET_PROGRESS]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            old[row[1] if len(row) > 1 else ''].append({
                'position': row[2] if len(row) > 2 else None,
                'remark': row[12] if len(row) > 12 else None,
                'result': row[9] if len(row) > 9 else None,
            })
    except Exception as e:
        print(f"⚠️ 读取旧进度表失败（将全量重建）：{e}")
    return old


def build():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ 找不到汇总表：{EXCEL_PATH}")
        return 1

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_MAIL] if SHEET_MAIL in wb.sheetnames else wb.active

    mails = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        date_str = str(row[0])
        if date_str[:10] < SINCE_DATE:
            continue                      # 只统计 8 月以来
        status = row[4] or ''
        etype = row[5] or ''
        if '用户不需要' in status:
            continue          # 用户明确不需要的邮件
        if etype == '宣讲会':
            continue          # 宣讲会不构成投递进度
        mails.append({
            'date': date_str,
            'subject': row[2] or '',
            'from': row[3] or '',
            'status': status,
            'type': etype,
            'link': row[6] or None,
        })

    # 按公司分组（平台邮件且无法归属公司的直接忽略）
    groups = defaultdict(list)
    for m in mails:
        company = extract_company(m['subject'], m['from'])
        if company is None:
            continue
        m['company'] = company
        groups[company].append(m)

    new_rows = []
    for company, ms in groups.items():
        agg = aggregate_company(ms)
        agg['company'] = company
        new_rows.append(agg)

    # 与旧表合并（保留用户手动字段）
    old_rows = load_old_rows(wb)
    merged = []
    used_old = set()

    for nr in sorted(new_rows, key=lambda r: r['apply_time'], reverse=True):
        company = nr['company']
        olds = old_rows.get(company, [])
        if not olds:
            merged.append(nr)
            continue
        # 复用旧表第一行，保留其手动字段
        o = olds[0]
        used_old.add(company)
        nr['position'] = nr['position'] or o['position']
        nr['remark'] = o['remark']
        if o['result'] and o['result'] not in AUTO_RESULT_VALUES:
            nr['result'] = o['result']   # 用户自定义结果优先
        merged.append(nr)
        # 旧表多余行（同公司多岗位）保留，自动字段同步更新
        for extra in olds[1:]:
            extra_row = dict(nr)
            extra_row['position'] = extra_row['position'] or extra['position']
            extra_row['remark'] = extra['remark']
            if extra['result'] and extra['result'] not in AUTO_RESULT_VALUES:
                extra_row['result'] = extra['result']
            merged.append(extra_row)

    # 旧表中有而新聚合没有的公司行（用户手动添加）→ 保留
    for company, olds in old_rows.items():
        if company in used_old or not olds:
            continue
        for o in olds:
            merged.append({
                'company': company,
                'position': o['position'],
                'apply_time': None,
                'assess_time': None,
                'i1': None, 'i2': None, 'i3': None, 'hr': None,
                'result': o['result'] or '',
                'dynamic': '',
                'link': None,
                'remark': o['remark'],
            })

    # 重建进度 sheet（替换旧 sheet，邮件 sheet 保留）
    if SHEET_PROGRESS in wb.sheetnames:
        del wb[SHEET_PROGRESS]
    out_ws = wb.create_sheet(SHEET_PROGRESS)

    ensure_headers(out_ws, HEADERS)
    style_header(out_ws)

    seq = 1
    for r in merged:
        out_ws.append([
            seq, r['company'], r['position'] or '', r['apply_time'] or '',
            r['assess_time'] or '', r['i1'] or '', r['i2'] or '', r['i3'] or '',
            r['hr'] or '', r['result'], r['dynamic'], r['link'] or '', r['remark'] or '',
        ])
        seq += 1

    # 美化：斑马纹 + 边框 + 行高留白 + 超链接，结果列彩色标签
    style_body(out_ws, ncols=len(HEADERS), left_cols=(11, 12, 13), link_col=12)
    for row in out_ws.iter_rows(min_row=2, min_col=10, max_col=10):
        style_result_cell(row[0])

    # 列宽
    widths = [6, 12, 12, 16, 16, 16, 16, 16, 16, 11, 34, 40, 20]
    for i, w in enumerate(widths, start=1):
        out_ws.column_dimensions[get_column_letter(i)].width = w

    refresh_filter(out_ws)
    # 结果列下拉列表 + 条件格式（用户可直接在表格里切换结果，颜色自动跟随）
    setup_result_column(out_ws)

    # 默认打开邮件 sheet
    wb.active = wb.sheetnames.index(SHEET_MAIL) if SHEET_MAIL in wb.sheetnames else 0

    wb.save(EXCEL_PATH)
    print(f"✅ 投递记录进度表已更新：{EXCEL_PATH}（sheet: {SHEET_PROGRESS}，{len(merged)} 家公司，仅统计 {SINCE_DATE} 以来）")
    return 0


if __name__ == '__main__':
    sys.exit(build())
