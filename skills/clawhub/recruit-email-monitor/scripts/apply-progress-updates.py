#!/usr/bin/env python3
"""
投递记录进度表 - Agent 增量更新器

读取 pending_judged.json 中 Agent 给出的 progress 更新指令，
增量更新总表「投递记录进度表」sheet（不重建，只更新指定行/字段）。

判定文件格式（Agent 生成，verdict=yes 时可携带 progress）:
[
  {
    "id": 1,
    "verdict": "yes",
    "type": "面试",
    "deadline": "8月12日 16:00",
    "progress": {
      "update": true,              # 是否更新进度表，默认 true；false 表示不更新
      "company": "中兴通讯",         # 公司名（缺省时由脚本自动提取）
      "position": "软件开发工程师",   # 岗位（可选，仅当表格该行岗位为空时写入）
      "stage": "一面",              # 更新阶段：投递|测评|一面|二面|三面|HR面|Offer|结果
      "time": "2026-08-11 10:56",  # 该阶段时间（可选，缺省用邮件日期）
      "link": "https://...",       # 投递链接（可选，仅当表格为空时写入）
      "result": "⏳ 进行中",         # 结果（stage=结果 或 Offer 时使用）
      "note": "面试改到8月12日"      # 备注（可选，追加到现有备注，保留用户手动内容）
    },
    "reason": "中兴通讯面试通知，更新为一面"
  }
]

阶段映射:
  投递 → 投递时间    测评 → 测评时间    一面 → 一面时间
  二面 → 二面时间    三面 → 三面时间    HR面 → HR面时间
  Offer → 结果列 = "✅ Offer"   结果 → 结果列 = progress.result

用法:
    python3 scripts/apply-progress-updates.py
"""

import json
import os
import sys
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from excel_styles import (ensure_headers, style_header, style_body, style_result_cell, refresh_filter,
                          setup_result_column, EXCEL_PATH, SHEET_MAIL, SHEET_PROGRESS,
                          MAIL_HEADERS, PROGRESS_HEADERS)
from company_extract import extract_company

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JUDGED_FILE = os.path.join(SCRIPT_DIR, 'pending_judged.json')
CANDIDATES_FILE = os.path.join(SCRIPT_DIR, 'pending_candidates.json')

# 只处理该日期（含）之后的邮件
SINCE_DATE = '2026-08-01'

# 列索引（1-based）
COL = {
    '序号': 1, '公司名称': 2, '投递岗位': 3, '投递时间': 4, '测评时间': 5,
    '一面时间': 6, '二面时间': 7, '三面时间': 8, 'HR面时间': 9, '结果': 10,
    '最近动态': 11, '投递链接': 12, '备注': 13,
}

# stage → (列名, 是否结果列)
STAGE_COLUMN = {
    '投递': ('投递时间', False),
    '测评': ('测评时间', False),
    '一面': ('一面时间', False),
    '二面': ('二面时间', False),
    '三面': ('三面时间', False),
    'HR面': ('HR面时间', False),
    'Offer': ('结果', True),
    '结果': ('结果', True),
}

HEADERS = PROGRESS_HEADERS


def ensure_table():
    """确保总表存在且包含进度 sheet，返回 workbook + worksheet"""
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if SHEET_PROGRESS in wb.sheetnames:
            ws = wb[SHEET_PROGRESS]
            headers = [c.value for c in ws[1]]
            if headers[:3] == ['序号', '公司名称', '投递岗位']:
                style_header(ws)
                return wb, ws
        # 文件存在但缺进度 sheet（旧版单表文件）→ 补建，保留邮件 sheet
        ws = wb.create_sheet(SHEET_PROGRESS)
        ensure_headers(ws, PROGRESS_HEADERS)
        return wb, ws
    # 文件不存在 → 创建总表（邮件 sheet + 进度 sheet）
    wb = openpyxl.Workbook()
    ws_mail = wb.active
    ws_mail.title = SHEET_MAIL
    ensure_headers(ws_mail, MAIL_HEADERS)
    ws = wb.create_sheet(SHEET_PROGRESS)
    ensure_headers(ws, PROGRESS_HEADERS)
    return wb, ws


def find_row(ws, company):
    """按公司名查找行号，返回行号或 None"""
    for row in ws.iter_rows(min_row=2):
        if row[1].value and str(row[1].value).strip() == company:
            return row[0].row
    return None


def append_row(ws, company, position):
    """追加新行，返回行号"""
    max_seq = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and isinstance(row[0], (int, float)):
            max_seq = max(max_seq, int(row[0]))
    ws.append([max_seq + 1, company, position or '', '', '', '', '', '', '', '', '', '', ''])
    return ws.max_row


def get_cell(ws, row_idx, col_idx):
    return ws.cell(row=row_idx, column=col_idx)


def apply_updates():
    if not os.path.exists(JUDGED_FILE):
        print("⚠️ 没有 pending_judged.json，跳过进度表更新")
        return 0

    with open(JUDGED_FILE, 'r', encoding='utf-8') as f:
        judged = json.load(f)

    candidates = []
    if os.path.exists(CANDIDATES_FILE):
        with open(CANDIDATES_FILE, 'r', encoding='utf-8') as f:
            candidates = json.load(f)
    cand_by_id = {c['id']: c for c in candidates}

    wb, ws = ensure_table()
    updated = 0
    created = 0

    for j in judged:
        if str(j.get('verdict', 'no')).strip().lower() not in ('yes', 'true', '1', 'y'):
            continue
        prog = j.get('progress')
        if not prog or prog.get('update') is False:
            continue

        cand = cand_by_id.get(j.get('id'))
        if cand:
            date_str = str(cand.get('date', ''))[:10]
            if date_str < SINCE_DATE:
                continue   # 8 月前的邮件不更新进度表

        # 公司名：progress.company 优先，否则自动提取
        company = (prog.get('company') or '').strip()
        if not company and cand:
            company = extract_company(cand.get('subject', ''), cand.get('from', ''))
        if not company or company == '未知公司':
            print(f"  ⏭️  判定 #{j.get('id')}：无法确定公司，跳过进度表更新")
            continue

        position = (prog.get('position') or '').strip()
        stage = (prog.get('stage') or '').strip()
        time_val = (prog.get('time') or '').strip() or (cand.get('date') if cand else '')
        link = (prog.get('link') or '').strip() or (cand.get('link') if cand else '')
        result = (prog.get('result') or '').strip()
        note = (prog.get('note') or '').strip()

        row_idx = find_row(ws, company)
        if row_idx is None:
            row_idx = append_row(ws, company, position)
            created += 1
            print(f"  🆕 新增公司行: {company}" + (f"（岗位：{position}）" if position else ""))

        # 岗位：仅当为空时写入
        pos_cell = get_cell(ws, row_idx, COL['投递岗位'])
        if position and not (pos_cell.value or '').strip():
            pos_cell.value = position

        # 阶段时间 / 结果
        if stage in STAGE_COLUMN:
            col_name, is_result = STAGE_COLUMN[stage]
            if is_result:
                result_value = '✅ Offer' if stage == 'Offer' else (result or '')
                if result_value:
                    get_cell(ws, row_idx, COL['结果']).value = result_value
            else:
                if time_val:
                    get_cell(ws, row_idx, COL[col_name]).value = time_val
                    print(f"  ✏️  {company} {col_name} ← {time_val}")
        # 无论 stage 是什么，只要提供了 result 就更新结果列
        if result and not (stage in STAGE_COLUMN and STAGE_COLUMN[stage][1] and stage == 'Offer'):
            get_cell(ws, row_idx, COL['结果']).value = result

        # 投递链接：仅当为空时写入
        link_cell = get_cell(ws, row_idx, COL['投递链接'])
        if link and not (link_cell.value or '').strip():
            link_cell.value = link

        # 备注：追加，保留用户手动内容，避免重复
        if note:
            remark_cell = get_cell(ws, row_idx, COL['备注'])
            old_remark = (remark_cell.value or '').strip()
            if note not in old_remark:
                remark_cell.value = f"{old_remark}；{note}" if old_remark else note

        # 最近动态：更新为当前邮件
        if cand:
            dynamic = f"{cand['date']} {j.get('type', '')}：{(cand.get('subject') or '')[:30]}"
            get_cell(ws, row_idx, COL['最近动态']).value = dynamic

        updated += 1

    if updated or created:
        # 美化：整表统一样式（斑马纹/边框/行高/超链接）+ 结果列彩色标签
        style_body(ws, ncols=len(HEADERS), left_cols=(11, 12, 13), link_col=12)
        for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
            style_result_cell(row[0])
        refresh_filter(ws)
        # 结果列下拉列表 + 条件格式（用户可直接在表格里切换结果，颜色自动跟随）
        setup_result_column(ws)
        wb.save(EXCEL_PATH)
        print(f"✅ 进度表更新完成：更新 {updated} 条，新增公司 {created} 家 → {EXCEL_PATH}（sheet: {SHEET_PROGRESS}）")
    else:
        print("ℹ️ 本次无进度表更新指令")
    return updated


if __name__ == '__main__':
    sys.exit(0 if apply_updates() >= 0 else 1)
