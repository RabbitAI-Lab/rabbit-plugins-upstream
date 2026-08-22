#!/usr/bin/env python3
"""
共享 Excel 样式模块 - 招聘邮件监控系统

统一美化总表两个 sheet（招聘邮件汇总 / 投递记录进度表）：
- 表头：深蓝底白字加粗居中，行高留白
- 正文：微软雅黑、隔行斑马纹、浅色边框、垂直居中、行高加高（解决文字拥挤）
- 状态/类型/结果列：语义化彩色标签（解决颜色单调）
- 链接列：可点击超链接（蓝色下划线）
- 冻结首行 + 自动筛选

用法：from excel_styles import ...
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList
from openpyxl.formatting.rule import Rule, DifferentialStyle
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles.differential import DifferentialStyleList

# ================== 表格路径 & 工作表名（合并后的总表） ==================
# 两个表格合并为一个 Excel 文件，通过切换工作表查看邮件列表 / 投递进度
EXCEL_PATH = '/home/erhao/shared/招聘邮件汇总.xlsx'
SHEET_MAIL = '招聘邮件汇总'      # sheet1：邮件列表（默认打开）
SHEET_PROGRESS = '投递记录进度表'  # sheet2：投递进度

MAIL_HEADERS = ['日期', '邮箱', '主题', '发件人', '状态', '类型', '链接', '截止日期']
PROGRESS_HEADERS = ['序号', '公司名称', '投递岗位', '投递时间', '测评时间', '一面时间', '二面时间',
                    '三面时间', 'HR面时间', '结果', '最近动态', '投递链接', '备注']

FONT_NAME = '微软雅黑'

# ---------- 基础配色 ----------
HEADER_FILL = '4472C4'        # 表头深蓝（Office 经典蓝）
ZEBRA_FILL = 'F2F7FC'         # 隔行斑马纹（极浅蓝）
BORDER_COLOR = 'C9D6E8'       # 浅蓝灰边框
LINK_COLOR = '0563C1'         # 超链接蓝

HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name=FONT_NAME, size=11, color='333333')

THIN = Side(style='thin', color=BORDER_COLOR)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

HEADER_ROW_HEIGHT = 28
DATA_ROW_HEIGHT = 24

# ---------- 状态列配色（汇总表）: (前缀, 填充色, 字体色, 加粗) ----------
STATUS_STYLES = [
    ('⏳', 'FFF2CC', '7F6000', True),            # 待处理：浅黄
    ('超期自动归档', 'F2F2F2', '7F7F7F', False),   # 超期归档：浅灰
    ('用户不需要', 'E7E6E6', '595959', False),     # 用户不需要：灰
    ('✅', 'E2EFDA', '375623', False),            # 已完成：浅绿
]

# ---------- 类型列配色（汇总表）: 精确值 -> (填充色, 字体色) ----------
TYPE_STYLES = {
    '笔试/测评':    ('DDEBF7', '1F4E79'),
    '面试':        ('FCE4D6', 'C55A11'),
    'Offer/录用':  ('E2EFDA', '375623'),
    '宣讲会':      ('E4DFEC', '5B4B8A'),
    '投递确认':    ('EAF1FB', '2E75B6'),
    '其他招聘相关': ('F2F2F2', '595959'),
}

# ---------- 结果列配色（进度表）: (前缀, 填充色, 字体色, 加粗) ----------
RESULT_STYLES = [
    ('✅', 'E2EFDA', '375623', True),    # Offer：浅绿
    ('⏳', 'FFF2CC', '7F6000', True),    # 进行中：浅黄
    ('🎤', 'DDEBF7', '1F4E79', False),   # 流程结束：浅蓝
    ('✍️', 'DDEBF7', '1F4E79', False),   # 流程结束：浅蓝
    ('📮', 'F2F2F2', '595959', False),   # 已投递：浅灰
]


def _match_style(value, styles):
    """按前缀匹配样式，返回 (fill, font_color, bold) 或 None"""
    if value is None:
        return None
    s = str(value).strip()
    for prefix, fill, color, bold in styles:
        if s.startswith(prefix):
            return fill, color, bold
    return None


def set_cell_style(cell, fill=None, color=None, bold=False, align=None):
    """给单个单元格套用统一字体/填充/边框/对齐"""
    cell.font = Font(name=FONT_NAME, size=11, bold=bold, color=color or '333333')
    if fill:
        cell.fill = PatternFill('solid', fgColor=fill)
    if align:
        cell.alignment = align
    cell.border = BORDER


def ensure_headers(ws, headers):
    """表格为空时写入表头并套样式；返回是否新建了表头"""
    first = ws.cell(row=1, column=1).value
    if first is None:
        # ⚠️ 不能用 ws.append() 写表头：上面的 ws.cell(row=1,column=1) 会把 openpyxl
        # 内部 _current_row 推到 1，此时 append 会写到第 2 行（新表头错位）。
        # 改为显式写入第 1 行。
        for i, h in enumerate(headers, start=1):
            ws.cell(row=1, column=i).value = h
        style_header(ws)
        return True
    return False


def style_header(ws, headers=None, height=HEADER_ROW_HEIGHT):
    """表头样式：深蓝底白字加粗居中，冻结首行"""
    if headers:
        for i, h in enumerate(headers, start=1):
            ws.cell(row=1, column=i).value = h
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = PatternFill('solid', fgColor=HEADER_FILL)
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = height
    ws.freeze_panes = 'A2'


def style_body(ws, ncols, left_cols=(), link_col=None):
    """正文样式：斑马纹 + 边框 + 垂直居中 + 行高留白；left_cols 左对齐，link_col 转超链接"""
    for row_idx in range(2, ws.max_row + 1):
        zebra = (row_idx % 2 == 0)
        for col_idx in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.fill = PatternFill('solid', fgColor=ZEBRA_FILL) if zebra else PatternFill('solid', fgColor='FFFFFF')
            cell.alignment = LEFT if col_idx in left_cols else CENTER
            cell.border = BORDER
            if link_col and col_idx == link_col and cell.value:
                cell.hyperlink = str(cell.value)
                cell.font = Font(name=FONT_NAME, size=11, color=LINK_COLOR, underline='single')
        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT


def refresh_filter(ws):
    """自动筛选范围跟随表格行数"""
    ws.auto_filter.ref = f'A1:{chr(64 + ws.max_column)}{ws.max_row}'


def style_status_cell(cell):
    """汇总表状态列：⏳待处理/✅已完成/超期归档/用户不需要 彩色标签"""
    st = _match_style(cell.value, STATUS_STYLES)
    if st:
        set_cell_style(cell, fill=st[0], color=st[1], bold=st[2], align=CENTER)


def style_type_cell(cell):
    """汇总表类型列：笔试/面试/Offer/宣讲会等 彩色标签"""
    key = str(cell.value or '').strip()
    if key in TYPE_STYLES:
        fill, color = TYPE_STYLES[key]
        set_cell_style(cell, fill=fill, color=color, align=CENTER)


def style_result_cell(cell):
    """进度表结果列：Offer/进行中/流程结束/已投递 彩色标签"""
    st = _match_style(cell.value, RESULT_STYLES)
    if st:
        set_cell_style(cell, fill=st[0], color=st[1], bold=st[2], align=CENTER)


# ================== 下拉列表 & 条件格式 ==================
# 下拉选项（与现有取值完全一致，保证脚本逻辑/样式能识别）
STATUS_DROPDOWN_OPTIONS = ['⏳ 待处理', '✅ 已完成', '✅ 已完成（超期自动归档）', '✅ 已完成（用户不需要）']
RESULT_DROPDOWN_OPTIONS = ['⏳ 进行中', '✅ Offer', '✅ 已完成', '🎤 流程结束', '✍️ 流程结束', '📮 已投递']

# 条件格式规则（按顺序匹配、先命中先生效 stopIfTrue）: (匹配文本, 填充色, 字体色, 加粗)
# 用 SEARCH 匹配中文关键词（避免 emoji 代理对在公式里匹配失败）
STATUS_CF_RULES = [
    ('待处理', 'FFF2CC', '7F6000', True),
    ('超期自动归档', 'F2F2F2', '7F7F7F', False),
    ('用户不需要', 'E7E6E6', '595959', False),
    ('已完成', 'E2EFDA', '375623', False),
]
RESULT_CF_RULES = [
    ('Offer', 'E2EFDA', '375623', True),
    ('已完成', 'E2EFDA', '375623', True),
    ('进行中', 'FFF2CC', '7F6000', True),
    ('流程结束', 'DDEBF7', '1F4E79', False),
    ('已投递', 'F2F2F2', '595959', False),
]

# 下拉/条件格式覆盖范围：表格行数 + 300，保底 1000 行（给未来自动追加的行预留）
DROP_EXTRA_ROWS = 300
DROP_MIN_ROWS = 1000


def _dropdown_last_row(ws):
    return max(ws.max_row + DROP_EXTRA_ROWS, DROP_MIN_ROWS)


def setup_status_column(ws, col_idx=5):
    """汇总表「状态」列：下拉列表 + 条件格式自动着色（幂等，保存前调用）"""
    _setup_column(ws, col_idx, STATUS_DROPDOWN_OPTIONS, STATUS_CF_RULES)


def setup_result_column(ws, col_idx=10):
    """进度表「结果」列：下拉列表 + 条件格式自动着色（幂等，保存前调用）"""
    _setup_column(ws, col_idx, RESULT_DROPDOWN_OPTIONS, RESULT_CF_RULES)


def _prune_orphan_dxfs(ws):
    """清理未被任何条件格式规则引用的孤儿 dxf（防止反复保存撑大文件）

    规则保存时会把 rule.dxf 重新 add 进 wb._differential_styles 并重新分配 dxfId，
    因此这里只保留当前所有 sheet 规则实际引用的 dxf 是安全的。
    """
    wb = ws.parent
    new_list = DifferentialStyleList()
    for sheet in wb.worksheets:
        for cf in sheet.conditional_formatting:
            for rule in cf.rules:
                if rule.dxf is not None:
                    new_list.add(rule.dxf)
    wb._differential_styles = new_list


def _setup_column(ws, col_idx, options, cf_rules):
    """给指定列加下拉列表 + 条件格式。每次重建（清旧防重复累积），幂等"""
    # 清掉旧的验证/条件格式，避免重复保存时累积
    ws.data_validations = DataValidationList()
    ws.conditional_formatting = ConditionalFormattingList()

    last_row = _dropdown_last_row(ws)
    letter = get_column_letter(col_idx)
    rng = f'{letter}2:{letter}{last_row}'

    # 下拉列表（showDropDown 保持默认 None = 显示下拉箭头）
    # 选项为内联字符串，逗号分隔；选项本身不含逗号，长度远小于 255 上限
    dv = DataValidation(type='list', formula1='"' + ','.join(options) + '"', allow_blank=True)
    dv.add(rng)
    ws.add_data_validation(dv)

    # 条件格式：值一变颜色自动跟随（用户在表格里手动切换下拉也实时变色）
    for text, fill, color, bold in cf_rules:
        formula = f'ISNUMBER(SEARCH("{text}",${letter}2))'
        # ⚠️ 坑：条件格式(dxf)的 solid 填充渲染用的是 bgColor，只设 fgColor 会不显示颜色。
        # 必须 start_color/end_color 都设（openpyxl 分别映射为 fgColor/bgColor）。
        dxf = DifferentialStyle(
            fill=PatternFill(start_color=fill, end_color=fill, fill_type='solid'),
            font=Font(color=color, bold=bold),
        )
        rule = Rule(type='expression', dxf=dxf, stopIfTrue=True)
        rule.formula = [formula]
        ws.conditional_formatting.add(rng, rule)

    # 清理孤儿 dxf（旧规则被清掉后残留的样式定义）
    _prune_orphan_dxfs(ws)
