#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""非模型模块自测：validate + ledger_builder + run_pipeline 导入与编排（不依赖 PaddleOCR/VLM）。"""
import os
import sys
import tempfile

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

# 1) run_pipeline 应能干净导入（重型依赖懒加载，不触发 paddleocr/requests 安装）
import run_pipeline  # noqa
import field_extractor  # noqa
import validate
import ledger_builder
import config

print("[1] run_pipeline / field_extractor 导入 OK（未触发 paddleocr/requests 安装）")

# 2) validate 校验逻辑
ok_fields = {
    "发票类型": "增值税专用发票", "发票号码": "12345678",
    "开票日期": "2026-07-08", "价税合计": 113.0, "不含税金额": 100.0,
    "税额": 13.0, "销售方名称": "测试科技有限公司",
}
st, rm, it = validate.validate(ok_fields)
assert st == "正常", f"正常票应判正常, 实得 {st}:{rm}"
assert it == "增值税专用发票"
print("[2.1] 正常票校验通过:", st, it, rm)

bad_fields = {
    "发票类型": "增值税专用发票", "发票号码": "123",
    "开票日期": "2026-07-08", "价税合计": 200.0, "不含税金额": 100.0,
    "税额": 13.0, "销售方名称": "测试科技有限公司",
}
st, rm, it = validate.validate(bad_fields)
assert st == "异常", f"勾稽/格式错误应判异常, 实得 {st}"
assert any("价税合计≠" in r for r in rm), rm
assert any("发票号码格式异常" in r for r in rm), rm
print("[2.2] 异常票校验捕获:", rm)

miss_fields = {"发票类型": "未知票", "价税合计": 50.0, "不含税金额": 45.0, "税额": 5.0}
st, rm, it = validate.validate(miss_fields)
assert it == "待确认", it
assert any("发票类型未能识别" in r for r in rm), rm
print("[2.3] 缺类型判待确认:", st, it, rm)

# 3) ledger_builder 生成并回读
records = [
    {"_src": "a.pdf", "开票日期": "2026-07-08", "发票类型": "增值税专用发票",
     "销售方名称": "甲科技", "不含税金额": 100.0, "税额": 13.0, "价税合计": 113.0,
     "发票号码": "12345678", "票据状态": "正常", "所属月份": "2026-07", "风险备注": ""},
    {"_src": "b.jpg", "开票日期": "2026-07-09", "发票类型": "电子普通发票",
     "销售方名称": "乙超市", "不含税金额": 200.0, "税额": 6.0, "价税合计": 206.0,
     "发票号码": "87654321", "票据状态": "正常", "所属月份": "2026-07", "风险备注": ""},
    {"_src": "c.png", "开票日期": "", "发票类型": "待确认", "销售方名称": "",
     "不含税金额": "", "税额": "", "价税合计": "", "发票号码": "12",
     "票据状态": "异常", "所属月份": "未知月份", "风险备注": "发票号码格式异常"},
]
tmp = tempfile.mkdtemp()
out = os.path.join(tmp, "月度财税台账.xlsx")
ledger_builder.build(records, out)
print(f"[3] 台账已生成: {out}")

import openpyxl
wb = openpyxl.load_workbook(out)
assert wb.sheetnames == ["月度财税台账", "异常票据清单"], wb.sheetnames
ws = wb["月度财税台账"]
rows = list(ws.iter_rows(values_only=True))
assert list(rows[0]) == config.LEDGER_HEADER, rows[0]
assert len(rows) == 4, f"应有1表头+2正常+1合计=4行, 实得 {len(rows)}"  # a,b + 合计
# 合计行校验：价税合计 113+206=319, 税额 13+6=19
total_row = rows[3]
assert total_row[5] == 319.0 and total_row[4] == 19.0, total_row
ab = wb["异常票据清单"]
abrows = list(ab.iter_rows(values_only=True))
assert len(abrows) == 2, abrows  # 表头 + c
assert abrows[1][0] == "c.png" and abrows[1][1] == "异常", abrows[1]
print("[3.1] 台账表头/正常行/合计行/异常sheet 全部正确")

# 4) run_pipeline 空目录分支
empty = tempfile.mkdtemp()
rc = run_pipeline.main2 if hasattr(run_pipeline, "main2") else None
# 直接调用内部函数验证空目录
files = run_pipeline.iter_invoice_files(empty)
assert files == [], "空目录应无文件"
print("[4] 空目录遍历正确（无文件即返回提示，不崩）")

print("\n全部自测通过 ✅")
