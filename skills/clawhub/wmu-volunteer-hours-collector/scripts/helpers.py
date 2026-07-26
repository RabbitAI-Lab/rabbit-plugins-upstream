#!/usr/bin/env python3
"""温州医科大学志愿时长统计辅助脚本

本脚本提供志愿时长统计流程中的关键操作函数。
由 AI agent 在对话工作流中调用，不独立运行。
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
import shutil
import os
from datetime import datetime, date


TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "志愿服务.xlsx")

HEADER_ROWS = 4
FOOTER_ROWS = 10

COL_MAP = {
    "活动名称": 1,      # A
    "活动时间起": 2,     # B
    "活动时间止": 3,     # C
    "序号": 4,          # D
    "报名来源": 5,       # E
    "姓名": 6,          # F
    "学院": 7,          # G
    "年级": 8,          # H
    "专业": 9,          # I
    "学号": 10,         # J
    "长号": 11,         # K
    "短号": 12,         # L
    "服务时长": 13,      # M
    "备注": 14,         # N
}

FONT_SONG = Font(name='宋体', size=11, charset=134)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)


def get_template_path():
    """返回模板文件路径"""
    return TEMPLATE_PATH


def copy_template(dest_dir: str, dest_name: str = None) -> str:
    """将模板文件复制到目标目录。

    Args:
        dest_dir: 目标目录路径
        dest_name: 目标文件名，默认为"志愿服务.xlsx"

    Returns:
        复制后的文件完整路径
    """
    os.makedirs(dest_dir, exist_ok=True)
    if dest_name is None:
        dest_name = "志愿服务.xlsx"
    dest_path = os.path.join(dest_dir, dest_name)
    shutil.copy2(TEMPLATE_PATH, dest_path)
    return dest_path


def load_template(path: str = None):
    """加载模板工作簿，返回 (workbook, worksheet)。

    Args:
        path: xlsx 文件路径，默认使用内置模板

    Returns:
        (openpyxl.Workbook, openpyxl.Worksheet)
    """
    if path is None:
        path = TEMPLATE_PATH
    wb = openpyxl.load_workbook(path)
    ws = wb['模板']
    return wb, ws


def get_data_range(ws) -> tuple:
    """获取表格数据区域的起止行号（表头之后、表尾之前）。

    Returns:
        (data_start_row, data_end_row): 数据区起止行号
    """
    total_rows = ws.max_row
    data_start = HEADER_ROWS + 1  # 第5行
    data_end = total_rows - FOOTER_ROWS  # 倒数第11行
    return data_start, data_end


def get_footer_start(ws) -> int:
    """获取表尾起始行号"""
    return ws.max_row - FOOTER_ROWS + 1


def adjust_rows(ws, needed_count: int):
    """调整数据区行数，使其恰好容纳 needed_count 条记录。

    在表尾前插入或删除行。保持表尾贴在数据区下方。

    插入行时，先保存表尾合并单元格并解除，插入完成后再在
    移位后的位置重建合并单元格。避免 openpyxl 在合并单元格
    附近 insert_rows 时产生不可写入的 MergedCell artifact。

    Args:
        ws: 模板工作表
        needed_count: 需要的条目行数
    """
    data_start, data_end = get_data_range(ws)
    current_count = data_end - data_start + 1
    footer_start = get_footer_start(ws)

    if needed_count > current_count:
        rows_to_add = needed_count - current_count
        # 保存表尾所有合并单元格范围，然后解除
        # ——避免 insert_rows 在合并单元格边界处产生 MergedCell artifact
        footer_merges = []
        for mc in list(ws.merged_cells.ranges):
            if mc.min_row >= footer_start:
                footer_merges.append(
                    (mc.min_row, mc.max_row, mc.min_col, mc.max_col)
                )
                ws.unmerge_cells(str(mc))
        for _ in range(rows_to_add):
            ws.insert_rows(footer_start)
        # 在移位后的表尾位置重建合并单元格
        for min_r, max_r, min_c, max_c in footer_merges:
            ws.merge_cells(
                start_row=min_r + rows_to_add,
                start_column=min_c,
                end_row=max_r + rows_to_add,
                end_column=max_c,
            )
    elif needed_count < current_count:
        rows_to_remove = current_count - needed_count
        for _ in range(rows_to_remove):
            ws.delete_rows(data_start)


def fill_row(ws, row_num: int, data: dict):
    """填充一行数据。

    Args:
        ws: 工作表
        row_num: 行号
        data: 键为 COL_MAP 中的键的字典
    """
    for key, value in data.items():
        col = COL_MAP.get(key)
        if col is None:
            continue
        cell = ws.cell(row=row_num, column=col)
        cell.value = value
        cell.font = FONT_SONG
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def fill_data(ws, records: list):
    """批量填充数据到表格数据区。

    每条记录是一个字典，键对应 COL_MAP 的键。
    自动调整行数并填入数据。

    Args:
        ws: 工作表
        records: 记录列表，每条记录包含：活动名称, 活动时间起, 活动时间止,
                 序号, 报名来源, 姓名, 学院, 年级, 专业, 学号, 长号, 短号, 服务时长, 备注
    """
    adjust_rows(ws, len(records))
    data_start, _ = get_data_range(ws)
    for i, rec in enumerate(records):
        fill_row(ws, data_start + i, rec)


def parse_excel_files(directory: str) -> list:
    """扫描目录中的所有表格文件，尝试用 pandas 解析并返回数据帧列表。

    Returns:
        [(filename, DataFrame), ...]
    """
    import pandas as pd
    results = []
    for fname in sorted(os.listdir(directory)):
        fpath = os.path.join(directory, fname)
        if not os.path.isfile(fpath):
            continue
        ext = os.path.splitext(fname)[1].lower()
        if ext not in ('.xlsx', '.xls', '.csv', '.xlsm'):
            continue
        try:
            if ext == '.csv':
                df = pd.read_csv(fpath, dtype=str).fillna('')
            else:
                df = pd.read_excel(fpath, dtype=str).fillna('')
            results.append((fname, df))
        except Exception as e:
            print(f"⚠ 无法解析 {fname}: {e}")
    return results


def date_to_chinese(d: date) -> str:
    """将 date 对象转换为中文日期格式，如 '2026年1月10日'"""
    return f"{d.year}年{d.month}月{d.day}日"


def parse_date(text: str) -> date:
    """尝试从各种文本格式解析日期。

    支持: '2026-01-10', '2026/1/10', '2026年1月10日', '1月10日', '20260110', 等
    """
    import re
    text = str(text).strip()
    if not text:
        return None

    # 中文格式
    m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', text)
    if m:
        return date(int(m[1]), int(m[2]), int(m[3]))

    # 只有月日
    m = re.match(r'(\d{1,2})月(\d{1,2})日', text)
    if m:
        today = date.today()
        return date(today.year, int(m[1]), int(m[2]))

    # YYYY-MM-DD 或 YYYY/MM/DD
    m = re.match(r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})', text)
    if m:
        return date(int(m[1]), int(m[2]), int(m[3]))

    # YYYYMMDD
    m = re.match(r'(\d{4})(\d{2})(\d{2})', text)
    if m:
        return date(int(m[1]), int(m[2]), int(m[3]))

    return None


def extract_grade(text: str, student_id: str = None) -> str:
    """从原始文本中提取年级后两位。

    支持: '2025', '2025级', '25级', '大二', 等。
    如果有学号，可以用于验证。

    Returns:
        年级后两位字符串，如 '25'
    """
    import re
    text = str(text).strip()

    # 4位年份格式: 2025, 2025级
    m = re.search(r'(\d{4})', text)
    if m:
        return m[1][2:]

    # 2位格式: 25, 25级
    m = re.search(r'(\d{2})', text)
    if m:
        return m[1]

    # 文字格式: 大一, 大二, etc.
    grade_map = {
        '大一': 1, '大二': 2, '大三': 3, '大四': 4, '大五': 5,
        '研一': 1, '研二': 2, '研三': 3,
    }
    today = date.today()
    for k, v in grade_map.items():
        if k in text:
            year = today.year - v + 1
            # 如果是8月之后，入学年份需要调整
            if today.month >= 9:
                year = today.year - v + 1
            else:
                year = today.year - v
            return str(year)[2:]

    return None


def resolve_college(raw: str) -> str:
    """根据原始文本解析学院简称。

    Returns:
        学院简称，或 None 表示无法确定
    """
    import re
    text = str(raw).strip().replace(' ', '').replace('\u3000', '')

    # 精确匹配简称
    abbr_map = {
        '眼生': '眼生', '临一': '临一', '临二': '临二', '药学院': '药学院',
        '中医药': '中医药', '精神': '精神', '检生': '检生', '公卫': '公卫',
        '口腔': '口腔', '康复': '康复', '护理': '护理', '外国语': '外国语',
        '基础': '基础', '文管': '文管', '阿尔伯塔': '阿尔伯塔', '仁济': '仁济',
    }
    for k, v in abbr_map.items():
        if text == k:
            return v

    # 关键词匹配
    if '阿尔伯塔' in text:
        return '阿尔伯塔'
    if '仁济' in text:
        return '仁济'
    if '眼视光' in text or '眼生' in text or '生物医学工程' in text:
        if '眼视光' in text and ('生物医学工程' in text or '生工' in text):
            return '眼生'
        if '眼' in text:
            return '眼生'
    if '临一' in text or '第一临床' in text or '第一临床医学院' in text or '信息与工程' in text:
        return '临一'
    if '临二' in text or '第二临床' in text or '二临' in text:
        return '临二'
    if '药学' in text or '药学院' in text:
        return '药学院'
    if '中医药' in text or '中医' in text:
        return '中医药'
    if '精神' in text:
        return '精神'
    if '检生' in text or '检验' in text or '生命科学' in text:
        return '检生'
    if '公卫' in text or '公共卫生' in text:
        return '公卫'
    if '口腔' in text:
        return '口腔'
    if '康复' in text:
        return '康复'
    if '护理' in text:
        return '护理'
    if '外国语' in text or '外语' in text:
        return '外国语'
    if '基础' in text:
        return '基础'
    if '文管' in text or '人文' in text or '医学人文' in text:
        return '文管'

    return None


def resolve_major(raw: str, college_abbr: str = None) -> dict:
    """根据原始文本解析专业简称。

    Args:
        raw: 原始专业/班级文本
        college_abbr: 已解析的学院简称

    Returns:
        {'major': 专业简称, 'certain': 是否确定, 'possible': [可能选项列表], 'is_graduate': 是否研究生}
    """
    import re
    text = str(raw).strip().replace(' ', '').replace('\u3000', '')

    result = {'major': None, 'certain': False, 'possible': [], 'is_graduate': False}

    # 生物医药数据科学（眼生学院较新专业，直接使用全称）
    if '生物医药数据科学' in text or '生医数据' in text or '生物数据科学' in text:
        result['major'] = '生物医药数据科学'
        result['certain'] = True
        return result

    # 研究生检测
    if '研' in text and any(c.isdigit() for c in text):
        result['major'] = text
        result['is_graduate'] = True
        result['certain'] = True
        return result

    # "护理" 单独出现时无法区分护本和护专，需要用户确认
    if text == '护理' or text == '护理专业':
        result['possible'] = ['护本（护理学）', '护专（护理）']
        return result

    # 班级名解析
    class_match = re.search(r'(\d{2})([^\d]+)\d*班?', text)
    if class_match:
        grade_part = class_match[1]
        rest = class_match.group(2)
        result['grade_hint'] = grade_part
        # 尝试从班级名前缀推断专业
        text = rest
        # 去除尾部数字
        text = re.sub(r'\d+$', '', text)

    # 精确匹配简称
    exact_abbr = {
        '眼专': '眼专', '眼本': '眼本', '眼八': '眼八', '生工': '生工',
        '眼新工': '眼新工', '眼新医': '眼新医',
        '临床': '临床', '临八': '临八', '临检': '临检', '临新医': '临新医',
        '信管': '信管', '影技': '影技', '影像': '影像', '影八': '影八',
        '儿八': '儿八', '儿科': '儿科', '麻醉': '麻醉', '麻八': '麻八',
        '药创': '药创', '药学': '药学', '制药': '制药', '临药': '临药',
        '中医': '中医', '中药': '中药',
        '精八': '精八', '精神': '精神', '老医': '老医', '心理': '心理',
        '医检': '医检', '卫检': '卫检', '生技': '生技',
        '预防': '预防', '放射': '放射', '放八': '放八',
        '口腔': '口腔',
        '言康': '言康', '运康': '运康', '康复': '康复',
        '助产': '助产', '护本': '护本', '护专': '护专',
        '英语': '英语', '日语': '日语',
        '基医': '基医',
        '营销': '营销', '社保': '社保', '康管': '康管', '公管': '公管',
        '临外': '临外',
        '法医': '法医', '眼视光': '眼视光', '全科': '全科',
    }
    if text in exact_abbr:
        result['major'] = exact_abbr[text]
        result['certain'] = True
        return result

    # 通过完整专业名匹配
    full_name_map = {
        '眼视光技术': '眼专', '眼视光医学': '眼本',
        '眼视光医学（5+3一体化）': '眼八', '生物医学工程': '生工',
        '生物医学工程（眼视光工程新工科班）': '眼新工',
        '眼视光医学（新医科班）': '眼新医',
        '临床医学': '临床', '临床医学（5+3一体化）': '临八',
        '临床医学（检验医师培养试验班）': '临检',
        '临床医学（新医科）': '临新医',
        '信息管理与信息系统': '信管', '医学影像技术': '影技',
        '医学影像学': '影像', '医学影像学（5+3一体化）': '影八',
        '儿科学（5+3）': '儿八', '儿科学（5年制）': '儿科',
        '麻醉学': '麻醉', '麻醉学（5+3一体化）': '麻八',
        '药学（生物药学创新实验班）': '药创', '药学': '药学',
        '生物制药': '制药', '临床药学': '临药',
        '中医学': '中医', '中药学': '中药',
        '精神医学（5+3一体化）': '精八', '精神医学（五年制）': '精神',
        '老年医学（5+3一体化）': '老医', '应用心理学': '心理',
        '医学检验技术': '医检', '卫生检验与检疫': '卫检', '生物技术': '生技',
        '预防医学': '预防', '放射医学（五年制）': '放射',
        '放射医学（5+3一体化）': '放八',
        '口腔医学': '口腔',
        '听力与言语康复学': '言康', '运动康复': '运康', '康复治疗学': '康复',
        '助产学': '助产', '护理学': '护本',
        '英语': '英语', '日语': '日语',
        '基础医学': '基医',
        '市场营销': '营销', '社会劳动保障': '社保',
        '健康服务与管理': '康管', '公共事业管理': '公管',
        '临床医学（中外合作办学）': '临外',
        '法医学': '法医', '眼视光学': '眼视光', '全科医学': '全科',
        '生物医药数据科学': '生物医药数据科学',
    }
    for k, v in full_name_map.items():
        if k in text or text in k:
            if result['major'] is None:
                result['major'] = v
                result['certain'] = True
                return result

    # 去除尾部数字后重试（如 "护本18" → "护本"）
    stripped = re.sub(r'\d+$', '', text)
    if stripped and stripped != text and stripped in exact_abbr:
        result['major'] = exact_abbr[stripped]
        result['certain'] = True
        return result

    # 模糊匹配
    if '眼视光' in text:
        if '5+3' in text or '八年' in text:
            result['major'] = '眼八'
        elif '新医科' in text:
            result['major'] = '眼新医'
        elif '技术' in text or '专' in text:
            result['major'] = '眼专'
        else:
            result['major'] = '眼本'
        result['certain'] = True
        return result

    if '临床' in text:
        if '5+3' in text or '八年' in text:
            result['major'] = '临八'
        elif '检验医师' in text:
            result['major'] = '临检'
        elif '新医科' in text:
            result['major'] = '临新医'
        elif '中外' in text:
            result['major'] = '临外'
        else:
            result['major'] = '临床'
        result['certain'] = True
        return result

    if '护理' in text:
        if '学' in text:
            result['major'] = '护本'
        elif '护专' in text:
            result['major'] = '护专'
        else:
            result['possible'] = ['护本（护理学）', '护专（护理）']
            return result
        result['certain'] = True
        return result

    if '麻醉' in text:
        result['major'] = '麻八' if ('5+3' in text or '八年' in text) else '麻醉'
        result['certain'] = True
        return result

    if '口腔' in text:
        result['major'] = '口腔'
        result['certain'] = True
        return result

    if '康复' in text:
        if '听力' in text or '言语' in text:
            result['major'] = '言康'
        elif '运动' in text:
            result['major'] = '运康'
        else:
            result['major'] = '康复'
        result['certain'] = True
        return result

    if '检验' in text:
        result['major'] = '医检'
        result['certain'] = True
        return result

    if '助产' in text:
        result['major'] = '助产'
        result['certain'] = True
        return result

    if '心理' in text:
        result['major'] = '心理'
        result['certain'] = True
        return result

    if '预防' in text:
        result['major'] = '预防'
        result['certain'] = True
        return result

    if '放射' in text:
        result['major'] = '放八' if ('5+3' in text or '八年' in text) else '放射'
        result['certain'] = True
        return result

    if '影像' in text:
        if '5+3' in text or '八年' in text:
            result['major'] = '影八'
        elif '技术' in text:
            result['major'] = '影技'
        else:
            result['major'] = '影像'
        result['certain'] = True
        return result

    if '药学' in text:
        if '创新' in text or '实验班' in text:
            result['major'] = '药创'
        elif '临床药学' in text:
            result['major'] = '临药'
        elif '制药' in text or '生物制药' in text:
            result['major'] = '制药'
        else:
            result['major'] = '药学'
        result['certain'] = True
        return result

    if '中医' in text:
        result['major'] = '中医'
        result['certain'] = True
        return result

    if '中药' in text:
        result['major'] = '中药'
        result['certain'] = True
        return result

    if '英语' in text:
        result['major'] = '英语'
        result['certain'] = True
        return result

    if '日语' in text:
        result['major'] = '日语'
        result['certain'] = True
        return result

    if '信管' in text or '信息管理' in text:
        result['major'] = '信管'
        result['certain'] = True
        return result

    if '公管' in text or '公共事业' in text:
        result['major'] = '公管'
        result['certain'] = True
        return result

    if '营销' in text or '市场' in text:
        result['major'] = '营销'
        result['certain'] = True
        return result

    if '社保' in text or '保障' in text:
        result['major'] = '社保'
        result['certain'] = True
        return result

    if '康管' in text or '健康服务' in text:
        result['major'] = '康管'
        result['certain'] = True
        return result

    if '基医' in text or '基础医学' in text:
        result['major'] = '基医'
        result['certain'] = True
        return result

    if '法医' in text:
        result['major'] = '法医'
        result['certain'] = True
        return result

    if '全科' in text:
        result['major'] = '全科'
        result['certain'] = True
        return result

    if '生技' in text or '生物技术' in text:
        result['major'] = '生技'
        result['certain'] = True
        return result

    if '儿科学' in text or '儿科' in text:
        result['major'] = '儿八' if ('5+3' in text) else '儿科'
        result['certain'] = True
        return result

    if '卫检' in text or '卫生检验' in text:
        result['major'] = '卫检'
        result['certain'] = True
        return result

    return result


def validate_grade_vs_student_id(grade: str, student_id: str) -> bool:
    """验证年级是否与学号前两位一致。
    
    学号格式通常为"YYYYXXXX"（12位）或"YYXXXX"（8位），
    分别检查前两位和3-4位。
    """
    if not grade or not student_id:
        return True
    sid = str(student_id).strip()
    grade = str(grade).strip()
    if len(sid) >= 2:
        if sid[:2] == grade:
            return True
    if len(sid) >= 4:
        if sid[2:4] == grade:
            return True
    return False

