# -*- coding: utf-8 -*-
"""
表格去重清洗脚本
用法:
    python dedup_clean.py <输入.xlsx> [输出.xlsx] [--key 列名1,列名2] [--keep first|last]

功能:
1. 去除首尾/中间多余空格、全角空格
2. 全角字母数字和常见全角标点转半角
3. 手机号标准化(去 +86、去分隔符)
4. 按 key 列去重(默认自动选择"姓名/手机号/编号"类列; keep=first 保留首条)
5. 输出清洗报告(清洗了多少单元格、删除了多少重复行)
"""
import sys
import re
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要 openpyxl, 请先安装: pip install openpyxl")
    sys.exit(1)

FULLWIDTH_MAP = {
    "０": "0", "１": "1", "２": "2", "３": "3", "４": "4",
    "５": "5", "６": "6", "７": "7", "８": "8", "９": "9",
    "Ａ": "A", "Ｂ": "B", "Ｃ": "C", "Ｄ": "D", "Ｅ": "E", "Ｆ": "F",
    "Ｇ": "G", "Ｈ": "H", "Ｉ": "I", "Ｊ": "J", "Ｋ": "K", "Ｌ": "L",
    "Ｍ": "M", "Ｎ": "N", "Ｏ": "O", "Ｐ": "P", "Ｑ": "Q", "Ｒ": "R",
    "Ｓ": "S", "Ｔ": "T", "Ｕ": "U", "Ｖ": "V", "Ｗ": "W", "Ｘ": "X",
    "Ｙ": "Y", "Ｚ": "Z",
    "ａ": "a", "ｂ": "b", "ｃ": "c", "ｄ": "d", "ｅ": "e", "ｆ": "f",
    "ｇ": "g", "ｈ": "h", "ｉ": "i", "ｊ": "j", "ｋ": "k", "ｌ": "l",
    "ｍ": "m", "ｎ": "n", "ｏ": "o", "ｐ": "p", "ｑ": "q", "ｒ": "r",
    "ｓ": "s", "ｔ": "t", "ｕ": "u", "ｖ": "v", "ｗ": "w", "ｘ": "x",
    "ｙ": "y", "ｚ": "z",
    "　": " ", "％": "%", "＃": "#", "＠": "@", "＆": "&",
    "－": "-", "（": "(", "）": ")", "，": ",", "。": ".",
    "：": ":", "；": ";", "！": "!", "？": "?", "＂": '"', "＇": "'",
}

PHONE_RE = re.compile(r"^(\+?86)?[-\s]?1[3-9]\d{9}$")
CN_ID_RE = re.compile(r"^\d{17}[\dXx]$")


def to_halfwidth(text: str) -> str:
    return "".join(FULLWIDTH_MAP.get(ch, ch) for ch in text)


def clean_cell(value, phone_cols=None, id_cols=None):
    """清洗单个单元格, 返回 (新值, 是否修改)"""
    if value is None or isinstance(value, (int, float, bool)):
        return value, False
    s = str(value)
    if not s.strip():
        return (None if s == "" else value), s != (value if value else s)

    original = s
    # 1. 全角转半角
    s = to_halfwidth(s)
    # 2. 合并中间空格 / 去首尾空格
    s = re.sub(r"\s+", " ", s).strip()

    # 3. 手机号标准化
    if phone_cols is not None and phone_cols:
        pass  # 在调用侧按列处理, 这里不处理

    if s != original:
        return s, True
    return s, False


def normalize_phone(s: str):
    digits = re.sub(r"[^0-9Xx+]", "", str(s))
    if PHONE_RE.match(digits):
        return digits[-11:], True
    return s, False


def normalize_id(s: str):
    s = str(s).strip().upper()
    if CN_ID_RE.match(s):
        return s, True
    return s, False


def detect_key_columns(headers):
    """自动猜去重 key 列: 姓名类 + 编号/手机号类"""
    name_cols, id_cols = [], []
    for i, h in enumerate(headers):
        h = str(h or "").strip()
        if h in ("姓名", "名字", "员工姓名", "客户姓名", "学员姓名"):
            name_cols.append(i)
        if any(k in h for k in ("手机", "电话", "工号", "编号", "学号", "身份证")):
            id_cols.append(i)
    if id_cols:
        return id_cols[:2]
    if name_cols:
        return name_cols[:1] + ([2] if len(headers) > 2 else [])
    return None


def main():
    parser = argparse.ArgumentParser(description="Excel 表格去重清洗")
    parser.add_argument("input", help="输入 xlsx 文件")
    parser.add_argument("output", nargs="?", help="输出 xlsx 文件 (默认在原文件名后加 _cleaned)")
    parser.add_argument("--key", default="", help="去重 key 列名, 逗号分隔, 默认自动检测")
    parser.add_argument("--keep", choices=["first", "last"], default="first", help="重复行保留哪条")
    parser.add_argument("--sheet", default="", help="处理的 sheet 名 (默认第一个)")
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        print(f"错误: 找不到文件 {src}")
        sys.exit(1)

    dst = Path(args.output) if args.output else src.with_name(src.stem + "_cleaned.xlsx")

    wb = load_workbook(src, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.worksheets[0]

    # 读取所有行
    rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    if len(rows) < 1:
        print("错误: 空表")
        sys.exit(1)

    # 定位表头行(默认第1行; 若第1行全空则下移)
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        if any(c is not None and str(c).strip() for c in row):
            header_idx = i
            break
    headers = [str(h).strip() if h is not None else "" for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    # 去掉整行为空的数据
    data_rows = [r for r in data_rows if any(c is not None and str(c).strip() for c in r)]
    total = len(data_rows)

    # 列语义识别
    phone_cols = [i for i, h in enumerate(headers) if "手机" in h or "电话" in h]
    idcard_cols = [i for i, h in enumerate(headers) if "身份证" in h]

    # 清洗
    cleaned_cells = 0
    for r_idx, row in enumerate(data_rows):
        new_row = list(row)
        for c_idx, val in enumerate(row):
            v, changed = clean_cell(val)
            if changed:
                cleaned_cells += 1
            new_row[c_idx] = v
        for i in phone_cols:
            if i < len(new_row) and new_row[i] is not None:
                v, changed = normalize_phone(new_row[i])
                if changed:
                    new_row[i] = v
                    cleaned_cells += 1
        for i in idcard_cols:
            if i < len(new_row) and new_row[i] is not None:
                v, changed = normalize_id(new_row[i])
                if changed:
                    new_row[i] = v
                    cleaned_cells += 1
        data_rows[r_idx] = new_row

    # 去重 key
    if args.key:
        key_names = [k.strip() for k in args.key.split(",") if k.strip()]
        key_idx = []
        for kn in key_names:
            if kn in headers:
                key_idx.append(headers.index(kn))
            else:
                print(f"警告: key 列 '{kn}' 不在表头中, 忽略")
        if not key_idx:
            print("错误: 没有可用的 key 列")
            sys.exit(1)
    else:
        key_idx = detect_key_columns(headers)
        if not key_idx:
            print("提示: 未自动识别到 key 列, 跳过去重 (只做清洗)")
    key_idx = [i for i in key_idx if i < len(headers)]

    dup_removed = 0
    if key_idx:
        seen = {}
        kept = []
        for row in data_rows:
            key = tuple(str(row[i]).strip().lower() if i < len(row) and row[i] is not None else "" for i in key_idx)
            if key in seen:
                dup_removed += 1
                if args.keep == "last":
                    kept[seen[key]] = row
                continue
            seen[key] = len(kept)
            kept.append(row)
        data_rows = kept

    # 写输出
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "清洗结果"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2563EB")
    out_ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = out_ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in data_rows:
        out_ws.append(list(row))

    # 列宽自适应
    for c_idx, h in enumerate(headers, 1):
        max_len = len(str(h or ""))
        for row in data_rows[:200]:
            if c_idx - 1 < len(row) and row[c_idx - 1] is not None:
                l = len(str(row[c_idx - 1]))
                if l > max_len:
                    max_len = l
        out_ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 4, 40)
    out_ws.freeze_panes = "A2"

    # 清洗报告 sheet
    report_ws = out_wb.create_sheet("清洗报告")
    report = [
        ["清洗报告", ""],
        ["源文件", str(src)],
        ["处理 sheet", ws.title],
        ["原始数据行", total],
        ["删除重复行", dup_removed],
        ["清洗后行数", len(data_rows)],
        ["清洗单元格数", cleaned_cells],
        ["去重 key 列", ", ".join(headers[i] for i in key_idx) if key_idx else "(无, 仅清洗)"],
        ["重复行保留策略", f"保留{ '首条' if args.keep == 'first' else '末条' }"],
    ]
    for r in report:
        report_ws.append(r)
    report_ws["A1"].font = Font(bold=True, size=14)
    report_ws.column_dimensions["A"].width = 20
    report_ws.column_dimensions["B"].width = 60

    out_wb.save(dst)

    print("=" * 46)
    print("清洗完成")
    print(f"  原始数据行 : {total}")
    print(f"  删除重复行 : {dup_removed}")
    print(f"  清洗单元格 : {cleaned_cells}")
    print(f"  去重 key 列: {', '.join(headers[i] for i in key_idx) if key_idx else '无(仅清洗)'}")
    print(f"  输出文件   : {dst}")
    print("=" * 46)


if __name__ == "__main__":
    main()
